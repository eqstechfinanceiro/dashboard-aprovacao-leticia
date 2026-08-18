import requests
import json
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

API_URL = "https://api.vexpenses.com"
API_KEY = "N2ntX8mUF7AvjVcyT6DZwnOQXRvYgel9pCirCGKKETO5R70HEZ8UdbQA0Nt8"
OUTPUT = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\BASE_PREST_API_FRESH.xlsx"

ITAU_PMID = "627401"
MAX_WORKERS = 1

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
money_fmt = 'R$ #,##0.00;[Red]-R$ #,##0.00'
date_fmt = 'DD/MM/YYYY'

headers = {
    "Authorization": API_KEY,
    "Accept": "application/json",
}

# Step 1: Fetch all reports
print("Fetching all reports from VExpenses API...", flush=True)
t0 = time.time()
for attempt in range(5):
    try:
        print(f"  Attempt {attempt+1}...", flush=True)
        resp = requests.get(f"{API_URL}/v2/reports?include=user", headers=headers, timeout=60)
        print(f"  Status: {resp.status_code} ({time.time()-t0:.1f}s)", flush=True)
        if resp.status_code == 429:
            wait = 30 * (attempt + 1)
            print(f"  429 rate limited, waiting {wait}s...", flush=True)
            time.sleep(wait)
            continue
        resp.raise_for_status()
        data = resp.json()
        all_reports = data.get("data", [])
        print(f"  Got {len(all_reports)} reports in {time.time()-t0:.1f}s", flush=True)
        break
    except Exception as e:
        print(f"  Error: {e} ({time.time()-t0:.1f}s)", flush=True)
        if attempt < 4:
            time.sleep(15)
            continue
        raise
else:
    print("  FAILED to fetch reports after 5 attempts", flush=True)
    exit(1)

# Extract user info from nested structure
def get_user_info(report):
    user = report.get("user", {})
    user_data = user.get("data", {}) if isinstance(user, dict) else {}
    return user_data.get("cpf", ""), user_data.get("name", "")

# Filter: APROVADO/ENVIADO, no FATURA/CARTAO/CARTÃO
def is_fatura(name):
    n = (name or "").upper()
    return "FATURA" in n or "CARTAO" in n or "CARTÃO" in n

def is_aprovado(status):
    s = (status or "").upper()
    return "APROVADO" in s or "ENVIADO" in s

filtered = []
for r in all_reports:
    cpf, nome = get_user_info(r)
    if not cpf:
        continue
    if not is_aprovado(r.get("status")):
        continue
    if is_fatura(r.get("description")):
        continue
    r["_cpf"] = str(cpf).strip()
    r["_nome"] = nome
    r["_name"] = r.get("description") or ""
    filtered.append(r)

print(f"  After filter (APROVADO/ENVIADO, no FATURA/CARTAO): {len(filtered)} reports", flush=True)

# Step 2: Fetch expenses concurrently
print(f"\nFetching expenses for {len(filtered)} reports ({MAX_WORKERS} workers)...", flush=True)
expenses_all = []
errors = []
lock = threading.Lock()
done_count = 0
t1 = time.time()

def fetch_expenses(report):
    global done_count
    rid = report["id"]
    for attempt in range(3):
        try:
            eresp = requests.get(
                f"{API_URL}/v2/reports/{rid}?include=expenses",
                headers=headers,
                timeout=30
            )
            if eresp.status_code == 429:
                time.sleep(3 * (attempt + 1))
                continue
            if eresp.status_code == 404:
                with lock:
                    errors.append(f"Report {rid}: 404 not found")
                return []
            eresp.raise_for_status()
            edata = eresp.json()
            report_expenses = edata.get("data", {}).get("expenses", {}).get("data", [])

            result = []
            for e in report_expenses:
                pmid = str(e.get("payment_method_id", ""))
                if pmid == ITAU_PMID:
                    continue
                result.append({
                    "report_id": rid,
                    "report_name": report.get("_name", ""),
                    "status": report.get("status", ""),
                    "cpf": report.get("_cpf", ""),
                    "nome": report.get("_nome", ""),
                    "approval_date": report.get("approval_date", ""),
                    "report_total": None,
                    "expense_id": e.get("id"),
                    "description": e.get("title") or "",
                    "value": float(e.get("value") or 0),
                    "date": e.get("date"),
                    "expense_status": e.get("validate", ""),
                    "payment_method_id": pmid,
                    "payment_method_name": "",
                })
            return result
        except Exception as ex:
            if attempt < 2:
                time.sleep(2 * (attempt + 1))
                continue
            with lock:
                errors.append(f"Report {rid}: {str(ex)[:100]}")
            return []
    with lock:
        errors.append(f"Report {rid}: max retries exceeded (429)")
    return []

with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
    futures = {executor.submit(fetch_expenses, r): r for r in filtered}
    for future in as_completed(futures):
        result = future.result()
        with lock:
            expenses_all.extend(result)
            done_count += 1
            if done_count % 200 == 0 or done_count == len(filtered):
                elapsed = time.time() - t1
                pct = done_count / len(filtered) * 100
                eta = elapsed / done_count * (len(filtered) - done_count)
                print(f"  [{done_count}/{len(filtered)}] {pct:.0f}% | {len(expenses_all)} expenses | {elapsed:.0f}s elapsed, ETA {eta:.0f}s | errors={len(errors)}", flush=True)

print(f"\nDone! {len(expenses_all)} expenses fetched in {time.time()-t1:.1f}s")
if errors:
    print(f"Errors: {len(errors)}")
    for e in errors[:10]:
        print(f"  {e}")

# Step 3: Build Excel
print("\nBuilding Excel...")

# CPF summary
cpf_summary = defaultdict(lambda: {"nome": "", "reports": set(), "count": 0, "total": 0.0})
for e in expenses_all:
    cpf = e["cpf"]
    cpf_summary[cpf]["nome"] = e["nome"]
    cpf_summary[cpf]["reports"].add(e["report_id"])
    cpf_summary[cpf]["count"] += 1
    cpf_summary[cpf]["total"] += e["value"]

# Report summary
report_summary = defaultdict(lambda: {"name": "", "status": "", "cpf": "", "nome": "", "approval_date": "", "count": 0, "total": 0.0})
for e in expenses_all:
    rid = e["report_id"]
    report_summary[rid]["name"] = e["report_name"]
    report_summary[rid]["status"] = e["status"]
    report_summary[rid]["cpf"] = e["cpf"]
    report_summary[rid]["nome"] = e["nome"]
    report_summary[rid]["approval_date"] = e["approval_date"]
    report_summary[rid]["count"] += 1
    report_summary[rid]["total"] += e["value"]

wb = Workbook()

# Sheet 1: CPF Summary
ws1 = wb.active
ws1.title = "CPF Summary"
h1 = ["CPF", "Nome", "Qtd Relatórios", "Qtd Despesas", "Total Prestação"]
for col, h in enumerate(h1, 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill
    c.alignment = Alignment(horizontal='center'); c.border = thin_border
sorted_cpfs = sorted(cpf_summary.keys(), key=lambda c: cpf_summary[c]["nome"])
for i, cpf in enumerate(sorted_cpfs, 2):
    d = cpf_summary[cpf]
    vals = [cpf, d["nome"], len(d["reports"]), d["count"], d["total"]]
    for col, v in enumerate(vals, 1):
        c = ws1.cell(row=i, column=col, value=v)
        c.border = thin_border
        if col == 5: c.number_format = money_fmt
for col, w in enumerate([15, 35, 15, 15, 18], 1):
    ws1.column_dimensions[get_column_letter(col)].width = w
ws1.auto_filter.ref = f"A1:E{len(sorted_cpfs)+1}"
ws1.freeze_panes = "A2"

# Sheet 2: Report Summary
ws2 = wb.create_sheet("Report Summary")
h2 = ["Report ID", "Report Name", "Status", "CPF", "Nome", "Approval Date", "Qtd Despesas", "Total Valor"]
for col, h in enumerate(h2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill
    c.alignment = Alignment(horizontal='center'); c.border = thin_border
sorted_reports = sorted(report_summary.keys(), key=lambda r: (report_summary[r]["cpf"], r))
for i, rid in enumerate(sorted_reports, 2):
    d = report_summary[rid]
    vals = [rid, d["name"], d["status"], d["cpf"], d["nome"], d["approval_date"], d["count"], d["total"]]
    for col, v in enumerate(vals, 1):
        c = ws2.cell(row=i, column=col, value=v)
        c.border = thin_border
        if col == 8: c.number_format = money_fmt
for col, w in enumerate([12, 25, 12, 15, 30, 22, 12, 15], 1):
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.auto_filter.ref = f"A1:H{len(sorted_reports)+1}"
ws2.freeze_panes = "A2"

# Sheet 3: All Expenses (BASE PREST)
ws3 = wb.create_sheet("BASE PREST (API)")
h3 = ["Report ID", "Report Name", "Status", "CPF", "Nome",
      "Approval Date", "Report Total",
      "Expense ID", "Expense Description", "Expense Value", "Expense Date", "Expense Status",
      "Payment Method ID", "Payment Method Name"]
for col, h in enumerate(h3, 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill
    c.alignment = Alignment(horizontal='center'); c.border = thin_border

for i, e in enumerate(expenses_all, 2):
    vals = [e["report_id"], e["report_name"], e["status"], e["cpf"], e["nome"],
            e["approval_date"], e["report_total"],
            e["expense_id"], e["description"], e["value"], e["date"], e["expense_status"],
            e["payment_method_id"], e["payment_method_name"]]
    for col, v in enumerate(vals, 1):
        c = ws3.cell(row=i, column=col, value=v)
        c.border = thin_border
        if col == 10: c.number_format = money_fmt
        if col == 11 and v: c.number_format = date_fmt

for col, w in enumerate([12, 25, 12, 15, 30, 22, 15, 12, 30, 15, 14, 15, 15, 20], 1):
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.auto_filter.ref = f"A1:N{len(expenses_all)+1}"
ws3.freeze_panes = "A2"

# Save
print(f"Saving to {OUTPUT}...")
wb.save(OUTPUT)
total_val = sum(e["value"] for e in expenses_all)
print(f"Done! {len(expenses_all)} expenses, {len(report_summary)} reports, {len(cpf_summary)} CPFs, Total: R$ {total_val:,.2f}")
