#!/usr/bin/env python3
"""
Script final para analisar a planilha Excel
"""

import pandas as pd
import openpyxl
import os
import sys
import traceback

output_file = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\output_analise.txt"
file_path = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\controle-api\data\CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsx"

def log(msg):
    msg_str = str(msg)
    print(msg_str)
    with open(output_file, 'a', encoding='utf-8') as f:
        f.write(msg_str + '\n')

# Limpar arquivo anterior
with open(output_file, 'w', encoding='utf-8') as f:
    f.write('')

try:
    log("=" * 80)
    log("ANALISE DA PLANILHA - VEXPENSES MAIO 2026")
    log("=" * 80)

    log(f"\nArquivo: {os.path.basename(file_path)}")
    log(f"Existe: {os.path.exists(file_path)}")

    if not os.path.exists(file_path):
        log("ERRO: Arquivo nao existe!")
        sys.exit(1)

    # 1. Listar abas
    log("\n" + "=" * 40)
    log("1. ABAS DISPONIVEIS")
    log("=" * 40)

    xl = pd.ExcelFile(file_path)
    for i, name in enumerate(xl.sheet_names, 1):
        log(f"  {i}. {name}")

    log(f"\nTotal: {len(xl.sheet_names)} abas")

    # 2. Para cada aba, detectar cabeçalho
    log("\n" + "=" * 40)
    log("2. CABEÇALHO POR ABA")
    log("=" * 40)

    for sheet in xl.sheet_names:
        try:
            df = pd.read_excel(file_path, sheet_name=sheet, nrows=5, header=None)
            if df.empty:
                log(f"\n  [ABA] {sheet}: (vazia)")
                continue

            log(f"\n  [ABA] {sheet}:")
            for i in range(min(5, len(df))):
                vals = [str(v)[:25] for v in df.iloc[i] if pd.notna(v)][:6]
                marker = " <-- CABECALHO?" if i < 3 and len(vals) >= 2 else ""
                log(f"     Linha {i+1}: {vals}{marker}")
        except Exception as e2:
            log(f"\n  [ABA] {sheet}: ERRO - {e2}")

    # 3. Analise da aba PAINEL
    if 'PAINEL' in xl.sheet_names:
        log("\n" + "=" * 60)
        log("3. ABA PAINEL - COLUNAS")
        log("=" * 60)

        df_painel = pd.read_excel(file_path, sheet_name='PAINEL', nrows=30)
        log(f"\nTotal de colunas detectadas: {len(df_painel.columns)}")
        log(f"Primeira coluna: {df_painel.columns[0]}")
        log("\nLista de colunas (indice: nome):")
        log("-" * 50)
        for i, col in enumerate(df_painel.columns, 1):
            log(f"  {i:2d}. {col}")

        # 4. Verificar colunas de SALDO
        log("\n" + "=" * 60)
        log("4. ANALISE DE COLUNAS DE SALDO - FORMULAS VS VALORES")
        log("=" * 60)

        saldo_cols = [c for c in df_painel.columns if 'SALDO' in str(c).upper()]
        log(f"\nColunas com 'SALDO' no nome: {saldo_cols}")

        # Carregar com openpyxl para ver formulas
        log("\nCarregando com openpyxl para verificar formulas...")
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb['PAINEL']

        log(f"Dimensoes da aba: {ws.max_row} linhas x {ws.max_column} colunas")

        # Encontrar cabeçalho
        header_row = 1
        for r in range(1, min(15, ws.max_row + 1)):
            cell_val = ws.cell(row=r, column=1).value
            if cell_val and str(cell_val).strip() == str(df_painel.columns[0]).strip():
                header_row = r
                break

        log(f"Cabecalho detectado na linha: {header_row}")

        # Mapear indices das colunas de SALDO
        col_mapping = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=header_row, column=col_idx).value
            if cell_val:
                for saldo_col in saldo_cols:
                    if str(cell_val).strip() == str(saldo_col).strip():
                        col_mapping[saldo_col] = col_idx

        log(f"\nMapeamento de colunas: {col_mapping}")

        data_start = header_row + 1

        for col_name, col_idx in col_mapping.items():
            log(f"\n[COLUNA] '{col_name}' (coluna {col_idx})")
            log("-" * 40)

            formulas = 0
            values = 0
            empty = 0
            sample_data = []
            all_numeric_values = []

            for row_idx in range(data_start, min(data_start + 30, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)

                if cell.data_type == 'f':  # formula
                    formulas += 1
                    if len(sample_data) < 2:
                        sample_data.append(f"FORMULA L{row_idx}: {cell.value}")
                elif cell.value is None:
                    empty += 1
                else:
                    values += 1
                    if len(sample_data) < 2:
                        sample_data.append(f"VALOR   L{row_idx}: {cell.value} (tipo: {type(cell.value).__name__})")

                    if isinstance(cell.value, (int, float)):
                        all_numeric_values.append(cell.value)

            log(f"  Formulas detectadas: {formulas}")
            log(f"  Valores estaticos: {values}")
            log(f"  Celulas vazias: {empty}")
            for s in sample_data:
                log(f"    {s}")

            if formulas > 0:
                log(f"  >>> RESULTADO: Contem FORMULAS (calculadas automaticamente)")
            elif values > 0:
                log(f"  >>> RESULTADO: VALORES ESTATICOS (inseridos manualmente)")
            else:
                log(f"  >>> RESULTADO: Sem dados nas primeiras linhas")

            # Estatisticas dos valores
            if all_numeric_values:
                log(f"\n  Estatisticas dos valores:")
                log(f"    Total valores: {len(all_numeric_values)}")
                log(f"    Min: {min(all_numeric_values):,.2f}")
                log(f"    Max: {max(all_numeric_values):,.2f}")
                log(f"    Media: {sum(all_numeric_values)/len(all_numeric_values):,.2f}")
                log(f"    Ultimos 5 valores: {[f'{v:,.2f}' for v in all_numeric_values[-5:]]}")

                zeros = sum(1 for v in all_numeric_values if v == 0)
                if zeros == len(all_numeric_values):
                    log(f"    !!! ALERTA: TODOS os valores sao ZERO !!!")
                elif zeros > 0:
                    log(f"    Aviso: {zeros} valores sao zero ({zeros/len(all_numeric_values)*100:.1f}%)")

        wb.close()

    log("\n" + "=" * 80)
    log("ANALISE CONCLUIDA COM SUCESSO")
    log("=" * 80)

except Exception as e:
    log(f"\nERRO: {e}")
    log(traceback.format_exc())
    log("\nANALISE FALHOU")

log(f"\nOutput salvo em: {output_file}")
