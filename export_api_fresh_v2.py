import requests
import json
import time
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

API_URL = "https://api.vexpenses.com"
API_KEY = "N2ntX8mUF7AvjVcyT6DZwnOQXRvYgel9pCirCGKKETO5R70HEZ8UdbQA0Nt8"
OUTPUT = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\BASE_PREST_API_FRESH_V2.xlsx"

ITAU_PMID = "627401"
CHUNK_DAYS = 7

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
money_fmt = 'R$ #,##0.00;[Red]-R$ #,##0.00'
date_fmt = 'DD/MM/YYYY'

api_headers = {"Authorization": API_KEY, "Accept": "application/json"}

# --- Step 1: Reports ---
print("[1/3] Fetching reports list...", flush=True)
t0 = time.time()
for attempt in range(5):
    try:
        resp = requests.get(f"{API_URL}/v2/reports?include=user", headers=api_headers, timeout=120)
        if resp.status_code == 429:
            print(f"  429, waiting {30*(attempt+1)}s", flush=True)
            time.sleep(30*(attempt+1))
            continue
        resp.raise_for_status()
        data = resp.json()
        all_reports = data.get("data", [])
        break
    except Exception as e:
        print(f"  Error: {e}", flush=True)
        time.sleep(10)
print(f"  Got {len(all_reports)} reports in {time.time()-t0:.1f}s", flush=True)

# Extract user info from nested
def get_user_info(report):
    user = report.get("user", {})
    user_data = user.get("data", {}) if isinstance(user, dict) else {}
    return user_data.get("cpf", ""), user_data.get("name", "")

def is_fatura(name):
    n = (name or "").upper()
    return "FATURA" in n or "CARTAO" in n or "CARTÃO" in n

def is_aprovado(status):
    s = (status or "").upper()
    return "APROVADO" in s or "ENVIADO" in s

# Build report index
reports_by_id = {}
for r in all_reports:
    cpf, nome = get_user_info(r)
    if not cpf:
        continue
    if not is_aprovado(r.get("status")):
        continue
    if is_fatura(r.get("description")):
        continue
    rid = r["id"]
    reports_by_id[rid] = {
        "id": rid,
        "name": r.get("description") or "",
        "status": r.get("status") or "",
        "cpf": str(cpf).strip(),
        "nome": nome,
        "approval_date": r.get("approval_date", ""),
    }

print(f"  Filtered {len(reports_by_id)} APROVADO/ENVIADO non-FATURA reports", flush=True)

# --- Step 2: Expenses by date chunks ---
def fetch_expenses_chunk(start, end):
    all_e = []
    page = 1
    per_page = 100
    t_chunk = time.time()
    while True:
        params = {
            "search": f"date:{start},{end}",
            "searchFields": "date:between",
            "paginate": "true",
            "page": str(page),
            "per_page": str(per_page),
        }
        for attempt in range(3):
            try:
                print(f"    >> {start}..{end} p{page} a{attempt+1}...", end=" ", flush=True)
                r = requests.get(f"{API_URL}/v2/expenses", headers=api_headers, params=params, timeout=(10, 30))
                print(f"{r.status_code} ({time.time()-t_chunk:.1f}s)", flush=True)
                if r.status_code == 429:
                    print(f"    429, wait {5*(attempt+1)}s", flush=True)
                    time.sleep(5*(attempt+1))
                    continue
                if r.status_code == 500:
                    print(f"    500, skipping chunk", flush=True)
                    return all_e
                r.raise_for_status()
                data = r.json()
                expenses = data.get("data", [])
                meta = data.get("meta", {})
                total = meta.get("total", 0)
                last_page = meta.get("last_page", 0)
                if not expenses:
                    print(f"    DONE {start}..{end}: {len(all_e)} expenses in {time.time()-t_chunk:.1f}s", flush=True)
                    return all_e
                all_e.extend(expenses)
                print(f"    [{len(all_e)}/{total}]", end=" ", flush=True)
                if last_page and page >= last_page:
                    print(f"    DONE {start}..{end}: {len(all_e)} expenses in {time.time()-t_chunk:.1f}s", flush=True)
                    return all_e
                if not last_page and len(expenses) < per_page:
                    print(f"    DONE {start}..{end}: {len(all_e)} expenses in {time.time()-t_chunk:.1f}s", flush=True)
                    return all_e
                page += 1
                break
            except requests.exceptions.Timeout:
                print(f"TIMEOUT", flush=True)
                if attempt < 2:
                    time.sleep(3*(attempt+1))
                    continue
                print(f"    GIVING UP on {start}..{end} p{page}", flush=True)
                return all_e
            except Exception as e:
                print(f"ERR: {e}", flush=True)
                if attempt < 2:
                    time.sleep(3*(attempt+1))
                    continue
                return all_e
    return all_e

# Generate 15-day chunks from 2025-01-01 to 2026-12-31
start_date = datetime(2025, 1, 1)
end_date = datetime(2026, 12, 31)
chunks = []
current = start_date
while current <= end_date:
    chunk_end = min(current + timedelta(days=CHUNK_DAYS-1), end_date)
    chunks.append((current.strftime("%Y-%m-%d"), chunk_end.strftime("%Y-%m-%d")))
    current = chunk_end + timedelta(days=1)

print(f"[2/3] Fetching expenses in {len(chunks)} chunks of {CHUNK_DAYS} days...", flush=True)
t1 = time.time()
all_expenses = []
for i, (s, e) in enumerate(chunks):
    chunk = fetch_expenses_chunk(s, e)
    all_expenses.extend(chunk)
    if (i + 1) % 5 == 0 or i == len(chunks) - 1:
        elapsed = time.time() - t1
        pct = (i + 1) / len(chunks) * 100
        eta = elapsed / (i + 1) * (len(chunks) - i - 1)
        print(f"  [{i+1}/{len(chunks)}] {pct:.0f}% | {len(all_expenses)} expenses | {elapsed:.0f}s elapsed, ETA {eta:.0f}s", flush=True)

print(f"  Total expenses fetched: {len(all_expenses)} in {time.time()-t1:.1f}s", flush=True)

# --- Step 3: Filter and build output ---
expenses_all = []
for e in all_expenses:
    pmid = str(e.get("payment_method_id", ""))
    if pmid == ITAU_PMID:
        continue
    rid = e.get("expense_id")
    if not rid:
        continue
    report = reports_by_id.get(rid)
    if not report:
        continue
    expenses_all.append({
        "report_id": rid,
        "report_name": report["name"],
        "status": report["status"],
        "cpf": report["cpf"],
        "nome": report["nome"],
        "approval_date": report["approval_date"],
        "report_total": None,
        "expense_id": e.get("id"),
        "description": e.get("title") or "",
        "value": float(e.get("value") or 0),
        "date": e.get("date"),
        "expense_status": e.get("validate", ""),
        "payment_method_id": pmid,
        "payment_method_name": "",
    })

print(f"[3/3] Matched {len(expenses_all)} expenses to reports (excluding Itau)", flush=True)

# Sort
expenses_all.sort(key=lambda x: (x["cpf"], x["report_id"], x["expense_id"] or 0))

# Build summaries
cpf_summary = defaultdict(lambda: {"nome": "", "reports": set(), "count": 0, "total": 0.0})
report_summary = defaultdict(lambda: {"name": "", "status": "", "cpf": "", "nome": "", "approval_date": "", "count": 0, "total": 0.0})

for e in expenses_all:
    cpf = e["cpf"]
    cpf_summary[cpf]["nome"] = e["nome"]
    cpf_summary[cpf]["reports"].add(e["report_id"])
    cpf_summary[cpf]["count"] += 1
    cpf_summary[cpf]["total"] += e["value"]

    rid = e["report_id"]
    rs = report_summary[rid]
    rs["name"] = e["report_name"]
    rs["status"] = e["status"]
    rs["cpf"] = e["cpf"]
    rs["nome"] = e["nome"]
    rs["approval_date"] = e["approval_date"]
    rs["count"] += 1
    rs["total"] += e["value"]

# Write Excel
wb = Workbook()

# CPF Summary
ws1 = wb.active
ws1.title = "CPF Summary"
h1 = ["CPF", "Nome", "Qtd Relatórios", "Qtd Despesas", "Total Prestação"]
for col, h in enumerate(h1, 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill
    c.alignment = Alignment(horizontal='center'); c.border = thin_border
sorted_cpfs = sorted(cpf_summary.keys(), key=lambda c: cpf_summary[c]["nome"])
for i, cpf in enumerate(sorted_cpfs, 2):
    d = cpf_summary[cpf]
    vals = [cpf, d["nome"], len(d["reports"]), d["count"], d["total"]]
    for col, v in enumerate(vals, 1):
        c = ws1.cell(row=i, column=col, value=v)
        c.border = thin_border
        if col == 5: c.number_format = money_fmt
for col, w in enumerate([15, 35, 15, 15, 18], 1):
    ws1.column_dimensions[get_column_letter(col)].width = w
ws1.auto_filter.ref = f"A1:E{len(sorted_cpfs)+1}"
ws1.freeze_panes = "A2"

# Report Summary
ws2 = wb.create_sheet("Report Summary")
h2 = ["Report ID", "Report Name", "Status", "CPF", "Nome", "Approval Date", "Qtd Despesas", "Total Valor"]
for col, h in enumerate(h2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill
    c.alignment = Alignment(horizontal='center'); c.border = thin_border
sorted_reports = sorted(report_summary.keys(), key=lambda r: (report_summary[r]["cpf"], r))
for i, rid in enumerate(sorted_reports, 2):
    d = report_summary[rid]
    vals = [rid, d["name"], d["status"], d["cpf"], d["nome"], d["approval_date"], d["count"], d["total"]]
    for col, v in enumerate(vals, 1):
        c = ws2.cell(row=i, column=col, value=v)
        c.border = thin_border
        if col == 8: c.number_format = money_fmt
for col, w in enumerate([12, 25, 12, 15, 30, 22, 12, 15], 1):
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.auto_filter.ref = f"A1:H{len(sorted_reports)+1}"
ws2.freeze_panes = "A2"

# All Expenses
ws3 = wb.create_sheet("BASE PREST (API)")
h3 = ["Report ID", "Report Name", "Status", "CPF", "Nome",
      "Approval Date", "Report Total",
      "Expense ID", "Expense Description", "Expense Value", "Expense Date", "Expense Status",
      "Payment Method ID", "Payment Method Name"]
for col, h in enumerate(h3, 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill
    c.alignment = Alignment(horizontal='center'); c.border = thin_border
for i, e in enumerate(expenses_all, 2):
    vals = [e["report_id"], e["report_name"], e["status"], e["cpf"], e["nome"],
            e["approval_date"], e["report_total"],
            e["expense_id"], e["description"], e["value"], e["date"], e["expense_status"],
            e["payment_method_id"], e["payment_method_name"]]
    for col, v in enumerate(vals, 1):
        c = ws3.cell(row=i, column=col, value=v)
        c.border = thin_border
        if col == 10: c.number_format = money_fmt
        if col == 11 and v: c.number_format = date_fmt
for col, w in enumerate([12, 25, 12, 15, 30, 22, 15, 12, 30, 15, 14, 15, 15, 20], 1):
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.auto_filter.ref = f"A1:N{len(expenses_all)+1}"
ws3.freeze_panes = "A2"

print(f"Saving to {OUTPUT}...", flush=True)
wb.save(OUTPUT)
total_val = sum(e["value"] for e in expenses_all)
print(f"DONE! {len(expenses_all)} expenses, {len(report_summary)} reports, {len(cpf_summary)} CPFs, Total: R$ {total_val:,.2f}", flush=True)
