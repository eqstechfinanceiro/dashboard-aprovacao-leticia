"""Compare API-only calculations against CARGA 2QZ JUNHO sheet."""
import openpyxl
import urllib.request
import json

# 1. Read the CARGA 2 QZ JUNHO sheet
wb = openpyxl.load_workbook(r'data\06 - JUNHO\CARGA 2 QZ JUNHO 26 VEXPENSES EQS.xlsx', data_only=True)
ws = wb.active

sheet_data = {}
for row in range(3, ws.max_row + 1):
    cpf = ws.cell(row, 3).value
    if not cpf:
        continue
    cpf = str(cpf).strip()
    sheet_data[cpf] = {
        'colaborador': ws.cell(row, 2).value,
        'saldo_final_sheet': float(ws.cell(row, 7).value) if ws.cell(row, 7).value is not None else 0,
        'col_qz_sheet': float(ws.cell(row, 8).value) if ws.cell(row, 8).value is not None else 0,
        'carga_final_sheet': float(ws.cell(row, 9).value) if ws.cell(row, 9).value is not None else 0,
    }

print(f"Sheet users: {len(sheet_data)}")

# 2. Load API data from JSON file (saved from browser)
with open('api_2qz_junho.json', 'r', encoding='utf-8') as f:
    api_data = json.loads(f.read())

print(f"API users: {len(api_data)}")

# 3. Compare
matches = 0
mismatches = []
exact_carga = 0
close_carga = 0
total_compared = 0

for api_user in api_data:
    cpf = api_user['c']
    if cpf not in sheet_data:
        continue

    total_compared += 1
    sheet = sheet_data[cpf]

    sf_api = round(api_user['sf'], 2)
    sf_sheet = round(sheet['saldo_final_sheet'], 2)
    cf_api = round(api_user['cf'], 2)
    cf_sheet = round(sheet['carga_final_sheet'], 2)

    sf_diff = round(sf_api - sf_sheet, 2)
    cf_diff = round(cf_api - cf_sheet, 2)

    if abs(cf_diff) < 0.01:
        exact_carga += 1
    elif abs(cf_diff) < 10:
        close_carga += 1

    if abs(sf_diff) < 0.01 and abs(cf_diff) < 0.01:
        matches += 1
    else:
        mismatches.append({
            'cpf': cpf,
            'colaborador': api_user['n'],
            'sf_api': sf_api,
            'sf_sheet': sf_sheet,
            'sf_diff': sf_diff,
            'cf_api': cf_api,
            'cf_sheet': cf_sheet,
            'cf_diff': cf_diff,
        })

print(f"\n=== COMPARISON RESULTS (2QZ JUNHO: API vs SHEET) ===")
print(f"Total compared: {total_compared}")
print(f"Exact matches (SF + CF): {matches} ({matches/total_compared*100:.1f}%)")
print(f"Exact carga_final: {exact_carga} ({exact_carga/total_compared*100:.1f}%)")
print(f"Close carga_final (<R$10): {close_carga} ({close_carga/total_compared*100:.1f}%)")
print(f"Mismatches: {len(mismatches)}")

mismatches.sort(key=lambda x: abs(x['cf_diff']), reverse=True)
print(f"\n=== TOP 20 MISMATCHES BY CARGA_FINAL DIFF ===")
for m in mismatches[:20]:
    name = m['colaborador'][:30] if m['colaborador'] else '???'
    print(f"  {name:30s} | CPF={m['cpf']} | SF: API={m['sf_api']:>12.2f} sheet={m['sf_sheet']:>12.2f} diff={m['sf_diff']:>10.2f} | CF: API={m['cf_api']:>10.2f} sheet={m['cf_sheet']:>10.2f} diff={m['cf_diff']:>10.2f}")

sf_diffs = [m['sf_diff'] for m in mismatches]
print(f"\n=== SALDO_FINAL DIFF SUMMARY ===")
print(f"  Max diff: R$ {max(sf_diffs):.2f}")
print(f"  Min diff: R$ {min(sf_diffs):.2f}")
print(f"  Avg abs diff: R$ {sum(abs(d) for d in sf_diffs)/len(sf_diffs):.2f}")

big_sf = [m for m in mismatches if abs(m['sf_diff']) > 100]
small_sf = [m for m in mismatches if abs(m['sf_diff']) <= 100 and abs(m['sf_diff']) > 0.01]
print(f"\n=== SALDO_FINAL MISMATCH CATEGORIES ===")
print(f"  Big diff (>R$100): {len(big_sf)} users")
print(f"  Small diff (<=R$100): {len(small_sf)} users")
pos = sum(1 for m in big_sf if m['sf_diff'] > 0)
neg = sum(1 for m in big_sf if m['sf_diff'] < 0)
print(f"  Big diffs: API higher={pos}, sheet higher={neg}")
