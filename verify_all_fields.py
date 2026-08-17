import openpyxl
import json
import subprocess
from collections import defaultdict

# Load CARGA 1 QZ sheet (the quinzena base)
wb = openpyxl.load_workbook('CARGA 1 QZ AGOSTO 26 VEXPENSES EQS.xlsx', read_only=True, data_only=True)
ws = wb.active

# Read headers to find column positions
headers = {}
for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
    for i, val in enumerate(row):
        if val and str(val).strip():
            key = str(val).strip().upper()
            if key not in headers:
                headers[key] = i

print("CARGA sheet headers found:")
for k, v in headers.items():
    print(f"  col {v}: {k}")

# Find the data rows (usually after header row)
# Let's scan for rows that have CPF values
carga_data = {}
data_start = False
for row in ws.iter_rows(min_row=1, values_only=True):
    # Check if this row has a CPF-like value
    cpf_col = headers.get('CPF', headers.get('Cpf', None))
    if cpf_col is not None and row[cpf_col]:
        cpf_val = str(row[cpf_col]).strip()
        if cpf_val and cpf_val != 'None' and len(cpf_val) >= 11:
            data_start = True
            carga_data[cpf_val] = {}
            for key, col in headers.items():
                val = row[col] if col < len(row) else None
                if val is not None and str(val).strip() and str(val).strip() != 'None':
                    try:
                        carga_data[cpf_val][key] = float(val)
                    except (ValueError, TypeError):
                        carga_data[cpf_val][key] = str(val).strip()

print(f"\nTotal CPFs in CARGA sheet: {len(carga_data)}")
if carga_data:
    first_cpf = list(carga_data.keys())[0]
    print(f"Sample ({first_cpf}): {carga_data[first_cpf]}")

# Now fetch quinzena-complete from API
ps_script = '''
$session = New-Object Microsoft.PowerShell.Commands.WebRequestSession
$body = @{ email = "italo.medrado@eqsengenharia.com.br"; password = "EQSeng4292@" } | ConvertTo-Json
Invoke-RestMethod -Uri "http://localhost:3000/api/auth/login" -Method POST -Body $body -ContentType "application/json" -WebSession $session | Out-Null
$r = Invoke-RestMethod -Uri "http://localhost:3000/api/quinzena-complete?year=2026&month=8&quinzena=1" -Method GET -WebSession $session
$r | ConvertTo-Json -Depth 10
'''
result = subprocess.run(['powershell.exe', '-NoProfile', '-Command', ps_script], capture_output=True, text=True, timeout=60)
try:
    api_data = json.loads(result.stdout)
except:
    print(f"ERROR parsing API response: {result.stderr[:300]}")
    print(f"STDOUT: {result.stdout[:500]}")
    exit(1)

api_rows = api_data.get('rows', [])
print(f"\nAPI returned {len(api_rows)} rows")

# Map API data by CPF
api_by_cpf = {}
for r in api_rows:
    cpf = r.get('cpf', '').strip()
    if cpf:
        api_by_cpf[cpf] = r

# Compare fields
fields_to_check = ['carga', 'transferencia', 'tarifa', 'prestacao', 'saldo_cartao', 'saldo_final']

print(f"\n{'='*100}")
print("FIELD-BY-FIELD COMPARISON: CARGA SHEET vs API")
print(f"{'='*100}")

for field in fields_to_check:
    # Find the matching column name in CARGA sheet (case-insensitive)
    carga_field = None
    for k in headers:
        if k.upper() == field.upper() or k.upper().replace('_', ' ') == field.upper().replace('_', ' '):
            carga_field = k
            break
    if not carga_field:
        # Try partial match
        for k in headers:
            if field.upper() in k.upper():
                carga_field = k
                break

    if not carga_field:
        print(f"\n{field}: NOT FOUND in CARGA sheet")
        continue

    total_carga = 0
    total_api = 0
    diffs = []
    match_count = 0

    for cpf, carga_row in carga_data.items():
        carga_val = carga_row.get(carga_field, 0)
        if isinstance(carga_val, str):
            try:
                carga_val = float(carga_val.replace(',', '.').replace('R$', '').strip())
            except:
                carga_val = 0

        api_val = api_by_cpf.get(cpf, {}).get(field, 0)
        if api_val is None:
            api_val = 0
        api_val = float(api_val)

        total_carga += carga_val
        total_api += api_val

        diff = carga_val - api_val
        if abs(diff) > 0.01:
            nome = api_by_cpf.get(cpf, {}).get('colaborador', carga_row.get('COLABORADOR', ''))
            diffs.append((cpf, nome, carga_val, api_val, diff))
        else:
            match_count += 1

    print(f"\n{field} (CARGA col='{carga_field}'):")
    print(f"  CARGA total: {total_carga:.2f}")
    print(f"  API total:   {total_api:.2f}")
    print(f"  Diff:        {total_carga - total_api:+.2f}")
    print(f"  Matched:     {match_count}/{len(carga_data)}")
    if diffs:
        diffs.sort(key=lambda x: abs(x[4]), reverse=True)
        print(f"  Top 10 diffs:")
        for cpf, nome, cv, av, d in diffs[:10]:
            print(f"    {str(nome)[:30]:30s} cpf={cpf} CARGA={cv:.2f} API={av:.2f} diff={d:+.2f}")
