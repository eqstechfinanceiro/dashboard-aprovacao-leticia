#!/usr/bin/env python3
"""
Analise especifica da aba PAINEL
"""

import pandas as pd
import openpyxl
import os
import traceback

output_file = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\output_painel.txt"
file_path = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\controle-api\data\CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsx"

def log(msg):
    msg_str = str(msg)
    print(msg_str)
    with open(output_file, 'a', encoding='utf-8') as f:
        f.write(msg_str + '\n')

# Limpar arquivo
with open(output_file, 'w', encoding='utf-8') as f:
    f.write('')

try:
    log("=" * 80)
    log("ANALISE ESPECIFICA DA ABA PAINEL")
    log("=" * 80)

    # Ler sem cabeçalho para ver estrutura
    log("\n--- Lendo sem cabeçalho (primeiras 15 linhas) ---")
    df_raw = pd.read_excel(file_path, sheet_name='PAINEL', nrows=15, header=None)

    for i in range(len(df_raw)):
        row_vals = []
        for j in range(min(10, len(df_raw.columns))):
            val = df_raw.iloc[i, j]
            if pd.notna(val):
                row_vals.append(f"[{j}]{str(val)[:30]}")
        log(f"Linha {i}: {row_vals}")

    # Tentar detectar onde esta o cabeçalho
    header_row_idx = None
    for i in range(len(df_raw)):
        row_texts = []
        for j in range(len(df_raw.columns)):
            val = df_raw.iloc[i, j]
            if pd.notna(val) and isinstance(val, str):
                row_texts.append(val.upper())

        # Verificar se tem palavras-chave de cabeçalho
        keywords = ['COLABORADOR', 'NOME', 'SALDO', 'CARTAO', 'PRESTACAO', 'CPF', 'STATUS']
        matches = sum(1 for k in keywords if any(k in t for t in row_texts))
is
        if matches >= 2:
            header_row_idx = i
            log(f"\n>>> CABECALHO DETECTADO NA LINHA {i} (indice {i}, linha Excel {i+1})")
            log(f"    Palavras-chave encontradas: {matches}")
            break

    if header_row_idx is None:
        header_row_idx = 0
        log(f"\n>>> Usando linha 0 como cabeçalho padrao")

    # Ler novamente com o cabeçalho correto
    log(f"\n--- Lendo com cabeçalho na linha {header_row_idx + 1} ---")
    df_painel = pd.read_excel(file_path, sheet_name='PAINEL', header=header_row_idx, nrows=20)

    log(f"\nTotal de colunas: {len(df_painel.columns)}")
    log("\nColunas encontradas:")
    for i, col in enumerate(df_painel.columns, 1):
        log(f"  {i:2d}. {col}")

    # Procurar colunas de SALDO
    saldo_keywords = ['SALDO', 'CARTAO', 'CARTÃO', 'PRESTACAO', 'PRESTAÇÃO', 'FINAL']
    saldo_cols = []

    for col in df_painel.columns:
        col_str = str(col).upper()
        for kw in saldo_keywords:
            if kw in col_str:
                saldo_cols.append(col)
                break

    log(f"\n--- COLUNAS DE SALDO ENCONTRADAS ---")
    log(f"Total: {len(saldo_cols)}")
    for c in saldo_cols:
        log(f"  - {c}")

    # Analisar com openpyxl
    if saldo_cols:
        log("\n" + "=" * 60)
        log("ANALISE DE FORMULAS VS VALORES")
        log("=" * 60)

        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb['PAINEL']

        log(f"Dimensoes: {ws.max_row} linhas x {ws.max_column} colunas")

        # Mapear colunas pelo nome
        col_mapping = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=header_row_idx + 1, column=col_idx).value
            if cell_val:
                for target_col in saldo_cols:
                    if str(cell_val).strip().upper() == str(target_col).strip().upper():
                        col_mapping[target_col] = col_idx

        log(f"\nMapeamento: {col_mapping}")

        data_start = header_row_idx + 2

        for col_name, col_idx in col_mapping.items():
            log(f"\n[COLUNA] '{col_name}' (coluna Excel {col_idx})")
            log("-" * 40)

            formulas = 0
            values = 0
            empty = 0
            samples = []
            all_values = []

            for row_idx in range(data_start, min(data_start + 30, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)

                if cell.data_type == 'f':
                    formulas += 1
                    if len(samples) < 2:
                        samples.append(f"FORMULA L{row_idx}: {str(cell.value)[:50]}")
                elif cell.value is None or cell.value == '':
                    empty += 1
                else:
                    values += 1
                    all_values.append(cell.value)
                    if len(samples) < 2:
                        samples.append(f"VALOR   L{row_idx}: {cell.value}")

            log(f"  Formulas: {formulas}")
            log(f"  Valores estaticos: {values}")
            log(f"  Vazios: {empty}")

            for s in samples:
                log(f"    {s}")

            if formulas > 0 and values == 0:
                log(f"  >>> RESULTADO: FORMULAS (calculado automaticamente)")
            elif values > 0 and formulas == 0:
                log(f"  >>> RESULTADO: VALORES ESTATICOS (inserido manualmente)")
            elif formulas > 0 and values > 0:
                log(f"  >>> RESULTADO: MISTO (formulas e valores)")
            else:
                log(f"  >>> RESULTADO: Sem dados")

            # Estatisticas
            numeric_vals = [v for v in all_values if isinstance(v, (int, float))]
            if numeric_vals:
                log(f"\n  Estatisticas (valores numericos):")
                log(f"    Quantidade: {len(numeric_vals)}")
                log(f"    Min: {min(numeric_vals):,.2f}")
                log(f"    Max: {max(numeric_vals):,.2f}")
                log(f"    Media: {sum(numeric_vals)/len(numeric_vals):,.2f}")

                # Verificar se parece atualizado
                zeros = sum(1 for v in numeric_vals if v == 0)
                if zeros == len(numeric_vals) and len(numeric_vals) > 0:
                    log(f"    !!! TODOS OS VALORES SAO ZERO - POSSIVELMENTE DESATUALIZADO !!!")
                elif zeros > len(numeric_vals) * 0.5:
                    log(f"    ! ATENCAO: {zeros}/{len(numeric_vals)} sao zero ({zeros/len(numeric_vals)*100:.1f}%)")
                else:
                    log(f"    OK: Apenas {zeros}/{len(numeric_vals)} sao zero")

        wb.close()

    log("\n" + "=" * 80)
    log("ANALISE CONCLUIDA")
    log("=" * 80)

except Exception as e:
    log(f"\nERRO: {e}")
    log(traceback.format_exc())
