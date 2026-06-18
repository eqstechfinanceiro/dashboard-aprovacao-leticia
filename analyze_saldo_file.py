#!/usr/bin/env python3
"""
Analise especifica das colunas de SALDO - salva em arquivo
"""

import openpyxl
import os
import sys

file_path = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\controle-api\data\CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsx"
output_path = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\relatorio_saldo.txt"

# Abrir arquivo de saida
with open(output_path, 'w', encoding='utf-8') as fout:
    def log(msg):
        print(msg)
        fout.write(str(msg) + '\n')

    log("=" * 80)
    log("ANALISE DAS COLUNAS DE SALDO - ABA PAINEL")
    log("=" * 80)
    log(f"\nArquivo analisado: {os.path.basename(file_path)}")

    try:
        # Carregar workbook
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb['PAINEL']

        log(f"\nDimensoes da aba PAINEL: {ws.max_row} linhas x {ws.max_column} colunas")

        # Cabeçalho esta na linha 11 (indice 10 + 1)
        header_row = 11
        data_start = 12

        log(f"\nCabeçalho identificado na linha: {header_row}")
        log(f"Dados comecam na linha: {data_start}")

        # Colunas especificas que o usuario pediu
        target_cols = {}

        # Encontrar indices das colunas
        log(f"\n--- LISTA DE TODAS AS COLUNAS (linha {header_row}) ---")
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=header_row, column=col_idx).value
            if cell_val:
                log(f"  Coluna {col_idx}: {cell_val}")

                # Verificar matches
                cell_str = str(cell_val).strip().upper()
                if 'SALDO FINAL' == cell_str:
                    target_cols['SALDO FINAL'] = col_idx
                elif 'SALDO CARTAO' in cell_str or '(-) SALDO CARTAO' in cell_str:
                    if 'SALDO CARTAO' not in target_cols:
                        target_cols['SALDO CARTAO'] = col_idx
                elif 'SALDO PRESTAÇÃO' in cell_str or 'SALDO PRESTACAO' in cell_str:
                    if 'SALDO PRESTAÇÃO' not in target_cols:
                        target_cols['SALDO PRESTAÇÃO'] = col_idx

        log("\n" + "=" * 60)
        log("MAPEAMENTO DAS COLUNAS DE INTERESSE:")
        log("=" * 60)
        for name, idx in target_cols.items():
            if idx:
                log(f"  {name}: coluna {idx}")

        # Analisar cada coluna
        log("\n" + "=" * 60)
        log("ANALISE DE FORMULAS VS VALORES")
        log("=" * 60)

        for col_name, col_idx in target_cols.items():
            if not col_idx:
                continue

            log(f"\n{'='*60}")
            log(f"COLUNA: {col_name}")
            log(f"Posicao: coluna {col_idx}")
            log(f"{'='*60}")

            formulas = 0
            values = 0
            empty = 0
            samples = []
            all_numeric = []

            # Verificar primeiras 50 linhas de dados
            for row_idx in range(data_start, min(data_start + 50, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)

                # Verificar tipo
                if cell.data_type == 'f':  # formula
                    formulas += 1
                    if len(samples) < 3:
                        samples.append(f"  FORMULA L{row_idx}: {str(cell.value)[:60]}")
                elif cell.value is None or cell.value == '':
                    empty += 1
                else:
                    values += 1
                    if isinstance(cell.value, (int, float)):
                        all_numeric.append(cell.value)
                    if len(samples) < 3:
                        samples.append(f"  VALOR   L{row_idx}: {cell.value} (tipo: {type(cell.value).__name__})")

            log(f"\nContagem nas primeiras 50 linhas de dados:")
            log(f"  Formulas detectadas: {formulas}")
            log(f"  Valores estaticos: {values}")
            log(f"  Celulas vazias: {empty}")

            if samples:
                log(f"\nAmostras:")
                for s in samples:
                    log(s)

            # Resultado
            log(f"\n>>> RESULTADO:")
            if formulas > 0 and values == 0:
                log(f"    FORMULAS (calculado automaticamente)")
            elif values > 0 and formulas == 0:
                log(f"    VALORES ESTATICOS (inserido manualmente)")
                log(f"    * Pode indicar dados nao atualizados se estiverem desatualizados")
            elif formulas > 0 and values > 0:
                log(f"    MISTO (formulas e valores)")
            else:
                log(f"    SEM DADOS nas linhas analisadas")

            # Estatisticas
            if all_numeric:
                log(f"\nEstatisticas dos valores numericos:")
                log(f"  Quantidade analisada: {len(all_numeric)}")
                log(f"  Valor minimo: {min(all_numeric):,.2f}")
                log(f"  Valor maximo: {max(all_numeric):,.2f}")
                log(f"  Media: {sum(all_numeric)/len(all_numeric):,.2f}")

                # Verificar atualidade
                zeros = sum(1 for v in all_numeric if v == 0)
                non_zeros = len(all_numeric) - zeros

                log(f"  Valores zero: {zeros}")
                log(f"  Valores nao-zero: {non_zeros}")

                if zeros == len(all_numeric) and len(all_numeric) > 0:
                    log(f"  *** ALERTA: TODOS OS VALORES SAO ZERO! ***")
                    log(f"  *** ISSO INDICA DADOS DESATUALIZADOS ***")
                elif zeros > len(all_numeric) * 0.5:
                    log(f"  * ATENCAO: Mais da metade dos valores sao zero")
                else:
                    log(f"  OK: Maioria dos valores esta preenchida")

                # Mostrar alguns valores
                log(f"\n  Primeiros 5 valores: {[f'{v:,.2f}' for v in all_numeric[:5]]}")
                log(f"  Ultimos 5 valores: {[f'{v:,.2f}' for v in all_numeric[-5:]]}")

        wb.close()

        log("\n" + "=" * 80)
        log("ANALISE CONCLUIDA COM SUCESSO")
        log("=" * 80)

    except Exception as e:
        log(f"\nERRO durante a analise: {e}")
        import traceback
        log(traceback.format_exc())

log(f"\nRelatorio salvo em: {output_path}")
