#!/usr/bin/env python3
"""
Analise especifica das colunas de SALDO
"""

import openpyxl
import os

file_path = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\controle-api\data\CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsx"

print("=" * 80)
print("ANALISE DAS COLUNAS DE SALDO - PAINEL")
print("=" * 80)
print(f"\nArquivo: {os.path.basename(file_path)}")

# Carregar workbook
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb['PAINEL']

print(f"\nDimensoes da aba PAINEL: {ws.max_row} linhas x {ws.max_column} colunas")

# Cabeçalho esta na linha 11 (indice 10 + 1)
header_row = 11
data_start = 12

# Colunas especificas que o usuario pediu
target_cols = {
    'SALDO FINAL': None,
    'SALDO CARTAO': None,
    'SALDO PRESTAÇÃO': None,
    '(-) SALDO CARTAO': None,
    'SALDO PRESTAÇÃO': None
}

# Encontrar indices das colunas
print(f"\nProcurando colunas na linha {header_row}...")
for col_idx in range(1, ws.max_column + 1):
    cell_val = ws.cell(row=header_row, column=col_idx).value
    if cell_val:
        cell_str = str(cell_val).strip().upper()
        print(f"  Col {col_idx}: {cell_val}")

        # Verificar matches
        if 'SALDO FINAL' in cell_str:
            target_cols['SALDO FINAL'] = col_idx
            print(f"    >>> MATCH: SALDO FINAL")
        elif 'SALDO CARTAO' in cell_str or '(-) SALDO CARTAO' in cell_str:
            target_cols['SALDO CARTAO'] = col_idx
            print(f"    >>> MATCH: SALDO CARTAO")
        elif 'SALDO PRESTAÇÃO' in cell_str or 'SALDO PRESTACAO' in cell_str:
            target_cols['SALDO PRESTAÇÃO'] = col_idx
            print(f"    >>> MATCH: SALDO PRESTAÇÃO")

print("\n" + "=" * 60)
print("MAPEAMENTO ENCONTRADO:")
print("=" * 60)
for name, idx in target_cols.items():
    if idx:
        print(f"  {name}: coluna {idx}")

# Analisar cada coluna
print("\n" + "=" * 60)
print("ANALISE DE FORMULAS VS VALORES")
print("=" * 60)

for col_name, col_idx in target_cols.items():
    if not col_idx:
        continue

    print(f"\n[COLUNA] {col_name} (coluna {col_idx})")
    print("-" * 50)

    formulas = 0
    values = 0
    empty = 0
    samples = []
    all_numeric = []

    for row_idx in range(data_start, min(data_start + 50, ws.max_row + 1)):
        cell = ws.cell(row=row_idx, column=col_idx)

        # Verificar tipo
        if cell.data_type == 'f':
            formulas += 1
            if len(samples) < 3:
                samples.append(f"FORMULA L{row_idx}: {str(cell.value)[:60]}")
        elif cell.value is None or cell.value == '':
            empty += 1
        else:
            values += 1
            if isinstance(cell.value, (int, float)):
                all_numeric.append(cell.value)
            if len(samples) < 3:
                samples.append(f"VALOR   L{row_idx}: {cell.value}")

    print(f"  Formulas: {formulas}")
    print(f"  Valores: {values}")
    print(f"  Vazios: {empty}")

    for s in samples:
        print(f"    {s}")

    # Resultado
    if formulas > 0 and values == 0:
        print(f"  >>> RESULTADO: FORMULAS (calculado automaticamente)")
    elif values > 0 and formulas == 0:
        print(f"  >>> RESULTADO: VALORES ESTATICOS (inserido manualmente)")
    elif formulas > 0 and values > 0:
        print(f"  >>> RESULTADO: MISTO (formulas e valores)")
    else:
        print(f"  >>> RESULTADO: Sem dados")

    # Estatisticas
    if all_numeric:
        print(f"\n  Estatisticas dos valores:")
        print(f"    Quantidade: {len(all_numeric)}")
        print(f"    Min: {min(all_numeric):,.2f}")
        print(f"    Max: {max(all_numeric):,.2f}")
        print(f"    Media: {sum(all_numeric)/len(all_numeric):,.2f}")
        print(f"    Ultimos 5: {[f'{v:,.2f}' for v in all_numeric[-5:]]}")

        zeros = sum(1 for v in all_numeric if v == 0)
        if zeros == len(all_numeric) and len(all_numeric) > 0:
            print(f"    !!! ALERTA: TODOS OS VALORES SAO ZERO !!!")
        elif zeros > len(all_numeric) * 0.3:
            print(f"    ! ATENCAO: {zeros}/{len(all_numeric)} valores sao zero")
        else:
            print(f"    OK: Dados parecem estar preenchidos")

wb.close()

print("\n" + "=" * 80)
print("ANALISE CONCLUIDA")
print("=" * 80)
