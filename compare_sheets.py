import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict
import re

OURS = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\BASE_PREST_API_FRESH_V2.xlsx"
REF = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\base-prest-referencia.xlsx"
OUTPUT = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\COMPARACAO_PREST.xlsx"

USERS = [
    ("ADSON ARAUJO DA SILVA", "94551804304", 40453.97, 34711.36, 5742.61),
    ("ALEX SANDRO PEREIRA MARQUES", "82877904091", 2732.36, 2575.56, 156.80),
    ("ALIRIO THOMAZ VALENTIM", "99335220000", 3093.40, 0, 3093.40),
    ("ANELISE MENDONCA DE AVILA", "02873491019", 1391.77, 4029.95, -2638.18),
    ("DEBORA ARAUJO FIGUEIREDO", "85225720030", 12957.94, 12937.94, 20.00),
    ("GUILHERME DOS SANTOS GOMES", "10489127959", 366.20, 0, 366.20),
    ("ILMAR FONTES VIEIRA", "22136444972", 3491.60, 3459.80, 31.80),
    ("JACKSON NATAN RODRIGUES DA SILVA", "01456786288", 3027.15, 2875.60, 151.55),
    ("JULIA FRANCISCO MADUREIRA", "15322365737", 1143.87, 467.69, 676.18),
    ("MARCIO ROBERTO DA SILVA", "96212993068", 2013.17, 2102.48, -89.31),
    ("MARCOS LUIZ SCHULTZ", "97607240063", 5860.59, 5678.29, 182.30),
    ("MATHEUS FERNANDES GUIMARAES", "17996820724", 43138.08, 42338.08, 800.00),
    ("MILTON FERREIRA DAS NEVES JUNIOR", "05100355581", 15183.06, 16582.65, -1399.59),
    ("NEDSON LUIZ DURAES VELOSO", "11498112692", 11527.02, 11487.02, 40.00),
    ("ORESTE BARBOSA DA SILVA JUNIOR", "25075562850", 87639.09, 81744.27, 5894.82),
    ("PATRICK FERNANDO GOULART ALVES", "07214272946", 16689.06, 16638.26, 50.80),
    ("RAFAEL RAMOS DAS NEVES", "06149576485", 1619.12, 819.00, 800.12),
    ("RENATA PALMA LEAL", "02449603050", 70.00, 0, 70.00),
    ("RENATO PARREIRAS", "97162485634", 12178.21, 11196.05, 982.16),
    ("RODRIGO CESAR DOS SANTOS", "07024923610", 35224.40, 34887.00, 337.40),
]

def norm_cpf(cpf):
    if cpf is None:
        return ""
    s = str(cpf).strip()
    s = re.sub(r'[^\d]', '', s)
    return s.zfill(11)

print("[1/4] Loading our sheet (BASE_PREST_API_FRESH_V2)...", flush=True)
wb_ours = openpyxl.load_workbook(OURS, read_only=True)
ws_ours = wb_ours["BASE PREST (API)"]

target_cpfs = {norm_cpf(c) for _, c, *_ in USERS}

ours_by_cpf = defaultdict(dict)  # cpf -> {expense_id: {report_id, report_name, value, description, date, status, pmid}}
for row in ws_ours.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    cpf = norm_cpf(row[3])
    if cpf not in target_cpfs:
        continue
    eid = str(row[7]) if row[7] else ""
    if not eid:
        continue
    ours_by_cpf[cpf][eid] = {
        "report_id": row[0],
        "report_name": row[1],
        "status": row[2],
        "nome": row[4],
        "approval_date": row[5],
        "expense_desc": row[8],
        "expense_value": float(row[9]) if row[9] else 0,
        "expense_date": row[10],
        "expense_status": row[11],
        "payment_method_id": row[12],
    }
wb_ours.close()
print(f"  Loaded {sum(len(v) for v in ours_by_cpf.values())} expenses for {len(ours_by_cpf)} users", flush=True)

print("[2/4] Loading reference sheet (base-prest-referencia)...", flush=True)
wb_ref = openpyxl.load_workbook(REF, read_only=True)
ws_ref = wb_ref["Planilha1"]

ref_by_cpf = defaultdict(dict)
for row in ws_ref.iter_rows(min_row=3, values_only=True):
    if not row[0]:
        continue
    cpf = norm_cpf(row[9])
    if cpf not in target_cpfs:
        continue
    eid = str(row[0]) if row[0] else ""
    if not eid:
        continue
    val = row[26]
    if val is None or val == "":
        val = row[29]
    try:
        val = float(val) if val else 0
    except:
        val = 0
    ref_by_cpf[cpf][eid] = {
        "report_id": row[1],
        "report_name": row[2],
        "data": row[3],
        "nome": row[4],
        "status": row[10],
        "expense_desc": row[12],
        "forma_pagamento": row[18],
        "expense_value": val,
    }
wb_ref.close()
print(f"  Loaded {sum(len(v) for v in ref_by_cpf.values())} expenses for {len(ref_by_cpf)} users", flush=True)

print("[3/4] Comparing...", flush=True)

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
only_ref_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
only_ours_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
diff_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

wb_out = openpyxl.Workbook()
ws_summary = wb_out.active
ws_summary.title = "RESUMO"

summary_headers = ["Colaborador", "CPF", "Prestação Referência", "Prestação Automação", "Diferença",
                   "Qtd Ref", "Qtd Automação", "Só na Ref", "Só na Automação", "Valores Diferentes",
                   "Soma Só Ref", "Soma Só Automação", "Diferença Valores"]
ws_summary.append(summary_headers)
for col in range(1, len(summary_headers) + 1):
    cell = ws_summary.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill

for nome, cpf, ref_total, our_total, diff in USERS:
    nc = norm_cpf(cpf)
    ours = ours_by_cpf.get(nc, {})
    refs = ref_by_cpf.get(nc, {})

    our_eids = set(ours.keys())
    ref_eids = set(refs.keys())

    only_ref = ref_eids - our_eids
    only_ours = our_eids - ref_eids
    common = our_eids & ref_eids
    diff_values = [(eid, refs[eid]["expense_value"], ours[eid]["expense_value"])
                   for eid in common
                   if abs(refs[eid]["expense_value"] - ours[eid]["expense_value"]) > 0.01]

    sum_only_ref = sum(refs[eid]["expense_value"] for eid in only_ref)
    sum_only_ours = sum(ours[eid]["expense_value"] for eid in only_ours)
    sum_diff = sum(refs[eid]["expense_value"] - ours[eid]["expense_value"] for eid, _, _ in diff_values)

    ws_summary.append([
        nome, cpf, ref_total, our_total, diff,
        len(refs), len(ours), len(only_ref), len(only_ours), len(diff_values),
        round(sum_only_ref, 2), round(sum_only_ours, 2), round(sum_diff, 2)
    ])

    safe_name = nome[:28].replace("/", "-")
    ws = wb_out.create_sheet(title=safe_name)

    ws.append(["=== SÓ NA REFERÊNCIA (não temos na automação) ==="])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12, color="FF0000")
    ws.append(["Expense ID", "Report ID", "Report Name", "Descrição", "Valor", "Status", "Forma Pgto", "Data"])
    for col in range(1, 9):
        ws.cell(row=ws.max_row, column=col).font = header_font
        ws.cell(row=ws.max_row, column=col).fill = header_fill
    for eid in sorted(only_ref, key=lambda x: refs[x]["expense_value"], reverse=True):
        e = refs[eid]
        ws.append([eid, e["report_id"], e["report_name"], e["expense_desc"],
                   e["expense_value"], e["status"], e["forma_pagamento"], e["data"]])
        for col in range(1, 9):
            ws.cell(row=ws.max_row, column=col).fill = only_ref_fill

    ws.append([])
    ws.append(["=== SÓ NA AUTOMAÇÃO (não tem na referência) ==="])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12, color="008000")
    ws.append(["Expense ID", "Report ID", "Report Name", "Descrição", "Valor", "Status", "PMID", "Data"])
    for col in range(1, 9):
        ws.cell(row=ws.max_row, column=col).font = header_font
        ws.cell(row=ws.max_row, column=col).fill = header_fill
    for eid in sorted(only_ours, key=lambda x: ours[x]["expense_value"], reverse=True):
        e = ours[eid]
        ws.append([eid, e["report_id"], e["report_name"], e["expense_desc"],
                   e["expense_value"], e["status"], e["payment_method_id"], e["expense_date"]])
        for col in range(1, 9):
            ws.cell(row=ws.max_row, column=col).fill = only_ours_fill

    ws.append([])
    ws.append(["=== VALORES DIFERENTES (mesmo expense ID, valor diferente) ==="])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12, color="FFA500")
    ws.append(["Expense ID", "Report ID (Ref)", "Report ID (Ours)", "Descrição (Ref)", "Descrição (Ours)",
               "Valor Ref", "Valor Automação", "Diferença"])
    for col in range(1, 9):
        ws.cell(row=ws.max_row, column=col).font = header_font
        ws.cell(row=ws.max_row, column=col).fill = header_fill
    for eid, ref_val, our_val in sorted(diff_values, key=lambda x: abs(x[1] - x[2]), reverse=True):
        ws.append([eid, refs[eid]["report_id"], ours[eid]["report_id"],
                   refs[eid]["expense_desc"], ours[eid]["expense_desc"],
                   ref_val, our_val, round(ref_val - our_val, 2)])
        for col in range(1, 9):
            ws.cell(row=ws.max_row, column=col).fill = diff_fill

    ws.append([])
    ws.append(["=== TOTAIS ==="])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
    ws.append(["", "", "", "", "Soma Só Ref", "Soma Só Automação", "Dif Valores", ""])
    ws.append(["", "", "", "", round(sum_only_ref, 2), round(sum_only_ours, 2), round(sum_diff, 2), ""])

    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 22
    ws.column_dimensions["D"].width = 35

for col_idx, col_name in enumerate(["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"], 1):
    ws_summary.column_dimensions[col_name].width = 18
ws_summary.column_dimensions["A"].width = 35

print("[4/4] Saving...", flush=True)
wb_out.save(OUTPUT)
print(f"DONE! Saved to {OUTPUT}", flush=True)
print(f"\nSummary:", flush=True)
for nome, cpf, ref_total, our_total, diff in USERS:
    nc = norm_cpf(cpf)
    ours = ours_by_cpf.get(nc, {})
    refs = ref_by_cpf.get(nc, {})
    only_ref = set(refs.keys()) - set(ours.keys())
    only_ours = set(ours.keys()) - set(refs.keys())
    common = set(ours.keys()) & set(refs.keys())
    diff_vals = sum(1 for eid in common if abs(refs[eid]["expense_value"] - ours[eid]["expense_value"]) > 0.01)
    sum_ref = sum(refs[eid]["expense_value"] for eid in only_ref)
    sum_ours = sum(ours[eid]["expense_value"] for eid in only_ours)
    print(f"  {nome[:30]:<30} | ref={len(refs):<5} our={len(ours):<5} | só_ref={len(only_ref):<3} (R${sum_ref:>10.2f}) | só_our={len(only_ours):<3} (R${sum_ours:>10.2f}) | diff_val={diff_vals}")
