#!/usr/bin/env python3
"""
Script para analisar completamente todas as tabs, colunas e linhas das duas planilhas
"""

import openpyxl
import pandas as pd
import json
from pathlib import Path

def analyze_xlsx(file_path):
    """Analisa arquivo .xlsx completamente"""
    print(f"\n{'='*80}")
    print(f"ANALISANDO: {file_path}")
    print(f"{'='*80}")

    wb = openpyxl.load_workbook(file_path, data_only=True)

    result = {
        "file": str(file_path),
        "type": "xlsx",
        "sheets": {}
    }

    for sheet_name in wb.sheetnames:
        print(f"\n--- SHEET: {sheet_name} ---")
        ws = wb[sheet_name]

        sheet_data = {
            "name": sheet_name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "columns": [],
            "sample_rows": []
        }

        # Ler todas as colunas da primeira linha
        print(f"Dimensões: {ws.max_row} linhas x {ws.max_column} colunas")

        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            sheet_data["columns"].append({
                "index": col_idx,
                "letter": openpyxl.utils.get_column_letter(col_idx),
                "name": str(cell_value) if cell_value else ""
            })
            print(f"  Coluna {col_idx} ({openpyxl.utils.get_column_letter(col_idx)}): {cell_value}")

        # Ler primeiras 5 linhas como amostra
        print(f"\nPrimeiras 5 linhas de dados:")
        for row_idx in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            sheet_data["sample_rows"].append(row_data)
            if row_idx <= 5:
                print(f"  Linha {row_idx}: {row_data[:10]}...")  # Primeiras 10 colunas

        result["sheets"][sheet_name] = sheet_data

    return result

def analyze_xlsb(file_path):
    """Analisa arquivo .xlsb completamente"""
    print(f"\n{'='*80}")
    print(f"ANALISANDO: {file_path}")
    print(f"{'='*80}")

    # Usar pandas para ler .xlsb
    xls = pd.ExcelFile(file_path)

    result = {
        "file": str(file_path),
        "type": "xlsb",
        "sheets": {}
    }

    for sheet_name in xls.sheet_names:
        print(f"\n--- SHEET: {sheet_name} ---")
        df = pd.read_excel(xls, sheet_name=sheet_name, header=None)

        sheet_data = {
            "name": sheet_name,
            "max_row": len(df),
            "max_column": len(df.columns),
            "columns": [],
            "sample_rows": []
        }

        print(f"Dimensões: {len(df)} linhas x {len(df.columns)} colunas")

        # Ler todas as colunas da primeira linha
        for col_idx in range(len(df.columns)):
            cell_value = df.iloc[0, col_idx]
            sheet_data["columns"].append({
                "index": col_idx + 1,
                "name": str(cell_value) if pd.notna(cell_value) else ""
            })
            print(f"  Coluna {col_idx + 1}: {cell_value}")

        # Ler primeiras 5 linhas como amostra
        print(f"\nPrimeiras 5 linhas de dados:")
        for row_idx in range(min(5, len(df))):
            row_data = []
            for col_idx in range(len(df.columns)):
                cell_value = df.iloc[row_idx, col_idx]
                row_data.append(str(cell_value) if pd.notna(cell_value) else "")
            sheet_data["sample_rows"].append(row_data)
            if row_idx < 5:
                print(f"  Linha {row_idx + 1}: {row_data[:10]}...")  # Primeiras 10 colunas

        result["sheets"][sheet_name] = sheet_data

    return result

def main():
    base_path = Path(r"c:/Users/italo.medrado/Desktop/Projects/Análise de dados/Leticia/dashboard-test/data")

    planilha1 = base_path / "1QZ ABRIL 2026 - VEXPENSES (1).xlsx"
    planilha2 = base_path / "CONTROLE - VEXPENSES - ABRIL- 2026 (1).xlsb"

    all_results = {
        "planilha1": analyze_xlsx(planilha1),
        "planilha2": analyze_xlsb(planilha2)
    }

    # Salvar resultado em JSON
    output_path = Path(r"c:/Users/italo.medrado/Desktop/Projects/Análise de dados/Leticia/dashboard-test/vexpenses-dashboard/spreadsheets_complete_analysis.json")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, indent=2, ensure_ascii=False)

    print(f"\n{'='*80}")
    print(f"Análise completa salva em: {output_path}")
    print(f"{'='*80}")

if __name__ == "__main__":
    main()
