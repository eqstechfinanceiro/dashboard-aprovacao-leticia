import openpyxl
import json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ============ STYLES ============
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
title_font = Font(bold=True, size=14)
sub_font = Font(bold=True, size=12, color="4472C4")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def auto_width(ws, max_width=45):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_width, max_len + 2)

# ============ LOAD DATA ============
print("Loading CONTROLE spreadsheet...")
wb_controle = openpyxl.load_workbook('CONTROLE - VEXPENSES - AGOSTO 2026.xlsx', read_only=True, data_only=True)

# PAINEL data per CPF
print("  Reading PAINEL...")
painel = {}
ws = wb_controle['PAINEL']
for row in ws.iter_rows(min_row=12, values_only=True):
    cpf = str(row[2]).strip() if row[2] else ''
    nome = str(row[1]).strip().upper() if row[1] else ''
    if not cpf or cpf == 'None':
        continue
    painel[cpf] = {
        'nome': nome,
        'carga': float(row[13]) if row[13] is not None else 0,
        'transferencia': abs(float(row[14])) if row[14] is not None else 0,
        'tarifa': abs(float(row[15])) if row[15] is not None else 0,
        'prestacao': float(row[16]) if row[16] is not None else 0,
        'saldo_prestacao': float(row[17]) if row[17] is not None else 0,
        'saldo_cartao': float(row[18]) if row[18] is not None else 0,
        'saldo_final': float(row[19]) if row[19] is not None else 0,
        'situacao': str(row[23]).strip() if row[23] else '',
    }
print(f"  PAINEL: {len(painel)} users")

# BASE PREST per CPF with report breakdown
print("  Reading BASE PREST...")
base_prest = defaultdict(lambda: {'total': 0, 'count': 0, 'aprovado': 0, 'enviado': 0, 'reports': {}})
ws = wb_controle['BASE PREST ']
for row in ws.iter_rows(min_row=4, values_only=True):
    cpf = str(row[9]).strip() if row[9] else ''
    valor = float(row[26]) if row[26] is not None else 0
    status = str(row[10]).strip() if row[10] else ''
    report_id = str(row[1]).strip() if row[1] else ''
    report_name = str(row[2]).strip() if row[2] else ''
    if not cpf or cpf == 'None':
        continue
    base_prest[cpf]['total'] += valor
    base_prest[cpf]['count'] += 1
    if status == 'Aprovado':
        base_prest[cpf]['aprovado'] += 1
    elif status == 'Enviado':
        base_prest[cpf]['enviado'] += 1
    if report_id and report_id != 'None':
        if report_id not in base_prest[cpf]['reports']:
            base_prest[cpf]['reports'][report_id] = {'name': report_name, 'status': status, 'total': 0, 'count': 0}
        base_prest[cpf]['reports'][report_id]['total'] += valor
        base_prest[cpf]['reports'][report_id]['count'] += 1
print(f"  BASE PREST: {len(base_prest)} CPFs")

# Load API frozen data
print("Loading API frozen data...")
with open('api_frozen.json', 'r', encoding='utf-8-sig') as f:
    api_data = json.load(f)
api = {}
for row in api_data:
    cpf = row['cpf'].strip()
    api[cpf] = row
print(f"  API: {len(api)} users")

# Load API prestacao debug
print("Loading API prestacao debug...")
with open('prestacao_debug.json', 'r', encoding='utf-8-sig') as f:
    debug_raw = json.load(f)

api_prestacao = {}
for item in debug_raw.get('cpf_summary', []):
    cpf = item['user_cpf'].strip()
    api_prestacao[cpf] = {
        'prestacao_total': float(item['prestacao_total']),
        'total_expenses': float(item['total_expenses']),
        'report_count': int(item['report_count']),
        'aprovado_count': int(item['aprovado_count']),
        'enviado_count': int(item['enviado_count']),
    }

# API expense details per CPF
api_reports = defaultdict(list)
for item in debug_raw.get('expense_details', []):
    cpf = item['user_cpf'].strip()
    api_reports[cpf].append({
        'report_id': item['report_id'],
        'report_name': item['report_name'],
        'status': item['status'],
        'expense_count': int(item['expense_count']),
        'total_value': float(item['total_value']),
        'excluded_value': float(item['excluded_value']),
    })

# Load CARGA 1 QZ AGOSTO
print("Loading CARGA 1 QZ AGOSTO...")
wb_carga = openpyxl.load_workbook('CARGA 1 QZ AGOSTO 26 VEXPENSES EQS.xlsx', read_only=True, data_only=True)
ws_carga = wb_carga['1 QZ AGOSTO']
carga_sheet = {}
for row in ws_carga.iter_rows(min_row=7, values_only=True):
    cpf = str(row[1]).strip() if row[1] else ''
    nome = str(row[0]).strip().upper() if row[0] else ''
    if not cpf or cpf == 'None':
        continue
    carga_sheet[cpf] = {
        'nome': nome,
        'saldo_reembolsar': float(row[7]) if row[7] is not None else 0,
        'saldo_final': float(row[8]) if row[8] is not None else 0,
        'col_qz': float(row[9]) if row[9] is not None else 0,
        'saldo_cartao': float(row[10]) if row[10] is not None else 0,
        'adiantamento': float(row[11]) if row[11] is not None else 0,
        'carga_parcial': float(row[12]) if row[12] is not None else 0,
        'reembolso': float(row[13]) if row[13] is not None else 0,
        'carga_final': float(row[14]) if row[14] is not None else 0,
    }
print(f"  CARGA: {len(carga_sheet)} users")

# Load expense gaps debug
print("Loading expense gaps debug...")
with open('expense_gaps.json', 'r', encoding='utf-8-sig') as f:
    gaps_raw = json.load(f)

# ============ CREATE EXCEL ============
print("\nCreating investigation Excel...")
wb = openpyxl.Workbook()

# ---- TAB 1: SUMMARY ----
ws1 = wb.active
ws1.title = "SUMMARY"

ws1['A1'] = "INVESTIGATION: PAINEL vs API vs CARGA - August 2026 1QZ"
ws1['A1'].font = title_font
ws1.merge_cells('A1:H1')

ws1['A3'] = "FIELD-LEVEL TOTALS"
ws1['A3'].font = sub_font

headers = ['Field', 'PAINEL Total', 'API Total', 'CARGA Total', 'Diff (API-PAINEL)', 'Diff %', 'Status', 'Root Cause']
for c, h in enumerate(headers, 1):
    ws1.cell(row=4, column=c, value=h)
style_header(ws1, 4, len(headers))

fields = [
    ('Carga', 'carga', 'carga', None, 'Missing extrato rows (download_extrato incomplete)'),
    ('Transferencia', 'transferencia', 'transferencia', None, 'Missing extrato rows (download_extrato incomplete)'),
    ('Tarifa', 'tarifa', 'tarifa', None, 'FIXED: Added Estorno de taxa + Pendencia de taxa to SQL'),
    ('Prestacao', 'prestacao', 'prestacao', None, 'download_expenses incomplete - 793 missing expense rows (~R$95k)'),
    ('Saldo Cartao', 'saldo_cartao', 'saldo_cartao', 'saldo_cartao', 'Snapshot timing differences'),
    ('Saldo Final', 'saldo_final', 'saldo_final', 'saldo_final', 'Derived from other fields'),
]

row_idx = 5
for label, painel_key, api_key, carga_key, cause in fields:
    painel_total = sum(v.get(painel_key, 0) for v in painel.values())
    api_total = sum(v.get(api_key, 0) for v in api.values())
    carga_total = sum(v.get(carga_key, 0) for v in carga_sheet.values()) if carga_key else None
    diff = api_total - painel_total
    pct = (diff / painel_total * 100) if painel_total else 0
    status = 'OK' if abs(diff) < 1 else ('WARN' if abs(pct) < 5 else 'ERROR')

    ws1.cell(row=row_idx, column=1, value=label)
    ws1.cell(row=row_idx, column=2, value=round(painel_total, 2))
    ws1.cell(row=row_idx, column=3, value=round(api_total, 2))
    ws1.cell(row=row_idx, column=4, value=round(carga_total, 2) if carga_total is not None else 'N/A')
    ws1.cell(row=row_idx, column=5, value=round(diff, 2))
    ws1.cell(row=row_idx, column=6, value=f"{pct:.2f}%")
    ws1.cell(row=row_idx, column=7, value=status)
    ws1.cell(row=row_idx, column=8, value=cause)

    if status == 'OK':
        for c in range(1, 9): ws1.cell(row=row_idx, column=c).fill = good_fill
    elif status == 'WARN':
        for c in range(1, 9): ws1.cell(row=row_idx, column=c).fill = warn_fill
    else:
        for c in range(1, 9): ws1.cell(row=row_idx, column=c).fill = error_fill

    row_idx += 1

# Prestacao breakdown
row_idx += 2
ws1.cell(row=row_idx, column=1, value="PRESTACAO BREAKDOWN").font = sub_font
row_idx += 1

prest_info = [
    ('BASE PREST total (all rows)', 7345177.21, 'Sum of all BASE PREST rows in CONTROLE'),
    ('BASE PREST FATURA/CARTAO rows', 1436.00, 'Rows with FATURA/CARTAO in report name'),
    ('BASE PREST normal rows', 7343741.21, 'BASE PREST minus FATURA/CARTAO'),
    ('PAINEL prestacao total', 7345089.21, 'Sum of PAINEL col 17 (prestacao)'),
    ('API prestacao (frozen)', 7179656.53, 'Sum of api_frozen.json prestacao field'),
    ('API prestacao (debug, current DB)', 7188872.60, 'Sum from debug-prestacao endpoint'),
    ('API non-FATURA reports total', 7249353.33, 'Sum of non-FATURA APROVADO+ENVIADO report expenses'),
    ('API total expense rows in DB', 112623, 'Total rows in prestacao_expenses table'),
    ('BASE PREST total rows', 77203, 'Total rows in BASE PREST sheet'),
    ('Missing expense rows', 793, 'BASE PREST rows - API non-FATURA expense rows'),
    ('Missing expense value', 95823.88, 'BASE PREST normal - API non-FATURA total'),
    ('FATURA reports filtered by API', 2661, 'Reports with FATURA/CARTAO in name (filtered by isFaturaOrCartao)'),
    ('FATURA value excluded (payment_method 627401)', 54247.74, 'Excluded by payment_method_id filter (RAFAEL only)'),
    ('Enviado reports (non-FATURA)', 201, 'Reports with ENVIADO status (may have partial expenses)'),
]

for label, value, note in prest_info:
    ws1.cell(row=row_idx, column=1, value=label)
    ws1.cell(row=row_idx, column=2, value=value)
    ws1.cell(row=row_idx, column=3, value=note)
    ws1.cell(row=row_idx, column=3).font = Font(italic=True, size=10)
    row_idx += 1

# Root cause summary
row_idx += 2
ws1.cell(row=row_idx, column=1, value="ROOT CAUSES IDENTIFIED").font = sub_font
row_idx += 1

causes = [
    "1. TARIFA: SQL only summed 'Taxa', missing 'Estorno de taxa' and 'Pendencia de taxa'. FIXED in quinzena-complete, quinzena-export, validate-calc routes.",
    "2. PRESTACAO: download_expenses step not re-run after reports were updated with new expenses. DB has 76,410 expenses vs 77,203 in BASE PREST (793 missing, ~R$95k).",
    "3. PRESTACAO: FATURA/CARTAO reports (2,661) correctly filtered by name + payment_method_id=627401. Only R$1,436 in BASE PREST is FATURA (negligible).",
    "4. CARGA/TRANSFERENCIA: Missing extrato rows in DB (download_extrato incomplete). DJONATAN missing 680, LUIS CARLOS missing 2613.86, SILVIO missing 1000.",
    "5. SALDO CARTAO: Snapshot timing differences - API uses last snapshot up to cutoff, PAINEL may use different date.",
    "6. PAINEL formulas use SUMIFS without date filtering, while API filters by financial_cutoff (2026-07-31).",
]

for cause in causes:
    ws1.cell(row=row_idx, column=1, value=cause)
    ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
    ws1.cell(row=row_idx, column=1).alignment = Alignment(wrap_text=True)
    row_idx += 1

auto_width(ws1)

# ---- TAB 2: PER-CPF COMPARISON ----
ws2 = wb.create_sheet("PER-CPF COMPARISON")

headers2 = [
    'CPF', 'Nome', 'Painel Carga', 'API Carga', 'Diff Carga',
    'Painel Transf', 'API Transf', 'Diff Transf',
    'Painel Tarifa', 'API Tarifa', 'Diff Tarifa',
    'Painel Prest', 'API Prest', 'Diff Prest',
    'Painel Saldo Cart', 'API Saldo Cart', 'Diff Saldo Cart',
    'Painel Saldo Final', 'API Saldo Final', 'Diff Saldo Final',
    'CARGA Saldo Final', 'CARGA Saldo Cart', 'CARGA Carga Final',
    'Painel Situacao', 'In API?', 'In CARGA?',
]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=c, value=h)
style_header(ws2, 1, len(headers2))

all_cpfs = set(painel.keys()) | set(api.keys()) | set(carga_sheet.keys())
row_idx = 2
for cpf in sorted(all_cpfs):
    p = painel.get(cpf, {})
    a = api.get(cpf, {})
    cg = carga_sheet.get(cpf, {})
    nome = p.get('nome') or a.get('colaborador', '') or cg.get('nome', '')

    vals = [
        cpf, nome,
        round(p.get('carga', 0), 2), round(a.get('carga', 0), 2), round(a.get('carga', 0) - p.get('carga', 0), 2),
        round(p.get('transferencia', 0), 2), round(a.get('transferencia', 0), 2), round(a.get('transferencia', 0) - p.get('transferencia', 0), 2),
        round(p.get('tarifa', 0), 2), round(a.get('tarifa', 0), 2), round(a.get('tarifa', 0) - p.get('tarifa', 0), 2),
        round(p.get('prestacao', 0), 2), round(a.get('prestacao', 0), 2), round(a.get('prestacao', 0) - p.get('prestacao', 0), 2),
        round(p.get('saldo_cartao', 0), 2), round(a.get('saldo_cartao', 0), 2), round(a.get('saldo_cartao', 0) - p.get('saldo_cartao', 0), 2),
        round(p.get('saldo_final', 0), 2), round(a.get('saldo_final', 0), 2), round(a.get('saldo_final', 0) - p.get('saldo_final', 0), 2),
        round(cg.get('saldo_final', 0), 2), round(cg.get('saldo_cartao', 0), 2), round(cg.get('carga_final', 0), 2),
        p.get('situacao', ''), 'YES' if cpf in api else 'NO', 'YES' if cpf in carga_sheet else 'NO',
    ]
    for c, v in enumerate(vals, 1):
        ws2.cell(row=row_idx, column=c, value=v)

    for diff_col in [5, 8, 11, 14, 17, 20]:
        val = ws2.cell(row=row_idx, column=diff_col).value
        if val and abs(val) > 0.01:
            ws2.cell(row=row_idx, column=diff_col).fill = error_fill

    row_idx += 1

auto_width(ws2)

# ---- TAB 3: PRESTACAO DETAILS (sorted by diff) ----
ws3 = wb.create_sheet("PRESTACAO DETAILS")

headers3 = [
    'CPF', 'Nome', 'PAINEL Prest', 'API Prest (frozen)', 'API Prest (debug)',
    'BASE PREST Total', 'BASE PREST Count', 'BASE PREST Aprovado', 'BASE PREST Enviado',
    'API Report Count', 'API Aprovado Count', 'API Enviado Count',
    'Diff (API-PAINEL)', 'Diff (BASE-API)', 'Diff %', 'Likely Cause',
    'VExpenses Link',
]
for c, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=c, value=h)
style_header(ws3, 1, len(headers3))

prest_rows = []
for cpf in set(painel.keys()) | set(api.keys()) | set(api_prestacao.keys()) | set(base_prest.keys()):
    p = painel.get(cpf, {})
    a = api.get(cpf, {})
    ap = api_prestacao.get(cpf, {})
    bp = base_prest.get(cpf, {'total': 0, 'count': 0, 'aprovado': 0, 'enviado': 0})
    nome = p.get('nome') or a.get('colaborador', '')

    painel_prest = p.get('prestacao', 0)
    api_frozen = a.get('prestacao', 0)
    api_debug = ap.get('prestacao_total', 0)
    base_total = bp['total']
    diff_api_painel = api_frozen - painel_prest
    diff_base_api = base_total - api_debug
    pct = (diff_api_painel / painel_prest * 100) if painel_prest else 0

    cause = ''
    if abs(diff_api_painel) < 1:
        cause = 'OK - matches'
    elif abs(diff_base_api) < 1:
        cause = 'BASE PREST matches API - diff is from frozen data being stale'
    elif diff_base_api > 1:
        cause = f'Missing expenses in DB (~R${diff_base_api:.2f}) - download_expenses needs re-run'
    elif diff_base_api < -1:
        cause = f'API has MORE than BASE PREST (~R${abs(diff_base_api):.2f}) - possible extra expenses or status change'

    prest_rows.append((
        cpf, nome, painel_prest, api_frozen, api_debug,
        base_total, bp['count'], bp['aprovado'], bp['enviado'],
        ap.get('report_count', 0), ap.get('aprovado_count', 0), ap.get('enviado_count', 0),
        diff_api_painel, diff_base_api, pct, cause,
        f"https://app.vexpenses.com/reports?cpf={cpf}" if cpf else '',
    ))

prest_rows.sort(key=lambda x: abs(x[12]), reverse=True)

for row_idx, vals in enumerate(prest_rows, 2):
    for c, v in enumerate(vals, 1):
        if isinstance(v, float):
            v = round(v, 2)
        ws3.cell(row=row_idx, column=c, value=v)

    diff_val = vals[12]
    if diff_val and abs(diff_val) > 0.01:
        ws3.cell(row=row_idx, column=13).fill = error_fill
    diff_base = vals[13]
    if diff_base and abs(diff_base) > 0.01:
        ws3.cell(row=row_idx, column=14).fill = warn_fill
    if vals[15] and 'OK' in str(vals[15]):
        ws3.cell(row=row_idx, column=16).fill = good_fill

auto_width(ws3)

# ---- TAB 4: REPORT DETAILS (top 50 users with biggest prestacao diffs) ----
ws4 = wb.create_sheet("REPORT DETAILS")

headers4 = [
    'CPF', 'Nome', 'Report ID', 'Report Name', 'Status',
    'API Expense Count', 'API Total Value', 'API Excluded Value',
    'BASE PREST Count', 'BASE PREST Total', 'Diff Count', 'Diff Value',
    'VExpenses Link',
]
for c, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=c, value=h)
style_header(ws4, 1, len(headers4))

top_users = prest_rows[:50]
row_idx = 2
for cpf, nome, _, _, _, bp_total, bp_count, _, _, _, _, _, diff, _, _, cause, _ in top_users:
    if abs(diff) < 1:
        continue

    api_reps = api_reports.get(cpf, [])
    bp_reports = base_prest.get(cpf, {}).get('reports', {})
    api_map = {str(r['report_id']): r for r in api_reps}

    for rid, bp_info in sorted(bp_reports.items(), key=lambda x: x[1]['total'], reverse=True):
        api_info = api_map.get(rid, {})
        api_count = api_info.get('expense_count', 0)
        api_total = api_info.get('total_value', 0)
        api_excluded = api_info.get('excluded_value', 0)
        bp_count_r = bp_info['count']
        bp_total_r = bp_info['total']
        diff_count = api_count - bp_count_r
        diff_value = api_total - bp_total_r

        link = f"https://app.vexpenses.com/reports/{rid}"
        vals = [
            cpf, nome, rid, bp_info['name'], bp_info['status'],
            api_count, round(api_total, 2), round(api_excluded, 2),
            bp_count_r, round(bp_total_r, 2), diff_count, round(diff_value, 2),
            link,
        ]
        for c, v in enumerate(vals, 1):
            ws4.cell(row=row_idx, column=c, value=v)

        if abs(diff_value) > 0.01:
            ws4.cell(row=row_idx, column=12).fill = error_fill
        if diff_count < 0:
            ws4.cell(row=row_idx, column=11).fill = error_fill

        row_idx += 1

    ws4.cell(row=row_idx, column=1, value=f"--- {nome} (total diff: R${diff:.2f}) ---")
    ws4.cell(row=row_idx, column=1).font = Font(bold=True, italic=True)
    row_idx += 1

auto_width(ws4)

# ---- TAB 5: EXTRATO MISMATCHES (carga, transf, tarifa) ----
ws5 = wb.create_sheet("EXTRATO MISMATCHES")

headers5 = [
    'CPF', 'Nome', 'Field', 'PAINEL Value', 'API Value', 'Diff',
    'Possible Cause', 'VExpenses Link',
]
for c, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=c, value=h)
style_header(ws5, 1, len(headers5))

row_idx = 2
for cpf in sorted(set(painel.keys()) | set(api.keys())):
    p = painel.get(cpf, {})
    a = api.get(cpf, {})
    nome = p.get('nome') or a.get('colaborador', '')

    for field_name, painel_key, api_key in [('Carga', 'carga', 'carga'), ('Transferencia', 'transferencia', 'transferencia'), ('Tarifa', 'tarifa', 'tarifa')]:
        pv = p.get(painel_key, 0)
        av = a.get(api_key, 0)
        diff = av - pv
        if abs(diff) > 0.01:
            cause = ''
            if field_name == 'Tarifa':
                cause = 'API may miss Estorno de taxa / Pendencia de taxa (FIXED - needs re-freeze)'
            elif field_name == 'Carga':
                cause = 'Missing extrato rows in DB (download_extrato incomplete)'
            elif field_name == 'Transferencia':
                cause = 'Missing extrato rows in DB (download_extrato incomplete)'

            vals = [cpf, nome, field_name, round(pv, 2), round(av, 2), round(diff, 2), cause, f"https://app.vexpenses.com/reports?cpf={cpf}"]
            for c, v in enumerate(vals, 1):
                ws5.cell(row=row_idx, column=c, value=v)
            ws5.cell(row=row_idx, column=6).fill = error_fill
            row_idx += 1

auto_width(ws5)

# ---- TAB 6: ENVIADO REPORTS (may have partial expenses) ----
ws6 = wb.create_sheet("ENVIADO REPORTS")

headers6 = ['Report ID', 'Report Name', 'User CPF', 'User Name', 'Status', 'API Expenses', 'API Total', 'VExpenses Link']
for c, h in enumerate(headers6, 1):
    ws6.cell(row=1, column=c, value=h)
style_header(ws6, 1, len(headers6))

enviado_reports = gaps_raw.get('enviado_reports', [])
for row_idx, rep in enumerate(enviado_reports, 2):
    if float(rep.get('total_value', 0)) > 0:
        vals = [
            rep.get('report_id', ''), rep.get('report_name', ''),
            rep.get('user_cpf', ''), rep.get('user_name', ''),
            rep.get('status', ''), rep.get('expense_count', 0),
            round(float(rep.get('total_value', 0)), 2),
            f"https://app.vexpenses.com/reports/{rep.get('report_id', '')}",
        ]
        for c, v in enumerate(vals, 1):
            ws6.cell(row=row_idx, column=c, value=v)

auto_width(ws6)

# ---- TAB 7: SUSPICIOUS APROVADO (<5 expenses but value > 0) ----
ws7 = wb.create_sheet("SUSPICIOUS APROVADO")

headers7 = ['Report ID', 'Report Name', 'User CPF', 'User Name', 'Status', 'API Expenses', 'API Total', 'VExpenses Link']
for c, h in enumerate(headers7, 1):
    ws7.cell(row=1, column=c, value=h)
style_header(ws7, 1, len(headers7))

suspicious = gaps_raw.get('suspicious_aprovado', [])
for row_idx, rep in enumerate(suspicious, 2):
    vals = [
        rep.get('report_id', ''), rep.get('report_name', ''),
        rep.get('user_cpf', ''), rep.get('user_name', ''),
        rep.get('status', ''), rep.get('expense_count', 0),
        round(float(rep.get('total_value', 0)), 2),
        f"https://app.vexpenses.com/reports/{rep.get('report_id', '')}",
    ]
    for c, v in enumerate(vals, 1):
        ws7.cell(row=row_idx, column=c, value=v)

auto_width(ws7)

# ---- TAB 8: FATURA/CARTAO REPORTS ----
ws8 = wb.create_sheet("FATURA CARTAO REPORTS")

headers8 = ['Report ID', 'Report Name', 'Status', 'User CPF', 'User Name', 'VExpenses Link']
for c, h in enumerate(headers8, 1):
    ws8.cell(row=1, column=c, value=h)
style_header(ws8, 1, len(headers8))

fatura_reports = debug_raw.get('fatura_reports', [])
for row_idx, rep in enumerate(fatura_reports[:500], 2):
    vals = [
        rep.get('id', ''), rep.get('name', ''), rep.get('status', ''),
        rep.get('user_cpf', ''), rep.get('user_name', ''),
        f"https://app.vexpenses.com/reports/{rep.get('id', '')}",
    ]
    for c, v in enumerate(vals, 1):
        ws8.cell(row=row_idx, column=c, value=v)

auto_width(ws8)
ws8.cell(row=1, column=7, value=f"Total: {len(fatura_reports)} reports (showing first 500)")

# ============ SAVE ============
output_file = 'investigacao_discrepancas.xlsx'
wb.save(output_file)
print(f"\nSaved: {output_file}")
print(f"Tabs: {wb.sheetnames}")
