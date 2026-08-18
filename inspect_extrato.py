import openpyxl

wb = openpyxl.load_workbook('CONTROLE - VEXPENSES - AGOSTO 2026.xlsx', read_only=True, data_only=True)

# Check EXTRATO sheet
ws = wb['EXTRATO']
print("EXTRATO sheet - first 5 rows:")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
    vals = [str(v)[:18] if v is not None else '' for v in row[:20]]
    print(f"Row {i+1}: {vals}")

# Check SALDO CARTAO sheet
ws2 = wb['SALDO CARTAO']
print("\nSALDO CARTAO sheet - first 5 rows:")
for i, row in enumerate(ws2.iter_rows(min_row=1, max_row=5, values_only=True)):
    vals = [str(v)[:18] if v is not None else '' for v in row[:20]]
    print(f"Row {i+1}: {vals}")
