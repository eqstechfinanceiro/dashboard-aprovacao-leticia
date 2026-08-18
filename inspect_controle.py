import openpyxl

# Check CONTROLE sheet headers
wb = openpyxl.load_workbook('CONTROLE - VEXPENSES - AGOSTO 2026.xlsx', read_only=True, data_only=True)
print("Sheets:", wb.sheetnames)

# Check PAINEL sheet (which has the quinzena comparison fields)
ws = wb['PAINEL']
print("\nPAINEL sheet - first 15 rows:")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
    vals = [str(v)[:18] if v is not None else '' for v in row[:20]]
    print(f"Row {i+1}: {vals}")
