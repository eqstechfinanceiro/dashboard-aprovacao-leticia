#!/usr/bin/env python3
"""
Simulate proposed filters and measure impact on BASE PREST divergence:
1. Current state (baseline divergence)
2. Add 'fartur' to exclusion regex
3. Add 'cartão corporativo' / 'cartao corporativo' to exclusion regex
4. Filter users by CARTÃO VEXPENSES = 'SIM' (from reference PAINEL)
5. All filters combined

For each scenario, compare:
- Report count (API vs ref)
- Expense count (API vs ref)
- Total value (API vs ref)
- Remaining divergent reports
"""
import os, re
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from pathlib import Path
import openpyxl
from collections import defaultdict

BASE = Path(__file__).parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"

def norm(s):
    return " ".join(str(s or "").upper().strip().split())

# ============================================================
# Load reference BASE PREST
# ============================================================
print("Loading reference BASE PREST...")
wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_bp = wb["BASE PREST "]

ref_report_ids = set()
ref_expense_ids = set()
ref_total_value = 0
ref_by_cpf = defaultdict(lambda: {"count": 0, "total": 0})
for row in ws_bp.iter_rows(min_row=4, values_only=True):
    if row[0] is None:
        continue
    eid = int(row[0])
    rid = int(row[1]) if row[1] else None
    cpf = str(row[9] or "").strip() if len(row) > 9 else ""
    valor = float(row[26] or 0) if len(row) > 26 else 0
    
    ref_expense_ids.add(eid)
    if rid:
        ref_report_ids.add(rid)
    ref_total_value += valor
    ref_by_cpf[cpf]["count"] += 1
    ref_by_cpf[cpf]["total"] += valor

print(f"  Reference: {len(ref_report_ids)} reports, {len(ref_expense_ids)} expenses, R$ {ref_total_value:,.2f}")

# ============================================================
# Load reference PAINEL — get CARTÃO VEXPENSES = 'SIM' CPFs
# ============================================================
ws_p = wb["PAINEL"]
vexpenses_cpfs = set()
all_painel_cpfs = set()
for row in ws_p.iter_rows(min_row=12, values_only=True):
    if row[2] is None:
        continue
    cpf = str(row[2] or "").strip()
    cartao_vx = str(row[12] or "").strip().upper() if len(row) > 12 else ""
    all_painel_cpfs.add(cpf)
    if cartao_vx == "SIM":
        vexpenses_cpfs.add(cpf)

print(f"  PAINEL: {len(all_painel_cpfs)} total CPFs, {len(vexpenses_cpfs)} with CARTÃO VEXPENSES='SIM'")
wb.close()

# ============================================================
# Load API data
# ============================================================
print("\nLoading API data...")
conn = psycopg2.connect(NEON_URL, connect_timeout=10)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cur.execute("""
    SELECT e.id, e.report_id, e.date, e.value, e.description, e.status as expense_status,
           r.name as report_name, r.status as report_status,
           r.user_cpf, r.user_name, r.created_at as report_created_at
    FROM prestacao_expenses e
    JOIN prestacao_reports r ON e.report_id = r.id
    WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND r.user_cpf IS NOT NULL
    ORDER BY r.user_name, r.name, e.date
""")
all_api = cur.fetchall()
conn.close()
print(f"  API (all, no filters): {len(all_api)} expenses, {len(set(e['report_id'] for e in all_api))} reports, R$ {sum(float(e['value']) for e in all_api):,.2f}")

# ============================================================
# Define filter scenarios
# ============================================================

# Current exclusion regex (baseline)
def filter_current(expenses):
    result = []
    for e in expenses:
        name = (e["report_name"] or "").strip()
        if re.match(r'^(fatu|farur|cart)', name, re.IGNORECASE):
            continue
        if re.search(r'(fatura|fatuar|fatut|farur)', name, re.IGNORECASE):
            continue
        result.append(e)
    return result

# Add 'fartur' to exclusion
def filter_fartur(expenses):
    result = []
    for e in expenses:
        name = (e["report_name"] or "").strip()
        if re.match(r'^(fatu|farur|cart)', name, re.IGNORECASE):
            continue
        if re.search(r'(fatura|fatuar|fatut|farur|fartur)', name, re.IGNORECASE):
            continue
        result.append(e)
    return result

# Add 'cartão corporativo' / 'cartao corporativo' to exclusion
def filter_cartao_corp(expenses):
    result = []
    for e in expenses:
        name = (e["report_name"] or "").strip()
        if re.match(r'^(fatu|farur|cart)', name, re.IGNORECASE):
            continue
        if re.search(r'(fatura|fatuar|fatut|farur)', name, re.IGNORECASE):
            continue
        if re.search(r'(cartão corporativo|cartao corporativo)', name, re.IGNORECASE):
            continue
        result.append(e)
    return result

# Filter by CARTÃO VEXPENSES = 'SIM' CPFs
def filter_vexpenses_card(expenses):
    result = []
    for e in expenses:
        if e["user_cpf"] in vexpenses_cpfs:
            result.append(e)
    return result

# All filters combined
def filter_all(expenses):
    result = []
    for e in expenses:
        name = (e["report_name"] or "").strip()
        if re.match(r'^(fatu|farur|cart)', name, re.IGNORECASE):
            continue
        if re.search(r'(fatura|fatuar|fatut|farur|fartur)', name, re.IGNORECASE):
            continue
        if re.search(r'(cartão corporativo|cartao corporativo)', name, re.IGNORECASE):
            continue
        if e["user_cpf"] not in vexpenses_cpfs:
            continue
        result.append(e)
    return result

# All filters + also exclude "itau" in report name (broader ITAU exclusion)
def filter_all_plus_itau(expenses):
    result = []
    for e in expenses:
        name = (e["report_name"] or "").strip()
        if re.match(r'^(fatu|farur|cart)', name, re.IGNORECASE):
            continue
        if re.search(r'(fatura|fatuar|fatut|farur|fartur)', name, re.IGNORECASE):
            continue
        if re.search(r'(cartão corporativo|cartao corporativo)', name, re.IGNORECASE):
            continue
        if re.search(r'itau|itaú', name, re.IGNORECASE):
            continue
        if e["user_cpf"] not in vexpenses_cpfs:
            continue
        result.append(e)
    return result

# ============================================================
# Run scenarios
# ============================================================

scenarios = [
    ("Baseline (current regex)", filter_current),
    ("+ 'fartur' exclusion", filter_fartur),
    ("+ 'cartão corporativo' exclusion", filter_cartao_corp),
    ("+ CARTÃO VEXPENSES='SIM' filter", filter_vexpenses_card),
    ("All filters combined", filter_all),
    ("All + ITAU name exclusion", filter_all_plus_itau),
]

print("\n" + "=" * 100)
print(f"{'Scenario':<40} {'Reports':>8} {'Expenses':>9} {'Total R$':>14} {'Δ Reports':>10} {'Δ Expenses':>11} {'Δ Value R$':>14}")
print("=" * 100)

for name, filter_fn in scenarios:
    filtered = filter_fn(all_api)
    
    api_report_ids = set(e["report_id"] for e in filtered)
    api_expense_ids = set(e["id"] for e in filtered)
    api_total = sum(float(e["value"]) for e in filtered)
    
    # Divergence vs reference
    new_reports = api_report_ids - ref_report_ids
    missing_reports = ref_report_ids - api_report_ids
    new_expenses = api_expense_ids - ref_expense_ids
    missing_expenses = ref_expense_ids - api_expense_ids
    
    delta_reports = len(new_reports) - len(missing_reports)
    delta_expenses = len(new_expenses) - len(missing_expenses)
    delta_value = api_total - ref_total_value
    
    print(f"{name:<40} {len(api_report_ids):>8} {len(api_expense_ids):>9} {api_total:>14,.2f} {delta_reports:>+10} {delta_expenses:>+11} {delta_value:>+14,.2f}")

# ============================================================
# Detailed breakdown for the "All filters combined" scenario
# ============================================================
print("\n" + "=" * 100)
print("DETAILED: All filters combined")
print("=" * 100)

filtered = filter_all(all_api)
api_report_ids = set(e["report_id"] for e in filtered)
api_expense_ids = set(e["id"] for e in filtered)
api_total = sum(float(e["value"]) for e in filtered)

new_reports = api_report_ids - ref_report_ids
missing_reports = ref_report_ids - api_report_ids
new_expenses = api_expense_ids - ref_expense_ids
missing_expenses = ref_expense_ids - api_expense_ids

print(f"\n  API after filters: {len(api_report_ids)} reports, {len(api_expense_ids)} expenses, R$ {api_total:,.2f}")
print(f"  Reference:         {len(ref_report_ids)} reports, {len(ref_expense_ids)} expenses, R$ {ref_total_value:,.2f}")
print(f"\n  New reports (in API, not in ref): {len(new_reports)}")
print(f"  Missing reports (in ref, not in API): {len(missing_reports)}")
print(f"  New expenses (in API, not in ref): {len(new_expenses)}")
print(f"  Missing expenses (in ref, not in API): {len(missing_expenses)}")
print(f"  Value gap: R$ {api_total - ref_total_value:,.2f}")

# Show new reports
if new_reports:
    print(f"\n  --- New reports (in API, NOT in reference) ---")
    new_report_data = {}
    for e in filtered:
        if e["report_id"] in new_reports:
            if e["report_id"] not in new_report_data:
                new_report_data[e["report_id"]] = {"name": e["report_name"], "user": e["user_name"], "cpf": e["user_cpf"], "count": 0, "total": 0}
            new_report_data[e["report_id"]]["count"] += 1
            new_report_data[e["report_id"]]["total"] += float(e["value"])
    
    for rid in sorted(new_report_data.keys()):
        r = new_report_data[rid]
        print(f"    {rid}  {str(r['name'])[:35]:<35}  {str(r['user'])[:25]:<25}  CPF={r['cpf']}  {r['count']:>3} items  R$ {r['total']:>10,.2f}")

# Show missing reports
if missing_reports:
    print(f"\n  --- Missing reports (in reference, NOT in API) ---")
    # Load ref report names
    wb2 = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
    ws2 = wb2["BASE PREST "]
    ref_report_info = {}
    for row in ws2.iter_rows(min_row=4, values_only=True):
        if row[0] is None:
            continue
        rid = int(row[1]) if row[1] else None
        if rid and rid in missing_reports:
            if rid not in ref_report_info:
                ref_report_info[rid] = {"name": row[2], "user": row[4], "cpf": row[9], "count": 0, "total": 0}
            ref_report_info[rid]["count"] += 1
            ref_report_info[rid]["total"] += float(row[26] or 0) if len(row) > 26 else 0
    wb2.close()
    
    for rid in sorted(ref_report_info.keys()):
        r = ref_report_info[rid]
        print(f"    {rid}  {str(r['name'])[:35]:<35}  {str(r['user'])[:25]:<25}  CPF={r['cpf']}  {r['count']:>3} items  R$ {r['total']:>10,.2f}")

# ============================================================
# Also check: per-CPF divergence with all filters
# ============================================================
print("\n" + "=" * 100)
print("Per-CPF divergence (All filters combined) — top 20 by absolute gap")
print("=" * 100)

api_by_cpf = defaultdict(lambda: {"count": 0, "total": 0})
for e in filtered:
    api_by_cpf[e["user_cpf"]]["count"] += 1
    api_by_cpf[e["user_cpf"]]["total"] += float(e["value"])

all_cpfs = set(api_by_cpf.keys()) | set(ref_by_cpf.keys())
gaps = []
for cpf in all_cpfs:
    api_val = api_by_cpf.get(cpf, {"total": 0})["total"]
    ref_val = ref_by_cpf.get(cpf, {"total": 0})["total"]
    gap = api_val - ref_val
    if abs(gap) > 0.01:
        gaps.append((cpf, gap, api_val, ref_val))

gaps.sort(key=lambda x: abs(x[1]), reverse=True)
print(f"\n  Total CPFs with divergence: {len(gaps)}")
print(f"  {'CPF':<15} {'API R$':>12} {'Ref R$':>12} {'Gap R$':>12}")
for cpf, gap, api_val, ref_val in gaps[:20]:
    print(f"  {cpf:<15} {api_val:>12,.2f} {ref_val:>12,.2f} {gap:>+12,.2f}")

# Summary
total_gap = sum(g[1] for g in gaps)
print(f"\n  Total gap: R$ {total_gap:,.2f}")
print(f"  Positive gaps (API > Ref): R$ {sum(g[1] for g in gaps if g[1] > 0):,.2f}")
print(f"  Negative gaps (API < Ref): R$ {sum(g[1] for g in gaps if g[1] < 0):,.2f}")
