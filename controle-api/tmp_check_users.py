#!/usr/bin/env python3
"""
Check if the 20 users with no reports in reference BASE PREST
appear in the reference PAINEL with a PRESTACAO value.
Also check if they appear in the reference BASE PREST with a different name.
"""
import openpyxl
from collections import defaultdict

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"

# Users with no reports in reference
users_no_ref = [
    "AFONSO FIORELLO CARVALHO",
    "AILTON MENDES AGUILAR",
    "AMARAL RODRIGUES NUNES",
    "ANA LUIZA OLIVEIRA CAMELO",
    "ATILA SILVA DOS SANTOS",
    "DANIEL PORFIRIO DE SOUSA",
    "ERASMO PEREIRA MONTEIRO",
    "FLAVIO PEREIRA DE SOUZA",
    "FRANCISCO MICHAEL CASTRO FERNANDES",
    "JACKSON CAROLINO CARNEIRO",
    "LEONARDO GONCALVES RIBEIRO FILHO",
    "MARCELO BEZERRA",
    "NELSON MOURA RIBEIRO",
    "OSMAN DOS SANTOS MORAIS JUNIOR",
    "PATRICK FERNANDO GOULART ALVES",
    "PEDRO LUIS PIRES DOS SANTOS",
    "ROBERTO ALIAGA",
    "SANDRA CRISTINA ALVES MACHADO",
    "WAGNER FERNANDES DA SILVA",
    "WESLEY CARLOS AUGUSTO",
]

# Normalize for matching
def norm(s):
    return " ".join(str(s or "").upper().strip().split())

users_norm = {norm(u): u for u in users_no_ref}

wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)

# 1. Check PAINEL
print("=" * 80)
print("Check: Are these 20 users in the reference PAINEL?")
print("=" * 80)
ws_p = wb["PAINEL"]
# PAINEL header is at row 11, data starts at row 12
# Col C = CPF, Col D = Colaborador, Col Q = PRESTACAO
painel_found = {}
for row in ws_p.iter_rows(min_row=12, values_only=True):
    if row[2] is None and row[3] is None:
        continue
    cpf = str(row[2] or "").strip()
    colaborador = str(row[3] or "").strip()
    prestacao = float(row[16] or 0) if len(row) > 16 else 0  # Col Q = index 16
    colab_norm = norm(colaborador)
    
    # Check if this matches any of our users
    for un, orig in users_norm.items():
        if un in colab_norm or colab_norm in un:
            painel_found[orig] = {"cpf": cpf, "prestacao": prestacao, "colaborador": colaborador}
            break

for user in users_no_ref:
    if user in painel_found:
        info = painel_found[user]
        print(f"  FOUND  {user[:35]:<35}  CPF={info['cpf']}  PRESTACAO=R$ {info['prestacao']:>10,.2f}")
    else:
        print(f"  MISSING {user[:35]:<35}  (not in PAINEL)")

# 2. Check BASE PREST for these users (by name, not by report ID)
print("\n" + "=" * 80)
print("Check: Are these users in the reference BASE PREST at all (by name)?")
print("=" * 80)
ws_bp = wb["BASE PREST "]
# Col 4 = Nome do membro de equipe, Col 9 = CPF/CNPJ, Col 26 = Valor
bp_found = defaultdict(lambda: {"count": 0, "total": 0, "cpfs": set()})
for row in ws_bp.iter_rows(min_row=4, values_only=True):
    if row[0] is None:
        continue
    user_name = str(row[4] or "").strip()
    cpf = str(row[9] or "").strip()
    valor = float(row[26] or 0) if len(row) > 26 else 0
    
    user_norm = norm(user_name)
    for un, orig in users_norm.items():
        if un in user_norm or user_norm in un:
            bp_found[orig]["count"] += 1
            bp_found[orig]["total"] += valor
            bp_found[orig]["cpfs"].add(cpf)
            break

for user in users_no_ref:
    if user in bp_found:
        info = bp_found[user]
        print(f"  FOUND  {user[:35]:<35}  {info['count']:>4} expenses  R$ {info['total']:>10,.2f}  CPFs={info['cpfs']}")
    else:
        print(f"  MISSING {user[:35]:<35}  (not in BASE PREST)")

# 3. Check: what are the PAINEL PRESTACAO values for users that ARE in both?
print("\n" + "=" * 80)
print("Check: PAINEL PRESTACAO for users with reports in both ref and API")
print("=" * 80)

# Users that have reports in ref but also new reports
users_partial = [
    "ADSON ARAUJO DA SILVA",
    "BRUNO EDUARDO DE SOUZA",
    "CLAUDINEI VICENTE DE LIMA",
    "EMERSON LEITE DE BARROS",
    "FABIO PROVESI",
    "HELDER ARCHANJO FERREIRA",
    "JOSE CARLOS BATISTA",
    "NIVALDO THIAGO DA CRUZ SANTOS",
    "TIAGO DE JESUS ARAUJO",
]

for row in ws_p.iter_rows(min_row=12, values_only=True):
    if row[2] is None and row[3] is None:
        continue
    colaborador = str(row[3] or "").strip()
    prestacao = float(row[16] or 0) if len(row) > 16 else 0
    colab_norm = norm(colaborador)
    
    for un in users_partial:
        if un in colab_norm or colab_norm in un:
            print(f"  {colaborador[:35]:<35}  PRESTACAO=R$ {prestacao:>10,.2f}")
            break

wb.close()
