#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Inspect carga_final and saldo_final columns in problematic sheets."""
import openpyxl
import os

DATA_DIR = r'C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\data'

SHEETS = [
    (1, 1, os.path.join(DATA_DIR, "01 - JANEIRO", "1QZ JANEIRO 2026 - VEXPENSES.xlsx"), "1 QZ VEXPENSES 01_2026", 6),
    (5, 1, os.path.join(DATA_DIR, "05 - MAIO", "CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx"), "Planilha1", 6),
    (5, 2, os.path.join(DATA_DIR, "05 - MAIO", "CARGA 2 QZ MAIO 26 VEXPENSES EQS.xlsx"), "2 QZ DE MAIO 26", 4),
    (6, 1, os.path.join(DATA_DIR, "06 - JUNHO", "CARGA 1 QZ JUNHO 26 VEXPENSES EQS.xlsx"), "1 QZ JUNHO", 6),
    (2, 1, os.path.join(DATA_DIR, "02 - FEVEREIRO", "1 QZN FEVEREIRO VEXPENSES 2026.xlsx"), "1 QZN FEV 2026", 6),
    (3, 1, os.path.join(DATA_DIR, "03 - MARÇO", "1 QZ MARÇO VEXPENSES 2026 (5).xlsx"), "QUINZENA MARÇO", 6),
]

for month, qz, path, sheet_name, header_row in SHEETS:
    print(f"\n{'='*80}")
    print(f"Month {month} QZ{qz}: {os.path.basename(path)}")
    print(f"  Sheet: '{sheet_name}', header row: {header_row}")
    print(f"{'='*80}")

    if not os.path.exists(path):
        print("  FILE NOT FOUND")
        continue

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"  Sheet not found. Available: {wb.sheetnames}")
        wb.close()
        continue

    ws = wb[sheet_name]

    # Print header row
    for i, row in enumerate(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), start=header_row):
        print(f"  Header (row {i}):")
        for j, val in enumerate(row):
            if val is not None and str(val).strip():
                print(f"    col {j} (xlsx {j+1}): {val}")

    # Print first 3 data rows
    for i, row in enumerate(ws.iter_rows(min_row=header_row+1, max_row=header_row+3, values_only=True), start=header_row+1):
        print(f"\n  Data row {i}:")
        for j, val in enumerate(row):
            if val is not None and str(val).strip() != '':
                print(f"    col {j} (xlsx {j+1}): {val}")

    wb.close()
