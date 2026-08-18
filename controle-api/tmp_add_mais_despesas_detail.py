#!/usr/bin/env python3
"""
Add a 'MAIS DESPESAS DETAIL' sheet to the existing itau excel with per-expense breakdown
for the 22 'mais despesas' reports, showing which extra expenses are Itaú vs not.
"""
import os, requests, json, time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from pathlib import Path

load_dotenv(Path(__file__).parent / ".env")
API_KEY = os.getenv("VEXPENSES_API_KEY", "")
BASE_URL = os.getenv("VEXPENSES_BASE_URL", "https://api.vexpenses.com")
HEADERS = {"Authorization": API_KEY, "Accept": "application/json"}

# Read MAIS DESPESAS from the generated excel
xlsx_in = Path(r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\controle-api\data\gap_analysis_itau_detail.xlsx")
wb = openpyxl.load_workbook(xlsx_in, read_only=True, data_only=True)
ws = wb["MAIS DESPESAS + ITAU"]
mais_neon = []
for row in ws.iter_rows(values_only=True):
    if row[0] is None or row[0] == "Report ID":
        continue
    mais_neon.append({
        "report_id": row[0], "name": row[1], "user": row[2], "status": row[3],
        "ref_count": row[4], "neon_count": row[5], "diff_count": row[6],
        "ref_total": row[7], "neon_total": row[8], "diff_value": row[9],
        "api_status": row[10], "n_itau": row[11], "n_non_itau": row[12], "category": row[13],
    })
wb.close()

# Styles
itau_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

def style_header(ws, ncols):
    for col in range(1, ncols+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def auto_width(ws, max_width=55):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value))
                if val_len > max_len:
                    max_len = val_len
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

# Open existing excel and add sheets
wb_out = openpyxl.load_workbook(xlsx_in)

# Sheet: MAIS DESPESAS EXPENSE DETAIL
ws_new = wb_out.create_sheet("MAIS DESPESAS EXPENSES")
headers = ["Report ID", "Report Name", "User", "API Status", "Expense ID", "Date", "Title", "Value", "Payment Method", "Is Itaú?", "Extra?"]
ws_new.append(headers)

for r in sorted(mais_neon, key=lambda x: -x["diff_value"]):
    rid = r["report_id"]
    time.sleep(0.4)
    try:
        resp = requests.get(f"{BASE_URL}/v2/reports/{rid}?include=expenses.payment_method", headers=HEADERS, timeout=30)
        if resp.status_code != 200:
            continue
        data = resp.json().get("data", {})
        api_status = data.get("status", "")
        expenses = data.get("expenses", {}).get("data", [])
        
        # Sort by value desc
        expenses.sort(key=lambda e: -float(e.get("value", 0) or 0))
        
        for exp in expenses:
            pm = exp.get("payment_method", {})
            if isinstance(pm, dict):
                pm_name = pm.get("data", {}).get("description", f"ID:{exp.get('payment_method_id')}")
            else:
                pm_name = f"ID:{exp.get('payment_method_id')}"
            is_itau = "itau" in pm_name.lower() or "itaú" in pm_name.lower()
            
            ws_new.append([rid, r["name"], r["user"], api_status,
                          exp.get("id"), exp.get("date", ""), exp.get("title", ""),
                          float(exp.get("value", 0) or 0), pm_name,
                          "YES" if is_itau else "NO", ""])
            
            row_idx = ws_new.max_row
            if is_itau:
                for col in range(1, len(headers)+1):
                    ws_new.cell(row=row_idx, column=col).fill = itau_fill
    except:
        continue

style_header(ws_new, len(headers))
auto_width(ws_new)

# Sheet: MAIS DESPESAS SUMMARY with Itaú breakdown
ws_sum = wb_out.create_sheet("MAIS DESPESAS SUMMARY")
headers2 = ["Report ID", "Name", "User", "Status", "API Status", "Ref Count", "Neon Count", "Diff Count", "Ref Total", "Neon Total", "Diff Value", "N Itaú Exp", "N Non-Itaú", "Extra Itaú?", "Explanation"]
ws_sum.append(headers2)

for r in sorted(mais_neon, key=lambda x: -x["diff_value"]):
    rid = r["report_id"]
    time.sleep(0.4)
    try:
        resp = requests.get(f"{BASE_URL}/v2/reports/{rid}?include=expenses.payment_method", headers=HEADERS, timeout=30)
        if resp.status_code != 200:
            continue
        data = resp.json().get("data", {})
        api_status = data.get("status", "")
        expenses = data.get("expenses", {}).get("data", [])
        
        itau_count = sum(1 for e in expenses if "itau" in str(e.get("payment_method", {}).get("data", {}).get("description", "")).lower())
        non_itau_count = len(expenses) - itau_count
        
        if api_status == "ENVIADO":
            explanation = "ENVIADO: expenses added after ref snapshot (report still pending)"
        elif itau_count == len(expenses) and len(expenses) > 0:
            explanation = "ALL ITAÚ: extra expenses are Itaú card (should be excluded)"
        elif itau_count > 0:
            explanation = f"MIXED: {itau_count} Itaú + {non_itau_count} non-Itaú — extra may or may not be Itaú"
        else:
            explanation = "NON-ITAÚ: extra expenses are genuine VExpenses (added after ref)"
        
        extra_itau = "YES (all)" if (itau_count == len(expenses) and len(expenses) > 0) else ("MIXED" if itau_count > 0 else "NO")
        
        ws_sum.append([rid, r["name"], r["user"], r["status"], api_status,
                      r["ref_count"], r["neon_count"], r["diff_count"],
                      r["ref_total"], r["neon_total"], r["diff_value"],
                      itau_count, non_itau_count, extra_itau, explanation])
    except:
        continue

style_header(ws_sum, len(headers2))
auto_width(ws_sum)

# Save
wb_out.save(xlsx_in)
print(f"Updated: {xlsx_in}")
print(f"Sheets: {wb_out.sheetnames}")
