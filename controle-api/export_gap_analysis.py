#!/usr/bin/env python3
"""
Export full gap analysis between API data and reference BASE PREST to Excel.
Uses the best filter (All + ITAU name exclusion) and exports every detail.
"""
import os, re
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from collections import defaultdict

BASE = Path(__file__).parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"
OUT_PATH = BASE.parent / "data" / "GAP_ANALYSIS_2QZ_JULHO_2026.xlsx"


CARD_PM_IDS = {'627401', '627508', '627721', '668240'}


def is_card_report(name):
    """Legacy name-based FATURA/CARTAO filter. Kept for 'Filtered Out' sheet analysis."""
    n = name.strip().upper()
    if not n:
        return False
    if 'CAIXA ITAU' in n or 'CAIXA ITAÚ' in n:
        return True
    if n.startswith('CAIXA'):
        return False
    if n.startswith(('FATURA', 'CARTAO', 'CARTÃO', 'FATUAR', 'FARTUR', 'FATUT', 'FARUR', 'FATUTR')):
        return True
    if 'CARTÃO DE CRÉDITO' in n or 'CARTAO DE CREDITO' in n or 'CARTÃO DE CREDITO' in n:
        return True
    if 'CARTÃO CORPORATIVO' in n:
        return True
    if ('ITAU' in n or 'ITAÚ' in n) and 'CAIXA' not in n:
        return True
    if 'DOLAR' in n or 'DÓLAR' in n:
        return True
    if n.startswith('DESPESA') and 'FATURA' in n:
        return True
    if n.startswith('COMPLEMENTAR') and 'FATURA' in n:
        return True
    if 'CARTÃO' in n and 'CRÉDITO' in n:
        return True
    if 'CARTAO' in n and 'CREDITO' in n:
        return True
    if n.startswith('CARTÃO VEXPENSES'):
        return True
    return False


def main():
    print("Loading reference BASE PREST...")
    wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
    ws_bp = wb["BASE PREST "]

    ref_report_ids = set()
    ref_expense_ids = set()
    ref_total_value = 0.0
    ref_by_cpf = defaultdict(lambda: {"count": 0, "total": 0.0})
    ref_expenses = []
    ref_report_info = {}

    for row in ws_bp.iter_rows(min_row=4, values_only=True):
        if row[0] is None:
            continue
        eid = int(row[0])
        rid = int(row[1]) if row[1] else None
        r_name = str(row[2] or "").strip() if len(row) > 2 else ""
        r_user = str(row[4] or "").strip() if len(row) > 4 else ""
        cpf = str(row[9] or "").strip() if len(row) > 9 else ""
        valor = float(row[26] or 0) if len(row) > 26 else 0.0

        ref_expense_ids.add(eid)
        if rid:
            ref_report_ids.add(rid)
            if rid not in ref_report_info:
                ref_report_info[rid] = {"name": r_name, "user": r_user, "cpf": cpf, "count": 0, "total": 0.0}
            ref_report_info[rid]["count"] += 1
            ref_report_info[rid]["total"] += valor
        ref_total_value += valor
        ref_by_cpf[cpf]["count"] += 1
        ref_by_cpf[cpf]["total"] += valor
        ref_expenses.append({
            "expense_id": eid, "report_id": rid, "report_name": r_name,
            "user_name": r_user, "cpf": cpf, "value": valor,
        })

    print(f"  Reference: {len(ref_report_ids)} reports, {len(ref_expense_ids)} expenses, R$ {ref_total_value:,.2f}")

    # Load PAINEL CPFs
    ws_p = wb["PAINEL"]
    vexpenses_cpfs = set()
    all_painel_cpfs = set()
    painel_names = {}
    for row in ws_p.iter_rows(min_row=12, values_only=True):
        if row[2] is None:
            continue
        cpf = str(row[2] or "").strip()
        cartao_vx = str(row[12] or "").strip().upper() if len(row) > 12 else ""
        nome = str(row[1] or "").strip() if len(row) > 1 else ""
        all_painel_cpfs.add(cpf)
        painel_names[cpf] = nome
        if cartao_vx == "SIM":
            vexpenses_cpfs.add(cpf)
    print(f"  PAINEL: {len(all_painel_cpfs)} total CPFs, {len(vexpenses_cpfs)} with CARTAO VEXPENSES='SIM'")
    wb.close()

    print("\nLoading API data...")
    conn = psycopg2.connect(NEON_URL, connect_timeout=10)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT e.id, e.report_id, e.date, e.value, e.description, e.status as expense_status,
               r.name as report_name, r.status as report_status,
               r.user_cpf, r.user_name, r.user_id, r.created_at as report_created_at,
               r.updated_at as report_updated_at, r.total_value as report_total_value,
               r.raw_data,
               e.raw_data->>'payment_method_id' as pm_id,
               e.raw_data->>'payment_method_name' as pm_name
        FROM prestacao_expenses e
        JOIN prestacao_reports r ON e.report_id = r.id
        WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
          AND r.user_cpf IS NOT NULL
        ORDER BY r.user_name, r.name, e.date
    """)
    all_api = cur.fetchall()
    conn.close()
    print(f"  API (all, no filters): {len(all_api)} expenses, {len(set(e['report_id'] for e in all_api))} reports, R$ {sum(float(e['value']) for e in all_api):,.2f}")

    # Build pm_id lookup per report for informational columns
    from collections import defaultdict as _dd
    report_pm_ids = _dd(set)
    for e in all_api:
        pm = e.get("pm_id")
        if pm:
            report_pm_ids[e["report_id"]].add(pm)

    # Apply filter: exclude FATURA/CARTAO statement reports (name-based)
    # + exclude individual expenses with pm_id=627401 (Cartão Itaú) even inside non-FATURA reports
    # + CARTAO VEXPENSES='SIM' PAINEL filter
    filtered = []
    for e in all_api:
        name = e["report_name"] or ""
        if is_card_report(name):
            continue
        if e.get("pm_id") == "627401":
            continue
        if e["user_cpf"] not in vexpenses_cpfs:
            continue
        filtered.append(e)

    api_report_ids = set(e["report_id"] for e in filtered)
    api_expense_ids = set(e["id"] for e in filtered)
    api_total = sum(float(e["value"]) for e in filtered)

    new_reports = api_report_ids - ref_report_ids
    missing_reports = ref_report_ids - api_report_ids
    new_expenses = api_expense_ids - ref_expense_ids
    missing_expenses = ref_expense_ids - api_expense_ids

    print(f"\n  API after filters: {len(api_report_ids)} reports, {len(api_expense_ids)} expenses, R$ {api_total:,.2f}")
    print(f"  Reference:         {len(ref_report_ids)} reports, {len(ref_expense_ids)} expenses, R$ {ref_total_value:,.2f}")
    print(f"  Gap: R$ {api_total - ref_total_value:,.2f}")

    # Build per-report data for new reports
    new_report_data = {}
    for e in filtered:
        if e["report_id"] in new_reports:
            if e["report_id"] not in new_report_data:
                raw = e["raw_data"] or {}
                if isinstance(raw, str):
                    import json as _json
                    raw = _json.loads(raw)
                new_report_data[e["report_id"]] = {
                    "name": e["report_name"], "user": e["user_name"], "cpf": e["user_cpf"],
                    "user_id": e["user_id"],
                    "count": 0, "total": 0.0, "status": e["report_status"],
                    "created_at": str(e["report_created_at"] or ""),
                    "updated_at": str(e["report_updated_at"] or ""),
                    "report_total_value": float(e["report_total_value"] or 0),
                    "pdf_link": raw.get("pdf_link", ""),
                    "excel_link": raw.get("excel_link", ""),
                    "approval_date": raw.get("approval_date", ""),
                    "observation": raw.get("observation", ""),
                    "justification": raw.get("justification", ""),
                }
            new_report_data[e["report_id"]]["count"] += 1
            new_report_data[e["report_id"]]["total"] += float(e["value"])

    # Per-CPF divergence
    api_by_cpf = defaultdict(lambda: {"count": 0, "total": 0.0})
    for e in filtered:
        api_by_cpf[e["user_cpf"]]["count"] += 1
        api_by_cpf[e["user_cpf"]]["total"] += float(e["value"])

    all_cpfs = set(api_by_cpf.keys()) | set(ref_by_cpf.keys())
    gaps = []
    for cpf in all_cpfs:
        api_val = api_by_cpf.get(cpf, {"total": 0.0})["total"]
        ref_val = ref_by_cpf.get(cpf, {"total": 0.0})["total"]
        gap = api_val - ref_val
        if abs(gap) > 0.01:
            gaps.append((cpf, gap, api_val, ref_val,
                        api_by_cpf.get(cpf, {"count": 0})["count"],
                        ref_by_cpf.get(cpf, {"count": 0})["count"]))
    gaps.sort(key=lambda x: abs(x[1]), reverse=True)

    # New expense details
    new_exp_details = [e for e in filtered if e["id"] in new_expenses]
    # Missing expense details
    missing_exp_details = [e for e in ref_expenses if e["expense_id"] in missing_expenses]

    # ================================================================
    # Write Excel
    # ================================================================
    print(f"\nWriting Excel: {OUT_PATH}")
    wb_out = openpyxl.Workbook()

    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", fgColor="1F4E79")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    money_fmt = '#,##0.00;[Red]-#,##0.00'
    pos_fill = PatternFill("solid", fgColor="D4EDDA")
    neg_fill = PatternFill("solid", fgColor="F8D7DA")

    def write_header(ws_out, headers, row=1):
        for ci, h in enumerate(headers, 1):
            cell = ws_out.cell(row=row, column=ci, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = halign

    def auto_width(ws_out):
        for col in ws_out.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except Exception:
                    pass
            ws_out.column_dimensions[col_letter].width = min(max_len + 2, 40)

    # --- Sheet 1: Summary ---
    ws1 = wb_out.active
    ws1.title = "Summary"
    write_header(ws1, ["Metric", "API", "Reference", "Gap"], row=1)
    rows_data = [
        ("Reports", len(api_report_ids), len(ref_report_ids), len(api_report_ids) - len(ref_report_ids)),
        ("Expenses", len(api_expense_ids), len(ref_expense_ids), len(api_expense_ids) - len(ref_expense_ids)),
        ("Total Value", api_total, ref_total_value, api_total - ref_total_value),
        ("", "", "", ""),
        ("New reports (API only)", len(new_reports), "", ""),
        ("Missing reports (Ref only)", len(missing_reports), "", ""),
        ("New expenses (API only)", len(new_expenses), "", ""),
        ("Missing expenses (Ref only)", len(missing_expenses), "", ""),
        ("", "", "", ""),
        ("CPFs with divergence", len(gaps), "", ""),
        ("Positive gap (API > Ref)", "", "", f"R$ {sum(g[1] for g in gaps if g[1] > 0):,.2f}"),
        ("Negative gap (API < Ref)", "", "", f"R$ {sum(g[1] for g in gaps if g[1] < 0):,.2f}"),
        ("Net gap", "", "", f"R$ {sum(g[1] for g in gaps):,.2f}"),
    ]
    for ri, row in enumerate(rows_data, 2):
        for ci, val in enumerate(row, 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            if ci == 4 and isinstance(val, str) and "R$" in str(val):
                cell.font = Font(bold=True)
    auto_width(ws1)

    # Count reports per CPF (needed by multiple sheets)
    api_reports_by_cpf = defaultdict(set)
    ref_reports_by_cpf = defaultdict(set)
    for e in filtered:
        api_reports_by_cpf[e["user_cpf"]].add(e["report_id"])
    for rid, rinfo in ref_report_info.items():
        ref_reports_by_cpf[rinfo["cpf"]].add(rid)

    # --- Sheet 2: New Reports (in API, not in Ref) ---
    ws2 = wb_out.create_sheet("New Reports (API only)")
    write_header(ws2, ["Report ID", "PDF Link", "Report Name", "User Name", "User ID", "CPF", "Expenses", "Expense Sum R$", "Report Total R$", "Status", "Created At", "Approval Date", "In PAINEL?", "CARTAO VEXPENSES", "Ref Total for CPF", "Ref Reports for CPF", "Observation", "Justification"])
    for ri, (rid, r) in enumerate(sorted(new_report_data.items(), key=lambda x: x[1]["total"], reverse=True), 2):
        in_painel = "YES" if r["cpf"] in all_painel_cpfs else "NO"
        has_card = "SIM" if r["cpf"] in vexpenses_cpfs else "NAO"
        ref_cpf_total = ref_by_cpf.get(r["cpf"], {}).get("total", 0.0)
        ref_cpf_reports = len(ref_reports_by_cpf.get(r["cpf"], set()))
        pdf_link = r.get("pdf_link", "")
        vals = [rid, pdf_link or "(no PDF)", r["name"], r["user"], r["user_id"], r["cpf"],
                r["count"], r["total"], r["report_total_value"], r["status"],
                r["created_at"], r.get("approval_date", ""),
                in_painel, has_card, ref_cpf_total, ref_cpf_reports,
                r.get("observation", ""), r.get("justification", "")]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            if ci in (8, 9, 15):
                cell.number_format = money_fmt
            if ci == 2 and pdf_link:
                cell.hyperlink = pdf_link
                cell.font = Font(color="0563C1", underline="single")
    auto_width(ws2)

    # --- Sheet 3: Missing Reports (in Ref, not in API) ---
    ws3 = wb_out.create_sheet("Missing Reports (Ref only)")
    write_header(ws3, ["Report ID", "Report Name", "User Name", "CPF", "Expenses", "Total R$", "In PAINEL?", "CARTAO VEXPENSES", "Possible Reason", "API Total for CPF", "API Reports for CPF"])
    for ri, rid in enumerate(sorted(missing_reports), 2):
        r = ref_report_info.get(rid, {"name": "?", "user": "?", "cpf": "", "count": 0, "total": 0})
        in_painel = "YES" if r["cpf"] in all_painel_cpfs else "NO"
        has_card = "SIM" if r["cpf"] in vexpenses_cpfs else "NAO"
        # Try to determine why it's missing
        r_name_upper = (r["name"] or "").strip().upper()
        if is_card_report(r_name_upper):
            reason = "Filtered as FATURA/CARTAO"
        elif r["cpf"] not in vexpenses_cpfs:
            reason = "CARTAO VEXPENSES != SIM"
        elif "DESATIVADO" in (r["user"] or "").upper():
            reason = "User DESATIVADO"
        else:
            reason = "Not in API (deleted or status changed)"
        api_cpf_total = api_by_cpf.get(r["cpf"], {}).get("total", 0.0)
        api_cpf_reports = len(api_reports_by_cpf.get(r["cpf"], set()))
        vals = [rid, r["name"], r["user"], r["cpf"], r["count"], r["total"],
                in_painel, has_card, reason, api_cpf_total, api_cpf_reports]
        for ci, v in enumerate(vals, 1):
            cell = ws3.cell(row=ri, column=ci, value=v)
            if ci in (6, 10):
                cell.number_format = money_fmt
    auto_width(ws3)

    # --- Sheet 4: Per-CPF Divergence ---
    ws4 = wb_out.create_sheet("Per-CPF Divergence")
    write_header(ws4, ["CPF", "Colaborador", "API Total R$", "Ref Total R$", "Gap R$", "API Expenses", "Ref Expenses", "API Reports", "Ref Reports", "In PAINEL?", "CARTAO VEXPENSES"])
    for ri, (cpf, gap, api_val, ref_val, api_cnt, ref_cnt) in enumerate(gaps, 2):
        nome = painel_names.get(cpf, "")
        in_painel = "YES" if cpf in all_painel_cpfs else "NO"
        has_card = "SIM" if cpf in vexpenses_cpfs else "NAO"
        api_rpts = len(api_reports_by_cpf.get(cpf, set()))
        ref_rpts = len(ref_reports_by_cpf.get(cpf, set()))
        vals = [cpf, nome, api_val, ref_val, gap, api_cnt, ref_cnt, api_rpts, ref_rpts, in_painel, has_card]
        for ci, v in enumerate(vals, 1):
            cell = ws4.cell(row=ri, column=ci, value=v)
            if ci in (3, 4, 5):
                cell.number_format = money_fmt
            if ci == 5:
                cell.fill = pos_fill if gap > 0 else neg_fill
    auto_width(ws4)

    # --- Sheet 5: New Expenses Detail (in API, not in Ref) ---
    ws5 = wb_out.create_sheet("New Expenses (API only)")
    write_header(ws5, ["Expense ID", "Report ID", "PDF Link", "Report Name", "User Name", "CPF", "Date", "Value R$", "Description", "Expense Status", "Report Status", "Approval Date"])
    for ri, e in enumerate(sorted(new_exp_details, key=lambda x: abs(float(x["value"])), reverse=True), 2):
        raw = e.get("raw_data") or {}
        if isinstance(raw, str):
            import json as _json
            raw = _json.loads(raw)
        pdf_link = raw.get("pdf_link", "") if raw else ""
        approval_date = raw.get("approval_date", "") if raw else ""
        vals = [e["id"], e["report_id"], pdf_link or "(no PDF)", e["report_name"], e["user_name"], e["user_cpf"],
                str(e["date"] or ""), float(e["value"]), e["description"], e["expense_status"], e["report_status"],
                approval_date]
        for ci, v in enumerate(vals, 1):
            cell = ws5.cell(row=ri, column=ci, value=v)
            if ci == 8:
                cell.number_format = money_fmt
            if ci == 3 and pdf_link:
                cell.hyperlink = pdf_link
                cell.font = Font(color="0563C1", underline="single")
    auto_width(ws5)

    # --- Sheet 6: Missing Expenses Detail (in Ref, not in API) ---
    ws6 = wb_out.create_sheet("Missing Expenses (Ref only)")
    write_header(ws6, ["Expense ID", "Report ID", "Report Link", "Report Name", "User Name", "CPF", "Value R$", "Possible Reason"])
    for ri, e in enumerate(sorted(missing_exp_details, key=lambda x: abs(x["value"]), reverse=True), 2):
        link = f"https://app.vexpenses.com/relatorios/{e['report_id']}" if e["report_id"] else ""
        r_name_upper = (e["report_name"] or "").strip().upper()
        if is_card_report(r_name_upper):
            reason = "Filtered as FATURA/CARTAO"
        elif e["cpf"] not in vexpenses_cpfs:
            reason = "CARTAO VEXPENSES != SIM"
        elif "DESATIVADO" in (e["user_name"] or "").upper():
            reason = "User DESATIVADO"
        else:
            reason = "Not in API (deleted or status changed)"
        vals = [e["expense_id"], e["report_id"], link, e["report_name"], e["user_name"], e["cpf"], e["value"], reason]
        for ci, v in enumerate(vals, 1):
            cell = ws6.cell(row=ri, column=ci, value=v)
            if ci == 7:
                cell.number_format = money_fmt
            if ci == 3 and link:
                cell.hyperlink = link

    # --- Sheet 7: Filtered Out by Card Filter ---
    ws7 = wb_out.create_sheet("Filtered Out (Card-Fatura)")
    write_header(ws7, ["Report ID", "PDF Link", "Report Name", "User Name", "CPF", "Status", "Expenses", "Total R$", "Reason", "PM IDs", "Approval Date"])
    card_filtered = {}
    for e in all_api:
        rid = e["report_id"]
        name = e["report_name"] or ""
        if not is_card_report(name):
            continue
        if e["user_cpf"] not in vexpenses_cpfs:
            continue
        if rid not in card_filtered:
            n = name.strip().upper()
            pms = report_pm_ids.get(rid, set())
            pm_str = ",".join(sorted(pms)) if pms else "(none)"
            reason = "FATURA/CARTAO" if n.startswith(('FATURA', 'CARTAO', 'CARTÃO', 'FATUAR', 'FARTUR', 'FATUT', 'FARUR', 'FATUTR')) else \
                     "ITAU (non-CAIXA)" if ('ITAU' in n or 'ITAÚ' in n) and 'CAIXA' not in n else \
                     "DOLAR" if 'DOLAR' in n or 'DÓLAR' in n else \
                     "CARTAO CORPORATIVO" if 'CARTÃO CORPORATIVO' in n else \
                     "CARTAO CREDITO" if ('CARTÃO' in n and 'CRÉDITO' in n) or ('CARTAO' in n and 'CREDITO' in n) else \
                     "CAIXA ITAU" if 'CAIXA ITAU' in n or 'CAIXA ITAÚ' in n else \
                     "OTHER"
            raw = e.get("raw_data") or {}
            if isinstance(raw, str):
                import json as _json
                raw = _json.loads(raw)
            card_filtered[rid] = {
                "name": name, "user": e["user_name"], "cpf": e["user_cpf"],
                "status": e["report_status"], "count": 0, "total": 0.0, "reason": reason,
                "pm_ids": pm_str,
                "pdf_link": raw.get("pdf_link", "") if raw else "",
                "approval_date": raw.get("approval_date", "") if raw else "",
            }
        card_filtered[rid]["count"] += 1
        card_filtered[rid]["total"] += float(e["value"])

    for ri, (rid, r) in enumerate(sorted(card_filtered.items(), key=lambda x: x[1]["total"], reverse=True), 2):
        pdf_link = r.get("pdf_link", "")
        vals = [rid, pdf_link or "(no PDF)", r["name"], r["user"], r["cpf"], r["status"], r["count"], r["total"], r["reason"], r.get("pm_ids", ""), r.get("approval_date", "")]
        for ci, v in enumerate(vals, 1):
            cell = ws7.cell(row=ri, column=ci, value=v)
            if ci == 8:
                cell.number_format = money_fmt
            if ci == 2 and pdf_link:
                cell.hyperlink = pdf_link
                cell.font = Font(color="0563C1", underline="single")
    auto_width(ws7)

    wb_out.save(OUT_PATH)
    print(f"\nDone! Saved to: {OUT_PATH}")
    print(f"  Sheets: Summary, New Reports, Missing Reports, Per-CPF Divergence, New Expenses, Missing Expenses, Filtered Out")


if __name__ == "__main__":
    main()
