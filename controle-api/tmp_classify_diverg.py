#!/usr/bin/env python3
"""
Classify ALL divergences between API and reference BASE PREST.
Apply proposed filters (fartur, cartão corporativo, CARTÃO VEXPENSES='SIM', ITAU name).
Then categorize each divergent report and expense to understand root causes.
"""
import os, re
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
import datetime as dt

BASE = Path(__file__).parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"
OUT_DIR = BASE.parent / "data"

def norm(s):
    return " ".join(str(s or "").upper().strip().split())

# ============================================================
# 1. Load reference BASE PREST
# ============================================================
print("Loading reference BASE PREST...")
wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_bp = wb["BASE PREST "]

ref_expenses = {}  # eid -> {rid, name, user, cpf, date, value}
ref_reports = {}   # rid -> {name, user, cpf, count, total}
for row in ws_bp.iter_rows(min_row=4, values_only=True):
    if row[0] is None:
        continue
    eid = int(row[0])
    rid = int(row[1]) if row[1] else None
    rname = str(row[2] or "")
    user_name = str(row[4] or "")
    user_cpf = str(row[9] or "").strip()
    expense_date = row[6]  # col 6 = data
    valor = float(row[26] or 0) if len(row) > 26 else 0
    
    ref_expenses[eid] = {"rid": rid, "name": rname, "user": user_name, "cpf": user_cpf, "value": valor}
    
    if rid:
        if rid not in ref_reports:
            ref_reports[rid] = {"name": rname, "user": user_name, "cpf": user_cpf, "count": 0, "total": 0}
        ref_reports[rid]["count"] += 1
        ref_reports[rid]["total"] += valor

print(f"  Reference: {len(ref_reports)} reports, {len(ref_expenses)} expenses, R$ {sum(e['value'] for e in ref_expenses.values()):,.2f}")

# ============================================================
# 2. Load PAINEL — CARTÃO VEXPENSES CPFs + user info
# ============================================================
ws_p = wb["PAINEL"]
vexpenses_cpfs = set()
painel_info = {}  # cpf -> {colaborador, situacao, cartao_vexp, cartao_itau, status_cartao}
for row in ws_p.iter_rows(min_row=12, values_only=True):
    if row[2] is None:
        continue
    cpf = str(row[2] or "").strip()
    colab = str(row[1] or "").strip()
    situacao = str(row[4] or "").strip()
    status_cartao = str(row[5] or "").strip()
    cartao_itau = str(row[6] or "").strip()
    cartao_vexp = str(row[12] or "").strip().upper()
    
    painel_info[cpf] = {
        "colaborador": colab,
        "situacao": situacao,
        "status_cartao": status_cartao,
        "cartao_itau": cartao_itau,
        "cartao_vexp": cartao_vexp,
    }
    if cartao_vexp == "SIM":
        vexpenses_cpfs.add(cpf)

print(f"  PAINEL: {len(painel_info)} users, {len(vexpenses_cpfs)} with CARTÃO VEXPENSES='SIM'")
wb.close()

# ============================================================
# 3. Load API data with ALL proposed filters
# ============================================================
print("\nLoading API data with all filters...")
conn = psycopg2.connect(NEON_URL, connect_timeout=10)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cur.execute("""
    SELECT e.id, e.report_id, e.date, e.value, e.description, e.status as expense_status,
           r.name as report_name, r.status as report_status,
           r.user_cpf, r.user_name, r.created_at as report_created_at,
           r.updated_at as report_updated_at
    FROM prestacao_expenses e
    JOIN prestacao_reports r ON e.report_id = r.id
    WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND r.user_cpf IS NOT NULL
    ORDER BY r.user_name, r.name, e.date
""")
all_api = cur.fetchall()
conn.close()

# Apply all filters
def apply_filters(expenses):
    result = []
    for e in expenses:
        name = (e["report_name"] or "").strip()
        # FATURA exclusion (with fartur)
        if re.match(r'^(fatu|farur|cart)', name, re.IGNORECASE):
            continue
        if re.search(r'(fatura|fatuar|fatut|farur|fartur)', name, re.IGNORECASE):
            continue
        # Cartão corporativo exclusion
        if re.search(r'(cartão corporativo|cartao corporativo)', name, re.IGNORECASE):
            continue
        # ITAU name exclusion
        if re.search(r'itau|itaú', name, re.IGNORECASE):
            continue
        # CARTÃO VEXPENSES = 'SIM' filter
        if e["user_cpf"] not in vexpenses_cpfs:
            continue
        result.append(e)
    return result

filtered = apply_filters(all_api)

api_expenses = {}
api_reports = {}
for e in filtered:
    eid = int(e["id"])
    rid = int(e["report_id"])
    api_expenses[eid] = {
        "rid": rid,
        "name": e["report_name"],
        "user": e["user_name"],
        "cpf": e["user_cpf"],
        "value": float(e["value"]),
        "date": e["date"],
        "status": e["report_status"],
        "created_at": e["report_created_at"],
        "updated_at": e["report_updated_at"],
    }
    if rid not in api_reports:
        api_reports[rid] = {
            "name": e["report_name"],
            "user": e["user_name"],
            "cpf": e["user_cpf"],
            "status": e["report_status"],
            "created_at": e["report_created_at"],
            "count": 0,
            "total": 0,
        }
    api_reports[rid]["count"] += 1
    api_reports[rid]["total"] += float(e["value"])

print(f"  API (filtered): {len(api_reports)} reports, {len(api_expenses)} expenses, R$ {sum(e['value'] for e in api_expenses.values()):,.2f}")

# ============================================================
# 4. Classify divergences
# ============================================================
print("\n" + "=" * 100)
print("CLASSIFYING DIVERGENCES")
print("=" * 100)

api_report_ids = set(api_reports.keys())
ref_report_ids = set(ref_reports.keys())
api_expense_ids = set(api_expenses.keys())
ref_expense_ids = set(ref_expenses.keys())

new_reports = api_report_ids - ref_report_ids
missing_reports = ref_report_ids - api_report_ids
common_reports = api_report_ids & ref_report_ids

new_expenses = api_expense_ids - ref_expense_ids
missing_expenses = ref_expense_ids - api_expense_ids

print(f"\n  New reports (API only): {len(new_reports)}")
print(f"  Missing reports (ref only): {len(missing_reports)}")
print(f"  Common reports: {len(common_reports)}")
print(f"  New expenses (API only): {len(new_expenses)}")
print(f"  Missing expenses (ref only): {len(missing_expenses)}")

# ============================================================
# 5. Classify NEW reports (in API, not in reference)
# ============================================================
print("\n" + "=" * 100)
print("CLASSIFYING NEW REPORTS (in API, not in reference)")
print("=" * 100)

categories = defaultdict(lambda: {"reports": [], "expense_count": 0, "total": 0})

for rid in sorted(new_reports):
    r = api_reports[rid]
    cpf = r["cpf"]
    user = r["user"]
    name = r["name"]
    status = r["status"]
    created = r["created_at"]
    total = r["total"]
    count = r["count"]
    
    # Get PAINEL info
    pinfo = painel_info.get(cpf, {})
    situacao = pinfo.get("situacao", "NOT IN PAINEL")
    cartao_vexp = pinfo.get("cartao_vexp", "?")
    status_cartao = pinfo.get("status_cartao", "?")
    
    # Classify
    if status and status.upper() == "ENVIADO":
        cat = "A. ENVIADO (not yet approved)"
    elif situacao and "INATIVO" in situacao.upper():
        cat = "B. User INATIVO in PAINEL"
    elif status_cartao and status_cartao != "Cartão ativo":
        cat = f"C. Card not active ({status_cartao})"
    elif created and created.date() > dt.date(2026, 6, 30):
        cat = "D. Created after cutoff (2026-06-30)"
    elif created and created.date() >= dt.date(2026, 1, 1):
        cat = "E. Created in 2026 (pre-cutoff)"
    elif created and created.date() >= dt.date(2025, 1, 1):
        cat = "F. Created in 2025"
    else:
        cat = "G. Older reports"
    
    categories[cat]["reports"].append({
        "rid": rid, "name": name, "user": user, "cpf": cpf,
        "status": status, "created": created, "count": count, "total": total,
        "situacao": situacao, "cartao_vexp": cartao_vexp, "status_cartao": status_cartao,
    })
    categories[cat]["expense_count"] += count
    categories[cat]["total"] += total

print(f"\n  {'Category':<45} {'Reports':>8} {'Expenses':>9} {'Total R$':>14}")
print("  " + "-" * 80)
for cat in sorted(categories.keys()):
    info = categories[cat]
    print(f"  {cat:<45} {len(info['reports']):>8} {info['expense_count']:>9} {info['total']:>14,.2f}")

# Show details for each category
for cat in sorted(categories.keys()):
    info = categories[cat]
    if not info["reports"]:
        continue
    print(f"\n  --- {cat} ({len(info['reports'])} reports, R$ {info['total']:,.2f}) ---")
    for r in sorted(info["reports"], key=lambda x: -x["total"]):
        created_str = r["created"].strftime("%Y-%m-%d") if r["created"] else "?"
        print(f"    {r['rid']}  {str(r['name'])[:30]:<30}  {str(r['user'])[:20]:<20}  CPF={r['cpf']}  {r['status']:<10}  created={created_str}  {r['count']:>3} items  R$ {r['total']:>10,.2f}  situacao={r['situacao']}  card={r['status_cartao']}")

# ============================================================
# 6. Classify MISSING reports (in reference, not in API)
# ============================================================
print("\n" + "=" * 100)
print("CLASSIFYING MISSING REPORTS (in reference, not in API)")
print("=" * 100)

missing_categories = defaultdict(lambda: {"reports": [], "expense_count": 0, "total": 0})

for rid in sorted(missing_reports):
    r = ref_reports[rid]
    cpf = r["cpf"]
    user = r["user"]
    name = r["name"]
    total = r["total"]
    count = r["count"]
    
    pinfo = painel_info.get(cpf, {})
    situacao = pinfo.get("situacao", "NOT IN PAINEL")
    
    # Classify
    if "DESATIVADO" in user.upper():
        cat = "X. [DESATIVADO] user in ref"
    elif "FATURA" in name.upper() or "FATU" in name.upper():
        cat = "Y. FATURA report in ref (should have been excluded)"
    elif cpf not in vexpenses_cpfs:
        cartao_vexp = pinfo.get("cartao_vexp", "?")
        cat = f"Z. No VExpenses card (cartao_vexp={cartao_vexp})"
    else:
        cat = "W. Other (in ref but not in API)"
    
    missing_categories[cat]["reports"].append({
        "rid": rid, "name": name, "user": user, "cpf": cpf,
        "count": count, "total": total, "situacao": situacao,
    })
    missing_categories[cat]["expense_count"] += count
    missing_categories[cat]["total"] += total

print(f"\n  {'Category':<50} {'Reports':>8} {'Expenses':>9} {'Total R$':>14}")
print("  " + "-" * 85)
for cat in sorted(missing_categories.keys()):
    info = missing_categories[cat]
    print(f"  {cat:<50} {len(info['reports']):>8} {info['expense_count']:>9} {info['total']:>14,.2f}")

for cat in sorted(missing_categories.keys()):
    info = missing_categories[cat]
    if not info["reports"]:
        continue
    print(f"\n  --- {cat} ({len(info['reports'])} reports, R$ {info['total']:,.2f}) ---")
    for r in sorted(info["reports"], key=lambda x: -x["total"]):
        print(f"    {r['rid']}  {str(r['name'])[:30]:<30}  {str(r['user'])[:25]:<25}  CPF={r['cpf']}  {r['count']:>3} items  R$ {r['total']:>10,.2f}")

# ============================================================
# 7. Common reports with value differences
# ============================================================
print("\n" + "=" * 100)
print("COMMON REPORTS WITH VALUE DIFFERENCES")
print("=" * 100)

value_diffs = []
for rid in common_reports:
    api_r = api_reports[rid]
    ref_r = ref_reports[rid]
    diff = api_r["total"] - ref_r["total"]
    if abs(diff) > 0.01:
        value_diffs.append((rid, diff, api_r, ref_r))

value_diffs.sort(key=lambda x: abs(x[1]), reverse=True)
print(f"\n  Common reports with value differences: {len(value_diffs)}")
print(f"  {'RID':<10} {'API R$':>12} {'Ref R$':>12} {'Diff R$':>12} {'API items':>10} {'Ref items':>10}  Report")
print("  " + "-" * 100)
for rid, diff, api_r, ref_r in value_diffs[:30]:
    print(f"  {rid:<10} {api_r['total']:>12,.2f} {ref_r['total']:>12,.2f} {diff:>+12,.2f} {api_r['count']:>10} {ref_r['count']:>10}  {str(api_r['name'])[:30]}")

total_diff = sum(d[1] for d in value_diffs)
print(f"\n  Total value diff in common reports: R$ {total_diff:,.2f}")

# ============================================================
# 8. Summary
# ============================================================
print("\n" + "=" * 100)
print("DIVERGENCE SUMMARY")
print("=" * 100)

new_total = sum(api_reports[rid]["total"] for rid in new_reports)
missing_total = sum(ref_reports[rid]["total"] for rid in missing_reports)
common_diff_total = sum(d[1] for d in value_diffs)

print(f"\n  New reports (API only):      {len(new_reports):>4} reports  R$ {new_total:>12,.2f}")
print(f"  Missing reports (ref only):  {len(missing_reports):>4} reports  R$ {missing_total:>12,.2f}")
print(f"  Common reports value diff:   {len(value_diffs):>4} reports  R$ {common_diff_total:>12,.2f}")
print(f"  ────────────────────────────────────────────")
print(f"  Net divergence:                           R$ {new_total - missing_total + common_diff_total:>12,.2f}")
