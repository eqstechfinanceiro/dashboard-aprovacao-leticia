"""Inspect CARGA 2 QZ JULHO sheet structure."""
import openpyxl
from pathlib import Path

SHEET_PATH = Path(r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CARGA 2 QZ JULHO 26 VEXPENSES EQS.xlsx")

wb = openpyxl.load_workbook(SHEET_PATH, data_only=True)
print(f"Sheets: {wb.sheetnames}")
ws = wb.active
print(f"Active: {ws.title}, rows={ws.max_row}, cols={ws.max_column}")

# Print first 12 rows, all columns up to 20
for row_idx in range(1, min(13, ws.max_row + 1)):
    vals = []
    for col_idx in range(1, min(ws.max_column + 1, 22)):
        v = ws.cell(row=row_idx, column=col_idx).value
        if v is not None:
            vals.append(f"[{col_idx}]={str(v)[:25]}")
    print(f"  Row {row_idx}: {' | '.join(vals) if vals else '(empty)'}")

# Also check if there's a second sheet
for sn in wb.sheetnames:
    ws2 = wb[sn]
    print(f"\n  Sheet '{sn}': {ws2.max_row} rows x {ws2.max_column} cols")
    # First 3 rows
    for r in range(1, min(4, ws2.max_row + 1)):
        vals = []
        for c in range(1, min(ws2.max_column + 1, 15)):
            v = ws2.cell(row=r, column=c).value
            if v is not None:
                vals.append(f"[{c}]={str(v)[:20]}")
        print(f"    Row {r}: {' | '.join(vals) if vals else '(empty)'}")

wb.close()
