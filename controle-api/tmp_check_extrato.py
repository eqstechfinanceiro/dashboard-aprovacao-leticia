#!/usr/bin/env python3
"""Check EXTRATO date range and raw_data fields."""
import os
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from pathlib import Path
import openpyxl

BASE = Path(__file__).parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")

# 1. Check reference EXTRATO last rows
REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"
wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws = wb["EXTRATO"]

# First and last data rows
rows = list(ws.iter_rows(min_row=9, max_row=15, values_only=True))
print("EXTRATO first 7 data rows:")
for r in rows:
    print("  {}".format([v for v in r if v is not None]))

# Last rows
last_rows = list(ws.iter_rows(min_row=ws.max_row-5, max_row=ws.max_row, values_only=True))
print("\nEXTRATO last 6 rows:")
for r in last_rows:
    print("  {}".format([v for v in r if v is not None]))

# Count by type
from collections import Counter
types = Counter()
for r in ws.iter_rows(min_row=9, values_only=True):
    if r[8] is not None:
        types[str(r[8]).strip()] += 1
print("\nEXTRATO types: {}".format(dict(types)))
wb.close()

# 2. Check DB extrato date range and raw_data
conn = psycopg2.connect(NEON_URL, connect_timeout=10)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cur.execute("""
    SELECT MIN(data) as min_date, MAX(data) as max_date, COUNT(*) as cnt,
           COUNT(*) FILTER (WHERE is_snapshot = FALSE) as non_snap,
           COUNT(*) FILTER (WHERE is_snapshot = TRUE) as snap
    FROM extrato_movimentacao
""")
r = cur.fetchone()
print("\nDB extrato_movimentacao:")
print("  Date range: {} to {}".format(r["min_date"], r["max_date"]))
print("  Total: {}, non-snap: {}, snap: {}".format(r["cnt"], r["non_snap"], r["snap"]))

# Check if any raw_data is not null in prestacao_expenses
cur.execute("""
    SELECT COUNT(*) as total,
           COUNT(*) FILTER (WHERE raw_data IS NOT NULL) as has_raw
    FROM prestacao_expenses
""")
r = cur.fetchone()
print("\nDB prestacao_expenses: total={}, has_raw_data={}".format(r["total"], r["has_raw"]))

# Check prestacao_reports raw_data
cur.execute("""
    SELECT COUNT(*) as total,
           COUNT(*) FILTER (WHERE raw_data IS NOT NULL) as has_raw
    FROM prestacao_reports
""")
r = cur.fetchone()
print("DB prestacao_reports: total={}, has_raw_data={}".format(r["total"], r["has_raw"]))

# Check a raw_data sample if any
cur.execute("SELECT raw_data FROM prestacao_expenses WHERE raw_data IS NOT NULL LIMIT 1")
r = cur.fetchone()
if r:
    print("\nSample expense raw_data keys: {}".format(list(r["raw_data"].keys()) if r["raw_data"] else "None"))
else:
    print("\nNo expense raw_data found")

cur.execute("SELECT raw_data FROM prestacao_reports WHERE raw_data IS NOT NULL LIMIT 1")
r = cur.fetchone()
if r:
    print("Sample report raw_data keys: {}".format(list(r["raw_data"].keys()) if r["raw_data"] else "None"))
else:
    print("No report raw_data found")

conn.close()
