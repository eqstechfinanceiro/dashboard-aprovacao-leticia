#!/usr/bin/env python3
"""
Deep investigation:
1. TARIFA: Check name matching for divergent CPFs
2. PRESTACAO: Categorize the 1051 extra API expenses
"""
import os, re
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from pathlib import Path
import openpyxl
from collections import defaultdict

BASE = Path(__file__).parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"

conn = psycopg2.connect(NEON_URL, connect_timeout=10)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# ============================================================
# PART 1: TARIFA — name matching issue
# ============================================================
print("=" * 80)
print("PART 1: TARIFA — Name matching investigation")
print("=" * 80)

# For GUILHERME (11178519740), check what name is in cadastro
# and what names exist in extrato for Taxa transactions
test_cpfs = [
    ("11178519740", "GUILHERME MOTTA RIBEIRO SILVA", +28.00),
    ("02027745203", "ABNER ANDRADE CAVALCANTE", +14.00),
    ("01050938232", "CHARLYTON COSTA ANDRADE", -29.93),
    ("06576198922", "LUDGERO HORACIO DE OLIVEIRA", +35.00),
]

for cpf, expected_name, diff in test_cpfs:
    # Get cadastro name
    cur.execute("SELECT colaborador, cpf FROM quinzena_cadastro WHERE cpf = %s LIMIT 1", (cpf,))
    r = cur.fetchone()
    cad_name = r["colaborador"] if r else "NOT FOUND"
    
    # Check what usuario names exist in extrato for Taxa, matching this name
    cur.execute("""
        SELECT DISTINCT usuario, COUNT(*) as cnt, SUM(valor) as total
        FROM extrato_movimentacao
        WHERE is_snapshot = FALSE
          AND data <= '2026-06-30'
          AND tipo = 'Taxa'
          AND UPPER(usuario) = UPPER(%s)
        GROUP BY usuario
    """, (cad_name,))
    db_matches = cur.fetchall()
    
    # Also check reference EXTRATO for this CPF
    wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
    ws = wb["EXTRATO"]
    ref_tarifa = []
    for row in ws.iter_rows(min_row=9, values_only=True):
        if row[8] is None:
            continue
        if str(row[8]).strip().upper() == "TARIFA":
            row_cpf = str(row[11] or "").strip().replace(".", "").replace("-", "").replace(" ", "").zfill(11) if row[11] else ""
            if row_cpf == cpf:
                ref_tarifa.append({"desc": str(row[9] or ""), "valor": float(row[10] or 0), "usuario": str(row[7] or "")})
    wb.close()
    
    print(f"\n  CPF={cpf}  Expected={expected_name}")
    print(f"    Cadastro name: '{cad_name}'")
    print(f"    Ref EXTRATO Taxa count: {len(ref_tarifa)}, total: R$ {sum(abs(t['valor']) for t in ref_tarifa):.2f}")
    if ref_tarifa:
        print(f"    Ref usuario names: {set(t['usuario'] for t in ref_tarifa)}")
    print(f"    DB extrato Taxa matches (by UPPER(usuario)=UPPER(cad_name)):")
    for m in db_matches:
        print(f"      '{m['usuario']}' — {m['cnt']} txns, total R$ {float(m['total']):.2f}")
    
    if not db_matches and ref_tarifa:
        # The name in ref extrato is different from cadastro
        ref_names = set(t['usuario'] for t in ref_tarifa)
        print(f"    *** MISMATCH: Ref has Taxa for this CPF under name(s): {ref_names}")
        print(f"    *** But cadastro name is '{cad_name}' — no match in DB extrato!")

# ============================================================
# PART 2: PRESTACAO — categorize 1051 extra API expenses
# ============================================================
print("\n" + "=" * 80)
print("PART 2: PRESTACAO — Categorize extra API expenses")
print("=" * 80)

wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_bp = wb["BASE PREST "]

ref_expense_ids = set()
ref_report_ids = set()
for row in ws_bp.iter_rows(min_row=4, values_only=True):
    if row[0] is None:
        continue
    ref_expense_ids.add(int(row[0]))
    ref_report_ids.add(int(row[1]))

# Get all API expenses (Aprovado+Enviado, excluding FATURA/CARTAO)
cur.execute("""
    SELECT e.id, e.report_id, e.value, r.user_cpf, r.name as report_name, r.status, r.user_name
    FROM prestacao_expenses e
    JOIN prestacao_reports r ON e.report_id = r.id
    WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND r.user_cpf IS NOT NULL
      AND TRIM(r.name) !~* '^(fatu|farur|cart)'
      AND TRIM(r.name) !~* '(fatura|fatuar|fatut|farur)'
""")
api_expenses = cur.fetchall()
api_expense_ids = set(int(r["id"]) for r in api_expenses)

extra_in_api = api_expense_ids - ref_expense_ids
extra_in_ref = ref_expense_ids - api_expense_ids

print(f"  Extra in API: {len(extra_in_api)}")
print(f"  Extra in ref: {len(extra_in_ref)}")

# Categorize extra API expenses
extra_rows = [r for r in api_expenses if int(r["id"]) in extra_in_api]

# By status
by_status = defaultdict(lambda: {"count": 0, "total": 0})
for r in extra_rows:
    by_status[r["status"]]["count"] += 1
    by_status[r["status"]]["total"] += float(r["value"])
print(f"\n  Extra by status:")
for s, info in sorted(by_status.items()):
    print(f"    {s}: {info['count']} items, R$ {info['total']:,.2f}")

# By report (new vs existing in ref)
new_report_extras = [r for r in extra_rows if int(r["report_id"]) not in ref_report_ids]
existing_report_extras = [r for r in extra_rows if int(r["report_id"]) in ref_report_ids]

print(f"\n  From NEW reports (not in ref): {len(new_report_extras)} expenses")
print(f"  From EXISTING reports (in ref but more expenses): {len(existing_report_extras)} expenses")

# Show new reports
new_report_ids = set(int(r["report_id"]) for r in new_report_extras)
cur.execute("""
    SELECT id, name, status, user_cpf, user_name, created_at, updated_at
    FROM prestacao_reports
    WHERE id = ANY(%s)
    ORDER BY updated_at DESC
""", (list(new_report_ids),))
new_reports = cur.fetchall()

print(f"\n  NEW reports ({len(new_reports)} total) — showing all:")
for r in new_reports:
    print(f"    {r['id']}  {str(r['name'])[:35]:<35}  {r['status']:<10}  {str(r['user_name'])[:25]:<25}  updated={r['updated_at']}")

# Show existing reports with extra expenses
existing_report_ids = set(int(r["report_id"]) for r in existing_report_extras)
cur.execute("""
    SELECT id, name, status, user_cpf, user_name
    FROM prestacao_reports
    WHERE id = ANY(%s)
    ORDER BY name
""", (list(existing_report_ids),))
existing_reports = cur.fetchall()

print(f"\n  EXISTING reports with extra expenses ({len(existing_reports)} reports):")
for r in existing_reports[:20]:
    # Count how many expenses in ref vs API for this report
    ref_cnt = sum(1 for row in ws_bp.iter_rows(min_row=4, values_only=True) if row[1] is not None and int(row[1]) == r["id"])
    api_cnt = sum(1 for e in api_expenses if int(e["report_id"]) == r["id"])
    extra = sum(1 for e in existing_report_extras if int(e["report_id"]) == r["id"])
    print(f"    {r['id']}  {str(r['name'])[:30]:<30}  ref={ref_cnt}  api={api_cnt}  extra={extra}")

# Check extra in ref (71 expenses not in API)
extra_ref_ids = list(extra_in_ref)[:20]
cur.execute("""
    SELECT e.id, e.report_id, e.value, r.user_cpf, r.name as report_name, r.status
    FROM prestacao_expenses e
    JOIN prestacao_reports r ON e.report_id = r.id
    WHERE e.id = ANY(%s)
""", (extra_ref_ids,))
ref_only = cur.fetchall()
not_in_api_at_all = [r for r in ref_only if r is None]
in_api_diff_status = [r for r in ref_only if r is not None]

print(f"\n  Extra in REF (not in API) — first 20:")
for eid in extra_ref_ids:
    # Check if this expense ID exists in API at all
    cur.execute("SELECT id, report_id, value FROM prestacao_expenses WHERE id = %s", (eid,))
    api_r = cur.fetchone()
    if api_r:
        cur.execute("SELECT status, name FROM prestacao_reports WHERE id = %s", (api_r["report_id"],))
        rep = cur.fetchone()
        print(f"    {eid}  EXISTS in API but report status={rep['status'] if rep else '?'}  report={rep['name'] if rep else '?'}")
    else:
        print(f"    {eid}  NOT IN API AT ALL")

wb.close()
conn.close()
