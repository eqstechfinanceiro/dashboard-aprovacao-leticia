#!/usr/bin/env python3
"""
Deep dive into the 54 reports in API but NOT in reference BASE PREST.
The user says both were exported at the same time, so timing is NOT the cause.
What do these 54 reports have in common?

Also simulate:
- APROVADO-only filter (to confirm it causes issues)
- SITUAÇÃO=ATIVO filter
- status_cartao='Cartão ativo' filter
- Various combinations to get to <10% error
"""
import os, re
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from pathlib import Path
import openpyxl
from collections import defaultdict
import datetime as dt

BASE = Path(__file__).parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"

# ============================================================
# Load reference BASE PREST report IDs
# ============================================================
print("Loading reference BASE PREST...")
wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_bp = wb["BASE PREST "]

ref_report_ids = set()
ref_report_info = {}
ref_total = 0
for row in ws_bp.iter_rows(min_row=4, values_only=True):
    if row[0] is None:
        continue
    eid = int(row[0])
    rid = int(row[1]) if row[1] else None
    rname = str(row[2] or "")
    user_name = str(row[4] or "")
    user_cpf = str(row[9] or "").strip()
    valor = float(row[26] or 0) if len(row) > 26 else 0
    ref_total += valor
    if rid:
        ref_report_ids.add(rid)
        if rid not in ref_report_info:
            ref_report_info[rid] = {"name": rname, "user": user_name, "cpf": user_cpf, "count": 0, "total": 0}
        ref_report_info[rid]["count"] += 1
        ref_report_info[rid]["total"] += valor

# Load PAINEL info
ws_p = wb["PAINEL"]
vexpenses_cpfs = set()
painel_info = {}
for row in ws_p.iter_rows(min_row=12, values_only=True):
    if row[2] is None:
        continue
    cpf = str(row[2] or "").strip()
    colab = str(row[1] or "").strip()
    situacao = str(row[4] or "").strip().upper()
    status_cartao = str(row[5] or "").strip()
    cartao_itau = str(row[6] or "").strip()
    cartao_vexp = str(row[12] or "").strip().upper()
    painel_info[cpf] = {
        "colaborador": colab, "situacao": situacao, "status_cartao": status_cartao,
        "cartao_itau": cartao_itau, "cartao_vexp": cartao_vexp,
    }
    if cartao_vexp == "SIM":
        vexpenses_cpfs.add(cpf)
wb.close()

print(f"  Reference: {len(ref_report_ids)} reports, R$ {ref_total:,.2f}")

# ============================================================
# Load API data
# ============================================================
print("\nLoading API data...")
conn = psycopg2.connect(NEON_URL, connect_timeout=10)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cur.execute("""
    SELECT e.id, e.report_id, e.date, e.value, e.description, e.status as expense_status,
           r.name as report_name, r.status as report_status,
           r.user_cpf, r.user_name, r.created_at as report_created_at,
           r.updated_at as report_updated_at, r.total_value
    FROM prestacao_expenses e
    JOIN prestacao_reports r ON e.report_id = r.id
    WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND r.user_cpf IS NOT NULL
    ORDER BY r.user_name, r.name, e.date
""")
all_api = cur.fetchall()

# Get report-level info
cur.execute("""
    SELECT DISTINCT id, name, status, user_cpf, user_name, created_at, updated_at, total_value
    FROM prestacao_reports
    WHERE (status ILIKE 'Aprovado' OR status ILIKE 'Enviado')
      AND user_cpf IS NOT NULL
""")
api_reports_raw = {r["id"]: r for r in cur.fetchall()}
conn.close()

# Apply current proposed filters (fartur, cartao corp, itau name, vexpenses card)
def apply_base_filters(expenses):
    result = []
    for e in expenses:
        name = (e["report_name"] or "").strip()
        if re.match(r'^(fatu|farur|cart)', name, re.IGNORECASE):
            continue
        if re.search(r'(fatura|fatuar|fatut|farur|fartur)', name, re.IGNORECASE):
            continue
        if re.search(r'(cartão corporativo|cartao corporativo)', name, re.IGNORECASE):
            continue
        if re.search(r'itau|itaú', name, re.IGNORECASE):
            continue
        if e["user_cpf"] not in vexpenses_cpfs:
            continue
        result.append(e)
    return result

filtered = apply_base_filters(all_api)

# Build API report set
api_report_ids = set()
api_report_data = {}
for e in filtered:
    rid = e["report_id"]
    api_report_ids.add(rid)
    if rid not in api_report_data:
        r = api_reports_raw.get(rid, {})
        api_report_data[rid] = {
            "name": e["report_name"],
            "user": e["user_name"],
            "cpf": e["user_cpf"],
            "status": e["report_status"],
            "created_at": r.get("created_at"),
            "updated_at": r.get("updated_at"),
            "total_value": float(r.get("total_value") or 0),
            "expense_count": 0,
            "expense_total": 0,
        }
    api_report_data[rid]["expense_count"] += 1
    api_report_data[rid]["expense_total"] += float(e["value"])

new_reports = api_report_ids - ref_report_ids
print(f"  API (filtered): {len(api_report_ids)} reports")
print(f"  New reports (API only): {len(new_reports)}")

# ============================================================
# 1. DEEP DIVE: What do the 54 reports have in common?
# ============================================================
print("\n" + "=" * 100)
print("DEEP DIVE: 54 reports in API but NOT in reference BASE PREST")
print("=" * 100)

# Collect all attributes
attributes = defaultdict(lambda: {"count": 0, "total": 0, "reports": []})

for rid in sorted(new_reports):
    r = api_report_data[rid]
    cpf = r["cpf"]
    pinfo = painel_info.get(cpf, {})
    situacao = pinfo.get("situacao", "NOT IN PAINEL")
    status_cartao = pinfo.get("status_cartao", "?")
    cartao_vexp = pinfo.get("cartao_vexp", "?")
    cartao_itau = pinfo.get("cartao_itau", "")
    
    # Check all possible common attributes
    status = r["status"].upper() if r["status"] else "?"
    created = r["created_at"]
    updated = r["updated_at"]
    
    # Group by status
    attributes[f"status={status}"]["count"] += 1
    attributes[f"status={status}"]["total"] += r["expense_total"]
    
    # Group by situacao
    attributes[f"situacao={situacao}"]["count"] += 1
    attributes[f"situacao={situacao}"]["total"] += r["expense_total"]
    
    # Group by status_cartao
    attributes[f"status_cartao={status_cartao}"]["count"] += 1
    attributes[f"status_cartao={status_cartao}"]["total"] += r["expense_total"]
    
    # Group by cartao_vexp
    attributes[f"cartao_vexp={cartao_vexp}"]["count"] += 1
    attributes[f"cartao_vexp={cartao_vexp}"]["total"] += r["expense_total"]
    
    # Group by created_at year-month
    if created:
        ym = created.strftime("%Y-%m")
        attributes[f"created={ym}"]["count"] += 1
        attributes[f"created={ym}"]["total"] += r["expense_total"]
    
    # Group by updated_at year-month
    if updated:
        ym = updated.strftime("%Y-%m")
        attributes[f"updated={ym}"]["count"] += 1
        attributes[f"updated={ym}"]["total"] += r["expense_total"]
    
    # Group by has_cartao_itau
    has_itau = "YES" if cartao_itau else "NO"
    attributes[f"has_cartao_itau={has_itau}"]["count"] += 1
    attributes[f"has_cartao_itau={has_itau}"]["total"] += r["expense_total"]

print("\n  Common attributes:")
print(f"  {'Attribute':<40} {'Count':>6} {'Total R$':>14} {'% of 54':>8}")
print("  " + "-" * 70)
for attr in sorted(attributes.keys()):
    info = attributes[attr]
    pct = info["count"] / len(new_reports) * 100
    print(f"  {attr:<40} {info['count']:>6} {info['total']:>14,.2f} {pct:>7.1f}%")

# ============================================================
# 2. Check: Do these 54 reports' users have ANY reports in the reference?
# ============================================================
print("\n" + "=" * 100)
print("Do the 54 reports' users have OTHER reports in the reference?")
print("=" * 100)

new_report_users = defaultdict(lambda: {"new_reports": [], "ref_reports": [], "new_total": 0, "ref_total": 0})

for rid in new_reports:
    r = api_report_data[rid]
    cpf = r["cpf"]
    new_report_users[cpf]["new_reports"].append(rid)
    new_report_users[cpf]["new_total"] += r["expense_total"]

for rid, r in ref_report_info.items():
    cpf = r["cpf"]
    if cpf in new_report_users:
        new_report_users[cpf]["ref_reports"].append(rid)
        new_report_users[cpf]["ref_total"] += r["total"]

print(f"\n  {'User':<30} {'CPF':<15} {'New':>5} {'Ref':>5} {'New R$':>12} {'Ref R$':>12}  Situacao")
print("  " + "-" * 100)
for cpf in sorted(new_report_users.keys(), key=lambda c: -new_report_users[c]["new_total"]):
    u = new_report_users[cpf]
    pinfo = painel_info.get(cpf, {})
    situacao = pinfo.get("situacao", "?")
    status_cartao = pinfo.get("status_cartao", "?")
    user_name = api_report_data[u["new_reports"][0]]["user"] if u["new_reports"] else "?"
    print(f"  {str(user_name)[:30]:<30} {cpf:<15} {len(u['new_reports']):>5} {len(u['ref_reports']):>5} {u['new_total']:>12,.2f} {u['ref_total']:>12,.2f}  {situacao}/{status_cartao}")

# ============================================================
# 3. Check: Are these reports in the reference under a DIFFERENT report ID?
# Sometimes exports create new report IDs. Check by name+user+value.
# ============================================================
print("\n" + "=" * 100)
print("Check: Are these reports in ref under a DIFFERENT report ID? (matching by name+user)")
print("=" * 100)

# Build ref lookup by (name, cpf)
ref_by_name_cpf = defaultdict(list)
for rid, r in ref_report_info.items():
    key = (r["name"].strip().upper(), r["cpf"])
    ref_by_name_cpf[key].append(rid)

matches_found = 0
for rid in sorted(new_reports):
    r = api_report_data[rid]
    key = (r["name"].strip().upper(), r["cpf"])
    if key in ref_by_name_cpf:
        ref_rids = ref_by_name_cpf[key]
        match_ref_total = sum(ref_report_info[rr]["total"] for rr in ref_rids)
        matches_found += 1
        print(f"  MATCH! API rid={rid} '{r['name']}' → ref rids={ref_rids}  API R$ {r['expense_total']:,.2f} vs Ref R$ {match_ref_total:,.2f}")

print(f"\n  Total matches by name+cpf: {matches_found} out of {len(new_reports)}")

# Also check by name only (in case CPF format differs)
ref_by_name = defaultdict(list)
for rid, r in ref_report_info.items():
    ref_by_name[r["name"].strip().upper()].append(rid)

name_only_matches = 0
for rid in sorted(new_reports):
    r = api_report_data[rid]
    name_upper = r["name"].strip().upper()
    if name_upper in ref_by_name:
        # Check if it's a different CPF
        ref_cpfs = set(ref_report_info[rr]["cpf"] for rr in ref_by_name[name_upper])
        if r["cpf"] not in ref_cpfs:
            name_only_matches += 1
            print(f"  NAME MATCH (diff CPF): API rid={rid} '{r['name']}' user={r['user']} CPF={r['cpf']} → ref CPFs={ref_cpfs}")

print(f"\n  Name-only matches with different CPF: {name_only_matches}")

# ============================================================
# 4. Check: updated_at vs created_at — were these reports recently modified?
# ============================================================
print("\n" + "=" * 100)
print("updated_at analysis for the 54 reports")
print("=" * 100)

print(f"\n  {'RID':<10} {'Name':<30} {'Created':>12} {'Updated':>12} {'Days btwn':>10} {'Status':<10}")
print("  " + "-" * 90)
for rid in sorted(new_reports, key=lambda x: api_report_data[x]["created_at"] or dt.datetime.min):
    r = api_report_data[rid]
    created = r["created_at"]
    updated = r["updated_at"]
    days = ""
    if created and updated:
        days = f"{(updated - created).days}"
    print(f"  {rid:<10} {str(r['name'])[:30]:<30} {str(created.strftime('%Y-%m-%d') if created else '?'):>12} {str(updated.strftime('%Y-%m-%d') if updated else '?'):>12} {days:>10} {r['status']:<10}")

# ============================================================
# 5. SIMULATE: APROVADO-only filter
# ============================================================
print("\n" + "=" * 100)
print("SIMULATION: Various filter combinations")
print("=" * 100)

ref_total_val = ref_total

def calc_divergence(expenses):
    api_rids = set(e["report_id"] for e in expenses)
    api_eids = set(e["id"] for e in expenses)
    api_total = sum(float(e["value"]) for e in expenses)
    
    new_r = api_rids - ref_report_ids
    missing_r = ref_report_ids - api_rids
    new_e = api_eids - set()  # we don't have ref expense IDs here, approximate
    gap = api_total - ref_total_val
    return len(api_rids), api_total, len(new_r), len(missing_r), gap

# Base filters
base = apply_base_filters(all_api)

# APROVADO only
aprovado_only = [e for e in base if e["report_status"].upper() == "APROVADO"]

# SITUAÇÃO=ATIVO only
ativo_cpfs = set(cpf for cpf, info in painel_info.items() if info["situacao"] == "ATIVO")
ativo_only = [e for e in base if e["user_cpf"] in ativo_cpfs]

# status_cartao='Cartão ativo' only
cartao_ativo_cpfs = set(cpf for cpf, info in painel_info.items() if info["status_cartao"] == "Cartão ativo")
cartao_ativo_only = [e for e in base if e["user_cpf"] in cartao_ativo_cpfs]

# APROVADO + ATIVO
aprovado_ativo = [e for e in base if e["report_status"].upper() == "APROVADO" and e["user_cpf"] in ativo_cpfs]

# APROVADO + Cartão ativo
aprovado_cartao = [e for e in base if e["report_status"].upper() == "APROVADO" and e["user_cpf"] in cartao_ativo_cpfs]

# ENVIADO only (to show what we'd lose)
enviado_only = [e for e in base if e["report_status"].upper() == "ENVIADO"]

scenarios = [
    ("Base (current proposed filters)", base),
    ("APROVADO only (no ENVIADO)", aprovado_only),
    ("SITUAÇÃO=ATIVO only", ativo_only),
    ("status_cartao='Cartão ativo'", cartao_ativo_only),
    ("APROVADO + ATIVO", aprovado_ativo),
    ("APROVADO + Cartão ativo", aprovado_cartao),
]

print(f"\n  {'Scenario':<40} {'Reports':>8} {'Total R$':>14} {'New':>5} {'Missing':>8} {'Gap R$':>14} {'% Gap':>7}")
print("  " + "-" * 100)
for name, expenses in scenarios:
    n_reports, total, n_new, n_missing, gap = calc_divergence(expenses)
    pct = abs(gap) / ref_total_val * 100
    print(f"  {name:<40} {n_reports:>8} {total:>14,.2f} {n_new:>5} {n_missing:>8} {gap:>+14,.2f} {pct:>6.1f}%")

# Show what ENVIADO reports we'd lose
print(f"\n  --- ENVIADO reports that would be excluded ({len(set(e['report_id'] for e in enviado_only))} reports) ---")
enviado_by_report = defaultdict(lambda: {"name": "", "user": "", "cpf": "", "count": 0, "total": 0, "created": None})
for e in enviado_only:
    rid = e["report_id"]
    enviado_by_report[rid]["name"] = e["report_name"]
    enviado_by_report[rid]["user"] = e["user_name"]
    enviado_by_report[rid]["cpf"] = e["user_cpf"]
    enviado_by_report[rid]["count"] += 1
    enviado_by_report[rid]["total"] += float(e["value"])
    r = api_reports_raw.get(rid, {})
    enviado_by_report[rid]["created"] = r.get("created_at")

for rid in sorted(enviado_by_report.keys(), key=lambda x: -enviado_by_report[x]["total"]):
    r = enviado_by_report[rid]
    created_str = r["created"].strftime("%Y-%m-%d") if r["created"] else "?"
    in_ref = "IN REF" if rid in ref_report_ids else "NOT IN REF"
    print(f"    {rid}  {str(r['name'])[:25]:<25}  {str(r['user'])[:20]:<20}  {r['count']:>4} items  R$ {r['total']:>10,.2f}  created={created_str}  {in_ref}")

# ============================================================
# 6. Best combination to get under 10%
# ============================================================
print("\n" + "=" * 100)
print("TARGET: Get divergence under 10% (R$ {:.2f})".format(ref_total_val * 0.10))
print("=" * 100)

# Try: ATIVO + Cartão ativo (no APROVADO filter)
ativo_cartao = [e for e in base if e["user_cpf"] in ativo_cpfs and e["user_cpf"] in cartao_ativo_cpfs]
n, t, nn, nm, g = calc_divergence(ativo_cartao)
pct = abs(g) / ref_total_val * 100
print(f"  ATIVO + Cartão ativo:         {n:>5} reports  R$ {t:>12,.2f}  gap=R$ {g:>+12,.2f}  ({pct:.1f}%)")

# Try: ATIVO only (keep both APROVADO and ENVIADO)
n, t, nn, nm, g = calc_divergence(ativo_only)
pct = abs(g) / ref_total_val * 100
print(f"  ATIVO only (keep ENVIADO):    {n:>5} reports  R$ {t:>12,.2f}  gap=R$ {g:>+12,.2f}  ({pct:.1f}%)")

# Try: Cartão ativo only (keep both)
n, t, nn, nm, g = calc_divergence(cartao_ativo_only)
pct = abs(g) / ref_total_val * 100
print(f"  Cartão ativo only:            {n:>5} reports  R$ {t:>12,.2f}  gap=R$ {g:>+12,.2f}  ({pct:.1f}%)")

# Try: exclude only INATIVO users
inativo_cpfs = set(cpf for cpf, info in painel_info.items() if info["situacao"] == "INATIVO")
not_inativo = [e for e in base if e["user_cpf"] not in inativo_cpfs]
n, t, nn, nm, g = calc_divergence(not_inativo)
pct = abs(g) / ref_total_val * 100
print(f"  Exclude INATIVO only:         {n:>5} reports  R$ {t:>12,.2f}  gap=R$ {g:>+12,.2f}  ({pct:.1f}%)")

# Try: exclude INATIVO + exclude non-active card
not_inativo_card = [e for e in base if e["user_cpf"] not in inativo_cpfs and e["user_cpf"] in cartao_ativo_cpfs]
n, t, nn, nm, g = calc_divergence(not_inativo_card)
pct = abs(g) / ref_total_val * 100
print(f"  Exclude INATIVO + Cartão ativo: {n:>5} reports  R$ {t:>12,.2f}  gap=R$ {g:>+12,.2f}  ({pct:.1f}%)")
