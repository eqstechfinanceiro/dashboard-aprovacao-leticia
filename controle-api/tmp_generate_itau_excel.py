#!/usr/bin/env python3
"""
Generate Excel with detailed Itaú payment method analysis for all 60 reports.
Includes per-expense breakdown showing payment_method_id and whether it's Itaú.
"""
import os, requests, json, time, re, unicodedata
import psycopg2, psycopg2.extras
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from pathlib import Path
from collections import Counter, defaultdict

load_dotenv(Path(__file__).parent / ".env")
API_KEY = os.getenv("VEXPENSES_API_KEY", "")
BASE_URL = os.getenv("VEXPENSES_BASE_URL", "https://api.vexpenses.com")
HEADERS = {"Authorization": API_KEY, "Accept": "application/json"}

# Payment method mapping
PM_MAP = {
    627401: "Cartão Corporativo Itaú",
    627721: "Unknown (627721)",
}

# Discover 627721
conn = psycopg2.connect(os.getenv("NEON_DATABASE_URL"), connect_timeout=10)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cur.execute("""
    SELECT e.id FROM prestacao_expenses e
    WHERE e.raw_data->'payment_method_id' = '627721'
    LIMIT 1
""")
row = cur.fetchone()
if row:
    try:
        resp = requests.get(f"{BASE_URL}/v2/expenses/{row['id']}?include=payment_method", headers=HEADERS, timeout=30)
        if resp.status_code == 200:
            data = resp.json().get("data", {})
            pm = data.get("payment_method", {}).get("data", {})
            PM_MAP[627721] = pm.get("description", "Unknown")
    except:
        pass

print(f"Payment methods: {PM_MAP}")

# Read gap file
xlsx_path = Path(r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\controle-api\data\gap entre referencia e neon ahahahahaah.xlsx")
wb_in = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
ws = wb_in["REPORTS SO NO NEON"]
rows = list(ws.iter_rows(values_only=True))
wb_in.close()

reports = []
for row in rows[1:]:
    if row[0] is None:
        continue
    reports.append({
        "report_id": row[0], "name": row[1], "user": row[2], "cpf": row[3],
        "status_db": row[4], "n_exp_db": row[5],
        "total_db": float(row[6]) if row[6] else 0,
    })

print(f"Total reports: {len(reports)}")

# Fetch all 60 reports from API with expenses + payment methods
print("Fetching from API...")
api_data = {}
for r in reports:
    rid = r["report_id"]
    time.sleep(0.4)
    try:
        resp = requests.get(f"{BASE_URL}/v2/reports/{rid}?include=expenses.payment_method", headers=HEADERS, timeout=30)
        if resp.status_code != 200:
            print(f"  ERROR rid={rid}: {resp.status_code}")
            continue
        data = resp.json().get("data", {})
        api_status = data.get("status", "")
        expenses = data.get("expenses", {}).get("data", [])
        
        exp_list = []
        for exp in expenses:
            pm = exp.get("payment_method", {})
            if isinstance(pm, dict):
                pm_data = pm.get("data", {})
                pm_name = pm_data.get("description", f"ID:{exp.get('payment_method_id')}")
            else:
                pm_name = f"ID:{exp.get('payment_method_id')}"
            
            exp_list.append({
                "exp_id": exp.get("id"),
                "value": float(exp.get("value", 0) or 0),
                "date": exp.get("date", ""),
                "title": exp.get("title", ""),
                "pm_id": exp.get("payment_method_id"),
                "pm_name": pm_name,
                "is_itau": "itau" in pm_name.lower() or "itaú" in pm_name.lower(),
            })
        
        api_data[rid] = {
            "status": api_status,
            "n_exp": len(expenses),
            "expenses": exp_list,
        }
        print(f"  rid={rid} | {r['name'][:30]} | {api_status} | {len(expenses)} exp")
    except Exception as e:
        print(f"  ERROR rid={rid}: {e}")

conn.close()

# Also read the other sheets from gap file
wb_in2 = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

def read_sheet_rows(ws, skip=1):
    """Read sheet rows, skipping header rows."""
    all_rows = list(ws.iter_rows(values_only=True))
    return all_rows[skip:]

# MAIS DESPESAS NO NEON
ws3 = wb_in2["MAIS DESPESAS NO NEON"]
mais_neon = []
for row in read_sheet_rows(ws3):
    if row[0] is None:
        continue
    try:
        mais_neon.append({
            "report_id": row[0], "name": row[1], "user": row[2], "cpf": row[3],
            "status": row[4], "ref_count": row[5], "neon_count": row[6],
            "diff_count": row[7], "ref_total": float(row[8]) if row[8] else 0,
            "neon_total": float(row[9]) if row[9] else 0, "diff_value": float(row[10]) if row[10] else 0,
        })
    except (ValueError, TypeError):
        continue

# MAIS DESPESAS NO REF
ws4 = wb_in2["MAIS DESPESAS NO REF"]
mais_ref = []
for row in read_sheet_rows(ws4):
    if row[0] is None:
        continue
    try:
        mais_ref.append({
            "report_id": row[0], "name": row[1], "cpf": row[3], "status": row[4],
            "ref_count": row[5], "neon_count": row[6], "diff_count": row[7],
            "ref_total": float(row[8]) if row[8] else 0, "neon_total": float(row[9]) if row[9] else 0,
            "diff_value": float(row[10]) if row[10] else 0,
        })
    except (ValueError, TypeError):
        continue

# VALOR DIFERENTE
ws5 = wb_in2["VALOR DIFERENTE"]
valor_dif = []
for row in read_sheet_rows(ws5):
    if row[0] is None:
        continue
    try:
        valor_dif.append({
            "exp_id": row[0], "report_id": row[1], "name": row[2],
            "cpf": row[4], "ref_value": float(row[5]) if row[5] else 0,
            "neon_value": float(row[6]) if row[6] else 0, "diff": float(row[7]) if row[7] else 0,
        })
    except (ValueError, TypeError):
        continue

wb_in2.close()

# Build output Excel
wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
itau_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

def style_header(ws, ncols):
    for col in range(1, ncols+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def auto_width(ws, max_width=55):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value))
                if val_len > max_len:
                    max_len = val_len
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

# Sheet 1: RESUMO
ws1 = wb.active
ws1.title = "RESUMO"

original_total = sum(r["total_db"] for r in reports)
itau_total = sum(r["total_db"] for r in reports if r["report_id"] in api_data and all(e["is_itau"] for e in api_data[r["report_id"]]["expenses"]))
status_excluded = sum(r["total_db"] for r in reports if r["report_id"] in api_data and api_data[r["report_id"]]["status"] != "APROVADO")
remaining_total = original_total - itau_total - status_excluded

summary = [
    ["Métrica", "Valor", "Quantidade", "% do Gap"],
    ["Gap original (60 reports)", f"R$ {original_total:.2f}", 60, "100.0%"],
    ["Excluídos por Itaú (100% despesas)", f"R$ {itau_total:.2f}", sum(1 for r in reports if r["report_id"] in api_data and all(e["is_itau"] for e in api_data[r["report_id"]]["expenses"])), f"{itau_total/original_total*100:.1f}%"],
    ["Excluídos por status (não APROVADO)", f"R$ {status_excluded:.2f}", sum(1 for r in reports if r["report_id"] in api_data and api_data[r["report_id"]]["status"] != "APROVADO"), f"{status_excluded/original_total*100:.1f}%"],
    ["REMANESCENTE (gap não explicado)", f"R$ {remaining_total:.2f}", sum(1 for r in reports if r["report_id"] in api_data and api_data[r["report_id"]]["status"] == "APROVADO" and not all(e["is_itau"] for e in api_data[r["report_id"]]["expenses"])), f"{remaining_total/original_total*100:.1f}%"],
    ["", "", "", ""],
    ["Redução do gap", f"R$ {original_total - remaining_total:.2f}", "", f"{(1 - remaining_total/original_total)*100:.1f}%"],
    ["", "", "", ""],
    ["MAIS DESPESAS NO NEON (reports)", f"R$ {sum(r['diff_value'] for r in mais_neon):.2f}", len(mais_neon), ""],
    ["MAIS DESPESAS NO REF (reports)", f"R$ {sum(r['diff_value'] for r in mais_ref):.2f}", len(mais_ref), ""],
    ["VALOR DIFERENTE (expenses)", f"R$ {sum(abs(r['diff']) for r in valor_dif):.2f}", len(valor_dif), ""],
]

for row_data in summary:
    ws1.append(row_data)
style_header(ws1, 4)
auto_width(ws1)

# Sheet 2: REPORTS DETAIL
ws2 = wb.create_sheet("REPORTS DETAIL")
headers = ["Report ID", "Name", "User", "CPF", "DB Status", "API Status", "DB N Exp", "API N Exp", "DB Total", "API Total", "N Itaú Exp", "N Non-Itaú Exp", "% Itaú", "Itaú Total", "Non-Itaú Total", "Category", "Action"]
ws2.append(headers)

for r in sorted(reports, key=lambda x: -x["total_db"]):
    rid = r["report_id"]
    api = api_data.get(rid, {})
    api_status = api.get("status", "N/A")
    api_n_exp = api.get("n_exp", 0)
    api_expenses = api.get("expenses", [])
    api_total = sum(e["value"] for e in api_expenses)
    itau_count = sum(1 for e in api_expenses if e["is_itau"])
    non_itau_count = len(api_expenses) - itau_count
    itau_total_r = sum(e["value"] for e in api_expenses if e["is_itau"])
    non_itau_total_r = sum(e["value"] for e in api_expenses if not e["is_itau"])
    pct_itau = f"{itau_count/len(api_expenses)*100:.0f}%" if api_expenses else "N/A"
    
    # Category
    if api_status != "APROVADO":
        category = f"STATUS: {api_status}"
        action = "Exclude (not approved)"
    elif itau_count == len(api_expenses) and len(api_expenses) > 0:
        category = "100% ITAÚ"
        action = "Exclude (all Itaú card)"
    elif itau_count > 0 and non_itau_count > 0:
        category = f"MIXED ({itau_count} Itaú + {non_itau_count} other)"
        action = "Keep, but filter Itaú expenses"
    else:
        category = "NON-ITAÚ"
        action = "Keep (genuine VExpenses)"
    
    ws2.append([rid, r["name"], r["user"], r["cpf"], r["status_db"], api_status,
                r["n_exp_db"], api_n_exp, r["total_db"], api_total,
                itau_count, non_itau_count, pct_itau, itau_total_r, non_itau_total_r,
                category, action])
    
    # Color rows
    row_idx = ws2.max_row
    if "100% ITAÚ" in category:
        for col in range(1, len(headers)+1):
            ws2.cell(row=row_idx, column=col).fill = itau_fill
    elif "STATUS" in category:
        for col in range(1, len(headers)+1):
            ws2.cell(row=row_idx, column=col).fill = warn_fill
    elif "NON-ITAÚ" in category:
        for col in range(1, len(headers)+1):
            ws2.cell(row=row_idx, column=col).fill = ok_fill

style_header(ws2, len(headers))
auto_width(ws2)

# Sheet 3: EXPENSE DETAIL (all expenses from all 60 reports)
ws3 = wb.create_sheet("EXPENSE DETAIL")
headers = ["Report ID", "Report Name", "User", "API Status", "Expense ID", "Date", "Title", "Value", "Payment Method ID", "Payment Method Name", "Is Itaú?", "Category"]
ws3.append(headers)

for r in sorted(reports, key=lambda x: -x["total_db"]):
    rid = r["report_id"]
    api = api_data.get(rid, {})
    api_status = api.get("status", "N/A")
    for exp in api.get("expenses", []):
        ws3.append([rid, r["name"], r["user"], api_status,
                    exp["exp_id"], exp["date"], exp["title"], exp["value"],
                    exp["pm_id"], exp["pm_name"], "YES" if exp["is_itau"] else "NO",
                    "EXCLUDE (Itaú)" if exp["is_itau"] else "KEEP"])
        
        row_idx = ws3.max_row
        if exp["is_itau"]:
            for col in range(1, len(headers)+1):
                ws3.cell(row=row_idx, column=col).fill = itau_fill

style_header(ws3, len(headers))
auto_width(ws3)

# Sheet 4: REMAINING GAP (non-Itaú, APROVADO)
ws4 = wb.create_sheet("REMAINING GAP")
headers = ["Report ID", "Name", "User", "CPF", "API Status", "N Exp", "Total (R$)", "N Itaú Exp", "N Non-Itaú", "Non-Itaú Total", "Notes"]
ws4.append(headers)

remaining_reports = []
for r in sorted(reports, key=lambda x: -x["total_db"]):
    rid = r["report_id"]
    api = api_data.get(rid, {})
    if api.get("status") != "APROVADO":
        continue
    api_expenses = api.get("expenses", [])
    itau_count = sum(1 for e in api_expenses if e["is_itau"])
    if itau_count == len(api_expenses) and len(api_expenses) > 0:
        continue
    non_itau_total = sum(e["value"] for e in api_expenses if not e["is_itau"])
    
    # Notes
    if r["name"].startswith("CAIXA 04/2026"):
        notes = "CAIXA 04/2026 batch — may be from a different ref period"
    elif itau_count > 0:
        notes = f"MIXED: {itau_count} Itaú + {len(api_expenses)-itau_count} non-Itaú expenses"
    else:
        notes = "Genuine VExpenses — needs investigation"
    
    ws4.append([rid, r["name"], r["user"], r["cpf"], api.get("status", ""),
                api.get("n_exp", 0), r["total_db"], itau_count,
                len(api_expenses)-itau_count, non_itau_total, notes])
    remaining_reports.append(r)

style_header(ws4, len(headers))
auto_width(ws4)

# Sheet 5: MAIS DESPESAS with Itaú info
ws5 = wb.create_sheet("MAIS DESPESAS + ITAU")
headers = ["Report ID", "Name", "User", "Status", "Ref Count", "Neon Count", "Diff Count", "Ref Total", "Neon Total", "Diff Value", "API Status", "N Itaú Exp", "N Non-Itaú", "Category"]
ws5.append(headers)

for r in sorted(mais_neon, key=lambda x: -x["diff_value"]):
    rid = r["report_id"]
    api = api_data.get(rid, {})
    api_status = api.get("status", "N/A")
    api_expenses = api.get("expenses", [])
    itau_count = sum(1 for e in api_expenses if e["is_itau"])
    non_itau_count = len(api_expenses) - itau_count
    
    if api_status != "APROVADO":
        category = f"STATUS: {api_status}"
    elif itau_count > 0:
        category = f"MIXED ({itau_count} Itaú + {non_itau_count} other)"
    else:
        category = "NON-ITAÚ (expenses added after ref)"
    
    ws5.append([rid, r["name"], r["user"], r["status"], r["ref_count"], r["neon_count"],
                r["diff_count"], r["ref_total"], r["neon_total"], r["diff_value"],
                api_status, itau_count, non_itau_count, category])

style_header(ws5, len(headers))
auto_width(ws5)

# Sheet 6: PAYMENT METHOD MAP
ws6 = wb.create_sheet("PAYMENT METHODS")
ws6.append(["Payment Method ID", "Description", "Is Itaú?", "Count in 60 reports"])
pm_counter = Counter()
for rid, api in api_data.items():
    for exp in api.get("expenses", []):
        pm_counter[(exp["pm_id"], exp["pm_name"])] += 1

for (pm_id, pm_name), cnt in pm_counter.most_common():
    is_itau = "YES" if "itau" in pm_name.lower() or "itaú" in pm_name.lower() else "NO"
    ws6.append([pm_id, pm_name, is_itau, cnt])
style_header(ws6, 4)
auto_width(ws6)

# Save
output_path = Path(r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\controle-api\data\gap_analysis_itau_detail.xlsx")
wb.save(output_path)
print(f"\nSaved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"\nRemaining gap: R$ {remaining_total:.2f} ({sum(1 for r in reports if r['report_id'] in api_data and api_data[r['report_id']]['status'] == 'APROVADO' and not all(e['is_itau'] for e in api_data[r['report_id']]['expenses']))} reports)")
