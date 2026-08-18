#!/usr/bin/env python3
"""Read the gap analysis Excel to understand all discrepancy categories."""
import openpyxl
from pathlib import Path

xlsx_path = Path(r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\controle-api\data\gap entre referencia e neon ahahahahaah.xlsx")
wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

print("=== SHEETS ===")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name} ({len(rows)} rows)")
    print(f"{'='*80}")
    if rows:
        # Print header
        header = rows[0]
        print(f"Headers: {header}")
        # Print first 5 data rows
        for i, row in enumerate(rows[1:6], 1):
            print(f"  Row {i}: {row}")
        if len(rows) > 6:
            print(f"  ... ({len(rows)-1} total data rows)")

wb.close()
