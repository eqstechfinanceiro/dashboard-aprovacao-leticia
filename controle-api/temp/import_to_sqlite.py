"""
Import data from Excel/XLSB files to SQLite database.
Handles formula detection (XLSX only), pre-header rows, and multiple tables per sheet.
"""
import os
import sqlite3
import json
import re
import openpyxl
import pyxlsb
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
DB_PATH = os.path.join(DATA_DIR, "spreadsheets.db")

XLSX_FILE = os.path.join(DATA_DIR, "CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx")
XLSB_FILE = os.path.join(DATA_DIR, "CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsb")


def sanitize_name(name):
    """Convert a column/table name to a valid SQLite identifier."""
    if not name:
        return "col_unnamed"
    name = str(name).strip()
    # Replace special chars
    name = re.sub(r'[^\w\s]', '_', name, flags=re.UNICODE)
    name = re.sub(r'\s+', '_', name)
    name = re.sub(r'_+', '_', name)
    name = name.strip('_')
    if not name:
        return "col_unnamed"
    # Don't start with digit
    if name[0].isdigit():
        name = 'col_' + name
    return name.lower()[:60]


def col_letter(n):
    """Convert 0-based column index to Excel column letter."""
    result = ''
    n += 1
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def create_metadata_tables(conn):
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS spreadsheet_info (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        file_name TEXT NOT NULL,
        sheet_name TEXT NOT NULL,
        table_name TEXT NOT NULL UNIQUE,
        header_row INTEGER,
        data_start_row INTEGER,
        total_rows INTEGER,
        pre_header_notes TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS column_info (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        spreadsheet_id INTEGER NOT NULL REFERENCES spreadsheet_info(id),
        column_letter TEXT,
        column_name TEXT NOT NULL,
        table_column_name TEXT NOT NULL,
        is_formula INTEGER NOT NULL DEFAULT 0,
        formula_sample TEXT,
        is_api_sourced INTEGER NOT NULL DEFAULT 0,
        col_order INTEGER,
        notes TEXT
    );
    """)
    conn.commit()


def safe_value(v):
    """Convert cell value to a safe SQLite-storable type."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, bool):
        return int(v)
    s = str(v).strip()
    if s == '':
        return None
    return s


def import_xlsx_planilha1(conn):
    """Import CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx - Planilha1"""
    print("\n--- Importing XLSX: Planilha1 ---")

    wb = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=True)
    ws = wb['Planilha1']

    # Also load with formulas to detect them
    wb_f = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=False)
    ws_f = wb_f['Planilha1']

    # Header row is 6, data starts at 7
    HEADER_ROW = 6
    DATA_START = 7

    # Read headers
    headers = {}  # col_index -> name
    for cell in ws[HEADER_ROW]:
        if cell.value is not None:
            headers[cell.column] = str(cell.value).strip()

    print(f"  Headers: {headers}")

    # Detect formula columns by checking data rows
    formula_cols = {}  # col_index -> sample formula
    for row in ws_f.iter_rows(min_row=DATA_START, max_row=DATA_START + 5):
        for cell in row:
            if hasattr(cell, 'column') and cell.column in headers:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    if cell.column not in formula_cols:
                        formula_cols[cell.column] = cell.value[:200]

    print(f"  Formula columns: {[headers.get(c, c) for c in formula_cols]}")

    # Pre-header notes
    pre_notes = []
    # Row 4: N4 = 0.5 (constant multiplier)
    pre_notes.append("Row 4: N4=0.5 (multiplier constant used in REEMBOLSO formula)")
    # Row 5: SUBTOTAL formulas
    pre_notes.append("Row 5: H5:O5 = SUBTOTAL(9,...) aggregation formulas")

    # Build sanitized column map
    col_map = {}  # col_index -> sanitized_name
    used_names = {}
    for col_idx, name in headers.items():
        san = sanitize_name(name)
        if san in used_names.values():
            san = san + f'_{col_idx}'
        col_map[col_idx] = san

    TABLE_NAME = 'carga_1qz_planilha1'

    # Drop and recreate table
    col_defs = [f'"{col_map[c]}" TEXT' for c in sorted(col_map.keys())]
    conn.execute(f'DROP TABLE IF EXISTS "{TABLE_NAME}"')
    conn.execute(f'CREATE TABLE "{TABLE_NAME}" ({", ".join(col_defs)})')

    # Insert data
    cols_sorted = sorted(col_map.keys())
    col_names = [col_map[c] for c in cols_sorted]
    placeholders = ', '.join(['?' for _ in cols_sorted])
    quoted_cols = ', '.join(f'"{c}"' for c in col_names)
    insert_sql = f'INSERT INTO "{TABLE_NAME}" ({quoted_cols}) VALUES ({placeholders})'

    row_count = 0
    batch = []
    for row in ws.iter_rows(min_row=DATA_START):
        row_dict = {cell.column: cell.value for cell in row if hasattr(cell, 'column')}
        if not any(row_dict.get(c) is not None for c in cols_sorted):
            continue
        values = [safe_value(row_dict.get(c)) for c in cols_sorted]
        batch.append(values)
        row_count += 1
        if len(batch) >= 200:
            conn.executemany(insert_sql, batch)
            batch = []

    if batch:
        conn.executemany(insert_sql, batch)
    conn.commit()
    print(f"  Inserted {row_count} rows")

    # Register in spreadsheet_info
    conn.execute('DELETE FROM spreadsheet_info WHERE table_name = ?', (TABLE_NAME,))
    cur = conn.execute(
        'INSERT INTO spreadsheet_info (file_name, sheet_name, table_name, header_row, data_start_row, total_rows, pre_header_notes) VALUES (?,?,?,?,?,?,?)',
        ('CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx', 'Planilha1', TABLE_NAME, HEADER_ROW, DATA_START, row_count, '\n'.join(pre_notes))
    )
    sheet_id = cur.lastrowid

    # Register columns
    conn.execute('DELETE FROM column_info WHERE spreadsheet_id = ?', (sheet_id,))
    for i, col_idx in enumerate(cols_sorted):
        name = headers[col_idx]
        san = col_map[col_idx]
        is_formula = 1 if col_idx in formula_cols else 0
        formula_sample = formula_cols.get(col_idx)
        conn.execute(
            'INSERT INTO column_info (spreadsheet_id, column_letter, column_name, table_column_name, is_formula, formula_sample, col_order) VALUES (?,?,?,?,?,?,?)',
            (sheet_id, col_letter(col_idx - 1), name, san, is_formula, formula_sample, i)
        )

    conn.commit()
    wb.close()
    wb_f.close()
    print(f"  Done: {TABLE_NAME}")
    return TABLE_NAME


def read_xlsb_sheet(sheet_name, header_row, data_start_row, col_start=None, col_end=None):
    """Read a sheet from the XLSB file. Returns (headers_dict, rows_list).
    headers_dict: {col_idx: header_name}
    rows_list: list of {col_idx: value} dicts
    col_start/col_end: 0-based column index filter (inclusive)
    """
    headers = {}
    rows = []

    with pyxlsb.open_workbook(XLSB_FILE) as wb:
        with wb.get_sheet(sheet_name) as ws:
            row_idx = 0
            for row in ws.rows():
                row_idx += 1
                if row_idx == header_row:
                    for cell in row:
                        c = cell.c  # 0-based
                        if col_start is not None and c < col_start:
                            continue
                        if col_end is not None and c > col_end:
                            continue
                        if cell.v is not None:
                            headers[c] = str(cell.v).strip()
                elif row_idx >= data_start_row:
                    row_dict = {}
                    has_data = False
                    for cell in row:
                        c = cell.c
                        if col_start is not None and c < col_start:
                            continue
                        if col_end is not None and c > col_end:
                            continue
                        if c in headers and cell.v is not None:
                            row_dict[c] = cell.v
                            has_data = True
                    if has_data:
                        rows.append(row_dict)
    return headers, rows


def import_xlsb_sheet(conn, sheet_name, table_name, header_row, data_start_row,
                       col_start=None, col_end=None, pre_notes=None, formula_cols_info=None):
    """Generic importer for XLSB sheets."""
    print(f"\n--- Importing XLSB: {sheet_name} -> {table_name} ---")

    headers, rows = read_xlsb_sheet(sheet_name, header_row, data_start_row, col_start, col_end)
    if not headers:
        print(f"  No headers found, skipping")
        return None

    print(f"  Headers: {list(headers.values())}")
    print(f"  Rows found: {len(rows)}")

    col_map = {}
    used = {}
    for c in sorted(headers.keys()):
        san = sanitize_name(headers[c])
        if san in used.values():
            san = san + f'_{c}'
        used[c] = san
        col_map[c] = san

    cols_sorted = sorted(col_map.keys())
    col_defs = [f'"{col_map[c]}" TEXT' for c in cols_sorted]
    conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
    conn.execute(f'CREATE TABLE "{table_name}" ({", ".join(col_defs)})')

    col_names = [col_map[c] for c in cols_sorted]
    placeholders = ', '.join(['?' for _ in cols_sorted])
    quoted_cols2 = ', '.join(f'"{c}"' for c in col_names)
    insert_sql = f'INSERT INTO "{table_name}" ({quoted_cols2}) VALUES ({placeholders})'

    batch = []
    for row_dict in rows:
        values = [safe_value(row_dict.get(c)) for c in cols_sorted]
        batch.append(values)
        if len(batch) >= 500:
            conn.executemany(insert_sql, batch)
            batch = []

    if batch:
        conn.executemany(insert_sql, batch)
    conn.commit()

    # Register metadata
    conn.execute('DELETE FROM spreadsheet_info WHERE table_name = ?', (table_name,))
    cur = conn.execute(
        'INSERT INTO spreadsheet_info (file_name, sheet_name, table_name, header_row, data_start_row, total_rows, pre_header_notes) VALUES (?,?,?,?,?,?,?)',
        ('CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsb', sheet_name, table_name,
         header_row, data_start_row, len(rows), pre_notes or '')
    )
    sheet_id = cur.lastrowid

    conn.execute('DELETE FROM column_info WHERE spreadsheet_id = ?', (sheet_id,))
    for i, col_idx in enumerate(cols_sorted):
        name = headers[col_idx]
        san = col_map[col_idx]
        finfo = (formula_cols_info or {}).get(col_idx, {})
        conn.execute(
            'INSERT INTO column_info (spreadsheet_id, column_letter, column_name, table_column_name, is_formula, formula_sample, notes, col_order) VALUES (?,?,?,?,?,?,?,?)',
            (sheet_id, col_letter(col_idx), name, san,
             finfo.get('is_formula', 0), finfo.get('formula_sample'),
             finfo.get('notes'), i)
        )

    conn.commit()
    print(f"  Done: {table_name}")
    return table_name


def main():
    print(f"Database: {DB_PATH}")

    # Remove existing database
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        print("Removed existing database")

    conn = sqlite3.connect(DB_PATH, timeout=60)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")

    create_metadata_tables(conn)

    # ---- XLSX ----
    import_xlsx_planilha1(conn)

    # ---- XLSB sheets ----

    # PAINEL: header row 11, data starts row 12, cols A(0)-AA(26)
    # Formula cols analysis: R=SALDO_PRESTACAO (N+O+P+Q), T=SALDO_FINAL (R-S)
    # W=ADICIONAIS (computed), X=SITUACAO_COLABORADOR (IF formula)
    painel_formula_cols = {
        17: {'is_formula': 1, 'notes': 'Calculated: CARGA+TRANSFERENCIA-TARIFA-PRESTACAO_CONTAS'},  # R (col index 17)
        19: {'is_formula': 1, 'notes': 'Calculated: SALDO_PRESTACAO - SALDO_CARTAO'},  # T
        22: {'is_formula': 1, 'notes': 'Calculated: sum of additional charges'},  # W
        23: {'is_formula': 1, 'notes': 'IF formula: status based on processing state'},  # X
    }
    import_xlsb_sheet(
        conn, 'PAINEL', 'controle_painel',
        header_row=11, data_start_row=12,
        col_start=0, col_end=26,
        pre_notes='Row 7-8: ANO/MÊS config (V7=ANO, W7=2026, V8=MÊS, W8=MAIO)\nRow 10: summary totals for numeric columns',
        formula_cols_info=painel_formula_cols
    )

    # SALDO CARTAO: Two tables on same sheet
    # Table 1: header row 4, cols B(1)-G(6)
    import_xlsb_sheet(
        conn, 'SALDO CARTAO', 'controle_saldo_cartao',
        header_row=4, data_start_row=5,
        col_start=1, col_end=6,
        pre_notes='Row 3: D3=total VALOR (sum), L3=total col L sum\nTwo tables side-by-side: this is Table 1 (cols B-G)'
    )
    # Table 2: header row 4, cols J(9)-N(13)
    import_xlsb_sheet(
        conn, 'SALDO CARTAO', 'controle_saldo_cartao_resumo',
        header_row=4, data_start_row=5,
        col_start=9, col_end=13,
        pre_notes='Table 2 (cols J-N): unique per-collaborator card balances'
    )

    # ADICIONAL ITAÚ: header row 4, cols B(1)-L(11)
    import_xlsb_sheet(
        conn, 'ADICIONAL ITAÚ', 'controle_adicional_itau',
        header_row=4, data_start_row=5,
        col_start=1, col_end=11,
        pre_notes='Row 3: F3=total ADICIONADO (sum of all additions)'
    )

    # ADICIONAIS: Two tables
    # Table 1: header row 3, cols B(1)-N(13)
    adicionais_formula_cols = {
        19: {'is_formula': 1, 'notes': 'T=difference VALOR-VLR.UTILIZADO (calculated)'},  # T
    }
    import_xlsb_sheet(
        conn, 'ADICIONAIS', 'controle_adicionais',
        header_row=3, data_start_row=4,
        col_start=1, col_end=13,
        pre_notes='Row 2: D2=total VALOR, R2=total PEDIDO, S2=VLR.UTILIZADO, T2=difference\nTwo tables: this is Table 1 (cols B-N)'
    )
    # Table 2: cols P(15)-S(18) - pedidos summary
    import_xlsb_sheet(
        conn, 'ADICIONAIS', 'controle_adicionais_pedidos',
        header_row=3, data_start_row=4,
        col_start=15, col_end=18,
        pre_notes='Table 2 (cols P-S): pedido summary with MÊS, PEDIDO, VALOR, VLR.UTILIZADO'
    )

    # QUINZENAS: header row 4, cols A(0)-I(8)
    import_xlsb_sheet(
        conn, 'QUINZENAS', 'controle_quinzenas',
        header_row=4, data_start_row=5,
        col_start=0, col_end=8,
        pre_notes='Row 3: C3=total VALOR (possibly a partial sum)'
    )

    # SALDOS ADM EQS: header row 4, cols A(0)-K(10)
    saldos_formula_cols = {
        8: {'is_formula': 1, 'notes': 'TOTAL QZ_2025 = 1QZ_DIRETORIA + 2QZ_DIRETORIA (sum)'},  # I
    }
    import_xlsb_sheet(
        conn, 'SALDOS ADM EQS', 'controle_saldos_adm',
        header_row=4, data_start_row=5,
        col_start=0, col_end=10,
        pre_notes='Row 1: D1=ANO, E1=year value\nRow 2: A2=sheet title, D2=MÊS, E2=month name, G2=VEXPENSES label\nRow 3: I3=total TOTAL_QZ_2025',
        formula_cols_info=saldos_formula_cols
    )

    # EXTRATO: header row 8, cols B(1)-M(12)
    import_xlsb_sheet(
        conn, 'EXTRATO', 'controle_extrato',
        header_row=8, data_start_row=9,
        col_start=1, col_end=12,
        pre_notes='Row 3: CARGA count/total; Row 4: TRANSFERÊNCIA count/total; Row 5: TARIFA count/total\nRow 7: L7=reference value (3500?)\nNote: Data/Hora columns are Excel serial numbers'
    )

    # PAINEL PRESTAÇÕES: pivot table, two side-by-side summaries
    # Table 1: header row 3, cols B(1)-C(2)
    import_xlsb_sheet(
        conn, 'PAINEL PRESTAÇÕES', 'controle_painel_prestacoes_desativados',
        header_row=3, data_start_row=4,
        col_start=1, col_end=2,
        pre_notes='Pivot table: desativated collaborators with sum of expense values'
    )
    # Table 2: cols G(6)-H(7)
    import_xlsb_sheet(
        conn, 'PAINEL PRESTAÇÕES', 'controle_painel_prestacoes_ativos',
        header_row=3, data_start_row=4,
        col_start=6, col_end=7,
        pre_notes='Pivot table: active collaborators with sum of expense values'
    )

    # BASE PREST: header row 3, cols A(0)-AE(30)
    import_xlsb_sheet(
        conn, 'BASE PREST ', 'controle_base_prestacoes',
        header_row=3, data_start_row=4,
        col_start=0, col_end=30,
        pre_notes='Row 2: AA2=total VALOR sum\nSource: VExpenses expense reports export'
    )

    # REEMBOLSO: header row 3, cols B(1)-J(9)
    import_xlsb_sheet(
        conn, 'REEMBOLSO', 'controle_reembolso',
        header_row=3, data_start_row=4,
        col_start=1, col_end=9,
        pre_notes='Row 2: D2=total VALOR (sum of all reimbursements)'
    )

    # ESTORNO - SAQUE: header row 1, cols A(0)-H(7) — no pre-header!
    import_xlsb_sheet(
        conn, 'ESTORNO - SAQUE', 'controle_estorno_saque',
        header_row=1, data_start_row=2,
        col_start=0, col_end=7,
        pre_notes='Headers start at row 1 (no pre-header rows)'
    )

    # Detalhes1/2/3: header row 3, cols A(0)-AC(28)
    for i, det in enumerate(['Detalhes1', 'Detalhes2', 'Detalhes3'], 1):
        import_xlsb_sheet(
            conn, det, f'controle_detalhes{i}',
            header_row=3, data_start_row=4,
            col_start=0, col_end=28,
            pre_notes=f'Row 1: pivot detail title\nDrill-down from PAINEL PRESTAÇÕES pivot'
        )

    # AUX: header row 2, cols B(1)-E(4)
    import_xlsb_sheet(
        conn, 'AUX', 'controle_aux',
        header_row=2, data_start_row=3,
        col_start=1, col_end=4,
        pre_notes='Reference table: REGIONAL -> GESTOR, DIRETOR mapping'
    )

    # Create useful indexes
    print("\n--- Creating indexes ---")
    indexes = [
        'CREATE INDEX IF NOT EXISTS idx_carga_cpf ON carga_1qz_planilha1(cpf)',
        'CREATE INDEX IF NOT EXISTS idx_painel_cpf ON controle_painel(cpf)',
        'CREATE INDEX IF NOT EXISTS idx_extrato_cpf ON controle_extrato(cpf)',
        'CREATE INDEX IF NOT EXISTS idx_base_cpf ON controle_base_prestacoes(cpf_cnpj)',
        'CREATE INDEX IF NOT EXISTS idx_quinzenas_cpf ON controle_quinzenas(cpf)',
        'CREATE INDEX IF NOT EXISTS idx_saldo_cartao_cpf ON controle_saldo_cartao(cpf)',
    ]
    for idx_sql in indexes:
        try:
            conn.execute(idx_sql)
        except Exception as e:
            print(f"  Index warning: {e}")
    conn.commit()

    # Final summary
    print("\n=== IMPORT SUMMARY ===")
    for row in conn.execute('SELECT table_name, sheet_name, total_rows FROM spreadsheet_info ORDER BY id'):
        print(f"  {row[0]}: {row[2]} rows (from sheet '{row[1]}')")

    conn.close()
    print(f"\nDatabase saved to: {DB_PATH}")


if __name__ == '__main__':
    main()
