#!/usr/bin/env python3
"""
Focused investigation:
1. Check status_cartao for the 20 missing users
2. Check status_cartao for ALL users in reference PAINEL (are they all "Cartão ativo"?)
3. Check PAINEL header structure to fix column mapping
4. Check: users with status_cartao != "Cartão ativo" that have expense reports
"""
import os
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from pathlib import Path
import openpyxl
from collections import defaultdict

BASE = Path(__file__).parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"

def norm(s):
    return " ".join(str(s or "").upper().strip().split())

conn = psycopg2.connect(NEON_URL, connect_timeout=10)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# ============================================================
# 1. Check PAINEL header structure
# ============================================================
print("=" * 80)
print("1. PAINEL header structure (row 11)")
print("=" * 80)

wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_p = wb["PAINEL"]
headers = []
for col_idx, cell in enumerate(ws_p[11]):
    headers.append((col_idx, cell.value))
    print(f"  col[{col_idx}] = {cell.value}")

# Read first data row (row 12)
print("\n  First data row (row 12):")
for col_idx, cell in enumerate(ws_p[12]):
    if cell.value is not None:
        print(f"  col[{col_idx}] = {cell.value}")

# Find CPF and colaborador columns
cpf_col = None
colab_col = None
for idx, h in headers:
    if h and "CPF" in str(h).upper():
        cpf_col = idx
    if h and "COLAB" in str(h).upper():
        colab_col = idx
print(f"\n  CPF column: {cpf_col}, Colaborador column: {colab_col}")

# ============================================================
# 2. Get reference PAINEL CPFs (using correct column)
# ============================================================
print("\n" + "=" * 80)
print("2. Reference PAINEL CPFs (using correct column)")
print("=" * 80)

painel_cpfs = set()
painel_colabs = {}
if cpf_col is not None:
    for row in ws_p.iter_rows(min_row=12, values_only=True):
        cpf = str(row[cpf_col] or "").strip() if len(row) > cpf_col else ""
        colab = str(row[colab_col] or "").strip() if colab_col is not None and len(row) > colab_col else ""
        if cpf and cpf != "None":
            painel_cpfs.add(cpf)
            painel_colabs[cpf] = colab
print(f"  Total PAINEL CPFs: {len(painel_cpfs)}")

# ============================================================
# 3. status_cartao for 20 missing users
# ============================================================
print("\n" + "=" * 80)
print("3. status_cartao for 20 'missing' users")
print("=" * 80)

users_no_ref = [
    "AFONSO FIORELLO CARVALHO", "AILTON MENDES AGUILAR", "AMARAL RODRIGUES NUNES",
    "ANA LUIZA OLIVEIRA CAMELO", "ATILA SILVA DOS SANTOS", "DANIEL PORFIRIO DE SOUSA",
    "ERASMO PEREIRA MONTEIRO", "FLAVIO PEREIRA DE SOUZA", "FRANCISCO MICHAEL CASTRO FERNANDES",
    "JACKSON CAROLINO CARNEIRO", "LEONARDO GONCALVES RIBEIRO FILHO", "MARCELO BEZERRA",
    "NELSON MOURA RIBEIRO", "OSMAN DOS SANTOS MORAIS JUNIOR", "PATRICK FERNANDO GOULART ALVES",
    "PEDRO LUIS PIRES DOS SANTOS", "ROBERTO ALIAGA", "SANDRA CRISTINA ALVES MACHADO",
    "WAGNER FERNANDS DA SILVA", "WESLEY CARLOS AUGUSTO",
]

cur.execute("SELECT colaborador, cpf, status_cartao, regional FROM quinzena_cadastro")
all_cad = cur.fetchall()

for user in users_no_ref:
    user_norm = norm(user)
    found = [c for c in all_cad if user_norm in norm(c["colaborador"]) or norm(c["colaborador"]) in user_norm]
    if found:
        for f in found:
            in_painel = f["cpf"] in painel_cpfs
            print(f"  {user[:35]:<35}  CPF={f['cpf']}  status_cartao={f['status_cartao']:<25}  in_painel={in_painel}")
    else:
        print(f"  {user[:35]:<35}  NOT FOUND in cadastro")

# ============================================================
# 4. status_cartao distribution for PAINEL users vs non-PAINEL users
# ============================================================
print("\n" + "=" * 80)
print("4. status_cartao: PAINEL users vs users with reports but NOT in PAINEL")
print("=" * 80)

# All users with expense reports
cur.execute("""
    SELECT DISTINCT r.user_cpf, r.user_name
    FROM prestacao_reports r
    JOIN prestacao_expenses e ON e.report_id = r.id
    WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND r.user_cpf IS NOT NULL
      AND TRIM(r.name) !~* '^(fatu|farur|cart)'
      AND TRIM(r.name) !~* '(fatura|fatuar|fatut|farur)'
""")
users_with_reports = cur.fetchall()

painel_status = defaultdict(int)
non_painel_status = defaultdict(int)

for u in users_with_reports:
    cpf = u["user_cpf"]
    # Find in cadastro
    cad = next((c for c in all_cad if c["cpf"] == cpf), None)
    status = cad["status_cartao"] if cad else "NOT IN CADASTRO"
    
    if cpf in painel_cpfs:
        painel_status[status] += 1
    else:
        non_painel_status[status] += 1

print("  PAINEL users with reports:")
for s, cnt in sorted(painel_status.items(), key=lambda x: -x[1]):
    print(f"    {s}: {cnt}")

print("\n  NON-PAINEL users with reports:")
for s, cnt in sorted(non_painel_status.items(), key=lambda x: -x[1]):
    print(f"    {s}: {cnt}")

# ============================================================
# 5. Key question: How many users have status_cartao != "Cartão ativo" 
#    but still have expense reports included in our prestacao?
# ============================================================
print("\n" + "=" * 80)
print("5. Users with status_cartao != 'Cartão ativo' that have expense reports")
print("=" * 80)

cur.execute("""
    SELECT r.user_cpf, r.user_name,
           COUNT(DISTINCT r.id) as report_count,
           SUM(e.value) as total_value
    FROM prestacao_reports r
    JOIN prestacao_expenses e ON e.report_id = r.id
    WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND r.user_cpf IS NOT NULL
      AND TRIM(r.name) !~* '^(fatu|farur|cart)'
      AND TRIM(r.name) !~* '(fatura|fatuar|fatut|farur)'
    GROUP BY r.user_cpf, r.user_name
    ORDER BY r.user_name
""")
all_users_reports = cur.fetchall()

non_active = []
for u in all_users_reports:
    cad = next((c for c in all_cad if c["cpf"] == u["user_cpf"]), None)
    status = cad["status_cartao"] if cad else "NOT IN CADASTRO"
    if status != "Cartão ativo":
        in_painel = u["user_cpf"] in painel_cpfs
        non_active.append({
            "name": u["user_name"],
            "cpf": u["user_cpf"],
            "status": status,
            "reports": u["report_count"],
            "total": float(u["total_value"]),
            "in_painel": in_painel,
        })

print(f"  Total users with reports: {len(all_users_reports)}")
print(f"  Users with status != 'Cartão ativo': {len(non_active)}")
print(f"  Total value from non-active: R$ {sum(u['total'] for u in non_active):,.2f}")

print(f"\n  Breakdown by status:")
by_status = defaultdict(lambda: {"count": 0, "total": 0, "in_painel": 0})
for u in non_active:
    by_status[u["status"]]["count"] += 1
    by_status[u["status"]]["total"] += u["total"]
    if u["in_painel"]:
        by_status[u["status"]]["in_painel"] += 1

for s, info in sorted(by_status.items(), key=lambda x: -x[1]["count"]):
    s_display = str(s) if s is not None else "None"
    print(f"    {s_display:<25}  {info['count']:>3} users  R$ {info['total']:>12,.2f}  in_painel={info['in_painel']}")

# Show the non-active users NOT in PAINEL
print(f"\n  Non-active users NOT in PAINEL ({sum(1 for u in non_active if not u['in_painel'])}):")
for u in sorted([u for u in non_active if not u["in_painel"]], key=lambda x: -x["total"])[:20]:
    status_display = str(u['status']) if u['status'] is not None else "None"
    print(f"    {u['name'][:30]:<30}  CPF={u['cpf']}  status={status_display:<25}  reports={u['reports']:>3}  R$ {u['total']:>10,.2f}")

# ============================================================
# 6. Check: if we filter by status_cartao = "Cartão ativo", how much do we exclude?
# ============================================================
print("\n" + "=" * 80)
print("6. Impact of filtering by status_cartao = 'Cartão ativo'")
print("=" * 80)

active_total = sum(float(u["total_value"]) for u in all_users_reports 
                   if next((c for c in all_cad if c["cpf"] == u["user_cpf"]), None) 
                   and next((c for c in all_cad if c["cpf"] == u["user_cpf"]), None)["status_cartao"] == "Cartão ativo")
non_active_total = sum(u["total"] for u in non_active)
print(f"  Total prestacao (current): R$ {sum(float(u['total_value']) for u in all_users_reports):,.2f}")
print(f"  From 'Cartão ativo' users: R$ {active_total:,.2f}")
print(f"  From non-active users:     R$ {non_active_total:,.2f}")
print(f"  Non-active in PAINEL:      R$ {sum(u['total'] for u in non_active if u['in_painel']):,.2f}")
print(f"  Non-active NOT in PAINEL:  R$ {sum(u['total'] for u in non_active if not u['in_painel']):,.2f}")

# ============================================================
# 7. Check "CARTÃO VEXPENSES" column (col[12]) in PAINEL
# ============================================================
print("\n" + "=" * 80)
print("7. PAINEL 'CARTÃO VEXPENSES' column (col[12]) distribution")
print("=" * 80)

cartao_vexp = defaultdict(int)
cartao_vexp_cpfs = defaultdict(set)
for row in ws_p.iter_rows(min_row=12, values_only=True):
    if row[2] is None:
        continue
    cpf = str(row[2] or "").strip()
    val = str(row[12] or "").strip() if len(row) > 12 else ""
    cartao_vexp[val] += 1
    cartao_vexp_cpfs[val].add(cpf)

for val, cnt in sorted(cartao_vexp.items(), key=lambda x: -x[1]):
    print(f"  '{val}': {cnt} users")

# ============================================================
# 8. Check "CARTÃO ITAU" column (col[6]) in PAINEL
# ============================================================
print("\n" + "=" * 80)
print("8. PAINEL 'CARTÃO ITAU' column (col[6]) distribution")
print("=" * 80)

cartao_itau = defaultdict(int)
for row in ws_p.iter_rows(min_row=12, values_only=True):
    if row[2] is None:
        continue
    val = str(row[6] or "").strip() if len(row) > 6 else ""
    cartao_itau[val] += 1

for val, cnt in sorted(cartao_itau.items(), key=lambda x: -x[1])[:10]:
    print(f"  '{val}': {cnt} users")

# ============================================================
# 9. Cross: CARTÃO VEXPENSES vs status_cartao for 20 missing users
# ============================================================
print("\n" + "=" * 80)
print("9. CARTÃO VEXPENSES column for 20 'missing' users")
print("=" * 80)

for row in ws_p.iter_rows(min_row=12, values_only=True):
    if row[2] is None:
        continue
    cpf = str(row[2] or "").strip()
    colab = str(row[1] or "").strip()
    cartao_vx = str(row[12] or "").strip() if len(row) > 12 else ""
    cartao_it = str(row[6] or "").strip() if len(row) > 6 else ""
    situacao = str(row[4] or "").strip() if len(row) > 4 else ""
    
    for user in users_no_ref:
        if norm(user) in norm(colab) or norm(colab) in norm(user):
            print(f"  {user[:35]:<35}  CPF={cpf}  cartao_vexp={cartao_vx:<5}  cartao_itau={cartao_it:<10}  situacao={situacao}")
            break

# ============================================================
# 10. Check: users with CARTÃO VEXPENSES = "SIM" vs others
# ============================================================
print("\n" + "=" * 80)
print("10. PAINEL users: CARTÃO VEXPENSES='SIM' vs others — prestacao values")
print("=" * 80)

sim_cpfs = cartao_vexp_cpfs.get("SIM", set())
nao_cpfs = cartao_vexp_cpfs.get("NÃO", set()) | cartao_vexp_cpfs.get("NAO", set())

sim_in_reports = sum(1 for u in all_users_reports if u["user_cpf"] in sim_cpfs)
nao_in_reports = sum(1 for u in all_users_reports if u["user_cpf"] in nao_cpfs)
sim_total = sum(float(u["total_value"]) for u in all_users_reports if u["user_cpf"] in sim_cpfs)
nao_total = sum(float(u["total_value"]) for u in all_users_reports if u["user_cpf"] in nao_cpfs)

print(f"  CARTÃO VEXPENSES='SIM': {len(sim_cpfs)} users in PAINEL, {sim_in_reports} with reports, R$ {sim_total:,.2f}")
print(f"  CARTÃO VEXPENSES='NÃO': {len(nao_cpfs)} users in PAINEL, {nao_in_reports} with reports, R$ {nao_total:,.2f}")

wb.close()
conn.close()
