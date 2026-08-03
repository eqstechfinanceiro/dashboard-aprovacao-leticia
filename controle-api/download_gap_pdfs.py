#!/usr/bin/env python3
"""Download PDFs for all reports in the gap (API only, not in reference)."""
import os, json
import psycopg2
import psycopg2.extras
import urllib.request
import urllib.parse
from dotenv import load_dotenv
from pathlib import Path
from collections import defaultdict
import openpyxl

BASE = Path(__file__).parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")
API_KEY = os.getenv("VEXPENSES_API_KEY", "")
LARAVEL_TOKEN = os.getenv("VEXPENSES_LARAVEL_TOKEN", "")

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"
OUT_DIR = BASE.parent / "data" / "gap_pdfs"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def is_card_report(name):
    n = name.strip().upper()
    if not n:
        return False
    if 'CAIXA ITAU' in n or 'CAIXA ITAÚ' in n:
        return True
    if n.startswith('CAIXA'):
        return False
    if n.startswith(('FATURA', 'CARTAO', 'CARTÃO', 'FATUAR', 'FARTUR', 'FATUT', 'FARUR', 'FATUTR')):
        return True
    if 'CARTÃO DE CRÉDITO' in n or 'CARTAO DE CREDITO' in n or 'CARTÃO DE CREDITO' in n:
        return True
    if 'CARTÃO CORPORATIVO' in n:
        return True
    if ('ITAU' in n or 'ITAÚ' in n) and 'CAIXA' not in n:
        return True
    if 'DOLAR' in n or 'DÓLAR' in n:
        return True
    if n.startswith('DESPESA') and 'FATURA' in n:
        return True
    if n.startswith('COMPLEMENTAR') and 'FATURA' in n:
        return True
    if 'CARTÃO' in n and 'CRÉDITO' in n:
        return True
    if 'CARTAO' in n and 'CREDITO' in n:
        return True
    if n.startswith('CARTÃO VEXPENSES'):
        return True
    return False


def main():
    print("Loading reference BASE PREST...")
    wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
    ws_bp = wb["BASE PREST "]
    ref_report_ids = set()
    for row in ws_bp.iter_rows(min_row=4, values_only=True):
        if row[0] is None:
            continue
        rid = int(row[1]) if row[1] else None
        if rid:
            ref_report_ids.add(rid)

    ws_p = wb["PAINEL"]
    vexpenses_cpfs = set()
    for row in ws_p.iter_rows(min_row=12, values_only=True):
        if row[2] is None:
            continue
        cpf = str(row[2] or "").strip()
        cartao_vx = str(row[12] or "").strip().upper() if len(row) > 12 else ""
        if cartao_vx == "SIM":
            vexpenses_cpfs.add(cpf)
    wb.close()
    print(f"  Reference: {len(ref_report_ids)} reports, {len(vexpenses_cpfs)} CPFs with CARTAO=SIM")

    print("\nLoading API data...")
    conn = psycopg2.connect(NEON_URL, connect_timeout=10)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT e.id, e.report_id, e.value,
               r.name, r.status, r.user_cpf, r.user_name, r.raw_data,
               e.raw_data->>'payment_method_id' as pm_id
        FROM prestacao_expenses e
        JOIN prestacao_reports r ON e.report_id = r.id
        WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
          AND r.user_cpf IS NOT NULL
    """)
    all_api = cur.fetchall()
    conn.close()
    print(f"  API expenses: {len(all_api)}")

    # Load ref expense IDs
    ref_expense_ids = set()
    wb2 = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
    ws_bp2 = wb2["BASE PREST "]
    for row in ws_bp2.iter_rows(min_row=4, values_only=True):
        if row[0] is None:
            continue
        ref_expense_ids.add(int(row[0]))
    wb2.close()

    # Filter: not card report, not pm_id=627401, CPF in vexpenses_cpfs, not in reference
    new_report_ids = set()
    report_info = {}
    for e in all_api:
        name = e["name"] or ""
        if is_card_report(name):
            continue
        if e.get("pm_id") == "627401":
            continue
        if e["user_cpf"] not in vexpenses_cpfs:
            continue
        if e["id"] in ref_expense_ids:
            continue
        rid = e["report_id"]
        if rid not in report_info:
            raw = e["raw_data"] or {}
            if isinstance(raw, str):
                raw = json.loads(raw)
            pdf_link = raw.get("pdf_link", "")
            report_info[rid] = {
                "id": rid,
                "name": name,
                "user": e["user_name"],
                "cpf": e["user_cpf"],
                "status": e["status"],
                "pdf_link": pdf_link,
            }
            new_report_ids.add(rid)

    new_reports = list(report_info.values())

    print(f"\nNew reports with PDF link: {len(new_reports)}")
    print(f"Downloading to: {OUT_DIR}\n")

    for i, r in enumerate(new_reports, 1):
        safe_name = "".join(c for c in r["name"] if c not in r'\/:*?"<>|')
        fname = f"{r['id']}_{safe_name}_{r['user']}.pdf"
        fpath = OUT_DIR / fname
        if fpath.exists():
            print(f"  [{i}/{len(new_reports)}] SKIP (exists): {fname}")
            continue
        try:
            print(f"  [{i}/{len(new_reports)}] Downloading: {fname}")
            req = urllib.request.Request(r["pdf_link"], headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Authorization": API_KEY,
                "Cookie": f"vexpenses_session={urllib.parse.unquote(LARAVEL_TOKEN)}",
            })
            with urllib.request.urlopen(req, timeout=60) as resp:
                data = resp.read()
                if len(data) < 1000 and b"html" in data[:200].lower():
                    print(f"    WARNING: Got HTML instead of PDF (auth required?)")
                    fpath.write_bytes(data)
                else:
                    fpath.write_bytes(data)
                    print(f"    OK: {len(data):,} bytes")
        except Exception as e:
            print(f"    ERROR: {e}")

    print(f"\nDone! {len(list(OUT_DIR.glob('*.pdf')))} PDFs in {OUT_DIR}")


if __name__ == "__main__":
    main()
