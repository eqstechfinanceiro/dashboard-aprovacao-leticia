#!/usr/bin/env python3
"""
Generate a comprehensive analysis Excel with all gap categories explained.
Each report gets a 'reason' column explaining WHY it's not in the ref.
"""
import os, psycopg2, psycopg2.extras, json, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from pathlib import Path
from datetime import datetime
from collections import Counter

load_dotenv(Path(__file__).parent / ".env")

REF_DATE = datetime(2026, 7, 27, 8, 0, 0)

# Read the original gap file
xlsx_path = Path(r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\controle-api\data\gap entre referencia e neon ahahahahaah.xlsx")
wb_in = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

# Read REPORTS SO NO NEON
ws = wb_in["REPORTS SO NO NEON"]
rows = list(ws.iter_rows(values_only=True))
reports_only_neon = []
for row in rows[1:]:
    if row[0] is None:
        continue
    reports_only_neon.append({
        "report_id": row[0], "name": row[1], "user": row[2], "cpf": row[3],
        "status": row[4], "n_expenses": row[5], "total": float(row[6]) if row[6] else 0,
    })

# Read SO NO NEON expenses
ws2 = wb_in["SO NO NEON"]
exp_rows = list(ws2.iter_rows(values_only=True))
so_no_neon_exp = []
for row in exp_rows[1:]:
    if row[0] is None:
        continue
    so_no_neon_exp.append({
        "exp_id": row[0], "report_id": row[1], "report_name": row[2],
        "user": row[3], "cpf": row[4], "status": row[5],
        "value": float(row[6]) if row[6] else 0, "date": row[7],
        "desc": row[8], "payment": row[9],
    })

# Read MAIS DESPESAS NO NEON
ws3 = wb_in["MAIS DESPESAS NO NEON"]
mais_neon_rows = list(ws3.iter_rows(values_only=True))
mais_neon = []
for row in mais_neon_rows[1:]:
    if row[0] is None:
        continue
    mais_neon.append({
        "report_id": row[0], "name": row[1], "user": row[2], "cpf": row[3],
        "status": row[4], "ref_count": row[5], "neon_count": row[6],
        "diff_count": row[7], "ref_total": float(row[8]) if row[8] else 0,
        "neon_total": float(row[9]) if row[9] else 0, "diff_value": float(row[10]) if row[10] else 0,
    })

# Read MAIS DESPESAS NO REF
ws4 = wb_in["MAIS DESPESAS NO REF"]
mais_ref_rows = list(ws4.iter_rows(values_only=True))
mais_ref = []
for row in mais_ref_rows[1:]:
    if row[0] is None:
        continue
    mais_ref.append({
        "report_id": row[0], "name": row[1], "cpf": row[3], "status": row[4],
        "ref_count": row[5], "neon_count": row[6], "diff_count": row[7],
        "ref_total": float(row[8]) if row[8] else 0, "neon_total": float(row[9]) if row[9] else 0,
        "diff_value": float(row[10]) if row[10] else 0,
    })

# Read VALOR DIFERENTE
ws5 = wb_in["VALOR DIFERENTE"]
vd_rows = list(ws5.iter_rows(values_only=True))
valor_dif = []
for row in vd_rows[1:]:
    if row[0] is None:
        continue
    valor_dif.append({
        "exp_id": row[0], "report_id": row[1], "name": row[2],
        "cpf": row[4], "ref_value": float(row[5]) if row[5] else 0,
        "neon_value": float(row[6]) if row[6] else 0, "diff": float(row[7]) if row[7] else 0,
    })

wb_in.close()

# Query Neon DB for report details
conn = psycopg2.connect(os.getenv("NEON_DATABASE_URL"), connect_timeout=10)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# Get all report raw_data
all_rids = set(r["report_id"] for r in reports_only_neon)
all_rids.update(r["report_id"] for r in mais_neon)
all_rids.update(r["report_id"] for r in mais_ref)
all_rids.update(r["report_id"] for r in valor_dif)

report_db = {}
for rid in all_rids:
    cur.execute("SELECT id, name, status, raw_data FROM prestacao_reports WHERE id = %s", (rid,))
    row = cur.fetchone()
    if row:
        raw = row["raw_data"]
        if isinstance(raw, str):
            raw = json.loads(raw)
        report_db[rid] = {
            "name": row["name"], "status": row["status"],
            "approval_date": raw.get("approval_date") if raw else None,
            "updated_at": raw.get("updated_at") if raw else None,
            "justification": raw.get("justification", "") if raw else "",
        }

conn.close()

def parse_dt(s):
    if not s:
        return None
    try:
        return datetime.strptime(str(s)[:19], "%Y-%m-%d %H:%M:%S")
    except:
        return None

def categorize_report(r, db_info):
    """Determine why a report is in Neon but not in ref."""
    status = r["status"]
    name = r["name"] or ""
    approval_date = parse_dt(db_info.get("approval_date")) if db_info else None
    updated_at = parse_dt(db_info.get("updated_at")) if db_info else None
    justification = (db_info.get("justification") or "") if db_info else ""
    
    reasons = []
    
    # Check naming convention
    is_standard_caixa = bool(re.match(r'^CAIXA\s+\d{2}/\d{4}$', name, re.IGNORECASE))
    
    # Check for Itaú/FATURA clues in justification
    just_upper = justification.upper()
    has_fatura_in_just = "FATURA" in just_upper or "ITAU" in just_upper or "ITAÚ" in just_upper
    
    if status == "ENVIADO":
        # Check timing
        if updated_at and updated_at > REF_DATE:
            reasons.append("TIMING: Was REPROVADO at ref time (~8am 27/07), reaberto/enviado after")
        else:
            reasons.append("GENUINE_ENVIADO: Not approved, correctly excluded from ref")
        
        # Check if was reprovado then reaberto
        if justification and ("REPROV" in just_upper or "recibo" in just_upper.lower()):
            reasons.append(f"Last action: {justification[:80]}")
    
    elif status == "APROVADO":
        if not is_standard_caixa:
            reasons.append(f"NON_STANDARD_NAME: Report named '{name}' instead of 'CAIXA XX/YYYY'")
        
        if has_fatura_in_just:
            reasons.append("ITAU_CARD: Approver asked to rename as FATURA (Itaú card)")
        
        if approval_date and approval_date.year == 2025:
            reasons.append(f"OLD_2025: Approved on {approval_date.strftime('%d/%m/%Y')} (old report)")
        
        if approval_date and approval_date > REF_DATE:
            reasons.append("APPROVED_AFTER_REF: Approved after ref was made")
        
        if justification and "REPROV" in just_upper:
            reasons.append(f"REPROVED_THEN_APPROVED: {justification[:80]}")
        
        if not reasons:
            reasons.append("UNCATEGORIZED: APROVADO, standard name, approved before ref - needs investigation")
    
    return " | ".join(reasons) if reasons else "UNKNOWN"

# Build output Excel
wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

def style_header(ws, ncols):
    for col in range(1, ncols+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def auto_width(ws, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value))
                if val_len > max_len:
                    max_len = val_len
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

# Sheet 1: RESUMO
ws1 = wb.active
ws1.title = "RESUMO"

# Calculate totals
reports_only_neon_total = sum(r["total"] for r in reports_only_neon)
mais_neon_diff = sum(r["diff_value"] for r in mais_neon)
mais_ref_diff = sum(r["diff_value"] for r in mais_ref)
valor_dif_total = sum(abs(r["diff"]) for r in valor_dif)

summary_data = [
    ["Categoria", "Quantidade", "Valor (R$)", "Descrição"],
    ["Reports só no Neon", len(reports_only_neon), reports_only_neon_total, "Reports inteiros que não estão no BASE PREST"],
    ["Despesas só no Neon (de reports só no Neon)", len([e for e in so_no_neon_exp if e["report_id"] in set(r["report_id"] for r in reports_only_neon)]), sum(e["value"] for e in so_no_neon_exp if e["report_id"] in set(r["report_id"] for r in reports_only_neon)), "Despesas de reports que não estão no ref"],
    ["Despesas só no Neon (de reports em ambos)", len([e for e in so_no_neon_exp if e["report_id"] not in set(r["report_id"] for r in reports_only_neon)]), sum(e["value"] for e in so_no_neon_exp if e["report_id"] not in set(r["report_id"] for r in reports_only_neon)), "Despesas extras em reports que estão no ref (ver aba MAIS DESPESAS)"],
    ["Mais despesas no Neon", len(mais_neon), mais_neon_diff, "Reports com mais despesas no Neon que no ref"],
    ["Mais despesas no Ref", len(mais_ref), mais_ref_diff, "Reports com mais despesas no ref que no Neon"],
    ["Valor diferente", len(valor_dif), valor_dif_total, "Mesma despesa, valor diferente"],
    ["", "", "", ""],
    ["DIVERGÊNCIA TOTAL (Neon > Ref)", "", reports_only_neon_total + mais_neon_diff + valor_dif_total, "Soma de tudo que está no Neon mas não no ref"],
    ["DIVERGÊNCIA INVERSA (Ref > Neon)", "", mais_ref_diff, "Soma de tudo que está no ref mas não no Neon"],
    ["DIVERGÊNCIA LÍQUIDA", "", reports_only_neon_total + mais_neon_diff + valor_dif_total - mais_ref_diff, "Neon - Ref"],
]

for row_data in summary_data:
    ws1.append(row_data)
style_header(ws1, 4)
auto_width(ws1)

# Sheet 2: REPORTS SO NO NEON - with reasons
ws2 = wb.create_sheet("REPORTS SO NO NEON")
headers = ["Report ID", "Name", "User", "CPF", "Status", "N Expenses", "Total (R$)", "Approval Date", "Updated At", "Justification", "REASON"]
ws2.append(headers)

for r in sorted(reports_only_neon, key=lambda x: -x["total"]):
    db = report_db.get(r["report_id"], {})
    reason = categorize_report(r, db)
    apd = db.get("approval_date", "") or ""
    upd = db.get("updated_at", "") or ""
    just = (db.get("justification") or "")[:200]
    ws2.append([r["report_id"], r["name"], r["user"], r["cpf"], r["status"],
                r["n_expenses"], r["total"], apd, upd, just, reason])
style_header(ws2, len(headers))
auto_width(ws2, 60)

# Sheet 3: MAIS DESPESAS NO NEON - with reasons
ws3 = wb.create_sheet("MAIS DESPESAS NO NEON")
headers = ["Report ID", "Name", "User", "CPF", "Status", "Ref Count", "Neon Count", "Diff Count", "Ref Total", "Neon Total", "Diff Value", "Approval Date", "REASON"]
ws3.append(headers)

for r in sorted(mais_neon, key=lambda x: -x["diff_value"]):
    db = report_db.get(r["report_id"], {})
    apd = db.get("approval_date", "") or ""
    # Simple reason for mais despesas
    status = r["status"]
    if status == "ENVIADO":
        reason = "ENVIADO: Expenses added after ref snapshot (report still pending)"
    elif status == "APROVADO":
        upd = parse_dt(db.get("updated_at"))
        if upd and upd > REF_DATE:
            reason = "TIMING: Expenses added/modified after ref snapshot"
        else:
            reason = "EXPENSES_ADDED: Neon has more expenses than ref snapshot (possible re-submission)"
    else:
        reason = "UNKNOWN"
    ws3.append([r["report_id"], r["name"], r["user"], r["cpf"], r["status"],
                r["ref_count"], r["neon_count"], r["diff_count"], r["ref_total"],
                r["neon_total"], r["diff_value"], apd, reason])
style_header(ws3, len(headers))
auto_width(ws3, 60)

# Sheet 4: MAIS DESPESAS NO REF
ws4 = wb.create_sheet("MAIS DESPESAS NO REF")
headers = ["Report ID", "Name", "CPF", "Status", "Ref Count", "Neon Count", "Diff Count", "Ref Total", "Neon Total", "Diff Value", "REASON"]
ws4.append(headers)

for r in sorted(mais_ref, key=lambda x: -x["diff_value"]):
    if r["status"] == "ENVIADO":
        reason = "ENVIADO: Expenses removed from report after ref was made (reprovado + resend with fewer expenses)"
    else:
        reason = "EXPENSES_REMOVED: Ref had more expenses than current Neon (expenses were removed)"
    ws4.append([r["report_id"], r["name"], r["cpf"], r["status"],
                r["ref_count"], r["neon_count"], r["diff_count"], r["ref_total"],
                r["neon_total"], r["diff_value"], reason])
style_header(ws4, len(headers))
auto_width(ws4, 60)

# Sheet 5: VALOR DIFERENTE
ws5 = wb.create_sheet("VALOR DIFERENTE")
headers = ["Expense ID", "Report ID", "Name", "CPF", "Ref Value", "Neon Value", "Diff", "REASON"]
ws5.append(headers)

for r in valor_dif:
    reason = "VALUE_CHANGED: Expense value was modified after ref was made"
    ws5.append([r["exp_id"], r["report_id"], r["name"], r["cpf"],
                r["ref_value"], r["neon_value"], r["diff"], reason])
style_header(ws5, len(headers))
auto_width(ws5, 60)

# Sheet 6: CATEGORY BREAKDOWN
ws6 = wb.create_sheet("CATEGORY BREAKDOWN")
ws6.append(["Category", "Count", "Total (R$)", "Explanation"])

cat_totals = Counter()
cat_values = Counter()
for r in reports_only_neon:
    db = report_db.get(r["report_id"], {})
    reason = categorize_report(r, db)
    # Extract first category
    cat = reason.split(":")[0].strip() if ":" in reason else reason.split("|")[0].strip()
    cat_totals[cat] += 1
    cat_values[cat] += r["total"]

cat_explanations = {
    "TIMING": "Report was REPROVADO when ref was made (~8am 27/07), but was REABERTO+ENVIADO after. Neon exported later, so shows ENVIADO.",
    "GENUINE_ENVIADO": "Report is genuinely ENVIADO (not approved). Correctly excluded from ref.",
    "NON_STANDARD_NAME": "APROVADO report with non-standard name (not 'CAIXA XX/YYYY'). May have been filtered out by financeiro.",
    "ITAU_CARD": "Report uses Cartão Itaú (credit card), not VExpenses. Approver asked to rename as FATURA. Not a prestação de contas.",
    "OLD_2025": "Report from 2025, approved months ago. May have been already processed in a previous ref period.",
    "REPROVED_THEN_APPROVED": "Report was reprovado, then re-approved. May have been excluded from ref during reproval period.",
    "APPROVED_AFTER_REF": "Report was approved AFTER the ref was made. Not in ref because it wasn't approved yet.",
    "UNCATEGORIZED": "APROVADO, standard name, approved before ref. Needs manual investigation.",
}

for cat in sorted(cat_totals.keys(), key=lambda x: -cat_values[x]):
    ws6.append([cat, cat_totals[cat], cat_values[cat], cat_explanations.get(cat, "")])
style_header(ws6, 4)
auto_width(ws6, 80)

# Sheet 7: SO NO NEON EXPENSES (all)
ws7 = wb.create_sheet("SO NO NEON EXPENSES")
headers = ["Expense ID", "Report ID", "Report Name", "User", "CPF", "Status", "Value", "Date", "Description", "Payment"]
ws7.append(headers)
for e in sorted(so_no_neon_exp, key=lambda x: -x["value"]):
    ws7.append([e["exp_id"], e["report_id"], e["report_name"], e["user"], e["cpf"],
                e["status"], e["value"], e["date"], e["desc"], e["payment"]])
style_header(ws7, len(headers))
auto_width(ws7, 50)

# Save
output_path = Path(r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\controle-api\data\gap_analysis_with_reasons.xlsx")
wb.save(output_path)
print(f"Saved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
