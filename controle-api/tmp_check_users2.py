#!/usr/bin/env python3
"""
Check:
1. Are the 20 "new" users in the reference EXTRATO?
2. Are they in the quinzena_cadastro table?
3. For PATRICK, ANA LUIZA, SANDRA — what report IDs does the ref BASE PREST have for them?
4. Compare those ref report IDs with API report IDs for same user
"""
import os
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from pathlib import Path
import openpyxl
from collections import defaultdict

BASE = Path(__file__).parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"

users_no_ref = [
    "AFONSO FIORELLO CARVALHO",
    "AILTON MENDES AGUILAR",
    "AMARAL RODRIGUES NUNES",
    "ANA LUIZA OLIVEIRA CAMELO",
    "ATILA SILVA DOS SANTOS",
    "DANIEL PORFIRIO DE SOUSA",
    "ERASMO PEREIRA MONTEIRO",
    "FLAVIO PEREIRA DE SOUZA",
    "FRANCISCO MICHAEL CASTRO FERNANDES",
    "JACKSON CAROLINO CARNEIRO",
    "LEONARDO GONCALVES RIBEIRO FILHO",
    "MARCELO BEZERRA",
    "NELSON MOURA RIBEIRO",
    "OSMAN DOS SANTOS MORAIS JUNIOR",
    "PATRICK FERNANDO GOULART ALVES",
    "PEDRO LUIS PIRES DOS SANTOS",
    "ROBERTO ALIAGA",
    "SANDRA CRISTINA ALVES MACHADO",
    "WAGNER FERNANDES DA SILVA",
    "WESLEY CARLOS AUGUSTO",
]

def norm(s):
    return " ".join(str(s or "").upper().strip().split())

users_norm = {norm(u): u for u in users_no_ref}

# ============================================================
# 1. Check reference EXTRATO for these users
# ============================================================
print("=" * 80)
print("1. Are these 20 users in the reference EXTRATO?")
print("=" * 80)

wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_ext = wb["EXTRATO"]
# Col 8 = Usuario, col 12 = CPF
ext_users = defaultdict(lambda: {"count": 0, "cpfs": set()})
for row in ws_ext.iter_rows(min_row=9, values_only=True):
    if row[8] is None:
        continue
    usuario = str(row[8]).strip()
    cpf = str(row[12] or "").strip() if len(row) > 12 else ""
    user_norm = norm(usuario)
    for un, orig in users_norm.items():
        if un in user_norm or user_norm in un:
            ext_users[orig]["count"] += 1
            ext_users[orig]["cpfs"].add(cpf)
            break

for user in users_no_ref:
    if user in ext_users:
        info = ext_users[user]
        print(f"  FOUND  {user[:35]:<35}  {info['count']:>4} txns  CPFs={info['cpfs']}")
    else:
        print(f"  MISSING {user[:35]:<35}")

# ============================================================
# 2. Check quinzena_cadastro in DB
# ============================================================
print("\n" + "=" * 80)
print("2. Are these users in quinzena_cadastro?")
print("=" * 80)

conn = psycopg2.connect(NEON_URL, connect_timeout=10)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cur.execute("""
    SELECT DISTINCT colaborador, cpf, regional
    FROM quinzena_cadastro
""")
cadastro_rows = cur.fetchall()

for user in users_no_ref:
    user_norm = norm(user)
    found = [c for c in cadastro_rows if user_norm in norm(c["colaborador"]) or norm(c["colaborador"]) in user_norm]
    if found:
        for f in found:
            print(f"  FOUND  {user[:35]:<35}  CPF={f['cpf']}  regional={f['regional']}")
    else:
        print(f"  MISSING {user[:35]:<35}")

# ============================================================
# 3. For PATRICK, ANA LUIZA, SANDRA — compare ref vs API report IDs
# ============================================================
print("\n" + "=" * 80)
print("3. Report ID comparison for users found in BASE PREST")
print("=" * 80)

ws_bp = wb["BASE PREST "]
# Col 1 = report ID, Col 2 = report name, Col 4 = user name, Col 9 = CPF, Col 26 = Valor
check_users = ["PATRICK FERNANDO GOULART ALVES", "ANA LUIZA OLIVEIRA CAMELO", "SANDRA CRISTINA ALVES MACHADO"]

for check_user in check_users:
    check_norm = norm(check_user)
    ref_reports = defaultdict(lambda: {"name": "", "count": 0, "total": 0})
    for row in ws_bp.iter_rows(min_row=4, values_only=True):
        if row[0] is None:
            continue
        user_name = str(row[4] or "").strip()
        if check_norm in norm(user_name) or norm(user_name) in check_norm:
            rid = int(row[1]) if row[1] else 0
            ref_reports[rid]["name"] = str(row[2] or "")
            ref_reports[rid]["count"] += 1
            ref_reports[rid]["total"] += float(row[26] or 0) if len(row) > 26 else 0
    
    # Get API reports for same user
    cur.execute("""
        SELECT r.id, r.name, r.status, COUNT(e.id) as count, SUM(e.value) as total
        FROM prestacao_reports r
        JOIN prestacao_expenses e ON e.report_id = r.id
        WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
          AND r.user_cpf IS NOT NULL
          AND TRIM(r.name) !~* '^(fatu|farur|cart)'
          AND TRIM(r.name) !~* '(fatura|fatuar|fatut|farur)'
          AND UPPER(r.user_name) LIKE %s
        GROUP BY r.id, r.name, r.status
        ORDER BY r.name
    """, (f"%{check_user.split()[0]}%{check_user.split()[1]}%",))
    api_reports = cur.fetchall()
    
    print(f"\n  {check_user}:")
    print(f"    Reference BASE PREST: {len(ref_reports)} reports")
    for rid in sorted(ref_reports.keys()):
        info = ref_reports[rid]
        in_api = "✓" if any(int(a["id"]) == rid for a in api_reports) else "✗"
        print(f"      {rid}  {info['name'][:35]:<35}  {info['count']:>3} items  R$ {info['total']:>10,.2f}  in_api={in_api}")
    
    print(f"    API: {len(api_reports)} reports")
    for a in api_reports:
        in_ref = "✓" if int(a["id"]) in ref_reports else "✗"
        print(f"      {a['id']}  {str(a['name'])[:35]:<35}  {a['status']:<10}  {a['count']:>3} items  R$ {float(a['total']):>10,.2f}  in_ref={in_ref}")

# ============================================================
# 4. Check: are the 17 truly missing users in the API cadastro?
# ============================================================
print("\n" + "=" * 80)
print("4. Check all cadastros for the 17 missing users")
print("=" * 80)

cur.execute("SELECT DISTINCT colaborador, cpf FROM quinzena_cadastro")
all_cad = cur.fetchall()

truly_missing = [u for u in users_no_ref if u not in ["PATRICK FERNANDO GOULART ALVES", "ANA LUIZA OLIVEIRA CAMELO", "SANDRA CRISTINA ALVES MACHADO"]]
for user in truly_missing:
    user_norm = norm(user)
    found = [c for c in all_cad if user_norm in norm(c["colaborador"]) or norm(c["colaborador"]) in user_norm]
    if found:
        for f in found:
            print(f"  FOUND  {user[:35]:<35}  CPF={f['cpf']}")
    else:
        print(f"  MISSING {user[:35]:<35}")

wb.close()
conn.close()
