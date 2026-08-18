#!/usr/bin/env python3
"""Focused analysis of the 54 reports and simulation results."""
import os, re
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from pathlib import Path
import openpyxl
from collections import defaultdict

BASE = Path(__file__).parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"

# Load reference
wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_bp = wb["BASE PREST "]
ref_report_ids = set()
ref_report_info = {}
ref_total = 0
for row in ws_bp.iter_rows(min_row=4, values_only=True):
    if row[0] is None: continue
    eid = int(row[0])
    rid = int(row[1]) if row[1] else None
    rname = str(row[2] or "")
    user_name = str(row[4] or "")
    user_cpf = str(row[9] or "").strip()
    valor = float(row[26] or 0) if len(row) > 26 else 0
    ref_total += valor
    if rid:
        ref_report_ids.add(rid)
        if rid not in ref_report_info:
            ref_report_info[rid] = {"name": rname, "user": user_name, "cpf": user_cpf, "count": 0, "total": 0}
        ref_report_info[rid]["count"] += 1
        ref_report_info[rid]["total"] += valor

ws_p = wb["PAINEL"]
vexpenses_cpfs = set()
painel_info = {}
for row in ws_p.iter_rows(min_row=12, values_only=True):
    if row[2] is None: continue
    cpf = str(row[2] or "").strip()
    colab = str(row[1] or "").strip()
    situacao = str(row[4] or "").strip().upper()
    status_cartao = str(row[5] or "").strip()
    cartao_vexp = str(row[12] or "").strip().upper()
    painel_info[cpf] = {"colaborador": colab, "situacao": situacao, "status_cartao": status_cartao, "cartao_vexp": cartao_vexp}
    if cartao_vexp == "SIM": vexpenses_cpfs.add(cpf)
wb.close()

# Load API
conn = psycopg2.connect(NEON_URL, connect_timeout=10)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cur.execute("""
    SELECT e.id, e.report_id, e.value, r.name as report_name, r.status as report_status,
           r.user_cpf, r.user_name, r.created_at, r.updated_at
    FROM prestacao_expenses e
    JOIN prestacao_reports r ON e.report_id = r.id
    WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado') AND r.user_cpf IS NOT NULL
""")
all_api = cur.fetchall()
cur.execute("SELECT DISTINCT id, name, status, user_cpf, user_name, created_at, updated_at, total_value FROM prestacao_reports WHERE (status ILIKE 'Aprovado' OR status ILIKE 'Enviado') AND user_cpf IS NOT NULL")
api_reports_raw = {r["id"]: r for r in cur.fetchall()}
conn.close()

def apply_base_filters(expenses):
    result = []
    for e in expenses:
        name = (e["report_name"] or "").strip()
        if re.match(r'^(fatu|farur|cart)', name, re.IGNORECASE): continue
        if re.search(r'(fatura|fatuar|fatut|farur|fartur)', name, re.IGNORECASE): continue
        if re.search(r'(cartão corporativo|cartao corporativo)', name, re.IGNORECASE): continue
        if re.search(r'itau|itaú', name, re.IGNORECASE): continue
        if e["user_cpf"] not in vexpenses_cpfs: continue
        result.append(e)
    return result

filtered = apply_base_filters(all_api)
api_report_ids = set()
api_report_data = {}
for e in filtered:
    rid = e["report_id"]
    api_report_ids.add(rid)
    if rid not in api_report_data:
        r = api_reports_raw.get(rid, {})
        api_report_data[rid] = {
            "name": e["report_name"], "user": e["user_name"], "cpf": e["user_cpf"],
            "status": e["report_status"], "created_at": r.get("created_at"),
            "expense_count": 0, "expense_total": 0,
        }
    api_report_data[rid]["expense_count"] += 1
    api_report_data[rid]["expense_total"] += float(e["value"])

new_reports = api_report_ids - ref_report_ids

# 1. User-level breakdown
print("=" * 110)
print("54 REPORTS IN API BUT NOT IN REFERENCE — User breakdown")
print("=" * 110)

new_by_user = defaultdict(lambda: {"new": [], "ref": [], "new_total": 0, "ref_total": 0})
for rid in new_reports:
    r = api_report_data[rid]
    cpf = r["cpf"]
    new_by_user[cpf]["new"].append(rid)
    new_by_user[cpf]["new_total"] += r["expense_total"]
for rid, r in ref_report_info.items():
    cpf = r["cpf"]
    if cpf in new_by_user:
        new_by_user[cpf]["ref"].append(rid)
        new_by_user[cpf]["ref_total"] += r["total"]

print(f"\n  {'User':<30} {'CPF':<15} {'New':>4} {'Ref':>4} {'New R$':>12} {'Ref R$':>12}  Situacao/Card")
print("  " + "-" * 105)
for cpf in sorted(new_by_user.keys(), key=lambda c: -new_by_user[c]["new_total"]):
    u = new_by_user[cpf]
    pinfo = painel_info.get(cpf, {})
    sit = pinfo.get("situacao", "?")
    card = pinfo.get("status_cartao", "?")
    name = api_report_data[u["new"][0]]["user"] if u["new"] else "?"
    has_ref = "HAS REF" if u["ref"] else "NO REF"
    print(f"  {str(name)[:30]:<30} {cpf:<15} {len(u['new']):>4} {len(u['ref']):>4} {u['new_total']:>12,.2f} {u['ref_total']:>12,.2f}  {sit}/{card}  {has_ref}")

# 2. The 3 name+cpf matches
print("\n" + "=" * 110)
print("NAME+CPF MATCHES (same report name+CPF in ref under different report ID)")
print("=" * 110)
ref_by_name_cpf = defaultdict(list)
for rid, r in ref_report_info.items():
    ref_by_name_cpf[(r["name"].strip().upper(), r["cpf"])].append(rid)

for rid in sorted(new_reports):
    r = api_report_data[rid]
    key = (r["name"].strip().upper(), r["cpf"])
    if key in ref_by_name_cpf:
        ref_rids = ref_by_name_cpf[key]
        match_total = sum(ref_report_info[rr]["total"] for rr in ref_rids)
        print(f"  API rid={rid} '{r['name']}' user={r['user']} → ref rids={ref_rids}  API R$ {r['expense_total']:,.2f} vs Ref R$ {match_total:,.2f}")

# 3. Key question: users with NO ref reports at all
no_ref_users = [(cpf, u) for cpf, u in new_by_user.items() if not u["ref"]]
print(f"\n  Users with NO reports in reference at all: {len(no_ref_users)}")
for cpf, u in sorted(no_ref_users, key=lambda x: -x[1]["new_total"]):
    pinfo = painel_info.get(cpf, {})
    name = api_report_data[u["new"][0]]["user"]
    print(f"    {str(name)[:30]:<30}  CPF={cpf}  {len(u['new'])} reports  R$ {u['new_total']:>10,.2f}  sit={pinfo.get('situacao','?')}  card={pinfo.get('status_cartao','?')}")

# 4. Simulation summary
print("\n" + "=" * 110)
print("SIMULATION RESULTS")
print("=" * 110)

def calc(expenses):
    api_rids = set(e["report_id"] for e in expenses)
    api_total = sum(float(e["value"]) for e in expenses)
    new_r = len(api_rids - ref_report_ids)
    missing_r = len(ref_report_ids - api_rids)
    gap = api_total - ref_total
    pct = abs(gap) / ref_total * 100
    return len(api_rids), api_total, new_r, missing_r, gap, pct

base = filtered
aprovado_only = [e for e in base if e["report_status"].upper() == "APROVADO"]
enviado_only = [e for e in base if e["report_status"].upper() == "ENVIADO"]

# How many ENVIADO are IN REF vs NOT IN REF
enviado_in_ref = sum(1 for rid in set(e["report_id"] for e in enviado_only) if rid in ref_report_ids)
enviado_not_ref = sum(1 for rid in set(e["report_id"] for e in enviado_only) if rid not in ref_report_ids)
enviado_in_ref_val = sum(float(e["value"]) for e in enviado_only if e["report_id"] in ref_report_ids)
enviado_not_ref_val = sum(float(e["value"]) for e in enviado_only if e["report_id"] not in ref_report_ids)

print(f"\n  Reference total: R$ {ref_total:,.2f}")
print(f"\n  {'Scenario':<45} {'Reports':>8} {'Total R$':>14} {'New':>5} {'Miss':>5} {'Gap R$':>14} {'% Gap':>7}")
print("  " + "-" * 100)
for name, expenses in [
    ("Base (current proposed filters)", base),
    ("APROVADO only (no ENVIADO)", aprovado_only),
]:
    n, t, nn, nm, g, p = calc(expenses)
    print(f"  {name:<45} {n:>8} {t:>14,.2f} {nn:>5} {nm:>5} {g:>+14,.2f} {p:>6.1f}%")

print(f"\n  ENVIADO reports: {enviado_in_ref} IN REF (R$ {enviado_in_ref_val:,.2f}), {enviado_not_ref} NOT IN REF (R$ {enviado_not_ref_val:,.2f})")
print(f"  → Removing ENVIADO would lose {enviado_in_ref} reports (R$ {enviado_in_ref_val:,.2f}) that ARE in the reference")
print(f"  → This confirms: we CANNOT remove ENVIADO reports")

# 5. What's the actual gap breakdown?
print("\n" + "=" * 110)
print("GAP BREAKDOWN (Base filters)")
print("=" * 110)

new_total = sum(api_report_data[rid]["expense_total"] for rid in new_reports)
missing_reports = ref_report_ids - api_report_ids
missing_total = sum(ref_report_info[rid]["total"] for rid in missing_reports)
common_reports = api_report_ids & ref_report_ids

# Common report diffs
common_api_total = sum(sum(float(e["value"]) for e in base if e["report_id"] == rid) for rid in common_reports)
common_ref_total = sum(ref_report_info[rid]["total"] for rid in common_reports)
common_diff = common_api_total - common_ref_total

print(f"\n  New reports (API only):     {len(new_reports):>5} reports  R$ {new_total:>+12,.2f}")
print(f"  Missing reports (Ref only): {len(missing_reports):>5} reports  R$ {-missing_total:>+12,.2f}")
print(f"  Common reports value diff:  {len(common_reports):>5} reports  R$ {common_diff:>+12,.2f}")
print(f"  ────────────────────────────────────────────────────────────")
print(f"  Net gap:                                      R$ {new_total - missing_total + common_diff:>+12,.2f}")
print(f"  Reference total:                              R$ {ref_total:>12,.2f}")
print(f"  Gap %:                                        {abs(new_total - missing_total + common_diff) / ref_total * 100:>11.1f}%")

# 6. Break down the 54 new reports by whether user has ref reports
print("\n" + "=" * 110)
print("54 NEW REPORTS: Users WITH ref reports vs WITHOUT")
print("=" * 110)

with_ref = [(cpf, u) for cpf, u in new_by_user.items() if u["ref"]]
without_ref = [(cpf, u) for cpf, u in new_by_user.items() if not u["ref"]]

with_ref_reports = sum(len(u["new"]) for _, u in with_ref)
without_ref_reports = sum(len(u["new"]) for _, u in without_ref)
with_ref_total = sum(u["new_total"] for _, u in with_ref)
without_ref_total = sum(u["new_total"] for _, u in without_ref)

print(f"\n  Users WITH other ref reports:  {len(with_ref):>3} users  {with_ref_reports:>3} reports  R$ {with_ref_total:>12,.2f}")
print(f"  Users WITHOUT any ref reports: {len(without_ref):>3} users  {without_ref_reports:>3} reports  R$ {without_ref_total:>12,.2f}")

print(f"\n  --- Users WITH ref reports (these users have SOME reports in ref, but not these specific ones) ---")
for cpf, u in sorted(with_ref, key=lambda x: -x[1]["new_total"]):
    pinfo = painel_info.get(cpf, {})
    name = api_report_data[u["new"][0]]["user"]
    print(f"    {str(name)[:30]:<30}  new={len(u['new'])}  ref={len(u['ref'])}  new R$ {u['new_total']:>10,.2f}  ref R$ {u['ref_total']:>10,.2f}  {pinfo.get('situacao','?')}/{pinfo.get('status_cartao','?')}")
    for rid in u["new"]:
        r = api_report_data[rid]
        created = r["created_at"].strftime("%Y-%m-%d") if r["created_at"] else "?"
        print(f"      → NEW: {rid} '{r['name']}'  {r['status']}  R$ {r['expense_total']:>10,.2f}  created={created}")
