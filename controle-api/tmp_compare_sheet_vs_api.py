#!/usr/bin/env python3
"""Compare financeiro CARGA 2 QZ JULHO sheet vs our Neon DB data."""
import os, sys, json, unicodedata, re
from pathlib import Path
import openpyxl
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env")

SHEET_PATH = Path(r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CARGA 2 QZ JULHO 26 VEXPENSES EQS.xlsx")

def clean_cpf(v):
    if v is None: return ''
    return re.sub(r'\D', '', str(v)).zfill(11)

def to_float(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace('.', '').replace(',', '.')
    try: return float(s)
    except: return 0.0

# ============================================================
# 1. Read financeiro Excel (header at row 6, data from row 7)
# ============================================================
print("=" * 100)
print("1. READING FINANCEIRO EXCEL")
print("=" * 100)

wb = openpyxl.load_workbook(SHEET_PATH, data_only=True)
ws = wb['2 QZ JULHO']

# Col mapping: 1=COLABORADOR, 2=CPF, 3=SITUAÇÃO, 4=REGIONAL, 5=CENTRO DE CUSTO,
# 6=GESTOR, 7=DIRETOR, 8=SALDO REEMBOLSAR, 9=SALDO FINAL, 10=2ª QZ,
# 11=SALDO CARTAO, 12=Adiantamento, 13=CARGA PARCIAL, 14=REEMBOLSO,
# 15=Carga Final, 16=obs, 17=STATUS DO CARTÃO

sheet_data = {}
for row_idx in range(7, ws.max_row + 1):
    nome = ws.cell(row=row_idx, column=1).value
    cpf_raw = ws.cell(row=row_idx, column=2).value
    cpf = clean_cpf(cpf_raw)
    if not cpf and not nome: continue
    if nome and 'TOTAL' in str(nome).upper(): break
    sheet_data[cpf] = {
        'colaborador': str(nome).strip() if nome else '',
        'cpf': cpf,
        'saldo_reembolsar': to_float(ws.cell(row=row_idx, column=8).value),
        'saldo_final': to_float(ws.cell(row=row_idx, column=9).value),
        'col_2qz': to_float(ws.cell(row=row_idx, column=10).value),
        'saldo_cartao': to_float(ws.cell(row=row_idx, column=11).value),
        'adiantamento': to_float(ws.cell(row=row_idx, column=12).value),
        'carga_parcial': to_float(ws.cell(row=row_idx, column=13).value),
        'reembolso': to_float(ws.cell(row=row_idx, column=14).value),
        'carga_final': to_float(ws.cell(row=row_idx, column=15).value),
    }
print(f"  Sheet rows: {len(sheet_data)}")
wb.close()

# ============================================================
# 2. Fetch data from Neon DB (same logic as quinzena-complete)
# ============================================================
print("\n" + "=" * 100)
print("2. FETCHING DATA FROM NEON DB")
print("=" * 100)

db_url = os.getenv("NEON_DATABASE_URL")
conn = psycopg2.connect(db_url)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# 2a. Load cadastro
cur.execute("""
    SELECT cpf, colaborador, situacao, status_cartao,
           regional, centro_custo, gestor, diretor
    FROM quinzena_cadastro
    ORDER BY colaborador ASC NULLS LAST
""")
cadastro = cur.fetchall()
print(f"  Cadastro: {len(cadastro)} colaboradores")

# 2b. Load manual inputs
cur.execute("""
    SELECT col_1qz::text, adiantamento::text, obs, cpf
    FROM quinzena_manual_inputs
    WHERE year = 2026 AND month = 7 AND quinzena = 2
""")
manuals = {r['cpf']: r for r in cur.fetchall() if r['cpf']}
print(f"  Manual inputs: {len(manuals)}")

# 2c. Load reembolso multiplier
cur.execute("""
    SELECT reembolso_multiplier::text
    FROM quinzena_config
    WHERE year = 2026 AND month = 7 AND quinzena = 2
""")
config_row = cur.fetchone()
reembolso_multiplier = float(config_row['reembolso_multiplier']) if config_row else 0.5
print(f"  Reembolso multiplier: {reembolso_multiplier}")

# 2d. Name normalization for extrato matching
def normalize_name(s):
    if not s: return ''
    s = str(s).upper().strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'\s+', ' ', s)
    return s

def fuzzy_ratio(a, b):
    if a == b: return 1.0
    if not a or not b: return 0.0
    bigrams_a = set(a[i:i+2] for i in range(len(a) - 1))
    bigrams_b = set(b[i:i+2] for i in range(len(b) - 1))
    if not bigrams_a or not bigrams_b: return 0.0
    inter = len(bigrams_a & bigrams_b)
    return (2 * inter) / (len(bigrams_a) + len(bigrams_b))

name_to_cpf = {}
for c in cadastro:
    norm = normalize_name(c['colaborador'] or '')
    if norm:
        name_to_cpf[norm] = c['cpf']

fuzzy_cache = {}
def resolve_cpf_by_name(extrato_name):
    normalized = normalize_name(extrato_name)
    exact = name_to_cpf.get(normalized)
    if exact: return exact
    cached = fuzzy_cache.get(normalized)
    if cached: return cached
    best_cpf = None
    best_ratio = 0.0
    for cad_name, cpf in name_to_cpf.items():
        ratio = fuzzy_ratio(normalized, cad_name)
        if ratio > best_ratio:
            best_ratio = ratio
            best_cpf = cpf
    if best_ratio >= 0.88 and best_cpf:
        fuzzy_cache[normalized] = best_cpf
        return best_cpf
    if len(normalized) >= 10:
        prefix15 = normalized[:15]
        for cad_name, cpf in name_to_cpf.items():
            if cad_name[:15] == prefix15: return cpf
        prefix10 = normalized[:10]
        for cad_name, cpf in name_to_cpf.items():
            if cad_name[:10] == prefix10: return cpf
    return None

# 2e. Extrato cumulativo (cutoff = 2026-07-01 for 2QZ)
financial_cutoff = '2026-07-01'
saldo_cartao_carga_date = '2026-07-25'

cur.execute("""
    WITH deduped AS (
        SELECT DISTINCT ON (
            UPPER(usuario), data, tipo, valor,
            COALESCE(NULLIF(codigo_transacao, ''), hora::text)
        )
            UPPER(usuario) AS usuario_up,
            data, tipo, valor, codigo_transacao
        FROM extrato_movimentacao
        WHERE is_snapshot = FALSE
          AND data <= %s
        ORDER BY UPPER(usuario), data, tipo, valor,
            COALESCE(NULLIF(codigo_transacao, ''), hora::text)
    )
    SELECT
        usuario_up,
        COALESCE(SUM(valor) FILTER(WHERE tipo = 'Transferência' AND valor > 0), 0) AS carga_raw,
        COALESCE(SUM(valor) FILTER(WHERE tipo = 'Transferência' AND valor < 0), 0) AS transf_raw,
        COALESCE(SUM(valor) FILTER(WHERE tipo = 'Taxa'), 0) AS tarifa_raw
    FROM deduped
    GROUP BY usuario_up
""", (financial_cutoff,))
extrato_rows = cur.fetchall()

carga_by_cpf = {}
transf_by_cpf = {}
tarifa_by_cpf = {}
for r in extrato_rows:
    cpf = resolve_cpf_by_name(r['usuario_up'])
    if cpf:
        carga_by_cpf[cpf] = float(r['carga_raw'] or 0)
        transf_by_cpf[cpf] = abs(float(r['transf_raw'] or 0))
        tarifa_by_cpf[cpf] = abs(float(r['tarifa_raw'] or 0))

# 2f. Somase (prestação de contas) — same filters as pipeline
def is_fatura_or_cartao(name):
    n = (name or '').strip().upper()
    if not n: return False
    if 'CAIXA ITAU' in n or 'CAIXA ITAÚ' in n: return True
    if n.startswith('CAIXA'): return False
    if re.match(r'^(FATURA|CARTAO|CARTÃO|FATUAR|FARTUR|FATUT|FARUR|FATUTR)', n): return True
    if 'CARTÃO DE CRÉDITO' in n or 'CARTAO DE CREDITO' in n or 'CARTÃO DE CREDITO' in n: return True
    if 'CARTÃO CORPORATIVO' in n: return True
    if ('ITAU' in n or 'ITAÚ' in n) and 'CAIXA' not in n: return True
    if 'DOLAR' in n or 'DÓLAR' in n: return True
    if n.startswith('DESPESA') and 'FATURA' in n: return True
    if n.startswith('COMPLEMENTAR') and 'FATURA' in n: return True
    if 'CARTÃO' in n and 'CRÉDITO' in n: return True
    if 'CARTAO' in n and 'CREDITO' in n: return True
    if n.startswith('CARTÃO VEXPENSES'): return True
    return False

cur.execute("""
    SELECT r.id, r.name
    FROM prestacao_reports r
    WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND r.user_cpf IS NOT NULL
""")
all_reports = cur.fetchall()
valid_ids = [r['id'] for r in all_reports if not is_fatura_or_cartao(r['name'] or '')]
print(f"  Reports: {len(all_reports)} total, {len(all_reports) - len(valid_ids)} FATURA/CARTAO excluded, {len(valid_ids)} valid")

somase_by_cpf = {}
if valid_ids:
    cur.execute("""
        SELECT pr.user_cpf, COALESCE(SUM(pe.value), 0) as total
        FROM prestacao_expenses pe
        JOIN prestacao_reports pr ON pe.report_id = pr.id
        WHERE pr.id = ANY(%s::bigint[])
          AND COALESCE(pe.raw_data->>'payment_method_id', '') != '627401'
        GROUP BY pr.user_cpf
    """, (valid_ids,))
    for r in cur.fetchall():
        if r['user_cpf']:
            somase_by_cpf[r['user_cpf']] = float(r['total'])

# 2g. Saldo cartão (carga date)
def get_saldo_cartao(cutoff_date):
    cur.execute("""
        WITH deduped AS (
            SELECT DISTINCT ON (UPPER(usuario), data, tipo, valor, codigo_transacao)
                UPPER(usuario) AS usuario_up, data, tipo, valor, codigo_transacao
            FROM extrato_movimentacao
            WHERE is_snapshot = FALSE
              AND data <= %s
            ORDER BY UPPER(usuario), data, tipo, valor, codigo_transacao
        ),
        latest_snap AS (
            SELECT DISTINCT ON (UPPER(usuario))
                UPPER(usuario) AS usuario_up,
                valor AS saldo, data AS snapshot_date
            FROM extrato_movimentacao
            WHERE is_snapshot = TRUE
              AND valor IS NOT NULL
              AND data <= %s
            ORDER BY UPPER(usuario), data DESC
        ),
        post_snap_txns AS (
            SELECT d.usuario_up, SUM(d.valor) AS adjustment
            FROM deduped d
            JOIN latest_snap s ON d.usuario_up = s.usuario_up
            WHERE d.data > s.snapshot_date
            GROUP BY d.usuario_up
        ),
        computed_balance AS (
            SELECT usuario_up, COALESCE(SUM(valor), 0) AS saldo
            FROM deduped
            GROUP BY usuario_up
        )
        SELECT COALESCE(s.usuario_up, c.usuario_up) AS usuario_up,
               COALESCE(s.saldo, 0) + COALESCE(p.adjustment, 0) AS snap_saldo,
               COALESCE(c.saldo, 0) AS computed_saldo,
               (s.usuario_up IS NOT NULL) AS has_snapshot
        FROM latest_snap s
        FULL OUTER JOIN post_snap_txns p ON p.usuario_up = s.usuario_up
        FULL OUTER JOIN computed_balance c ON c.usuario_up = COALESCE(s.usuario_up, p.usuario_up)
    """, (cutoff_date, cutoff_date))
    result = {}
    for r in cur.fetchall():
        cpf = resolve_cpf_by_name(r['usuario_up'])
        if cpf:
            has_snap = r['has_snapshot']
            snap_saldo = float(r['snap_saldo'] or 0)
            computed_saldo = float(r['computed_saldo'] or 0)
            result[cpf] = round(snap_saldo if has_snap else computed_saldo, 2)
    return result

saldo_controle = get_saldo_cartao(financial_cutoff)
saldo_carga = get_saldo_cartao(saldo_cartao_carga_date)

# 2h. Build API-equivalent data
api_data = {}
for c in cadastro:
    cpf = c['cpf']
    manual = manuals.get(cpf, {})
    col_1qz = float(manual['col_1qz']) if manual.get('col_1qz') else None
    adiantamento = float(manual['adiantamento']) if manual.get('adiantamento') else 0.0

    carga = carga_by_cpf.get(cpf, 0)
    transf = transf_by_cpf.get(cpf, 0)
    tarifa = tarifa_by_cpf.get(cpf, 0)
    prestacao = somase_by_cpf.get(cpf, 0)
    saldo_prestacao = round(carga - transf - tarifa - prestacao, 2)

    sc_controle = saldo_controle.get(cpf, 0)
    sc_carga = saldo_carga.get(cpf, 0)

    saldo_final = round(saldo_prestacao - sc_controle, 2)
    saldo_final_carga = max(saldo_final, 0)
    saldo_reembolsar = max(-saldo_final, 0)

    is_pendente = 'pendente' in (c['status_cartao'] or '').lower() if c['status_cartao'] else False
    if is_pendente:
        carga_parcial = 0
        reembolso = 0
        carga_final = 0
    else:
        col_qz_efetivo = col_1qz if col_1qz is not None else 0
        carga_parcial = round(col_qz_efetivo - saldo_final_carga - sc_carga - adiantamento, 2)
        reembolso = 0  # 2QZ never has reembolso
        carga_final = round(max(0, carga_parcial) + reembolso, 2)

    api_data[cpf] = {
        'colaborador': c['colaborador'] or '',
        'cpf': cpf,
        'saldo_final': saldo_final_carga,
        'saldo_final_raw': saldo_final,
        'saldo_reembolsar': saldo_reembolsar,
        'saldo_cartao': sc_carga,
        'saldo_cartao_controle': sc_controle,
        'carga_parcial': carga_parcial,
        'reembolso': reembolso,
        'carga_final': carga_final,
        'col_2qz': col_1qz if col_1qz is not None else 0,
        'adiantamento': adiantamento,
        'carga': carga,
        'transferencia': transf,
        'tarifa': tarifa,
        'prestacao': prestacao,
        'saldo_prestacao': saldo_prestacao,
    }

conn.close()
print(f"  API-equivalent data rows: {len(api_data)}")

# ============================================================
# 3. Compare
# ============================================================
print("\n" + "=" * 100)
print("3. COMPARISON")
print("=" * 100)

all_cpfs = set(sheet_data.keys()) | set(api_data.keys())
sheet_only = set(sheet_data.keys()) - set(api_data.keys())
api_only = set(api_data.keys()) - set(sheet_data.keys())
both = set(sheet_data.keys()) & set(api_data.keys())

print(f"  Total CPFs: sheet={len(sheet_data)}, api={len(api_data)}")
print(f"  In both: {len(both)}, Sheet only: {len(sheet_only)}, API only: {len(api_only)}")

if sheet_only:
    print(f"\n  --- CPFs in SHEET but NOT in API ({len(sheet_only)}) ---")
    for cpf in sorted(sheet_only):
        r = sheet_data[cpf]
        print(f"    {cpf}  {r['colaborador'][:35]:<35}  CF={r['carga_final']:>10,.2f}")

if api_only:
    with_carga = [cpf for cpf in api_only if api_data[cpf]['carga_final'] > 0]
    print(f"\n  --- CPFs in API but NOT in SHEET: {len(api_only)} total, {len(with_carga)} with carga_final > 0 ---")
    for cpf in sorted(with_carga)[:20]:
        r = api_data[cpf]
        print(f"    {cpf}  {r['colaborador'][:35]:<35}  CF={r['carga_final']:>10,.2f}")

# Compare fields for CPFs in both
# Sheet SALDO REEMBOLSAR is negative (min(sf, 0)), API is positive (max(-sf, 0))
COMPARE_FIELDS = [
    ('saldo_final', 'saldo_final', 'saldo_final'),
    ('saldo_reembolsar', 'saldo_reembolsar', 'saldo_reembolsar'),
    ('saldo_cartao', 'saldo_cartao', 'saldo_cartao'),
    ('col_2qz', 'col_2qz', 'col_2qz'),
    ('adiantamento', 'adiantamento', 'adiantamento'),
    ('carga_parcial', 'carga_parcial', 'carga_parcial'),
    ('reembolso', 'reembolso', 'reembolso'),
    ('carga_final', 'carga_final', 'carga_final'),
]

diffs = []
for cpf in both:
    s = sheet_data[cpf]
    a = api_data[cpf]
    row_diffs = {}
    for sheet_key, api_key, label in COMPARE_FIELDS:
        sv = s[sheet_key]
        av = a[api_key]
        if sheet_key == 'saldo_reembolsar':
            sv = abs(sv)
        gap = round(sv - av, 2)
        if abs(gap) > 0.01:
            row_diffs[label] = {'sheet': sv, 'api': av, 'gap': gap}
    if row_diffs:
        diffs.append((cpf, s['colaborador'], row_diffs))

diffs.sort(key=lambda x: max([abs(v['gap']) for v in x[2].values()], default=0), reverse=True)

print(f"\n  --- Per-CPF differences (abs gap > R$ 0,01) ---")
print(f"  Total CPFs with differences: {len(diffs)}")
print(f"  Total CPFs matching exactly: {len(both) - len(diffs)}")

for cpf, nome, row_diffs in diffs[:50]:
    print(f"\n    {cpf}  {nome[:35]:<35}")
    for field, vals in row_diffs.items():
        print(f"      {field:<20}  sheet={vals['sheet']:>12,.2f}  api={vals['api']:>12,.2f}  gap={vals['gap']:>10,.2f}")

if len(diffs) > 50:
    print(f"\n    ... and {len(diffs) - 50} more")

# ============================================================
# 4. Totals
# ============================================================
print("\n" + "=" * 100)
print("4. TOTALS COMPARISON (all rows)")
print("=" * 100)

for sheet_key, api_key, label in COMPARE_FIELDS:
    sheet_total = sum(abs(s[sheet_key]) if sheet_key == 'saldo_reembolsar' else s[sheet_key] for s in sheet_data.values())
    api_total = sum(a[api_key] for a in api_data.values())
    gap = round(sheet_total - api_total, 2)
    status = "OK" if abs(gap) < 1.0 else "XX"
    print(f"  [{status}] {label:<20}  sheet={sheet_total:>14,.2f}  api={api_total:>14,.2f}  gap={gap:>10,.2f}")

print(f"\n  --- Totals for CPFs in both ({len(both)} CPFs) ---")
for sheet_key, api_key, label in COMPARE_FIELDS:
    sheet_total = sum(abs(sheet_data[cpf][sheet_key]) if sheet_key == 'saldo_reembolsar' else sheet_data[cpf][sheet_key] for cpf in both)
    api_total = sum(api_data[cpf][api_key] for cpf in both)
    gap = round(sheet_total - api_total, 2)
    status = "OK" if abs(gap) < 1.0 else "XX"
    print(f"  [{status}] {label:<20}  sheet={sheet_total:>14,.2f}  api={api_total:>14,.2f}  gap={gap:>10,.2f}")

# ============================================================
# 5. Categorize differences
# ============================================================
if diffs:
    print("\n" + "=" * 100)
    print("5. DIFFERENCE ANALYSIS")
    print("=" * 100)
    
    field_counts = {}
    for _, _, row_diffs in diffs:
        for field in row_diffs:
            field_counts[field] = field_counts.get(field, 0) + 1
    
    print("\n  Fields with differences (count of CPFs):")
    for field, count in sorted(field_counts.items(), key=lambda x: -x[1]):
        print(f"    {field:<20}  {count} CPFs")
    
    for field_name in ['col_2qz', 'saldo_cartao', 'saldo_final', 'carga_parcial', 'carga_final']:
        field_diffs = [(cpf, nome, rd) for cpf, nome, rd in diffs if field_name in rd]
        if field_diffs:
            print(f"\n  --- {field_name} differences ({len(field_diffs)} CPFs) ---")
            for cpf, nome, rd in field_diffs[:25]:
                d = rd[field_name]
                print(f"    {cpf}  {nome[:30]:<30}  sheet={d['sheet']:>10,.2f}  api={d['api']:>10,.2f}  gap={d['gap']:>8,.2f}")
