#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Inspect all available Excel sheets: validity, sheet names, header rows."""
import os
import sys
import glob
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))

def find_xlsx(base):
    out = []
    for dp, dn, fn in os.walk(base):
        for f in fn:
            if f.lower().endswith('.xlsx') and not f.startswith('~$'):
                out.append(os.path.join(dp, f))
    return out

files = []
files += find_xlsx(os.path.join(ROOT, 'data'))
# also dashboard-test/data
dt = os.path.dirname(ROOT)
files += find_xlsx(os.path.join(dt, 'data'))
# Downloads
dl = os.path.expanduser('~/Downloads')
files += [f for f in find_xlsx(dl) if any(k in f.upper() for k in ('VEXPENSES','CARGA','CONTROLE','QZ'))]

seen = set()
for f in sorted(files):
    if f in seen:
        continue
    seen.add(f)
    sz = os.path.getsize(f)
    print(f"\n{'='*80}\nFILE: {f}\nSIZE: {sz} bytes")
    if sz < 1000:
        print("  -> TOO SMALL (likely corrupted/error text)")
        try:
            with open(f, 'r', errors='replace') as fh:
                print("  CONTENT:", fh.read()[:200])
        except Exception:
            pass
        continue
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        print("  SHEETS:", wb.sheetnames)
        for ws in wb.worksheets[:3]:
            print(f"  -- sheet '{ws.title}' dims={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}")
            # print first 6 rows
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True)):
                # truncate long
                vals = [str(c)[:18] if c is not None else '' for c in row[:16]]
                print(f"     r{i+1}: {vals}")
        wb.close()
    except Exception as e:
        print(f"  ERROR opening: {e}")
