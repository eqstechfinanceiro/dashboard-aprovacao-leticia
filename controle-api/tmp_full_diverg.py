#!/usr/bin/env python3
"""Comprehensive divergence analysis between reference and our generated PAINEL."""
import openpyxl
from collections import defaultdict

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"
OURS_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\data\CONTROLE - VEXPENSES - JULHO 2026 - 2QZ - API.xlsx"

wb_ref = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_ref = wb_ref["PAINEL"]

wb_ours = openpyxl.load_workbook(OURS_PATH, read_only=True, data_only=True)
ws_ours = wb_ours["PAINEL"]

ref = {}
for row in ws_ref.iter_rows(min_row=12, values_only=True):
    cpf = str(row[2] or "").strip().replace(".", "").replace("-", "").replace(" ", "").zfill(11) if row[2] else ""
    if not cpf or cpf == "00000000000":
        continue
    ref[cpf] = {
        "nome": str(row[1] or ""),
        "carga": float(row[13] or 0),
        "transf": abs(float(row[14] or 0)),
        "tarifa": abs(float(row[15] or 0)),
        "prestacao": abs(float(row[16] or 0)),
        "saldo_prest": float(row[17] or 0),
        "saldo_cartao": abs(float(row[18] or 0)),
        "saldo_final": float(row[19] or 0),
    }
wb_ref.close()

ours = {}
for row in ws_ours.iter_rows(min_row=12, values_only=True):
    cpf = str(row[2] or "").strip().replace(".", "").replace("-", "").replace(" ", "").zfill(11) if row[2] else ""
    if not cpf or cpf == "00000000000":
        continue
    ours[cpf] = {
        "nome": str(row[1] or ""),
        "carga": float(row[13] or 0),
        "transf": abs(float(row[14] or 0)),
        "tarifa": abs(float(row[15] or 0)),
        "prestacao": abs(float(row[16] or 0)),
        "saldo_prest": float(row[17] or 0),
        "saldo_cartao": abs(float(row[18] or 0)),
        "saldo_final": float(row[19] or 0),
    }
wb_ours.close()

ref_only = set(ref.keys()) - set(ours.keys())
ours_only = set(ours.keys()) - set(ref.keys())
both = set(ref.keys()) & set(ours.keys())

print("=" * 80)
print("CPF COVERAGE")
print("  Reference: {} CPFs".format(len(ref)))
print("  Ours:      {} CPFs".format(len(ours)))
print("  In both:   {} CPFs".format(len(both)))
print("  Ref only:  {} CPFs".format(len(ref_only)))
print("  Ours only: {} CPFs".format(len(ours_only)))

if ref_only:
    print("\n  CPFs only in reference:")
    for cpf in sorted(ref_only):
        print("    {} {}".format(cpf, ref[cpf]["nome"][:40]))

if ours_only:
    print("\n  CPFs only in ours:")
    for cpf in sorted(ours_only):
        print("    {} {}".format(cpf, ours[cpf]["nome"][:40]))

# Field-by-field analysis
fields = ["carga", "transf", "tarifa", "prestacao", "saldo_prest", "saldo_cartao", "saldo_final"]
all_cpfs = set(ref.keys()) | set(ours.keys())

for field in fields:
    divs = []
    for cpf in all_cpfs:
        rv = ref.get(cpf, {}).get(field, 0)
        ov = ours.get(cpf, {}).get(field, 0)
        diff = ov - rv
        if abs(diff) > 0.50:
            divs.append((cpf, rv, ov, diff))
    
    total_gap = sum(abs(d[3]) for d in divs)
    print("\n" + "=" * 80)
    print("FIELD: {} — {} divergences, total gap R$ {:,.2f}".format(field.upper(), len(divs), total_gap))
    print("-" * 80)
    
    if field == "saldo_cartao" and not divs:
        print("  PERFECT MATCH!")
        continue
    
    for cpf, rv, ov, d in sorted(divs, key=lambda x: abs(x[3]), reverse=True):
        rname = ref.get(cpf, {}).get("nome", ours.get(cpf, {}).get("nome", "?"))
        in_ref = cpf in ref
        in_ours = cpf in ours
        tag = ""
        if not in_ref:
            tag = " [NOT IN REF]"
        elif not in_ours:
            tag = " [NOT IN OURS]"
        print("  {} {:<30} ref={:>12,.2f}  ours={:>12,.2f}  diff={:>+10,.2f}{}".format(
            cpf, rname[:30], rv, ov, d, tag))

print("\n" + "=" * 80)
print("SUMMARY: Which fields have the most divergences?")
print("-" * 80)
for field in fields:
    divs = []
    for cpf in all_cpfs:
        rv = ref.get(cpf, {}).get(field, 0)
        ov = ours.get(cpf, {}).get(field, 0)
        if abs(ov - rv) > 0.50:
            divs.append(abs(ov - rv))
    print("  {:<20}: {:>3} divs, total gap R$ {:>12,.2f}".format(field, len(divs), sum(divs)))
