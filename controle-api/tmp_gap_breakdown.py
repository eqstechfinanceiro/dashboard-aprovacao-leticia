#!/usr/bin/env python3
"""
Complete gap breakdown: investigate every R$ of the 86k gap.
Components:
  A) 54 new reports (API only) — R$ +90,824
  B) 14 missing reports (Ref only) — R$ -2,709
  C) Common reports value diff — R$ -1,985
"""
import os, re
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from pathlib import Path
import openpyxl
from collections import defaultdict
from datetime import datetime, date

BASE = Path(__file__).parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"

# ============================================================
# Load reference BASE PREST
# ============================================================
wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_bp = wb["BASE PREST "]

ref_expenses = []  # list of dicts
ref_report_ids = set()
ref_report_info = {}  # rid -> {name, user, cpf, count, total, expense_ids}
ref_by_cpf = defaultdict(lambda: {"count": 0, "total": 0, "report_ids": set()})
ref_total = 0

for row in ws_bp.iter_rows(min_row=4, values_only=True):
    if row[0] is None:
        continue
    eid = int(row[0])
    rid = int(row[1]) if row[1] else None
    rname = str(row[2] or "")
    user_name = str(row[4] or "")
    user_cpf = str(row[9] or "").strip()
    valor = float(row[26] or 0) if len(row) > 26 else 0
    ref_total += valor
    ref_expenses.append({"id": eid, "rid": rid, "name": rname, "user": user_name, "cpf": user_cpf, "value": valor})
    if rid:
        ref_report_ids.add(rid)
        if rid not in ref_report_info:
            ref_report_info[rid] = {"name": rname, "user": user_name, "cpf": user_cpf, "count": 0, "total": 0, "expense_ids": set()}
        ref_report_info[rid]["count"] += 1
        ref_report_info[rid]["total"] += valor
        ref_report_info[rid]["expense_ids"].add(eid)
        ref_by_cpf[user_cpf]["count"] += 1
        ref_by_cpf[user_cpf]["total"] += valor
        ref_by_cpf[user_cpf]["report_ids"].add(rid)

# Load PAINEL
ws_p = wb["PAINEL"]
vexpenses_cpfs = set()
painel_info = {}
for row in ws_p.iter_rows(min_row=12, values_only=True):
    if row[2] is None: continue
    cpf = str(row[2] or "").strip()
    colab = str(row[1] or "").strip()
    situacao = str(row[4] or "").strip().upper()
    status_cartao = str(row[5] or "").strip()
    cartao_vexp = str(row[12] or "").strip().upper()
    painel_info[cpf] = {"colaborador": colab, "situacao": situacao, "status_cartao": status_cartao, "cartao_vexp": cartao_vexp}
    if cartao_vexp == "SIM": vexpenses_cpfs.add(cpf)
wb.close()

# ============================================================
# Load API data
# ============================================================
conn = psycopg2.connect(NEON_URL, connect_timeout=10)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cur.execute("""
    SELECT e.id, e.report_id, e.value, e.date, e.description, e.status as exp_status,
           r.name as report_name, r.status as report_status,
           r.user_cpf, r.user_name, r.created_at, r.updated_at
    FROM prestacao_expenses e
    JOIN prestacao_reports r ON e.report_id = r.id
    WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado') AND r.user_cpf IS NOT NULL
""")
all_api = cur.fetchall()
conn.close()

def apply_base_filters(expenses):
    result = []
    for e in expenses:
        name = (e["report_name"] or "").strip()
        if re.match(r'^(fatu|farur|cart)', name, re.IGNORECASE): continue
        if re.search(r'(fatura|fatuar|fatut|farur|fartur)', name, re.IGNORECASE): continue
        if re.search(r'(cartão corporativo|cartao corporativo)', name, re.IGNORECASE): continue
        if re.search(r'itau|itaú', name, re.IGNORECASE): continue
        if e["user_cpf"] not in vexpenses_cpfs: continue
        result.append(e)
    return result

filtered = apply_base_filters(all_api)

# Build API report data
api_report_ids = set()
api_report_data = {}
api_expense_by_rid = defaultdict(list)
for e in filtered:
    rid = e["report_id"]
    api_report_ids.add(rid)
    api_expense_by_rid[rid].append(e)
    if rid not in api_report_data:
        api_report_data[rid] = {
            "name": e["report_name"], "user": e["user_name"], "cpf": e["user_cpf"],
            "status": e["report_status"], "created_at": e["created_at"],
            "expense_count": 0, "expense_total": 0, "expense_ids": set(),
        }
    api_report_data[rid]["expense_count"] += 1
    api_report_data[rid]["expense_total"] += float(e["value"])
    api_report_data[rid]["expense_ids"].add(e["id"])

new_reports = api_report_ids - ref_report_ids
missing_reports = ref_report_ids - api_report_ids
common_reports = api_report_ids & ref_report_ids

print("=" * 120)
print(f"COMPLETE GAP BREAKDOWN — Reference total: R$ {ref_total:,.2f}")
print("=" * 120)

# ============================================================
# A) 54 NEW REPORTS (API only) — R$ +90,824
# ============================================================
print(f"\n{'='*120}")
print(f"A) NEW REPORTS (API only): {len(new_reports)} reports, R$ {sum(api_report_data[r]['expense_total'] for r in new_reports):,.2f}")
print(f"{'='*120}")

# Categorize each new report
categories = defaultdict(lambda: {"reports": [], "total": 0, "count": 0})

for rid in sorted(new_reports):
    r = api_report_data[rid]
    cpf = r["cpf"]
    pinfo = painel_info.get(cpf, {})
    situacao = pinfo.get("situacao", "NOT IN PAINEL")
    status_cartao = pinfo.get("status_cartao", "?")
    created = r["created_at"]
    created_str = created.strftime("%Y-%m-%d") if created else "UNKNOWN"
    created_month = created.strftime("%Y-%m") if created else "UNKNOWN"
    status = r["status"].upper()
    user_has_ref = cpf in ref_by_cpf and len(ref_by_cpf[cpf]["report_ids"]) > 0

    # Categorize
    if situacao == "INATIVO":
        cat = "INATIVO_USER"
    elif status_cartao and status_cartao != "Cartão ativo":
        cat = f"CARD_ISSUE ({status_cartao})"
    elif not user_has_ref:
        cat = "ATIVO_NO_REF_AT_ALL"
    elif status == "ENVIADO":
        cat = "ENVIADO_USER_HAS_REF"
    elif created and created.year == 2025:
        cat = "OLD_2025_APROVADO"
    else:
        cat = "OTHER"

    categories[cat]["reports"].append(rid)
    categories[cat]["total"] += r["expense_total"]
    categories[cat]["count"] += 1

print(f"\n  {'Category':<35} {'Count':>6} {'Total R$':>14}  Details")
print(f"  {'-'*100}")
for cat in sorted(categories.keys(), key=lambda c: -categories[c]["total"]):
    c = categories[cat]
    pct = c["total"] / 90823.69 * 100
    print(f"  {cat:<35} {c['count']:>6} {c['total']:>14,.2f}  ({pct:.1f}% of new)")

    # Show details for each report in this category
    for rid in c["reports"]:
        r = api_report_data[rid]
        created = r["created_at"]
        created_str = created.strftime("%Y-%m-%d") if created else "?"
        print(f"    → {rid}  '{r['name'][:25]}'  {r['user'][:25]}  {r['status']:<8}  R$ {r['expense_total']:>10,.2f}  created={created_str}  {r['expense_count']} items")

# ============================================================
# B) 14 MISSING REPORTS (Ref only) — R$ -2,709
# ============================================================
print(f"\n{'='*120}")
print(f"B) MISSING REPORTS (Ref only): {len(missing_reports)} reports, R$ {sum(ref_report_info[r]['total'] for r in missing_reports):,.2f}")
print(f"{'='*120}")

for rid in sorted(missing_reports):
    r = ref_report_info[rid]
    pinfo = painel_info.get(r["cpf"], {})
    print(f"  → rid={rid}  '{r['name'][:30]}'  {r['user'][:25]}  CPF={r['cpf']}  R$ {r['total']:>10,.2f}  {r['count']} items  sit={pinfo.get('situacao','?')}  card={pinfo.get('status_cartao','?')}  vexp={pinfo.get('cartao_vexp','?')}")

# Check if these missing reports match by name+CPF to API reports
print("\n  Checking if missing ref reports match API reports by name+CPF:")
ref_by_name_cpf = defaultdict(list)
for rid, r in ref_report_info.items():
    ref_by_name_cpf[(r["name"].strip().upper(), r["cpf"])].append(rid)

api_by_name_cpf = defaultdict(list)
for rid, r in api_report_data.items():
    api_by_name_cpf[(r["name"].strip().upper(), r["cpf"])].append(rid)

for rid in sorted(missing_reports):
    r = ref_report_info[rid]
    key = (r["name"].strip().upper(), r["cpf"])
    if key in api_by_name_cpf:
        api_rids = api_by_name_cpf[key]
        api_total = sum(api_report_data[ar]["expense_total"] for ar in api_rids)
        print(f"    MATCH! ref rid={rid} → API rids={api_rids}  Ref R$ {r['total']:,.2f} vs API R$ {api_total:,.2f}")
    else:
        print(f"    NO MATCH: ref rid={rid} '{r['name']}' CPF={r['cpf']}  R$ {r['total']:,.2f}")

# ============================================================
# C) COMMON REPORTS VALUE DIFF — R$ -1,985
# ============================================================
print(f"\n{'='*120}")
print(f"C) COMMON REPORTS VALUE DIFF: {len(common_reports)} reports")
print(f"{'='*120}")

diffs = []
total_api_common = 0
total_ref_common = 0
for rid in common_reports:
    api_total = api_report_data[rid]["expense_total"]
    ref_total_r = ref_report_info[rid]["total"]
    diff = api_total - ref_total_r
    total_api_common += api_total
    total_ref_common += ref_total_r
    if abs(diff) > 0.01:
        diffs.append((rid, diff, api_total, ref_total_r, api_report_data[rid], ref_report_info[rid]))

total_common_diff = total_api_common - total_ref_common
print(f"\n  Total API (common):  R$ {total_api_common:,.2f}")
print(f"  Total Ref (common):  R$ {total_ref_common:,.2f}")
print(f"  Diff:                R$ {total_common_diff:,.2f}")
print(f"  Reports with diffs:  {len(diffs)} out of {len(common_reports)}")

# Sort by absolute diff
diffs.sort(key=lambda x: -abs(x[1]))
print(f"\n  {'RID':>10}  {'Diff R$':>12}  {'API R$':>12}  {'Ref R$':>12}  {'Report':<25}  {'User':<25}  API_items  Ref_items")
print(f"  {'-'*115}")
for rid, diff, api_t, ref_t, api_r, ref_r in diffs[:30]:
    print(f"  {rid:>10}  {diff:>+12,.2f}  {api_t:>12,.2f}  {ref_t:>12,.2f}  {api_r['name'][:25]:<25}  {api_r['user'][:25]:<25}  {api_r['expense_count']:>5}  {ref_r['count']:>5}")

# For top 10 diffs, show expense-level breakdown
print(f"\n  --- Top 10 diffs: expense-level breakdown ---")
for rid, diff, api_t, ref_t, api_r, ref_r in diffs[:10]:
    print(f"\n  RID={rid}  '{api_r['name']}'  {api_r['user']}  diff=R$ {diff:,.2f}")
    print(f"    API: {api_r['expense_count']} expenses, R$ {api_t:,.2f}")
    print(f"    Ref: {ref_r['count']} expenses, R$ {ref_t:,.2f}")

    # Check expense ID overlap
    api_eids = api_r["expense_ids"]
    ref_eids = ref_r["expense_ids"]
    common_eids = api_eids & ref_eids
    api_only_eids = api_eids - ref_eids
    ref_only_eids = ref_eids - api_eids

    print(f"    Common expense IDs: {len(common_eids)}")
    print(f"    API-only expense IDs: {len(api_only_eids)}")
    print(f"    Ref-only expense IDs: {len(ref_only_eids)}")

    if api_only_eids:
        api_only_total = sum(float(e["value"]) for e in api_expense_by_rid[rid] if e["id"] in api_only_eids)
        print(f"    API-only expenses total: R$ {api_only_total:,.2f}")
        for e in api_expense_by_rid[rid]:
            if e["id"] in api_only_eids:
                print(f"      → eid={e['id']}  value=R$ {float(e['value']):,.2f}  date={e['date']}  desc={str(e['description'] or '')[:40]}  status={e['exp_status']}")

    if ref_only_eids:
        ref_only_total = sum(re["value"] for re in ref_expenses if re["rid"] == rid and re["id"] in ref_only_eids)
        print(f"    Ref-only expenses total: R$ {ref_only_total:,.2f}")
        for re in ref_expenses:
            if re["rid"] == rid and re["id"] in ref_only_eids:
                print(f"      → eid={re['id']}  value=R$ {re['value']:,.2f}")

    # Check value mismatches in common expense IDs
    api_by_eid = {e["id"]: e for e in api_expense_by_rid[rid]}
    ref_by_eid = {re["id"]: re for re in ref_expenses if re["rid"] == rid}
    value_mismatches = []
    for eid in common_eids:
        av = float(api_by_eid[eid]["value"])
        rv = ref_by_eid[eid]["value"]
        if abs(av - rv) > 0.01:
            value_mismatches.append((eid, av, rv, av - rv))

    if value_mismatches:
        print(f"    Value mismatches in common IDs: {len(value_mismatches)}")
        for eid, av, rv, vd in value_mismatches:
            print(f"      → eid={eid}  API=R$ {av:,.2f}  Ref=R$ {rv:,.2f}  diff=R$ {vd:,.2f}")

# ============================================================
# D) SUMMARY: Where does the 86k come from?
# ============================================================
print(f"\n{'='*120}")
print(f"D) SUMMARY: Complete gap decomposition")
print(f"{'='*120}")

new_total = sum(api_report_data[r]["expense_total"] for r in new_reports)
missing_total = sum(ref_report_info[r]["total"] for r in missing_reports)
common_diff = total_api_common - total_ref_common

print(f"\n  Component                              R$")
print(f"  {'─'*50}")
print(f"  New reports (API only, +)         {new_total:>+14,.2f}")
print(f"  Missing reports (Ref only, -)     {-missing_total:>+14,.2f}")
print(f"  Common reports value diff         {common_diff:>+14,.2f}")
print(f"  {'─'*50}")
print(f"  NET GAP                            {new_total - missing_total + common_diff:>+14,.2f}")
print(f"  Reference total                    {ref_total:>14,.2f}")
print(f"  Gap %                              {abs(new_total - missing_total + common_diff) / ref_total * 100:>13.1f}%")

print(f"\n  --- New reports breakdown by category ---")
for cat in sorted(categories.keys(), key=lambda c: -categories[c]["total"]):
    c = categories[cat]
    print(f"  {cat:<35}  R$ {c['total']:>12,.2f}  ({c['total']/new_total*100:.1f}%)")

print(f"\n  --- Actionable items ---")
inativo_total = categories.get("INATIVO_USER", {"total": 0})["total"]
card_issue_total = sum(c["total"] for cat, c in categories.items() if cat.startswith("CARD_ISSUE"))
enviado_total = categories.get("ENVIADO_USER_HAS_REF", {"total": 0})["total"]
old_total = categories.get("OLD_2025_APROVADO", {"total": 0})["total"]
ativo_no_ref = categories.get("ATIVO_NO_REF_AT_ALL", {"total": 0})["total"]
other_total = categories.get("OTHER", {"total": 0})["total"]

print(f"  Filter INATIVO users:           -R$ {inativo_total:>12,.2f}  (but ref includes some INATIVO users too)")
print(f"  Filter card issues:             -R$ {card_issue_total:>12,.2f}  (não vinculado, cadastro pendente)")
print(f"  ENVIADO with ref user:          -R$ {enviado_total:>12,.2f}  (can't filter - ref has ENVIADO too)")
print(f"  Old 2025 APROVADO:              -R$ {old_total:>12,.2f}  (very old reports)")
print(f"  ATIVO+Cartão ativo, no ref:     -R$ {ativo_no_ref:>12,.2f}  (should be in ref?)")
print(f"  Other:                          -R$ {other_total:>12,.2f}")

# Check: does reference include any INATIVO users?
ref_inativo = 0
ref_inativo_total = 0
for cpf, info in ref_by_cpf.items():
    pinfo = painel_info.get(cpf, {})
    if pinfo.get("situacao") == "INATIVO":
        ref_inativo += 1
        ref_inativo_total += info["total"]
print(f"\n  Reference includes {ref_inativo} INATIVO users with R$ {ref_inativo_total:,.2f} in expenses")
print(f"  → Filtering INATIVO would also remove these from our side, widening the gap")

# Check: what % of the 86k can we safely eliminate?
safely_removable = inativo_total + card_issue_total + old_total
print(f"\n  Safely removable (INATIVO + card issues + old 2025): R$ {safely_removable:,.2f}")
print(f"  Remaining gap after safe removals: R$ {86429.61 - safely_removable:,.2f}")
print(f"  Remaining gap %: {(86429.61 - safely_removable) / ref_total * 100:.1f}%")
