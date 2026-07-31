#!/usr/bin/env python3
"""
Check SO NO NEON expenses: which reports do they belong to?
Are they all from the 60 'reports only in Neon', or are some from reports in the ref?
"""
import openpyxl
from pathlib import Path
from collections import Counter

xlsx_path = Path(r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\controle-api\data\gap entre referencia e neon ahahahahaah.xlsx")
wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

# Read SO NO NEON sheet
ws = wb["SO NO NEON"]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
print(f"SO NO NEON headers: {header}")
print(f"Total rows: {len(rows)-1}")

# Count by report_id
report_counts = Counter()
report_totals = {}
report_info = {}
for row in rows[1:]:
    if row[0] is None:
        continue
    rid = row[1]  # Report ID
    report_counts[rid] += 1
    val = float(row[6]) if row[6] else 0
    if rid not in report_totals:
        report_totals[rid] = 0
        report_info[rid] = {"name": row[2], "user": row[3], "cpf": row[4], "status": row[5]}
    report_totals[rid] += val

print(f"\n=== SO NO NEON expenses by report ({len(report_counts)} unique reports) ===")
print(f"{'Report ID':<12} {'Name':<35} {'User':<30} {'Status':<10} {'#Exp':>5} {'Total':>12}")
print("-" * 110)

# Sort by total descending
for rid, total in sorted(report_totals.items(), key=lambda x: -x[1]):
    info = report_info[rid]
    cnt = report_counts[rid]
    print(f"{rid:<12} {str(info['name'])[:34]:<35} {str(info['user'])[:29]:<30} {str(info['status']):<10} {cnt:>5} R$ {total:>10.2f}")

# Now check: how many of these reports are in the "REPORTS SO NO NEON" sheet?
ws2 = wb["REPORTS SO NO NEON"]
report_only_rows = list(ws2.iter_rows(values_only=True))
report_only_rids = set()
for row in report_only_rows[1:]:
    if row[0] is not None:
        report_only_rids.add(row[0])

print(f"\n=== Cross-reference ===")
in_both = set(report_counts.keys()) & report_only_rids
only_in_expenses = set(report_counts.keys()) - report_only_rids
print(f"Reports in SO NO NEON expenses AND in REPORTS SO NO NEON: {len(in_both)}")
print(f"Reports in SO NO NEON expenses but NOT in REPORTS SO NO NEON: {len(only_in_expenses)}")

if only_in_expenses:
    print(f"\n=== Expenses from reports NOT in 'REPORTS SO NO NEON' (these are in ref but have extra expenses in Neon) ===")
    for rid in only_in_expenses:
        info = report_info[rid]
        cnt = report_counts[rid]
        total = report_totals[rid]
        print(f"  rid={rid} | {info['name']} | {info['user']} | {info['status']} | {cnt} exp | R$ {total:.2f}")

# Also check MAIS DESPESAS NO NEON
ws3 = wb["MAIS DESPESAS NO NEON"]
mais_rows = list(ws3.iter_rows(values_only=True))
print(f"\n=== MAIS DESPESAS NO NEON ({len(mais_rows)-1} reports) ===")
for row in mais_rows[1:]:
    if row[0] is None:
        continue
    print(f"  rid={row[0]} | {row[1]} | {row[2]} | {row[4]} | ref={row[5]} neon={row[6]} diff={row[7]} | ref_R$={row[8]} neon_R$={row[9]} diff_R$={row[10]}")

# Check MAIS DESPESAS NO REF
ws4 = wb["MAIS DESPESAS NO REF"]
mais_ref_rows = list(ws4.iter_rows(values_only=True))
print(f"\n=== MAIS DESPESAS NO REF ({len(mais_ref_rows)-1} reports) ===")
for row in mais_ref_rows[1:]:
    if row[0] is None:
        continue
    print(f"  rid={row[0]} | {row[1]} | {row[3]} | {row[4]} | ref={row[5]} neon={row[6]} diff={row[7]} | ref_R$={row[8]} neon_R$={row[9]} diff_R$={row[10]}")

# VALOR DIFERENTE
ws5 = wb["VALOR DIFERENTE"]
vd_rows = list(ws5.iter_rows(values_only=True))
print(f"\n=== VALOR DIFERENTE ({len(vd_rows)-1} expenses) ===")
for row in vd_rows[1:]:
    if row[0] is None:
        continue
    print(f"  exp_id={row[0]} | rid={row[1]} | {row[2]} | ref_R$={row[5]} neon_R$={row[6]} diff={row[7]}")

wb.close()
