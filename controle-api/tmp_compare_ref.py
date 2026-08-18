#!/usr/bin/env python3
"""Compare reference CONTROLE with our generated sheet."""
import openpyxl

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"
OURS_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\data\CONTROLE - VEXPENSES - JULHO 2026 - 2QZ - API.xlsx"

wb_ref = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_ref = wb_ref["PAINEL"]

wb_ours = openpyxl.load_workbook(OURS_PATH, read_only=True, data_only=True)
ws_ours = wb_ours["PAINEL"]

ref = {}
for row in ws_ref.iter_rows(min_row=12, values_only=True):
    cpf = str(row[2] or "").strip().replace(".", "").replace("-", "").replace(" ", "").zfill(11) if row[2] else ""
    if not cpf or cpf == "00000000000":
        continue
    ref[cpf] = {
        "nome": str(row[1] or ""),
        "carga": float(row[13] or 0),
        "transf": abs(float(row[14] or 0)),
        "tarifa": abs(float(row[15] or 0)),
        "prestacao": abs(float(row[16] or 0)),
        "saldo_prest": float(row[17] or 0),
        "saldo_cartao": abs(float(row[18] or 0)),
        "saldo_final": float(row[19] or 0),
    }
wb_ref.close()

ours = {}
for row in ws_ours.iter_rows(min_row=12, values_only=True):
    cpf = str(row[2] or "").strip().replace(".", "").replace("-", "").replace(" ", "").zfill(11) if row[2] else ""
    if not cpf or cpf == "00000000000":
        continue
    ours[cpf] = {
        "nome": str(row[1] or ""),
        "carga": float(row[13] or 0),
        "transf": abs(float(row[14] or 0)),
        "tarifa": abs(float(row[15] or 0)),
        "prestacao": abs(float(row[16] or 0)),
        "saldo_prest": float(row[17] or 0),
        "saldo_cartao": abs(float(row[18] or 0)),
        "saldo_final": float(row[19] or 0),
    }
wb_ours.close()

print("Reference: {} CPFs".format(len(ref)))
print("Ours: {} CPFs".format(len(ours)))

fields = ["carga", "transf", "tarifa", "prestacao", "saldo_prest", "saldo_cartao", "saldo_final"]
all_cpfs = set(ref.keys()) | set(ours.keys())

for field in fields:
    match = 0
    divs = []
    for cpf in all_cpfs:
        rv = ref.get(cpf, {}).get(field, 0)
        ov = ours.get(cpf, {}).get(field, 0)
        if abs(ov - rv) <= 0.50:
            match += 1
        else:
            divs.append((cpf, rv, ov, ov - rv))
    pct = match / len(all_cpfs) * 100
    total_gap = sum(abs(d[3]) for d in divs)
    print("\n  {:<15}: {}/{} = {:.1f}%  ({} divs, total gap R$ {:,.2f})".format(
        field, match, len(all_cpfs), pct, len(divs), total_gap))
    for cpf, rv, ov, d in sorted(divs, key=lambda x: abs(x[3]), reverse=True)[:5]:
        rname = ref.get(cpf, {}).get("nome", ours.get(cpf, {}).get("nome", "?"))
        print("    {} {:<25} ref={:>12,.2f} ours={:>12,.2f} diff={:>+10,.2f}".format(
            cpf, rname[:25], rv, ov, d))
