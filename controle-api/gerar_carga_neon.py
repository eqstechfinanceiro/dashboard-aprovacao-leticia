#!/usr/bin/env python3
"""
Gerador da planilha Carga Quinzenal — VERSÃO NEON (100% API).

Lê dados do Neon PostgreSQL (mesma fonte do dashboard Next.js) e gera
uma planilha Excel com as colunas da CARGA QZ.

Filtros aplicados (idênticos ao pipeline e quinzena-complete):
- Exclui Cartão Itaú (pm_id=627401)
- Exclui relatórios FATURA/CARTÃO por nome
- Inclui APROVADO + ENVIADO
- Usa approval_date do raw_data para cutoff (updated_at é NULL)

Uso:
    python gerar_carga_neon.py --year 2026 --month 7 --quinzena 2
    python gerar_carga_neon.py --year 2026 --month 7 --quinzena 2 --output data/carga_jul_2qz.xlsx
"""

import argparse
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl não instalado. Instale com: pip install openpyxl")
    sys.exit(1)


# ---- FATURA/CARTAO filter (must match isFaturaOrCartao in pipeline.ts) --------

def is_fatura_or_cartao(name: str) -> bool:
    n = (name or '').strip().upper()
    if not n:
        return False
    if 'CAIXA ITAU' in n or 'CAIXA ITAÚ' in n:
        return True
    if n.startswith('CAIXA'):
        return False
    if re.match(r'^(FATURA|CARTAO|CARTÃO|FATUAR|FARTUR|FATUT|FARUR|FATUTR)', n):
        return True
    if 'CARTÃO DE CRÉDITO' in n or 'CARTAO DE CREDITO' in n or 'CARTÃO DE CREDITO' in n:
        return True
    if 'CARTÃO CORPORATIVO' in n:
        return True
    if ('ITAU' in n or 'ITAÚ' in n) and 'CAIXA' not in n:
        return True
    if 'DOLAR' in n or 'DÓLAR' in n:
        return True
    if n.startswith('DESPESA') and 'FATURA' in n:
        return True
    if n.startswith('COMPLEMENTAR') and 'FATURA' in n:
        return True
    if 'CARTÃO' in n and 'CRÉDITO' in n:
        return True
    if 'CARTAO' in n and 'CREDITO' in n:
        return True
    if n.startswith('CARTÃO VEXPENSES'):
        return True
    return False


# ---- Name normalization for extrato matching ---------------------------------

def normalize_name(s: str) -> str:
    if not s:
        return ''
    import unicodedata
    s = s.upper().strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'\s+', ' ', s)
    return s


def fuzzy_ratio(a: str, b: str) -> float:
    if a == b:
        return 1.0
    if not a or not b:
        return 0.0
    bigrams_a = set(a[i:i+2] for i in range(len(a) - 1))
    bigrams_b = set(b[i:i+2] for i in range(len(b) - 1))
    if not bigrams_a or not bigrams_b:
        return 0.0
    inter = len(bigrams_a & bigrams_b)
    return (2 * inter) / (len(bigrams_a) + len(bigrams_b))


def resolve_cpf_by_name(
    extrato_name: str,
    name_to_cpf: dict,
    fuzzy_cache: dict,
) -> Optional[str]:
    normalized = normalize_name(extrato_name)
    exact = name_to_cpf.get(normalized)
    if exact:
        return exact
    cached = fuzzy_cache.get(normalized)
    if cached:
        return cached
    best_cpf = None
    best_ratio = 0.0
    for cad_name, cpf in name_to_cpf.items():
        ratio = fuzzy_ratio(normalized, cad_name)
        if ratio > best_ratio:
            best_ratio = ratio
            best_cpf = cpf
    if best_ratio >= 0.88 and best_cpf:
        fuzzy_cache[normalized] = best_cpf
        return best_cpf
    if len(normalized) >= 10:
        prefix15 = normalized[:15]
        for cad_name, cpf in name_to_cpf.items():
            if cad_name[:15] == prefix15:
                return cpf
        prefix10 = normalized[:10]
        for cad_name, cpf in name_to_cpf.items():
            if cad_name[:10] == prefix10:
                return cpf
    return None


# ---- Quinzena date helpers ---------------------------------------------------

MONTH_NAMES = {
    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
    5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
    9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro',
}


def get_quinzena_dates(year: int, month: int, quinzena: int):
    if quinzena == 1:
        start_date = f"{year}-{month - 1:02d}-26" if month > 1 else f"{year - 1}-12-26"
        end_date = f"{year}-{month:02d}-10"
        fechamento = f"{year}-{month:02d}-11"
        financial_cutoff = f"{year}-{month:02d}-01"
        saldo_cartao_controle_date = f"{year}-{month:02d}-01"
        saldo_cartao_carga_date = f"{year}-{month:02d}-11"
    else:
        start_date = f"{year}-{month:02d}-11"
        end_date = f"{year}-{month:02d}-25"
        fechamento = f"{year}-{month:02d}-25"
        financial_cutoff = f"{year}-{month:02d}-01"
        saldo_cartao_controle_date = f"{year}-{month:02d}-01"
        saldo_cartao_carga_date = f"{year}-{month:02d}-25"
    return {
        'start_date': start_date,
        'end_date': end_date,
        'fechamento': fechamento,
        'financial_cutoff': financial_cutoff,
        'saldo_cartao_controle_date': saldo_cartao_controle_date,
        'saldo_cartao_carga_date': saldo_cartao_carga_date,
    }


# ---- Main generator ----------------------------------------------------------

def gerar_carga_neon(year: int, month: int, quinzena: int, output_path: Path):
    load_dotenv(Path(__file__).parent / ".env")
    db_url = os.getenv("NEON_DATABASE_URL")
    if not db_url:
        print("ERRO: NEON_DATABASE_URL não configurado")
        sys.exit(1)

    dates = get_quinzena_dates(year, month, quinzena)
    print(f"Gerando CARGA QZ: {MONTH_NAMES[month]} {year} - {quinzena}a Quinzena")
    print(f"  Período: {dates['start_date']} → {dates['end_date']}")
    print(f"  Saldo cartão controle: {dates['saldo_cartao_controle_date']}")
    print(f"  Saldo cartão carga: {dates['saldo_cartao_carga_date']}")

    conn = psycopg2.connect(db_url)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # 1. Load cadastro
    print("\n[1/6] Carregando cadastro...")
    cur.execute("""
        SELECT cpf, colaborador, situacao, status_cartao,
               regional, centro_custo, gestor, diretor
        FROM quinzena_cadastro
        ORDER BY colaborador ASC NULLS LAST
    """)
    cadastro = cur.fetchall()
    print(f"  {len(cadastro)} colaboradores")

    # 2. Load manual inputs
    print("[2/6] Carregando inputs manuais...")
    cur.execute("""
        SELECT col_1qz::text, adiantamento::text, obs, cpf
        FROM quinzena_manual_inputs
        WHERE year = %s AND month = %s AND quinzena = %s
    """, (year, month, quinzena))
    manuals = {r['cpf']: r for r in cur.fetchall() if r['cpf']}
    print(f"  {len(manuals)} inputs manuais")

    # 3. Load reembolso multiplier
    cur.execute("""
        SELECT reembolso_multiplier::text
        FROM quinzena_config
        WHERE year = %s AND month = %s AND quinzena = %s
    """, (year, month, quinzena))
    config_row = cur.fetchone()
    reembolso_multiplier = float(config_row['reembolso_multiplier']) if config_row else 0.5
    print(f"  Reembolso multiplier: {reembolso_multiplier}")

    # 4. Extrato cumulativo (carga, transferencia, tarifa)
    print("[3/6] Carregando extrato...")
    name_to_cpf = {}
    for c in cadastro:
        norm = normalize_name(c['colaborador'] or '')
        if norm:
            name_to_cpf[norm] = c['cpf']
    fuzzy_cache = {}

    cur.execute("""
        WITH deduped AS (
            SELECT DISTINCT ON (
                UPPER(usuario), data, tipo, valor,
                COALESCE(NULLIF(codigo_transacao, ''), hora::text)
            )
                UPPER(usuario) AS usuario_up,
                data, tipo, valor, codigo_transacao
            FROM extrato_movimentacao
            WHERE is_snapshot = FALSE
              AND data <= %s
            ORDER BY UPPER(usuario), data, tipo, valor,
                COALESCE(NULLIF(codigo_transacao, ''), hora::text)
        )
        SELECT
            usuario_up,
            COALESCE(SUM(valor) FILTER(WHERE tipo = 'Transferência' AND valor > 0), 0) AS carga_raw,
            COALESCE(SUM(valor) FILTER(WHERE tipo = 'Transferência' AND valor < 0), 0) AS transf_raw,
            COALESCE(SUM(valor) FILTER(WHERE tipo = 'Taxa'), 0) AS tarifa_raw
        FROM deduped
        GROUP BY usuario_up
    """, (dates['financial_cutoff'],))
    extrato_rows = cur.fetchall()

    carga_by_cpf = {}
    transf_by_cpf = {}
    tarifa_by_cpf = {}
    for r in extrato_rows:
        cpf = resolve_cpf_by_name(r['usuario_up'], name_to_cpf, fuzzy_cache)
        if cpf:
            carga_by_cpf[cpf] = float(r['carga_raw'] or 0)
            transf_by_cpf[cpf] = abs(float(r['transf_raw'] or 0))
            tarifa_by_cpf[cpf] = abs(float(r['tarifa_raw'] or 0))

    # 5. Somase (prestação de contas) — same filters as pipeline
    print("[4/6] Calculando somase (prestação de contas)...")
    cur.execute("""
        SELECT r.id, r.name
        FROM prestacao_reports r
        WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
          AND r.user_cpf IS NOT NULL
    """)
    all_reports = cur.fetchall()
    valid_ids = [r['id'] for r in all_reports if not is_fatura_or_cartao(r['name'] or '')]
    excluded = len(all_reports) - len(valid_ids)
    print(f"  {len(all_reports)} relatórios, {excluded} FATURA/CARTAO excluídos, {len(valid_ids)} válidos")

    somase_by_cpf = {}
    if valid_ids:
        cur.execute("""
            SELECT pr.user_cpf, COALESCE(SUM(pe.value), 0) as total
            FROM prestacao_expenses pe
            JOIN prestacao_reports pr ON pe.report_id = pr.id
            WHERE pr.id = ANY(%s::bigint[])
              AND COALESCE(pe.raw_data->>'payment_method_id', '') != '627401'
            GROUP BY pr.user_cpf
        """, (valid_ids,))
        for r in cur.fetchall():
            if r['user_cpf']:
                somase_by_cpf[r['user_cpf']] = float(r['total'])

    # 6. Saldo cartão (controle + carga)
    print("[5/6] Calculando saldo cartão...")

    def get_saldo_cartao(cutoff_date: str) -> dict:
        cur.execute("""
            WITH deduped AS (
                SELECT DISTINCT ON (UPPER(usuario), data, tipo, valor, codigo_transacao)
                    UPPER(usuario) AS usuario_up, data, tipo, valor, codigo_transacao
                FROM extrato_movimentacao
                WHERE is_snapshot = FALSE
                  AND data <= %s
                ORDER BY UPPER(usuario), data, tipo, valor, codigo_transacao
            ),
            latest_snap AS (
                SELECT DISTINCT ON (UPPER(usuario))
                    UPPER(usuario) AS usuario_up,
                    valor AS saldo, data AS snapshot_date
                FROM extrato_movimentacao
                WHERE is_snapshot = TRUE
                  AND valor IS NOT NULL
                  AND data <= %s
                ORDER BY UPPER(usuario), data DESC
            ),
            post_snap_txns AS (
                SELECT d.usuario_up, SUM(d.valor) AS adjustment
                FROM deduped d
                JOIN latest_snap s ON d.usuario_up = s.usuario_up
                WHERE d.data > s.snapshot_date
                GROUP BY d.usuario_up
            ),
            computed_balance AS (
                SELECT usuario_up, COALESCE(SUM(valor), 0) AS saldo
                FROM deduped
                GROUP BY usuario_up
            )
            SELECT COALESCE(s.usuario_up, c.usuario_up) AS usuario_up,
                   COALESCE(s.saldo, 0) + COALESCE(p.adjustment, 0) AS snap_saldo,
                   COALESCE(c.saldo, 0) AS computed_saldo,
                   (s.usuario_up IS NOT NULL) AS has_snapshot
            FROM latest_snap s
            FULL OUTER JOIN post_snap_txns p ON p.usuario_up = s.usuario_up
            FULL OUTER JOIN computed_balance c ON c.usuario_up = COALESCE(s.usuario_up, p.usuario_up)
        """, (cutoff_date, cutoff_date))
        result = {}
        for r in cur.fetchall():
            cpf = resolve_cpf_by_name(r['usuario_up'], name_to_cpf, fuzzy_cache)
            if cpf:
                has_snap = r['has_snapshot']
                snap_saldo = float(r['snap_saldo'] or 0)
                computed_saldo = float(r['computed_saldo'] or 0)
                result[cpf] = round(snap_saldo if has_snap else computed_saldo, 2)
        return result

    saldo_controle = get_saldo_cartao(dates['saldo_cartao_controle_date'])
    saldo_carga = get_saldo_cartao(dates['saldo_cartao_carga_date'])

    # Build rows
    print("[6/6] Gerando planilha...")
    rows = []
    for c in cadastro:
        cpf = c['cpf']
        manual = manuals.get(cpf, {})
        col_1qz = float(manual['col_1qz']) if manual.get('col_1qz') else None
        adiantamento = float(manual['adiantamento']) if manual.get('adiantamento') else 0.0
        obs = manual.get('obs')

        carga = carga_by_cpf.get(cpf, 0)
        transf = transf_by_cpf.get(cpf, 0)
        tarifa = tarifa_by_cpf.get(cpf, 0)
        prestacao = somase_by_cpf.get(cpf, 0)
        saldo_prestacao = round(carga - transf - tarifa - prestacao, 2)

        sc_controle = saldo_controle.get(cpf, 0)
        sc_carga = saldo_carga.get(cpf, 0)

        saldo_final = round(saldo_prestacao - sc_controle, 2)
        saldo_final_carga = max(saldo_final, 0)
        saldo_reembolsar = max(-saldo_final, 0)

        is_pendente = 'pendente' in (c['status_cartao'] or '').lower() if c['status_cartao'] else False
        if is_pendente:
            carga_parcial = 0
            reembolso = 0
            carga_final = 0
        else:
            col_qz_efetivo = col_1qz if col_1qz is not None else 0
            carga_parcial = round(col_qz_efetivo - saldo_final_carga - sc_carga - adiantamento, 2)
            reembolso = round(max(0, saldo_reembolsar) * reembolso_multiplier, 2) if quinzena == 1 else 0
            carga_final = round(max(0, carga_parcial) + reembolso, 2)

        rows.append({
            'cpf': cpf,
            'colaborador': c['colaborador'] or '',
            'situacao': c['situacao'] or '',
            'regional': c['regional'] or '',
            'centro_custo': c['centro_custo'] or '',
            'gestor': c['gestor'] or '',
            'diretor': c['diretor'] or '',
            'status_cartao': c['status_cartao'] or '',
            'carga': carga,
            'transferencia': transf,
            'tarifa': tarifa,
            'prestacao': prestacao,
            'saldo_prestacao': saldo_prestacao,
            'saldo_final': saldo_final,
            'saldo_final_carga': saldo_final_carga,
            'saldo_reembolsar': saldo_reembolsar,
            'saldo_cartao': sc_controle,
            'saldo_cartao_carga': sc_carga,
            'col_qz': col_1qz,
            'adiantamento': adiantamento,
            'obs': obs,
            'carga_parcial': carga_parcial,
            'reembolso': reembolso,
            'carga_final': carga_final,
        })

    conn.close()

    # Sort by colaborador
    rows.sort(key=lambda r: r['colaborador'])

    # Generate Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"QZ{quinzena} {MONTH_NAMES[month][:3]}"

    # Column definitions
    COLS = [
        ('CPF', 'cpf', 15),
        ('COLABORADOR', 'colaborador', 34),
        ('SITUACAO', 'situacao', 10),
        ('REGIONAL', 'regional', 16),
        ('CENTRO DE CUSTO', 'centro_custo', 30),
        ('GESTOR', 'gestor', 28),
        ('DIRETOR', 'diretor', 28),
        ('STATUS CARTAO', 'status_cartao', 16),
        ('CARGA', 'carga', 14),
        ('TRANSFERENCIA', 'transferencia', 14),
        ('TARIFA', 'tarifa', 12),
        ('PRESTACAO', 'prestacao', 14),
        ('SALDO PRESTACAO', 'saldo_prestacao', 16),
        ('SALDO FINAL', 'saldo_final_carga', 14),
        ('SALDO REEMBOLSAR', 'saldo_reembolsar', 16),
        ('SALDO CARTAO', 'saldo_cartao_carga', 14),
        (f'{quinzena}a QZ', 'col_qz', 16),
        ('ADIANTAMENTO', 'adiantamento', 14),
        ('CARGA PARCIAL', 'carga_parcial', 14),
        ('REEMBOLSO', 'reembolso', 14),
        ('CARGA FINAL', 'carga_final', 14),
        ('OBS', 'obs', 24),
    ]

    numeric_keys = {'carga', 'transferencia', 'tarifa', 'prestacao', 'saldo_prestacao',
                    'saldo_final_carga', 'saldo_reembolsar', 'saldo_cartao_carga',
                    'col_qz', 'adiantamento', 'carga_parcial', 'reembolso', 'carga_final'}

    # Title row
    title = f"{MONTH_NAMES[month]} {year} - {quinzena}a Quinzena ({dates['start_date']} a {dates['end_date']})"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, color='FFFFFF', size=12)
    cell.fill = PatternFill('solid', fgColor='1E40AF')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Header row
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='1F4E79')
    thin_border = Border(
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
    )

    for col_idx, (header, _, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, row_data in enumerate(rows, start=3):
        for col_idx, (_, key, _) in enumerate(COLS, start=1):
            val = row_data.get(key)
            if val is None:
                val = ''
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if key in numeric_keys and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

    # Totals row
    totals_row = len(rows) + 3
    ws.cell(row=totals_row, column=2, value=f"TOTAL ({len(rows)})").font = Font(bold=True)
    for col_idx, (_, key, _) in enumerate(COLS, start=1):
        if key in numeric_keys:
            total = sum(r.get(key, 0) or 0 for r in rows)
            cell = ws.cell(row=totals_row, column=col_idx, value=round(total, 2))
            cell.number_format = '#,##0.00'
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1E40AF')
            cell.alignment = Alignment(horizontal='right')
            cell.border = thin_border

    # Freeze panes
    ws.freeze_panes = 'A3'

    wb.save(output_path)
    print(f"\n✅ Planilha gerada: {output_path}")
    print(f"   {len(rows)} colaboradores")

    # Summary
    total_carga_final = sum(r['carga_final'] for r in rows)
    total_saldo_final = sum(r['saldo_final'] for r in rows)
    total_col_qz = sum(r['col_qz'] or 0 for r in rows)
    print(f"\n   Total CARGA FINAL: R$ {total_carga_final:,.2f}")
    print(f"   Total SALDO FINAL: R$ {total_saldo_final:,.2f}")
    print(f"   Total COL QZ: R$ {total_col_qz:,.2f}")

    return output_path


def main():
    parser = argparse.ArgumentParser(description='Gerar planilha Carga QZ do Neon')
    parser.add_argument('--year', type=int, required=True, help='Ano (ex: 2026)')
    parser.add_argument('--month', type=int, required=True, help='Mês (1-12)')
    parser.add_argument('--quinzena', type=int, required=True, choices=[1, 2], help='Quinzena (1 ou 2)')
    parser.add_argument('--output', type=Path, default=None, help='Caminho de saída (.xlsx)')
    args = parser.parse_args()

    if args.output:
        output_path = args.output
    else:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = Path(__file__).parent / "data" / f"carga_{args.year}_{args.month:02d}_q{args.quinzena}_{timestamp}.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    gerar_carga_neon(args.year, args.month, args.quinzena, output_path)


if __name__ == '__main__':
    main()
