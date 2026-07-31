#!/usr/bin/env python3
"""
Generate Excel with two tabs:
1. "BASE PREST COMPLETA" — Our complete BASE PREST from API data
2. "DIFERENÇAS" — Only the reports not in the reference BASE PREST
"""
import os, re
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

BASE = Path(__file__).parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"
OUT_PATH = BASE.parent / "data" / "BASE_PREST_COMPARACAO.xlsx"

# Load reference report IDs
wb_ref = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_bp = wb_ref["BASE PREST "]
ref_report_ids = set()
ref_report_names = {}
for row in ws_bp.iter_rows(min_row=4, values_only=True):
    if row[0] is not None and row[1] is not None:
        rid = int(row[1])
        ref_report_ids.add(rid)
        ref_report_names[rid] = str(row[2] or "")
wb_ref.close()

# Load all API expenses (Aprovado+Enviado, excluding FATURA/CARTAO)
conn = psycopg2.connect(NEON_URL, connect_timeout=10)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cur.execute("""
    SELECT e.id, e.report_id, e.date, e.value, e.description, e.status as expense_status,
           r.name as report_name, r.status as report_status,
           r.user_cpf, r.user_name, r.created_at as report_created_at
    FROM prestacao_expenses e
    JOIN prestacao_reports r ON e.report_id = r.id
    WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND r.user_cpf IS NOT NULL
      AND TRIM(r.name) !~* '^(fatu|farur|cart)'
      AND TRIM(r.name) !~* '(fatura|fatuar|fatut|farur)'
    ORDER BY r.user_name, r.name, e.date
""")
api_expenses = cur.fetchall()
conn.close()

print(f"API expenses (Aprovado+Enviado, excl FATURA): {len(api_expenses)}")
print(f"Reference report IDs: {len(ref_report_ids)}")

# Split into "in ref" and "not in ref"
in_ref = [e for e in api_expenses if int(e["report_id"]) in ref_report_ids]
not_in_ref = [e for e in api_expenses if int(e["report_id"]) not in ref_report_ids]

print(f"In reference: {len(in_ref)} expenses")
print(f"Not in reference: {len(not_in_ref)} expenses ({len(set(e['report_id'] for e in not_in_ref))} reports)")

# Create Excel
wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, size=9, name="Calibri")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=9, name="Calibri", color="FFFFFF")
data_font = Font(size=9, name="Calibri")
diff_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
new_report_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

headers = [
    "ID da Despesa", "ID do Relatório", "Nome do relatório", "Data",
    "Nome do membro de equipe", "CPF/CNPJ", "Status", "Descrição da despesa",
    "Valor", "Report Created At", "In Reference?"
]

# ============================================================
# Tab 1: BASE PREST COMPLETA
# ============================================================
ws1 = wb.active
ws1.title = "BASE PREST COMPLETA"

# Header
for col_idx, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col_idx, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Data
for row_idx, e in enumerate(api_expenses, 2):
    in_ref_flag = "SIM" if int(e["report_id"]) in ref_report_ids else "NÃO"
    values = [
        int(e["id"]),
        int(e["report_id"]),
        e["report_name"],
        e["date"],
        e["user_name"],
        e["user_cpf"],
        e["report_status"],
        e["description"],
        float(e["value"]),
        str(e["report_created_at"])[:10] if e["report_created_at"] else "",
        in_ref_flag,
    ]
    for col_idx, v in enumerate(values, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=v)
        cell.font = data_font
        cell.border = thin_border
        if col_idx == 9:  # Valor
            cell.number_format = '#,##0.00'
        if col_idx == 11 and in_ref_flag == "NÃO":
            cell.fill = new_report_fill

ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:K{len(api_expenses)+1}"

# Column widths
widths = [12, 12, 30, 12, 25, 15, 12, 30, 12, 14, 12]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# ============================================================
# Tab 2: DIFERENÇAS (only reports not in reference)
# ============================================================
ws2 = wb.create_sheet("DIFERENÇAS")

# Header
for col_idx, h in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Data — only expenses from reports NOT in reference
for row_idx, e in enumerate(not_in_ref, 2):
    values = [
        int(e["id"]),
        int(e["report_id"]),
        e["report_name"],
        e["date"],
        e["user_name"],
        e["user_cpf"],
        e["report_status"],
        e["description"],
        float(e["value"]),
        str(e["report_created_at"])[:10] if e["report_created_at"] else "",
        "NÃO",
    ]
    for col_idx, v in enumerate(values, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=v)
        cell.font = data_font
        cell.border = thin_border
        if col_idx == 9:
            cell.number_format = '#,##0.00'
        if col_idx == 11:
            cell.fill = diff_fill

ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:K{len(not_in_ref)+1}"

for i, w in enumerate(widths, 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Save
OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT_PATH)
print(f"\nFile saved: {OUT_PATH}")
print(f"  Tab 1 'BASE PREST COMPLETA': {len(api_expenses)} expenses")
print(f"  Tab 2 'DIFERENÇAS': {len(not_in_ref)} expenses from {len(set(e['report_id'] for e in not_in_ref))} reports")
