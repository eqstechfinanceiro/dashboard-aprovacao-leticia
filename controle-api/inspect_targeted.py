#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Targeted inspection of specific sheets and columns."""
import os
import openpyxl
import pyxlsb

ROOT = r'C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test'
DATA = os.path.join(ROOT, 'data')
OUT = os.path.join(ROOT, 'controle-api', 'inspect_targeted.txt')

def truncate(v, n=25):
    if v is None:
        return ''
    return str(v)[:n]

with open(OUT, 'w', encoding='utf-8') as fh:
    # 1. CONTROLE PAINEL - full columns (18-25)
    controle = os.path.join(DATA, 'CONTROLE - VEXPENSES - JULHO 2026.xlsb')
    fh.write("=== CONTROLE PAINEL (full columns) ===\n")
    with pyxlsb.open_workbook(controle) as wb:
        with wb.get_sheet('PAINEL') as ws:
            for i, row in enumerate(ws.rows()):
                if i >= 13:
                    break
                vals = [truncate(c.v, 25) for c in row[:26]]
                fh.write(f"  r{i+1}: {vals}\n")

    # 2. March 1QZ
    mar = os.path.join(DATA, '03 - MARÇO', '1 QZ MARÇO VEXPENSES 2026 (5).xlsx')
    fh.write(f"\n=== MAR 1QZ ===\n")
    wb = openpyxl.load_workbook(mar, read_only=True, data_only=True)
    for ws in wb.worksheets[:2]:
        fh.write(f"  sheet '{ws.title}' max_row={ws.max_row} max_col={ws.max_column}\n")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
            vals = [truncate(c, 22) for c in row[:22]]
            fh.write(f"    r{i+1}: {vals}\n")
    wb.close()

    # 3. Jan 2QZ - full columns
    jan2 = os.path.join(DATA, '01 - JANEIRO', '2QZ JANEIRO 2026 - VEXPENSES.xlsx')
    fh.write(f"\n=== JAN 2QZ (full cols) ===\n")
    wb = openpyxl.load_workbook(jan2, read_only=True, data_only=True)
    ws = wb['2 QZ VEXPENSES 01_2026']
    for i, row in enumerate(ws.iter_rows(min_row=5, max_row=8, values_only=True)):
        vals = [truncate(c, 22) for c in row[:25]]
        fh.write(f"    r{i+5}: {vals}\n")
    wb.close()

    # 4. Feb 2QZ - full columns
    feb2 = os.path.join(DATA, '02 - FEVEREIRO', '2QZ FEVEREIRO 2026 - VEXPENSES EQS.xlsx')
    fh.write(f"\n=== FEB 2QZ (full cols) ===\n")
    wb = openpyxl.load_workbook(feb2, read_only=True, data_only=True)
    ws = wb['2 QZ VEXPENSES 02_2026']
    for i, row in enumerate(ws.iter_rows(min_row=5, max_row=8, values_only=True)):
        vals = [truncate(c, 22) for c in row[:25]]
        fh.write(f"    r{i+5}: {vals}\n")
    wb.close()

    # 5. Apr 1QZ
    apr1 = os.path.join(DATA, '04 - ABRIL', '1QZ ABRIL 2026 - VEXPENSES.xlsx')
    fh.write(f"\n=== APR 1QZ ===\n")
    wb = openpyxl.load_workbook(apr1, read_only=True, data_only=True)
    ws = wb['1 QZ VEXPENSES 04_2026']
    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=8, values_only=True)):
        vals = [truncate(c, 22) for c in row[:22]]
        fh.write(f"    r{i+4}: {vals}\n")
    wb.close()

    # 6. Check multipliers in all sheets (cell row 4, various cols)
    fh.write(f"\n=== MULTIPLIERS (searching row 4) ===\n")
    sheets_to_check = [
        ('Jan 1QZ', os.path.join(DATA, '01 - JANEIRO', '1QZ JANEIRO 2026 - VEXPENSES.xlsx'), '1 QZ VEXPENSES 01_2026'),
        ('Jan 2QZ', os.path.join(DATA, '01 - JANEIRO', '2QZ JANEIRO 2026 - VEXPENSES.xlsx'), '2 QZ VEXPENSES 01_2026'),
        ('Feb 1QZ', os.path.join(DATA, '02 - FEVEREIRO', '1 QZN FEVEREIRO VEXPENSES 2026.xlsx'), '1 QZN FEV 2026'),
        ('Feb 2QZ', os.path.join(DATA, '02 - FEVEREIRO', '2QZ FEVEREIRO 2026 - VEXPENSES EQS.xlsx'), '2 QZ VEXPENSES 02_2026'),
        ('Mar 1QZ', os.path.join(DATA, '03 - MARÇO', '1 QZ MARÇO VEXPENSES 2026 (5).xlsx'), None),
        ('Apr 1QZ', os.path.join(DATA, '04 - ABRIL', '1QZ ABRIL 2026 - VEXPENSES.xlsx'), '1 QZ VEXPENSES 04_2026'),
        ('May 1QZ', os.path.join(DATA, '05 - MAIO', 'CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx'), 'Planilha1'),
        ('May 2QZ', os.path.join(DATA, '05 - MAIO', 'CARGA 2 QZ MAIO 26 VEXPENSES EQS.xlsx'), '2 QZ DE MAIO 26'),
        ('Jun 1QZ', os.path.join(DATA, '06 - JUNHO', 'CARGA 1 QZ JUNHO 26 VEXPENSES EQS.xlsx'), '1 QZ JUNHO'),
        ('Jun 2QZ', os.path.join(DATA, '06 - JUNHO', 'CARGA 2 QZ JUNHO 26 VEXPENSES EQS.xlsx'), '2 QZ JUNHO'),
    ]
    for label, path, sname in sheets_to_check:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            if sname is None:
                sname = wb.sheetnames[0]
            ws = wb[sname]
            # check rows 1-6 for any numeric value that looks like a multiplier (0.1-1.0)
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True)):
                for j, c in enumerate(row):
                    if c is not None and isinstance(c, (int, float)) and 0.1 <= float(c) <= 1.0:
                        fh.write(f"  {label}: r{i+1}c{j+1} = {c}\n")
            wb.close()
        except Exception as e:
            fh.write(f"  {label}: ERROR {e}\n")

print(f"WROTE {OUT}")
