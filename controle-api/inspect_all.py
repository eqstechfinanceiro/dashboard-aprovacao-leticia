#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Inspect all available Excel sheets (.xlsx via openpyxl, .xlsb via pyxlsb).
Writes findings to inspect_output.txt to avoid console encoding issues."""
import os
import sys
import openpyxl
import pyxlsb

ROOT = r'C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test'
DATA = os.path.join(ROOT, 'data')
OUT = os.path.join(ROOT, 'controle-api', 'inspect_output.txt')

def find_xlsx(base):
    out = []
    for dp, dn, fn in os.walk(base):
        for f in fn:
            if f.lower().endswith(('.xlsx', '.xlsb')) and not f.startswith('~$'):
                out.append(os.path.join(dp, f))
    return out

def truncate(v, n=22):
    if v is None:
        return ''
    s = str(v)
    return s[:n]

def inspect_xlsx(path, fh):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    fh.write(f"  SHEETS: {wb.sheetnames}\n")
    for ws in wb.worksheets[:5]:
        fh.write(f"  -- sheet '{ws.title}' max_row={ws.max_row} max_col={ws.max_column}\n")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True)):
            vals = [truncate(c) for c in row[:18]]
            fh.write(f"     r{i+1}: {vals}\n")
    wb.close()

def inspect_xlsb(path, fh):
    with pyxlsb.open_workbook(path) as wb:
        fh.write(f"  SHEETS: {wb.sheets}\n")
        for sname in wb.sheets[:5]:
            with wb.get_sheet(sname) as ws:
                fh.write(f"  -- sheet '{sname}' (first 12 rows):\n")
                for i, row in enumerate(ws.rows()):
                    if i >= 12:
                        break
                    vals = [truncate(c.v) for c in row[:18]]
                    fh.write(f"     r{i+1}: {vals}\n")

files = sorted(find_xlsx(DATA))
with open(OUT, 'w', encoding='utf-8') as fh:
    fh.write(f"=== SHEET INSPECTION ===\nFound {len(files)} files\n\n")
    for f in files:
        sz = os.path.getsize(f)
        rel = os.path.relpath(f, DATA)
        fh.write(f"\n{'='*80}\nFILE: {rel}\nSIZE: {sz} bytes\n")
        if sz < 1000:
            fh.write("  -> TOO SMALL (likely corrupted)\n")
            continue
        try:
            if f.lower().endswith('.xlsb'):
                inspect_xlsb(f, fh)
            else:
                inspect_xlsx(f, fh)
        except Exception as e:
            fh.write(f"  ERROR: {e}\n")
print(f"WROTE {OUT}")
