#!/usr/bin/env python3
"""
Corrected investigation with proper EXTRATO column indices.
EXTRATO columns (0-indexed): 0=empty, 1=ANO, 2=MES, 3=Data, 4=Hora, 5=CodTrans, 
6=NumCartao, 7=Grupo, 8=Usuario, 9=Tipo, 10=Descricao, 11=Valor, 12=CPF
"""
import os, re
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from pathlib import Path
import openpyxl
from collections import defaultdict

BASE = Path(__file__).parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"

conn = psycopg2.connect(NEON_URL, connect_timeout=10)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_ext = wb["EXTRATO"]

# ============================================================
# PART 1: TARIFA — correct column indices
# ============================================================
print("=" * 80)
print("PART 1: TARIFA comparison (corrected columns)")
print("=" * 80)

# Build ref tarifa by CPF (col 9=Tipo, col 11=Valor, col 12=CPF, col 8=Usuario)
ref_tarifa_by_cpf = defaultdict(list)
ref_tarifa_by_name = defaultdict(list)
for row in ws_ext.iter_rows(min_row=9, values_only=True):
    if row[9] is None:
        continue
    tipo = str(row[9]).strip().upper()
    if tipo == "TARIFA":
        cpf = str(row[12] or "").strip().replace(".", "").replace("-", "").replace(" ", "").zfill(11) if row[12] else ""
        usuario = str(row[8] or "").strip()
        valor = float(row[11] or 0)
        desc = str(row[10] or "")
        data = row[3]
        hora = str(row[4] or "")
        ref_tarifa_by_cpf[cpf].append({"valor": valor, "desc": desc, "data": data, "hora": hora, "usuario": usuario})
        ref_tarifa_by_name[usuario.upper()].append({"valor": valor, "desc": desc, "data": data, "hora": hora, "cpf": cpf})

# Test CPFs with divergences
test_cpfs = [
    ("02027745203", "ABNER ANDRADE CAVALCANTE", +14.00),
    ("11178519740", "GUILHERME MOTTA RIBEIRO SILVA", +28.00),
    ("01050938232", "CHARLYTON COSTA ANDRADE", -29.93),
    ("06576198922", "LUDGERO HORACIO DE OLIVEIRA", +35.00),
    ("01677920599", "RAFAEL AMORIM VELLO", +14.00),
]

for cpf, name, diff in test_cpfs:
    ref_txns = ref_tarifa_by_cpf.get(cpf, [])
    ref_total = sum(abs(t["valor"]) for t in ref_txns)
    
    # Get DB total for this person
    cur.execute("""
        WITH deduped AS (
            SELECT DISTINCT ON (
                UPPER(usuario), data, tipo, valor,
                COALESCE(NULLIF(codigo_transacao, ''), hora::text)
            )
                UPPER(usuario) AS usuario_up, data, tipo, valor, descricao, hora
            FROM extrato_movimentacao
            WHERE is_snapshot = FALSE
              AND tipo = 'Taxa'
              AND data <= '2026-06-30'
              AND UPPER(usuario) = UPPER(%s)
            ORDER BY UPPER(usuario), data, tipo, valor,
                COALESCE(NULLIF(codigo_transacao, ''), hora::text)
        )
        SELECT COUNT(*) as cnt, SUM(ABS(valor)) as total,
               MAX(data) as max_date, MIN(data) as min_date
        FROM deduped
    """, (name,))
    r = cur.fetchone()
    
    print(f"\n  {cpf} {name[:30]} (diff={diff:+.2f})")
    print(f"    Ref:  {len(ref_txns):>3} txns, total R$ {ref_total:>10.2f}, dates {ref_txns[0]['data'] if ref_txns else '?'} to {ref_txns[-1]['data'] if ref_txns else '?'}")
    print(f"    DB:   {r['cnt']:>3} txns, total R$ {float(r['total']):>10.2f}, dates {r['min_date']} to {r['max_date']}")
    
    if abs(ref_total - float(r["total"])) > 0.01 or len(ref_txns) != r["cnt"]:
        # Show last few from each to find the mismatch
        print(f"    --- Last 5 ref txns ---")
        for t in ref_txns[-5:]:
            print(f"      {t['data']} {t['hora']:<10} R$ {t['valor']:>8.2f} {t['desc'][:25]}")
        
        cur.execute("""
            WITH deduped AS (
                SELECT DISTINCT ON (
                    UPPER(usuario), data, tipo, valor,
                    COALESCE(NULLIF(codigo_transacao, ''), hora::text)
                )
                    data, valor, descricao, hora
                FROM extrato_movimentacao
                WHERE is_snapshot = FALSE
                  AND tipo = 'Taxa'
                  AND data <= '2026-06-30'
                  AND UPPER(usuario) = UPPER(%s)
                ORDER BY UPPER(usuario), data, tipo, valor,
                    COALESCE(NULLIF(codigo_transacao, ''), hora::text)
            )
            SELECT data, valor, descricao, hora
            FROM deduped
            ORDER BY data, hora
            LIMIT 5
            OFFSET (SELECT COUNT(*) FROM deduped) - 5
        """, (name,))
        print(f"    --- Last 5 DB txns ---")
        for t in cur.fetchall():
            print(f"      {t['data']} {str(t['hora']):<10} R$ {float(t['valor']):>8.2f} {str(t['descricao'] or '')[:25]}")

# ============================================================
# PART 2: "Fartura" exclusion check
# ============================================================
print("\n" + "=" * 80)
print("PART 2: FATURA/FARTURA exclusion check")
print("=" * 80)

cur.execute("""
    SELECT r.id, r.name, r.status, r.user_name,
           COUNT(e.id) as expense_count, SUM(e.value) as total
    FROM prestacao_reports r
    JOIN prestacao_expenses e ON e.report_id = r.id
    WHERE TRIM(r.name) ILIKE 'fat%' OR TRIM(r.name) ILIKE 'cart%'
       OR TRIM(r.name) ILIKE 'fart%'
    GROUP BY r.id, r.name, r.status, r.user_name
    ORDER BY r.name
""")
fatura_reports = cur.fetchall()
print("FATURA/CARTAO/FARTURA reports:")
for r in fatura_reports:
    name = str(r["name"]).strip()
    # Our current regex
    matches_current = bool(re.match(r'^(fatu|farur|cart)', name, re.IGNORECASE) or re.search(r'(fatura|fatuar|fatut|farur)', name, re.IGNORECASE))
    # Should "fartur" be added?
    matches_fartur = bool(re.search(r'fartur', name, re.IGNORECASE))
    print(f"  {r['id']}  {name:<35}  {r['status']:<10}  {r['expense_count']:>3} items  R$ {float(r['total']):>10,.2f}  excluded={matches_current}  fartur={matches_fartur}")

# ============================================================
# PART 3: PRESTACAO — categorize new reports by date
# ============================================================
print("\n" + "=" * 80)
print("PART 3: PRESTACAO — new reports by expense date range")
print("=" * 80)

ws_bp = wb["BASE PREST "]
ref_report_ids = set()
for row in ws_bp.iter_rows(min_row=4, values_only=True):
    if row[0] is not None and row[1] is not None:
        ref_report_ids.add(int(row[1]))

cur.execute("""
    SELECT r.id, r.name, r.status, r.user_cpf, r.user_name,
           MIN(e.date) as min_date, MAX(e.date) as max_date,
           COUNT(e.id) as expense_count, SUM(e.value) as total_value
    FROM prestacao_reports r
    JOIN prestacao_expenses e ON e.report_id = r.id
    WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND r.user_cpf IS NOT NULL
      AND TRIM(r.name) !~* '^(fatu|farur|cart)'
      AND TRIM(r.name) !~* '(fatura|fatuar|fatut|farur)'
    GROUP BY r.id, r.name, r.status, r.user_cpf, r.user_name
""")
all_api_reports = cur.fetchall()
new_reports = [r for r in all_api_reports if int(r["id"]) not in ref_report_ids]

# Categorize by max_date
pre_cutoff = [r for r in new_reports if r["max_date"] and str(r["max_date"]) <= "2026-06-30"]
post_cutoff = [r for r in new_reports if r["max_date"] and str(r["max_date"]) > "2026-06-30"]

print(f"New reports with ALL expenses up to 2026-06-30: {len(pre_cutoff)} reports, R$ {sum(float(r['total_value']) for r in pre_cutoff):,.2f}")
print(f"New reports with expenses AFTER 2026-06-30: {len(post_cutoff)} reports, R$ {sum(float(r['total_value']) for r in post_cutoff):,.2f}")

print(f"\nPre-cutoff new reports (should be in ref but aren't):")
for r in sorted(pre_cutoff, key=lambda x: x["min_date"] or "9999"):
    print(f"  {r['id']}  {str(r['name'])[:35]:<35}  {r['status']:<10}  {str(r['user_name'])[:25]:<25}  {r['expense_count']:>3} items  R$ {float(r['total_value']):>10,.2f}  dates={r['min_date']} to {r['max_date']}")

print(f"\nPost-cutoff new reports (expected to be new):")
for r in sorted(post_cutoff, key=lambda x: x["min_date"] or "9999"):
    print(f"  {r['id']}  {str(r['name'])[:35]:<35}  {r['status']:<10}  {str(r['user_name'])[:25]:<25}  {r['expense_count']:>3} items  R$ {float(r['total_value']):>10,.2f}  dates={r['min_date']} to {r['max_date']}")

# ============================================================
# PART 4: Check if "Fartura" report is in reference BASE PREST
# ============================================================
print("\n" + "=" * 80)
print("PART 4: Is 'Fartura 05/2026' (report 10383163) in reference BASE PREST?")
print("=" * 80)

found_in_ref = False
for row in ws_bp.iter_rows(min_row=4, values_only=True):
    if row[1] is not None and int(row[1]) == 10383163:
        found_in_ref = True
        print(f"  FOUND in ref: expense_id={row[0]}, report_id={row[1]}, report_name={row[2]}")
        break
if not found_in_ref:
    print("  NOT found in reference BASE PREST")
    # Check if it's under a different name
    for row in ws_bp.iter_rows(min_row=4, values_only=True):
        if row[2] is not None and 'fartura' in str(row[2]).lower():
            print(f"  Found 'fartura' in ref: expense_id={row[0]}, report_name={row[2]}")
            break

wb.close()
conn.close()
