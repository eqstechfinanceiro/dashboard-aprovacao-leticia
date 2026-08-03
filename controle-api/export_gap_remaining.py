"""Export the remaining gap expenses (new in API, not in reference) with PDF links."""
import os, psycopg2, psycopg2.extras, openpyxl
from dotenv import load_dotenv
from pathlib import Path
from collections import defaultdict
from openpyxl.styles import Font, PatternFill

load_dotenv(Path(__file__).parent / ".env")
conn = psycopg2.connect(os.getenv("NEON_DATABASE_URL"))
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"

# Load reference expense IDs
wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_bp = wb["BASE PREST "]
ws_p = wb["PAINEL"]

vexpenses_cpfs = set()
for row in ws_p.iter_rows(min_row=12, values_only=True):
    if row[2] is None:
        continue
    cpf = str(row[2] or "").strip()
    cartao_vx = str(row[12] or "").strip().upper() if len(row) > 12 else ""
    if cartao_vx == "SIM":
        vexpenses_cpfs.add(cpf)

ref_expense_ids = set()
ref_report_ids = set()
for row in ws_bp.iter_rows(min_row=4, values_only=True):
    if row[0] is None:
        continue
    eid = int(row[0])
    rid = int(row[1]) if row[1] else None
    ref_expense_ids.add(eid)
    if rid:
        ref_report_ids.add(rid)
wb.close()

def is_card_report(name):
    n = name.strip().upper()
    if not n:
        return False
    if 'CAIXA ITAU' in n or 'CAIXA ITAÚ' in n:
        return True
    if n.startswith('CAIXA'):
        return False
    if n.startswith(('FATURA', 'CARTAO', 'CARTÃO', 'FATUAR', 'FARTUR', 'FATUT', 'FARUR', 'FATUTR')):
        return True
    if 'CARTÃO DE CRÉDITO' in n or 'CARTAO DE CREDITO' in n or 'CARTÃO DE CREDITO' in n:
        return True
    if 'CARTÃO CORPORATIVO' in n:
        return True
    if ('ITAU' in n or 'ITAÚ' in n) and 'CAIXA' not in n:
        return True
    if 'DOLAR' in n or 'DÓLAR' in n:
        return True
    if n.startswith('DESPESA') and 'FATURA' in n:
        return True
    if n.startswith('COMPLEMENTAR') and 'FATURA' in n:
        return True
    if 'CARTÃO' in n and 'CRÉDITO' in n:
        return True
    if 'CARTAO' in n and 'CREDITO' in n:
        return True
    if n.startswith('CARTÃO VEXPENSES'):
        return True
    return False

cur.execute("""
    SELECT e.id, e.report_id, e.value, e.description, e.date,
           r.name as report_name, r.user_cpf, r.user_name, r.status as report_status,
           r.raw_data,
           e.raw_data->>'payment_method_id' as pm_id,
           e.raw_data->>'payment_method_name' as pm_name
    FROM prestacao_expenses e
    JOIN prestacao_reports r ON e.report_id = r.id
    WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND r.user_cpf IS NOT NULL
    ORDER BY e.value DESC
""")
all_api = cur.fetchall()
conn.close()

# Apply same filter as export_gap_analysis
filtered = []
for e in all_api:
    name = e["report_name"] or ""
    if is_card_report(name):
        continue
    if e.get("pm_id") == "627401":
        continue
    if e["user_cpf"] not in vexpenses_cpfs:
        continue
    filtered.append(e)

# Gap expenses: in API filtered but not in reference
new_expenses = [e for e in filtered if e["id"] not in ref_expense_ids]
missing_expense_ids = ref_expense_ids - set(e["id"] for e in filtered)

print(f"API filtered: {len(filtered)} expenses")
print(f"New expenses (API only): {len(new_expenses)} totaling R$ {sum(float(e['value']) for e in new_expenses):,.2f}")
print(f"Missing expenses (Ref only): {len(missing_expense_ids)}")

# Build Excel
wb_out = openpyxl.Workbook()
hf = Font(bold=True, color="FFFFFF", size=10)
hfill = PatternFill("solid", fgColor="1F4E79")
money_fmt = '#,##0.00;[Red]-#,##0.00'

# --- Sheet 1: New Expenses (API only) ---
ws1 = wb_out.active
ws1.title = "New Expenses (API only)"
headers = ["Expense ID", "Report ID", "Report Name", "User Name", "CPF", "Date",
           "Value R$", "Description", "PM ID", "PM Name", "Report Status",
           "PDF Link", "In Ref Reports?"]
for ci, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=ci, value=h)
    cell.font = hf
    cell.fill = hfill

for ri, e in enumerate(sorted(new_expenses, key=lambda x: abs(float(x["value"])), reverse=True), 2):
    raw = e.get("raw_data") or {}
    if isinstance(raw, str):
        import json as _json
        raw = _json.loads(raw)
    pdf_link = raw.get("pdf_link", "") if raw else ""
    in_ref = "YES" if e["report_id"] in ref_report_ids else "NO"
    vals = [e["id"], e["report_id"], e["report_name"], e["user_name"], e["user_cpf"],
            str(e["date"] or ""), float(e["value"]), e["description"] or "",
            e.get("pm_id") or "", e.get("pm_name") or "", e["report_status"],
            pdf_link, in_ref]
    for ci, v in enumerate(vals, 1):
        cell = ws1.cell(row=ri, column=ci, value=v)
        if ci == 7:
            cell.number_format = money_fmt
        if ci == 12 and pdf_link:
            cell.hyperlink = pdf_link
            cell.font = Font(color="0563C1", underline="single")

for col in ws1.columns:
    max_len = max(len(str(c.value or "")) for c in col)
    ws1.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

# --- Sheet 2: Summary by PM ---
ws2 = wb_out.create_sheet("Summary by PM")
from collections import Counter
pm_counts = Counter()
pm_totals = defaultdict(float)
for e in new_expenses:
    pm = e.get("pm_id") or "NULL"
    pm_counts[pm] += 1
    pm_totals[pm] += float(e["value"])

ws2.cell(row=1, column=1, value="PM ID").font = hf
ws2.cell(row=1, column=2, value="PM Name").font = hf
ws2.cell(row=1, column=3, value="Count").font = hf
ws2.cell(row=1, column=4, value="Total R$").font = hf
for c in range(1, 5):
    ws2.cell(row=1, column=c).fill = hfill

PM_NAMES = {"627401": "Cartão Corporativo Itaú", "627508": "Cartão VExpenses",
            "627721": "Saque VExpenses", "668240": "Pix VExpenses",
            "630113": "Recurso Próprio", "627726": "Desconto de Rescisão",
            "627741": "Tarifa de Saque"}
for ri, (pm, cnt) in enumerate(pm_counts.most_common(), 2):
    ws2.cell(row=ri, column=1, value=pm)
    ws2.cell(row=ri, column=2, value=PM_NAMES.get(pm, ""))
    ws2.cell(row=ri, column=3, value=cnt)
    cell = ws2.cell(row=ri, column=4, value=pm_totals[pm])
    cell.number_format = money_fmt

# --- Sheet 3: Summary by Report ---
ws3 = wb_out.create_sheet("Summary by Report")
report_data = defaultdict(lambda: {"count": 0, "total": 0.0, "name": "", "user": "", "pdf": ""})
for e in new_expenses:
    rid = e["report_id"]
    report_data[rid]["count"] += 1
    report_data[rid]["total"] += float(e["value"])
    report_data[rid]["name"] = e["report_name"]
    report_data[rid]["user"] = e["user_name"]
    raw = e.get("raw_data") or {}
    if isinstance(raw, str):
        import json as _json
        raw = _json.loads(raw)
    if not report_data[rid]["pdf"]:
        report_data[rid]["pdf"] = raw.get("pdf_link", "") if raw else ""

ws3.cell(row=1, column=1, value="Report ID").font = hf
ws3.cell(row=1, column=2, value="Report Name").font = hf
ws3.cell(row=1, column=3, value="User").font = hf
ws3.cell(row=1, column=4, value="Expenses").font = hf
ws3.cell(row=1, column=5, value="Total R$").font = hf
ws3.cell(row=1, column=6, value="In Ref?").font = hf
ws3.cell(row=1, column=7, value="PDF Link").font = hf
for c in range(1, 8):
    ws3.cell(row=1, column=c).fill = hfill

for ri, (rid, r) in enumerate(sorted(report_data.items(), key=lambda x: x[1]["total"], reverse=True), 2):
    in_ref = "YES" if rid in ref_report_ids else "NO"
    vals = [rid, r["name"], r["user"], r["count"], r["total"], in_ref, r["pdf"]]
    for ci, v in enumerate(vals, 1):
        cell = ws3.cell(row=ri, column=ci, value=v)
        if ci == 5:
            cell.number_format = money_fmt
        if ci == 7 and r["pdf"]:
            cell.hyperlink = r["pdf"]
            cell.font = Font(color="0563C1", underline="single")

for col in ws3.columns:
    max_len = max(len(str(c.value or "")) for c in col)
    ws3.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

out = Path(__file__).parent.parent / "data" / "GAP_REMAINING_EXPENSES.xlsx"
out.parent.mkdir(exist_ok=True)
wb_out.save(out)
print(f"\nSaved to {out}")
