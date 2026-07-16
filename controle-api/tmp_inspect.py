import openpyxl

def parse(v):
    if v is None:
        return 0.0
    if isinstance(v, str):
        v = v.strip()
        if v == '':
            return 0.0
    return float(v)

def load_excel_carga(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[1]:
            continue
        cpf = str(r[1]).replace('.', '').replace('-', '').zfill(11)
        rows[cpf] = {
            'colaborador': r[0],
            'saldo_reembolsar': parse(r[7]),
            'saldo_final': parse(r[8]),
            'col_qz': parse(r[9]),
            'saldo_cartao': parse(r[10]),
            'carga_parcial': parse(r[12]),
            'reembolso': parse(r[13]),
            'carga_final': parse(r[14]),
        }
    return rows

def load_excel_ref(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = {}
    for r in ws.iter_rows(min_row=7, values_only=True):
        if not r[1]:
            continue
        cpf = str(r[1]).replace('.', '').replace('-', '').zfill(11)
        rows[cpf] = {
            'colaborador': r[0],
            'saldo_reembolsar': parse(r[7]),
            'saldo_final': parse(r[8]),
            'col_qz': parse(r[9]),
            'saldo_cartao': parse(r[10]),
            'adiantamento': parse(r[11]),
            'carga_parcial': parse(r[12]),
            'reembolso': parse(r[13]),
            'carga_final': parse(r[14]),
        }
    return rows

ref = load_excel_ref(r'..\data\06 - JUNHO\CARGA 1 QZ JUNHO 26 VEXPENSES EQS.xlsx')
calc = load_excel_carga('data/carga_1qz_junho_2026.xlsx')

print(f"Ref: {len(ref)} CPFs | Calc: {len(calc)} CPFs")
common = set(ref.keys()) & set(calc.keys())
print(f"Common: {len(common)}")

diffs = []
for cpf in common:
    c = calc[cpf]
    r = ref[cpf]
    # Recalcula os valores vivos da CARGA a partir das colunas de entrada
    # (a coluna Carga Final da planilha pode estar stale)
    r_cp = max(r['col_qz'] - r['saldo_final'] - r['saldo_cartao'] - r['adiantamento'], 0.0)
    r_reem = round(r['saldo_reembolsar'] * 0.6, 2)
    r_cf = r_cp + r_reem
    if (abs(c['saldo_final'] - r['saldo_final']) > 0.01 or
        abs(c['saldo_reembolsar'] - r['saldo_reembolsar']) > 0.01 or
        abs(c['col_qz'] - r['col_qz']) > 0.01 or
        abs(c['saldo_cartao'] - r['saldo_cartao']) > 0.01 or
        abs(c['carga_parcial'] - r_cp) > 0.01 or
        abs(c['reembolso'] - r_reem) > 0.01 or
        abs(c['carga_final'] - r_cf) > 0.01):
        diffs.append({
            'cpf': cpf,
            'nome': r['colaborador'],
            'c_sf': c['saldo_final'], 'r_sf': r['saldo_final'],
            'c_sr': c['saldo_reembolsar'], 'r_sr': r['saldo_reembolsar'],
            'c_qz': c['col_qz'], 'r_qz': r['col_qz'],
            'c_sc': c['saldo_cartao'], 'r_sc': r['saldo_cartao'],
            'c_cp': c['carga_parcial'], 'r_cp': r_cp,
            'c_reem': c['reembolso'], 'r_reem': r_reem,
            'c_cf': c['carga_final'], 'r_cf': r_cf,
        })

print(f"Divergencias: {len(diffs)}")
for d in diffs[:10]:
    print(d)
print('...')
for d in diffs[-5:]:
    print(d)

total_c = sum(c['carga_final'] for c in calc.values() if c['carga_final'])
total_r = sum(r['carga_final'] for r in ref.values() if r['carga_final'])
print(f"Total CARGA FINAL calc: R$ {total_c:,.2f}")
print(f"Total CARGA FINAL ref:  R$ {total_r:,.2f}")
print(f"Diferenca: R$ {total_c - total_r:,.2f}")

