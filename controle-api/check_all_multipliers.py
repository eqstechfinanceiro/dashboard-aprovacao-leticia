import openpyxl, os
ROOT = r'C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test'

# Check all January sheets for the reembolso multiplier in Row 4 Col O
files = [
    ("Jan QZ1", os.path.join(ROOT, 'data', '01 - JANEIRO', '1QZ JANEIRO 2026 - VEXPENSES.xlsx'), "1 QZ VEXPENSES 01_2026"),
    ("Jan QZ2", os.path.join(ROOT, 'data', '01 - JANEIRO', '2QZ JANEIRO 2026 - VEXPENSES.xlsx'), "2 QZ VEXPENSES 01_2026"),
    ("Feb QZ1", os.path.join(ROOT, 'data', '02 - FEVEREIRO', '1 QZN FEVEREIRO VEXPENSES 2026.xlsx'), "1 QZN FEV 2026"),
    ("Feb QZ2", os.path.join(ROOT, 'data', '02 - FEVEREIRO', '2QZ FEVEREIRO 2026 - VEXPENSES EQS.xlsx'), "2 QZ VEXPENSES 02_2026"),
    ("Mar QZ1", os.path.join(ROOT, 'data', '03 - MARÇO', '1 QZ MARÇO VEXPENSES 2026 (5).xlsx'), "QUINZENA MARÇO"),
    ("Apr QZ1", os.path.join(ROOT, 'data', '04 - ABRIL', '1QZ ABRIL 2026 - VEXPENSES.xlsx'), "1 QZ VEXPENSES 04_2026"),
    ("May QZ1", os.path.join(ROOT, 'data', '05 - MAIO', 'CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx'), "Planilha1"),
    ("May QZ2", os.path.join(ROOT, 'data', '05 - MAIO', 'CARGA 2 QZ MAIO 26 VEXPENSES EQS.xlsx'), "2 QZ DE MAIO 26"),
    ("Jun QZ1", os.path.join(ROOT, 'data', '06 - JUNHO', 'CARGA 1 QZ JUNHO 26 VEXPENSES EQS.xlsx'), "1 QZ JUNHO"),
    ("Jun QZ2", os.path.join(ROOT, 'data', '06 - JUNHO', 'CARGA 2 QZ JUNHO 26 VEXPENSES EQS.xlsx'), "2 QZ JUNHO"),
]

for label, path, sheet_name in files:
    if not os.path.exists(path):
        print(f"{label}: FILE NOT FOUND")
        continue
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"{label}: SHEET '{sheet_name}' NOT FOUND, available: {wb.sheetnames}")
        wb.close()
        continue
    ws = wb[sheet_name]
    
    # Check row 4 for multiplier (Col O = index 14)
    row4 = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))
    multiplier = row4[0][14] if row4 and len(row4[0]) > 14 else None
    
    # Check row 6 for headers
    row6 = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))
    headers = row6[0] if row6 else []
    
    # Find reembolso and saldo_reembolsar columns from headers
    reembolso_col = None
    sr_col = None
    for j, h in enumerate(headers):
        if h and 'REEMBOLSO' in str(h).upper() and 'SEM' not in str(h).upper() and 'FINAL' not in str(h).upper():
            reembolso_col = j
        if h and 'SALDO REEMBOLSAR' in str(h).upper():
            sr_col = j
    
    # Sample users with non-zero reembolso
    samples = []
    for i, row in enumerate(ws.iter_rows(min_row=7, values_only=True)):
        if len(samples) >= 5:
            break
        if reembolso_col is not None and sr_col is not None:
            re = row[reembolso_col] if reembolso_col < len(row) else None
            sr = row[sr_col] if sr_col < len(row) else None
            if re and sr and float(sr) != 0 and float(re) != 0:
                ratio = float(re) / abs(float(sr))
                samples.append(f"  SR={float(sr):.2f}, RE={float(re):.2f}, ratio={ratio:.4f}")
    
    print(f"\n{label}:")
    print(f"  Row4 Col O (multiplier): {multiplier}")
    print(f"  SR col={sr_col}, RE col={reembolso_col}")
    for s in samples:
        print(s)
    
    wb.close()
