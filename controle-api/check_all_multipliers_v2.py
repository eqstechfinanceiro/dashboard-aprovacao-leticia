import openpyxl, os
ROOT = r'C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test'
OUT = os.path.join(ROOT, 'controle-api', 'all_multipliers.txt')
f = open(OUT, 'w', encoding='utf-8')

# All QZ1 sheets with their config
sheets = [
    ("Jan QZ1", os.path.join(ROOT, 'data', '01 - JANEIRO', '1QZ JANEIRO 2026 - VEXPENSES.xlsx'), "1 QZ VEXPENSES 01_2026", 6, 8, 14),
    ("Feb QZ1", os.path.join(ROOT, 'data', '02 - FEVEREIRO', '1 QZN FEVEREIRO VEXPENSES 2026.xlsx'), "1 QZN FEV 2026", 6, 8, 14),
    ("Mar QZ1", os.path.join(ROOT, 'data', '03 - MARÇO', '1 QZ MARÇO VEXPENSES 2026 (5).xlsx'), "QUINZENA MARÇO", 6, 10, 17),
    ("Apr QZ1", os.path.join(ROOT, 'data', '04 - ABRIL', '1QZ ABRIL 2026 - VEXPENSES.xlsx'), "1 QZ VEXPENSES 04_2026", 6, 9, 15),
    ("May QZ1", os.path.join(ROOT, 'data', '05 - MAIO', 'CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx'), "Planilha1", 6, 7, 13),
    ("Jun QZ1", os.path.join(ROOT, 'data', '06 - JUNHO', 'CARGA 1 QZ JUNHO 26 VEXPENSES EQS.xlsx'), "1 QZ JUNHO", 6, 7, 13),
]

for label, path, sheet_name, header_row, sr_col, re_col in sheets:
    if not os.path.exists(path):
        f.write(f"{label}: FILE NOT FOUND\n")
        continue
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        f.write(f"{label}: SHEET NOT FOUND\n")
        wb.close()
        continue
    ws = wb[sheet_name]
    
    # Check row 4 for multiplier (Col O = index 14)
    row4 = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))
    multiplier_cell = row4[0][14] if row4 and len(row4[0]) > 14 else None
    
    # Collect all users with non-zero SR and RE
    ratios = []
    for row in ws.iter_rows(min_row=header_row+1, values_only=True):
        sr = row[sr_col] if sr_col < len(row) else None
        re = row[re_col] if re_col < len(row) else None
        if sr and re and float(sr) != 0 and float(re) != 0:
            ratio = float(re) / abs(float(sr))
            ratios.append(ratio)
    
    if ratios:
        # Find the most common ratio
        from collections import Counter
        rounded = [round(r, 4) for r in ratios]
        most_common = Counter(rounded).most_common(3)
        f.write(f"{label}: cell={multiplier_cell}, users={len(ratios)}, most_common={most_common}\n")
    else:
        f.write(f"{label}: cell={multiplier_cell}, no users with non-zero SR+RE\n")
    
    wb.close()

f.close()
print("Done: " + OUT)
