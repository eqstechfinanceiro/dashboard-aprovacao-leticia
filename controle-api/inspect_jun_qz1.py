import openpyxl, os
DATA = r'C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\data'
path = os.path.join(DATA, "06 - JUNHO", "CARGA 1 QZ JUNHO 26 VEXPENSES EQS.xlsx")
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
ws = wb["1 QZ JUNHO"]

# Check specific CPFs
target_cpfs = {"29758827855", "07690263623", "09276240446", "99114160030"}
# Also check a matching user
# First find a user that matches (col_qz > 0, reembolso = 0)

header = None
for i, row in enumerate(ws.iter_rows(min_row=6, max_row=6, values_only=True)):
    header = row
    print("HEADER:")
    for j, v in enumerate(row):
        if v: print(f"  col {j}: {v}")

print("\n--- Target users (mismatches) ---")
for i, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
    cpf_raw = row[1] if len(row) > 1 else None
    if not cpf_raw: continue
    cpf = str(cpf_raw).strip().replace(".","").replace("-","").zfill(11)
    if cpf in target_cpfs:
        print(f"\nRow {i}: CPF={cpf}")
        print(f"  COLABORADOR (col 0): {row[0]}")
        print(f"  SALDO REEMBOLSAR (col 7): {row[7]}")
        print(f"  SALDO FINAL (col 8): {row[8]}")
        print(f"  1QZ (col 9): {row[9]}")
        print(f"  SALDO CARTAO (col 10): {row[10]}")
        print(f"  ADIANTAMENTO (col 11): {row[11]}")
        print(f"  CARGA PARCIAL (col 12): {row[12]}")
        print(f"  REEMBOLSO (col 13): {row[13]}")
        print(f"  Carga Final (col 14): {row[14]}")
        print(f"  STATUS CARTAO (col 16): {row[16]}")

# Find users with col_qz > 0 and reembolso > 0 (should match)
print("\n--- Users with col_qz > 0 AND reembolso > 0 ---")
count = 0
for i, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
    cpf_raw = row[1] if len(row) > 1 else None
    if not cpf_raw: continue
    col_qz = row[9]
    reembolso = row[13]
    if col_qz and col_qz > 0 and reembolso and reembolso > 0:
        print(f"\nRow {i}: CPF={str(cpf_raw).strip()}")
        print(f"  COLABORADOR: {row[0]}")
        print(f"  SALDO FINAL: {row[8]}")
        print(f"  1QZ (col_qz): {col_qz}")
        print(f"  CARGA PARCIAL: {row[12]}")
        print(f"  REEMBOLSO: {reembolso}")
        print(f"  Carga Final: {row[14]}")
        count += 1
        if count >= 5: break

# Find users with col_qz = 0 and reembolso > 0
print("\n--- Users with col_qz = 0 AND reembolso > 0 ---")
count = 0
for i, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
    cpf_raw = row[1] if len(row) > 1 else None
    if not cpf_raw: continue
    col_qz = row[9]
    reembolso = row[13]
    if (col_qz is None or col_qz == 0) and reembolso and reembolso > 0:
        print(f"\nRow {i}: CPF={str(cpf_raw).strip()}")
        print(f"  COLABORADOR: {row[0]}")
        print(f"  SALDO REEMBOLSAR: {row[7]}")
        print(f"  SALDO FINAL: {row[8]}")
        print(f"  1QZ (col_qz): {col_qz}")
        print(f"  CARGA PARCIAL: {row[12]}")
        print(f"  REEMBOLSO: {reembolso}")
        print(f"  Carga Final: {row[14]}")
        count += 1
        if count >= 5: break

wb.close()
