#!/usr/bin/env python3
"""
Investigate root causes:
1. TARIFA ±R$7 pattern (dedup issue with Taxa de saque?)
2. PRESTACAO gap (compare reference BASE PREST vs API data expense-by-expense)
3. CARGA/TRANSF CPF resolution issues
"""
import os, re
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from pathlib import Path
import openpyxl
from collections import defaultdict

BASE = Path(__file__).parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"

# ============================================================
# PART 1: Investigate TARIFA ±R$7 pattern
# ============================================================
print("=" * 80)
print("PART 1: TARIFA ±R$7 INVESTIGATION")
print("=" * 80)

# Pick a few CPFs with +R$7 tarifa diff and check their Taxa transactions
tarifa_div_cpfs = [
    ("02027745203", "ABNER", +14.00),
    ("01677920599", "RAFAEL VELLO", +14.00),
    ("06576198922", "LUDGERO", +35.00),
    ("01050938232", "CHARLYTON", -29.93),
    ("11178519740", "GUILHERME", +28.00),
]

conn = psycopg2.connect(NEON_URL, connect_timeout=10)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# Load reference EXTRATO Taxa transactions by CPF
wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_ext = wb["EXTRATO"]

ref_tarifa_by_cpf = defaultdict(list)
for row in ws_ext.iter_rows(min_row=9, values_only=True):
    if row[8] is None:
        continue
    tipo = str(row[8]).strip().upper()
    if tipo == "TARIFA":
        cpf = str(row[11] or "").strip().replace(".", "").replace("-", "").replace(" ", "").zfill(11) if row[11] else ""
        valor = float(row[10] or 0)
        desc = str(row[9] or "")
        data = row[2]  # Excel date serial
        hora = str(row[3] or "")
        ref_tarifa_by_cpf[cpf].append({"valor": valor, "desc": desc, "data": data, "hora": hora})

# Load our DB Taxa transactions by CPF (deduped)
for cpf, name, diff in tarifa_div_cpfs:
    # Get from DB
    cur.execute("""
        WITH deduped AS (
            SELECT DISTINCT ON (
                UPPER(usuario), data, tipo, valor,
                COALESCE(NULLIF(codigo_transacao, ''), hora::text)
            )
                UPPER(usuario) AS usuario_up,
                data, tipo, valor, descricao, hora, codigo_transacao
            FROM extrato_movimentacao
            WHERE is_snapshot = FALSE
              AND data <= '2026-06-30'
              AND tipo = 'Taxa'
            ORDER BY UPPER(usuario), data, tipo, valor,
                COALESCE(NULLIF(codigo_transacao, ''), hora::text)
        )
        SELECT data, valor, descricao, hora, codigo_transacao
        FROM deduped d
        WHERE UPPER(d.usuario_up) = (
            SELECT UPPER(colaborador) FROM quinzena_cadastro WHERE cpf = %s LIMIT 1
        )
        ORDER BY data, hora
    """, (cpf,))
    db_rows = cur.fetchall()
    
    ref_rows = ref_tarifa_by_cpf.get(cpf, [])
    
    ref_total = sum(abs(r["valor"]) for r in ref_rows)
    db_total = sum(abs(float(r["valor"])) for r in db_rows)
    
    print(f"\n  {cpf} {name} (diff={diff:+.2f})")
    print(f"    Ref: {len(ref_rows)} txns, total R$ {ref_total:.2f}")
    print(f"    DB:  {len(db_rows)} txns, total R$ {db_total:.2f}")
    
    if len(ref_rows) != len(db_rows) or abs(ref_total - db_total) > 0.01:
        # Show all ref
        print(f"    --- REF Taxa transactions ---")
        for r in ref_rows:
            print(f"      {r['desc'][:30]:<30} R$ {r['valor']:>8.2f}  data={r['data']}  hora={r['hora']}")
        print(f"    --- DB Taxa transactions ---")
        for r in db_rows:
            print(f"      {str(r['descricao'] or '')[:30]:<30} R$ {float(r['valor']):>8.2f}  data={r['data']}  hora={r['hora']}  cod={r['codigo_transacao']}")

# ============================================================
# PART 2: PRESTACAO — compare reference BASE PREST vs API data
# ============================================================
print("\n" + "=" * 80)
print("PART 2: PRESTACAO — REFERENCE BASE PREST vs API")
print("=" * 80)

# Load reference BASE PREST: CPF -> set of expense IDs
ws_bp = wb["BASE PREST "]
ref_expense_ids = set()
ref_expense_by_cpf = defaultdict(set)
ref_total_by_cpf = defaultdict(float)
for row in ws_bp.iter_rows(min_row=4, values_only=True):
    if row[0] is None:
        continue
    eid = int(row[0])
    cpf = str(row[9] or "").strip().replace(".", "").replace("-", "").replace(" ", "").zfill(11) if row[9] else ""
    val = float(row[26] or 0)
    ref_expense_ids.add(eid)
    ref_expense_by_cpf[cpf].add(eid)
    ref_total_by_cpf[cpf] += val

print(f"  Reference BASE PREST: {len(ref_expense_ids)} expense IDs")

# Load API expenses (Aprovado+Enviado, excluding FATURA/CARTAO)
cur.execute("""
    SELECT e.id, e.report_id, e.value, r.user_cpf, r.name as report_name, r.status
    FROM prestacao_expenses e
    JOIN prestacao_reports r ON e.report_id = r.id
    WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND r.user_cpf IS NOT NULL
      AND TRIM(r.name) !~* '^(fatu|farur|cart)'
      AND TRIM(r.name) !~* '(fatura|fatuar|fatut|farur)'
""")
api_expenses = cur.fetchall()
api_expense_ids = set()
api_expense_by_cpf = defaultdict(set)
api_total_by_cpf = defaultdict(float)
for r in api_expenses:
    eid = int(r["id"])
    cpf = str(r["user_cpf"] or "").strip().replace(".", "").replace("-", "").replace(" ", "").zfill(11)
    api_expense_ids.add(eid)
    api_expense_by_cpf[cpf].add(eid)
    api_total_by_cpf[cpf] += float(r["value"])

print(f"  API expenses: {len(api_expense_ids)} expense IDs")

# Compare
extra_in_api = api_expense_ids - ref_expense_ids
extra_in_ref = ref_expense_ids - api_expense_ids
common = api_expense_ids & ref_expense_ids

print(f"  Common: {len(common)}")
print(f"  Extra in API (not in ref): {len(extra_in_api)}")
print(f"  Extra in ref (not in API): {len(extra_in_ref)}")

# Group extra-in-API by report name to see patterns
cur.execute("""
    SELECT e.id, e.report_id, e.value, r.user_cpf, r.name as report_name, r.status
    FROM prestacao_expenses e
    JOIN prestacao_reports r ON e.report_id = r.id
    WHERE e.id = ANY(%s)
    ORDER BY r.name
""", (list(extra_in_api[:500]),))
extra_rows = cur.fetchall()

extra_by_report = defaultdict(lambda: {"count": 0, "total": 0, "status": set()})
for r in extra_rows:
    rn = str(r["report_name"] or "")[:40]
    extra_by_report[rn]["count"] += 1
    extra_by_report[rn]["total"] += float(r["value"])
    extra_by_report[rn]["status"].add(str(r["status"]))

print(f"\n  Extra in API — top 20 report names:")
for rn, info in sorted(extra_by_report.items(), key=lambda x: -x[1]["count"])[:20]:
    print(f"    {rn:<40} {info['count']:>4} items, R$ {info['total']:>10,.2f}  status={info['status']}")

# Check if extra expenses are from reports that exist in ref but with fewer expenses
# Or from reports that don't exist in ref at all
ref_report_ids = set()
for row in ws_bp.iter_rows(min_row=4, values_only=True):
    if row[0] is None:
        continue
    ref_report_ids.add(int(row[1]))

extra_report_ids = set(int(r["report_id"]) for r in extra_rows)
new_reports = extra_report_ids - ref_report_ids
existing_reports = extra_report_ids & ref_report_ids

print(f"\n  Extra expenses from NEW reports (not in ref): {len(new_reports)} reports")
print(f"  Extra expenses from EXISTING reports (in ref but more expenses): {len(existing_reports)} reports")

# Show new reports
if new_reports:
    cur.execute("""
        SELECT id, name, status, user_cpf, user_name
        FROM prestacao_reports
        WHERE id = ANY(%s)
        ORDER BY name
        LIMIT 30
    """, (list(new_reports[:100]),))
    print(f"\n  NEW reports (in API but not in ref BASE PREST):")
    for r in cur.fetchall():
        print(f"    {r['id']}  {str(r['name'])[:35]:<35}  {r['status']:<10}  {str(r['user_name'])[:25]}")

wb.close()
conn.close()
