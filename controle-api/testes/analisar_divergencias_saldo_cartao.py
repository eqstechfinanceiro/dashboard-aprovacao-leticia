#!/usr/bin/env python3
"""
Categoriza as divergencias de SALDO CARTAO (fechamento 2026-05-10) e testa a
hipotese: divergencias = CPFs ausentes na aba 'SALDO CARTAO' do CONTROLE
(VLOOKUP falhou -> IFERROR retornou 0).
"""
import os, sys
from pathlib import Path
from dotenv import load_dotenv
import psycopg2
import openpyxl

BASE = Path(__file__).parent.parent
sys.path.insert(0, str(BASE / "src"))
from name_matcher import NameMatcher, normalizar  # noqa: E402

load_dotenv(BASE / ".env")
CARGA = BASE / "data" / "CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx"
CONTROLE = BASE / "data" / "CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsx"
FECH = "2026-05-10"

# --- gabarito carga ---
wb = openpyxl.load_workbook(CARGA, data_only=True)
ws = wb["Planilha1"]
carga_rows = []
for r in range(7, 347):
    nome = ws.cell(row=r, column=1).value
    saldo = ws.cell(row=r, column=11).value
    if nome:
        try:
            carga_rows.append((str(nome).strip(), float(saldo) if saldo is not None else 0.0))
        except (ValueError, TypeError):
            carga_rows.append((str(nome).strip(), 0.0))
wb.close()

# --- nomes presentes na aba 'SALDO CARTAO' do CONTROLE (col B nome, col C valor) ---
print("Lendo aba 'SALDO CARTAO' do CONTROLE (nomes presentes)...")
wbc = openpyxl.load_workbook(CONTROLE, data_only=True, read_only=True)
wssc = wbc["SALDO CARTAO"]
nomes_saldo_sheet = set()
for row in wssc.iter_rows(min_row=5, values_only=True):
    nome = row[1] if len(row) > 1 else None  # col B
    if nome:
        nomes_saldo_sheet.add(normalizar(str(nome)))
wbc.close()
print(f"  {len(nomes_saldo_sheet)} nomes na aba SALDO CARTAO do controle")

# --- snapshots Neon <= FECH: ultimo por usuario (1 query) ---
conn = psycopg2.connect(os.getenv("NEON_DATABASE_URL"))
cur = conn.cursor()
cur.execute("""
    SELECT DISTINCT ON (usuario) usuario, valor
    FROM extrato_movimentacao
    WHERE is_snapshot AND data <= %s AND usuario IS NOT NULL
    ORDER BY usuario, data DESC
""", (FECH,))
neon_saldo = {u: float(v) for u, v in cur.fetchall()}
cur.close(); conn.close()
print(f"  {len(neon_saldo)} usuarios com snapshot <= {FECH} no Neon\n")

matcher = NameMatcher(list(neon_saldo.keys()), threshold=0.88)

cat = {"ok": 0, "gap_controle": 0, "carga0_no_sheet": 0, "real_diff": 0, "nao_match": 0}
real_diffs = []
for nome, saldo_carga in carga_rows:
    res = matcher.match(nome)
    if not res:
        cat["nao_match"] += 1
        continue
    saldo_neon = neon_saldo.get(res.nome_match, 0.0)
    if abs(saldo_neon - saldo_carga) < 0.01:
        cat["ok"] += 1
        continue
    # divergencia: o nome estava na aba SALDO CARTAO do controle?
    no_sheet = normalizar(nome) in nomes_saldo_sheet
    if abs(saldo_carga) < 0.01 and not no_sheet:
        cat["gap_controle"] += 1   # carga=0 E ausente da aba -> VLOOKUP falhou
    elif abs(saldo_carga) < 0.01 and no_sheet:
        cat["carga0_no_sheet"] += 1  # carga=0 mas estava na aba (valor 0 no controle)
    else:
        cat["real_diff"] += 1
        real_diffs.append((nome, saldo_carga, saldo_neon, no_sheet))

total = len(carga_rows)
print("=" * 80)
print(f"  CATEGORIZACAO DAS {total} LINHAS (fechamento {FECH})")
print("=" * 80)
print(f"  OK (bate exato)................: {cat['ok']}")
print(f"  GAP CONTROLE (carga=0, ausente da aba SALDO CARTAO; Neon tem valor): {cat['gap_controle']}")
print(f"  carga=0 mas presente na aba (controle tinha 0).....: {cat['carga0_no_sheet']}")
print(f"  DIVERGENCIA REAL (ambos != 0, valores diferentes)..: {cat['real_diff']}")
print(f"  Nome nao encontrado no Neon........................: {cat['nao_match']}")
print(f"\n  >> Match efetivo (OK + gaps explicados pelo controle) = "
      f"{cat['ok'] + cat['gap_controle'] + cat['carga0_no_sheet']}/{total - cat['nao_match']} matched")

if real_diffs:
    print(f"\n  DIVERGENCIAS REAIS a investigar ({len(real_diffs)}):")
    for nome, sc, sn, ns in real_diffs[:25]:
        print(f"     {nome[:32]:32} | carga={sc:>10.2f} | neon={sn:>10.2f} "
              f"| na_aba={'sim' if ns else 'NAO'}")
