#!/usr/bin/env python3
"""
Extrai as FORMULAS REAIS (nao cache) das planilhas CARGA e CONTROLE,
celula a celula, para descobrir a origem definitiva de cada valor.
"""
import openpyxl
from pathlib import Path

BASE = Path(__file__).parent.parent
CARGA = BASE / "data" / "CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx"
CONTROLE = BASE / "data" / "CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsx"

def col_letter(idx):
    return openpyxl.utils.get_column_letter(idx)

print("=" * 90)
print("CARGA 1 QZ — FORMULAS REAIS (data_only=False)")
print("=" * 90)
wb_f = openpyxl.load_workbook(CARGA, data_only=False)
wb_v = openpyxl.load_workbook(CARGA, data_only=True)
print(f"Abas: {wb_f.sheetnames}")
ws_f = wb_f["Planilha1"]
ws_v = wb_v["Planilha1"]

# Cabecalho linha 6
print("\n--- Cabecalho (linha 6) ---")
header = {}
for c in range(1, 20):
    v = ws_f.cell(row=6, column=c).value
    if v is not None:
        header[c] = str(v).strip()
        print(f"  {col_letter(c)}6 = {v!r}")

# Linhas 4, 5 (parametros / subtotais)
print("\n--- Linha 4 (parametros) ---")
for c in range(1, 20):
    v = ws_f.cell(row=4, column=c).value
    if v is not None:
        print(f"  {col_letter(c)}4 = {v!r}")
print("\n--- Linha 5 (subtotais) ---")
for c in range(1, 20):
    v = ws_f.cell(row=5, column=c).value
    if v is not None:
        print(f"  {col_letter(c)}5 = {v!r}")

# Primeiras 3 linhas de dados: formula + valor cacheado
for row in (7, 8, 9):
    print(f"\n--- Linha de dados {row}: FORMULA | VALOR ---")
    for c in range(1, 20):
        f = ws_f.cell(row=row, column=c).value
        v = ws_v.cell(row=row, column=c).value
        if f is not None or v is not None:
            nome = header.get(c, "?")
            print(f"  {col_letter(c)}{row} [{nome[:18]:18}] FORMULA={f!r:50} | VALOR={v!r}")

wb_f.close(); wb_v.close()

print("\n\n" + "=" * 90)
print("CONTROLE — aba PAINEL — FORMULAS REAIS")
print("=" * 90)
wb_cf = openpyxl.load_workbook(CONTROLE, data_only=False, read_only=True)
ws_p = wb_cf["PAINEL"]

# Cabecalho linha 11
print("\n--- Cabecalho PAINEL (linha 11) ---")
hp = {}
for row in ws_p.iter_rows(min_row=11, max_row=11):
    for cell in row:
        if cell.value is not None:
            hp[cell.column] = str(cell.value).strip()
            print(f"  {cell.coordinate} = {cell.value!r}")

# Primeira linha de dados (linha 12): formulas
print("\n--- PAINEL linha 12: FORMULAS ---")
for row in ws_p.iter_rows(min_row=12, max_row=12):
    for cell in row:
        if cell.value is not None:
            nome = hp.get(cell.column, "?")
            print(f"  {cell.coordinate} [{nome[:22]:22}] = {cell.value!r}")
wb_cf.close()
