#!/usr/bin/env python3
"""
Fecha as lacunas:
1. Aba EXTRATO do CONTROLE: header, tipos distintos, amostra -> comparar com extrato v3
2. Aba SALDO CARTAO do CONTROLE: confirmar que e snapshot
3. CARGA: achar caso com SALDO FINAL negativo p/ confirmar SALDO REEMBOLSAR
"""
import openpyxl
from collections import Counter
from pathlib import Path

BASE = Path(__file__).parent.parent
CARGA = BASE / "data" / "CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx"
CONTROLE = BASE / "data" / "CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsx"

print("=" * 90)
print("1. ABA 'EXTRATO' DO CONTROLE")
print("=" * 90)
wb = openpyxl.load_workbook(CONTROLE, data_only=True, read_only=True)
print(f"Abas: {wb.sheetnames}")

ws = wb["EXTRATO"]
# Cabecalho linha 2
rows_iter = ws.iter_rows(min_row=2, max_row=2, values_only=True)
header = next(rows_iter)
print(f"\nCabecalho (linha 2): {[h for h in header if h is not None]}")

# Amostra de 8 linhas + tipos distintos nos primeiros 3000 registros
print("\nAmostra de dados (linhas 3-12):")
tipos = Counter()
count = 0
for row in ws.iter_rows(min_row=3, max_row=3002, values_only=True):
    count += 1
    # acha coluna TIPO (J = idx 9 conforme doc) e VALOR (L = idx 11)
    if count <= 10:
        vals = [v for v in row[:13]]
        print(f"  {vals}")
    tipo = row[9] if len(row) > 9 else None
    if tipo is not None:
        tipos[str(tipo).strip()] += 1
print(f"\nTipos distintos (primeiros {count} registros):")
for t, n in tipos.most_common():
    print(f"   {t!r}: {n}")

print("\n" + "=" * 90)
print("2. ABA 'SALDO CARTAO' DO CONTROLE")
print("=" * 90)
ws2 = wb["SALDO CARTAO"]
h2 = next(ws2.iter_rows(min_row=5, max_row=5, values_only=True))
print(f"Cabecalho (linha 5): {[h for h in h2 if h is not None]}")
print("Amostra (linhas 6-13):")
for row in ws2.iter_rows(min_row=6, max_row=13, values_only=True):
    print(f"  {[v for v in row[:14]]}")
wb.close()

print("\n" + "=" * 90)
print("3. CARGA: casos com SALDO FINAL negativo (confirmar SALDO REEMBOLSAR)")
print("=" * 90)
wbc = openpyxl.load_workbook(CARGA, data_only=True)
wsc = wbc["Planilha1"]
# H=8 SALDO REEMBOLSAR, I=9 SALDO FINAL
achados = 0
for r in range(7, 347):
    reemb = wsc.cell(row=r, column=8).value
    sfinal = wsc.cell(row=r, column=9).value
    nome = wsc.cell(row=r, column=1).value
    try:
        if sfinal is not None and float(sfinal) < 0:
            print(f"  {nome[:30]:30} | SALDO FINAL={float(sfinal):10.2f} | SALDO REEMBOLSAR={reemb}")
            achados += 1
            if achados >= 12:
                break
    except (ValueError, TypeError):
        pass
if achados == 0:
    print("  Nenhum SALDO FINAL negativo encontrado na CARGA.")
    # Mostra casos onde SALDO REEMBOLSAR != 0
    print("  Casos com SALDO REEMBOLSAR != 0:")
    for r in range(7, 347):
        reemb = wsc.cell(row=r, column=8).value
        sfinal = wsc.cell(row=r, column=9).value
        nome = wsc.cell(row=r, column=1).value
        try:
            if reemb is not None and float(reemb) != 0:
                print(f"  {nome[:30]:30} | SALDO FINAL={sfinal} | SALDO REEMBOLSAR={float(reemb):.2f}")
                achados += 1
                if achados >= 12:
                    break
        except (ValueError, TypeError):
            pass
wbc.close()
