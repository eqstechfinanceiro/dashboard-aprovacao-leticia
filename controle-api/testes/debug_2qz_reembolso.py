"""Verifica o problema do reembolso na 2QZ MAIO."""
import openpyxl
from pathlib import Path

BASE = Path(__file__).parent.parent
wb = openpyxl.load_workbook(
    BASE / "data" / "CARGA 2 QZ MAIO 26 VEXPENSES EQS (1).xlsx",
    read_only=True, data_only=True
)
ws = wb["2 QZ DE MAIO 26"]

# col7=SALDO PENDENTE PARCIAL (saldo_reembolsar), col14=REEMBOLSO, col15=CF
print("Linhas com SALDO PENDENTE > 0 na ref 2QZ:")
count = 0
for row in ws.iter_rows(min_row=5, values_only=True):
    if row[2] is None:
        continue
    try:
        sr = float(row[7] or 0)
    except (TypeError, ValueError):
        sr = 0.0
    if sr > 0.01:
        try:
            reem = float(row[14] or 0)
            cf   = float(row[15] or 0)
        except (TypeError, ValueError):
            reem = cf = 0.0
        cpf = str(row[2]).strip().replace(".","").replace("-","").zfill(11)
        print(f"  {cpf} {str(row[1])[:22]:<22} SR={sr:.2f} Reem={reem:.2f} CF={cf:.2f}")
        count += 1
        if count >= 15:
            break

print(f"\nTotal com SR>0 mostrados: {count}")

# Ver estrutura do cabecalho linha 4
print("\nCabecalho linha 4 (indices relevantes):")
header = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
for i, v in enumerate(header[:18]):
    if v:
        print(f"  col{i}: {v!r}")
wb.close()
