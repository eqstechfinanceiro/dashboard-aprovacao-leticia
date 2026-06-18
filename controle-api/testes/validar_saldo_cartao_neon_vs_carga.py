#!/usr/bin/env python3
"""
TAREFA 4+5: Validar SALDO CARTAO por colaborador.
Compara o SALDO CARTAO da planilha CARGA 1 QZ (gabarito) com o calculado do
extrato Neon (ultimo snapshot <= data de fechamento), usando matching por NOME.
"""
import os
import sys
from pathlib import Path
from dotenv import load_dotenv
import psycopg2
import openpyxl

BASE = Path(__file__).parent.parent
sys.path.insert(0, str(BASE / "src"))
from name_matcher import NameMatcher  # noqa: E402

load_dotenv(BASE / ".env")
CARGA = BASE / "data" / "CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx"
FECHAMENTOS = ["2026-05-10", "2026-05-11"]

# 1. Ler gabarito da CARGA (nome col A, saldo_cartao col K)
wb = openpyxl.load_workbook(CARGA, data_only=True)
ws = wb["Planilha1"]
carga_rows = []  # (nome, saldo_cartao)
for r in range(7, 347):
    nome = ws.cell(row=r, column=1).value
    saldo = ws.cell(row=r, column=11).value
    if nome:
        try:
            carga_rows.append((str(nome).strip(), float(saldo) if saldo is not None else 0.0))
        except (ValueError, TypeError):
            carga_rows.append((str(nome).strip(), 0.0))
wb.close()
print(f"CARGA 1 QZ: {len(carga_rows)} colaboradores")

# 2. Usuarios distintos no Neon (com algum snapshot ate maio)
conn = psycopg2.connect(os.getenv("NEON_DATABASE_URL"))
cur = conn.cursor()
cur.execute("""
    SELECT DISTINCT usuario FROM extrato_movimentacao
    WHERE usuario IS NOT NULL AND data <= '2026-05-11'
""")
neon_usuarios = [row[0] for row in cur.fetchall()]
print(f"Neon: {len(neon_usuarios)} usuarios distintos")

matcher = NameMatcher(neon_usuarios, threshold=0.88)

for FECH in FECHAMENTOS:
    print("\n" + "=" * 80)
    print(f"  FECHAMENTO = ultimo snapshot <= {FECH}")
    print("=" * 80)

    n_match, n_exato, n_fuzzy, n_nao_match = 0, 0, 0, 0
    n_valor_ok, n_valor_diff = 0, 0
    nao_encontrados = []
    divergencias = []

    for nome, saldo_carga in carga_rows:
        res = matcher.match(nome)
        if not res:
            n_nao_match += 1
            nao_encontrados.append(nome)
            continue
        n_match += 1
        if res.metodo == "exato":
            n_exato += 1
        else:
            n_fuzzy += 1

        # saldo do Neon = ultimo snapshot <= FECH
        cur.execute("""
            SELECT valor FROM extrato_movimentacao
            WHERE usuario = %s AND is_snapshot AND data <= %s
            ORDER BY data DESC LIMIT 1
        """, (res.nome_match, FECH))
        row = cur.fetchone()
        saldo_neon = float(row[0]) if row else 0.0

        if abs(saldo_neon - saldo_carga) < 0.01:
            n_valor_ok += 1
        else:
            n_valor_diff += 1
            divergencias.append((nome, saldo_carga, saldo_neon, res.metodo, res.score))

    print(f"  Matching de nome : {n_match}/{len(carga_rows)} "
          f"(exato={n_exato}, fuzzy={n_fuzzy}, NAO encontrados={n_nao_match})")
    print(f"  Valor SALDO CARTAO: OK={n_valor_ok} | DIFERENTE={n_valor_diff} "
          f"({100*n_valor_ok/max(n_match,1):.1f}% dos matched)")

    if nao_encontrados:
        print(f"\n  Nomes NAO encontrados no Neon ({len(nao_encontrados)}):")
        for nm in nao_encontrados[:15]:
            print(f"     - {nm}")

    if divergencias:
        print(f"\n  Amostra de divergencias de valor ({len(divergencias)}):")
        for nome, sc, sn, met, score in divergencias[:15]:
            print(f"     {nome[:32]:32} | carga={sc:>10.2f} | neon={sn:>10.2f} "
                  f"| diff={sn-sc:>+10.2f} | {met}({score})")

cur.close(); conn.close()
