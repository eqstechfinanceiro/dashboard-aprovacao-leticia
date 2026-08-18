#!/usr/bin/env python3
"""Entender o que alimenta (-) PRESTACAO DE CONTAS: aba BASE PREST + valores cacheados do PAINEL."""
import openpyxl
from pathlib import Path
from collections import Counter

BASE = Path(__file__).parent.parent
CONTROLE = BASE / "data" / "CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsx"

wb = openpyxl.load_workbook(CONTROLE, data_only=True, read_only=True)

print("=" * 90)
print("ABA 'BASE PREST ' — cabecalho + amostra")
print("=" * 90)
# nome da aba pode ter espaco no fim
nome_aba = next((s for s in wb.sheetnames if s.strip().upper() == "BASE PREST"), None)
print(f"Aba encontrada: {nome_aba!r}")
ws = wb[nome_aba]

# cabecalho linha 2
header = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
print("\nColunas (linha 2):")
for i, h in enumerate(header):
    if h is not None:
        print(f"  [{i:2}] {openpyxl.utils.get_column_letter(i+1)}: {h!r}")

print("\nAmostra (linhas 3-8) — colunas-chave:")
status_col = Counter()
count = 0
for row in ws.iter_rows(min_row=3, max_row=5002, values_only=True):
    count += 1
    if count <= 6:
        print(f"  linha {count+2}: {[v for v in row[:12]]}")
    # tentar achar coluna status
print(f"\nTotal de linhas lidas: {count}")

# valor total da coluna AA (idx 26) - usada no SUMIF
print("\nSoma coluna AA (idx 26, usada no SUMIF do PAINEL):")
total_aa = 0.0
n_aa = 0
for row in ws.iter_rows(min_row=3, values_only=True):
    if len(row) > 26 and isinstance(row[26], (int, float)):
        total_aa += row[26]; n_aa += 1
print(f"  {n_aa} valores, soma = R$ {total_aa:,.2f}")

print("\n" + "=" * 90)
print("PAINEL: (-) PRESTACAO DE CONTAS (col Q=17) cacheado p/ amostra")
print("=" * 90)
wsp = wb["PAINEL"]
count = 0
for row in wsp.iter_rows(min_row=12, max_row=22, values_only=True):
    nome = row[1]; cpf = row[2]; prest = row[16] if len(row) > 16 else None  # Q = idx 16
    print(f"  {str(nome)[:30]:30} | CPF={cpf} | (-)PRESTACAO={prest}")
wb.close()
