import openpyxl, os
ROOT = r'C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test'
OUT = os.path.join(ROOT, 'controle-api', 'jun_qz1_headers.txt')
f = open(OUT, 'w', encoding='utf-8')

path = os.path.join(ROOT, 'data', '06 - JUNHO', 'CARGA 1 QZ JUNHO 26 VEXPENSES EQS.xlsx')
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
f.write(f"Sheet names: {wb.sheetnames}\n")

ws = wb["1 QZ JUNHO"]
# Check rows 1-8
for rn in range(1, 9):
    rows = list(ws.iter_rows(min_row=rn, max_row=rn, values_only=True))
    if rows:
        f.write(f"\n=== Row {rn} ===\n")
        for j, v in enumerate(rows[0]):
            if v is not None:
                f.write(f"  [{j}] (Col {chr(65+j) if j<26 else '?'}) = {repr(v)}\n")

wb.close()
f.close()
print("Done: " + OUT)
