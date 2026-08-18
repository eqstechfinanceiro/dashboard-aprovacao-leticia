#!/usr/bin/env python3
"""Check reference BASE PREST date range and report count."""
import openpyxl
from collections import defaultdict
import datetime as dt_module

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"

wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws = wb["BASE PREST "]

# Collect all data
report_ids = set()
report_names = {}
expense_dates = []
expense_by_month = defaultdict(int)
expense_by_report = defaultdict(lambda: {"count": 0, "total": 0, "name": ""})

for row in ws.iter_rows(min_row=4, values_only=True):
    if row[0] is None:
        continue
    eid = int(row[0])
    rid = int(row[1]) if row[1] else None
    rname = str(row[2] or "")
    edate = row[3]
    
    if rid:
        report_ids.add(rid)
        report_names[rid] = rname
        expense_by_report[rid]["count"] += 1
        expense_by_report[rid]["name"] = rname
    
    if edate is not None:
        if isinstance(edate, dt_module.datetime):
            expense_dates.append(edate)
            expense_by_month[edate.strftime("%Y-%m")] += 1
        elif isinstance(edate, dt_module.date):
            expense_dates.append(edate)
            expense_by_month[edate.strftime("%Y-%m")] += 1
        elif isinstance(edate, (int, float)):
            try:
                d = dt_module.date.fromordinal(int(edate) + 693594)
                expense_dates.append(d)
                expense_by_month[d.strftime("%Y-%m")] += 1
            except:
                pass

print(f"Reference BASE PREST:")
print(f"  Total expenses: {len(expense_dates)}")
print(f"  Unique reports: {len(report_ids)}")
if expense_dates:
    print(f"  Date range: {min(expense_dates)} to {max(expense_dates)}")
else:
    print("  No dates parsed!")

if expense_by_month:
    print(f"\nExpenses by month:")
    for month in sorted(expense_by_month.keys()):
        print(f"  {month}: {expense_by_month[month]:>5} expenses")


# Check: does the reference BASE PREST have a "Status" column and what statuses are there?
print(f"\nStatus distribution in reference BASE PREST:")
statuses = defaultdict(int)
for row in ws.iter_rows(min_row=4, values_only=True):
    if row[0] is None:
        continue
    # Status is col[10] based on earlier header check
    status = str(row[10] or "").strip() if len(row) > 10 else ""
    statuses[status] += 1
for s, c in sorted(statuses.items(), key=lambda x: -x[1]):
    print(f"  {s}: {c}")

# Check: what's the value column? Let me verify
print(f"\nColumn headers (row 3):")
for col_idx, cell in enumerate(ws.iter_rows(min_row=3, max_row=3, values_only=True).__next__()):
    if cell is not None:
        print(f"  col[{col_idx}] = {repr(cell)[:50]}")

# Check value column - find which column has the expense values
print(f"\nFirst data row values:")
for row in ws.iter_rows(min_row=4, max_row=4, values_only=True):
    for j, v in enumerate(row):
        if v is not None:
            print(f"  col[{j}] = {repr(v)[:60]}")

wb.close()
