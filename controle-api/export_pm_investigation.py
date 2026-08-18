"""Export expenses with NULL or unknown payment_method_id to Excel for investigation."""
import os, psycopg2, psycopg2.extras
from dotenv import load_dotenv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

load_dotenv(Path(__file__).parent / ".env")
conn = psycopg2.connect(os.getenv("NEON_DATABASE_URL"))
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cur.execute("""
    SELECT e.id, e.report_id, e.value, e.description, e.date,
           r.name as report_name, r.user_name, r.user_cpf,
           e.raw_data->>'payment_method_id' as pm_id,
           e.raw_data->>'payment_method_name' as pm_name,
           r.raw_data->>'pdf_link' as pdf_link,
           r.raw_data->>'approval_date' as approval_date,
           r.raw_data->>'observation' as observation
    FROM prestacao_expenses e
    JOIN prestacao_reports r ON e.report_id = r.id
    WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND r.user_cpf IS NOT NULL
      AND (
        e.raw_data->>'payment_method_id' IS NULL
        OR e.raw_data->>'payment_method_id' = '627726'
      )
    ORDER BY e.raw_data->>'payment_method_id' NULLS LAST, e.value DESC
""")
rows = cur.fetchall()
print(f"Found {len(rows)} expenses to investigate")

wb = Workbook()
ws = wb.active
ws.title = "PM Investigation"

headers = ["Expense ID", "Report ID", "Report Name", "User", "CPF", "Date",
           "Value", "Description", "PM ID", "PM Name", "PDF Link", "Approval Date", "Observation"]
ws.append(headers)
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")

for r in rows:
    pdf = r.get("pdf_link") or ""
    if pdf:
        pdf_url = pdf if pdf.startswith("http") else f"https://app.vexpenses.com{pdf}"
    else:
        pdf_url = ""
    ws.append([
        r["id"],
        r["report_id"],
        r["report_name"],
        r["user_name"],
        r["user_cpf"],
        str(r["date"]) if r["date"] else "",
        float(r["value"]) if r["value"] else 0,
        r["description"] or "",
        r["pm_id"] or "NULL",
        r["pm_name"] or "",
        pdf_url,
        str(r["approval_date"]) if r.get("approval_date") else "",
        r.get("observation") or "",
    ])

# Make PDF link clickable
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=11)
    if cell.value:
        cell.hyperlink = cell.value
        cell.font = Font(color="0563C1", underline="single")

# Auto-width
for col in ws.columns:
    max_len = max(len(str(c.value or "")) for c in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

out = Path(__file__).parent / "data" / "PM_INVESTIGATION.xlsx"
out.parent.mkdir(exist_ok=True)
wb.save(out)
print(f"Saved to {out}")

conn.close()
