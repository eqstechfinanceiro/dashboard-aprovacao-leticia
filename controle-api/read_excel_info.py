import openpyxl

path = r'C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\controle-api\data\CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx'

# ── 1. Carrega COM fórmulas ──────────────────────────────────────────────────
wb_formulas = openpyxl.load_workbook(path, data_only=False)
wb_values   = openpyxl.load_workbook(path, data_only=True)

print("=" * 60)
print("1. ABAS DISPONÍVEIS")
print("=" * 60)
for name in wb_formulas.sheetnames:
    print(f"  - {name}")
print()

for sheet_name in wb_formulas.sheetnames:
    ws_f = wb_formulas[sheet_name]
    ws_v = wb_values[sheet_name]

    print("=" * 60)
    print(f"ABA: '{sheet_name}'")
    print(f"Dimensões: max_row={ws_f.max_row}, max_col={ws_f.max_column}")
    print("=" * 60)

    # ── Encontra a linha do cabeçalho (primeira linha com múltiplas células não-nulas) ──
    header_row = None
    header_cols = []
    for row_idx in range(1, min(20, ws_f.max_row + 1)):
        values = [ws_f.cell(row=row_idx, column=c).value for c in range(1, ws_f.max_column + 1)]
        non_null = [v for v in values if v is not None]
        if len(non_null) >= 3:
            header_row = row_idx
            header_cols = values
            break

    print(f"\n2. LINHA DO CABEÇALHO: {header_row}")
    print()
    print("3. COLUNAS ENCONTRADAS:")
    col_map = {}  # nome -> índice de coluna
    for idx, col_name in enumerate(header_cols):
        if col_name is not None:
            col_letter = openpyxl.utils.get_column_letter(idx + 1)
            print(f"  [{col_letter}] {col_name}")
            col_map[str(col_name).strip().upper()] = idx + 1
    print()

    # ── Colunas de interesse ─────────────────────────────────────────────────
    targets = ["SALDO FINAL", "SALDO CARTAO", "CARGA PARCIAL", "CARGA FINAL", "REEMBOLSO"]

    print("4-8. FÓRMULAS DAS COLUNAS DE INTERESSE:")
    print("-" * 60)

    for target in targets:
        # Busca parcial (case-insensitive) no cabeçalho
        found_col = None
        found_name = None
        for col_name_key, col_idx in col_map.items():
            if target in col_name_key:
                found_col = col_idx
                found_name = col_name_key
                break

        if found_col is None:
            print(f"\n  ► {target}: NÃO ENCONTRADA")
            # Tenta busca mais ampla
            for col_name_key, col_idx in col_map.items():
                for word in target.split():
                    if word in col_name_key:
                        print(f"    (Possível match: '{col_name_key}' na coluna {col_idx})")
            continue

        col_letter = openpyxl.utils.get_column_letter(found_col)
        print(f"\n  ► {target} (coluna {col_letter} - '{found_name}'):")

        # Coleta fórmulas das primeiras linhas de dados (após o cabeçalho)
        formulas_found = {}
        for row_idx in range(header_row + 1, min(header_row + 20, ws_f.max_row + 1)):
            cell = ws_f.cell(row=row_idx, column=found_col)
            cell_val = ws_v.cell(row=row_idx, column=found_col)
            if cell.value is not None:
                formula = str(cell.value)
                if formula not in formulas_found:
                    formulas_found[formula] = {"row": row_idx, "value": cell_val.value}
                if len(formulas_found) >= 3:
                    break

        if formulas_found:
            for formula, info in formulas_found.items():
                print(f"    Linha {info['row']}: {formula}")
                if info['value'] is not None:
                    print(f"             (valor calculado: {info['value']})")
        else:
            print(f"    Nenhum dado encontrado abaixo do cabeçalho")

    print()

print("\nScript finalizado com sucesso.")
