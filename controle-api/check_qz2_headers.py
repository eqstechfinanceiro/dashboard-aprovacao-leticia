import openpyxl, os, sys
ROOT = r'C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test'
OUT = os.path.join(ROOT, 'controle-api', 'qz2_headers.txt')
f = open(OUT, 'w', encoding='utf-8')

# Check QZ2 sheets for reembolso columns
files = [
    ("Jan QZ2", os.path.join(ROOT, 'data', '01 - JANEIRO', '2QZ JANEIRO 2026 - VEXPENSES.xlsx'), "2 QZ VEXPENSES 01_2026", 6),
    ("Feb QZ2", os.path.join(ROOT, 'data', '02 - FEVEREIRO', '2QZ FEVEREIRO 2026 - VEXPENSES EQS.xlsx'), "2 QZ VEXPENSES 02_2026", 6),
    ("May QZ2", os.path.join(ROOT, 'data', '05 - MAIO', 'CARGA 2 QZ MAIO 26 VEXPENSES EQS.xlsx'), "2 QZ DE MAIO 26", 4),
    ("Jun QZ2", os.path.join(ROOT, 'data', '06 - JUNHO', 'CARGA 2 QZ JUNHO 26 VEXPENSES EQS.xlsx'), "2 QZ JUNHO", 6),
]

for label, path, sheet_name, header_row in files:
    if not os.path.exists(path):
        f.write(f"{label}: FILE NOT FOUND\n")
        continue
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        f.write(f"{label}: SHEET NOT FOUND, available: {wb.sheetnames}\n")
        wb.close()
        continue
    ws = wb[sheet_name]
    
    # Print header row
    headers = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    if headers:
        f.write(f"\n{label} (header row {header_row}):\n")
        for j, h in enumerate(headers[0]):
            if h is not None:
                f.write(f"  [{j}] (Col {chr(65+j) if j<26 else '?'}) = {repr(h)}\n")
    
    # Print first 3 data rows
    for dr in range(header_row+1, header_row+4):
        data = list(ws.iter_rows(min_row=dr, max_row=dr, values_only=True))
        if data:
            f.write(f"  Data row {dr}:\n")
            for j, v in enumerate(data[0]):
                if v is not None:
                    f.write(f"    [{j}] = {repr(v)}\n")
    
    wb.close()

f.close()
print("Done: " + OUT)
