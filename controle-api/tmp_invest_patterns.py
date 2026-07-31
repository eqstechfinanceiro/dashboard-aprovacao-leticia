#!/usr/bin/env python3
"""
Deep investigation into the 67 reports not in reference BASE PREST.
Look for patterns: users, date ranges, report types, statuses, etc.
"""
import os, re
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from pathlib import Path
import openpyxl
from collections import defaultdict

BASE = Path(__file__).parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"

conn = psycopg2.connect(NEON_URL, connect_timeout=10)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# Load reference report IDs
wb_ref = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_bp = wb_ref["BASE PREST "]
ref_report_ids = set()
ref_report_by_user = defaultdict(set)  # user_name -> set of report IDs
ref_report_by_cpf = defaultdict(set)   # cpf -> set of report IDs
for row in ws_bp.iter_rows(min_row=4, values_only=True):
    if row[0] is not None and row[1] is not None:
        rid = int(row[1])
        ref_report_ids.add(rid)
        user_name = str(row[4] or "").strip()
        cpf = str(row[9] or "").strip()
        ref_report_by_user[user_name].add(rid)
        ref_report_by_cpf[cpf].add(rid)
wb_ref.close()

# Get all API reports (Aprovado+Enviado, excluding FATURA)
cur.execute("""
    SELECT r.id, r.name, r.status, r.user_cpf, r.user_name, r.created_at,
           COUNT(e.id) as expense_count, SUM(e.value) as total_value,
           MIN(e.date) as min_exp_date, MAX(e.date) as max_exp_date
    FROM prestacao_reports r
    JOIN prestacao_expenses e ON e.report_id = r.id
    WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND r.user_cpf IS NOT NULL
      AND TRIM(r.name) !~* '^(fatu|farur|cart)'
      AND TRIM(r.name) !~* '(fatura|fatuar|fatut|farur)'
    GROUP BY r.id, r.name, r.status, r.user_cpf, r.user_name, r.created_at
    ORDER BY r.user_name, r.name
""")
all_api_reports = cur.fetchall()

new_reports = [r for r in all_api_reports if int(r["id"]) not in ref_report_ids]
in_ref_reports = [r for r in all_api_reports if int(r["id"]) in ref_report_ids]

print("=" * 80)
print("PATTERN ANALYSIS: 67 reports in API but NOT in reference BASE PREST")
print("=" * 80)

# 1. By user — are these users that have OTHER reports in the reference?
print("\n--- By User ---")
new_by_user = defaultdict(list)
for r in new_reports:
    new_by_user[str(r["user_name"])].append(r)

for user in sorted(new_by_user.keys()):
    new_r = new_by_user[user]
    ref_count = len(ref_report_by_user.get(user, set()))
    api_count = sum(1 for r in all_api_reports if str(r["user_name"]) == user)
    total_new = sum(float(r["total_value"]) for r in new_r)
    print(f"  {user[:30]:<30}  new={len(new_r):>2} reports  ref={ref_count:>3}  api_total={api_count:>3}  new_value=R$ {total_new:>10,.2f}")

# 2. By status
print("\n--- By Status ---")
by_status = defaultdict(lambda: {"count": 0, "total": 0})
for r in new_reports:
    by_status[r["status"]]["count"] += 1
    by_status[r["status"]]["total"] += float(r["total_value"])
for s, info in sorted(by_status.items()):
    print(f"  {s}: {info['count']} reports, R$ {info['total']:,.2f}")

# 3. By created_at month
print("\n--- By Created At Month ---")
by_month = defaultdict(lambda: {"count": 0, "total": 0, "reports": []})
for r in new_reports:
    if r["created_at"]:
        month = str(r["created_at"])[:7]
    else:
        month = "unknown"
    by_month[month]["count"] += 1
    by_month[month]["total"] += float(r["total_value"])
    by_month[month]["reports"].append(r)
for m in sorted(by_month.keys()):
    info = by_month[m]
    print(f"  {m}: {info['count']:>2} reports, R$ {info['total']:>10,.2f}")

# 4. Check: do these users have ANY reports in the reference?
print("\n--- Users with NO reports in reference at all ---")
users_with_no_ref = []
for user in sorted(new_by_user.keys()):
    ref_count = len(ref_report_by_user.get(user, set()))
    if ref_count == 0:
        users_with_no_ref.append(user)
        total_new = sum(float(r["total_value"]) for r in new_by_user[user])
        print(f"  {user[:35]:<35}  {len(new_by_user[user]):>2} new reports  R$ {total_new:>10,.2f}")
print(f"  Total: {len(users_with_no_ref)} users with NO reports in reference")

# 5. Check: for users that DO have reports in reference, compare report names
print("\n--- Users that HAVE reports in reference but also have new reports ---")
for user in sorted(new_by_user.keys()):
    ref_count = len(ref_report_by_user.get(user, set()))
    if ref_count > 0:
        new_r = new_by_user[user]
        total_new = sum(float(r["total_value"]) for r in new_r)
        print(f"\n  {user[:35]:<35}  ref={ref_count} reports, new={len(new_r)} reports, R$ {total_new:>10,.2f}")
        for r in new_r:
            print(f"    {r['id']}  {str(r['name'])[:35]:<35}  {r['status']:<10}  created={str(r['created_at'])[:10]}  expenses={r['expense_count']:>3}  R$ {float(r['total_value']):>10,.2f}")

# 6. Check: are any of these "CARTÃO CORPORATIVO ITAÚ" reports?
print("\n--- Report name patterns ---")
patterns = defaultdict(lambda: {"count": 0, "total": 0})
for r in new_reports:
    name = str(r["name"]).strip().upper()
    if "CART" in name and "ITAU" in name:
        cat = "CARTÃO ITAÚ"
    elif "CAIXA" in name:
        cat = "CAIXA"
    elif "MATERIAL" in name:
        cat = "MATERIAL"
    elif "ITAU" in name or "ITAÚ" in name:
        cat = "ITAU"
    elif "DOLAR" in name or "DÓLAR" in name:
        cat = "DOLAR"
    elif "FARTUR" in name:
        cat = "FARTURA (typo)"
    else:
        cat = "OTHER"
    patterns[cat]["count"] += 1
    patterns[cat]["total"] += float(r["total_value"])
for cat, info in sorted(patterns.items(), key=lambda x: -x[1]["count"]):
    print(f"  {cat:<20}  {info['count']:>2} reports  R$ {info['total']:>10,.2f}")

# 7. Check: reference BASE PREST — does it have a "CARTÃO CORPORATIVO ITAÚ" report for ANA LUIZA?
print("\n--- Check: does reference have 'CARTÃO CORPORATIVO' reports? ---")
for row_data in list(openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)["BASE PREST "].iter_rows(min_row=4, values_only=True))[:100]:
    if row_data[2] and "CART" in str(row_data[2]).upper():
        print(f"  {row_data[1]}  {row_data[2]}  {row_data[4]}")
        break
else:
    print("  No 'CARTÃO' reports found in reference BASE PREST (first 100 rows)")

# Check all CART reports in reference
wb_ref2 = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_bp2 = wb_ref2["BASE PREST "]
cart_in_ref = []
for row in ws_bp2.iter_rows(min_row=4, values_only=True):
    if row[2] and "CART" in str(row[2]).upper():
        cart_in_ref.append((row[1], row[2], row[4]))
wb_ref2.close()
print(f"\n  Total 'CART' reports in reference: {len(cart_in_ref)}")
for rid, name, user in cart_in_ref[:10]:
    print(f"    {rid}  {name}  {user}")

# 8. Check: does the reference BASE PREST include reports with "CARTÃO CORPORATIVO ITAÚ" in the name?
# These might be card billing reports that should be excluded
print("\n--- 'CARTÃO CORPORATIVO' reports in API (all statuses) ---")
cur.execute("""
    SELECT id, name, status, user_name, created_at
    FROM prestacao_reports
    WHERE TRIM(name) ILIKE '%cartão corporativo%' OR TRIM(name) ILIKE '%cartao corporativo%'
    ORDER BY user_name, name
""")
cartao_reports = cur.fetchall()
for r in cartao_reports:
    in_ref_flag = "IN REF" if int(r["id"]) in ref_report_ids else "NOT IN REF"
    print(f"  {r['id']}  {str(r['name'])[:35]:<35}  {r['status']:<10}  {str(r['user_name'])[:25]:<25}  {in_ref_flag}")

conn.close()
