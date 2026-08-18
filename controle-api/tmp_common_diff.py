#!/usr/bin/env python3
"""
Deep dive into common reports with value differences.
For each, find which expenses are in API but not ref, and vice versa.
"""
import os, re
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from pathlib import Path
import openpyxl
from collections import defaultdict

BASE = Path(__file__).parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"

# Load reference BASE PREST expenses by report
wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_bp = wb["BASE PREST "]

ref_by_report = defaultdict(lambda: {"expenses": {}, "total": 0})
for row in ws_bp.iter_rows(min_row=4, values_only=True):
    if row[0] is None:
        continue
    eid = int(row[0])
    rid = int(row[1]) if row[1] else None
    valor = float(row[26] or 0) if len(row) > 26 else 0
    desc = str(row[7] or "") if len(row) > 7 else ""
    date_val = row[6] if len(row) > 6 else None
    if rid:
        ref_by_report[rid]["expenses"][eid] = {"value": valor, "desc": desc, "date": date_val}
        ref_by_report[rid]["total"] += valor
wb.close()

# Load API with filters
conn = psycopg2.connect(NEON_URL, connect_timeout=10)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# Get PAINEL vexpenses CPFs
wb2 = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_p = wb2["PAINEL"]
vexpenses_cpfs = set()
for row in ws_p.iter_rows(min_row=12, values_only=True):
    if row[2] is None:
        continue
    cpf = str(row[2] or "").strip()
    cartao_vx = str(row[12] or "").strip().upper() if len(row) > 12 else ""
    if cartao_vx == "SIM":
        vexpenses_cpfs.add(cpf)
wb2.close()

cur.execute("""
    SELECT e.id, e.report_id, e.date, e.value, e.description, e.status as expense_status,
           r.name as report_name, r.status as report_status,
           r.user_cpf, r.user_name, r.created_at as report_created_at
    FROM prestacao_expenses e
    JOIN prestacao_reports r ON e.report_id = r.id
    WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND r.user_cpf IS NOT NULL
    ORDER BY r.user_name, r.name, e.date
""")
all_api = cur.fetchall()
conn.close()

# Apply filters
filtered = []
for e in all_api:
    name = (e["report_name"] or "").strip()
    if re.match(r'^(fatu|farur|cart)', name, re.IGNORECASE):
        continue
    if re.search(r'(fatura|fatuar|fatut|farur|fartur)', name, re.IGNORECASE):
        continue
    if re.search(r'(cartão corporativo|cartao corporativo)', name, re.IGNORECASE):
        continue
    if re.search(r'itau|itaú', name, re.IGNORECASE):
        continue
    if e["user_cpf"] not in vexpenses_cpfs:
        continue
    filtered.append(e)

api_by_report = defaultdict(lambda: {"expenses": {}, "total": 0, "name": "", "user": "", "cpf": "", "status": ""})
for e in filtered:
    eid = int(e["id"])
    rid = int(e["report_id"])
    api_by_report[rid]["expenses"][eid] = {
        "value": float(e["value"]),
        "desc": e["description"] or "",
        "date": e["date"],
    }
    api_by_report[rid]["total"] += float(e["value"])
    api_by_report[rid]["name"] = e["report_name"]
    api_by_report[rid]["user"] = e["user_name"]
    api_by_report[rid]["cpf"] = e["user_cpf"]
    api_by_report[rid]["status"] = e["report_status"]

# Find common reports with value differences
common_diffs = []
for rid in set(api_by_report.keys()) & set(ref_by_report.keys()):
    api_r = api_by_report[rid]
    ref_r = ref_by_report[rid]
    diff = api_r["total"] - ref_r["total"]
    if abs(diff) > 0.01:
        common_diffs.append((rid, diff, api_r, ref_r))

common_diffs.sort(key=lambda x: abs(x[1]), reverse=True)

print("=" * 100)
print(f"COMMON REPORTS WITH VALUE DIFFERENCES: {len(common_diffs)}")
print("=" * 100)

for rid, diff, api_r, ref_r in common_diffs:
    api_eids = set(api_r["expenses"].keys())
    ref_eids = set(ref_r["expenses"].keys())
    
    extra_in_api = api_eids - ref_eids
    extra_in_ref = ref_eids - api_eids
    common_eids = api_eids & ref_eids
    
    # Check value differences in common expenses
    value_mismatches = []
    for eid in common_eids:
        api_val = api_r["expenses"][eid]["value"]
        ref_val = ref_r["expenses"][eid]["value"]
        if abs(api_val - ref_val) > 0.01:
            value_mismatches.append((eid, api_val, ref_val))
    
    print(f"\n  Report {rid}: {api_r['name']}")
    print(f"    User: {api_r['user']}  CPF: {api_r['cpf']}  Status: {api_r['status']}")
    print(f"    API: {len(api_eids)} items, R$ {api_r['total']:,.2f}")
    print(f"    Ref: {len(ref_eids)} items, R$ {ref_r['total']:,.2f}")
    print(f"    Diff: R$ {diff:+,.2f}")
    print(f"    Extra in API: {len(extra_in_api)} items, Extra in Ref: {len(extra_in_ref)} items, Value mismatches: {len(value_mismatches)}")
    
    if extra_in_api:
        total_extra_api = sum(api_r["expenses"][eid]["value"] for eid in extra_in_api)
        print(f"    → Extra in API (R$ {total_extra_api:,.2f}):")
        for eid in sorted(extra_in_api)[:10]:
            e = api_r["expenses"][eid]
            print(f"      eid={eid}  R$ {e['value']:>10,.2f}  {str(e['desc'])[:40]}  date={e['date']}")
        if len(extra_in_api) > 10:
            print(f"      ... and {len(extra_in_api) - 10} more")
    
    if extra_in_ref:
        total_extra_ref = sum(ref_r["expenses"][eid]["value"] for eid in extra_in_ref)
        print(f"    → Extra in Ref (R$ {total_extra_ref:,.2f}):")
        for eid in sorted(extra_in_ref)[:10]:
            e = ref_r["expenses"][eid]
            print(f"      eid={eid}  R$ {e['value']:>10,.2f}  {str(e['desc'])[:40]}  date={e['date']}")
        if len(extra_in_ref) > 10:
            print(f"      ... and {len(extra_in_ref) - 10} more")
    
    if value_mismatches:
        print(f"    → Value mismatches in common expenses:")
        for eid, av, rv in value_mismatches[:5]:
            print(f"      eid={eid}  API=R$ {av:,.2f}  Ref=R$ {rv:,.2f}  diff=R$ {av-rv:+,.2f}")

# Summary
print("\n" + "=" * 100)
print("SUMMARY: Common report value differences")
print("=" * 100)

total_extra_api = sum(len(set(api_r["expenses"].keys()) - set(ref_r["expenses"].keys())) for _, _, api_r, ref_r in common_diffs)
total_extra_ref = sum(len(set(ref_r["expenses"].keys()) - set(api_r["expenses"].keys())) for _, _, api_r, ref_r in common_diffs)
total_value_mismatch = sum(len([1 for eid in set(api_r["expenses"].keys()) & set(ref_r["expenses"].keys()) if abs(api_r["expenses"][eid]["value"] - ref_r["expenses"][eid]["value"]) > 0.01]) for _, _, api_r, ref_r in common_diffs)

print(f"  Total extra expenses in API: {total_extra_api}")
print(f"  Total extra expenses in Ref: {total_extra_ref}")
print(f"  Total value mismatches: {total_value_mismatch}")
print(f"  Net value difference: R$ {sum(d[1] for d in common_diffs):,.2f}")
