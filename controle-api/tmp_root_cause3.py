#!/usr/bin/env python3
"""
Focused investigation:
1. EXTRATO column structure verification
2. PRESTACAO: Why 66 new reports are in API but not in reference BASE PREST
3. TARIFA: Check if we have post-cutoff transactions
4. Check "Fartura" typo exclusion
"""
import os, re
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from pathlib import Path
import openpyxl
from collections import defaultdict

BASE = Path(__file__).parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - JULHO 2026.xlsx"

conn = psycopg2.connect(NEON_URL, connect_timeout=10)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# ============================================================
# PART 1: EXTRATO column structure
# ============================================================
print("=" * 80)
print("PART 1: EXTRATO column structure")
print("=" * 80)

wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws = wb["EXTRATO"]

# Check header row (row 8)
print("Header row (row 8):")
for col_idx, cell in enumerate(ws.iter_rows(min_row=8, max_row=8, values_only=True).__next__()):
    if cell is not None:
        print("  col[{}] = {}".format(col_idx, repr(cell)[:60]))

# Check a few data rows with different structures
print("\nFirst 3 data rows (raw):")
for i, row in enumerate(ws.iter_rows(min_row=9, max_row=11, values_only=True)):
    print("  row {}: len={}, values={}".format(i, len(row), [repr(v)[:30] for v in row]))

# Check row lengths
lengths = defaultdict(int)
for row in ws.iter_rows(min_row=9, values_only=True):
    lengths[len(row)] += 1
print("\nRow length distribution:")
for l, c in sorted(lengths.items()):
    print("  len={}: {} rows".format(l, c))

# ============================================================
# PART 2: Check reference BASE PREST structure and date range
# ============================================================
print("\n" + "=" * 80)
print("PART 2: Reference BASE PREST structure")
print("=" * 80)

ws_bp = wb["BASE PREST "]
# Header row
print("Header row (row 3):")
for col_idx, cell in enumerate(ws_bp.iter_rows(min_row=3, max_row=3, values_only=True).__next__()):
    if cell is not None:
        print("  col[{}] = {}".format(col_idx, repr(cell)[:50]))

# Check date range of expenses in reference BASE PREST
# Find the date column
print("\nFirst 3 data rows:")
for i, row in enumerate(ws_bp.iter_rows(min_row=4, max_row=6, values_only=True)):
    print("  row {}: len={}".format(i, len(row)))
    for j, v in enumerate(row):
        if v is not None:
            print("    [{}] = {}".format(j, repr(v)[:50]))

# Check unique report statuses in reference BASE PREST
# Need to find the status column
print("\nChecking columns for status...")
# Print all non-None values for first data row
for i, row in enumerate(ws_bp.iter_rows(min_row=4, max_row=4, values_only=True)):
    for j, v in enumerate(row):
        if v is not None:
            print("  col[{}] = {}".format(j, repr(v)[:60]))

wb.close()

# ============================================================
# PART 3: Check if new API reports have expenses in ref date range
# ============================================================
print("\n" + "=" * 80)
print("PART 3: New API reports — expense date ranges")
print("=" * 80)

# Get the 66 new report IDs
wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_bp = wb["BASE PREST "]
ref_report_ids = set()
for row in ws_bp.iter_rows(min_row=4, values_only=True):
    if row[0] is not None and row[1] is not None:
        ref_report_ids.add(int(row[1]))
wb.close()

# Get API expenses for new reports
cur.execute("""
    SELECT r.id, r.name, r.status, r.user_cpf, r.user_name,
           MIN(e.date) as min_date, MAX(e.date) as max_date,
           COUNT(e.id) as expense_count, SUM(e.value) as total_value
    FROM prestacao_reports r
    JOIN prestacao_expenses e ON e.report_id = r.id
    WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND r.user_cpf IS NOT NULL
      AND TRIM(r.name) !~* '^(fatu|farur|cart)'
      AND TRIM(r.name) !~* '(fatura|fatuar|fatut|farur)'
    GROUP BY r.id, r.name, r.status, r.user_cpf, r.user_name
""")
all_api_reports = cur.fetchall()

new_reports = [r for r in all_api_reports if int(r["id"]) not in ref_report_ids]

print("New API reports (not in ref BASE PREST):")
for r in sorted(new_reports, key=lambda x: x["min_date"] or "9999"):
    print("  {}  {:<35}  {:<10}  {:<25}  expenses={:>3}  total=R$ {:>10,.2f}  dates={} to {}".format(
        r["id"], str(r["name"])[:35], r["status"], str(r["user_name"])[:25],
        r["expense_count"], float(r["total_value"]), r["min_date"], r["max_date"]))

# ============================================================
# PART 4: Check "Fartura" typo — is it being excluded?
# ============================================================
print("\n" + "=" * 80)
print("PART 4: FATURA/FARTURA exclusion check")
print("=" * 80)

# Check all report names that match FATURA pattern
cur.execute("""
    SELECT id, name, status, user_name, COUNT(e.id) as expense_count, SUM(e.value) as total
    FROM prestacao_reports r
    JOIN prestacao_expenses e ON e.report_id = r.id
    WHERE TRIM(r.name) ILIKE 'fat%' OR TRIM(r.name) ILIKE 'cart%'
    GROUP BY r.id, r.name, r.status, r.user_name
    ORDER BY r.name
""")
fatura_reports = cur.fetchall()
print("FATURA/CARTAO reports (should be EXCLUDED):")
for r in fatura_reports:
    # Check if our regex catches it
    name = str(r["name"]).strip()
    matches_exclude = re.match(r'^(fatu|farur|cart)', name, re.IGNORECASE) or re.search(r'(fatura|fatuar|fatut|farur)', name, re.IGNORECASE)
    print("  {}  {:<35}  {:<10}  expenses={:>3}  total=R$ {:>10,.2f}  excluded={}".format(
        r["id"], name[:35], r["status"], r["expense_count"], float(r["total"]), bool(matches_exclude)))

# ============================================================
# PART 5: TARIFA — check for post-cutoff transactions
# ============================================================
print("\n" + "=" * 80)
print("PART 5: TARIFA — post-cutoff check for ABNER")
print("=" * 80)

cur.execute("""
    WITH deduped AS (
        SELECT DISTINCT ON (
            UPPER(usuario), data, tipo, valor,
            COALESCE(NULLIF(codigo_transacao, ''), hora::text)
        )
            UPPER(usuario) AS usuario_up, data, tipo, valor, descricao, hora
        FROM extrato_movimentacao
        WHERE is_snapshot = FALSE
          AND tipo = 'Taxa'
          AND UPPER(usuario) = 'ABNER ANDRADE CAVALCANTE'
        ORDER BY UPPER(usuario), data, tipo, valor,
            COALESCE(NULLIF(codigo_transacao, ''), hora::text)
    )
    SELECT data, valor, descricao, hora
    FROM deduped
    WHERE data > '2026-06-30'
    ORDER BY data, hora
""")
post_cutoff = cur.fetchall()
print("Post-cutoff (after 2026-06-30) TARIFA for ABNER: {} rows".format(len(post_cutoff)))
for r in post_cutoff:
    print("  {} {} R$ {:.2f} {}".format(r["data"], r["hora"], float(r["valor"]), r["descricao"]))

# Also check total tarifa for ABNER up to cutoff
cur.execute("""
    WITH deduped AS (
        SELECT DISTINCT ON (
            UPPER(usuario), data, tipo, valor,
            COALESCE(NULLIF(codigo_transacao, ''), hora::text)
        )
            UPPER(usuario) AS usuario_up, data, tipo, valor
        FROM extrato_movimentacao
        WHERE is_snapshot = FALSE
          AND tipo = 'Taxa'
          AND data <= '2026-06-30'
          AND UPPER(usuario) = 'ABNER ANDRADE CAVALCANTE'
        ORDER BY UPPER(usuario), data, tipo, valor,
            COALESCE(NULLIF(codigo_transacao, ''), hora::text)
    )
    SELECT COUNT(*) as cnt, SUM(ABS(valor)) as total
    FROM deduped
""")
r = cur.fetchone()
print("ABNER TARIFA up to 2026-06-30: {} txns, total R$ {:.2f}".format(r["cnt"], float(r["total"])))

# Check reference PAINEL value for ABNER
wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws_p = wb["PAINEL"]
for row in ws_p.iter_rows(min_row=12, values_only=True):
    cpf = str(row[2] or "").strip().replace(".", "").replace("-", "").replace(" ", "").zfill(11) if row[2] else ""
    if cpf == "02027745203":
        print("Ref PAINEL ABNER: TARIFA={:.2f}".format(abs(float(row[15] or 0))))
        break
wb.close()

conn.close()
