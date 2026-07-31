#!/usr/bin/env python3
"""Quick check: compare ref extrato transactions vs neon extrato transactions for a few CPFs."""
import os, sys, unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path
import openpyxl

BASE = Path(__file__).parent.parent

def nc(raw):
    if raw is None: return ""
    s = str(raw).strip().replace(".","").replace("-","").replace("/","").replace(" ","")
    return s.zfill(11)

def nf(raw):
    try: return round(float(raw), 2) if raw is not None else 0.0
    except: return 0.0

def norm(name):
    if not name: return ""
    s = str(name).upper().strip()
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn").strip()

# Load ref extrato with CPF
wb = openpyxl.load_workbook(BASE / "data" / "ref_dump.xlsx", read_only=True, data_only=True)
ws = wb["extrato_ref"]
ref_by_cpf = defaultdict(lambda: {"CARGA": 0, "TRANSFERÊNCIA": 0, "TARIFA": 0, "count": 0})
ref_name_to_cpf = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    cpf = nc(row[0])
    nome = str(row[1] or "").upper()
    tipo = str(row[2] or "")
    valor = nf(row[3])
    if cpf and cpf != "00000000000":
        ref_name_to_cpf[nome] = cpf
        ref_by_cpf[cpf][tipo] += abs(valor)
        ref_by_cpf[cpf]["count"] += 1
wb.close()

# Load neon extrato raw
wb = openpyxl.load_workbook(BASE / "data" / "neon_dump.xlsx", read_only=True, data_only=True)
ws = wb["extrato_raw"]
neon_by_name = defaultdict(lambda: {"carga": 0, "transferencia": 0, "tarifa": 0, "count": 0})
neon_names = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    nome = str(row[0] or "").upper()
    tipo = str(row[2] or "")
    valor = nf(row[3])
    d = row[1]
    # Filter to <= June 30
    if hasattr(d, 'date'):
        d = d.date() if hasattr(d, 'date') else d
    elif isinstance(d, str):
        d = date.fromisoformat(d)
    if d and d <= date(2026, 6, 30):
        # Map Neon tipos to categories
        if tipo == "Transferência" and valor > 0:
            neon_by_name[nome]["carga"] += abs(valor)
        elif tipo == "Transferência" and valor < 0:
            neon_by_name[nome]["transferencia"] += abs(valor)
        elif tipo == "Taxa":
            neon_by_name[nome]["tarifa"] += abs(valor)
        neon_by_name[nome]["count"] += 1
        neon_names.add(nome)

# Load cadastro for name mapping
ws2 = wb["cadastro"]
cad_name_to_cpf = {}
for row in ws2.iter_rows(min_row=2, values_only=True):
    cpf = nc(row[0])
    nome = str(row[1] or "")
    cad_name_to_cpf[norm(nome)] = cpf
    cad_name_to_cpf[norm(nome)[:20]] = cpf
wb.close()

# Build best mapping
def map_name(nome):
    if nome in ref_name_to_cpf: return ref_name_to_cpf[nome]
    n = norm(nome)
    if n in cad_name_to_cpf: return cad_name_to_cpf[n]
    if n[:20] in cad_name_to_cpf: return cad_name_to_cpf[n[:20]]
    if n[:15] in cad_name_to_cpf: return cad_name_to_cpf[n[:15]]
    if n[:10] in cad_name_to_cpf: return cad_name_to_cpf[n[:10]]
    return None

# Compare per CPF for CARGA
print("=== CARGA comparison (top divergences) ===")
divs = []
for nome in neon_names:
    cpf = map_name(nome)
    if not cpf or cpf not in ref_by_cpf: continue
    n_carga = neon_by_name[nome]["carga"]
    r_carga = ref_by_cpf[cpf].get("CARGA", 0)
    diff = n_carga - r_carga
    if abs(diff) > 0.05:
        divs.append((cpf, nome[:25], r_carga, n_carga, diff, neon_by_name[nome]["count"], ref_by_cpf[cpf]["count"]))

for cpf, nome, rc, nc_, diff, ncnt, rcnt in sorted(divs, key=lambda x: abs(x[4]), reverse=True)[:15]:
    print(f"  {cpf} {nome:<25} ref={rc:>12.2f}({rcnt}tx) neon={nc_:>12.2f}({ncnt}tx) diff={diff:>+12.2f}")

print(f"\n  Total divergences: {len(divs)}")
print(f"  Total ref CARGA: {sum(ref_by_cpf[c]['CARGA'] for c in ref_by_cpf):,.2f}")
print(f"  Total neon CARGA (mapped): {sum(neon_by_name[n]['carga'] for n in neon_names if map_name(n) and map_name(n) in ref_by_cpf):,.2f}")

# Check if neon has transactions that ref doesn't (by date range)
print("\n=== Date range check ===")
wb = openpyxl.load_workbook(BASE / "data" / "neon_dump.xlsx", read_only=True, data_only=True)
ws = wb["extrato_raw"]
neon_dates = defaultdict(int)
for row in ws.iter_rows(min_row=2, values_only=True):
    d = row[1]
    if hasattr(d, 'date'):
        d = d.date() if hasattr(d, 'date') else d
    elif isinstance(d, str):
        d = date.fromisoformat(d)
    if d and d <= date(2026, 6, 30):
        month_key = f"{d.year}-{d.month:02d}"
        neon_dates[month_key] += 1
wb.close()

wb = openpyxl.load_workbook(BASE / "data" / "ref_dump.xlsx", read_only=True, data_only=True)
ws = wb["extrato_ref"]
ref_dates = defaultdict(int)
for row in ws.iter_rows(min_row=2, values_only=True):
    d = row[4]
    if d:
        parts = d.split("-")
        if len(parts) == 3:
            month_key = f"{parts[0]}-{parts[1]}"
            ref_dates[month_key] += 1
wb.close()

print("  Month | Ref | Neon")
all_months = sorted(set(list(ref_dates.keys()) + list(neon_dates.keys())))
for m in all_months:
    r = ref_dates.get(m, 0)
    n = neon_dates.get(m, 0)
    flag = " <<<" if abs(n - r) > 100 else ""
    print(f"  {m} | {r:>6} | {n:>6}{flag}")
