#!/usr/bin/env python3
"""
Direct comparison of reference extrato vs Neon extrato (deduplicated).
Shows exactly which transactions are missing/different per CPF.
"""
import unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path
import openpyxl

BASE = Path(__file__).parent.parent

def nc(raw):
    if raw is None: return ""
    s = str(raw).strip().replace(".","").replace("-","").replace("/","").replace(" ","")
    return s.zfill(11)

def nf(raw):
    try: return round(float(raw), 2) if raw is not None else 0.0
    except: return 0.0

def norm(name):
    if not name: return ""
    s = str(name).upper().strip()
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn").strip()

def bigrams(s):
    return set(s[i:i+2] for i in range(len(s)-1))

def fuzzy_ratio(a, b):
    if a == b: return 1.0
    if not a or not b: return 0.0
    ba, bb = bigrams(a), bigrams(b)
    inter = len(ba & bb)
    return (2 * inter) / (len(ba) + len(bb))

# === Load reference extrato ===
wb = openpyxl.load_workbook(BASE / "data" / "ref_dump.xlsx", read_only=True, data_only=True)
ws = wb["extrato_ref"]
ref_by_cpf = defaultdict(lambda: {"CARGA": 0, "TRANSFERÊNCIA": 0, "TARIFA": 0})
ref_types = set()
ref_dates = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    cpf = nc(row[0])
    usuario = str(row[1] or "").upper()
    tipo = str(row[2] or "").upper().strip()
    valor = nf(row[3])
    d = str(row[4] or "")
    ref_types.add(tipo)
    ref_dates.add(d[:10] if len(d) >= 10 else d)
    if cpf and cpf != "00000000000":
        if tipo == "CARGA":
            ref_by_cpf[cpf]["CARGA"] += valor
        elif tipo == "TRANSFERÊNCIA":
            ref_by_cpf[cpf]["TRANSFERÊNCIA"] += abs(valor)
        elif tipo == "TARIFA":
            ref_by_cpf[cpf]["TARIFA"] += abs(valor)
wb.close()
print(f"Ref extrato: {len(ref_by_cpf)} CPFs, types={ref_types}")
print(f"Ref date range: {min(ref_dates)} to {max(ref_dates)}")
print(f"Ref totals: CARGA={sum(v['CARGA'] for v in ref_by_cpf.values()):,.2f} "
      f"TRANSF={sum(v['TRANSFERÊNCIA'] for v in ref_by_cpf.values()):,.2f} "
      f"TARIFA={sum(v['TARIFA'] for v in ref_by_cpf.values()):,.2f}")

# === Load Neon extrato (deduplicated) ===
wb = openpyxl.load_workbook(BASE / "data" / "neon_dump.xlsx", read_only=True, data_only=True)
ws = wb["extrato_raw"]
seen = set()
neon_by_name = defaultdict(lambda: {"CARGA": 0, "TRANSFERÊNCIA": 0, "TARIFA": 0})
neon_types = set()
neon_dates = set()
cutoff = date(2026, 6, 30)

for row in ws.iter_rows(min_row=2, values_only=True):
    nome = str(row[0] or "").upper()
    d = row[1]
    if hasattr(d, 'date'):
        d = d.date() if hasattr(d, 'date') else d
    elif isinstance(d, str):
        d = date.fromisoformat(d)
    tipo = str(row[2] or "")
    valor = nf(row[3])
    cod = str(row[4] or "") if len(row) > 4 else ""
    
    if not d or d > cutoff:
        continue
    
    dedup_key = (nome, d, tipo, valor, cod)
    if dedup_key in seen:
        continue
    seen.add(dedup_key)
    
    neon_types.add(tipo)
    neon_dates.add(str(d))
    
    if tipo == "Transferência" and valor > 0:
        neon_by_name[nome]["CARGA"] += valor
    elif tipo == "Transferência" and valor < 0:
        neon_by_name[nome]["TRANSFERÊNCIA"] += abs(valor)
    elif tipo == "Taxa":
        neon_by_name[nome]["TARIFA"] += abs(valor)
wb.close()
print(f"\nNeon extrato: {len(seen)} unique rows, types={neon_types}")
print(f"Neon date range: {min(neon_dates)} to {max(neon_dates)}")
print(f"Neon totals: CARGA={sum(v['CARGA'] for v in neon_by_name.values()):,.2f} "
      f"TRANSF={sum(v['TRANSFERÊNCIA'] for v in neon_by_name.values()):,.2f} "
      f"TARIFA={sum(v['TARIFA'] for v in neon_by_name.values()):,.2f}")

# === Build name mapping ===
wb = openpyxl.load_workbook(BASE / "data" / "neon_dump.xlsx", read_only=True, data_only=True)
ws_c = wb["cadastro"]
cad_name_to_cpf = {}
for row in ws_c.iter_rows(min_row=2, values_only=True):
    cpf = nc(row[0])
    nome = str(row[1] or "")
    cad_name_to_cpf[norm(nome)] = cpf
wb.close()

wb = openpyxl.load_workbook(BASE / "data" / "ref_dump.xlsx", read_only=True, data_only=True)
ws_p = wb["painel"]
ref_name_to_cpf = {}
for row in ws_p.iter_rows(min_row=2, values_only=True):
    cpf = nc(row[0])
    nome = str(row[1] or "")
    ref_name_to_cpf[norm(nome)] = cpf
wb.close()

def map_name(nome):
    n = norm(nome)
    if n in ref_name_to_cpf: return ref_name_to_cpf[n]
    if n in cad_name_to_cpf: return cad_name_to_cpf[n]
    if len(n) >= 15:
        p15 = n[:15]
        for cn, cpf in cad_name_to_cpf.items():
            if cn[:15] == p15: return cpf
    if len(n) >= 10:
        p10 = n[:10]
        for cn, cpf in cad_name_to_cpf.items():
            if cn[:10] == p10: return cpf
    best_cpf = None
    best_ratio = 0
    for cn, cpf in cad_name_to_cpf.items():
        r = fuzzy_ratio(n, cn)
        if r > best_ratio:
            best_ratio = r
            best_cpf = cpf
    if best_ratio >= 0.88 and best_cpf:
        return best_cpf
    return None

# === Map neon by name to by CPF ===
neon_by_cpf = defaultdict(lambda: {"CARGA": 0, "TRANSFERÊNCIA": 0, "TARIFA": 0})
unmapped = set()
for nome, vals in neon_by_name.items():
    cpf = map_name(nome)
    if cpf:
        neon_by_cpf[cpf]["CARGA"] += vals["CARGA"]
        neon_by_cpf[cpf]["TRANSFERÊNCIA"] += vals["TRANSFERÊNCIA"]
        neon_by_cpf[cpf]["TARIFA"] += vals["TARIFA"]
    else:
        unmapped.add(nome)

print(f"\nMapped neon: {len(neon_by_cpf)} CPFs, unmapped: {len(unmapped)} names")

# === Compare per CPF ===
all_cpfs = set(ref_by_cpf.keys()) | set(neon_by_cpf.keys())
fields = ["CARGA", "TRANSFERÊNCIA", "TARIFA"]
match = {f: 0 for f in fields}
divs = {f: [] for f in fields}

for cpf in all_cpfs:
    rv = ref_by_cpf.get(cpf, {"CARGA": 0, "TRANSFERÊNCIA": 0, "TARIFA": 0})
    nv = neon_by_cpf.get(cpf, {"CARGA": 0, "TRANSFERÊNCIA": 0, "TARIFA": 0})
    for f in fields:
        diff = abs(nv[f] - rv[f])
        if diff <= 0.50:
            match[f] += 1
        else:
            divs[f].append((cpf, rv[f], nv[f], diff))

print(f"\n=== Match Results (tolerance R$0.50) ===")
print(f"Total CPFs: {len(all_cpfs)}")
for f in fields:
    pct = match[f] / len(all_cpfs) * 100
    print(f"  {f:<15}: {match[f]}/{len(all_cpfs)} = {pct:.1f}%")

print(f"\n=== Top divergences ===")
for f in fields:
    d = divs[f]
    if not d: continue
    print(f"\n  {f} ({len(d)} divergences):")
    for cpf, rv, nv, diff in sorted(d, key=lambda x: x[3], reverse=True)[:10]:
        print(f"    {cpf} ref={rv:>12.2f} neon={nv:>12.2f} diff={diff:>+10.2f}")

if unmapped:
    print(f"\n=== Unmapped names ({len(unmapped)}) ===")
    for n in sorted(unmapped)[:10]:
        print(f"  {n}")
