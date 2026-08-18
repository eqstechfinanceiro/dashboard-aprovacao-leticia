#!/usr/bin/env python3
"""
compare_with_controle.py
------------------------
Compara o snapshot calculado com a planilha CONTROLE de referência.
Mostra divergências campo por campo.
"""
import argparse
import logging
import os
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

BASE = Path(__file__).parent.parent
load_dotenv(BASE / ".env")

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
logger = logging.getLogger(__name__)

NEON_URL = os.getenv("NEON_DATABASE_URL")


def _r2(v: float) -> float:
    return float(Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def nc(raw) -> str:
    if raw is None:
        return ""
    s = str(raw).strip().replace(".", "").replace("-", "").replace("/", "").replace(" ", "")
    return s.zfill(11)


def nf(raw) -> float:
    try:
        return round(float(raw), 2) if raw is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


def main():
    parser = argparse.ArgumentParser(description="Compara snapshot calculado com planilha CONTROLE")
    parser.add_argument("--ref", type=str, required=True, help="Caminho para planilha CONTROLE")
    parser.add_argument("--ano", type=int, required=True)
    parser.add_argument("--mes", type=int, required=True)
    parser.add_argument("--quinzena", type=int, choices=[1, 2], required=True)
    args = parser.parse_args()

    # Load reference planilha
    ref_path = args.ref
    if not os.path.isabs(ref_path):
        ref_path = str(BASE / ref_path)

    logger.info(f"Carregando planilha: {ref_path}")
    wb = openpyxl.load_workbook(ref_path, read_only=True, data_only=True)
    ws = wb["PAINEL"]

    # Header at row 11, data from row 12
    # Col mapping (0-indexed) from header:
    # 0=EMPRESA, 1=COLABORADOR, 2=CPF, 3=CHAVE, 4=SITUAÇÃO, 5=STATUS CARTÃO,
    # 6=CARTÃO ITAÚ, 7=TERMO, 8=REGIONAL, 9=CENTRO CUSTO, 10=GESTOR, 11=DIRETOR,
    # 12=CARTÃO VEXPENSES, 13=CARGA, 14=TRANSFERENCIA, 15=TARIFA,
    # 16=PRESTAÇÃO, 17=SALDO PRESTAÇÃO, 18=SALDO CARTAO, 19=SALDO FINAL

    header = list(ws.iter_rows(min_row=11, max_row=11, values_only=True))[0]
    print(f"\nHeader PAINEL: {[str(h)[:20] if h else None for h in header[:20]]}")

    ref = {}
    for row in ws.iter_rows(min_row=12, values_only=True):
        cpf = nc(row[2])
        if not cpf or cpf == "00000000000":
            continue
        ref[cpf] = {
            "colaborador": str(row[1] or "")[:30],
            "carga": nf(row[13]),
            "transferencia": nf(row[14]),
            "tarifa": nf(row[15]),
            "prestacao": nf(row[16]),
            "saldo_prestacao": nf(row[17]),
            "saldo_cartao": nf(row[18]),
            "saldo_final": nf(row[19]),
        }
    wb.close()

    logger.info(f"Planilha: {len(ref)} CPFs")

    # Load calculated snapshot from DB
    sys.path.insert(0, str(BASE / "src"))
    from calcular_quinzena_neon import calcular_quinzena

    linhas = calcular_quinzena(args.ano, args.mes, args.quinzena)
    calc = {l["cpf"]: l for l in linhas}

    logger.info(f"Calculado: {len(calc)} CPFs")

    # Compare
    cpfs_ref = set(ref.keys())
    cpfs_calc = set(calc.keys())
    tol = 0.05

    campos = ["carga", "transferencia", "tarifa", "prestacao", "saldo_prestacao", "saldo_cartao", "saldo_final"]
    divergencias = {c: [] for c in campos}
    total_comparados = 0

    for cpf in sorted(cpfs_ref & cpfs_calc):
        total_comparados += 1
        r = ref[cpf]
        c = calc[cpf]
        for campo in campos:
            v_ref = r[campo]
            v_calc = c.get(f"_{campo}", c.get(campo, 0.0))
            # Map calc fields to ref fields
            if campo == "carga":
                v_calc = c["_carga_extrato"]
            elif campo == "transferencia":
                v_calc = c["_transferencia_extrato"]
            elif campo == "tarifa":
                v_calc = c["_tarifa_extrato"]
            elif campo == "prestacao":
                v_calc = c["_prestacao_somase"]
            elif campo == "saldo_prestacao":
                v_calc = c["_saldo_prestacao"]
            elif campo == "saldo_cartao":
                v_calc = c["_saldo_cartao_controle"]
            elif campo == "saldo_final":
                v_calc = c["saldo_final"]

            diff = abs(v_calc - v_ref)
            if diff > tol:
                divergencias[campo].append((cpf, c["colaborador"][:25], v_ref, v_calc, v_calc - v_ref))

    print()
    print("=" * 80)
    print(f"  COMPARAÇÃO — {args.quinzena}ª QZ {args.mes:02d}/{args.ano}")
    print(f"  Planilha: {ref_path}")
    print(f"  CPFs na planilha: {len(cpfs_ref)} | CPFs calculados: {len(cpfs_calc)} | Comparados: {total_comparados}")
    print(f"  Apenas na planilha: {len(cpfs_ref - cpfs_calc)} | Apenas no cálculo: {len(cpfs_calc - cpfs_ref)}")
    print("=" * 80)

    for campo in campos:
        divs = divergencias[campo]
        match = total_comparados - len(divs)
        pct = match / total_comparados * 100 if total_comparados else 0
        status = "✅" if len(divs) == 0 else ("⚠️ " if len(divs) <= 10 else "❌")
        print(f"  {status} {campo:<22}: {match:>4}/{total_comparados}  ({pct:5.1f}%)  divergências: {len(divs)}")
        if divs:
            for cpf, nome, vr, vc, d in sorted(divs, key=lambda x: abs(x[4]), reverse=True)[:10]:
                print(f"       {cpf} {nome:<25} ref={vr:>12.2f}  calc={vc:>12.2f}  diff={d:>+12.2f}")

    print("=" * 80)

    # Totals
    total_carga_ref = sum(r["carga"] for r in ref.values())
    total_carga_calc = sum(c["_carga_extrato"] for c in calc.values() if c["cpf"] in ref)
    total_prest_ref = sum(r["prestacao"] for r in ref.values())
    total_prest_calc = sum(c["_prestacao_somase"] for c in calc.values() if c["cpf"] in ref)
    total_sp_ref = sum(r["saldo_prestacao"] for r in ref.values())
    total_sp_calc = sum(c["_saldo_prestacao"] for c in calc.values() if c["cpf"] in ref)
    total_sf_ref = sum(r["saldo_final"] for r in ref.values())
    total_sf_calc = sum(c["saldo_final"] for c in calc.values() if c["cpf"] in ref)

    print(f"  Total CARGA          ref: R$ {total_carga_ref:>14,.2f}  calc: R$ {total_carga_calc:>14,.2f}  diff: R$ {total_carga_calc - total_carga_ref:>+14,.2f}")
    print(f"  Total PRESTAÇÃO      ref: R$ {total_prest_ref:>14,.2f}  calc: R$ {total_prest_calc:>14,.2f}  diff: R$ {total_prest_calc - total_prest_ref:>+14,.2f}")
    print(f"  Total SALDO PREST    ref: R$ {total_sp_ref:>14,.2f}  calc: R$ {total_sp_calc:>14,.2f}  diff: R$ {total_sp_calc - total_sp_ref:>+14,.2f}")
    print(f"  Total SALDO FINAL    ref: R$ {total_sf_ref:>14,.2f}  calc: R$ {total_sf_calc:>14,.2f}  diff: R$ {total_sf_calc - total_sf_ref:>+14,.2f}")
    print("=" * 80)


if __name__ == "__main__":
    main()
