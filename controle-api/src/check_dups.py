#!/usr/bin/env python3
"""Check for duplicates in Neon extrato and compare types with ref."""
import os, sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
import openpyxl

BASE = Path(__file__).parent.parent

# Load neon extrato raw
wb = openpyxl.load_workbook(BASE / "data" / "neon_dump.xlsx", read_only=True, data_only=True)
ws = wb["extrato_raw"]

# Count by tipo
tipo_count = Counter()
all_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    nome = str(row[0] or "").upper()
    d = row[1]
    if hasattr(d, 'date'):
        d = d.date() if hasattr(d, 'date') else d
    elif isinstance(d, str):
        d = date.fromisoformat(d)
    tipo = str(row[2] or "")
    valor = row[3]
    tipo_count[tipo] += 1
    all_rows.append((nome, d, tipo, valor))

print("=== Neon extrato types ===")
for t, c in tipo_count.most_common():
    print(f"  {t:<30} {c:>6}")

# Check for duplicates: same (nome, data, tipo, valor)
row_count = Counter((r[0], r[1], r[2], r[3]) for r in all_rows)
dups = {k: v for k, v in row_count.items() if v > 1}
print(f"\n=== Duplicate check ===")
print(f"  Total rows: {len(all_rows)}")
print(f"  Unique rows: {len(row_count)}")
print(f"  Duplicate keys: {len(dups)}")
print(f"  Duplicate rows (extra): {sum(v-1 for v in dups.values())}")

if dups:
    print(f"\n  Top 10 duplicates:")
    for (nome, d, tipo, val), cnt in sorted(dups.items(), key=lambda x: x[1], reverse=True)[:10]:
        print(f"    {d} {nome[:20]:<20} {tipo:<15} {val:>10.2f} x{cnt}")

# Now check ref types
wb.close()
wb = openpyxl.load_workbook(BASE / "data" / "ref_dump.xlsx", read_only=True, data_only=True)
ws = wb["extrato_ref"]
ref_tipo_count = Counter()
for row in ws.iter_rows(min_row=2, values_only=True):
    tipo = str(row[2] or "")
    ref_tipo_count[tipo] += 1

print(f"\n=== Ref extrato types ===")
for t, c in ref_tipo_count.most_common():
    print(f"  {t:<30} {c:>6}")
wb.close()

# Filter neon to only same types as ref and check counts
ref_types = set(ref_tipo_count.keys())
neon_filtered = [r for r in all_rows if r[2] in ref_types or 
                 (r[2] == "Transferência" and "TRANSFERÊNCIA" in ref_types) or
                 (r[2] == "Taxa" and "TARIFA" in ref_types)]
print(f"\n=== Neon filtered to ref types ===")
print(f"  Total: {len(neon_filtered)}")
# Deduplicate
neon_dedup = set((r[0], r[1], r[2], r[3]) for r in neon_filtered)
print(f"  Unique: {len(neon_dedup)}")
