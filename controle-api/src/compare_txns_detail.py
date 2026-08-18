#!/usr/bin/env python3
"""
Transaction-level comparison for specific CPFs.
Shows exactly which transactions are in ref but not in neon (or vice versa).
"""
import unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path
import openpyxl

BASE = Path(__file__).parent.parent

def nc(raw):
    if raw is None: return ""
    s = str(raw).strip().replace(".","").replace("-","").replace("/","").replace(" ","")
    return s.zfill(11)

def nf(raw):
    try: return round(float(raw), 2) if raw is not None else 0.0
    except: return 0.0

def norm(name):
    if not name: return ""
    s = str(name).upper().strip()
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn").strip()

def bigrams(s):
    return set(s[i:i+2] for i in range(len(s)-1))

def fuzzy_ratio(a, b):
    if a == b: return 1.0
    if not a or not b: return 0.0
    ba, bb = bigrams(a), bigrams(b)
    inter = len(ba & bb)
    return (2 * inter) / (len(ba) + len(bb))

# Target CPFs to investigate
TARGET_CPFS = {
    "02027745203": "ABNER",
    "01677920599": "RAFAEL AMORIM",
    "60023937220": "ANA CRISTINA",
    "68187483920": "LUIS CARLOS BROERING",
    "74190920959": "LUIZ CARLOS NETO",
}

# === Load ref extrato for target CPFs ===
wb = openpyxl.load_workbook(BASE / "data" / "ref_dump.xlsx", read_only=True, data_only=True)
ws = wb["extrato_ref"]
ref_txns = defaultdict(list)  # cpf -> [(data, tipo, valor)]
for row in ws.iter_rows(min_row=2, values_only=True):
    cpf = nc(row[0])
    if cpf not in TARGET_CPFS:
        continue
    usuario = str(row[1] or "").upper()
    tipo = str(row[2] or "").upper().strip()
    valor = nf(row[3])
    d = str(row[4] or "")[:10]
    ref_txns[cpf].append((d, tipo, valor))
wb.close()

# === Load cadastro for name mapping ===
wb = openpyxl.load_workbook(BASE / "data" / "neon_dump.xlsx", read_only=True, data_only=True)
ws_c = wb["cadastro"]
cpf_to_name = {}
cad_name_to_cpf = {}
for row in ws_c.iter_rows(min_row=2, values_only=True):
    cpf = nc(row[0])
    nome = str(row[1] or "")
    cpf_to_name[cpf] = norm(nome)
    cad_name_to_cpf[norm(nome)] = cpf

# Also load ref names
wb_ref = openpyxl.load_workbook(BASE / "data" / "ref_dump.xlsx", read_only=True, data_only=True)
ws_p = wb_ref["painel"]
ref_name_to_cpf = {}
for row in ws_p.iter_rows(min_row=2, values_only=True):
    cpf = nc(row[0])
    nome = str(row[1] or "")
    ref_name_to_cpf[norm(nome)] = cpf
wb_ref.close()

def map_name(nome):
    n = norm(nome)
    if n in ref_name_to_cpf: return ref_name_to_cpf[n]
    if n in cad_name_to_cpf: return cad_name_to_cpf[n]
    if len(n) >= 15:
        p15 = n[:15]
        for cn, cpf in cad_name_to_cpf.items():
            if cn[:15] == p15: return cpf
    if len(n) >= 10:
        p10 = n[:10]
        for cn, cpf in cad_name_to_cpf.items():
            if cn[:10] == p10: return cpf
    best_cpf = None
    best_ratio = 0
    for cn, cpf in cad_name_to_cpf.items():
        r = fuzzy_ratio(n, cn)
        if r > best_ratio:
            best_ratio = r
            best_cpf = cpf
    if best_ratio >= 0.88 and best_cpf:
        return best_cpf
    return None

# === Load Neon extrato for target CPFs ===
ws = wb["extrato_raw"]
seen = set()
neon_txns = defaultdict(list)  # cpf -> [(data, tipo, valor)]
cutoff = date(2026, 6, 30)

for row in ws.iter_rows(min_row=2, values_only=True):
    nome = str(row[0] or "").upper()
    d = row[1]
    if hasattr(d, 'date'):
        d = d.date() if hasattr(d, 'date') else d
    elif isinstance(d, str):
        d = date.fromisoformat(d)
    tipo = str(row[2] or "")
    valor = nf(row[3])
    cod = str(row[4] or "") if len(row) > 4 else ""
    
    if not d or d > cutoff:
        continue
    
    dedup_key = (nome, d, tipo, valor, cod)
    if dedup_key in seen:
        continue
    seen.add(dedup_key)
    
    cpf = map_name(nome)
    if cpf not in TARGET_CPFS:
        continue
    
    # Map neon types to ref types
    if tipo == "Transferência" and valor > 0:
        neon_txns[cpf].append((str(d), "CARGA", valor))
    elif tipo == "Transferência" and valor < 0:
        neon_txns[cpf].append((str(d), "TRANSFERÊNCIA", abs(valor)))
    elif tipo == "Taxa":
        neon_txns[cpf].append((str(d), "TARIFA", abs(valor)))

wb.close()

# === Compare for each target CPF ===
for cpf, label in TARGET_CPFS.items():
    ref_list = ref_txns.get(cpf, [])
    neon_list = neon_txns.get(cpf, [])
    
    print(f"\n{'='*70}")
    print(f"CPF: {cpf} ({label})")
    print(f"Ref: {len(ref_list)} txns, Neon: {len(neon_list)} txns")
    
    # Create sets of (data, tipo, valor) for comparison
    ref_set = set((d, t, round(v, 2)) for d, t, v in ref_list)
    neon_set = set((d, t, round(v, 2)) for d, t, v in neon_list)
    
    only_ref = ref_set - neon_set
    only_neon = neon_set - ref_set
    
    ref_total = sum(v for _, _, v in ref_list)
    neon_total = sum(v for _, _, v in neon_list)
    print(f"Ref total: R$ {ref_total:,.2f}, Neon total: R$ {neon_total:,.2f}, Diff: R$ {ref_total - neon_total:,.2f}")
    
    if only_ref:
        print(f"\n  ONLY IN REF ({len(only_ref)} txns):")
        for d, t, v in sorted(only_ref):
            print(f"    {d} {t:<15} R$ {v:>10.2f}")
    
    if only_neon:
        print(f"\n  ONLY IN NEON ({len(only_neon)} txns):")
        for d, t, v in sorted(only_neon):
            print(f"    {d} {t:<15} R$ {v:>10.2f}")
    
    if not only_ref and not only_neon:
        print("  PERFECT MATCH!")
