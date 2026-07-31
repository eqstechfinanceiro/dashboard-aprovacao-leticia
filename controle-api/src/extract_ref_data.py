#!/usr/bin/env python3
"""
extract_ref_data.py
-------------------
Extracts EXTRATO (with CPF), BASE PREST, and PAINEL from the reference CONTROLE sheet
into a local Excel file for offline comparison.
"""
import os
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE = Path(__file__).parent.parent

def nc(raw):
    if raw is None:
        return ""
    s = str(raw).strip().replace(".", "").replace("-", "").replace("/", "").replace(" ", "")
    return s.zfill(11)

def nf(raw):
    try:
        return round(float(raw), 2) if raw is not None else 0.0
    except (ValueError, TypeError):
        return 0.0

def excel_serial_to_date(serial):
    if isinstance(serial, (int, float)):
        # Excel epoch: 1900-01-01 = 1, but with the 1900 leap year bug
        base = date(1899, 12, 30)
        days = int(serial)
        try:
            return base.toordinal() + days
        except:
            return None
    return None

def main():
    ref_path = sys.argv[1] if len(sys.argv) > 1 else str(BASE / "data" / "CONTROLE - VEXPENSES - JULHO 2026 - ATUALIZADA PARA COMPARAR.xlsx")
    if not os.path.isabs(ref_path):
        ref_path = str(BASE / ref_path)

    print(f"Loading: {ref_path}")
    wb = openpyxl.load_workbook(ref_path, read_only=True, data_only=True)

    out_path = BASE / "data" / "ref_dump.xlsx"
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="1F4E79")

    def write_sheet(name, rows, headers):
        ws = wb_out.create_sheet(name)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center")
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
        print(f"  {name}: {len(rows)} rows")

    # 1. PAINEL — already known: header row 11, data from row 12
    ws = wb["PAINEL"]
    painel = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        cpf = nc(row[2])
        if not cpf or cpf == "00000000000":
            continue
        painel.append((
            cpf,
            str(row[1] or ""),
            nf(row[13]),  # CARGA
            nf(row[14]),  # TRANSFERENCIA
            nf(row[15]),  # TARIFA
            nf(row[16]),  # PRESTAÇÃO
            nf(row[17]),  # SALDO PRESTAÇÃO
            nf(row[18]),  # SALDO CARTAO
            nf(row[19]),  # SALDO FINAL
            nf(row[20]),  # 1ª QZ
            nf(row[21]),  # 2ª QZ
        ))
    write_sheet("painel", painel,
                ["cpf", "colaborador", "carga", "transferencia", "tarifa",
                 "prestacao", "saldo_prestacao", "saldo_cartao", "saldo_final",
                 "col_1qz", "col_2qz"])

    # 2. EXTRATO — header at row 8, data from row 9
    # Cols: (None), ANO, MÊS, Data, Hora, Código, Cartão, Grupo, Usuário, Tipo, Descrição, Valor, CPF
    ws = wb["EXTRATO"]
    extrato = []
    for row in ws.iter_rows(min_row=9, values_only=True):
        if row[1] is None and row[2] is None:
            continue
        cpf = nc(row[12])
        nome = str(row[8] or "").upper()
        tipo = str(row[9] or "")
        valor = nf(row[11])
        data_serial = row[3]
        data_val = None
        if isinstance(data_serial, (int, float)):
            from datetime import timedelta
            base = datetime(1899, 12, 30)
            data_val = (base + timedelta(days=int(data_serial))).date()
        extrato.append((cpf, nome, tipo, valor, str(data_val) if data_val else ""))
    write_sheet("extrato_ref", extrato,
                ["cpf", "usuario", "tipo", "valor", "data"])

    # 3. BASE PREST — header at row 2 (index 2)
    # col0=ID Despesa, col1=ID Relatorio, col2=Nome Relatorio, col3=Data,
    # col9=CPF/CNPJ, col11=Data Pagamento, col26=Valor
    ws = wb["BASE PREST "]
    base_prest = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        cpf = nc(row[9])
        valor = nf(row[26])
        base_prest.append((cpf, valor, str(row[3] or ""), str(row[0])))
    write_sheet("base_prest_ref", base_prest,
                ["cpf", "valor", "data", "id_despesa"])

    # 4. SALDO CARTAO sheet
    ws = wb["SALDO CARTAO"]
    # Find header row
    header_row = None
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        if row[0] and "CPF" in str(row[0]).upper():
            header_row = ri
            break
    print(f"  SALDO CARTAO header at row {header_row}")
    saldo_cartao = []
    if header_row:
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            cpf = nc(row[0])
            if not cpf:
                continue
            # Find valor column — try col 1 and col 2
            valor = nf(row[1]) if row[1] is not None else nf(row[2])
            saldo_cartao.append((cpf, valor))
    write_sheet("saldo_cartao_ref", saldo_cartao,
                ["cpf", "valor"])

    wb.close()

    wb_out.save(out_path)
    print(f"\n✅ Saved: {out_path}")

    # Print summary
    print("\n=== SUMMARY ===")
    print(f"  PAINEL: {len(painel)} CPFs")
    print(f"  EXTRATO: {len(extrato)} rows, {len(set(e[0] for e in extrato if e[0]))} CPFs")
    print(f"  BASE PREST: {len(base_prest)} rows, {len(set(b[0] for b in base_prest if b[0]))} CPFs")
    print(f"  SALDO CARTAO: {len(saldo_cartao)} CPFs")

    # Totals from extrato
    total_carga = sum(e[3] for e in extrato if e[2] == "CARGA")
    total_transf = sum(abs(e[3]) for e in extrato if e[2] == "TRANSFERÊNCIA")
    total_tarifa = sum(abs(e[3]) for e in extrato if e[2] == "TARIFA")
    print(f"  EXTRATO totals: CARGA={total_carga:,.2f} TRANSF={total_transf:,.2f} TARIFA={total_tarifa:,.2f}")

    # Totals from base prest
    total_prest = sum(b[1] for b in base_prest)
    print(f"  BASE PREST total: {total_prest:,.2f}")


if __name__ == "__main__":
    main()
