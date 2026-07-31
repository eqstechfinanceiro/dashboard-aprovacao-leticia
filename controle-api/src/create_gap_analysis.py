#!/usr/bin/env python3
"""Create comprehensive Excel comparison of ref vs Neon prestação data."""
import os, sys, json, unicodedata
from collections import defaultdict
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
sys.path.insert(0, str(BASE / "src"))

REF_PATH = r"C:\Users\italo.medrado\Desktop\Projects\planilha de carga\data\CONTROLE - VEXPENSES - JULHO 2026 - ATUALIZADA PARA COMPARAR.xlsx"
NEON_DUMP = BASE.parent / "data" / "neon_dump.xlsx"
OUTPUT = BASE.parent / "data" / "gap_analysis_ref_vs_neon.xlsx"

def nc(raw):
    if raw is None: return ""
    s = str(raw).strip().replace(".", "").replace("-", "").replace("/", "").replace(" ", "")
    return s.zfill(11)

def nf(raw):
    try: return round(float(raw), 2) if raw is not None else 0.0
    except: return 0.0

def norm_name(s):
    if not s: return ""
    s = str(s).upper().strip()
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn").strip()

def _is_fatura_or_cartao(name):
    n = name.strip().upper()
    if "CAIXA ITAU" in n or "CAIXA ITAÚ" in n: return True
    if n.startswith("CAIXA"): return False
    if n.startswith(("FATURA","CARTAO","CARTÃO","FATUAR","FARTUR","FATUT","FARUR","FATUTR")): return True
    if "CARTÃO DE CRÉDITO" in n or "CARTAO DE CREDITO" in n or "CARTÃO DE CREDITO" in n: return True
    if "CARTÃO CORPORATIVO" in n: return True
    if ("ITAU" in n or "ITAÚ" in n) and "CAIXA" not in n: return True
    if "DOLAR" in n or "DÓLAR" in n: return True
    if n.startswith("DESPESA") and "FATURA" in n: return True
    if n.startswith("COMPLEMENTAR") and "FATURA" in n: return True
    if "CARTÃO" in n and "CRÉDITO" in n: return True
    if "CARTAO" in n and "CREDITO" in n: return True
    if n.startswith("CARTÃO VEXPENSES"): return True
    return False

# ============================================================
# 1. Load REF BASE PREST (expense level)
# ============================================================
print("Loading ref BASE PREST...")
wb_ref = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)
ws = wb_ref["BASE PREST "]

ref_expenses = {}  # eid -> expense data
for row in ws.iter_rows(min_row=4, values_only=True):
    if not row[0]: continue
    eid = int(row[0])
    rid = int(row[1]) if row[1] else None
    report_name = str(row[2] or "").strip()
    status = str(row[10] or "").strip().upper()
    is_fatura = report_name.upper().startswith("FATURA") or report_name.upper().startswith("CARTAO")
    if is_fatura: continue
    
    ref_expenses[eid] = {
        "report_id": rid,
        "report_name": report_name,
        "user_name": str(row[5] or "").strip(),
        "user_cpf": nc(row[9]),
        "status": status,
        "value": nf(row[26]),
        "date": str(row[6] or "")[:10],
        "description": str(row[7] or "").strip(),
        "payment": str(row[18] or "").strip(),
        "cost_center": str(row[17] or "").strip(),
    }
wb_ref.close()
print(f"  Ref expenses (non-FATURA): {len(ref_expenses)}")

# ============================================================
# 2. Load NEON (expense level) — direct from DB for full details
# ============================================================
print("Loading Neon from database...")
import psycopg2, psycopg2.extras
from dotenv import load_dotenv
load_dotenv(BASE.parent / ".env")
conn = psycopg2.connect(os.getenv("NEON_DATABASE_URL"), connect_timeout=10)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

cur.execute("""
    SELECT r.id as rid, r.name, r.status, r.user_cpf, r.user_name, r.created_at
    FROM prestacao_reports r
    WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND r.user_cpf IS NOT NULL
""")
neon_reports = {}
for row in cur.fetchall():
    rid = row["rid"]
    name = str(row["name"] or "").strip()
    if _is_fatura_or_cartao(name): continue
    neon_reports[rid] = {
        "name": name, "status": str(row["status"] or "").upper(),
        "cpf": nc(row["user_cpf"]), "user_name": str(row["user_name"] or "").strip(),
        "created": str(row["created_at"] or "")[:10],
    }

if neon_reports:
    placeholders = ",".join(["%s"] * len(neon_reports))
    cur.execute(f"""
        SELECT e.id, e.report_id, e.value, e.date, e.description, e.status, e.raw_data
        FROM prestacao_expenses e
        WHERE e.report_id IN ({placeholders})
    """, list(neon_reports.keys()))
else:
    cur.execute("SELECT 1 WHERE FALSE")

neon_expenses = {}
for row in cur.fetchall():
    eid = row["id"]
    rid = row["report_id"]
    if rid not in neon_reports: continue
    raw = row.get("raw_data")
    payment = ""
    if raw:
        try:
            d = json.loads(raw) if isinstance(raw, str) else raw
            payment = str(d.get("payment_type", "") or d.get("paymentMethod", "") or "")
        except: pass
    neon_expenses[eid] = {
        "report_id": rid,
        "report_name": neon_reports[rid]["name"],
        "report_status": neon_reports[rid]["status"],
        "user_name": neon_reports[rid]["user_name"],
        "user_cpf": neon_reports[rid]["cpf"],
        "value": nf(row["value"]),
        "date": str(row["date"] or "")[:10],
        "description": str(row["description"] or "").strip(),
        "payment": payment,
    }
conn.close()
print(f"  Neon expenses (APROVADO+ENVIADO, non-FATURA): {len(neon_expenses)}")

# ============================================================
# 3. Compare expense by expense
# ============================================================
print("Comparing...")

ref_eids = set(ref_expenses.keys())
neon_eids = set(neon_expenses.keys())

only_in_ref = ref_eids - neon_eids
only_in_neon = neon_eids - ref_eids
in_both = ref_eids & neon_eids

# Value mismatches for expenses in both
value_mismatches = []
for eid in in_both:
    rv = ref_expenses[eid]["value"]
    nv = neon_expenses[eid]["value"]
    if abs(rv - nv) > 0.01:
        value_mismatches.append(eid)

# Report-level comparison
ref_by_report = defaultdict(lambda: {"count": 0, "total": 0.0, "name": "", "user": "", "cpf": "", "status": ""})
for eid, e in ref_expenses.items():
    rid = e["report_id"]
    ref_by_report[rid]["count"] += 1
    ref_by_report[rid]["total"] += e["value"]
    ref_by_report[rid]["name"] = e["report_name"]
    ref_by_report[rid]["user"] = e["user_name"]
    ref_by_report[rid]["cpf"] = e["user_cpf"]
    ref_by_report[rid]["status"] = e["status"]

neon_by_report = defaultdict(lambda: {"count": 0, "total": 0.0, "name": "", "user": "", "cpf": "", "status": ""})
for eid, e in neon_expenses.items():
    rid = e["report_id"]
    neon_by_report[rid]["count"] += 1
    neon_by_report[rid]["total"] += e["value"]
    neon_by_report[rid]["name"] = e["report_name"]
    neon_by_report[rid]["user"] = e["user_name"]
    neon_by_report[rid]["cpf"] = e["user_cpf"]
    neon_by_report[rid]["status"] = e["report_status"]

reports_only_ref = set(ref_by_report.keys()) - set(neon_by_report.keys())
reports_only_neon = set(neon_by_report.keys()) - set(ref_by_report.keys())
reports_both = set(ref_by_report.keys()) & set(neon_by_report.keys())

reports_more_neon = []  # reports where Neon has more expenses
reports_more_ref = []   # reports where ref has more expenses
reports_diff_value = [] # reports with same count but different total

for rid in reports_both:
    rc = ref_by_report[rid]["count"]
    nc_ = neon_by_report[rid]["count"]
    rt = ref_by_report[rid]["total"]
    nt = neon_by_report[rid]["total"]
    if nc_ > rc:
        reports_more_neon.append(rid)
    elif rc > nc_:
        reports_more_ref.append(rid)
    elif abs(rt - nt) > 0.01:
        reports_diff_value.append(rid)

print(f"  Expenses only in ref: {len(only_in_ref)}")
print(f"  Expenses only in Neon: {len(only_in_neon)}")
print(f"  Expenses in both (same value): {len(in_both) - len(value_mismatches)}")
print(f"  Value mismatches: {len(value_mismatches)}")
print(f"  Reports only in ref: {len(reports_only_ref)}")
print(f"  Reports only in Neon: {len(reports_only_neon)}")
print(f"  Reports with more expenses in Neon: {len(reports_more_neon)}")
print(f"  Reports with more expenses in ref: {len(reports_more_ref)}")
print(f"  Reports with different values (same count): {len(reports_diff_value)}")

# ============================================================
# 4. Create Excel
# ============================================================
print("Creating Excel...")
wb = openpyxl.Workbook()

header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
money_fmt = '#,##0.00'

def write_sheet(ws, headers, rows, col_widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if isinstance(val, float):
                cell.number_format = money_fmt
    if col_widths:
        for c, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

# --- Tab 1: Summary ---
ws = wb.active
ws.title = "RESUMO"
summary_rows = [
    ["Categoria", "Quantidade", "Valor Total (R$)", "Descrição"],
    ["Despesas só no REF", len(only_in_ref), sum(ref_expenses[e]["value"] for e in only_in_ref), "No BASE PREST mas não no Neon"],
    ["Despesas só no NEON", len(only_in_neon), sum(neon_expenses[e]["value"] for e in only_in_neon), "No Neon mas não no BASE PREST"],
    ["Despesas em ambos (mesmo valor)", len(in_both) - len(value_mismatches), sum(ref_expenses[e]["value"] for e in in_both if e not in value_mismatches), "Coincidem perfeitamente"],
    ["Despesas com valor diferente", len(value_mismatches), "", "Mesmo ID mas valores diferentes"],
    ["", "", "", ""],
    ["Relatórios só no REF", len(reports_only_ref), sum(ref_by_report[r]["total"] for r in reports_only_ref), "Relatórios inteiros só no BASE PREST"],
    ["Relatórios só no NEON", len(reports_only_neon), sum(neon_by_report[r]["total"] for r in reports_only_neon), "Relatórios inteiros só no Neon"],
    ["Relatórios com mais despesas no NEON", len(reports_more_neon), "", "Mesmo relatório, Neon tem mais despesas"],
    ["Relatórios com mais despesas no REF", len(reports_more_ref), "", "Mesmo relatório, REF tem mais despesas"],
    ["Relatórios com valor diferente", len(reports_diff_value), "", "Mesmo nº despesas, valores diferentes"],
    ["", "", "", ""],
    ["TOTAL REF", len(ref_expenses), sum(e["value"] for e in ref_expenses.values()), ""],
    ["TOTAL NEON", len(neon_expenses), sum(e["value"] for e in neon_expenses.values()), ""],
    ["DIFERENÇA", "", sum(e["value"] for e in neon_expenses.values()) - sum(e["value"] for e in ref_expenses.values()), "Neon - Ref"],
]
for r, row_data in enumerate(summary_rows, 1):
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=val)
        if r == 1:
            cell.font = header_font_white
            cell.fill = header_fill
        cell.border = thin_border
        if isinstance(val, float):
            cell.number_format = money_fmt
ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 45

# --- Tab 2: Só no REF ---
ws = wb.create_sheet("SO NO REF")
headers = ["Expense ID", "Report ID", "Report Name", "User", "CPF", "Status", "Value", "Date", "Description", "Payment", "Cost Center"]
rows = []
for eid in sorted(only_in_ref):
    e = ref_expenses[eid]
    rows.append([eid, e["report_id"], e["report_name"], e["user_name"], e["user_cpf"],
                 e["status"], e["value"], e["date"], e["description"], e["payment"], e["cost_center"]])
rows.sort(key=lambda x: abs(x[6]), reverse=True)
write_sheet(ws, headers, rows, [14, 12, 35, 30, 14, 10, 12, 12, 40, 20, 25])

# --- Tab 3: Só no NEON ---
ws = wb.create_sheet("SO NO NEON")
headers = ["Expense ID", "Report ID", "Report Name", "User", "CPF", "Report Status", "Value", "Date", "Description", "Payment"]
rows = []
for eid in sorted(only_in_neon):
    e = neon_expenses[eid]
    rows.append([eid, e["report_id"], e["report_name"], e["user_name"], e["user_cpf"],
                 e["report_status"], e["value"], e["date"], e["description"], e["payment"]])
rows.sort(key=lambda x: abs(x[6]), reverse=True)
write_sheet(ws, headers, rows, [14, 12, 35, 30, 14, 12, 12, 12, 40, 20])

# --- Tab 4: Reports só no REF ---
ws = wb.create_sheet("REPORTS SO NO REF")
headers = ["Report ID", "Report Name", "User", "CPF", "Status", "N Expenses", "Total Value"]
rows = []
for rid in sorted(reports_only_ref):
    r = ref_by_report[rid]
    rows.append([rid, r["name"], r["user"], r["cpf"], r["status"], r["count"], r["total"]])
rows.sort(key=lambda x: abs(x[6]), reverse=True)
write_sheet(ws, headers, rows, [12, 40, 30, 14, 10, 12, 15])

# --- Tab 5: Reports só no NEON ---
ws = wb.create_sheet("REPORTS SO NO NEON")
headers = ["Report ID", "Report Name", "User", "CPF", "Status", "N Expenses", "Total Value"]
rows = []
for rid in sorted(reports_only_neon):
    r = neon_by_report[rid]
    rows.append([rid, r["name"], r["user"], r["cpf"], r["status"], r["count"], r["total"]])
rows.sort(key=lambda x: abs(x[6]), reverse=True)
write_sheet(ws, headers, rows, [12, 40, 30, 14, 10, 12, 15])

# --- Tab 6: Reports com mais despesas no NEON ---
ws = wb.create_sheet("MAIS DESPESAS NO NEON")
headers = ["Report ID", "Report Name", "User", "CPF", "Status", "Ref Count", "Neon Count", "Diff Count", "Ref Total", "Neon Total", "Diff Value"]
rows = []
for rid in reports_more_neon:
    r = ref_by_report[rid]
    n = neon_by_report[rid]
    rows.append([rid, n["name"], n["user"], n["cpf"], n["status"],
                 r["count"], n["count"], n["count"] - r["count"],
                 r["total"], n["total"], n["total"] - r["total"]])
rows.sort(key=lambda x: abs(x[10]), reverse=True)
write_sheet(ws, headers, rows, [12, 40, 30, 14, 10, 10, 10, 10, 14, 14, 14])

# --- Tab 7: Reports com mais despesas no REF ---
ws = wb.create_sheet("MAIS DESPESAS NO REF")
headers = ["Report ID", "Report Name", "User", "CPF", "Status", "Ref Count", "Neon Count", "Diff Count", "Ref Total", "Neon Total", "Diff Value"]
rows = []
for rid in reports_more_ref:
    r = ref_by_report[rid]
    n = neon_by_report[rid]
    rows.append([rid, r["name"], r["user"], r["cpf"], r["status"],
                 r["count"], n["count"], r["count"] - n["count"],
                 r["total"], n["total"], r["total"] - n["total"]])
rows.sort(key=lambda x: abs(x[10]), reverse=True)
write_sheet(ws, headers, rows, [12, 40, 30, 14, 10, 10, 10, 10, 14, 14, 14])

# --- Tab 8: Valor diferente (mesmo expense ID) ---
ws = wb.create_sheet("VALOR DIFERENTE")
headers = ["Expense ID", "Report ID", "Report Name", "User", "CPF", "Ref Value", "Neon Value", "Diff"]
rows = []
for eid in value_mismatches:
    r = ref_expenses[eid]
    n = neon_expenses[eid]
    rows.append([eid, r["report_id"], r["report_name"], r["user_name"], r["user_cpf"],
                 r["value"], n["value"], n["value"] - r["value"]])
rows.sort(key=lambda x: abs(x[7]), reverse=True)
write_sheet(ws, headers, rows, [14, 12, 35, 30, 14, 12, 12, 12])

# --- Tab 9: CPF-level summary ---
ws = wb.create_sheet("RESUMO POR CPF")
ref_by_cpf = defaultdict(float)
for e in ref_expenses.values():
    ref_by_cpf[e["user_cpf"]] += e["value"]
neon_by_cpf = defaultdict(float)
for e in neon_expenses.values():
    neon_by_cpf[e["user_cpf"]] += e["value"]

all_cpfs = set(ref_by_cpf.keys()) | set(neon_by_cpf.keys())
headers = ["CPF", "User", "Ref Total", "Neon Total", "Diff", "Status"]
rows = []
for cpf in all_cpfs:
    rv = ref_by_cpf.get(cpf, 0)
    nv = neon_by_cpf.get(cpf, 0)
    diff = nv - rv
    if abs(diff) <= 0.50: continue
    # Find user name
    user = ""
    for e in ref_expenses.values():
        if e["user_cpf"] == cpf: user = e["user_name"]; break
    if not user:
        for e in neon_expenses.values():
            if e["user_cpf"] == cpf: user = e["user_name"]; break
    status = "NEON > REF" if diff > 0 else "REF > NEON"
    rows.append([cpf, user, rv, nv, diff, status])
rows.sort(key=lambda x: abs(x[4]), reverse=True)
write_sheet(ws, headers, rows, [14, 35, 15, 15, 15, 12])

wb.save(OUTPUT)
print(f"\nSaved: {OUTPUT}")
print(f"Tabs: RESUMO, SO NO REF, SO NO NEON, REPORTS SO NO REF, REPORTS SO NO NEON, MAIS DESPESAS NO NEON, MAIS DESPESAS NO REF, VALOR DIFERENTE, RESUMO POR CPF")
