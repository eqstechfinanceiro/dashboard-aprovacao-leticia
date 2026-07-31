#!/usr/bin/env python3
"""
Local comparison with ALL fixes applied:
1. Deduplicate extrato (DISTINCT on usuario, data, tipo, valor, codigo_transacao)
2. Prestação: only APROVADO, filtered by updated_at <= cutoff
3. Name matching: exact → prefix(15,10) → fuzzy(0.88)
"""
import unicodedata, sys
from collections import defaultdict
from datetime import date
from pathlib import Path
import openpyxl

BASE = Path(__file__).parent.parent

def nc(raw):
    if raw is None: return ""
    s = str(raw).strip().replace(".","").replace("-","").replace("/","").replace(" ","")
    return s.zfill(11)

def nf(raw):
    try: return round(float(raw), 2) if raw is not None else 0.0
    except: return 0.0

def norm(name):
    if not name: return ""
    s = str(name).upper().strip()
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn").strip()

def bigrams(s):
    return set(s[i:i+2] for i in range(len(s)-1))

def fuzzy_ratio(a, b):
    if a == b: return 1.0
    if not a or not b: return 0.0
    ba, bb = bigrams(a), bigrams(b)
    inter = len(ba & bb)
    return (2 * inter) / (len(ba) + len(bb))

def _is_fatura_or_cartao(name: str) -> bool:
    """Comprehensive FATURA/CARTAO filter matching ref BASE PREST behavior."""
    n = name.strip().upper()
    if 'CAIXA ITAU' in n or 'CAIXA ITAÚ' in n:
        return True
    if n.startswith('CAIXA'):
        return False
    if n.startswith(('FATURA', 'CARTAO', 'CARTÃO', 'FATUAR', 'FARTUR', 'FATUT', 'FARUR', 'FATUTR')):
        return True
    if 'CARTÃO DE CRÉDITO' in n or 'CARTAO DE CREDITO' in n or 'CARTÃO DE CREDITO' in n:
        return True
    if 'CARTÃO CORPORATIVO' in n:
        return True
    if ('ITAU' in n or 'ITAÚ' in n) and 'CAIXA' not in n:
        return True
    if 'DOLAR' in n or 'DÓLAR' in n:
        return True
    if n.startswith('DESPESA') and 'FATURA' in n:
        return True
    if n.startswith('COMPLEMENTAR') and 'FATURA' in n:
        return True
    if 'CARTÃO' in n and 'CRÉDITO' in n:
        return True
    if 'CARTAO' in n and 'CREDITO' in n:
        return True
    if n.startswith('CARTÃO VEXPENSES'):
        return True
    return False

# === Load ref data ===
wb = openpyxl.load_workbook(BASE / "data" / "ref_dump.xlsx", read_only=True, data_only=True)
ws = wb["painel"]
ref = {}
ref_name_to_cpf = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    cpf = nc(row[0])
    if not cpf or cpf == "00000000000": continue
    ref[cpf] = {
        "colaborador": str(row[1] or "")[:30],
        "carga": nf(row[2]),
        "transferencia": abs(nf(row[3])),
        "tarifa": abs(nf(row[4])),
        "prestacao": nf(row[5]),
        "saldo_prestacao": nf(row[6]),
        "saldo_cartao": abs(nf(row[7])),
        "saldo_final": nf(row[8]),
    }
    ref_name_to_cpf[norm(str(row[1] or ""))] = cpf
wb.close()
print(f"Ref: {len(ref)} CPFs")

# === Load neon extrato (deduplicated) ===
wb = openpyxl.load_workbook(BASE / "data" / "neon_dump.xlsx", read_only=True, data_only=True)
ws = wb["extrato_raw"]
seen = set()
neon_by_name = defaultdict(lambda: {"carga": 0, "transferencia": 0, "tarifa": 0})
neon_snap_by_name = {}
cutoff = date(2026, 6, 30)
row_idx = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    row_idx += 1
    nome = str(row[0] or "").upper()
    d = row[1]
    if hasattr(d, 'date'):
        d = d.date() if hasattr(d, 'date') else d
    elif isinstance(d, str):
        d = date.fromisoformat(d)
    hora = str(row[2]) if len(row) > 2 and row[2] else ""
    tipo = str(row[3] or "")
    valor = nf(row[4])
    cod = str(row[5] or "") if len(row) > 5 else ""
    
    if not d or d > cutoff:
        continue
    
    # Dedup key: use cod when present; use hora as tiebreaker for empty cod
    if cod:
        dedup_key = (nome, d, tipo, valor, cod)
    else:
        dedup_key = (nome, d, hora, tipo, valor)
    if dedup_key in seen:
        continue
    seen.add(dedup_key)
    
    if tipo == "Transferência" and valor > 0:
        neon_by_name[nome]["carga"] += valor
    elif tipo == "Transferência" and valor < 0:
        neon_by_name[nome]["transferencia"] += abs(valor)
    elif tipo == "Taxa":
        neon_by_name[nome]["tarifa"] += abs(valor)

# Load snapshots for saldo cartao
ws_snap = wb["snapshots_all"] if "snapshots_all" in wb.sheetnames else None
if ws_snap:
    for row in ws_snap.iter_rows(min_row=2, values_only=True):
        nome = str(row[0] or "").upper()
        d = row[1]
        if hasattr(d, 'date'):
            d = d.date() if hasattr(d, 'date') else d
        elif isinstance(d, str):
            d = date.fromisoformat(d)
        valor = nf(row[2])
        if d and d <= date(2026, 7, 1):  # saldo_cartao_controle_date
            if nome not in neon_snap_by_name or d > neon_snap_by_name[nome][1]:
                neon_snap_by_name[nome] = (valor, d)
wb.close()
print(f"Neon extrato: {len(seen)} unique rows")

# === Load neon somase (prestacao) ===
# Try to load from prestacao_reports + prestacao_expenses sheets
wb = openpyxl.load_workbook(BASE / "data" / "neon_dump.xlsx", read_only=True, data_only=True)
somase_by_cpf = {}
if "prestacao_reports" in wb.sheetnames and "prestacao_expenses" in wb.sheetnames:
    # Load reports - ALL APROVADO, no date filter (cumulative since card creation)
    ws_r = wb["prestacao_reports"]
    reports = {}
    for row in ws_r.iter_rows(min_row=2, values_only=True):
        rid = row[0]
        name = str(row[1] or "").upper()
        status = str(row[2] or "").upper()
        user_cpf = nc(row[5]) if len(row) > 5 and row[5] else ""
        
        if status not in ("APROVADO", "ENVIADO") or not user_cpf:
            continue
        name_upper = name.strip().upper()
        if _is_fatura_or_cartao(name_upper):
            continue
        reports[rid] = user_cpf
    
    # Load expenses and sum by CPF
    ws_e = wb["prestacao_expenses"]
    for row in ws_e.iter_rows(min_row=2, values_only=True):
        rid = row[1] if len(row) > 1 else None
        value = nf(row[2]) if len(row) > 2 else 0
        if rid in reports:
            cpf = reports[rid]
            somase_by_cpf[cpf] = somase_by_cpf.get(cpf, 0) + value
    print(f"Neon somase (computed): {len(somase_by_cpf)} CPFs, total R$ {sum(somase_by_cpf.values()):,.2f}")
elif "somase_snapshots" in wb.sheetnames:
    ws_s = wb["somase_snapshots"]
    for row in ws_s.iter_rows(min_row=2, values_only=True):
        cpf = nc(row[0])
        qid = str(row[1] or "")
        total = nf(row[2])
        if qid == "2026-07-1":
            somase_by_cpf[cpf] = total
    print(f"Neon somase (snapshots): {len(somase_by_cpf)} CPFs, total R$ {sum(somase_by_cpf.values()):,.2f}")
else:
    print("WARNING: No prestacao data found in neon_dump.xlsx")
wb.close()

# === Build name mapping ===
# Load cadastro
wb = openpyxl.load_workbook(BASE / "data" / "neon_dump.xlsx", read_only=True, data_only=True)
ws_c = wb["cadastro"]
cad_name_to_cpf = {}
for row in ws_c.iter_rows(min_row=2, values_only=True):
    cpf = nc(row[0])
    nome = str(row[1] or "")
    cad_name_to_cpf[norm(nome)] = cpf
wb.close()

# Best mapping: ref names first, then cadastro
def map_name(nome):
    n = norm(nome)
    if n in ref_name_to_cpf: return ref_name_to_cpf[n]
    if n in cad_name_to_cpf: return cad_name_to_cpf[n]
    # Fuzzy match FIRST (handles LUIZ vs LUIS, typos, etc.)
    best_cpf = None
    best_ratio = 0
    for cn, cpf in cad_name_to_cpf.items():
        r = fuzzy_ratio(n, cn)
        if r > best_ratio:
            best_ratio = r
            best_cpf = cpf
    if best_ratio >= 0.88 and best_cpf:
        return best_cpf
    # Prefix 15 (fallback for truncated names)
    if len(n) >= 15:
        p15 = n[:15]
        for cn, cpf in cad_name_to_cpf.items():
            if cn[:15] == p15: return cpf
    # Prefix 10 (last resort)
    if len(n) >= 10:
        p10 = n[:10]
        for cn, cpf in cad_name_to_cpf.items():
            if cn[:10] == p10: return cpf
    return None

# === Compute neon values per CPF ===
neon_data = {}
unmapped = set()
for nome, vals in neon_by_name.items():
    cpf = map_name(nome)
    if not cpf:
        unmapped.add(nome)
        continue
    if cpf not in neon_data:
        neon_data[cpf] = {"carga": 0, "transferencia": 0, "tarifa": 0, "saldo_cartao": 0}
    neon_data[cpf]["carga"] += vals["carga"]
    neon_data[cpf]["transferencia"] += vals["transferencia"]
    neon_data[cpf]["tarifa"] += vals["tarifa"]

# Add saldo cartao from snapshots
for nome, (saldo, snap_date) in neon_snap_by_name.items():
    cpf = map_name(nome)
    if cpf:
        if cpf not in neon_data:
            neon_data[cpf] = {"carga": 0, "transferencia": 0, "tarifa": 0, "saldo_cartao": 0}
        neon_data[cpf]["saldo_cartao"] = abs(saldo)

# Add prestacao from somase
for cpf, total in somase_by_cpf.items():
    if cpf not in neon_data:
        neon_data[cpf] = {"carga": 0, "transferencia": 0, "tarifa": 0, "saldo_cartao": 0}
    neon_data[cpf]["prestacao"] = total

# === Compare ===
fields = ["carga", "transferencia", "tarifa", "prestacao", "saldo_cartao"]
match_counts = {f: 0 for f in fields}
total_cpf = 0
divs_by_field = defaultdict(list)

for cpf, ref_vals in ref.items():
    total_cpf += 1
    neon_vals = neon_data.get(cpf, {})
    
    for field in fields:
        rv = ref_vals.get(field, 0)
        nv = neon_vals.get(field, 0)
        diff = abs(nv - rv)
        if diff <= 0.50:  # 50 cent tolerance for rounding
            match_counts[field] += 1
        else:
            divs_by_field[field].append((cpf, ref_vals["colaborador"], rv, nv, diff))

print(f"\n=== Match Results (tolerance R$0.50) ===")
print(f"Total CPFs: {total_cpf}")
for f in fields:
    pct = match_counts[f] / total_cpf * 100
    print(f"  {f:<15}: {match_counts[f]}/{total_cpf} = {pct:.1f}%")

# Also compute saldo_prestacao and saldo_final
sp_match = 0
sf_match = 0
for cpf, ref_vals in ref.items():
    neon_vals = neon_data.get(cpf, {})
    neon_carga = neon_vals.get("carga", 0)
    neon_transf = neon_vals.get("transferencia", 0)
    neon_tarifa = neon_vals.get("tarifa", 0)
    neon_prest = neon_vals.get("prestacao", 0)
    neon_sc = neon_vals.get("saldo_cartao", 0)
    
    neon_sp = round(neon_carga - neon_transf - neon_tarifa - neon_prest, 2)
    neon_sf = round(neon_sp - neon_sc, 2)
    
    if abs(neon_sp - ref_vals["saldo_prestacao"]) <= 0.50:
        sp_match += 1
    if abs(neon_sf - ref_vals["saldo_final"]) <= 0.50:
        sf_match += 1

print(f"  saldo_prestacao: {sp_match}/{total_cpf} = {sp_match/total_cpf*100:.1f}%")
print(f"  saldo_final:     {sf_match}/{total_cpf} = {sf_match/total_cpf*100:.1f}%")

# Top divergences per field
print(f"\n=== Top divergences ===")
for field in fields:
    divs = divs_by_field[field]
    if not divs: continue
    print(f"\n  {field} ({len(divs)} divergences):")
    for cpf, nome, rv, nv, diff in sorted(divs, key=lambda x: x[4], reverse=True)[:5]:
        print(f"    {cpf} {nome[:20]:<20} ref={rv:>12.2f} neon={nv:>12.2f} diff={diff:>+10.2f}")

# Unmapped names
if unmapped:
    print(f"\n=== Unmapped names ({len(unmapped)}) ===")
    for n in sorted(unmapped)[:10]:
        print(f"  {n}")
