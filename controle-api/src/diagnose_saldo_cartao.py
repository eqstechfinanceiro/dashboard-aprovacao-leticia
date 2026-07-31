#!/usr/bin/env python3
"""Diagnose saldo_cartao: check which snapshot names fail to map to CPFs."""
import unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path
import openpyxl

BASE = Path(__file__).parent.parent

def nc(raw):
    if raw is None: return ""
    s = str(raw).strip().replace(".","").replace("-","").replace("/","").replace(" ","")
    return s.zfill(11)

def norm(name):
    if not name: return ""
    s = str(name).upper().strip()
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn").strip()

# Load ref saldo_cartao
wb = openpyxl.load_workbook(BASE / "data" / "ref_dump.xlsx", read_only=True, data_only=True)
ws = wb["saldo_cartao_ref"]
ref_sc = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    cpf = nc(row[0])
    val = row[1]
    if cpf and cpf != "00000000000":
        ref_sc[cpf] = val if val else 0
wb.close()
print(f"Ref saldo_cartao: {len(ref_sc)} CPFs")

# Load neon snapshots
wb = openpyxl.load_workbook(BASE / "data" / "neon_dump.xlsx", read_only=True, data_only=True)
ws = wb["snapshots_all"]
# Get latest snapshot per user up to July 1
neon_snap = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    nome = str(row[0] or "").upper()
    d = row[1]
    if hasattr(d, 'date'):
        d = d.date() if hasattr(d, 'date') else d
    elif isinstance(d, str):
        d = date.fromisoformat(d)
    valor = row[2]
    if d and d <= date(2026, 7, 1):
        if nome not in neon_snap or d > neon_snap[nome][1]:
            neon_snap[nome] = (valor, d)
wb.close()
print(f"Neon snapshots: {len(neon_snap)} users (up to July 1)")

# Load cadastro for name mapping
wb = openpyxl.load_workbook(BASE / "data" / "neon_dump.xlsx", read_only=True, data_only=True)
ws_c = wb["cadastro"]
cad_name_to_cpf = {}
for row in ws_c.iter_rows(min_row=2, values_only=True):
    cpf = nc(row[0])
    nome = str(row[1] or "")
    cad_name_to_cpf[norm(nome)] = cpf
wb.close()

# Also load ref names
wb = openpyxl.load_workbook(BASE / "data" / "ref_dump.xlsx", read_only=True, data_only=True)
ws_p = wb["painel"]
ref_name_to_cpf = {}
for row in ws_p.iter_rows(min_row=2, values_only=True):
    cpf = nc(row[0])
    nome = str(row[1] or "")
    ref_name_to_cpf[norm(nome)] = cpf
wb.close()

# Map snapshot names to CPFs
def map_name(nome):
    n = norm(nome)
    if n in ref_name_to_cpf: return ref_name_to_cpf[n]
    if n in cad_name_to_cpf: return cad_name_to_cpf[n]
    # Prefix 15
    if len(n) >= 15:
        p15 = n[:15]
        for cn, cpf in cad_name_to_cpf.items():
            if cn[:15] == p15: return cpf
    # Prefix 10
    if len(n) >= 10:
        p10 = n[:10]
        for cn, cpf in cad_name_to_cpf.items():
            if cn[:10] == p10: return cpf
    # Fuzzy
    best_cpf = None
    best_ratio = 0
    for cn, cpf in cad_name_to_cpf.items():
        ba = set(n[i:i+2] for i in range(len(n)-1))
        bb = set(cn[i:i+2] for i in range(len(cn)-1))
        if not ba or not bb: continue
        inter = len(ba & bb)
        r = (2 * inter) / (len(ba) + len(bb))
        if r > best_ratio:
            best_ratio = r
            best_cpf = cpf
    if best_ratio >= 0.88 and best_cpf:
        return best_cpf
    return None

mapped = 0
unmapped = []
neon_sc_by_cpf = {}
for nome, (saldo, snap_date) in neon_snap.items():
    cpf = map_name(nome)
    if cpf:
        mapped += 1
        neon_sc_by_cpf[cpf] = saldo
    else:
        unmapped.append((nome, saldo, snap_date))

print(f"Mapped: {mapped}/{len(neon_snap)}")
print(f"Unmapped: {len(unmapped)}")

# Check ref CPFs with no neon snapshot
missing_in_neon = 0
zero_when_ref_has_value = 0
for cpf, ref_val in ref_sc.items():
    if cpf not in neon_sc_by_cpf:
        missing_in_neon += 1
    elif abs(neon_sc_by_cpf[cpf]) < 0.01 and abs(ref_val) > 0.01:
        zero_when_ref_has_value += 1

print(f"\nRef CPFs missing in neon: {missing_in_neon}")
print(f"Neon has 0 but ref has value: {zero_when_ref_has_value}")

# Show unmapped names with their saldo values
if unmapped:
    print(f"\n=== Top unmapped snapshot names (by abs value) ===")
    for nome, saldo, d in sorted(unmapped, key=lambda x: abs(x[1] or 0), reverse=True)[:15]:
        print(f"  {d} {nome[:35]:<35} saldo={saldo:>10.2f}")

# Show top saldo_cartao divergences
divs = []
for cpf, ref_val in ref_sc.items():
    neon_val = neon_sc_by_cpf.get(cpf, 0)
    diff = abs(neon_val - ref_val)
    if diff > 0.50:
        divs.append((cpf, ref_val, neon_val, diff))

print(f"\n=== Saldo cartao divergences: {len(divs)} ===")
for cpf, rv, nv, diff in sorted(divs, key=lambda x: x[3], reverse=True)[:10]:
    print(f"  {cpf} ref={rv:>10.2f} neon={nv:>10.2f} diff={diff:>+10.2f}")
