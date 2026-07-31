#!/usr/bin/env python3
"""Check tarifa divergences: compare specific user's tarifa transactions in Neon vs ref."""
from datetime import date
from collections import defaultdict
from pathlib import Path
import openpyxl

BASE = Path(__file__).parent.parent

def nf(raw):
    try: return float(raw) if raw is not None else 0.0
    except: return 0.0

# Load ref painel
wb = openpyxl.load_workbook(BASE / "data" / "ref_dump.xlsx", read_only=True, data_only=True)
ws = wb["painel"]
ref_tarifa = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    cpf = str(row[0] or "").strip().replace(".","").replace("-","").replace(" ","").zfill(11)
    nome = str(row[1] or "")
    tarifa = abs(nf(row[4]))
    if tarifa > 0:
        ref_tarifa[cpf] = (nome, tarifa)
wb.close()

# Load neon extrato raw (deduplicated)
wb = openpyxl.load_workbook(BASE / "data" / "neon_dump.xlsx", read_only=True, data_only=True)
ws = wb["extrato_raw"]
neon_tarifa_by_user = defaultdict(list)
seen = set()
cutoff = date(2026, 6, 30)

for row in ws.iter_rows(min_row=2, values_only=True):
    nome = str(row[0] or "").upper()
    d = row[1]
    if hasattr(d, 'date'):
        d = d.date() if hasattr(d, 'date') else d
    elif isinstance(d, str):
        d = date.fromisoformat(d)
    tipo = str(row[2] or "")
    valor = abs(nf(row[3]))
    cod = str(row[4] or "") if len(row) > 4 else ""
    
    if not d or d > cutoff:
        continue
    
    dedup_key = (nome, d, tipo, valor, cod)
    if dedup_key in seen:
        continue
    seen.add(dedup_key)
    
    if tipo == "Taxa":
        neon_tarifa_by_user[nome].append((d, valor))

wb.close()

# Check top tarifa divergences
print("=== Top tarifa divergences (ref > neon) ===")
divs = []
for cpf, (nome, ref_val) in ref_tarifa.items():
    # Find this user in neon by name
    neon_total = 0
    for user, txns in neon_tarifa_by_user.items():
        if nome.upper()[:15] in user or user[:15] in nome.upper():
            neon_total = sum(v for _, v in txns)
            break
    if abs(neon_total - ref_val) > 0.50:
        divs.append((cpf, nome[:25], ref_val, neon_total, abs(neon_total - ref_val)))

for cpf, nome, rv, nv, diff in sorted(divs, key=lambda x: x[4], reverse=True)[:10]:
    print(f"  {cpf} {nome:<25} ref={rv:>10.2f} neon={nv:>10.2f} diff={diff:>+10.2f}")

# Check total tarifa
total_neon = sum(sum(v for _, v in txns) for txns in neon_tarifa_by_user.values())
total_ref = sum(v for _, v in ref_tarifa.values())
print(f"\nTotal tarifa - Ref: R$ {total_ref:,.2f} | Neon: R$ {total_neon:,.2f} | Gap: R$ {total_ref - total_neon:,.2f}")

# Check if there are tarifa transactions after cutoff in Neon
after_cutoff = 0
after_cutoff_total = 0
for user, txns in neon_tarifa_by_user.items():
    for d, v in txns:
        if d > cutoff:
            after_cutoff += 1
            after_cutoff_total += v
print(f"Tarifa transactions AFTER June 30: {after_cutoff}, total R$ {after_cutoff_total:,.2f}")

# Check a specific user: RAFAEL AMORIM VELLO
print("\n=== RAFAEL AMORIM VELLO tarifa transactions ===")
for user, txns in neon_tarifa_by_user.items():
    if "RAFAEL AMORIM" in user:
        print(f"  User: {user}")
        for d, v in sorted(txns):
            print(f"    {d} R$ {v:.2f}")
        print(f"  Total: R$ {sum(v for _, v in txns):.2f}")
        break
