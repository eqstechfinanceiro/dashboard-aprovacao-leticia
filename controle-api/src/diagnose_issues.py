#!/usr/bin/env python3
"""
diagnose_issues.py
------------------
Diagnoses the 3 key issues:
1. Name→CPF matching (extrato names are truncated ~20 chars)
2. Prestação: our somase R$ 12.1M vs ref R$ 6.8M — why?
3. Tarifa: 41 small divergences
"""
import os
import sys
import unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl

BASE = Path(__file__).parent.parent
NEON_DUMP = BASE / "data" / "neon_dump.xlsx"
REF_DUMP = BASE / "data" / "ref_dump.xlsx"


def nc(raw):
    if raw is None:
        return ""
    s = str(raw).strip().replace(".", "").replace("-", "").replace("/", "").replace(" ", "")
    return s.zfill(11)

def nf(raw):
    try:
        return round(float(raw), 2) if raw is not None else 0.0
    except (ValueError, TypeError):
        return 0.0

def normalize_name(name):
    """Remove accents, uppercase, strip."""
    if not name:
        return ""
    s = str(name).upper().strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.strip()


def load_neon_dump():
    wb = openpyxl.load_workbook(NEON_DUMP, read_only=True, data_only=True)

    # Cadastro
    ws = wb["cadastro"]
    cadastro = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cpf = nc(row[0])
        if cpf:
            cadastro[cpf] = str(row[1] or "")

    # Extrato acumulado (filtered to June 30)
    ws = wb["extrato_acumulado"]
    extrato = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = str(row[0] or "").upper()
        extrato[nome] = {
            "carga": nf(row[1]),
            "transferencia": nf(row[2]),
            "tarifa": nf(row[3]),
        }

    # Extrato raw (all dates)
    ws = wb["extrato_raw"]
    extrato_raw = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        extrato_raw.append({
            "usuario": str(row[0] or "").upper(),
            "data": row[1],
            "tipo": str(row[2] or ""),
            "valor": nf(row[3]),
        })

    # Snapshots
    ws = wb["snapshots_all"]
    snapshots = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = str(row[0] or "").upper()
        d = row[1]
        if hasattr(d, "strftime"):
            d = d.date() if hasattr(d, "date") else d
        elif isinstance(d, str):
            d = date.fromisoformat(d)
        snapshots[nome].append((d, nf(row[2])))
    for nome in snapshots:
        snapshots[nome].sort(key=lambda x: x[0])

    # Somase
    ws = wb["somase_snapshots"]
    somase = defaultdict(dict)
    for row in ws.iter_rows(min_row=2, values_only=True):
        cpf = nc(row[0])
        qid = str(row[1])
        somase[qid][cpf] = nf(row[2])

    wb.close()
    return {
        "cadastro": cadastro,
        "extrato": extrato,
        "extrato_raw": extrato_raw,
        "snapshots": snapshots,
        "somase": somase,
    }


def load_ref_dump():
    wb = openpyxl.load_workbook(REF_DUMP, read_only=True, data_only=True)

    # PAINEL
    ws = wb["painel"]
    painel = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cpf = nc(row[0])
        if cpf:
            painel[cpf] = {
                "colaborador": str(row[1] or ""),
                "carga": nf(row[2]),
                "transferencia": abs(nf(row[3])),
                "tarifa": abs(nf(row[4])),
                "prestacao": nf(row[5]),
                "saldo_prestacao": nf(row[6]),
                "saldo_cartao": abs(nf(row[7])),
                "saldo_final": nf(row[8]),
            }

    # EXTRATO ref — name→CPF mapping
    ws = wb["extrato_ref"]
    name_to_cpf = {}
    extrato_by_cpf = defaultdict(lambda: {"carga": 0.0, "transferencia": 0.0, "tarifa": 0.0})
    for row in ws.iter_rows(min_row=2, values_only=True):
        cpf = nc(row[0])
        nome = str(row[1] or "").upper()
        tipo = str(row[2] or "")
        valor = nf(row[3])
        if cpf and cpf != "00000000000":
            name_to_cpf[nome] = cpf
            if tipo == "CARGA" and valor > 0:
                extrato_by_cpf[cpf]["carga"] += valor
            elif tipo == "TRANSFERÊNCIA":
                extrato_by_cpf[cpf]["transferencia"] += abs(valor)
            elif tipo == "TARIFA":
                extrato_by_cpf[cpf]["tarifa"] += abs(valor)

    # BASE PREST ref
    ws = wb["base_prest_ref"]
    prest_by_cpf = defaultdict(float)
    prest_count_by_cpf = defaultdict(int)
    for row in ws.iter_rows(min_row=2, values_only=True):
        cpf = nc(row[0])
        valor = nf(row[1])
        if cpf:
            prest_by_cpf[cpf] += valor
            prest_count_by_cpf[cpf] += 1

    wb.close()
    return {
        "painel": painel,
        "name_to_cpf": name_to_cpf,
        "extrato_by_cpf": dict(extrato_by_cpf),
        "prest_by_cpf": dict(prest_by_cpf),
        "prest_count_by_cpf": dict(prest_count_by_cpf),
    }


def main():
    print("Loading dumps...")
    neon = load_neon_dump()
    ref = load_ref_dump()

    cadastro = neon["cadastro"]
    ref_name_to_cpf = ref["name_to_cpf"]

    # ============================================================
    # ISSUE 1: Name→CPF matching
    # ============================================================
    print("\n" + "=" * 80)
    print("  ISSUE 1: Name→CPF matching")
    print("=" * 80)

    # Build mapping from cadastro: normalized_name(20) → cpf
    cad_name_to_cpf = {}
    cad_name_norm_to_cpf = {}
    for cpf, nome in cadastro.items():
        nome_up = nome.upper().strip()
        nome_norm = normalize_name(nome)
        cad_name_to_cpf[nome_up] = cpf
        cad_name_norm_to_cpf[nome_norm] = cpf
        # Also try truncated to 20 chars
        cad_name_norm_to_cpf[nome_norm[:20]] = cpf

    # Check how many extrato names match
    extrato_names = set(neon["extrato"].keys())
    matched_exact = 0
    matched_norm = 0
    matched_trunc = 0
    unmatched = []

    for nome in extrato_names:
        if nome in cad_name_to_cpf:
            matched_exact += 1
        elif normalize_name(nome) in cad_name_norm_to_cpf:
            matched_norm += 1
        elif normalize_name(nome)[:20] in cad_name_norm_to_cpf:
            matched_trunc += 1
        else:
            # Try ref mapping
            if nome in ref_name_to_cpf:
                matched_trunc += 1  # ref has the mapping
            else:
                unmatched.append(nome)

    print(f"  Extrato names: {len(extrato_names)}")
    print(f"  Matched exact (uppercase): {matched_exact}")
    print(f"  Matched normalized: {matched_norm}")
    print(f"  Matched truncated/ref: {matched_trunc}")
    print(f"  Unmatched: {len(unmatched)}")
    if unmatched:
        print(f"\n  Unmatched names (first 20):")
        for n in sorted(unmatched)[:20]:
            # Find closest cadastro name
            n_norm = normalize_name(n)
            best = None
            best_score = 0
            for cn in cadastro.values():
                cn_norm = normalize_name(cn)
                # Simple prefix match score
                if cn_norm[:15] == n_norm[:15]:
                    best = cn
                    best_score = 100
                    break
                if cn_norm[:10] == n_norm[:10]:
                    best = cn
                    best_score = 80
            print(f"    '{n}' → closest: '{best}' (score={best_score})")

    # Build the BEST mapping: use ref mapping first, then cadastro
    best_mapping = {}
    for nome in extrato_names:
        if nome in ref_name_to_cpf:
            best_mapping[nome] = ref_name_to_cpf[nome]
        elif nome in cad_name_to_cpf:
            best_mapping[nome] = cad_name_to_cpf[nome]
        elif normalize_name(nome) in cad_name_norm_to_cpf:
            best_mapping[nome] = cad_name_norm_to_cpf[normalize_name(nome)]
        elif normalize_name(nome)[:20] in cad_name_norm_to_cpf:
            best_mapping[nome] = cad_name_norm_to_cpf[normalize_name(nome)[:20]]
        else:
            # Try prefix matching (first 15 chars normalized)
            n_norm = normalize_name(nome)[:15]
            found = False
            for cn, cpf in cadastro.items():
                if normalize_name(cn)[:15] == n_norm:
                    best_mapping[nome] = cpf
                    found = True
                    break
            if not found:
                # Try first 10 chars
                n_norm10 = normalize_name(nome)[:10]
                for cn, cpf in cadastro.items():
                    if normalize_name(cn)[:10] == n_norm10:
                        best_mapping[nome] = cpf
                        found = True
                        break
            if not found:
                best_mapping[nome] = None

    mapped = sum(1 for v in best_mapping.values() if v is not None)
    print(f"\n  Best mapping: {mapped}/{len(extrato_names)} names mapped to CPF")

    # ============================================================
    # ISSUE 2: Prestação divergences
    # ============================================================
    print("\n" + "=" * 80)
    print("  ISSUE 2: Prestação — our somase vs ref BASE PREST")
    print("=" * 80)

    somase_q = neon["somase"].get("2026-07-1", {})
    prest_ref = ref["prest_by_cpf"]

    # Compare per CPF
    all_cpfs = set(somase_q.keys()) | set(prest_ref.keys())
    matched_prest = 0
    divergent_prest = []
    only_somase = []
    only_ref = []

    for cpf in sorted(all_cpfs):
        v_som = somase_q.get(cpf, 0.0)
        v_ref = prest_ref.get(cpf, 0.0)
        if abs(v_som - v_ref) <= 0.05:
            matched_prest += 1
        else:
            divergent_prest.append((cpf, v_som, v_ref, v_som - v_ref))
            if v_ref == 0 and v_som > 0:
                only_somase.append((cpf, v_som))
            elif v_som == 0 and v_ref > 0:
                only_ref.append((cpf, v_ref))

    print(f"  CPFs in somase: {len(somase_q)}")
    print(f"  CPFs in ref BASE PREST: {len(prest_ref)}")
    print(f"  Matched (tol 0.05): {matched_prest}")
    print(f"  Divergent: {len(divergent_prest)}")
    print(f"    Only in somase (ref=0): {len(only_somase)}")
    print(f"    Only in ref (somase=0): {len(only_ref)}")
    print(f"    Both have values but differ: {len(divergent_prest) - len(only_somase) - len(only_ref)}")

    total_somase = sum(v for _, v in only_somase)
    total_ref_only = sum(v for _, v in only_ref)
    print(f"  Total only in somase: R$ {total_somase:,.2f}")
    print(f"  Total only in ref: R$ {total_ref_only:,.2f}")

    if divergent_prest:
        print(f"\n  Top 15 divergences:")
        for cpf, vs, vr, d in sorted(divergent_prest, key=lambda x: abs(x[3]), reverse=True)[:15]:
            nome = cadastro.get(cpf, "?")[:25]
            print(f"    {cpf} {nome:<25} somase={vs:>12.2f}  ref={vr:>12.2f}  diff={d:>+12.2f}")

    # ============================================================
    # ISSUE 3: Tarifa divergences
    # ============================================================
    print("\n" + "=" * 80)
    print("  ISSUE 3: Tarifa — 41 small divergences")
    print("=" * 80)

    # Compare ref extrato tarifa vs neon extrato tarifa (by name)
    ref_tarifa_by_name = defaultdict(float)
    for row in ref["extrato_by_cpf"].values():
        pass  # Already by CPF

    # Check tarifa in ref extrato vs neon extrato for matched users
    tarifa_divs = []
    for nome, cpf in best_mapping.items():
        if cpf is None or cpf not in ref["extrato_by_cpf"]:
            continue
        t_neon = neon["extrato"][nome]["tarifa"]
        t_ref = ref["extrato_by_cpf"][cpf]["tarifa"]
        if abs(t_neon - t_ref) > 0.05:
            tarifa_divs.append((cpf, nome[:20], t_ref, t_neon, t_neon - t_ref))

    print(f"  Tarifa divergences (matched users): {len(tarifa_divs)}")
    if tarifa_divs:
        print(f"  All divergences:")
        for cpf, nome, tr, tn, d in sorted(tarifa_divs, key=lambda x: abs(x[4]), reverse=True):
            print(f"    {cpf} {nome:<20} ref={tr:>8.2f}  neon={tn:>8.2f}  diff={d:>+8.2f}")

    # Check if the extra tarifa is from July 1 transactions
    print(f"\n  Checking extrato_raw for July tarifa transactions:")
    july_tarifa = [e for e in neon["extrato_raw"] if e["tipo"] == "TARIFA" and hasattr(e["data"], 'month') and e["data"].month == 7 and e["data"].year == 2026]
    print(f"  July tarifa transactions: {len(july_tarifa)}")
    if july_tarifa:
        total_july_tarifa = sum(abs(e["valor"]) for e in july_tarifa)
        print(f"  Total July tarifa: R$ {total_july_tarifa:,.2f}")
        print(f"  Sample:")
        for e in july_tarifa[:5]:
            print(f"    {e['data']} {e['usuario'][:20]} {e['valor']:>8.2f}")

    # ============================================================
    # SUMMARY: What would match % be with fixes?
    # ============================================================
    print("\n" + "=" * 80)
    print("  PROJECTED MATCH WITH FIXES")
    print("=" * 80)

    # Recalculate using best mapping + ref prestação
    painel = ref["painel"]
    saldo_controle_date = date(2026, 7, 1)

    # Build extrato by CPF using best mapping
    extrato_by_cpf = {}
    for nome, ext in neon["extrato"].items():
        cpf = best_mapping.get(nome)
        if cpf:
            if cpf not in extrato_by_cpf:
                extrato_by_cpf[cpf] = {"carga": 0.0, "transferencia": 0.0, "tarifa": 0.0}
            extrato_by_cpf[cpf]["carga"] += ext["carga"]
            extrato_by_cpf[cpf]["transferencia"] += ext["transferencia"]
            extrato_by_cpf[cpf]["tarifa"] += ext["tarifa"]

    # Also get snapshots by CPF
    snap_by_cpf = {}
    for nome, snaps in neon["snapshots"].items():
        cpf = best_mapping.get(nome)
        if cpf:
            # Get latest snapshot <= July 1
            best_val = 0.0
            for d, v in snaps:
                if d <= saldo_controle_date:
                    best_val = v
                else:
                    break
            snap_by_cpf[cpf] = best_val

    # Calculate with: neon extrato (mapped) + ref prestação + neon snapshots
    campos = ["carga", "transferencia", "tarifa", "prestacao", "saldo_prestacao", "saldo_cartao", "saldo_final"]
    divs_count = {c: 0 for c in campos}
    total_cmp = 0

    for cpf in set(painel.keys()) & set(cadastro.keys()):
        total_cmp += 1
        r = painel[cpf]

        ext = extrato_by_cpf.get(cpf, {"carga": 0.0, "transferencia": 0.0, "tarifa": 0.0})
        prest = prest_ref.get(cpf, 0.0)  # Use ref prestação
        sc = snap_by_cpf.get(cpf, 0.0)

        sp = round(ext["carga"] - ext["transferencia"] - ext["tarifa"] - prest, 2)
        sf = round(sp - sc, 2)

        vals = {
            "carga": ext["carga"],
            "transferencia": ext["transferencia"],
            "tarifa": ext["tarifa"],
            "prestacao": prest,
            "saldo_prestacao": sp,
            "saldo_cartao": sc,
            "saldo_final": sf,
        }

        for campo in campos:
            if abs(vals[campo] - r[campo]) > 0.05:
                divs_count[campo] += 1

    print(f"  Using: NEON extrato (mapped) + REF prestação + NEON snapshots")
    print(f"  Compared: {total_cmp}")
    for campo in campos:
        match = total_cmp - divs_count[campo]
        pct = match / total_cmp * 100 if total_cmp else 0
        print(f"    {campo:<22}: {match:>4}/{total_cmp}  ({pct:5.1f}%)  diverg: {divs_count[campo]}")

    # Now with NEON prestação (our somase)
    print(f"\n  Using: NEON extrato (mapped) + NEON prestação (somase) + NEON snapshots")
    divs_count2 = {c: 0 for c in campos}
    for cpf in set(painel.keys()) & set(cadastro.keys()):
        r = painel[cpf]
        ext = extrato_by_cpf.get(cpf, {"carga": 0.0, "transferencia": 0.0, "tarifa": 0.0})
        prest = somase_q.get(cpf, 0.0)  # Use our somase
        sc = snap_by_cpf.get(cpf, 0.0)

        sp = round(ext["carga"] - ext["transferencia"] - ext["tarifa"] - prest, 2)
        sf = round(sp - sc, 2)

        vals = {
            "carga": ext["carga"],
            "transferencia": ext["transferencia"],
            "tarifa": ext["tarifa"],
            "prestacao": prest,
            "saldo_prestacao": sp,
            "saldo_cartao": sc,
            "saldo_final": sf,
        }

        for campo in campos:
            if abs(vals[campo] - r[campo]) > 0.05:
                divs_count2[campo] += 1

    for campo in campos:
        match = total_cmp - divs_count2[campo]
        pct = match / total_cmp * 100 if total_cmp else 0
        print(f"    {campo:<22}: {match:>4}/{total_cmp}  ({pct:5.1f}%)  diverg: {divs_count2[campo]}")

    print("=" * 80)


if __name__ == "__main__":
    main()
