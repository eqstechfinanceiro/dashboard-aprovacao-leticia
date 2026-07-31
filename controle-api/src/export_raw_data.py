#!/usr/bin/env python3
"""
export_raw_data.py
------------------
Exporta dados brutos (carga, transferencia, tarifa, prestacao, saldo cartao)
em arquivos Excel separados para comparacao manual.

Para cada quinzena, gera:
  1. extrato_acumulado.xlsx  — carga, transferencia, tarifa por usuario (ate cutoff)
  2. saldo_cartao_controle.xlsx — snapshot ate dia 1 do mes atual
  3. saldo_cartao_carga.xlsx — snapshot ate data de fechamento (11 ou 25)
  4. prestacao_contas.xlsx — somase por CPF (se disponivel)
  5. resumo_completo.xlsx — todos os dados calculados em uma sheet
"""
import argparse
import logging
import os
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

BASE = Path(__file__).parent.parent
load_dotenv(BASE / ".env")

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
logger = logging.getLogger(__name__)

NEON_URL = os.getenv("NEON_DATABASE_URL")


def _r2(v: float) -> float:
    return float(Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def get_periodo(ano: int, mes: int, quinzena: int):
    mes_ant = mes - 1 if mes > 1 else 12
    ano_ant = ano if mes > 1 else ano - 1
    if quinzena == 1:
        inicio = f"{ano_ant}-{mes_ant:02d}-26"
        fim = f"{ano}-{mes:02d}-10"
        fechamento = f"{ano}-{mes:02d}-11"
    else:
        inicio = f"{ano}-{mes:02d}-11"
        fim = f"{ano}-{mes:02d}-25"
        fechamento = f"{ano}-{mes:02d}-25"
    cutoff_fin = f"{ano_ant}-{mes_ant:02d}-30"
    saldo_controle = f"{ano}-{mes:02d}-01"
    return inicio, fim, fechamento, cutoff_fin, saldo_controle


def buscar_extrato_acumulado(conn, cutoff: str):
    cur = conn.cursor()
    cur.execute("""
        SELECT
            UPPER(usuario) AS u,
            COALESCE(SUM(valor) FILTER(WHERE tipo = 'Transferência' AND valor > 0), 0) AS carga,
            COALESCE(ABS(SUM(valor) FILTER(WHERE tipo = 'Transferência' AND valor < 0)), 0) AS transferencia,
            COALESCE(ABS(SUM(valor) FILTER(WHERE tipo = 'Taxa')), 0) AS tarifa
        FROM extrato_movimentacao
        WHERE data <= %s
          AND is_snapshot = FALSE
        GROUP BY UPPER(usuario)
        ORDER BY UPPER(usuario)
    """, (cutoff,))
    return cur.fetchall()


def buscar_saldo_cartao(conn, data_limite: str, quinzena: int):
    op = "<=" if quinzena == 1 else "<="
    cur = conn.cursor()
    cur.execute(f"""
        SELECT UPPER(m.usuario), m.valor, m.data
        FROM extrato_movimentacao m
        WHERE m.is_snapshot = TRUE
          AND m.valor IS NOT NULL
          AND m.data = (
              SELECT MAX(m2.data)
              FROM extrato_movimentacao m2
              WHERE UPPER(m2.usuario) = UPPER(m.usuario)
                AND m2.is_snapshot = TRUE
                AND m2.valor IS NOT NULL
                AND m2.data {op} %s
          )
        ORDER BY UPPER(m.usuario)
    """, (data_limite,))
    return cur.fetchall()


def buscar_somase(conn, quinzena_id: str):
    cur = conn.cursor()
    cur.execute("SELECT user_cpf, total FROM somase_snapshots WHERE quinzena = %s ORDER BY user_cpf", (quinzena_id,))
    return cur.fetchall()


def buscar_cadastro(conn):
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT cpf, colaborador, situacao, status_cartao, regional,
               centro_custo, gestor, diretor
        FROM quinzena_cadastro
        ORDER BY colaborador ASC NULLS LAST
    """)
    return cur.fetchall()


def salvar_excel(rows, headers, output_path: Path, sheet_name: str = "Dados"):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("openpyxl nao instalado")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="1F4E79")

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(len(h) + 4, 14)

    wb.save(output_path)
    logger.info(f"Salvo: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Exporta dados brutos para Excel separados")
    parser.add_argument("--ano", type=int, required=True)
    parser.add_argument("--mes", type=int, required=True)
    parser.add_argument("--quinzena", type=int, choices=[1, 2], required=True)
    args = parser.parse_args()

    inicio, fim, fechamento, cutoff_fin, saldo_controle_date = get_periodo(args.ano, args.mes, args.quinzena)
    quinzena_id = f"{args.ano}-{args.mes:02d}-{args.quinzena}"

    print(f"\n{'='*65}")
    print(f"  {args.quinzena}ª QZ {args.mes:02d}/{args.ano}")
    print(f"  Cutoff financeiro: {cutoff_fin}")
    print(f"  Saldo cartão controle: {saldo_controle_date}")
    print(f"  Saldo cartão carga: {fechamento}")
    print(f"{'='*65}\n")

    if not NEON_URL:
        raise RuntimeError("NEON_DATABASE_URL nao configurada")

    conn = psycopg2.connect(NEON_URL)
    out_dir = BASE / "data" / f"raw_qz{args.quinzena}_{args.mes:02d}_{args.ano}"
    out_dir.mkdir(parents=True, exist_ok=True)

    # 1. Extrato acumulado (carga, transferencia, tarifa)
    logger.info("Buscando extrato acumulado...")
    extrato = buscar_extrato_acumulado(conn, cutoff_fin)
    salvar_excel(
        extrato,
        ["USUARIO", "CARGA", "TRANSFERENCIA", "TARIFA"],
        out_dir / "extrato_acumulado.xlsx",
        "Extrato Acumulado"
    )

    # 2. Saldo cartão controle
    logger.info("Buscando saldo cartão controle...")
    sc_controle = buscar_saldo_cartao(conn, saldo_controle_date, 1)
    salvar_excel(
        sc_controle,
        ["USUARIO", "SALDO_CARTAO", "DATA_SNAPSHOT"],
        out_dir / "saldo_cartao_controle.xlsx",
        "Saldo Cartao Controle"
    )

    # 3. Saldo cartão carga
    logger.info("Buscando saldo cartão carga...")
    sc_carga = buscar_saldo_cartao(conn, fechamento, args.quinzena)
    salvar_excel(
        sc_carga,
        ["USUARIO", "SALDO_CARTAO", "DATA_SNAPSHOT"],
        out_dir / "saldo_cartao_carga.xlsx",
        "Saldo Cartao Carga"
    )

    # 4. Prestação de contas (somase)
    logger.info("Buscando prestação de contas (somase)...")
    somase = buscar_somase(conn, quinzena_id)
    salvar_excel(
        somase,
        ["CPF", "TOTAL_PRESTACAO"],
        out_dir / "prestacao_contas.xlsx",
        "Prestacao Contas"
    )

    # 5. Cadastro completo
    logger.info("Buscando cadastro...")
    cadastro = buscar_cadastro(conn)
    cadastro_rows = [
        (r["cpf"], r["colaborador"], r["situacao"], r["status_cartao"],
        r["regional"], r["centro_custo"], r["gestor"], r["diretor"])
        for r in cadastro
    ]
    salvar_excel(
        cadastro_rows,
        ["CPF", "COLABORADOR", "SITUACAO", "STATUS_CARTAO", "REGIONAL", "CENTRO_CUSTO", "GESTOR", "DIRETOR"],
        out_dir / "cadastro.xlsx",
        "Cadastro"
    )

    conn.close()
    print(f"\n✅ Arquivos salvos em: {out_dir}")


if __name__ == "__main__":
    main()
