#!/usr/bin/env python3
"""
generate_sheets.py
------------------
Generates CONTROLE (PAINEL) and CARGA Excel sheets from API/Neon data.
For 2QZ July 2026 (closing date: July 25, 2026).

Formulas validated per quinzena-complete/route.ts:
  carga         = SUM(Transferência > 0) up to 2026-06-30
  transferencia = ABS(SUM(Transferência < 0)) up to 2026-06-30
  tarifa        = ABS(SUM(Taxa)) up to 2026-06-30
  prestacao     = SUM(expenses from Aprovado+Enviado, excluding FATURA/CARTAO) — cumulative
  saldo_prestacao = carga - transferencia - tarifa - prestacao
  saldo_cartao_controle = last snapshot up to 2026-07-01 + post-snapshot txns
  saldo_cartao_carga    = last snapshot up to 2026-07-25 + post-snapshot txns
  saldo_final  = saldo_prestacao - saldo_cartao_controle
  saldo_reembolsar = max(-saldo_final, 0)
  saldo_final_carga = max(saldo_final, 0)
  carga_parcial = col_qz_efetivo - saldo_final_carga - saldo_cartao_carga - adiantamento
  reembolso = 0 (2ª QZ)
  carga_final = max(0, carga_parcial) + reembolso
"""
import os
import sys
import re
import unicodedata
from pathlib import Path
from datetime import date
from collections import defaultdict

import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

BASE = Path(__file__).parent.parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")

# === Quinzena config ===
YEAR = 2026
MONTH = 7
QUINZENA = 2
FINANCIAL_CUTOFF = "2026-06-30"
SALDO_CARTAO_CONTROLE_DATE = "2026-07-01"
SALDO_CARTAO_CARGA_DATE = "2026-07-25"

OUT_DIR = BASE.parent / "data"

# === Helpers ===
def r2(v):
    return round(v * 100) / 100

def norm(name):
    if not name:
        return ""
    s = str(name).upper().strip()
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn").strip()

def nc(raw):
    if raw is None:
        return ""
    s = str(raw).strip().replace(".", "").replace("-", "").replace("/", "").replace(" ", "")
    return s.zfill(11)

def bigrams(s):
    if len(s) < 2:
        return set()
    return set(s[i:i+2] for i in range(len(s) - 1))

def fuzzy_ratio(a, b):
    if a == b:
        return 1.0
    if not a or not b:
        return 0.0
    ba, bb = bigrams(a), bigrams(b)
    if not ba or not bb:
        return 0.0
    return (2 * len(ba & bb)) / (len(ba) + len(bb))

def is_card_report(name):
    n = name.upper().strip()
    if re.match(r'^(FATU|FARUR|CART)', n):
        return True
    if re.search(r'(FATURA|FATUAR|FATUT|FARUR)', n):
        return True
    return False

# === Main ===
def main():
    conn = psycopg2.connect(NEON_URL, connect_timeout=10)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    print("=" * 70)
    print(f"  GENERATE SHEETS — {YEAR}-{MONTH:02d}-QZ{QUINZENA}")
    print(f"  Financial cutoff: {FINANCIAL_CUTOFF}")
    print(f"  Saldo cartão controle: {SALDO_CARTAO_CONTROLE_DATE}")
    print(f"  Saldo cartão carga: {SALDO_CARTAO_CARGA_DATE}")
    print("=" * 70)

    # 1. Load cadastro
    cur.execute("""
        SELECT cpf, colaborador, situacao, status_cartao,
               regional, centro_custo, gestor, diretor
        FROM quinzena_cadastro
        ORDER BY colaborador ASC NULLS LAST
    """)
    cadastro = cur.fetchall()
    print(f"  Cadastro: {len(cadastro)} rows")

    # Build name→cpf map
    cad_name_to_cpf = {}
    for c in cadastro:
        n = norm(c["colaborador"])
        if n:
            cad_name_to_cpf[n] = c["cpf"]

    # 2. Load manual inputs
    cur.execute("""
        SELECT col_1qz::text, adiantamento::text, obs, cpf
        FROM quinzena_manual_inputs
        WHERE year = %s AND month = %s AND quinzena = %s
    """, (YEAR, MONTH, QUINZENA))
    manuals = cur.fetchall()
    manual_by_cpf = {}
    for m in manuals:
        if m["cpf"]:
            manual_by_cpf[m["cpf"]] = m
    print(f"  Manual inputs: {len(manuals)} rows")

    # 3. Load reembolso multiplier
    cur.execute("""
        SELECT reembolso_multiplier::text
        FROM quinzena_config
        WHERE year = %s AND month = %s AND quinzena = %s
    """, (YEAR, MONTH, QUINZENA))
    config = cur.fetchone()
    reembolso_mult = float(config["reembolso_multiplier"]) if config else 0.5
    print(f"  Reembolso multiplier: {reembolso_mult}")

    # 4. Extrato cumulativo (deduped with hora)
    cur.execute("""
        WITH deduped AS (
            SELECT DISTINCT ON (
                UPPER(usuario), data, tipo, valor,
                COALESCE(NULLIF(codigo_transacao, ''), hora::text)
            )
                UPPER(usuario) AS usuario_up,
                data, tipo, valor, codigo_transacao
            FROM extrato_movimentacao
            WHERE is_snapshot = FALSE
              AND data <= %s
            ORDER BY UPPER(usuario), data, tipo, valor,
                COALESCE(NULLIF(codigo_transacao, ''), hora::text)
        )
        SELECT
            usuario_up,
            COALESCE(SUM(valor) FILTER(WHERE tipo = 'Transferência' AND valor > 0), 0) AS carga_raw,
            COALESCE(SUM(valor) FILTER(WHERE tipo = 'Transferência' AND valor < 0), 0) AS transf_raw,
            COALESCE(SUM(valor) FILTER(WHERE tipo = 'Taxa'), 0) AS tarifa_raw
        FROM deduped
        GROUP BY usuario_up
    """, (FINANCIAL_CUTOFF,))
    extrato = cur.fetchall()
    print(f"  Extrato (deduped): {len(extrato)} users")

    # 5. Somase (prestação) — Aprovado+Enviado, excluding FATURA/CARTAO
    cur.execute("""
        SELECT
            r.user_cpf,
            COALESCE(SUM(e.value), 0) AS total
        FROM prestacao_reports r
        JOIN prestacao_expenses e ON e.report_id = r.id
        WHERE (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
          AND r.user_cpf IS NOT NULL
          AND TRIM(r.name) !~* '^(fatu|farur|cart)'
          AND TRIM(r.name) !~* '(fatura|fatuar|fatut|farur)'
        GROUP BY r.user_cpf
    """)
    somase_rows = cur.fetchall()
    somase_by_cpf = {r["user_cpf"]: float(r["total"]) for r in somase_rows}
    print(f"  Somase: {len(somase_by_cpf)} CPFs, R$ {sum(somase_by_cpf.values()):,.2f}")

    # 6. Saldo cartão (controle + carga) — snapshot + post-snapshot txns
    def compute_saldo_cartao(cutoff_date):
        cur.execute("""
            WITH deduped AS (
                SELECT DISTINCT ON (
                    UPPER(usuario), data, tipo, valor,
                    COALESCE(NULLIF(codigo_transacao, ''), hora::text)
                )
                    UPPER(usuario) AS usuario_up, data, tipo, valor, codigo_transacao
                FROM extrato_movimentacao
                WHERE is_snapshot = FALSE
                  AND data <= %s
                ORDER BY UPPER(usuario), data, tipo, valor,
                    COALESCE(NULLIF(codigo_transacao, ''), hora::text)
            ),
            latest_snap AS (
                SELECT DISTINCT ON (UPPER(usuario))
                    UPPER(usuario) AS usuario_up,
                    valor AS saldo,
                    data AS snapshot_date
                FROM extrato_movimentacao
                WHERE is_snapshot = TRUE
                  AND valor IS NOT NULL
                  AND data <= %s
                ORDER BY UPPER(usuario), data DESC
            ),
            post_snap_txns AS (
                SELECT d.usuario_up, SUM(d.valor) AS adjustment
                FROM deduped d
                JOIN latest_snap s ON d.usuario_up = s.usuario_up
                WHERE d.data > s.snapshot_date
                GROUP BY d.usuario_up
            ),
            computed_balance AS (
                SELECT usuario_up, COALESCE(SUM(valor), 0) AS saldo
                FROM deduped
                GROUP BY usuario_up
            )
            SELECT COALESCE(s.usuario_up, c.usuario_up) AS usuario_up,
                   COALESCE(s.saldo, 0) + COALESCE(p.adjustment, 0) AS snap_saldo,
                   COALESCE(c.saldo, 0) AS computed_saldo,
                   (s.usuario_up IS NOT NULL) AS has_snapshot
            FROM latest_snap s
            FULL OUTER JOIN post_snap_txns p ON p.usuario_up = s.usuario_up
            FULL OUTER JOIN computed_balance c ON c.usuario_up = COALESCE(s.usuario_up, p.usuario_up)
        """, (cutoff_date, cutoff_date))
        result = {}
        for r in cur.fetchall():
            has_snap = r["has_snapshot"]
            snap_saldo = float(r["snap_saldo"] or 0)
            computed_saldo = float(r["computed_saldo"] or 0)
            result[r["usuario_up"]] = r2(snap_saldo if has_snap else computed_saldo)
        return result

    saldo_controle = compute_saldo_cartao(SALDO_CARTAO_CONTROLE_DATE)
    saldo_carga = compute_saldo_cartao(SALDO_CARTAO_CARGA_DATE)
    print(f"  Saldo cartão controle: {len(saldo_controle)} users")
    print(f"  Saldo cartão carga: {len(saldo_carga)} users")

    conn.close()

    # 7. Resolve names to CPFs (fuzzy match first, then prefix)
    fuzzy_cache = {}
    def resolve_cpf(nome):
        n = norm(nome)
        if n in cad_name_to_cpf:
            return cad_name_to_cpf[n]
        if n in fuzzy_cache:
            return fuzzy_cache[n]
        best_cpf = None
        best_ratio = 0
        for cn, cpf in cad_name_to_cpf.items():
            r = fuzzy_ratio(n, cn)
            if r > best_ratio:
                best_ratio = r
                best_cpf = cpf
        if best_ratio >= 0.88 and best_cpf:
            fuzzy_cache[n] = best_cpf
            return best_cpf
        # Prefix fallback
        if len(n) >= 10:
            p15 = n[:15]
            for cn, cpf in cad_name_to_cpf.items():
                if cn[:15] == p15:
                    return cpf
            p10 = n[:10]
            for cn, cpf in cad_name_to_cpf.items():
                if cn[:10] == p10:
                    return cpf
        return None

    # 8. Build extrato by CPF
    carga_by_cpf = {}
    transf_by_cpf = {}
    tarifa_by_cpf = {}
    for r in extrato:
        cpf = resolve_cpf(r["usuario_up"])
        if cpf:
            carga = float(r["carga_raw"] or 0)
            transf = abs(float(r["transf_raw"] or 0))
            tarifa = abs(float(r["tarifa_raw"] or 0))
            carga_by_cpf[cpf] = carga
            transf_by_cpf[cpf] = transf
            tarifa_by_cpf[cpf] = tarifa

    # 9. Resolve saldo cartão by CPF
    saldo_controle_by_cpf = {}
    for nome, val in saldo_controle.items():
        cpf = resolve_cpf(nome)
        if cpf:
            saldo_controle_by_cpf[cpf] = val

    saldo_carga_by_cpf = {}
    for nome, val in saldo_carga.items():
        cpf = resolve_cpf(nome)
        if cpf:
            saldo_carga_by_cpf[cpf] = val

    # 10. Build rows
    rows = []
    for snap in cadastro:
        cpf = snap["cpf"]
        manual = manual_by_cpf.get(cpf)

        carga = carga_by_cpf.get(cpf, 0)
        transf = transf_by_cpf.get(cpf, 0)
        tarifa = tarifa_by_cpf.get(cpf, 0)
        prestacao = somase_by_cpf.get(cpf, 0)

        sp = r2(carga - transf - tarifa - prestacao)
        sc_controle = saldo_controle_by_cpf.get(cpf, 0)
        sc_carga = saldo_carga_by_cpf.get(cpf, 0)

        sf = r2(sp - sc_controle)
        saldo_reembolsar = max(-sf, 0)
        saldo_final_carga = max(sf, 0)

        col_qz_manual = None
        if manual and manual["col_1qz"] is not None:
            col_qz_manual = float(manual["col_1qz"])

        adiantamento = 0
        if manual and manual["adiantamento"] is not None:
            adiantamento = float(manual["adiantamento"])

        col_qz_efetivo = col_qz_manual if col_qz_manual is not None else 0

        status_cartao = snap["status_cartao"] or ""
        is_pendente = "pendente" in status_cartao.lower()

        if is_pendente:
            carga_parcial = 0
            reembolso = 0
            carga_final = 0
        else:
            carga_parcial = r2(col_qz_efetivo - saldo_final_carga - sc_carga - adiantamento)
            reembolso = 0 if QUINZENA == 2 else r2(max(0, saldo_reembolsar) * reembolso_mult)
            carga_final = r2(max(0, carga_parcial) + reembolso)

        rows.append({
            "cpf": cpf,
            "colaborador": snap["colaborador"] or "",
            "situacao": snap["situacao"] or "",
            "status_cartao": status_cartao,
            "regional": snap["regional"] or "",
            "centro_custo": snap["centro_custo"] or "",
            "gestor": snap["gestor"] or "",
            "diretor": snap["diretor"] or "",
            "carga": carga,
            "transferencia": transf,
            "tarifa": tarifa,
            "prestacao": prestacao,
            "saldo_prestacao": sp,
            "saldo_cartao": sc_controle,
            "saldo_cartao_carga": sc_carga,
            "saldo_final": sf,
            "saldo_reembolsar": saldo_reembolsar,
            "saldo_final_carga": saldo_final_carga,
            "col_qz_manual": col_qz_manual,
            "adiantamento": adiantamento,
            "obs": manual["obs"] if manual else None,
            "carga_parcial": carga_parcial,
            "reembolso": reembolso,
            "carga_final": carga_final,
        })

    print(f"\n  Total rows: {len(rows)}")
    print(f"  Ativos: {sum(1 for r in rows if r['situacao'].upper() == 'ATIVO')}")
    print(f"  Com carga: {sum(1 for r in rows if r['carga_final'] > 0)}")
    print(f"  Total carga final: R$ {sum(r['carga_final'] for r in rows):,.2f}")
    print(f"  Total saldo final: R$ {sum(r['saldo_final'] for r in rows):,.2f}")

    # === Generate CONTROLE sheet (PAINEL) ===
    generate_controle(rows)
    
    # === Generate CARGA sheet ===
    generate_carga(rows)

    print(f"\n  Done! Files saved to: {OUT_DIR}")


def generate_controle(rows):
    """Generate CONTROLE sheet matching the reference PAINEL structure."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PAINEL"

    # Styles
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", fgColor="1F4E79")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    money_fmt = '#,##0.00;[Red]-#,##0.00'

    # Headers (row 11, matching reference)
    headers = [
        "EMPRESA", "COLABORADOR", "CPF", "CHAVE", "SITUAÇÃO", "STATUS DO CARTÃO",
        "CARTÃO ITAU", "TERMO", "REGIONAL", "CENTRO DE CUSTO", "GESTOR", "DIRETOR",
        "CARTÃO VEXPENSES", "CARGA", "TRANSFERENCIA", "(-) TARIFA",
        "(-) PRESTAÇÃO DE CONTAS", "SALDO PRESTAÇÃO", "(-) SALDO CARTAO",
        "SALDO FINAL", "1ª QZ", "2ª QZ", "ADICIONAIS",
        "SITUAÇÃO COLABORADOR", "CARTÃO CRED. ITAU", "ITAU", "ADICIONAL ITAU"
    ]

    header_row = 11
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = halign
        cell.border = thin_border

    # Data rows
    for ri, r in enumerate(rows, header_row + 1):
        values = [
            "EQS",  # EMPRESA
            r["colaborador"],
            r["cpf"],
            "",  # CHAVE
            r["situacao"],
            r["status_cartao"],
            "",  # CARTÃO ITAU
            "",  # TERMO
            r["regional"],
            r["centro_custo"],
            r["gestor"],
            r["diretor"],
            "SIM",  # CARTÃO VEXPENSES
            r["carga"],
            -r["transferencia"],  # negative
            -r["tarifa"],  # negative
            -r["prestacao"],  # negative
            r["saldo_prestacao"],
            -r["saldo_cartao"],  # negative
            r["saldo_final"],
            "",  # 1ª QZ (manual, not set for this QZ)
            r["col_qz_manual"] if r["col_qz_manual"] is not None else "",
            "",  # ADICIONAIS
            "",  # SITUAÇÃO COLABORADOR
            "",  # CARTÃO CRED. ITAU
            "",  # ITAU
            "",  # ADICIONAL ITAU
        ]
        for ci, v in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin_border
            # Money format for columns 14-20 (CARGA through SALDO FINAL)
            if 14 <= ci <= 20:
                cell.number_format = money_fmt
            elif ci in (21, 22):  # QZ columns
                cell.number_format = money_fmt

    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

    # Freeze panes
    ws.freeze_panes = "B12"

    fname = OUT_DIR / f"CONTROLE - VEXPENSES - JULHO 2026 - 2QZ - API.xlsx"
    wb.save(fname)
    print(f"  ✅ CONTROLE saved: {fname}")


def generate_carga(rows):
    """Generate CARGA sheet matching the reference structure."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"2 QZ VEXPENSES 07_2026"

    # Styles
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", fgColor="1F4E79")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    money_fmt = '#,##0.00;[Red]-#,##0.00'

    # Headers (row 5, matching reference carga sheet structure)
    headers = [
        "",  # col 0 (empty)
        "PORTADOR",  # col 1
        "CPF",  # col 2
        "STATUS COLAB",  # col 3
        "CENTRO CUSTO",  # col 4
        "COD CENTRO CUSTO",  # col 5
        "GESTOR",  # col 6
        "DIREÇÃO",  # col 7
        "SALDO REEMBOLSAR",  # col 8
        "SALDO FINAL",  # col 9
        f"{QUINZENA}QZ {MONTH:02d}/{YEAR}",  # col 10 — 2QZ JULHO 26
        "SALDO CARTAO",  # col 11
        "ADIANTAMENTO",  # col 12
        "CARGA PARCIAL",  # col 13
        "REEMBOLSO",  # col 14
        "CARGA FINAL",  # col 15
        "STATUS DO CARTAO",  # col 16
        "OBS",  # col 17
    ]

    header_row = 5
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = halign
        cell.border = thin_border

    # Data rows
    for ri, r in enumerate(rows, header_row + 1):
        values = [
            "",  # col 0
            r["colaborador"],  # PORTADOR
            r["cpf"],  # CPF
            r["situacao"],  # STATUS COLAB
            r["centro_custo"],  # CENTRO CUSTO
            "",  # COD CENTRO CUSTO
            r["gestor"],  # GESTOR
            r["diretor"],  # DIREÇÃO
            r["saldo_reembolsar"],  # SALDO REEMBOLSAR
            r["saldo_final"],  # SALDO FINAL
            r["col_qz_manual"] if r["col_qz_manual"] is not None else "",  # 2QZ
            r["saldo_cartao_carga"],  # SALDO CARTAO (carga date)
            r["adiantamento"] if r["adiantamento"] > 0 else "",  # ADIANTAMENTO
            r["carga_parcial"],  # CARGA PARCIAL
            r["reembolso"],  # REEMBOLSO
            r["carga_final"],  # CARGA FINAL
            r["status_cartao"],  # STATUS DO CARTAO
            r["obs"] or "",  # OBS
        ]
        for ci, v in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin_border
            # Money format for columns 9-16
            if 9 <= ci <= 16:
                cell.number_format = money_fmt

    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

    # Freeze panes
    ws.freeze_panes = "B6"

    fname = OUT_DIR / f"CARGA - 2QZ JULHO 2026 - VEXPENSES - API.xlsx"
    wb.save(fname)
    print(f"  ✅ CARGA saved: {fname}")


if __name__ == "__main__":
    main()
