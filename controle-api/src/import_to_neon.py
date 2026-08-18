"""
import_to_neon.py
-----------------
Importa dados das planilhas de Controle e de Carga para o banco Neon.

Popula quinzena_controle_snapshot com:

  DO CONTROLE (aba PAINEL) — dados mensais, iguais para QZ1 e QZ2:
    colaborador, cpf, situacao, status_cartao, regional, centro_custo,
    gestor, diretor, saldo_prestacao, saldo_cartao, saldo_final

  DA CARGA (aba principal) — dados específicos da quinzena:
    col_qz          (1ª QZ ou 2ª QZ — valor que o colaborador recebe)
    saldo_reembolsar (SALDO REEMBOLSAR da CARGA — base para o reembolso)

Estrutura das planilhas confirmada por inspect_all.py:

  CONTROLE — aba PAINEL:
    header = linha 11, dados a partir da linha 12
    col  2 = COLABORADOR  (idx 1)
    col  3 = CPF          (idx 2)
    col  5 = SITUACAO     (idx 4)
    col  6 = STATUS CARTAO(idx 5)
    col  9 = REGIONAL     (idx 8)
    col 10 = CENTRO CUSTO (idx 9)
    col 11 = GESTOR       (idx 10)
    col 12 = DIRETOR      (idx 11)
    col 18 = SALDO PRESTACAO (idx 17)
    col 19 = (-) SALDO CARTAO(idx 18)
    col 20 = SALDO FINAL  (idx 19)

  CARGA 1QZ — aba Planilha1:
    header = linha 6, dados a partir da linha 7
    col  1 = COLABORADOR  (idx 0)
    col  2 = CPF          (idx 1)
    col  8 = SALDO REEMBOLSAR (idx 7)
    col 10 = 1a QZ        (idx 9)

  CARGA 2QZ — aba "2 QZ DE MAIO 26" (ou similar, detectado automaticamente):
    header = linha 4, dados a partir da linha 5
    col  2 = COLABORADOR  (idx 1)
    col  3 = CPF          (idx 2)
    col  8 = SALDO PENDENTE PARCIAL (idx 7)  — equivale ao SALDO REEMBOLSAR
    col 11 = 2a QZ        (idx 10)

Uso:
  # Importar so o Controle (saldos mensais):
  python src/import_to_neon.py --controle "data/CONTROLE ..." --year 2026 --month 5 --quinzena 1

  # Importar Controle + Carga (recomendado, dados completos):
  python src/import_to_neon.py --controle "data/CONTROLE ..." --carga "data/CARGA 1 QZ ..." --year 2026 --month 5 --quinzena 1
  python src/import_to_neon.py --controle "data/CONTROLE ..." --carga "data/CARGA 2 QZ ..." --year 2026 --month 5 --quinzena 2

  # Dry-run (sem gravar):
  python src/import_to_neon.py ... --dry-run
"""

import argparse
import os
import sys
from decimal import Decimal, InvalidOperation

import openpyxl
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))

NEON_DATABASE_URL = os.getenv("NEON_DATABASE_URL")

# =============================================================================
# Mapeamento de colunas (indices 0-based, confirmados por inspect_all.py)
# =============================================================================

# --- CONTROLE, aba PAINEL ----------------------------------------------------
PAINEL_HEADER_ROW = 11
PAINEL_DATA_START = 12
PAINEL_COLS = {
    "colaborador":          1,   # col  2
    "cpf":                  2,   # col  3
    "situacao":             4,   # col  5
    "status_cartao":        5,   # col  6
    "regional":             8,   # col  9
    "centro_custo":         9,   # col 10
    "gestor":              10,   # col 11
    "diretor":             11,   # col 12
    "saldo_prestacao":     17,   # col 18
    "saldo_cartao_painel": 18,   # col 19  (-) SALDO CARTAO
    "saldo_final":         19,   # col 20
}

# --- CONTROLE, aba SALDO CARTAO (tabela resumo, lado direito) ----------------
SALDO_RESUMO_HEADER_ROW = 4
SALDO_RESUMO_DATA_START = 5
SALDO_RESUMO_COLS = {
    "portador": 9,   # col 10
    "cpf":     10,   # col 11
    "valor":   11,   # col 12
}

# --- CARGA 1QZ, aba Planilha1 ------------------------------------------------
# Estrutura confirmada por inspecao direta (inspect_all.py):
#   L6 = cabecalho: col0=COLABORADOR col1=CPF col2=SITUACAO col3=REGIONAL
#                   col4=CENTRO col5=GESTOR col6=DIRETOR
#                   col7=SALDO REEMBOLSAR  col8=SALDO FINAL  col9=1a QZ
#                   col10=SALDO CARTAO     col11=Adiantamento
#                   col12=CARGA PARCIAL    col13=REEMBOLSO   col14=Carga Final
CARGA1_SHEET  = "Planilha1"
CARGA1_HEADER = 6
CARGA1_DATA   = 7
CARGA1_COLS = {
    "colaborador":      0,   # col  1 = COLABORADOR
    "cpf":              1,   # col  2 = CPF
    "saldo_reembolsar": 7,   # col  8 = SALDO REEMBOLSAR
    "saldo_final_carga":8,   # col  9 = SALDO FINAL (usado no calculo da carga)
    "col_qz":           9,   # col 10 = 1a QZ
    "saldo_cartao_carga":10, # col 11 = SALDO CARTAO
}

# --- CARGA 2QZ (aba detectada automaticamente — nome muda por mes) -----------
# Estrutura confirmada por inspecao direta:
#   L4 = cabecalho: col1=COLABORADOR col2=CPF col3=SITUACAO col4=CENTRO
#                   col5=GESTOR col6=DIRETOR
#                   col7=SALDO PENDENTE PARCIAL  col8=CARGA 1 QZ
#                   col9=SALDO FINAL  col10=2a QZ  col11=SALDO CARTAO
#                   col12=Adiantamento  col13=CARGA PARCIAL
#                   col14=REEMBOLSO  col15=Carga Final
CARGA2_HEADER = 4
CARGA2_DATA   = 5
CARGA2_COLS = {
    "colaborador":       1,   # col  2 = COLABORADOR
    "cpf":               2,   # col  3 = CPF
    "saldo_reembolsar":  7,   # col  8 = SALDO PENDENTE PARCIAL
    "saldo_final_carga": 9,   # col 10 = SALDO FINAL
    "col_qz":           10,   # col 11 = 2a QZ
    "saldo_cartao_carga":11,  # col 12 = SALDO CARTAO
}


# =============================================================================
# Helpers
# =============================================================================

def normalize_cpf(raw) -> "str | None":
    if raw is None:
        return None
    s = str(raw).strip().replace(".", "").replace("-", "").replace("/", "").replace(" ", "")
    if "." in s:
        s = s.split(".")[0]
    return s.zfill(11) if s.isdigit() else None


def safe_decimal(raw) -> "Decimal | None":
    if raw is None:
        return None
    try:
        return Decimal(str(raw)).quantize(Decimal("0.0001"))
    except (InvalidOperation, ValueError):
        return None


def get_cell(row_values: list, col_idx: int):
    return row_values[col_idx] if col_idx < len(row_values) else None


# =============================================================================
# Leitores de planilha
# =============================================================================

def read_painel(wb: openpyxl.Workbook) -> "dict[str, dict]":
    """Le aba PAINEL do Controle. Retorna {cpf: record}."""
    ws = wb["PAINEL"]
    records: "dict[str, dict]" = {}

    # Validacao do cabecalho
    header = list(ws.iter_rows(
        min_row=PAINEL_HEADER_ROW, max_row=PAINEL_HEADER_ROW, values_only=True
    ))[0]
    expected = {
        PAINEL_COLS["colaborador"]:  "COLABORADOR",
        PAINEL_COLS["cpf"]:          "CPF",
        PAINEL_COLS["status_cartao"]: "STATUS",
        PAINEL_COLS["regional"]:     "REGIONAL",
        PAINEL_COLS["saldo_final"]:  "SALDO FINAL",
    }
    for idx, name in expected.items():
        actual = str(header[idx]).strip() if idx < len(header) and header[idx] else ""
        if name.upper() not in actual.upper():
            print(f"  [AVISO] PAINEL col {idx+1}: esperado '{name}', encontrado '{actual}'")

    for row in ws.iter_rows(min_row=PAINEL_DATA_START, values_only=True):
        cpf = normalize_cpf(get_cell(row, PAINEL_COLS["cpf"]))
        if not cpf:
            continue
        records[cpf] = {
            "colaborador":   str(get_cell(row, PAINEL_COLS["colaborador"]) or "").strip(),
            "situacao":      str(get_cell(row, PAINEL_COLS["situacao"]) or "").strip(),
            "status_cartao": str(get_cell(row, PAINEL_COLS["status_cartao"]) or "").strip(),
            "regional":      str(get_cell(row, PAINEL_COLS["regional"]) or "").strip(),
            "centro_custo":  str(get_cell(row, PAINEL_COLS["centro_custo"]) or "").strip(),
            "gestor":        str(get_cell(row, PAINEL_COLS["gestor"]) or "").strip(),
            "diretor":       str(get_cell(row, PAINEL_COLS["diretor"]) or "").strip(),
            "saldo_prestacao":      safe_decimal(get_cell(row, PAINEL_COLS["saldo_prestacao"])),
            "saldo_cartao":         safe_decimal(get_cell(row, PAINEL_COLS["saldo_cartao_painel"])),
            "saldo_final":          safe_decimal(get_cell(row, PAINEL_COLS["saldo_final"])),
        }

    print(f"  PAINEL: {len(records)} colaboradores.")
    return records


def read_saldo_cartao_resumo(wb: openpyxl.Workbook) -> "dict[str, Decimal | None]":
    """Le tabela resumo da aba SALDO CARTAO (colunas J-L). Retorna {cpf: valor}."""
    ws = wb["SALDO CARTAO"]
    resumo: "dict[str, Decimal | None]" = {}
    for row in ws.iter_rows(min_row=SALDO_RESUMO_DATA_START, values_only=True):
        cpf = normalize_cpf(get_cell(row, SALDO_RESUMO_COLS["cpf"]))
        valor = safe_decimal(get_cell(row, SALDO_RESUMO_COLS["valor"]))
        if cpf and cpf not in resumo:
            resumo[cpf] = valor
    print(f"  SALDO CARTAO resumo: {len(resumo)} CPFs.")
    return resumo


def read_carga(wb: openpyxl.Workbook, quinzena: int) -> "dict[str, dict]":
    """
    Le a planilha de Carga e retorna {cpf: {col_qz, saldo_reembolsar}}.
    Detecta automaticamente a aba correta para a quinzena informada.
    """
    if quinzena == 1:
        sheet_name = CARGA1_SHEET
        header_row = CARGA1_HEADER
        data_start = CARGA1_DATA
        cols = CARGA1_COLS
    else:
        # Para 2QZ o nome da aba muda por mes; pega a primeira que nao e STATUS
        sheet_name = None
        for name in wb.sheetnames:
            if "STATUS" not in name.upper():
                sheet_name = name
                break
        if sheet_name is None:
            print("  [AVISO] Nenhuma aba de dados encontrada na CARGA 2QZ.")
            return {}
        header_row = CARGA2_HEADER
        data_start = CARGA2_DATA
        cols = CARGA2_COLS

    ws = wb[sheet_name]

    # Validacao do cabecalho
    header = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    cpf_col_name = str(header[cols["cpf"]]).strip() if len(header) > cols["cpf"] and header[cols["cpf"]] else ""
    if "CPF" not in cpf_col_name.upper():
        print(f"  [AVISO] CARGA aba '{sheet_name}' col {cols['cpf']+1}: "
              f"esperado 'CPF', encontrado '{cpf_col_name}'")

    records: "dict[str, dict]" = {}
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        cpf = normalize_cpf(get_cell(row, cols["cpf"]))
        if not cpf:
            continue
        records[cpf] = {
            "col_qz":            safe_decimal(get_cell(row, cols["col_qz"])),
            "saldo_reembolsar":  safe_decimal(get_cell(row, cols["saldo_reembolsar"])),
            "saldo_final_carga": safe_decimal(get_cell(row, cols["saldo_final_carga"])),
            "saldo_cartao_carga":safe_decimal(get_cell(row, cols["saldo_cartao_carga"])),
        }

    label = f"{quinzena}QZ"
    print(f"  CARGA {label} (aba '{sheet_name}'): {len(records)} colaboradores.")
    return records


# =============================================================================
# Importacao principal
# =============================================================================

def import_to_neon(
    controle_path: str,
    carga_path: "str | None",
    year: int,
    month: int,
    quinzena: int,
    dry_run: bool = False,
) -> dict:
    filename_controle = os.path.basename(controle_path)
    filename_carga    = os.path.basename(carga_path) if carga_path else "N/A"

    print(f"\nImportando: {filename_controle}")
    if carga_path:
        print(f"  + Carga:   {filename_carga}")
    print(f"Periodo: {year}/{month:02d} — Quinzena {quinzena}")
    print(f"Modo: {'DRY RUN' if dry_run else 'GRAVANDO no Neon'}")

    # --- Le planilhas --------------------------------------------------------
    print("\nCarregando planilha de Controle...")
    wb_controle = openpyxl.load_workbook(controle_path, read_only=True, data_only=True)
    painel      = read_painel(wb_controle)
    saldo_cartao_resumo = read_saldo_cartao_resumo(wb_controle)
    wb_controle.close()

    carga: "dict[str, dict]" = {}
    if carga_path:
        print("Carregando planilha de Carga...")
        wb_carga = openpyxl.load_workbook(carga_path, read_only=True, data_only=True)
        carga = read_carga(wb_carga, quinzena)
        wb_carga.close()

    # --- Merge ---------------------------------------------------------------
    rows_to_insert: list[dict] = []
    for cpf, rec in painel.items():
        saldo_cartao = saldo_cartao_resumo.get(cpf, rec["saldo_cartao"])
        carga_rec    = carga.get(cpf, {})

        rows_to_insert.append({
            "year":               year,
            "month":              month,
            "quinzena":           quinzena,
            "cpf":                cpf,
            "colaborador":        rec["colaborador"] or None,
            "situacao":           rec["situacao"] or None,
            "status_cartao":      rec["status_cartao"] or None,
            "regional":           rec["regional"] or None,
            "centro_custo":       rec["centro_custo"] or None,
            "gestor":             rec["gestor"] or None,
            "diretor":            rec["diretor"] or None,
            # Controle: campos cadastrais/informativos
            "saldo_prestacao":    rec["saldo_prestacao"],
            "saldo_cartao":       saldo_cartao,
            "saldo_final":        rec["saldo_final"],
            # Carga: valores financeiros usados nas formulas
            "col_qz":             carga_rec.get("col_qz"),
            "saldo_reembolsar":   carga_rec.get("saldo_reembolsar"),
            "saldo_final_carga":  carga_rec.get("saldo_final_carga"),
            "saldo_cartao_carga": carga_rec.get("saldo_cartao_carga"),
            "import_source":      f"{filename_controle} + {filename_carga}",
        })

    stats = {"rows_imported": 0, "rows_skipped": 0, "rows_failed": 0}

    if dry_run:
        print(f"\n[DRY RUN] {len(rows_to_insert)} linhas seriam importadas.")
        # Mostra as primeiras linhas com CPF conhecido para validacao
        cpfs_amostra = ('02027745203', '01932662537', '85087572634', '06223031980')
        for s in rows_to_insert:
            if s['cpf'] in cpfs_amostra:
                print(f"  CPF={s['cpf']} | {s['colaborador'][:30]}")
                print(f"    col_qz={s['col_qz']} | saldo_reembolsar={s['saldo_reembolsar']}")
                print(f"    saldo_final_carga={s['saldo_final_carga']} | saldo_cartao_carga={s['saldo_cartao_carga']}")
        return stats

    if not NEON_DATABASE_URL:
        print("ERRO: NEON_DATABASE_URL nao configurada no .env")
        sys.exit(1)

    conn = psycopg2.connect(NEON_DATABASE_URL)
    conn.autocommit = False
    cur  = conn.cursor()

    upsert_sql = """
        INSERT INTO quinzena_controle_snapshot
          (year, month, quinzena, cpf,
           colaborador, situacao, status_cartao,
           regional, centro_custo, gestor, diretor,
           saldo_prestacao, saldo_cartao, saldo_final,
           col_qz, saldo_reembolsar, saldo_final_carga, saldo_cartao_carga,
           import_source, imported_at)
        VALUES
          (%(year)s, %(month)s, %(quinzena)s, %(cpf)s,
           %(colaborador)s, %(situacao)s, %(status_cartao)s,
           %(regional)s, %(centro_custo)s, %(gestor)s, %(diretor)s,
           %(saldo_prestacao)s, %(saldo_cartao)s, %(saldo_final)s,
           %(col_qz)s, %(saldo_reembolsar)s, %(saldo_final_carga)s, %(saldo_cartao_carga)s,
           %(import_source)s, NOW())
        ON CONFLICT ON CONSTRAINT uq_snapshot
        DO UPDATE SET
          colaborador         = EXCLUDED.colaborador,
          situacao            = EXCLUDED.situacao,
          status_cartao       = EXCLUDED.status_cartao,
          regional            = EXCLUDED.regional,
          centro_custo        = EXCLUDED.centro_custo,
          gestor              = EXCLUDED.gestor,
          diretor             = EXCLUDED.diretor,
          saldo_prestacao     = EXCLUDED.saldo_prestacao,
          saldo_cartao        = EXCLUDED.saldo_cartao,
          saldo_final         = EXCLUDED.saldo_final,
          col_qz              = EXCLUDED.col_qz,
          saldo_reembolsar    = EXCLUDED.saldo_reembolsar,
          saldo_final_carga   = EXCLUDED.saldo_final_carga,
          saldo_cartao_carga  = EXCLUDED.saldo_cartao_carga,
          import_source       = EXCLUDED.import_source,
          imported_at         = NOW()
    """

    log_sql = """
        INSERT INTO quinzena_import_log
          (year, month, quinzena, filename, rows_imported, rows_skipped, rows_failed, status)
        VALUES
          (%(year)s, %(month)s, %(quinzena)s, %(filename)s,
           %(rows_imported)s, %(rows_skipped)s, %(rows_failed)s, %(status)s)
    """

    try:
        for row in rows_to_insert:
            try:
                cur.execute(upsert_sql, row)
                stats["rows_imported"] += 1
            except Exception as e:
                stats["rows_failed"] += 1
                print(f"  [ERRO] CPF {row['cpf']}: {e}")
                conn.rollback()
                conn.autocommit = False

        cur.execute(log_sql, {
            "year": year, "month": month, "quinzena": quinzena,
            "filename": f"{filename_controle} + {filename_carga}",
            "rows_imported": stats["rows_imported"],
            "rows_skipped":  stats["rows_skipped"],
            "rows_failed":   stats["rows_failed"],
            "status": "success" if stats["rows_failed"] == 0 else "partial",
        })

        conn.commit()
        print(f"\nConcluido: {stats['rows_imported']} importados, "
              f"{stats['rows_skipped']} ignorados, {stats['rows_failed']} falhas.")

    except Exception as e:
        conn.rollback()
        print(f"\nERRO fatal: {e}")
        stats["rows_failed"] = len(rows_to_insert)
        raise
    finally:
        cur.close()
        conn.close()

    return stats


# =============================================================================
# CLI
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Importa planilhas de Controle e Carga VExpenses para o Neon."
    )
    parser.add_argument("--controle", required=True,
                        help="Caminho para o CONTROLE - VEXPENSES - *.xlsx")
    parser.add_argument("--carga", default=None,
                        help="Caminho para a planilha de CARGA da quinzena (opcional, mas recomendado)")
    parser.add_argument("--year",     type=int, required=True)
    parser.add_argument("--month",    type=int, required=True)
    parser.add_argument("--quinzena", type=int, choices=[1, 2], required=True)
    parser.add_argument("--dry-run",  action="store_true")
    args = parser.parse_args()

    def resolve(path):
        if path and not os.path.isabs(path):
            path = os.path.join(os.path.dirname(__file__), "..", path)
        return os.path.normpath(path) if path else None

    controle_path = resolve(args.controle)
    carga_path    = resolve(args.carga)

    if not os.path.exists(controle_path):
        print(f"ERRO: arquivo nao encontrado: {controle_path}")
        sys.exit(1)
    if carga_path and not os.path.exists(carga_path):
        print(f"ERRO: arquivo nao encontrado: {carga_path}")
        sys.exit(1)

    import_to_neon(
        controle_path=controle_path,
        carga_path=carga_path,
        year=args.year,
        month=args.month,
        quinzena=args.quinzena,
        dry_run=args.dry_run,
    )


if __name__ == "__main__":
    main()
