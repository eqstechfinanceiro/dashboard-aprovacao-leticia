#!/usr/bin/env python3
"""
calcular_quinzena_neon.py
--------------------------
Pipeline completo 100% API/Neon -> planilha de Carga Quinzenal.

Nao precisa de nenhuma planilha .xlsx como entrada.
Fontes de dados:
  - extrato_movimentacao  (Neon) -> CARGA, TRANSFERENCIA, TARIFA, SALDO CARTAO
  - somase_snapshots      (Neon) -> PRESTACAO DE CONTAS (acumulado por CPF)
  - quinzena_controle_snapshot (Neon) -> col_qz (valor da quinzena), dados cadastrais

Uso:
  # Calcular e comparar (modo validacao, sem salvar):
  python src/calcular_quinzena_neon.py --ano 2026 --mes 5 --quinzena 1
  python src/calcular_quinzena_neon.py --ano 2026 --mes 5 --quinzena 2

  # Gerar planilha Excel:
  python src/calcular_quinzena_neon.py --ano 2026 --mes 5 --quinzena 1 --output data/carga_1qz_maio.xlsx

  # Comparar com planilha de referencia:
  python src/calcular_quinzena_neon.py --ano 2026 --mes 5 --quinzena 1 --ref "data/CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx"
"""

import argparse
import logging
import os
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Optional

import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

BASE = Path(__file__).parent.parent
load_dotenv(BASE / ".env")

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
logger = logging.getLogger(__name__)

NEON_URL = os.getenv("NEON_DATABASE_URL")


# =============================================================================
# Regra de quinzena (validada em regra_quinzena.json)
# =============================================================================

def get_periodo(ano: int, mes: int, quinzena: int) -> tuple[str, str, str]:
    """Retorna (data_inicio, data_fim, data_fechamento) para a quinzena.

    Regra validada em jul/2026:
      1ª QZ: período 26 do mês anterior → 10 do mês atual  (fechamento dia 11)
      2ª QZ: período 11 → 25 do mês atual                  (fechamento dia 25)

    Cutoff financeiro (carga/transf/tarifa/prestação): dia 30 do mês anterior
    Saldo cartão CONTROLE: último snapshot até dia 1 do mês atual
    Saldo cartão CARGA: último snapshot até a data de fechamento (11 ou 25)
    """
    mes_ant = mes - 1 if mes > 1 else 12
    ano_ant = ano if mes > 1 else ano - 1
    if quinzena == 1:
        inicio = f"{ano_ant}-{mes_ant:02d}-26"
        fim = f"{ano}-{mes:02d}-10"
        fechamento = f"{ano}-{mes:02d}-11"
    else:
        inicio = f"{ano}-{mes:02d}-11"
        fim = f"{ano}-{mes:02d}-25"
        fechamento = f"{ano}-{mes:02d}-25"
    return inicio, fim, fechamento


def get_cutoff_financeiro(ano: int, mes: int) -> str:
    """Cutoff financeiro: dia 30 do mês anterior (mesmo para QZ1 e QZ2)."""
    mes_ant = mes - 1 if mes > 1 else 12
    ano_ant = ano if mes > 1 else ano - 1
    return f"{ano_ant}-{mes_ant:02d}-30"


def get_saldo_cartao_controle_date(ano: int, mes: int) -> str:
    """Saldo cartão CONTROLE: último snapshot até dia 1 do mês atual."""
    return f"{ano}-{mes:02d}-01"


def get_quinzena_id(ano: int, mes: int, quinzena: int) -> str:
    """Chave usada em somase_snapshots ex: '2026-05-1'."""
    return f"{ano}-{mes:02d}-{quinzena}"


def get_quinzena_anterior(ano: int, mes: int, quinzena: int) -> tuple[int, int, int]:
    """Retorna (ano, mes, qz) da quinzena imediatamente anterior."""
    if quinzena == 2:
        return ano, mes, 1
    # qz == 1: volta para qz 2 do mes anterior
    if mes == 1:
        return ano - 1, 12, 2
    return ano, mes - 1, 2


# =============================================================================
# Queries ao Neon
# =============================================================================

def buscar_extrato_quinzena(conn, inicio: str, fim: str) -> dict:
    """
    Retorna dict {usuario_upper: {carga, transferencia, tarifa}} para o periodo.
    Usuarios em UPPERCASE para facilitar join posterior.
    """
    cur = conn.cursor()
    cur.execute("""
        SELECT
            UPPER(usuario) AS u,
            COALESCE(SUM(valor) FILTER(WHERE tipo = 'Transferência' AND valor > 0), 0) AS carga,
            COALESCE(ABS(SUM(valor) FILTER(WHERE tipo = 'Transferência' AND valor < 0)), 0) AS transferencia,
            COALESCE(ABS(SUM(valor) FILTER(WHERE tipo = 'Taxa')), 0) AS tarifa
        FROM extrato_movimentacao
        WHERE data BETWEEN %s AND %s
          AND is_snapshot = FALSE
        GROUP BY UPPER(usuario)
    """, (inicio, fim))
    return {r[0]: {"carga": float(r[1]), "transferencia": float(r[2]), "tarifa": float(r[3])}
            for r in cur.fetchall()}


def buscar_extrato_acumulado(conn, fim: str) -> dict:
    """
    Retorna dict {usuario_upper: {carga, transferencia, tarifa}} acumulado até fim.
    Usado para calcular SALDO PRESTAÇÃO direto no fechamento da quinzena.
    """
    cur = conn.cursor()
    cur.execute("""
        SELECT
            UPPER(usuario) AS u,
            COALESCE(SUM(valor) FILTER(WHERE tipo = 'Transferência' AND valor > 0), 0) AS carga,
            COALESCE(ABS(SUM(valor) FILTER(WHERE tipo = 'Transferência' AND valor < 0)), 0) AS transferencia,
            COALESCE(ABS(SUM(valor) FILTER(WHERE tipo = 'Taxa')), 0) AS tarifa
        FROM extrato_movimentacao
        WHERE data <= %s
          AND is_snapshot = FALSE
        GROUP BY UPPER(usuario)
    """, (fim,))
    return {r[0]: {"carga": float(r[1]), "transferencia": float(r[2]), "tarifa": float(r[3])}
            for r in cur.fetchall()}


def buscar_saldo_cartao(conn, fechamento: str, quinzena: int) -> dict:
    """
    Retorna dict {usuario_upper: saldo_cartao} usando o ultimo snapshot <= fechamento.
    Para 2a QZ, ignora o proprio dia 25 (usa < fechamento) pois snapshots do dia
    de fechamento ja refletem a carga da proxima quinzena.
    """
    cur = conn.cursor()
    op = "<=" if quinzena == 1 else "<"
    cur.execute(f"""
        SELECT UPPER(m.usuario), m.valor
        FROM extrato_movimentacao m
        WHERE m.is_snapshot = TRUE
          AND m.valor IS NOT NULL
          AND m.data = (
              SELECT MAX(m2.data)
              FROM extrato_movimentacao m2
              WHERE UPPER(m2.usuario) = UPPER(m.usuario)
                AND m2.is_snapshot = TRUE
                AND m2.valor IS NOT NULL
                AND m2.data {op} %s
          )
    """, (fechamento,))
    result = {}
    for usuario_up, valor in cur.fetchall():
        result[usuario_up] = float(valor)
    return result


def buscar_somase(conn, quinzena_id: str) -> dict:
    """Retorna dict {cpf: total_prestacao} da somase_snapshots."""
    cur = conn.cursor()
    cur.execute("SELECT user_cpf, total FROM somase_snapshots WHERE quinzena = %s", (quinzena_id,))
    return {r[0]: float(r[1]) for r in cur.fetchall()}


def buscar_controle_snapshot(conn, ano: int, mes: int, quinzena: int) -> dict:
    """
    Retorna dict {cpf: record} da quinzena_cadastro.
    Inclui dados cadastrais. Para July 2026+ (100% API), usa cadastro direto.
    """
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT cpf, colaborador, situacao, status_cartao, regional,
               centro_custo, gestor, diretor
        FROM quinzena_cadastro
        ORDER BY colaborador ASC NULLS LAST
    """)
    return {r["cpf"]: dict(r) for r in cur.fetchall()}


def buscar_reembolso_multiplier(conn, ano: int, mes: int, quinzena: int) -> float:
    """Retorna o multiplicador de reembolso da quinzena (default 0.5)."""
    cur = conn.cursor()
    cur.execute("""
        SELECT reembolso_multiplier
        FROM quinzena_config
        WHERE year = %s AND month = %s AND quinzena = %s
    """, (ano, mes, quinzena))
    row = cur.fetchone()
    return float(row[0]) if row else 0.5


# =============================================================================
# Pipeline principal
# =============================================================================

def _r2(v: float) -> float:
    """Arredonda para 2 casas decimais (ROUND_HALF_UP como o Excel)."""
    return float(Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def salvar_quinzena_snapshot(conn, ano: int, mes: int, quinzena: int, linhas: list[dict]) -> None:
    """Atualiza quinzena_controle_snapshot com os valores calculados."""
    cur = conn.cursor()
    values = []
    for l in linhas:
        values.append((
            l["_saldo_prestacao"],
            l["saldo_cartao"],
            l["_saldo_final_painel"],
            l["saldo_reembolsar"],
            l["saldo_cartao"],
            l["saldo_final"],
            l["col_qz"],
            "calc",
            ano, mes, quinzena, l["cpf"],
        ))
    psycopg2.extras.execute_batch(cur, """
        UPDATE quinzena_controle_snapshot
        SET saldo_prestacao = COALESCE(%s, saldo_prestacao),
            saldo_cartao = COALESCE(%s, saldo_cartao),
            saldo_final = COALESCE(%s, saldo_final),
            saldo_reembolsar = COALESCE(%s, saldo_reembolsar),
            saldo_cartao_carga = COALESCE(%s, saldo_cartao_carga),
            saldo_final_carga = COALESCE(%s, saldo_final_carga),
            col_qz = COALESCE(%s, col_qz),
            import_source = COALESCE(NULLIF(import_source, ''), %s)
        WHERE year = %s AND month = %s AND quinzena = %s AND cpf = %s
    """, values)
    conn.commit()
    cur.close()
    logger.info(f"Snapshot salvo: {len(values)} linhas")


def calcular_quinzena(
    ano: int,
    mes: int,
    quinzena: int,
    manuais: Optional[dict] = None,
    save: bool = False,
) -> list[dict]:
    """
    Calcula a Carga Quinzenal completa usando apenas dados do Neon.

    Para quinzenas importadas de planilha, usa os valores importados.
    Para quinzenas geradas 100% API (import_source='api'), usa a formula
    incremental (ancora + delta) a partir da quinzena anterior.
    """
    manuais = manuais or {}

    inicio, fim, fechamento = get_periodo(ano, mes, quinzena)
    cutoff_fin = get_cutoff_financeiro(ano, mes)
    saldo_controle_date = get_saldo_cartao_controle_date(ano, mes)
    quinzena_id = get_quinzena_id(ano, mes, quinzena)
    prev_ano, prev_mes, prev_qz = get_quinzena_anterior(ano, mes, quinzena)
    prev_quinzena_id = get_quinzena_id(prev_ano, prev_mes, prev_qz)
    prevprev_ano, prevprev_mes, prevprev_qz = get_quinzena_anterior(prev_ano, prev_mes, prev_qz)
    prevprev_quinzena_id = get_quinzena_id(prevprev_ano, prevprev_mes, prevprev_qz)

    logger.info(f"Calculando {quinzena}ª QZ {mes:02d}/{ano} | periodo {inicio} -> {fim} | fech {fechamento}")
    logger.info(f"  Cutoff financeiro: {cutoff_fin} | Saldo cartão controle: {saldo_controle_date} | Saldo cartão carga: {fechamento}")

    if not NEON_URL:
        raise RuntimeError("NEON_DATABASE_URL nao configurada no .env")

    conn = psycopg2.connect(NEON_URL)

    logger.info("Buscando dados do Neon...")
    # Carga/transf/tarifa: acumulado até cutoff financeiro (dia 30 do mês anterior)
    extrato_acum    = buscar_extrato_acumulado(conn, cutoff_fin)
    # Saldo cartão CONTROLE: último snapshot até dia 1 do mês atual
    snapshots_controle = buscar_saldo_cartao(conn, saldo_controle_date, 1)
    # Saldo cartão CARGA: último snapshot até data de fechamento (11 ou 25)
    snapshots_carga = buscar_saldo_cartao(conn, fechamento, quinzena)
    somase          = buscar_somase(conn, quinzena_id)
    prev_somase     = buscar_somase(conn, prev_quinzena_id)
    prevprev_somase = buscar_somase(conn, prevprev_quinzena_id)
    controle        = buscar_controle_snapshot(conn, ano, mes, quinzena)
    prev_controle   = buscar_controle_snapshot(conn, prev_ano, prev_mes, prev_qz)
    reembolso_multiplier = buscar_reembolso_multiplier(conn, ano, mes, quinzena)

    logger.info(f"  Extrato acumulado: {len(extrato_acum)} usuarios | Snapshots controle: {len(snapshots_controle)} | Snapshots carga: {len(snapshots_carga)} | Somase: {len(somase)} CPFs | Controle: {len(controle)} CPFs")
    logger.info(f"  Multiplicador reembolso: {reembolso_multiplier}")

    linhas = []
    sem_extrato = 0
    sem_somase = 0
    calculados = 0

    for cpf, snap in controle.items():
        nome_up = snap["colaborador"].upper() if snap["colaborador"] else ""
        man = manuais.get(cpf, {})

        # --- Extrato acumulado até cutoff financeiro (dia 30 do mês anterior) ---
        ext_cum = extrato_acum.get(nome_up, {"carga": 0.0, "transferencia": 0.0, "tarifa": 0.0})
        if not extrato_acum.get(nome_up):
            sem_extrato += 1

        # --- SALDO CARTÃO CONTROLE (snapshot até dia 1 do mês atual) ---
        saldo_cartao_controle = snapshots_controle.get(nome_up, 0.0)

        # --- SALDO CARTÃO CARGA (snapshot até data de fechamento 11 ou 25) ---
        saldo_cartao_carga = snapshots_carga.get(nome_up, 0.0)

        # --- PRESTACAO DE CONTAS: somase acumulado ---
        prestacao_atual = somase.get(cpf, 0.0)
        if cpf not in somase:
            sem_somase += 1

        # --- SALDO PRESTACAO (PAINEL) ---
        # Cumulative: carga - transf - tarifa - prestacao (tudo até cutoff financeiro)
        saldo_prestacao = _r2(
            ext_cum["carga"]
            - ext_cum["transferencia"]
            - ext_cum["tarifa"]
            - somase.get(cpf, 0.0)
        )
        calculados += 1

        # --- SALDO FINAL e SALDO REEMBOLSAR ---
        # Saldo final uses CONTROLE saldo cartão (snapshot up to day 1)
        # Carga calculations use CARGA saldo cartão (snapshot up to closing date)
        saldo_final_painel = _r2(saldo_prestacao - saldo_cartao_controle)
        saldo_final      = _r2(max(saldo_final_painel, 0.0))
        saldo_reembolsar = _r2(max(-saldo_final_painel, 0.0))

        # --- Entradas manuais ---
        col_qz = 0.0
        if "col_qz" in man:
            col_qz = float(man["col_qz"] or 0)

        adiantamento = 0.0
        if "adiantamento" in man:
            adiantamento = float(man.get("adiantamento") or 0)

        # --- Formulas vivas da CARGA ---
        reembolso     = _r2(saldo_reembolsar * reembolso_multiplier)
        carga_parcial = _r2(col_qz - saldo_final - saldo_cartao_carga - adiantamento)
        carga_final   = _r2(max(carga_parcial, 0.0) + reembolso)

        # Regra de negocio: cartao com "Cadastro pendente" nao recebe carga
        status_c = (snap.get("status_cartao") or "").strip()
        if "pendente" in status_c.lower():
            carga_parcial = 0.0
            carga_final   = 0.0

        # Regra de negocio: REEMBOLSO so e pago na 1a QZ (reembolso mensal unico)
        if quinzena == 2:
            reembolso   = 0.0
            carga_final = _r2(max(carga_parcial, 0.0))

        linhas.append({
            "colaborador":    snap.get("colaborador", ""),
            "cpf":            cpf,
            "situacao":       snap.get("situacao", ""),
            "regional":       snap.get("regional", ""),
            "centro_custo":   snap.get("centro_custo", ""),
            "gestor":         snap.get("gestor", ""),
            "diretor":        snap.get("diretor", ""),
            "saldo_reembolsar": saldo_reembolsar,
            "saldo_final":    saldo_final,
            "col_qz":         col_qz,
            "saldo_cartao":   saldo_cartao_controle,
            "saldo_cartao_carga": saldo_cartao_carga,
            "adiantamento":   adiantamento,
            "carga_parcial":  carga_parcial,
            "reembolso":      reembolso,
            "carga_final":    carga_final,
            "obs":            man.get("obs", snap.get("obs", "")),
            "status_cartao":  snap.get("status_cartao", ""),
            # Campos auxiliares para diagnostico
            "_carga_extrato":       ext_cum["carga"],
            "_transferencia_extrato": ext_cum["transferencia"],
            "_tarifa_extrato":       ext_cum["tarifa"],
            "_prestacao_somase":     prestacao_atual,
            "_saldo_prestacao":      saldo_prestacao,
            "_saldo_cartao_controle": saldo_cartao_controle,
            "_saldo_cartao_carga":   saldo_cartao_carga,
            "_saldo_final_painel":   saldo_final_painel,
        })

    linhas.sort(key=lambda r: r["colaborador"] or "")
    logger.info(f"Linhas calculadas: {len(linhas)} | Sem extrato: {sem_extrato} | Sem somase: {sem_somase} | Calculados: {calculados}")

    if save:
        salvar_quinzena_snapshot(conn, ano, mes, quinzena, linhas)

    conn.close()
    return linhas


# =============================================================================
# Comparacao com planilha de referencia
# =============================================================================

def comparar_com_planilha(linhas: list[dict], ref_path: str, quinzena: int) -> None:
    """Compara resultado calculado com planilha de referencia e exibe divergencias."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl nao instalado")
        return

    logger.info(f"Comparando com planilha de referencia: {ref_path}")
    wb = openpyxl.load_workbook(ref_path, read_only=True, data_only=True)

    if quinzena == 1:
        ws = wb["Planilha1"]
        header_row = 6
        data_start = 7
        col_cpf    = 1   # idx 0-based
        col_sf     = 8
        col_qz     = 9
        col_sc     = 10
        col_cp     = 12
        col_reem   = 13
        col_cf     = 14
        col_sr     = 7
    else:
        sheet = [s for s in wb.sheetnames if "STATUS" not in s.upper()][0]
        ws = wb[sheet]
        header_row = 4
        data_start = 5
        col_cpf    = 2
        col_sf     = 9   # SALDO FINAL
        col_qz     = 10  # 2a QZ
        col_sc     = 11  # SALDO CARTAO
        col_cp     = 13  # CARGA PARCIAL
        col_reem   = 14  # REEMBOLSO (sempre 0 na 2QZ)
        col_cf     = 15  # Carga Final
        col_sr     = None  # 2QZ nao tem SALDO REEMBOLSAR como coluna separada

    def nc(raw) -> str:
        if raw is None:
            return ""
        s = str(raw).strip().replace(".", "").replace("-", "").replace("/", "").replace(" ", "")
        if "." in s:
            s = s.split(".")[0]
        return s.zfill(11)

    def nf(raw) -> float:
        try:
            return round(float(raw), 2) if raw is not None else 0.0
        except (ValueError, TypeError):
            return 0.0

    ref = {}
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        cpf = nc(row[col_cpf])
        if not cpf or cpf == "00000000000":
            continue
        ref[cpf] = {
            "saldo_reembolsar": nf(row[col_sr]) if col_sr is not None else 0.0,
            "saldo_final":      nf(row[col_sf]),
            "col_qz":           nf(row[col_qz]),
            "saldo_cartao":     nf(row[col_sc]),
            "carga_parcial":    nf(row[col_cp]),
            "reembolso":        nf(row[col_reem]),
            "carga_final":      nf(row[col_cf]),
        }
    wb.close()

    calc = {l["cpf"]: l for l in linhas}
    cpfs_ref = set(ref.keys())
    cpfs_calc = set(calc.keys())
    tol = 0.05

    logger.info(f"  Planilha ref: {len(cpfs_ref)} CPFs | Calculado: {len(cpfs_calc)} CPFs")
    logger.info(f"  Apenas na ref: {len(cpfs_ref - cpfs_calc)} | Apenas no calc: {len(cpfs_calc - cpfs_ref)}")

    # Na 2QZ nao ha coluna SALDO REEMBOLSAR separada — comparar apenas campos disponiveis
    if quinzena == 1:
        campos = ["saldo_reembolsar", "saldo_final", "saldo_cartao", "carga_parcial", "reembolso", "carga_final"]
    else:
        campos = ["saldo_final", "saldo_cartao", "carga_parcial", "carga_final"]
    divergencias = {c: [] for c in campos}
    total_comparados = 0

    for cpf in sorted(cpfs_ref & cpfs_calc):
        total_comparados += 1
        r = ref[cpf]
        c = calc[cpf]
        for campo in campos:
            v_ref  = r[campo]
            v_calc = c[campo]
            diff = abs(v_calc - v_ref)
            if diff > tol:
                divergencias[campo].append((cpf, c["colaborador"], v_ref, v_calc, v_calc - v_ref))

    print()
    print("=" * 72)
    print(f"  COMPARACAO COM PLANILHA DE REFERENCIA — {quinzena}ª QZ {ref_path}")
    print(f"  CPFs comparados: {total_comparados}")
    print("=" * 72)

    total_match = total_comparados
    for campo in campos:
        divs = divergencias[campo]
        match = total_comparados - len(divs)
        pct = match / total_comparados * 100 if total_comparados else 0
        status = "✅" if len(divs) == 0 else ("⚠️ " if len(divs) <= 5 else "❌")
        print(f"  {status} {campo:<22}: {match:>3}/{total_comparados}  ({pct:5.1f}%)")
        if divs:
            total_match = min(total_match, match)
            for cpf, nome, vr, vc, d in sorted(divs, key=lambda x: abs(x[4]), reverse=True)[:5]:
                print(f"       {cpf} {nome[:25]:<25} ref={vr:>10.2f}  calc={vc:>10.2f}  diff={d:>+10.2f}")

    print("=" * 72)

    # Totais de carga_final
    total_cf_ref  = sum(r["carga_final"] for r in ref.values())
    total_cf_calc = sum(c["carga_final"] for c in calc.values() if c["cpf"] in ref)
    print(f"  Total CARGA FINAL ref:  R$ {total_cf_ref:>12,.2f}")
    print(f"  Total CARGA FINAL calc: R$ {total_cf_calc:>12,.2f}")
    print(f"  Diferenca:              R$ {total_cf_calc - total_cf_ref:>+12,.2f}")
    print("=" * 72)


# =============================================================================
# Salvar Excel
# =============================================================================

COLUNAS_OUTPUT = [
    ("colaborador",    "COLABORADOR"),
    ("cpf",            "CPF"),
    ("situacao",       "SITUAÇÃO"),
    ("regional",       "REGIONAL"),
    ("centro_custo",   "CENTRO DE CUSTO"),
    ("gestor",         "GESTOR"),
    ("diretor",        "DIRETOR"),
    ("saldo_reembolsar", "SALDO REEMBOLSAR"),
    ("saldo_final",    "SALDO FINAL"),
    ("col_qz",         "QZ"),
    ("saldo_cartao",   "SALDO CARTAO (CONTROLE)"),
    ("saldo_cartao_carga", "SALDO CARTAO (CARGA)"),
    ("adiantamento",   "ADIANTAMENTO"),
    ("carga_parcial",  "CARGA PARCIAL"),
    ("reembolso",      "REEMBOLSO"),
    ("carga_final",    "CARGA FINAL"),
    ("obs",            "OBS"),
    ("status_cartao",  "STATUS DO CARTÃO"),
]


def salvar_excel(linhas: list[dict], output_path: Path, titulo: str = "") -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("openpyxl nao instalado")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo or "CARGA QZ"

    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="1F4E79")
    mfill = PatternFill("solid", fgColor="FFF2CC")  # amarelo = entrada manual
    MANUAIS = {"col_qz", "adiantamento", "obs"}

    for ci, (campo, label) in enumerate(COLUNAS_OUTPUT, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = hf
        cell.fill = hfill if campo not in MANUAIS else PatternFill("solid", fgColor="7F6000")
        cell.alignment = Alignment(horizontal="center")

    for ri, linha in enumerate(linhas, 2):
        for ci, (campo, _) in enumerate(COLUNAS_OUTPUT, 1):
            val = linha.get(campo)
            cell = ws.cell(row=ri, column=ci, value=val)
            if campo in MANUAIS:
                cell.fill = mfill

    for ci, (campo, label) in enumerate(COLUNAS_OUTPUT, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(len(label) + 4, 12)

    wb.save(output_path)
    logger.info(f"Salvo: {output_path}")


# =============================================================================
# CLI
# =============================================================================

def main() -> None:
    parser = argparse.ArgumentParser(description="Calcula Carga Quinzenal 100% via Neon")
    parser.add_argument("--ano",      type=int, required=True)
    parser.add_argument("--mes",      type=int, required=True)
    parser.add_argument("--quinzena", type=int, choices=[1, 2], required=True)
    parser.add_argument("--output",   type=Path, default=None,
                        help="Salvar resultado em .xlsx")
    parser.add_argument("--ref",      type=str, default=None,
                        help="Planilha de referencia para comparacao")
    parser.add_argument("--manuais",  type=Path, default=None,
                        help="JSON {cpf: {col_qz, adiantamento, obs}}")
    parser.add_argument("--save", action="store_true",
                        help="Salvar os valores calculados em quinzena_controle_snapshot")
    args = parser.parse_args()

    manuais: dict = {}
    if args.manuais and args.manuais.exists():
        import json
        with open(args.manuais, encoding="utf-8") as f:
            manuais = json.load(f)
        logger.info(f"Manuais: {len(manuais)} CPFs")

    inicio, fim, fechamento = get_periodo(args.ano, args.mes, args.quinzena)
    print(f"\n{'='*65}")
    print(f"  {args.quinzena}ª QZ {args.mes:02d}/{args.ano}")
    print(f"  Periodo extrato: {inicio} -> {fim}")
    print(f"  Data fechamento: {fechamento}")
    print(f"{'='*65}\n")

    linhas = calcular_quinzena(args.ano, args.mes, args.quinzena, manuais, save=args.save)

    print(f"\n✅ Calculadas {len(linhas)} linhas")

    if args.ref:
        ref_path = args.ref
        if not os.path.isabs(ref_path):
            ref_path = str(BASE / ref_path)
        comparar_com_planilha(linhas, ref_path, args.quinzena)

    if args.output:
        titulo = f"{args.quinzena}QZ {args.mes:02d}-{args.ano}"
        salvar_excel(linhas, args.output, titulo)
        print(f"\n✅ Salvo: {args.output}")


if __name__ == "__main__":
    main()
