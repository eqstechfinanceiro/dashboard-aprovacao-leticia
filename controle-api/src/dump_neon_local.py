#!/usr/bin/env python3
"""
dump_neon_local.py
------------------
Dumps all Neon data needed for quinzena calculation into a single local Excel file.
Run once, then use compare_local.py for fast offline comparisons.
"""
import os
import sys
from pathlib import Path

import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE = Path(__file__).parent.parent
load_dotenv(BASE / ".env")
NEON_URL = os.getenv("NEON_DATABASE_URL")

OUT = BASE / "data" / "neon_dump.xlsx"

def main():
    conn = psycopg2.connect(NEON_URL)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="1F4E79")

    def write_sheet(name, rows, headers):
        ws = wb.create_sheet(name)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center")
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
        print(f"  {name}: {len(rows)} rows")

    # 1. Cadastro
    cur.execute("""
        SELECT cpf, colaborador, situacao, status_cartao, regional,
               centro_custo, gestor, diretor
        FROM quinzena_cadastro
        ORDER BY colaborador
    """)
    rows = cur.fetchall()
    write_sheet("cadastro", [tuple(r.values()) for r in rows],
                ["cpf", "colaborador", "situacao", "status_cartao", "regional",
                 "centro_custo", "gestor", "diretor"])

    # 2. Extrato acumulado (all non-snapshot, grouped by user, up to June 30)
    cur.execute("""
        SELECT
            UPPER(usuario) AS usuario,
            COALESCE(SUM(valor) FILTER(WHERE tipo = 'Transferência' AND valor > 0), 0) AS carga,
            COALESCE(ABS(SUM(valor) FILTER(WHERE tipo = 'Transferência' AND valor < 0)), 0) AS transferencia,
            COALESCE(ABS(SUM(valor) FILTER(WHERE tipo = 'Taxa')), 0) AS tarifa
        FROM extrato_movimentacao
        WHERE is_snapshot = FALSE
          AND data <= '2026-06-30'
        GROUP BY UPPER(usuario)
        ORDER BY UPPER(usuario)
    """)
    rows = cur.fetchall()
    write_sheet("extrato_acumulado", [tuple(r.values()) for r in rows],
                ["usuario", "carga", "transferencia", "tarifa"])

    # 2b. Raw extrato with dates (for debugging / date filtering / dedup)
    cur.execute("""
        SELECT UPPER(usuario) AS usuario, data, hora, tipo, valor, codigo_transacao, descricao
        FROM extrato_movimentacao
        WHERE is_snapshot = FALSE
        ORDER BY data DESC, UPPER(usuario), hora
    """)
    rows = cur.fetchall()
    write_sheet("extrato_raw", [tuple(r.values()) for r in rows],
                ["usuario", "data", "hora", "tipo", "valor", "codigo_transacao", "descricao"])

    # 3. All snapshots (is_snapshot=TRUE) — latest per user per date
    cur.execute("""
        SELECT UPPER(usuario) AS usuario, data, valor
        FROM extrato_movimentacao
        WHERE is_snapshot = TRUE AND valor IS NOT NULL
        ORDER BY UPPER(usuario), data
    """)
    rows = cur.fetchall()
    write_sheet("snapshots_all", [tuple(r.values()) for r in rows],
                ["usuario", "data", "valor"])

    # 4. Somase snapshots
    cur.execute("""
        SELECT user_cpf, quinzena, total
        FROM somase_snapshots
        ORDER BY quinzena, user_cpf
    """)
    rows = cur.fetchall()
    write_sheet("somase_snapshots", [tuple(r.values()) for r in rows],
                ["user_cpf", "quinzena", "total"])

    # 5. Manual inputs
    cur.execute("""
        SELECT year, month, quinzena, cpf, col_1qz, adiantamento, obs
        FROM quinzena_manual_inputs
        ORDER BY year, month, quinzena, cpf
    """)
    rows = cur.fetchall()
    write_sheet("manual_inputs", [tuple(r.values()) for r in rows],
                ["year", "month", "quinzena", "cpf", "col_1qz", "adiantamento", "obs"])

    # 6. Quinzena config
    cur.execute("""
        SELECT year, month, quinzena, reembolso_multiplier
        FROM quinzena_config
        ORDER BY year, month, quinzena
    """)
    rows = cur.fetchall()
    write_sheet("quinzena_config", [tuple(r.values()) for r in rows],
                ["year", "month", "quinzena", "reembolso_multiplier"])

    # 7. Prestacao reports (full, for cutoff-filtered somase computation)
    cur.execute("""
        SELECT id, name, status, user_id, user_name, user_cpf,
               total_value, created_at, updated_at
        FROM prestacao_reports
        ORDER BY id
    """)
    rows = cur.fetchall()
    write_sheet("prestacao_reports", [tuple(r.values()) for r in rows],
                ["id", "name", "status", "user_id", "user_name", "user_cpf",
                 "total_value", "created_at", "updated_at"])

    # 7b. Prestacao expenses (full)
    cur.execute("""
        SELECT id, report_id, value
        FROM prestacao_expenses
        ORDER BY id
    """)
    rows = cur.fetchall()
    write_sheet("prestacao_expenses", [tuple(r.values()) for r in rows],
                ["id", "report_id", "value"])

    # 7c. Prestacao reports summary
    cur.execute("""
        SELECT status, COUNT(*) as cnt
        FROM prestacao_reports
        GROUP BY status
        ORDER BY COUNT(*) DESC
    """)
    rows = cur.fetchall()
    write_sheet("prestacao_summary", [tuple(r.values()) for r in rows],
                ["status", "count"])

    conn.close()

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"\n✅ Saved: {OUT}")


if __name__ == "__main__":
    main()
