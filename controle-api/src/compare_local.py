#!/usr/bin/env python3
"""
compare_local.py
----------------
Compara dados locais (neon_dump.xlsx) com planilha CONTROLE de referência.
100% offline — sem queries no Neon. Rápido.
"""
import argparse
import os
import sys
from collections import defaultdict
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl

BASE = Path(__file__).parent.parent
DUMP = BASE / "data" / "neon_dump.xlsx"


def _r2(v):
    return float(Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def nc(raw):
    if raw is None:
        return ""
    s = str(raw).strip().replace(".", "").replace("-", "").replace("/", "").replace(" ", "")
    return s.zfill(11)


def nf(raw):
    try:
        return round(float(raw), 2) if raw is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


def load_neon_dump():
    """Load all sheets from neon_dump.xlsx into dicts."""
    wb = openpyxl.load_workbook(DUMP, read_only=True, data_only=True)

    # Cadastro
    ws = wb["cadastro"]
    cadastro = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cpf = nc(row[0])
        if cpf:
            cadastro[cpf] = {
                "colaborador": str(row[1] or ""),
                "situacao": str(row[2] or ""),
                "status_cartao": str(row[3] or ""),
                "regional": str(row[4] or ""),
                "centro_custo": str(row[5] or ""),
                "gestor": str(row[6] or ""),
                "diretor": str(row[7] or ""),
            }

    # Extrato acumulado
    ws = wb["extrato_acumulado"]
    extrato = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = str(row[0] or "").upper()
        extrato[nome] = {
            "carga": nf(row[1]),
            "transferencia": nf(row[2]),
            "tarifa": nf(row[3]),
        }

    # Snapshots — build index: {user: [(date, valor), ...]}
    ws = wb["snapshots_all"]
    snapshots = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = str(row[0] or "").upper()
        d = row[1]
        if hasattr(d, "strftime"):
            d = d.date() if hasattr(d, "date") else d
        elif isinstance(d, str):
            d = date.fromisoformat(d)
        snapshots[nome].append((d, nf(row[2])))
    # Sort by date for each user
    for nome in snapshots:
        snapshots[nome].sort(key=lambda x: x[0])

    # Somase
    ws = wb["somase_snapshots"]
    somase = defaultdict(dict)  # {quinzena_id: {cpf: total}}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cpf = nc(row[0])
        qid = str(row[1])
        somase[qid][cpf] = nf(row[2])

    # Manual inputs
    ws = wb["manual_inputs"]
    manuais = defaultdict(dict)  # {(year,month,qz): {cpf: {col_qz, adiantamento}}}
    for row in ws.iter_rows(min_row=2, values_only=True):
        y, m, q = int(row[0]), int(row[1]), int(row[2])
        cpf = nc(row[3])
        manuais[(y, m, q)][cpf] = {
            "col_qz": nf(row[4]),
            "adiantamento": nf(row[5]),
            "obs": str(row[6] or ""),
        }

    # Quinzena config
    ws = wb["quinzena_config"]
    config = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        y, m, q = int(row[0]), int(row[1]), int(row[2])
        config[(y, m, q)] = nf(row[3])

    wb.close()

    return {
        "cadastro": cadastro,
        "extrato": extrato,
        "snapshots": snapshots,
        "somase": somase,
        "manuais": manuais,
        "config": config,
    }


def get_saldo_cartao(snapshots, nome_upper, cutoff_date):
    """Get latest snapshot for user up to cutoff_date."""
    snaps = snapshots.get(nome_upper, [])
    best = 0.0
    for d, v in snaps:
        if d <= cutoff_date:
            best = v
        else:
            break
    return best


def get_periodo(ano, mes, quinzena):
    mes_ant = mes - 1 if mes > 1 else 12
    ano_ant = ano if mes > 1 else ano - 1
    if quinzena == 1:
        fechamento = date(ano, mes, 11)
    else:
        fechamento = date(ano, mes, 25)
    cutoff_fin = date(ano_ant, mes_ant, 30)
    saldo_controle = date(ano, mes, 1)
    return fechamento, cutoff_fin, saldo_controle


def calcular_local(data, ano, mes, quinzena):
    """Calculate quinzena from local dump data."""
    cadastro = data["cadastro"]
    extrato = data["extrato"]
    snapshots = data["snapshots"]
    somase = data["somase"]
    manuais = data["manuais"]
    config = data["config"]

    fechamento, cutoff_fin, saldo_controle_date = get_periodo(ano, mes, quinzena)
    quinzena_id = f"{ano}-{mes:02d}-{quinzena}"
    multiplier = config.get((ano, mes, quinzena), 0.5)
    somase_q = somase.get(quinzena_id, {})
    manuais_q = manuais.get((ano, mes, quinzena), {})

    linhas = []
    for cpf, cad in cadastro.items():
        nome_up = cad["colaborador"].upper()
        ext = extrato.get(nome_up, {"carga": 0.0, "transferencia": 0.0, "tarifa": 0.0})
        sc_controle = get_saldo_cartao(snapshots, nome_up, saldo_controle_date)
        sc_carga = get_saldo_cartao(snapshots, nome_up, fechamento)
        prest = somase_q.get(cpf, 0.0)

        saldo_prestacao = _r2(ext["carga"] - ext["transferencia"] - ext["tarifa"] - prest)
        saldo_final_painel = _r2(saldo_prestacao - sc_controle)
        saldo_final = _r2(max(saldo_final_painel, 0.0))
        saldo_reembolsar = _r2(max(-saldo_final_painel, 0.0))

        man = manuais_q.get(cpf, {})
        col_qz = man.get("col_qz", 0.0)
        adiantamento = man.get("adiantamento", 0.0)

        reembolso = _r2(saldo_reembolsar * multiplier)
        if quinzena == 2:
            reembolso = 0.0

        carga_parcial = _r2(col_qz - saldo_final - sc_carga - adiantamento)
        carga_final = _r2(max(carga_parcial, 0.0) + reembolso)

        status_c = cad.get("status_cartao", "").strip()
        if "pendente" in status_c.lower():
            carga_parcial = 0.0
            carga_final = 0.0

        linhas.append({
            "cpf": cpf,
            "colaborador": cad["colaborador"],
            "carga": ext["carga"],
            "transferencia": ext["transferencia"],
            "tarifa": ext["tarifa"],
            "prestacao": prest,
            "saldo_prestacao": saldo_prestacao,
            "saldo_cartao_controle": sc_controle,
            "saldo_cartao_carga": sc_carga,
            "saldo_final": saldo_final,
            "saldo_reembolsar": saldo_reembolsar,
            "col_qz": col_qz,
            "adiantamento": adiantamento,
            "reembolso": reembolso,
            "carga_parcial": carga_parcial,
            "carga_final": carga_final,
        })

    return linhas


def main():
    parser = argparse.ArgumentParser(description="Compare local neon dump with CONTROLE sheet")
    parser.add_argument("--ref", type=str, required=True, help="Path to CONTROLE xlsx")
    parser.add_argument("--ano", type=int, required=True)
    parser.add_argument("--mes", type=int, required=True)
    parser.add_argument("--quinzena", type=int, choices=[1, 2], required=True)
    args = parser.parse_args()

    ref_path = args.ref
    if not os.path.isabs(ref_path):
        ref_path = str(BASE / ref_path)

    print("Loading neon dump...")
    data = load_neon_dump()
    print(f"  cadastro: {len(data['cadastro'])} | extrato: {len(data['extrato'])} | snapshots: {len(data['snapshots'])} | somase: {sum(len(v) for v in data['somase'].values())}")

    print("Loading reference sheet...")
    wb = openpyxl.load_workbook(ref_path, read_only=True, data_only=True)
    ws = wb["PAINEL"]

    # Header at row 11
    header = list(ws.iter_rows(min_row=11, max_row=11, values_only=True))[0]
    print(f"  Header: {[str(h)[:20] if h else None for h in header[:20]]}")

    ref = {}
    for row in ws.iter_rows(min_row=12, values_only=True):
        cpf = nc(row[2])
        if not cpf or cpf == "00000000000":
            continue
        ref[cpf] = {
            "colaborador": str(row[1] or "")[:30],
            "carga": nf(row[13]),
            "transferencia": nf(row[14]),
            "tarifa": nf(row[15]),
            "prestacao": nf(row[16]),
            "saldo_prestacao": nf(row[17]),
            "saldo_cartao": nf(row[18]),
            "saldo_final": nf(row[19]),
        }
    wb.close()
    print(f"  Reference: {len(ref)} CPFs")

    print("Calculating locally...")
    linhas = calcular_local(data, args.ano, args.mes, args.quinzena)
    calc = {l["cpf"]: l for l in linhas}
    print(f"  Calculated: {len(calc)} CPFs")

    # Compare
    cpfs_ref = set(ref.keys())
    cpfs_calc = set(calc.keys())
    tol = 0.05
    campos = ["carga", "transferencia", "tarifa", "prestacao", "saldo_prestacao", "saldo_cartao", "saldo_final"]
    divergencias = {c: [] for c in campos}
    total_comparados = 0

    for cpf in sorted(cpfs_ref & cpfs_calc):
        total_comparados += 1
        r = ref[cpf]
        c = calc[cpf]
        for campo in campos:
            v_ref = r[campo]
            v_calc = c.get(campo, 0.0)
            diff = abs(v_calc - v_ref)
            if diff > tol:
                divergencias[campo].append((cpf, c["colaborador"][:25], v_ref, v_calc, v_calc - v_ref))

    print()
    print("=" * 80)
    print(f"  COMPARAÇÃO — {args.quinzena}ª QZ {args.mes:02d}/{args.ano}")
    print(f"  Planilha: {ref_path}")
    print(f"  CPFs na planilha: {len(cpfs_ref)} | CPFs calculados: {len(cpfs_calc)} | Comparados: {total_comparados}")
    print(f"  Apenas na planilha: {len(cpfs_ref - cpfs_calc)} | Apenas no cálculo: {len(cpfs_calc - cpfs_ref)}")
    print("=" * 80)

    for campo in campos:
        divs = divergencias[campo]
        match = total_comparados - len(divs)
        pct = match / total_comparados * 100 if total_comparados else 0
        status = "✅" if len(divs) == 0 else ("⚠️ " if len(divs) <= 10 else "❌")
        print(f"  {status} {campo:<22}: {match:>4}/{total_comparados}  ({pct:5.1f}%)  divergências: {len(divs)}")
        if divs:
            for cpf, nome, vr, vc, d in sorted(divs, key=lambda x: abs(x[4]), reverse=True)[:10]:
                print(f"       {cpf} {nome:<25} ref={vr:>12.2f}  calc={vc:>12.2f}  diff={d:>+12.2f}")

    print("=" * 80)

    # Totals
    total_carga_ref = sum(r["carga"] for r in ref.values())
    total_carga_calc = sum(c["carga"] for c in calc.values() if c["cpf"] in ref)
    total_prest_ref = sum(r["prestacao"] for r in ref.values())
    total_prest_calc = sum(c["prestacao"] for c in calc.values() if c["cpf"] in ref)
    total_sp_ref = sum(r["saldo_prestacao"] for r in ref.values())
    total_sp_calc = sum(c["saldo_prestacao"] for c in calc.values() if c["cpf"] in ref)
    total_sf_ref = sum(r["saldo_final"] for r in ref.values())
    total_sf_calc = sum(c["saldo_final"] for c in calc.values() if c["cpf"] in ref)

    print(f"  Total CARGA          ref: R$ {total_carga_ref:>14,.2f}  calc: R$ {total_carga_calc:>14,.2f}  diff: R$ {total_carga_calc - total_carga_ref:>+14,.2f}")
    print(f"  Total PRESTAÇÃO      ref: R$ {total_prest_ref:>14,.2f}  calc: R$ {total_prest_calc:>14,.2f}  diff: R$ {total_prest_calc - total_prest_ref:>+14,.2f}")
    print(f"  Total SALDO PREST    ref: R$ {total_sp_ref:>14,.2f}  calc: R$ {total_sp_calc:>14,.2f}  diff: R$ {total_sp_calc - total_sp_ref:>+14,.2f}")
    print(f"  Total SALDO FINAL    ref: R$ {total_sf_ref:>14,.2f}  calc: R$ {total_sf_calc:>14,.2f}  diff: R$ {total_sf_calc - total_sf_ref:>+14,.2f}")
    print("=" * 80)


if __name__ == "__main__":
    main()
