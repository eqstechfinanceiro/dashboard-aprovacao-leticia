"""
Gerador automático da planilha de Carga Quinzenal (CARGA QZ).

Gera todas as colunas automatizáveis a partir do banco SQLite.
As 3 colunas manuais (col_1ª_qz, adiantamento, obs) são aceitas como
entrada opcional — se não fornecidas, ficam vazias no output.

Uso:
    python src/gerar_carga_qz.py
    python src/gerar_carga_qz.py --output data/carga_qz_gerada.xlsx
    python src/gerar_carga_qz.py --manuais data/manuais.json --output data/carga.xlsx

Formato do arquivo manuais.json:
    {
        "01696239478": {"col_1qz": 1750, "adiantamento": 0, "obs": ""},
        "07024923610": {"col_1qz": 700,  "adiantamento": 0, "obs": ""}
    }
"""

import argparse
import json
import logging
import sqlite3
from pathlib import Path
from typing import Optional

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

DB_PATH = Path(__file__).parent.parent / "data" / "spreadsheets.db"

COLUNAS_ORDEM = [
    "colaborador", "cpf", "situação", "regional", "centro_de_custo",
    "gestor", "diretor", "saldo_reembolsar", "saldo_final",
    "col_1ª_qz", "saldo_cartao", "adiantamento",
    "carga_parcial", "reembolso", "carga_final",
    "obs", "status_do_cartão",
]


def _sf(v) -> float:
    try:
        return float(v) if v is not None and str(v).strip() not in ("", "None") else 0.0
    except (ValueError, TypeError):
        return 0.0


def _load_painel(conn: sqlite3.Connection) -> dict[str, dict]:
    rows = conn.execute("SELECT * FROM controle_painel").fetchall()
    return {r["cpf"]: dict(r) for r in rows}


def _load_saldo_cartao(conn: sqlite3.Connection) -> dict[str, float]:
    """Retorna o saldo_cartao mais recente por CPF (maior data)."""
    rows = conn.execute(
        "SELECT cpf, valor, data FROM controle_saldo_cartao_resumo"
    ).fetchall()
    best: dict[str, tuple[float, float]] = {}  # cpf -> (data, valor)
    for r in rows:
        cpf = r["cpf"]
        data_val = _sf(r["data"])
        valor = _sf(r["valor"])
        if cpf not in best or data_val > best[cpf][0]:
            best[cpf] = (data_val, valor)
    return {cpf: v for cpf, (_, v) in best.items()}


def gerar_carga(
    manuais: Optional[dict[str, dict]] = None,
    db_path: Path = DB_PATH,
) -> list[dict]:
    """
    Gera as linhas da carga_qz calculando todas as colunas automáticas.

    Args:
        manuais: dict cpf -> {col_1qz, adiantamento, obs}. Opcional.
        db_path: caminho para o banco SQLite.

    Returns:
        Lista de dicts com todas as colunas na ordem correta.
    """
    manuais = manuais or {}
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    logger.info("Carregando controle_painel...")
    painel = _load_painel(conn)

    logger.info("Carregando controle_saldo_cartao_resumo...")
    saldo_cartao_map = _load_saldo_cartao(conn)

    conn.close()

    logger.info(f"Gerando carga para {len(painel)} colaboradores...")

    linhas = []
    for cpf, p in painel.items():
        man = manuais.get(cpf, {})
        col_1qz = _sf(man.get("col_1qz"))
        adiantamento = _sf(man.get("adiantamento"))
        obs = man.get("obs", None)

        # Saldos do painel
        painel_sf = _sf(p.get("saldo_final"))
        saldo_final = max(painel_sf, 0.0)
        saldo_reembolsar = abs(painel_sf) if painel_sf < 0 else 0.0

        # Saldo cartão
        saldo_cartao = saldo_cartao_map.get(cpf, 0.0)

        # Cálculos
        reembolso = round(saldo_reembolsar / 2, 2)
        carga_parcial = round(col_1qz - saldo_final - saldo_cartao - adiantamento, 2)
        carga_final = round(max(carga_parcial + reembolso, 0.0), 2)

        linhas.append({
            "colaborador":     p.get("colaborador", ""),
            "cpf":             cpf,
            "situação":        p.get("situação", ""),
            "regional":        p.get("regional", ""),
            "centro_de_custo": p.get("centro_de_custo", ""),
            "gestor":          p.get("gestor", ""),
            "diretor":         p.get("diretor", ""),
            "saldo_reembolsar": saldo_reembolsar,
            "saldo_final":     saldo_final,
            "col_1ª_qz":       col_1qz if col_1qz else None,
            "saldo_cartao":    saldo_cartao,
            "adiantamento":    adiantamento if adiantamento else None,
            "carga_parcial":   carga_parcial,
            "reembolso":       reembolso,
            "carga_final":     carga_final,
            "obs":             obs,
            "status_do_cartão": p.get("status_do_cartão", ""),
        })

    linhas.sort(key=lambda r: r["colaborador"] or "")
    logger.info(f"Geradas {len(linhas)} linhas.")
    return linhas


def salvar_excel(linhas: list[dict], output_path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("openpyxl não instalado. Instale com: pip install openpyxl")
        raise

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CARGA QZ"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    manual_fill = PatternFill("solid", fgColor="FFF2CC")  # amarelo para manuais

    MANUAIS = {"col_1ª_qz", "adiantamento", "obs"}

    # Cabeçalho
    for col_idx, col_name in enumerate(COLUNAS_ORDEM, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper().replace("_", " "))
        cell.font = header_font
        cell.fill = header_fill if col_name not in MANUAIS else PatternFill("solid", fgColor="7F6000")
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for row_idx, linha in enumerate(linhas, start=2):
        for col_idx, col_name in enumerate(COLUNAS_ORDEM, start=1):
            val = linha.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_name in MANUAIS:
                cell.fill = manual_fill

    # Larguras automáticas
    for col_idx, col_name in enumerate(COLUNAS_ORDEM, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = max(len(col_name) + 4, 14)

    wb.save(output_path)
    logger.info(f"Salvo em: {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera planilha de Carga QZ")
    parser.add_argument(
        "--manuais", type=Path, default=None,
        help="JSON com col_1qz/adiantamento/obs por CPF"
    )
    parser.add_argument(
        "--output", type=Path,
        default=Path(__file__).parent.parent / "data" / "carga_qz_gerada.xlsx",
        help="Caminho de saída (.xlsx)"
    )
    parser.add_argument(
        "--db", type=Path, default=DB_PATH,
        help="Caminho do banco SQLite"
    )
    args = parser.parse_args()

    manuais: dict = {}
    if args.manuais and args.manuais.exists():
        with open(args.manuais, encoding="utf-8") as f:
            manuais = json.load(f)
        logger.info(f"Manuais carregados: {len(manuais)} CPFs")

    linhas = gerar_carga(manuais=manuais, db_path=args.db)
    salvar_excel(linhas, args.output)
    print(f"\n✅ Carga QZ gerada: {len(linhas)} colaboradores → {args.output}")


if __name__ == "__main__":
    main()
