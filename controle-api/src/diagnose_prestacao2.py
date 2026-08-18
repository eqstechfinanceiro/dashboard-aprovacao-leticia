#!/usr/bin/env python3
"""Diagnose prestacao gap: check APROVADO reports with/without CPF, missing expenses."""
from pathlib import Path
import openpyxl

BASE = Path(__file__).parent.parent

def nc(raw):
    if raw is None: return ""
    s = str(raw).strip().replace(".","").replace("-","").replace("/","").replace(" ","")
    return s.zfill(11)

def nf(raw):
    try: return float(raw) if raw is not None else 0.0
    except: return 0.0

wb = openpyxl.load_workbook(BASE / "data" / "neon_dump.xlsx", read_only=True, data_only=True)

# Check prestacao_reports
ws = wb["prestacao_reports"]
statuses = {}
aprovado_with_cpf = 0
aprovado_without_cpf = 0
aprovado_fatura = 0
aprovado_valid = 0
valid_report_ids = set()

for row in ws.iter_rows(min_row=2, values_only=True):
    rid = row[0]
    name = str(row[1] or "").upper()
    status = str(row[2] or "").upper()
    user_cpf = nc(row[5]) if len(row) > 5 and row[5] else ""
    
    statuses[status] = statuses.get(status, 0) + 1
    
    if status == "APROVADO":
        is_fatura = name.startswith("FATURA") or name.startswith("CARTAO") or name.startswith("CARTÃO")
        if is_fatura:
            aprovado_fatura += 1
        elif user_cpf:
            aprovado_with_cpf += 1
            valid_report_ids.add(rid)
        else:
            aprovado_without_cpf += 1

print("=== Prestacao Reports ===")
print(f"Statuses: {statuses}")
print(f"APROVADO with CPF (valid): {aprovado_with_cpf}")
print(f"APROVADO without CPF: {aprovado_without_cpf}")
print(f"APROVADO FATURA/CARTAO: {aprovado_fatura}")
print(f"Valid report IDs: {len(valid_report_ids)}")

# Check expenses
ws_e = wb["prestacao_expenses"]
total_expenses = 0
matched_expenses = 0
unmatched_expenses = 0
expenses_by_cpf = {}

for row in ws_e.iter_rows(min_row=2, values_only=True):
    rid = row[1] if len(row) > 1 else None
    value = nf(row[2]) if len(row) > 2 else 0
    total_expenses += value
    if rid in valid_report_ids:
        matched_expenses += value
    else:
        unmatched_expenses += value

print(f"\n=== Expenses ===")
print(f"Total expenses value: R$ {total_expenses:,.2f}")
print(f"Matched (valid APROVADO): R$ {matched_expenses:,.2f}")
print(f"Unmatched: R$ {unmatched_expenses:,.2f}")

# Check ref prestacao total
wb_ref = openpyxl.load_workbook(BASE / "data" / "ref_dump.xlsx", read_only=True, data_only=True)
ws_ref = wb_ref["painel"]
ref_total = 0
ref_cpf_count = 0
for row in ws_ref.iter_rows(min_row=2, values_only=True):
    cpf = nc(row[0])
    prestacao = nf(row[5])
    if prestacao > 0:
        ref_total += prestacao
        ref_cpf_count += 1
wb_ref.close()

print(f"\n=== Reference ===")
print(f"Ref prestacao total: R$ {ref_total:,.2f} ({ref_cpf_count} CPFs with prestacao > 0)")
print(f"Neon prestacao total: R$ {matched_expenses:,.2f}")
print(f"Gap: R$ {ref_total - matched_expenses:,.2f}")

# Check which ref CPFs are missing from neon
# Rebuild somase_by_cpf
reports_map = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    rid = row[0]
    name = str(row[1] or "").upper()
    status = str(row[2] or "").upper()
    user_cpf = nc(row[5]) if len(row) > 5 and row[5] else ""
    if status == "APROVADO" and user_cpf:
        if not (name.startswith("FATURA") or name.startswith("CARTAO") or name.startswith("CARTÃO")):
            reports_map[rid] = user_cpf

neon_prest_by_cpf = {}
for row in ws_e.iter_rows(min_row=2, values_only=True):
    rid = row[1] if len(row) > 1 else None
    value = nf(row[2]) if len(row) > 2 else 0
    if rid in reports_map:
        cpf = reports_map[rid]
        neon_prest_by_cpf[cpf] = neon_prest_by_cpf.get(cpf, 0) + value

wb.close()

# Compare
wb_ref = openpyxl.load_workbook(BASE / "data" / "ref_dump.xlsx", read_only=True, data_only=True)
ws_ref = wb_ref["painel"]
missing_cpf = []
for row in ws_ref.iter_rows(min_row=2, values_only=True):
    cpf = nc(row[0])
    nome = str(row[1] or "")[:25]
    ref_prest = nf(row[5])
    neon_prest = neon_prest_by_cpf.get(cpf, 0)
    if abs(ref_prest - neon_prest) > 0.50:
        missing_cpf.append((cpf, nome, ref_prest, neon_prest, abs(ref_prest - neon_prest)))
wb.close()

print(f"\n=== Top prestacao divergences ({len(missing_cpf)} total) ===")
for cpf, nome, rv, nv, diff in sorted(missing_cpf, key=lambda x: x[4], reverse=True)[:15]:
    print(f"  {cpf} {nome:<25} ref={rv:>12.2f} neon={nv:>12.2f} diff={diff:>+10.2f}")

# Check if ref has CPFs that are not in neon at all
missing_in_neon = [(cpf, nome, rv) for cpf, nome, rv, nv, diff in missing_cpf if nv == 0]
print(f"\n=== Ref CPFs with ZERO in neon: {len(missing_in_neon)} ===")
for cpf, nome, rv in sorted(missing_in_neon, key=lambda x: x[2], reverse=True)[:10]:
    print(f"  {cpf} {nome:<25} ref={rv:>12.2f}")
