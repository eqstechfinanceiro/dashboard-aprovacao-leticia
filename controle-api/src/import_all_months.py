#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
import_all_months.py
====================
Comprehensive import script for all quinzenas Jan-Jun 2026.

Uses the universal CONTROLE (JULHO 2026.xlsb) for cadastro data,
and each month's CARGA sheet for per-quinzena financial data.

Sheet structure variations are handled via per-month column mappings
discovered by inspecting each sheet individually.
"""
import os
import sys
import argparse
from decimal import Decimal, InvalidOperation
from collections import defaultdict

import openpyxl
import pyxlsb
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

# Load .env from controle-api directory
# Script is at: controle-api/src/import_all_months.py
# ROOT = controle-api/src/, PROJECT = controle-api/, DASHBOARD = dashboard-test/
ROOT = os.path.dirname(os.path.abspath(__file__))
PROJECT = os.path.dirname(ROOT)          # controle-api/
DASHBOARD = os.path.dirname(PROJECT)      # dashboard-test/
load_dotenv(os.path.join(PROJECT, ".env"))

NEON_DATABASE_URL = os.getenv("NEON_DATABASE_URL")

# Data paths — CARGA sheets are in dashboard-test/data/
DATA_DIR = os.path.join(DASHBOARD, "data")
CONTROLE_PATH = os.path.join(DATA_DIR, "CONTROLE - VEXPENSES - JULHO 2026.xlsb")

# =============================================================================
# CONTROLE PAINEL column mapping (0-indexed)
# Header row 11 (0-indexed: 10), data starts row 12 (0-indexed: 11)
# =============================================================================
PAINEL_HEADER_ROW = 11  # 1-indexed
PAINEL_DATA_START = 12  # 1-indexed
PAINEL_COLS = {
    "colaborador":   1,   # COLABORADOR
    "cpf":           2,   # CPF
    "situacao":      4,   # SITUAÇÃO
    "status_cartao": 5,   # STATUS DO CARTÃO
    "regional":      8,   # REGIONAL
    "centro_custo":  9,   # CENTRO DE CUSTO
    "gestor":       10,   # GESTOR
    "diretor":      11,   # DIRETOR
    "saldo_prestacao":     17,  # SALDO PRESTAÇÃO
    "saldo_cartao_painel": 18,  # (-) SALDO CARTAO
    "saldo_final":         19,  # SALDO FINAL
}

# =============================================================================
# CARGA sheet column mappings per month/quinzena (0-indexed)
# Discovered by inspecting each sheet's header row
# =============================================================================
CARGA_CONFIG = {
    # (month, quinzena): {sheet_name, header_row, data_start, cols}
    # reembolso_col: column index for the sheet's REEMBOLSO column (1QZ only)
    # For 2QZ, reembolso is always 0, so reembolso_col is None
    (1, 1): {
        "sheet": "1 QZ VEXPENSES 01_2026",
        "header": 6, "data": 7,
        "cols": {
            "cpf": 2, "colaborador": 1, "status_colab": 3, "status_cartao": 17,
            "saldo_reembolsar": 8, "saldo_final_carga": 9,
            "col_qz": 10, "saldo_cartao_carga": 11, "adiantamento": 12,
            "reembolso": 14,  # REEMBOLSO column
        },
    },
    (1, 2): {
        "sheet": "2 QZ VEXPENSES 01_2026",
        "header": 6, "data": 7,
        "cols": {
            "cpf": 2, "colaborador": 1, "status_colab": 3, "status_cartao": 22,
            "saldo_reembolsar": 8, "saldo_final_carga": 12,  # SALDO PENDENTE FINAL
            "col_qz": 14, "saldo_cartao_carga": 15, "adiantamento": 16,
            "reembolso": None,  # 2QZ - no reembolso
        },
    },
    (2, 1): {
        "sheet": "1 QZN FEV 2026",
        "header": 6, "data": 7,
        "cols": {
            "cpf": 2, "colaborador": 1, "status_colab": 3, "status_cartao": 16,
            "saldo_reembolsar": 8, "saldo_final_carga": 9,  # SALDO PENDENTE
            "col_qz": 10, "saldo_cartao_carga": 11, "adiantamento": 12,
            "reembolso": 14,  # REEMBOLSO column
        },
    },
    (2, 2): {
        "sheet": "2 QZ VEXPENSES 02_2026",
        "header": 6, "data": 7,
        "cols": {
            "cpf": 2, "colaborador": 1, "status_colab": 3, "status_cartao": 18,
            "saldo_reembolsar": 8, "saldo_final_carga": 11,  # SALDO PENDENTE FINAL
            "col_qz": 12, "saldo_cartao_carga": 13, "adiantamento": 14,
            "reembolso": None,  # 2QZ - no reembolso
        },
    },
    (3, 1): {
        "sheet": "QUINZENA MARÇO",
        "header": 6, "data": 7,
        "cols": {
            "cpf": 2, "colaborador": 1, "status_colab": 3, "status_cartao": 20,
            "saldo_reembolsar": 8, "saldo_final_carga": 9,  # SALDO PENDENTE COM CX
            "col_qz": 12, "saldo_cartao_carga": 13, "adiantamento": 14,
            "reembolso": 17,  # REEMBOLSO column
        },
    },
    (4, 1): {
        "sheet": "1 QZ VEXPENSES 04_2026",
        "header": 6, "data": 7,
        "cols": {
            "cpf": 3, "colaborador": 2, "status_colab": 4, "status_cartao": 17,
            "saldo_reembolsar": 9, "saldo_final_carga": 10,
            "col_qz": 11, "saldo_cartao_carga": 12, "adiantamento": 13,
            "reembolso": 15,  # REEMBOLSO column
        },
    },
    (5, 1): {
        "sheet": "Planilha1",
        "header": 6, "data": 7,
        "cols": {
            "cpf": 1, "colaborador": 0, "status_colab": 2, "status_cartao": 16,
            "saldo_reembolsar": 7, "saldo_final_carga": 8,
            "col_qz": 9, "saldo_cartao_carga": 10, "adiantamento": 11,
            "reembolso": 13,  # REEMBOLSO column
        },
    },
    (5, 2): {
        "sheet": "2 QZ DE MAIO 26",
        "header": 4, "data": 5,
        "cols": {
            "cpf": 2, "colaborador": 1, "status_colab": 3, "status_cartao": 17,
            "saldo_reembolsar": 7, "saldo_final_carga": 9,
            "col_qz": 10, "saldo_cartao_carga": 11, "adiantamento": 12,
            "reembolso": None,  # 2QZ - no reembolso
        },
    },
    (6, 1): {
        "sheet": "1 QZ JUNHO",
        "header": 6, "data": 7,
        "cols": {
            "cpf": 1, "colaborador": 0, "status_colab": 2, "status_cartao": 16,
            "saldo_reembolsar": 7, "saldo_final_carga": 8,
            "col_qz": 9, "saldo_cartao_carga": 10, "adiantamento": 11,
            "reembolso": 13,  # REEMBOLSO column
        },
    },
    (6, 2): {
        "sheet": "2 QZ JUNHO",
        "header": 6, "data": 7,
        "cols": {
            "cpf": 1, "colaborador": 0, "status_colab": 2, "status_cartao": 15,
            "saldo_reembolsar": None,  # Not present in simplified 2QZ sheet
            "saldo_final_carga": 7,
            "col_qz": 8, "saldo_cartao_carga": 9, "adiantamento": 10,
            "reembolso": None,  # 2QZ - no reembolso
        },
    },
}

# Reembolso multipliers per month (from quinzena_config)
REEMBOLSO_MULTIPLIERS = {
    1: Decimal("0.2"),   # January
    2: Decimal("0.5"),   # February
    3: Decimal("0.5"),   # March
    4: Decimal("0.5"),   # April
    5: Decimal("0.5"),   # May
    6: Decimal("0.6"),   # June
}

# CARGA file paths per month
CARGA_FILES = {
    (1, 1): os.path.join(DATA_DIR, "01 - JANEIRO", "1QZ JANEIRO 2026 - VEXPENSES.xlsx"),
    (1, 2): os.path.join(DATA_DIR, "01 - JANEIRO", "2QZ JANEIRO 2026 - VEXPENSES.xlsx"),
    (2, 1): os.path.join(DATA_DIR, "02 - FEVEREIRO", "1 QZN FEVEREIRO VEXPENSES 2026.xlsx"),
    (2, 2): os.path.join(DATA_DIR, "02 - FEVEREIRO", "2QZ FEVEREIRO 2026 - VEXPENSES EQS.xlsx"),
    (3, 1): os.path.join(DATA_DIR, "03 - MARÇO", "1 QZ MARÇO VEXPENSES 2026 (5).xlsx"),
    (4, 1): os.path.join(DATA_DIR, "04 - ABRIL", "1QZ ABRIL 2026 - VEXPENSES.xlsx"),
    (5, 1): os.path.join(DATA_DIR, "05 - MAIO", "CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx"),
    (5, 2): os.path.join(DATA_DIR, "05 - MAIO", "CARGA 2 QZ MAIO 26 VEXPENSES EQS.xlsx"),
    (6, 1): os.path.join(DATA_DIR, "06 - JUNHO", "CARGA 1 QZ JUNHO 26 VEXPENSES EQS.xlsx"),
    (6, 2): os.path.join(DATA_DIR, "06 - JUNHO", "CARGA 2 QZ JUNHO 26 VEXPENSES EQS.xlsx"),
}

# =============================================================================
# Helpers
# =============================================================================

def normalize_cpf(raw):
    if raw is None:
        return None
    s = str(raw).strip().replace(".", "").replace("-", "").replace("/", "").replace(" ", "")
    # Handle float-formatted CPFs (e.g., 2873491019.0)
    if "." in s:
        try:
            s = str(int(float(s)))
        except (ValueError, TypeError):
            pass
    return s.zfill(11) if s.isdigit() and len(s) <= 11 else None


def safe_decimal(raw):
    if raw is None:
        return None
    try:
        return Decimal(str(raw)).quantize(Decimal("0.0001"))
    except (InvalidOperation, ValueError, TypeError):
        return None


def get_cell(row_values, col_idx):
    if col_idx is None:
        return None
    return row_values[col_idx] if col_idx < len(row_values) else None


# =============================================================================
# Readers
# =============================================================================

def read_controle_painel(path):
    """Read CONTROLE PAINEL from .xlsb file. Returns {cpf: cadastro_record}."""
    records = {}
    with pyxlsb.open_workbook(path) as wb:
        with wb.get_sheet("PAINEL") as ws:
            for i, row in enumerate(ws.rows()):
                row_num = i + 1  # 1-indexed
                if row_num < PAINEL_DATA_START:
                    continue
                vals = [c.v for c in row]
                cpf = normalize_cpf(get_cell(vals, PAINEL_COLS["cpf"]))
                if not cpf:
                    continue
                records[cpf] = {
                    "colaborador":   str(get_cell(vals, PAINEL_COLS["colaborador"]) or "").strip(),
                    "situacao":      str(get_cell(vals, PAINEL_COLS["situacao"]) or "").strip(),
                    "status_cartao": str(get_cell(vals, PAINEL_COLS["status_cartao"]) or "").strip(),
                    "regional":      str(get_cell(vals, PAINEL_COLS["regional"]) or "").strip(),
                    "centro_custo":  str(get_cell(vals, PAINEL_COLS["centro_custo"]) or "").strip(),
                    "gestor":        str(get_cell(vals, PAINEL_COLS["gestor"]) or "").strip(),
                    "diretor":       str(get_cell(vals, PAINEL_COLS["diretor"]) or "").strip(),
                    "saldo_prestacao":      safe_decimal(get_cell(vals, PAINEL_COLS["saldo_prestacao"])),
                    "saldo_cartao":         safe_decimal(get_cell(vals, PAINEL_COLS["saldo_cartao_painel"])),
                    "saldo_final":          safe_decimal(get_cell(vals, PAINEL_COLS["saldo_final"])),
                }
    print(f"  CONTROLE PAINEL: {len(records)} colaboradores loaded.")
    return records


def read_carga(path, month, quinzena):
    """Read a CARGA sheet. Returns {cpf: financial_record}."""
    cfg = CARGA_CONFIG.get((month, quinzena))
    if cfg is None:
        print(f"  [SKIP] No config for month={month} quinzena={quinzena}")
        return {}

    if not os.path.exists(path):
        print(f"  [SKIP] File not found: {path}")
        return {}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if cfg["sheet"] not in wb.sheetnames:
        print(f"  [AVISO] Sheet '{cfg['sheet']}' not found in {os.path.basename(path)}")
        print(f"         Available sheets: {wb.sheetnames}")
        wb.close()
        return {}

    ws = wb[cfg["sheet"]]
    cols = cfg["cols"]
    records = {}
    multiplier = REEMBOLSO_MULTIPLIERS.get(month, Decimal("0.5"))

    for i, row in enumerate(ws.iter_rows(min_row=cfg["data"], values_only=True)):
        cpf = normalize_cpf(get_cell(row, cols["cpf"]))
        if not cpf:
            continue

        # Skip duplicate/summary rows where key financial fields are all None
        # (Some sheets have a second row per CPF with only Carga Final populated)
        saldo_reembolsar_raw = safe_decimal(get_cell(row, cols["saldo_reembolsar"]))
        saldo_final_carga = safe_decimal(get_cell(row, cols["saldo_final_carga"]))
        col_qz = safe_decimal(get_cell(row, cols["col_qz"]))
        saldo_cartao_carga = safe_decimal(get_cell(row, cols["saldo_cartao_carga"]))
        adiantamento = safe_decimal(get_cell(row, cols["adiantamento"]))

        # If this CPF already has data and the current row has no col_qz and no saldo_final,
        # it's a duplicate/summary row — skip it
        if cpf in records and col_qz is None and saldo_final_carga is None:
            continue

        # Read REEMBOLSO column directly from sheet (1QZ only)
        # This is the authoritative value computed by the sheet's formula
        reembolso_col = cols.get("reembolso")
        reembolso_from_sheet = safe_decimal(get_cell(row, reembolso_col)) if reembolso_col is not None else None

        # --- Derive painel_saldo_final (the PAINEL saldo_final value) ---
        # Uses the RAW sheet values, NOT the reverse-computed saldo_reembolsar
        if saldo_reembolsar_raw is not None and saldo_reembolsar_raw > 0:
            painel_saldo_final = -saldo_reembolsar_raw
        elif saldo_final_carga is not None:
            painel_saldo_final = saldo_final_carga
        else:
            painel_saldo_final = Decimal("0")

        # --- Compute effective saldo_reembolsar for DB ---
        # The API formula is: reembolso = max(0, saldo_reembolsar) * multiplier
        # We want: max(0, stored_saldo_reembolsar) * multiplier = sheet's REEMBOLSO
        # So: stored_saldo_reembolsar = sheet's REEMBOLSO / multiplier
        if reembolso_from_sheet is not None and reembolso_from_sheet > 0 and multiplier > 0:
            saldo_reembolsar = reembolso_from_sheet / multiplier
        elif quinzena == 2:
            # 2QZ: reembolso is always 0
            saldo_reembolsar = Decimal("0")
        elif saldo_reembolsar_raw is not None and saldo_reembolsar_raw > 0:
            # 1QZ with positive SALDO REEMBOLSAR (normal case, matches API formula)
            saldo_reembolsar = saldo_reembolsar_raw
        else:
            # 1QZ with no reembolso
            saldo_reembolsar = Decimal("0")

        # Derive saldo_prestacao = saldo_final + saldo_cartao
        sc = saldo_cartao_carga if saldo_cartao_carga is not None else Decimal("0")
        saldo_prestacao_derived = painel_saldo_final + sc

        records[cpf] = {
            "colaborador":        str(get_cell(row, cols["colaborador"]) or "").strip(),
            "situacao":           str(get_cell(row, cols["status_colab"]) or "").strip(),
            "status_cartao":      str(get_cell(row, cols["status_cartao"]) or "").strip(),
            "saldo_prestacao":    saldo_prestacao_derived,
            "saldo_cartao":       sc,
            "saldo_final":        painel_saldo_final,
            "col_qz":             col_qz,
            "saldo_reembolsar":   saldo_reembolsar,
            "saldo_final_carga":  saldo_final_carga if saldo_final_carga is not None else Decimal("0"),
            "saldo_cartao_carga": sc,
            "adiantamento":       adiantamento,
        }

    wb.close()
    print(f"  CARGA {month}/{quinzena} (sheet '{cfg['sheet']}'): {len(records)} colaboradores.")
    return records


# =============================================================================
# Import to Neon
# =============================================================================

def import_to_neon(controle_data, carga_data, year, month, quinzena, dry_run=False):
    """Merge controle + carga data and upsert into quinzena_controle_snapshot."""
    filename_carga = os.path.basename(CARGA_FILES.get((month, quinzena), "N/A"))
    import_source = f"CONTROLE-JULHO-2026.xlsb + {filename_carga}"

    rows_to_insert = []
    # Start with all CPFs from CONTROLE (for full cadastro coverage)
    all_cpfs = set(controle_data.keys()) | set(carga_data.keys())

    for cpf in all_cpfs:
        ctrl = controle_data.get(cpf, {})
        carg = carga_data.get(cpf, {})

        # Cadastro from CONTROLE (full names, complete data)
        colaborador = ctrl.get("colaborador") or carg.get("colaborador") or None
        regional = ctrl.get("regional") or None
        centro_custo = ctrl.get("centro_custo") or None
        gestor = ctrl.get("gestor") or None
        diretor = ctrl.get("diretor") or None

        # Situacao and status_cartao from CARGA (per-month accurate) with CONTROLE fallback
        situacao = carg.get("situacao") or ctrl.get("situacao") or None
        status_cartao = carg.get("status_cartao") or ctrl.get("status_cartao") or None

        # Financial data from CARGA (per-quinzena accurate)
        if carg:
            saldo_prestacao = carg.get("saldo_prestacao")
            saldo_cartao = carg.get("saldo_cartao")
            saldo_final = carg.get("saldo_final")
            col_qz = carg.get("col_qz")
            saldo_reembolsar = carg.get("saldo_reembolsar")
            saldo_final_carga = carg.get("saldo_final_carga")
            saldo_cartao_carga = carg.get("saldo_cartao_carga")
        else:
            # No CARGA data for this CPF — set financial fields to NULL
            # (don't use July CONTROLE values as they're not per-month accurate)
            saldo_prestacao = None
            saldo_cartao = None
            saldo_final = None
            col_qz = None
            saldo_reembolsar = None
            saldo_final_carga = None
            saldo_cartao_carga = None

        rows_to_insert.append({
            "year": year, "month": month, "quinzena": quinzena,
            "cpf": cpf,
            "colaborador": colaborador,
            "situacao": situacao,
            "status_cartao": status_cartao,
            "regional": regional,
            "centro_custo": centro_custo,
            "gestor": gestor,
            "diretor": diretor,
            "saldo_prestacao": saldo_prestacao,
            "saldo_cartao": saldo_cartao,
            "saldo_final": saldo_final,
            "col_qz": col_qz,
            "saldo_reembolsar": saldo_reembolsar,
            "saldo_final_carga": saldo_final_carga,
            "saldo_cartao_carga": saldo_cartao_carga,
            "import_source": import_source,
        })

    print(f"  Prepared {len(rows_to_insert)} rows for {year}/{month:02d} QZ{quinzena}")
    carga_count = sum(1 for r in rows_to_insert if r["col_qz"] is not None)
    print(f"    (with col_qz from CARGA: {carga_count}, without: {len(rows_to_insert) - carga_count})")

    if dry_run:
        print("  [DRY RUN] No data written to DB.")
        # Show sample
        for r in rows_to_insert[:3]:
            print(f"    CPF={r['cpf']} | {r['colaborador'][:30] if r['colaborador'] else 'N/A'}")
            print(f"      col_qz={r['col_qz']} | saldo_final={r['saldo_final']} | saldo_final_carga={r['saldo_final_carga']}")
        return {"rows_imported": 0, "rows_total": len(rows_to_insert)}

    if not NEON_DATABASE_URL:
        print("ERRO: NEON_DATABASE_URL nao configurada")
        sys.exit(1)

    conn = psycopg2.connect(NEON_DATABASE_URL)
    conn.autocommit = False
    cur = conn.cursor()

    upsert_sql = """
        INSERT INTO quinzena_controle_snapshot
          (year, month, quinzena, cpf,
           colaborador, situacao, status_cartao,
           regional, centro_custo, gestor, diretor,
           saldo_prestacao, saldo_cartao, saldo_final,
           col_qz, saldo_reembolsar, saldo_final_carga, saldo_cartao_carga,
           import_source, imported_at)
        VALUES
          (%(year)s, %(month)s, %(quinzena)s, %(cpf)s,
           %(colaborador)s, %(situacao)s, %(status_cartao)s,
           %(regional)s, %(centro_custo)s, %(gestor)s, %(diretor)s,
           %(saldo_prestacao)s, %(saldo_cartao)s, %(saldo_final)s,
           %(col_qz)s, %(saldo_reembolsar)s, %(saldo_final_carga)s, %(saldo_cartao_carga)s,
           %(import_source)s, NOW())
        ON CONFLICT ON CONSTRAINT uq_snapshot
        DO UPDATE SET
          colaborador         = EXCLUDED.colaborador,
          situacao            = EXCLUDED.situacao,
          status_cartao       = EXCLUDED.status_cartao,
          regional            = EXCLUDED.regional,
          centro_custo        = EXCLUDED.centro_custo,
          gestor              = EXCLUDED.gestor,
          diretor             = EXCLUDED.diretor,
          saldo_prestacao     = EXCLUDED.saldo_prestacao,
          saldo_cartao        = EXCLUDED.saldo_cartao,
          saldo_final         = EXCLUDED.saldo_final,
          col_qz              = EXCLUDED.col_qz,
          saldo_reembolsar    = EXCLUDED.saldo_reembolsar,
          saldo_final_carga   = EXCLUDED.saldo_final_carga,
          saldo_cartao_carga  = EXCLUDED.saldo_cartao_carga,
          import_source       = EXCLUDED.import_source,
          imported_at         = NOW()
    """

    try:
        psycopg2.extras.execute_batch(cur, upsert_sql, rows_to_insert, page_size=100)
        conn.commit()
        print(f"  ✓ Imported {len(rows_to_insert)} rows to Neon.")
    except Exception as e:
        conn.rollback()
        print(f"  ✗ ERROR: {e}")
        raise
    finally:
        cur.close()
        conn.close()

    return {"rows_imported": len(rows_to_insert), "rows_total": len(rows_to_insert)}


def import_manual_inputs(carga_data, year, month, quinzena, dry_run=False):
    """Import col_qz and adiantamento into quinzena_manual_inputs."""
    if not NEON_DATABASE_URL or dry_run:
        print("  [DRY RUN] Skipping manual inputs.")
        return

    conn = psycopg2.connect(NEON_DATABASE_URL)
    conn.autocommit = False
    cur = conn.cursor()

    upsert_sql = """
        INSERT INTO quinzena_manual_inputs (cpf, year, month, quinzena, col_1qz, adiantamento, updated_at)
        VALUES (%(cpf)s, %(year)s, %(month)s, %(quinzena)s, %(col_1qz)s, %(adiantamento)s, NOW())
        ON CONFLICT (cpf, year, month, quinzena) WHERE cpf IS NOT NULL
        DO UPDATE SET
          col_1qz = EXCLUDED.col_1qz,
          adiantamento = EXCLUDED.adiantamento,
          updated_at = NOW()
    """

    rows = []
    for cpf, rec in carga_data.items():
        col_qz = rec.get("col_qz")
        adiantamento = rec.get("adiantamento")
        # Only import if at least one is non-null and non-zero
        if col_qz is None and adiantamento is None:
            continue
        rows.append({
            "cpf": cpf,
            "year": year, "month": month, "quinzena": quinzena,
            "col_1qz": col_qz,
            "adiantamento": adiantamento,
        })

    if rows:
        try:
            psycopg2.extras.execute_batch(cur, upsert_sql, rows, page_size=100)
            conn.commit()
            print(f"  ✓ Imported {len(rows)} manual inputs.")
        except Exception as e:
            conn.rollback()
            print(f"  ✗ ERROR importing manual inputs: {e}")
    else:
        print("  No manual inputs to import.")

    cur.close()
    conn.close()


# =============================================================================
# Main
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description="Import all months to Neon")
    parser.add_argument("--dry-run", action="store_true", help="Don't write to DB")
    parser.add_argument("--months", type=str, default="1,2,3,4,5,6",
                        help="Comma-separated months to import (default: 1,2,3,4,5,6)")
    parser.add_argument("--no-manual", action="store_true", help="Skip manual inputs")
    args = parser.parse_args()

    months = [int(m) for m in args.months.split(",")]
    dry = args.dry_run

    print(f"\n{'='*80}")
    print(f"IMPORT ALL MONTHS — Jan-Jun 2026")
    print(f"Mode: {'DRY RUN' if dry else 'GRAVANDO no Neon'}")
    print(f"Months: {months}")
    print(f"{'='*80}\n")

    # 1. Load CONTROLE PAINEL (universal cadastro)
    print("Loading CONTROLE PAINEL (JULHO 2026.xlsb)...")
    if not os.path.exists(CONTROLE_PATH):
        print(f"ERRO: CONTROLE not found: {CONTROLE_PATH}")
        sys.exit(1)
    controle_data = read_controle_painel(CONTROLE_PATH)

    # 2. Import each month/quinzena
    results = {}
    for month in months:
        for quinzena in [1, 2]:
            key = (month, quinzena)
            carga_path = CARGA_FILES.get(key)
            if carga_path is None or not os.path.exists(carga_path):
                print(f"\n--- Month {month} QZ{quinzena}: NO CARGA SHEET — skipping ---")
                results[key] = "skipped (no sheet)"
                continue

            print(f"\n--- Month {month} QZ{quinzena} ---")
            print(f"  CARGA file: {os.path.basename(carga_path)}")

            carga_data = read_carga(carga_path, month, quinzena)
            if not carga_data:
                print(f"  No data read from CARGA sheet — skipping.")
                results[key] = "skipped (no data)"
                continue

            stats = import_to_neon(controle_data, carga_data, 2026, month, quinzena, dry_run=dry)
            results[key] = stats

            if not args.no_manual:
                import_manual_inputs(carga_data, 2026, month, quinzena, dry_run=dry)

    # 3. Summary
    print(f"\n{'='*80}")
    print("SUMMARY:")
    for key, val in sorted(results.items()):
        m, q = key
        print(f"  {m:02d}/{2026} QZ{q}: {val}")
    print(f"{'='*80}")


if __name__ == "__main__":
    main()
