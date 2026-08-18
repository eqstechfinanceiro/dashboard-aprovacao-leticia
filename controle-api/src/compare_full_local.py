#!/usr/bin/env python3
"""
compare_full_local.py
---------------------
Full offline comparison using neon_dump.xlsx + ref_dump.xlsx.
Uses reference EXTRATO's CPF mapping instead of name matching.
Compares absolute values (sign-agnostic).
"""
import argparse
import os
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl

BASE = Path(__file__).parent.parent
NEON_DUMP = BASE / "data" / "neon_dump.xlsx"
REF_DUMP = BASE / "data" / "ref_dump.xlsx"


def _r2(v):
    return float(Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def nc(raw):
    if raw is None:
        return ""
    s = str(raw).strip().replace(".", "").replace("-", "").replace("/", "").replace(" ", "")
    return s.zfill(11)


def nf(raw):
    try:
        return round(float(raw), 2) if raw is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


def load_neon_dump():
    wb = openpyxl.load_workbook(NEON_DUMP, read_only=True, data_only=True)

    # Cadastro
    ws = wb["cadastro"]
    cadastro = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cpf = nc(row[0])
        if cpf:
            cadastro[cpf] = {
                "colaborador": str(row[1] or ""),
                "situacao": str(row[2] or ""),
                "status_cartao": str(row[3] or ""),
            }

    # Extrato acumulado (already filtered to June 30)
    ws = wb["extrato_acumulado"]
    extrato_by_name = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = str(row[0] or "").upper()
        extrato_by_name[nome] = {
            "carga": nf(row[1]),
            "transferencia": nf(row[2]),
            "tarifa": nf(row[3]),
        }

    # Snapshots
    ws = wb["snapshots_all"]
    snapshots = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = str(row[0] or "").upper()
        d = row[1]
        if hasattr(d, "strftime"):
            d = d.date() if hasattr(d, "date") else d
        elif isinstance(d, str):
            d = date.fromisoformat(d)
        snapshots[nome].append((d, nf(row[2])))
    for nome in snapshots:
        snapshots[nome].sort(key=lambda x: x[0])

    # Somase
    ws = wb["somase_snapshots"]
    somase = defaultdict(dict)
    for row in ws.iter_rows(min_row=2, values_only=True):
        cpf = nc(row[0])
        qid = str(row[1])
        somase[qid][cpf] = nf(row[2])

    # Manual inputs
    ws = wb["manual_inputs"]
    manuais = defaultdict(dict)
    for row in ws.iter_rows(min_row=2, values_only=True):
        y, m, q = int(row[0]), int(row[1]), int(row[2])
        cpf = nc(row[3])
        manuais[(y, m, q)][cpf] = {
            "col_qz": nf(row[4]),
            "adiantamento": nf(row[5]),
        }

    # Quinzena config
    ws = wb["quinzena_config"]
    config = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        y, m, q = int(row[0]), int(row[1]), int(row[2])
        config[(y, m, q)] = nf(row[3])

    wb.close()
    return {
        "cadastro": cadastro,
        "extrato_by_name": extrato_by_name,
        "snapshots": snapshots,
        "somase": somase,
        "manuais": manuais,
        "config": config,
    }


def load_ref_dump():
    wb = openpyxl.load_workbook(REF_DUMP, read_only=True, data_only=True)

    # PAINEL
    ws = wb["painel"]
    painel = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cpf = nc(row[0])
        if cpf:
            painel[cpf] = {
                "colaborador": str(row[1] or ""),
                "carga": nf(row[2]),
                "transferencia": abs(nf(row[3])),  # store as positive
                "tarifa": abs(nf(row[4])),  # store as positive
                "prestacao": nf(row[5]),
                "saldo_prestacao": nf(row[6]),
                "saldo_cartao": abs(nf(row[7])),  # store as positive
                "saldo_final": nf(row[8]),
                "col_1qz": nf(row[9]),
                "col_2qz": nf(row[10]),
            }

    # EXTRATO ref — build name→CPF mapping and per-CPF totals
    ws = wb["extrato_ref"]
    name_to_cpf = {}
    extrato_by_cpf = defaultdict(lambda: {"carga": 0.0, "transferencia": 0.0, "tarifa": 0.0})
    for row in ws.iter_rows(min_row=2, values_only=True):
        cpf = nc(row[0])
        nome = str(row[1] or "").upper()
        tipo = str(row[2] or "")
        valor = nf(row[3])
        if cpf and cpf != "00000000000":
            name_to_cpf[nome] = cpf
            if tipo == "CARGA" and valor > 0:
                extrato_by_cpf[cpf]["carga"] += valor
            elif tipo == "TRANSFERÊNCIA":
                extrato_by_cpf[cpf]["transferencia"] += abs(valor)
            elif tipo == "TARIFA":
                extrato_by_cpf[cpf]["tarifa"] += abs(valor)

    # BASE PREST ref — per-CPF totals
    ws = wb["base_prest_ref"]
    prest_by_cpf = defaultdict(float)
    for row in ws.iter_rows(min_row=2, values_only=True):
        cpf = nc(row[0])
        valor = nf(row[1])
        if cpf:
            prest_by_cpf[cpf] += valor

    wb.close()
    return {
        "painel": painel,
        "name_to_cpf": name_to_cpf,
        "extrato_by_cpf": dict(extrato_by_cpf),
        "prest_by_cpf": dict(prest_by_cpf),
    }


def get_saldo_cartao(snapshots, nome_upper, cutoff_date):
    snaps = snapshots.get(nome_upper, [])
    best = 0.0
    for d, v in snaps:
        if d <= cutoff_date:
            best = v
        else:
            break
    return best


def main():
    parser = argparse.ArgumentParser(description="Full local comparison")
    parser.add_argument("--ano", type=int, default=2026)
    parser.add_argument("--mes", type=int, default=7)
    parser.add_argument("--quinzena", type=int, choices=[1, 2], default=1)
    args = parser.parse_args()

    print("Loading neon dump...")
    neon = load_neon_dump()
    print(f"  cadastro: {len(neon['cadastro'])} | extrato: {len(neon['extrato_by_name'])} | snapshots: {len(neon['snapshots'])}")

    print("Loading ref dump...")
    ref = load_ref_dump()
    print(f"  painel: {len(ref['painel'])} | name_to_cpf: {len(ref['name_to_cpf'])} | extrato_by_cpf: {len(ref['extrato_by_cpf'])} | prest_by_cpf: {len(ref['prest_by_cpf'])}")

    # Build name→CPF mapping from ref, fallback to cadastro name
    name_to_cpf = dict(ref["name_to_cpf"])
    # Also build from cadastro (name → cpf)
    cpf_to_name = {}
    for cpf, cad in neon["cadastro"].items():
        nome_up = cad["colaborador"].upper()
        cpf_to_name[cpf] = nome_up
        if nome_up not in name_to_cpf:
            name_to_cpf[nome_up] = cpf

    # Calculate using REF data (extrato by CPF + prest by CPF + neon snapshots)
    cutoff_fin = date(args.ano, args.mes - 1, 30)  # June 30
    saldo_controle_date = date(args.ano, args.mes, 1)  # July 1
    quinzena_id = f"{args.ano}-{args.mes:02d}-{args.quinzena}"
    multiplier = neon["config"].get((args.ano, args.mes, args.quinzena), 0.5)
    somase_q = neon["somase"].get(quinzena_id, {})
    manuais_q = neon["manuais"].get((args.ano, args.mes, args.quinzena), {})

    print(f"\nCalculating using REF extrato + REF prestação + Neon snapshots...")
    print(f"  Cutoff: {cutoff_fin} | Saldo controle: {saldo_controle_date} | Multiplier: {multiplier}")

    # Use ref extrato and ref prestação, but neon snapshots and cadastro
    calc = {}
    for cpf, cad in neon["cadastro"].items():
        nome_up = cad["colaborador"].upper()

        # Use ref extrato by CPF if available, else neon extrato by name
        if cpf in ref["extrato_by_cpf"]:
            ext = ref["extrato_by_cpf"][cpf]
        else:
            ext = neon["extrato_by_name"].get(nome_up, {"carga": 0.0, "transferencia": 0.0, "tarifa": 0.0})

        # Use ref prestação by CPF if available, else neon somase
        if cpf in ref["prest_by_cpf"]:
            prest = ref["prest_by_cpf"][cpf]
        else:
            prest = somase_q.get(cpf, 0.0)

        # Saldo cartão from neon snapshots (by name)
        sc_controle = get_saldo_cartao(neon["snapshots"], nome_up, saldo_controle_date)

        # Calculate
        saldo_prestacao = _r2(ext["carga"] - ext["transferencia"] - ext["tarifa"] - prest)
        saldo_final_painel = _r2(saldo_prestacao - sc_controle)
        saldo_final = _r2(max(saldo_final_painel, 0.0))
        saldo_reembolsar = _r2(max(-saldo_final_painel, 0.0))

        man = manuais_q.get(cpf, {})
        col_qz = man.get("col_qz", 0.0)
        adiantamento = man.get("adiantamento", 0.0)

        reembolso = _r2(saldo_reembolsar * multiplier) if args.quinzena == 1 else 0.0

        calc[cpf] = {
            "colaborador": cad["colaborador"],
            "carga": ext["carga"],
            "transferencia": ext["transferencia"],
            "tarifa": ext["tarifa"],
            "prestacao": prest,
            "saldo_prestacao": saldo_prestacao,
            "saldo_cartao": sc_controle,
            "saldo_final": saldo_final,
        }

    # Compare with PAINEL
    painel = ref["painel"]
    cpfs_ref = set(painel.keys())
    cpfs_calc = set(calc.keys())
    tol = 0.05
    campos = ["carga", "transferencia", "tarifa", "prestacao", "saldo_prestacao", "saldo_cartao", "saldo_final"]
    divergencias = {c: [] for c in campos}
    total_comparados = 0

    for cpf in sorted(cpfs_ref & cpfs_calc):
        total_comparados += 1
        r = painel[cpf]
        c = calc[cpf]
        for campo in campos:
            v_ref = r[campo]
            v_calc = c.get(campo, 0.0)
            diff = abs(v_calc - v_ref)
            if diff > tol:
                divergencias[campo].append((cpf, c["colaborador"][:25], v_ref, v_calc, v_calc - v_ref))

    print()
    print("=" * 80)
    print(f"  COMPARAÇÃO (REF data) — {args.quinzena}ª QZ {args.mes:02d}/{args.ano}")
    print(f"  CPFs na planilha: {len(cpfs_ref)} | CPFs calculados: {len(cpfs_calc)} | Comparados: {total_comparados}")
    print(f"  Apenas na planilha: {len(cpfs_ref - cpfs_calc)} | Apenas no cálculo: {len(cpfs_calc - cpfs_ref)}")
    print("=" * 80)

    for campo in campos:
        divs = divergencias[campo]
        match = total_comparados - len(divs)
        pct = match / total_comparados * 100 if total_comparados else 0
        status = "OK" if len(divs) == 0 else ("~" if len(divs) <= 10 else "X")
        print(f"  [{status}] {campo:<22}: {match:>4}/{total_comparados}  ({pct:5.1f}%)  diverg: {len(divs)}")
        if divs:
            for cpf, nome, vr, vc, d in sorted(divs, key=lambda x: abs(x[4]), reverse=True)[:10]:
                print(f"       {cpf} {nome:<25} ref={vr:>12.2f}  calc={vc:>12.2f}  diff={d:>+12.2f}")

    print("=" * 80)

    # Totals
    total_carga_ref = sum(r["carga"] for r in painel.values())
    total_carga_calc = sum(calc[cpf]["carga"] for cpf in (cpfs_ref & cpfs_calc))
    total_prest_ref = sum(r["prestacao"] for r in painel.values())
    total_prest_calc = sum(calc[cpf]["prestacao"] for cpf in (cpfs_ref & cpfs_calc))
    total_sp_ref = sum(r["saldo_prestacao"] for r in painel.values())
    total_sp_calc = sum(calc[cpf]["saldo_prestacao"] for cpf in (cpfs_ref & cpfs_calc))
    total_sf_ref = sum(r["saldo_final"] for r in painel.values())
    total_sf_calc = sum(calc[cpf]["saldo_final"] for cpf in (cpfs_ref & cpfs_calc))
    total_sc_ref = sum(r["saldo_cartao"] for r in painel.values())
    total_sc_calc = sum(calc[cpf]["saldo_cartao"] for cpf in (cpfs_ref & cpfs_calc))

    print(f"  Total CARGA          ref: R$ {total_carga_ref:>14,.2f}  calc: R$ {total_carga_calc:>14,.2f}  diff: R$ {total_carga_calc - total_carga_ref:>+14,.2f}")
    print(f"  Total TRANSF         ref: R$ {sum(r['transferencia'] for r in painel.values()):>14,.2f}  calc: R$ {sum(calc[cpf]['transferencia'] for cpf in (cpfs_ref & cpfs_calc)):>14,.2f}")
    print(f"  Total TARIFA         ref: R$ {sum(r['tarifa'] for r in painel.values()):>14,.2f}  calc: R$ {sum(calc[cpf]['tarifa'] for cpf in (cpfs_ref & cpfs_calc)):>14,.2f}")
    print(f"  Total PRESTAÇÃO      ref: R$ {total_prest_ref:>14,.2f}  calc: R$ {total_prest_calc:>14,.2f}  diff: R$ {total_prest_calc - total_prest_ref:>+14,.2f}")
    print(f"  Total SALDO PREST    ref: R$ {total_sp_ref:>14,.2f}  calc: R$ {total_sp_calc:>14,.2f}  diff: R$ {total_sp_calc - total_sp_ref:>+14,.2f}")
    print(f"  Total SALDO CARTAO   ref: R$ {total_sc_ref:>14,.2f}  calc: R$ {total_sc_calc:>14,.2f}  diff: R$ {total_sc_calc - total_sc_ref:>+14,.2f}")
    print(f"  Total SALDO FINAL    ref: R$ {total_sf_ref:>14,.2f}  calc: R$ {total_sf_calc:>14,.2f}  diff: R$ {total_sf_calc - total_sf_ref:>+14,.2f}")
    print("=" * 80)

    # Now calculate using NEON data (our actual calculation) for comparison
    print(f"\n  --- Now comparing NEON calc (name-matched) vs REF ---")

    calc_neon = {}
    for cpf, cad in neon["cadastro"].items():
        nome_up = cad["colaborador"].upper()
        ext = neon["extrato_by_name"].get(nome_up, {"carga": 0.0, "transferencia": 0.0, "tarifa": 0.0})
        prest = somase_q.get(cpf, 0.0)
        sc_controle = get_saldo_cartao(neon["snapshots"], nome_up, saldo_controle_date)

        saldo_prestacao = _r2(ext["carga"] - ext["transferencia"] - ext["tarifa"] - prest)
        saldo_final_painel = _r2(saldo_prestacao - sc_controle)
        saldo_final = _r2(max(saldo_final_painel, 0.0))

        calc_neon[cpf] = {
            "colaborador": cad["colaborador"],
            "carga": ext["carga"],
            "transferencia": ext["transferencia"],
            "tarifa": ext["tarifa"],
            "prestacao": prest,
            "saldo_prestacao": saldo_prestacao,
            "saldo_cartao": sc_controle,
            "saldo_final": saldo_final,
        }

    # Compare neon calc vs ref
    divs_neon = {c: [] for c in campos}
    for cpf in sorted(cpfs_ref & cpfs_calc):
        r = painel[cpf]
        c = calc_neon[cpf]
        for campo in campos:
            v_ref = r[campo]
            v_calc = c.get(campo, 0.0)
            diff = abs(v_calc - v_ref)
            if diff > tol:
                divs_neon[campo].append((cpf, c["colaborador"][:25], v_ref, v_calc, v_calc - v_ref))

    print()
    print("=" * 80)
    print(f"  NEON CALC vs REF — {args.quinzena}ª QZ {args.mes:02d}/{args.ano}")
    print("=" * 80)

    for campo in campos:
        divs = divs_neon[campo]
        match = total_comparados - len(divs)
        pct = match / total_comparados * 100 if total_comparados else 0
        status = "OK" if len(divs) == 0 else ("~" if len(divs) <= 10 else "X")
        print(f"  [{status}] {campo:<22}: {match:>4}/{total_comparados}  ({pct:5.1f}%)  diverg: {len(divs)}")
        if divs and campo in ("carga", "saldo_cartao", "saldo_final"):
            for cpf, nome, vr, vc, d in sorted(divs, key=lambda x: abs(x[4]), reverse=True)[:5]:
                print(f"       {cpf} {nome:<25} ref={vr:>12.2f}  calc={vc:>12.2f}  diff={d:>+12.2f}")

    print("=" * 80)
    total_carga_neon = sum(calc_neon[cpf]["carga"] for cpf in (cpfs_ref & cpfs_calc))
    total_prest_neon = sum(calc_neon[cpf]["prestacao"] for cpf in (cpfs_ref & cpfs_calc))
    total_sc_neon = sum(calc_neon[cpf]["saldo_cartao"] for cpf in (cpfs_ref & cpfs_calc))
    total_sf_neon = sum(calc_neon[cpf]["saldo_final"] for cpf in (cpfs_ref & cpfs_calc))
    print(f"  Total CARGA          ref: R$ {total_carga_ref:>14,.2f}  neon: R$ {total_carga_neon:>14,.2f}  diff: R$ {total_carga_neon - total_carga_ref:>+14,.2f}")
    print(f"  Total PRESTAÇÃO      ref: R$ {total_prest_ref:>14,.2f}  neon: R$ {total_prest_neon:>14,.2f}  diff: R$ {total_prest_neon - total_prest_ref:>+14,.2f}")
    print(f"  Total SALDO CARTAO   ref: R$ {total_sc_ref:>14,.2f}  neon: R$ {total_sc_neon:>14,.2f}  diff: R$ {total_sc_neon - total_sc_ref:>+14,.2f}")
    print(f"  Total SALDO FINAL    ref: R$ R$ {total_sf_ref:>14,.2f}  neon: R$ {total_sf_neon:>14,.2f}  diff: R$ {total_sf_neon - total_sf_ref:>+14,.2f}")
    print("=" * 80)


if __name__ == "__main__":
    main()
