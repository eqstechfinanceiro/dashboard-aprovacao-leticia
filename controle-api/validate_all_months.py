#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
validate_all_months.py
======================
Validates API data against CARGA sheets for all months.
Compares saldo_final, col_qz, and carga_final.
"""
import os
import json
import openpyxl
from decimal import Decimal

ROOT = r'C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test'
DATA_DIR = os.path.join(ROOT, 'data')
API_JSON = os.path.join(ROOT, 'api_all_quinzenas.json')
OUT = os.path.join(ROOT, 'controle-api', 'validation_results.txt')

# Reuse the same CARGA config from import_all_months.py
CARGA_CONFIG = {
    (1, 1): {"sheet": "1 QZ VEXPENSES 01_2026", "header": 6, "data": 7,
             "cols": {"cpf": 2, "colaborador": 1, "saldo_final_carga": 9, "col_qz": 10,
                      "saldo_cartao_carga": 11, "adiantamento": 12, "saldo_reembolsar": 8,
                      "carga_final": 16}},  # CARGA FINAL REEMBOLSO (includes reembolso)
    (1, 2): {"sheet": "2 QZ VEXPENSES 01_2026", "header": 6, "data": 7,
             "cols": {"cpf": 2, "colaborador": 1, "saldo_final_carga": 12, "col_qz": 14,
                      "saldo_cartao_carga": 15, "adiantamento": 16, "saldo_reembolsar": 8,
                      "carga_final": 21}},  # CARGA FINAL COM CX REP
    (2, 1): {"sheet": "1 QZN FEV 2026", "header": 6, "data": 7,
             "cols": {"cpf": 2, "colaborador": 1, "saldo_final_carga": 9, "col_qz": 10,
                      "saldo_cartao_carga": 11, "adiantamento": 12, "saldo_reembolsar": 8,
                      "carga_final": 15}},  # CARGA FINAL
    (2, 2): {"sheet": "2 QZ VEXPENSES 02_2026", "header": 6, "data": 7,
             "cols": {"cpf": 2, "colaborador": 1, "saldo_final_carga": 11, "col_qz": 12,
                      "saldo_cartao_carga": 13, "adiantamento": 14, "saldo_reembolsar": 8,
                      "carga_final": 17}},  # CARGA FINAL
    (3, 1): {"sheet": "QUINZENA MARÇO", "header": 6, "data": 7,
             "cols": {"cpf": 2, "colaborador": 1, "saldo_final_carga": 9, "col_qz": 12,
                      "saldo_cartao_carga": 13, "adiantamento": 14, "saldo_reembolsar": 8,
                      "carga_final": 19}},  # CARGA FINAL SEM CX REPROV.
    (4, 1): {"sheet": "1 QZ VEXPENSES 04_2026", "header": 6, "data": 7,
             "cols": {"cpf": 3, "colaborador": 2, "saldo_final_carga": 10, "col_qz": 11,
                      "saldo_cartao_carga": 12, "adiantamento": 13, "saldo_reembolsar": 9,
                      "carga_final": 16}},  # CARGA FINAL
    (5, 1): {"sheet": "Planilha1", "header": 6, "data": 7,
             "cols": {"cpf": 1, "colaborador": 0, "saldo_final_carga": 8, "col_qz": 9,
                      "saldo_cartao_carga": 10, "adiantamento": 11, "saldo_reembolsar": 7,
                      "carga_final": 14}},  # Carga Final
    (5, 2): {"sheet": "2 QZ DE MAIO 26", "header": 4, "data": 5,
             "cols": {"cpf": 2, "colaborador": 1, "saldo_final_carga": 9, "col_qz": 10,
                      "saldo_cartao_carga": 11, "adiantamento": 12, "saldo_reembolsar": 7,
                      "carga_final": 15}},  # Carga Final
    (6, 1): {"sheet": "1 QZ JUNHO", "header": 6, "data": 7,
             "cols": {"cpf": 1, "colaborador": 0, "saldo_final_carga": 8, "col_qz": 9,
                      "saldo_cartao_carga": 10, "adiantamento": 11, "saldo_reembolsar": 7,
                      "carga_final": 14}},  # Carga Final
    (6, 2): {"sheet": "2 QZ JUNHO", "header": 6, "data": 7,
             "cols": {"cpf": 1, "colaborador": 0, "saldo_final_carga": 7, "col_qz": 8,
                      "saldo_cartao_carga": 9, "adiantamento": 10, "saldo_reembolsar": None,
                      "carga_final": 13}},  # Carga Final
}

CARGA_FILES = {
    (1, 1): os.path.join(DATA_DIR, "01 - JANEIRO", "1QZ JANEIRO 2026 - VEXPENSES.xlsx"),
    (1, 2): os.path.join(DATA_DIR, "01 - JANEIRO", "2QZ JANEIRO 2026 - VEXPENSES.xlsx"),
    (2, 1): os.path.join(DATA_DIR, "02 - FEVEREIRO", "1 QZN FEVEREIRO VEXPENSES 2026.xlsx"),
    (2, 2): os.path.join(DATA_DIR, "02 - FEVEREIRO", "2QZ FEVEREIRO 2026 - VEXPENSES EQS.xlsx"),
    (3, 1): os.path.join(DATA_DIR, "03 - MARÇO", "1 QZ MARÇO VEXPENSES 2026 (5).xlsx"),
    (4, 1): os.path.join(DATA_DIR, "04 - ABRIL", "1QZ ABRIL 2026 - VEXPENSES.xlsx"),
    (5, 1): os.path.join(DATA_DIR, "05 - MAIO", "CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx"),
    (5, 2): os.path.join(DATA_DIR, "05 - MAIO", "CARGA 2 QZ MAIO 26 VEXPENSES EQS.xlsx"),
    (6, 1): os.path.join(DATA_DIR, "06 - JUNHO", "CARGA 1 QZ JUNHO 26 VEXPENSES EQS.xlsx"),
    (6, 2): os.path.join(DATA_DIR, "06 - JUNHO", "CARGA 2 QZ JUNHO 26 VEXPENSES EQS.xlsx"),
}


def normalize_cpf(raw):
    if raw is None:
        return None
    s = str(raw).strip().replace(".", "").replace("-", "").replace("/", "").replace(" ", "")
    if "." in s:
        try:
            s = str(int(float(s)))
        except (ValueError, TypeError):
            pass
    return s.zfill(11) if s.isdigit() and len(s) <= 11 else None


def to_float(v):
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def read_carga_for_validation(path, month, quinzena):
    """Read CARGA sheet and return {cpf: {saldo_final, col_qz, carga_final, ...}}"""
    cfg = CARGA_CONFIG.get((month, quinzena))
    if cfg is None or not os.path.exists(path):
        return {}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if cfg["sheet"] not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb[cfg["sheet"]]
    cols = cfg["cols"]
    records = {}

    for row in ws.iter_rows(min_row=cfg["data"], values_only=True):
        cpf = normalize_cpf(row[cols["cpf"]] if cols["cpf"] < len(row) else None)
        if not cpf:
            continue

        # Skip duplicate/summary rows (same logic as import script)
        col_qz_val = row[cols["col_qz"]] if cols["col_qz"] < len(row) else None
        sf_val = row[cols["saldo_final_carga"]] if cols["saldo_final_carga"] < len(row) else None
        if cpf in records and col_qz_val is None and sf_val is None:
            continue

        carga_final_idx = cols.get("carga_final")
        carga_final = to_float(row[carga_final_idx]) if carga_final_idx is not None and carga_final_idx < len(row) else None

        records[cpf] = {
            "colaborador": str(row[cols["colaborador"]] or "").strip() if cols["colaborador"] < len(row) else "",
            "saldo_final_sheet": to_float(row[cols["saldo_final_carga"]] if cols["saldo_final_carga"] < len(row) else None),
            "col_qz_sheet": to_float(row[cols["col_qz"]] if cols["col_qz"] < len(row) else None),
            "carga_final_sheet": carga_final if carga_final is not None else 0.0,
        }

    wb.close()
    return records


def validate_quinzena(api_data, sheet_data, month, quinzena, fh):
    """Compare API data against sheet data."""
    fh.write(f"\n{'='*80}\n")
    fh.write(f"VALIDATION: {month:02d}/2026 QZ{quinzena}\n")
    fh.write(f"{'='*80}\n")

    if api_data.get("error"):
        fh.write(f"  API ERROR: {api_data['error']}\n")
        return

    data_mode = api_data.get("data_mode", "unknown")
    fh.write(f"  API data_mode: {data_mode}\n")
    fh.write(f"  API total rows: {api_data.get('statistics', {}).get('total_rows', '?')}\n")
    fh.write(f"  Sheet total rows: {len(sheet_data)}\n")

    if data_mode != "snapshot" and data_mode != "calculado":
        fh.write(f"  [WARNING] Unknown data_mode '{data_mode}' — skipping.\n")
        return

    api_rows = api_data.get("data", [])
    api_by_cpf = {r["cpf"]: r for r in api_rows}

    matches_sf = 0
    matches_cq = 0
    matches_cf = 0
    matches_all = 0
    total_compared = 0
    mismatches = []

    for cpf, sheet in sheet_data.items():
        if cpf not in api_by_cpf:
            continue

        api = api_by_cpf[cpf]
        total_compared += 1

        sf_api = round(api.get("saldo_final_carga", 0) or 0, 2)
        sf_sheet = round(sheet["saldo_final_sheet"], 2)
        # In calculado mode, col_qz is null but col_qz_manual has the value
        cq_api = round(api.get("col_qz_manual") or api.get("col_qz") or 0, 2)
        cq_sheet = round(sheet["col_qz_sheet"], 2)
        cf_api = round(api.get("carga_final", 0), 2)
        cf_sheet = round(sheet["carga_final_sheet"], 2)

        sf_match = abs(sf_api - sf_sheet) < 0.02
        cq_match = abs(cq_api - cq_sheet) < 0.02
        # For carga_final: API clamps to max(0, carga_parcial) + reembolso (always >= 0)
        # Sheet may have negative values — treat as match when API=0 and sheet<0
        if cf_api == 0 and cf_sheet < 0:
            cf_match = True
        else:
            cf_match = abs(cf_api - cf_sheet) < 0.02

        if sf_match:
            matches_sf += 1
        if cq_match:
            matches_cq += 1
        if cf_match:
            matches_cf += 1
        if sf_match and cq_match and cf_match:
            matches_all += 1
        else:
            mismatches.append({
                "cpf": cpf,
                "name": (api.get("colaborador") or "")[:25],
                "sf_api": sf_api, "sf_sheet": sf_sheet, "sf_diff": round(sf_api - sf_sheet, 2),
                "cq_api": cq_api, "cq_sheet": cq_sheet, "cq_diff": round(cq_api - cq_sheet, 2),
                "cf_api": cf_api, "cf_sheet": cf_sheet, "cf_diff": round(cf_api - cf_sheet, 2),
            })

    if total_compared == 0:
        fh.write(f"  No matching CPFs between API and sheet!\n")
        return

    fh.write(f"\n  Compared: {total_compared} CPFs\n")
    fh.write(f"  saldo_final matches:   {matches_sf}/{total_compared} ({matches_sf/total_compared*100:.1f}%)\n")
    fh.write(f"  col_qz matches:        {matches_cq}/{total_compared} ({matches_cq/total_compared*100:.1f}%)\n")
    fh.write(f"  carga_final matches:   {matches_cf}/{total_compared} ({matches_cf/total_compared*100:.1f}%)\n")
    fh.write(f"  ALL 3 match:           {matches_all}/{total_compared} ({matches_all/total_compared*100:.1f}%)\n")

    if mismatches:
        mismatches.sort(key=lambda x: abs(x["cf_diff"]), reverse=True)
        fh.write(f"\n  TOP 15 MISMATCHES (by carga_final diff):\n")
        for m in mismatches[:15]:
            fh.write(f"    {m['name']:25s} CPF={m['cpf']} | SF: api={m['sf_api']:>10.2f} sheet={m['sf_sheet']:>10.2f} d={m['sf_diff']:>8.2f} | CQ: api={m['cq_api']:>8.2f} sheet={m['cq_sheet']:>8.2f} d={m['cq_diff']:>8.2f} | CF: api={m['cf_api']:>10.2f} sheet={m['cf_sheet']:>10.2f} d={m['cf_diff']:>8.2f}\n")

    return {
        "month": month, "quinzena": quinzena,
        "compared": total_compared,
        "sf_match": matches_sf, "cq_match": matches_cq, "cf_match": matches_cf, "all_match": matches_all,
        "sf_pct": matches_sf/total_compared*100 if total_compared else 0,
        "cq_pct": matches_cq/total_compared*100 if total_compared else 0,
        "cf_pct": matches_cf/total_compared*100 if total_compared else 0,
        "all_pct": matches_all/total_compared*100 if total_compared else 0,
    }


def main():
    with open(API_JSON, 'r', encoding='utf-8') as f:
        api_data = json.loads(f.read())

    all_results = []

    with open(OUT, 'w', encoding='utf-8') as fh:
        fh.write("=== VALIDATION RESULTS: API vs CARGA sheets ===\n")
        fh.write(f"Date: 2026-07-14\n")

        for (month, quinzena), carga_path in sorted(CARGA_FILES.items()):
            if not os.path.exists(carga_path):
                fh.write(f"\n[SKIP] {month:02d}/2026 QZ{quinzena}: No CARGA sheet\n")
                continue

            sheet_data = read_carga_for_validation(carga_path, month, quinzena)
            api_key = f"{month}_{quinzena}"
            api = api_data.get(api_key, {})

            result = validate_quinzena(api, sheet_data, month, quinzena, fh)
            if result:
                all_results.append(result)

        # Summary table
        fh.write(f"\n{'='*80}\n")
        fh.write(f"SUMMARY TABLE\n")
        fh.write(f"{'='*80}\n")
        fh.write(f"{'Month':>6} {'QZ':>3} {'Compared':>8} {'SF%':>6} {'CQ%':>6} {'CF%':>6} {'ALL%':>6}\n")
        fh.write(f"{'-'*6} {'-'*3} {'-'*8} {'-'*6} {'-'*6} {'-'*6} {'-'*6}\n")
        for r in all_results:
            fh.write(f"{r['month']:>6} {r['quinzena']:>3} {r['compared']:>8} {r['sf_pct']:>5.1f}% {r['cq_pct']:>5.1f}% {r['cf_pct']:>5.1f}% {r['all_pct']:>5.1f}%\n")

    print(f"WROTE {OUT}")
    # Also print summary to console
    for r in all_results:
        print(f"  {r['month']:02d}/2026 QZ{r['quinzena']}: compared={r['compared']} SF={r['sf_pct']:.1f}% CQ={r['cq_pct']:.1f}% CF={r['cf_pct']:.1f}% ALL={r['all_pct']:.1f}%")


if __name__ == "__main__":
    main()
