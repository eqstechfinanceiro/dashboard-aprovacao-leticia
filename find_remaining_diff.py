import openpyxl
import json
from collections import defaultdict

# Load BASE PREST per CPF
wb = openpyxl.load_workbook('CONTROLE - VEXPENSES - AGOSTO 2026.xlsx', read_only=True, data_only=True)
ws = wb['BASE PREST ']

bp_per_cpf = defaultdict(float)
bp_fatura_per_cpf = defaultdict(float)
bp_reports = defaultdict(lambda: defaultdict(lambda: {'total': 0, 'count': 0, 'status': '', 'name': ''}))
for row in ws.iter_rows(min_row=4, values_only=True):
    cpf = str(row[9]).strip() if row[9] else ''
    valor = float(row[26]) if row[26] is not None else 0
    report_id = str(row[1]).strip() if row[1] else ''
    report_name = str(row[2]).strip().upper() if row[2] else ''
    status = str(row[10]).strip() if row[10] else ''
    if not cpf or cpf == 'None':
        continue
    bp_per_cpf[cpf] += valor
    if 'FATURA' in report_name or 'CARTAO' in report_name or 'CARTAO' in report_name:
        bp_fatura_per_cpf[cpf] += valor
    if report_id and report_id != 'None':
        bp_reports[cpf][report_id]['total'] += valor
        bp_reports[cpf][report_id]['count'] += 1
        bp_reports[cpf][report_id]['status'] = status
        bp_reports[cpf][report_id]['name'] = report_name

# Load API prestacao debug (fresh)
with open('prestacao_debug.json', 'r', encoding='utf-8-sig') as f:
    debug = json.load(f)

api_per_cpf = {}
for item in debug.get('cpf_summary', []):
    cpf = item['user_cpf'].strip()
    api_per_cpf[cpf] = {
        'prestacao_total': float(item['prestacao_total']),
        'total_expenses': float(item['total_expenses']),
        'report_count': int(item['report_count']),
    }

# Load API report details
api_reports = defaultdict(list)
for item in debug.get('expense_details', []):
    cpf = item['user_cpf'].strip()
    api_reports[cpf].append({
        'report_id': int(item['report_id']),
        'report_name': item['report_name'],
        'status': item['status'],
        'expense_count': int(item['expense_count']),
        'total_value': float(item['total_value']),
        'excluded_value': float(item['excluded_value']),
    })

# Load PAINEL prestacao
ws_painel = wb['PAINEL']
painel_prest = {}
painel_names = {}
for row in ws_painel.iter_rows(min_row=12, values_only=True):
    cpf = str(row[2]).strip() if row[2] else ''
    prest = float(row[16]) if row[16] is not None else 0
    nome = str(row[1]).strip() if row[1] else ''
    if cpf and cpf != 'None':
        painel_prest[cpf] = prest
        painel_names[cpf] = nome

# Compare per CPF: BASE PREST (non-FATURA) vs API prestacao
diffs = []
for cpf in set(bp_per_cpf.keys()) | set(api_per_cpf.keys()):
    bp_total = bp_per_cpf.get(cpf, 0) - bp_fatura_per_cpf.get(cpf, 0)  # non-FATURA only
    api_total = api_per_cpf.get(cpf, {}).get('prestacao_total', 0)
    diff = bp_total - api_total
    if abs(diff) > 1:
        diffs.append((cpf, painel_names.get(cpf, ''), bp_total, api_total, diff))

diffs.sort(key=lambda x: abs(x[4]), reverse=True)

print(f"Total remaining diff: {sum(d[4] for d in diffs):.2f}")
print(f"Users with diff: {len(diffs)}")
print()
print("Top 30 diffs (BASE_PREST_normal - API_prestacao):")
for cpf, nome, bp, api, diff in diffs[:30]:
    print(f"  {nome[:35]:35s} cpf={cpf} BP={bp:.2f} API={api:.2f} diff={diff:+.2f}")

# For top 5, show report-level comparison
print()
for cpf, nome, bp, api, diff in diffs[:5]:
    print(f"\n=== {nome} (cpf={cpf}) BP={bp:.2f} API={api:.2f} diff={diff:+.2f} ===")
    bp_reps = bp_reports.get(cpf, {})
    api_reps = {r['report_id']: r for r in api_reports.get(cpf, [])}

    # Reports in BASE PREST but not in API (or different values)
    all_rids = set(bp_reps.keys()) | set(api_reps.keys())
    for rid in sorted(all_rids, key=lambda x: abs(bp_reps.get(x, {}).get('total', 0) - api_reps.get(int(x) if str(x).isdigit() else 0, {}).get('total_value', 0)), reverse=True):
        bp_r = bp_reps.get(rid, {'total': 0, 'count': 0, 'status': '', 'name': ''})
        api_r = api_reps.get(int(rid) if str(rid).isdigit() else -1, {'report_name': '', 'status': '', 'expense_count': 0, 'total_value': 0, 'excluded_value': 0})
        rdiff = bp_r['total'] - api_r['total_value']
        if abs(rdiff) > 0.01:
            print(f"  rid={rid} bp_name={bp_r['name']} bp_status={bp_r['status']} bp_total={bp_r['total']:.2f} api_name={api_r['report_name']} api_status={api_r['status']} api_total={api_r['total_value']:.2f} api_excluded={api_r['excluded_value']:.2f} diff={rdiff:+.2f}")
