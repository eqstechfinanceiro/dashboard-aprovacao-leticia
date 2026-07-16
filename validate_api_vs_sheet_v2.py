"""Validate API-only calculations against CARGA sheets for 1QZ and 2QZ Junho 2026."""
import openpyxl

def read_carga_sheet(filepath):
    """Read CARGA sheet and return dict by CPF."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    header_row = None
    for row in range(1, 6):
        if ws.cell(row, 2).value == 'COLABORADOR':
            header_row = row
            break
    if not header_row:
        print(f"Could not find header row in {filepath}")
        return {}
    print(f"Header at row {header_row}")
    for col in range(1, 15):
        h = ws.cell(header_row, col).value
        if h:
            print(f"  Col {col}: {h}")
    data = {}
    for row in range(header_row + 1, ws.max_row + 1):
        cpf = ws.cell(row, 3).value
        if not cpf:
            continue
        cpf = str(cpf).strip()
        colaborador = ws.cell(row, 2).value
        saldo_final = ws.cell(row, 7).value
        col_qz = ws.cell(row, 8).value
        carga_final = ws.cell(row, 9).value
        data[cpf] = {
            'colaborador': colaborador,
            'saldo_final': float(saldo_final) if saldo_final is not None else 0,
            'col_qz': float(col_qz) if col_qz is not None else 0,
            'carga_final': float(carga_final) if carga_final is not None else 0,
        }
    return data

print("=== CARGA 1 QZ JUNHO ===")
sheet1qz = read_carga_sheet(r'data\06 - JUNHO\CARGA 1 QZ JUNHO 26 VEXPENSES EQS.xlsx')
print(f"\nTotal users in 1QZ sheet: {len(sheet1qz)}")

print("\n=== CARGA 2 QZ JUNHO ===")
sheet2qz = read_carga_sheet(r'data\06 - JUNHO\CARGA 2 QZ JUNHO 26 VEXPENSES EQS.xlsx')
print(f"\nTotal users in 2QZ sheet: {len(sheet2qz)}")

abner_cpf = '02027745203'
print(f"\n=== ABNER ({abner_cpf}) in sheets ===")
if abner_cpf in sheet1qz:
    print(f"  1QZ: {sheet1qz[abner_cpf]}")
if abner_cpf in sheet2qz:
    print(f"  2QZ: {sheet2qz[abner_cpf]}")
