"""Validate API-only calculations against CARGA sheet for 2QZ Junho 2026."""
import openpyxl
import json

# Read the CARGA 2 QZ JUNHO sheet
wb = openpyxl.load_workbook(r'data\06 - JUNHO\CARGA 2 QZ JUNHO 26 VEXPENSES EQS.xlsx', data_only=True)
ws = wb.active

# Print all headers in row 2
print("=== Sheet Headers (Row 2) ===")
headers = {}
for col in range(1, 30):
    v = ws.cell(row=2, column=col).value
    if v:
        headers[col] = str(v).strip()
        col_letter = chr(64 + col) if col <= 26 else '??'
        print(f"  Col {col} ({col_letter}): {v}")

# Print ABNER's full row (row 3)
print("\n=== ABNER (row 3) - Full data ===")
for col in range(1, 20):
    v = ws.cell(row=3, column=col).value
    h = ws.cell(row=2, column=col).value
    print(f"  {h or f'col{col}'}: {v}")

# Print a few more users
print("\n=== Rows 3-12 (first 10 users) ===")
for row in range(3, 13):
    colaborador = ws.cell(row=row, column=2).value
    cpf = ws.cell(row=row, column=3).value
    saldo_final = ws.cell(row=row, column=7).value
    col_qz = ws.cell(row=row, column=8).value
    carga_final = ws.cell(row=row, column=9).value
    print(f"  {colaborador} | CPF={cpf} | SF={saldo_final} | QZ={col_qz} | CF={carga_final}")

# Check if there are more columns beyond 9
print("\n=== Checking columns 10-20 for row 3 ===")
for col in range(10, 20):
    v = ws.cell(row=3, column=col).value
    h = ws.cell(row=2, column=col).value
    if v is not None or h:
        print(f"  {h or f'col{col}'}: {v}")
