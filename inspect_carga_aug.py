import openpyxl

wb = openpyxl.load_workbook('CARGA 1 QZ AGOSTO 26 VEXPENSES EQS.xlsx', read_only=True, data_only=True)
ws = wb['1 QZ AGOSTO']
for r in range(1, 12):
    vals = []
    for c in range(1, 18):
        cell = ws.cell(row=r, column=c)
        v = cell.value
        if v is not None:
            vals.append(f'col{c}={v}')
    if vals:
        print(f'Row {r}: {"  |  ".join(vals)}')

print("\n=== STATUS CARTAO ===")
ws2 = wb['STATUS CARTAO']
for r in range(1, 6):
    vals = []
    for c in range(1, 11):
        cell = ws2.cell(row=r, column=c)
        v = cell.value
        if v is not None:
            vals.append(f'col{c}={v}')
    if vals:
        print(f'Row {r}: {"  |  ".join(vals)}')
