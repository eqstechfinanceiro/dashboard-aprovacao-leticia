import psycopg2
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DB_URL = "postgresql://neondb_owner:npg_iItZN95svyEG@ep-restless-voice-amrrz188-pooler.c-5.us-east-1.aws.neon.tech/neondb?sslmode=require"

print("Connecting to DB...")
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()

# ===== Query 1: CPF Summary =====
print("Querying CPF summary...")
t0 = time.time()
cur.execute("""
    SELECT r.user_cpf, r.user_name,
           COUNT(DISTINCT r.id) as report_count,
           COUNT(e.id) as expense_count,
           COALESCE(SUM(e.value), 0) as total_prestacao
    FROM prestacao_reports r
    LEFT JOIN prestacao_expenses e ON e.report_id = r.id
    WHERE r.user_cpf IS NOT NULL
      AND (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND NOT (r.name ILIKE '%FATURA%' OR r.name ILIKE '%CARTAO%' OR r.name ILIKE '%CARTÃO%')
    GROUP BY r.user_cpf, r.user_name
    ORDER BY r.user_name
""")
cpf_summary = cur.fetchall()
print(f"  Got {len(cpf_summary)} CPFs in {time.time()-t0:.1f}s")

# ===== Query 2: Report Summary =====
print("Querying report summary...")
t0 = time.time()
cur.execute("""
    SELECT r.id, r.name, r.status, r.user_cpf, r.user_name,
           r.raw_data->>'approval_date' as approval_date,
           COUNT(e.id) as expense_count,
           COALESCE(SUM(e.value), 0) as total_value
    FROM prestacao_reports r
    LEFT JOIN prestacao_expenses e ON e.report_id = r.id
    WHERE r.user_cpf IS NOT NULL
      AND (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND NOT (r.name ILIKE '%FATURA%' OR r.name ILIKE '%CARTAO%' OR r.name ILIKE '%CARTÃO%')
    GROUP BY r.id, r.name, r.status, r.user_cpf, r.user_name, r.raw_data
    ORDER BY r.user_cpf, r.id
""")
report_summary = cur.fetchall()
print(f"  Got {len(report_summary)} reports in {time.time()-t0:.1f}s")

# ===== Query 3: All Expenses (using server-side cursor for memory efficiency) =====
print("Querying all expenses (streaming)...")
t0 = time.time()
cur.execute("""
    SELECT r.id, r.name, r.status, r.user_cpf, r.user_name,
           r.created_at, r.updated_at,
           r.raw_data->>'approval_date' as approval_date,
           r.total_value,
           e.id, e.description, e.value, e.date, e.status
    FROM prestacao_reports r
    LEFT JOIN prestacao_expenses e ON e.report_id = r.id
    WHERE r.user_cpf IS NOT NULL
    ORDER BY r.user_cpf, r.id, e.id
""")

# Create workbook
wb = Workbook()
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

# ===== Sheet 1: CPF Summary =====
ws1 = wb.active
ws1.title = "CPF Summary"
h1 = ["CPF", "Nome", "Qtd Relatórios", "Qtd Despesas", "Total Prestação"]
for col, h in enumerate(h1, 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill
    c.alignment = Alignment(horizontal='center'); c.border = thin_border
for i, row in enumerate(cpf_summary, 2):
    for col, val in enumerate(row, 1):
        c = ws1.cell(row=i, column=col, value=val)
        c.border = thin_border
        if col == 5: c.number_format = '#,##0.00'
for col, w in enumerate([15, 35, 15, 15, 18], 1):
    ws1.column_dimensions[get_column_letter(col)].width = w
ws1.auto_filter.ref = f"A1:E{len(cpf_summary)+1}"

# ===== Sheet 2: Report Summary =====
ws2 = wb.create_sheet("Report Summary")
h2 = ["Report ID", "Report Name", "Status", "CPF", "Nome", "Approval Date", "Qtd Despesas", "Total Valor"]
for col, h in enumerate(h2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill
    c.alignment = Alignment(horizontal='center'); c.border = thin_border
for i, row in enumerate(report_summary, 2):
    for col, val in enumerate(row, 1):
        c = ws2.cell(row=i, column=col, value=val)
        c.border = thin_border
        if col == 8: c.number_format = '#,##0.00'
for col, w in enumerate([12, 25, 12, 15, 30, 22, 12, 15], 1):
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.auto_filter.ref = f"A1:H{len(report_summary)+1}"

# ===== Sheet 3: All Expenses (streaming write) =====
ws3 = wb.create_sheet("BASE PREST (DB)")
h3 = ["Report ID", "Report Name", "Status", "CPF", "Nome",
      "Created At", "Updated At", "Approval Date", "Report Total",
      "Expense ID", "Expense Description", "Expense Value", "Expense Date", "Expense Status"]
for col, h in enumerate(h3, 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill
    c.alignment = Alignment(horizontal='center'); c.border = thin_border

# Stream rows
BATCH = 5000
row_idx = 2
total = 0
batch_count = 0
query_time = time.time() - t0
print(f"  Query done in {query_time:.1f}s, writing rows...")

while True:
    rows = cur.fetchmany(BATCH)
    if not rows:
        break
    for row in rows:
        for col, val in enumerate(row, 1):
            c = ws3.cell(row=row_idx, column=col, value=val)
            c.border = thin_border
            if col in (9, 12): c.number_format = '#,##0.00'
        row_idx += 1
    total += len(rows)
    batch_count += 1
    elapsed = time.time() - t0
    print(f"  Wrote {total:>7d} rows ({elapsed:.0f}s)", flush=True)

print(f"  Total: {total} expense rows in {time.time()-t0:.1f}s")

for col, w in enumerate([12, 25, 12, 15, 30, 22, 22, 22, 15, 12, 30, 15, 14, 15], 1):
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.auto_filter.ref = f"A1:N{row_idx-1}"
ws3.freeze_panes = "A2"

# Save
output = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\BASE_PREST_DB_EXPORT_V2.xlsx"
print(f"Saving to {output}...")
wb.save(output)
print(f"Done! {total} expenses, {len(report_summary)} reports, {len(cpf_summary)} CPFs")

cur.close()
conn.close()
