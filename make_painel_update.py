import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import time

SOURCE = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\BASE_PREST_DB_EXPORT_V2.xlsx"
CONTROLE = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - AGOSTO 2026.xlsx"
OUTPUT = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\PRESTACAO_PAINEL_UPDATE.xlsx"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
money_fmt = 'R$ #,##0.00;[Red]-R$ #,##0.00'

# Load DB CPF summary (only APROVADO/ENVIADO, no FATURA/CARTAO)
print("Loading DB CPF Summary...")
t0 = time.time()
wb_src = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
ws = wb_src['CPF Summary']
db_data = {}
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    cpf = str(row[0] or '').strip()
    nome = row[1] or ''
    qtd_rep = int(row[2] or 0)
    qtd_desp = int(row[3] or 0)
    total = float(row[4] or 0)
    if cpf:
        db_data[cpf] = {'nome': nome, 'qtd_rep': qtd_rep, 'qtd_desp': qtd_desp, 'total': total}
wb_src.close()
print(f"  {len(db_data)} CPFs from DB in {time.time()-t0:.1f}s")

# Load PAINEL to get the full list of CPFs and their order/names
print("Loading PAINEL...")
wb_pn = openpyxl.load_workbook(CONTROLE, read_only=True, data_only=True)
ws_pn = wb_pn['PAINEL']
painel_rows = []
for row in ws_pn.iter_rows(min_row=12, values_only=True):
    cpf = str(row[2]).strip() if len(row) > 2 and row[2] else ''
    if not cpf or cpf == 'None':
        continue
    painel_rows.append({
        'cpf': cpf,
        'nome': str(row[1]).strip() if row[1] else '',
        'painel_prestacao': float(row[16]) if row[16] is not None else 0,
    })
wb_pn.close()
print(f"  {len(painel_rows)} CPFs in PAINEL")

# Build output - match PAINEL order, show DB total side by side
print("Building output...")
wb = Workbook()

# Sheet 1: Ready to paste (CPF + Nome + DB Prestação) in PAINEL order
ws1 = wb.active
ws1.title = "PASTE_COL"
headers = ["CPF", "NOME", "PRESTAÇÃO DB (NOVO)"]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center')
    c.border = thin_border

r = 2
for p in painel_rows:
    cpf = p['cpf']
    db = db_data.get(cpf)
    nome = db['nome'] if db else p['nome']
    total = db['total'] if db else 0
    ws1.cell(row=r, column=1, value=cpf).border = thin_border
    ws1.cell(row=r, column=2, value=nome).border = thin_border
    c = ws1.cell(row=r, column=3, value=total)
    c.number_format = money_fmt
    c.border = thin_border
    r += 1

# Add CPFs in DB but not in PAINEL
extra = set(db_data.keys()) - set(p['cpf'] for p in painel_rows)
if extra:
    r += 1
    ws1.cell(row=r, column=1, value="--- CPFs no DB mas não no PAINEL ---").font = Font(bold=True, italic=True)
    r += 1
    for cpf in sorted(extra, key=lambda c: db_data[c]['nome']):
        d = db_data[cpf]
        ws1.cell(row=r, column=1, value=cpf).border = thin_border
        ws1.cell(row=r, column=2, value=d['nome']).border = thin_border
        c = ws1.cell(row=r, column=3, value=d['total'])
        c.number_format = money_fmt
        c.border = thin_border
        r += 1

ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 35
ws1.column_dimensions['C'].width = 22
ws1.freeze_panes = "A2"

# Sheet 2: Comparison side-by-side
ws2 = wb.create_sheet("COMPARACAO")
headers2 = ["CPF", "NOME", "PAINEL (R$)", "DB NOVO (R$)", "DIFERENÇA (R$)"]
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center')
    c.border = thin_border

r = 2
total_painel = 0
total_db = 0
for p in painel_rows:
    cpf = p['cpf']
    db = db_data.get(cpf)
    pn_val = p['painel_prestacao']
    db_val = db['total'] if db else 0
    diff = db_val - pn_val
    nome = db['nome'] if db else p['nome']
    total_painel += pn_val
    total_db += db_val

    ws2.cell(row=r, column=1, value=cpf).border = thin_border
    ws2.cell(row=r, column=2, value=nome).border = thin_border
    c = ws2.cell(row=r, column=3, value=pn_val); c.number_format = money_fmt; c.border = thin_border
    c = ws2.cell(row=r, column=4, value=db_val); c.number_format = money_fmt; c.border = thin_border
    c = ws2.cell(row=r, column=5, value=diff); c.number_format = money_fmt; c.border = thin_border
    if abs(diff) > 0.01:
        fill_color = "FFC7CE" if diff > 0 else "C6EFCE"
        c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    r += 1

# Totals row
c = ws2.cell(row=r, column=2, value="TOTAL"); c.font = Font(bold=True); c.border = thin_border
c = ws2.cell(row=r, column=3, value=total_painel); c.number_format = money_fmt; c.font = Font(bold=True); c.border = thin_border
c = ws2.cell(row=r, column=4, value=total_db); c.number_format = money_fmt; c.font = Font(bold=True); c.border = thin_border
c = ws2.cell(row=r, column=5, value=total_db - total_painel); c.number_format = money_fmt; c.font = Font(bold=True); c.border = thin_border

ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 18
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:E{r-1}"

print(f"Saving to {OUTPUT}...")
wb.save(OUTPUT)
print(f"Done! PAINEL total: R$ {total_painel:,.2f} | DB total: R$ {total_db:,.2f} | Diff: R$ {total_db - total_painel:+,.2f}")
