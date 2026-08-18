import openpyxl
import json
from datetime import datetime, timedelta
from collections import defaultdict

# Excel serial to date
def excel_serial_to_date(serial):
    if serial is None:
        return None
    if isinstance(serial, datetime):
        return serial.date()
    try:
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=float(serial))).date()
    except (ValueError, TypeError):
        return None

cutoff = datetime(2026, 7, 31).date()

wb = openpyxl.load_workbook('CONTROLE - VEXPENSES - AGOSTO 2026.xlsx', read_only=True, data_only=True)
ws = wb['EXTRATO']

# Sum per user with July 31 cutoff
extrato_sums = defaultdict(lambda: {'CARGA': 0, 'TRANSFERÊNCIA': 0, 'TARIFA': 0})
extrato_counts = defaultdict(lambda: {'CARGA': 0, 'TRANSFERÊNCIA': 0, 'TARIFA': 0})
total_by_month = defaultdict(lambda: defaultdict(int))

for row in ws.iter_rows(min_row=9, values_only=True):
    usuario = str(row[8]).strip().upper() if row[8] else ''
    tipo = str(row[9]).strip() if row[9] else ''
    valor = float(row[11]) if row[11] is not None else 0
    data_serial = row[3]  # col D = Data (Excel serial)
    mes = str(row[2]).strip() if row[2] else ''

    if not usuario or not tipo:
        continue

    row_date = excel_serial_to_date(data_serial)
    total_by_month[mes]['all'] += 1
    if row_date and row_date > cutoff:
        total_by_month[mes]['after_cutoff'] += 1
        continue  # Skip rows after cutoff

    # Normalize tipo
    tipo_upper = tipo.upper()
    if tipo_upper == 'CARGA':
        extrato_sums[usuario]['CARGA'] += valor
        extrato_counts[usuario]['CARGA'] += 1
    elif tipo_upper == 'TRANSFERÊNCIA':
        extrato_sums[usuario]['TRANSFERÊNCIA'] += abs(valor)
        extrato_counts[usuario]['TRANSFERÊNCIA'] += 1
    elif tipo_upper == 'TARIFA':
        extrato_sums[usuario]['TARIFA'] += abs(valor)
        extrato_counts[usuario]['TARIFA'] += 1

# Print month/cutoff summary
print("=== EXTRATO rows by month (with cutoff filtering) ===")
for mes in sorted(total_by_month.keys()):
    total = total_by_month[mes]['all']
    after = total_by_month[mes]['after_cutoff']
    before = total - after
    print(f"  {mes}: total={total}, before_cutoff={before}, after_cutoff={after}")
print()

# Read PAINEL for CPF mapping
ws_painel = wb['PAINEL']
painel_users = {}
for row in ws_painel.iter_rows(min_row=12, values_only=True):
    cpf = str(row[2]).strip() if row[2] else ''
    nome = str(row[1]).strip().upper() if row[1] else ''
    if not cpf or cpf == 'None':
        continue
    painel_users[cpf] = nome

# Read API data
with open('api_frozen.json', 'r', encoding='utf-8-sig') as f:
    api_raw = json.load(f)
api = {}
for row in api_raw:
    cpf = row['cpf'].strip()
    api[cpf] = row

# Compare with cutoff
for field_name, ext_key, api_key in [('carga', 'CARGA', 'carga'), ('transferencia', 'TRANSFERÊNCIA', 'transferencia'), ('tarifa', 'TARIFA', 'tarifa')]:
    print(f"=== {field_name.upper()} MISMATCHES (EXTRATO with July 31 cutoff vs API) ===")
    mismatches = []
    exact = 0
    for cpf, painel_name in painel_users.items():
        ext = extrato_sums.get(painel_name)
        api_row = api.get(cpf)
        if not ext or not api_row:
            continue
        ext_val = abs(ext[ext_key])
        api_val = api_row.get(api_key, 0)
        diff = api_val - ext_val
        if abs(diff) > 0.01:
            mismatches.append((painel_name[:35], cpf, ext_val, api_val, diff, extrato_counts[painel_name][ext_key]))
        else:
            exact += 1

    mismatches.sort(key=lambda x: abs(x[4]), reverse=True)
    print(f"  Exact matches: {exact}")
    print(f"  Mismatches: {len(mismatches)}")
    total_diff = sum(m[4] for m in mismatches)
    print(f"  Total diff: {total_diff:+.2f}")
    for name, cpf, ext_v, api_v, diff, cnt in mismatches[:15]:
        print(f"  {name} ({cpf}): EXTRATO={ext_v:.2f} ({cnt} rows), API={api_v:.2f}, diff={diff:+.2f}")
    print()
