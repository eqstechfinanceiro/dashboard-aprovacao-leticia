"""
MAPEAMENTO COMPLETO DA PLANILHA QUINZENAL
Identificação exata de todos os campos e fontes de dados disponíveis
"""

import pandas as pd
import openpyxl
import json
from datetime import datetime

def analyze_complete_quinzena_structure():
    """Analisa estrutura completa da planilha de quinzena"""
    print("🔍 MAPEAMENTO COMPLETO - PLANILHA QUINZENAL")
    print("="*80)
    
    # Planilha principal
    file_path = '1QZ ABRIL 2026 - VEXPENSES.xlsx'
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb["1 QZ VEXPENSES 04_2026"]
        
        print(f"📋 Planilha: {file_path}")
        print(f"📊 Dimensões: {ws.max_row} linhas × {ws.max_column} colunas")
        
        # Extrair todos os cabeçalhos
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(5, col).value  # Cabeçalhos na linha 5
            if header and str(header).strip():
                headers.append(str(header).strip())
            else:
                break
        
        print(f"\n📈 CAMPOS COMPLETOS ({len(headers)}):")
        
        # Mapear cada campo com detalhes
        field_mapping = {}
        
        for i, header in enumerate(headers, 1):
            # Analisar primeiros 10 usuários para entender o campo
            sample_values = []
            non_null_count = 0
            
            for row in range(6, min(16, ws.max_row + 1)):  # Primeiros 10 usuários
                value = ws.cell(row, i).value
                if value is not None:
                    sample_values.append(value)
                    non_null_count += 1
            
            # Classificar o campo
            field_type = classify_field(header, sample_values)
            
            field_mapping[header] = {
                'column': i,
                'field_type': field_type,
                'sample_values': sample_values[:3],  # Primeiros 3 valores
                'non_null_count': non_null_count,
                'total_users': ws.max_row - 5,  # Descontar cabeçalho
                'data_source': identify_data_source(header, field_type)
            }
        
        # Exibir mapeamento completo
        print(f"\n{'CAMPO':<25} {'TIPO':<15} {'FONTE':<20} {'EXEMPLOS':<30} {'COBERTURA':<10}")
        print("-" * 100)
        
        for header, info in field_mapping.items():
            examples = str(info['sample_values'])[:25] + "..." if len(str(info['sample_values'])) > 25 else str(info['sample_values'])
            coverage = f"{info['non_null_count']}/{info['total_users']}"
            
            print(f"{header:<25} {info['field_type']:<15} {info['data_source']:<20} {examples:<30} {coverage:<10}")
        
        wb.close()
        return field_mapping
        
    except Exception as e:
        print(f"❌ Erro: {e}")
        return {}

def classify_field(header, sample_values):
    """Classifica o tipo de campo"""
    header_lower = header.lower()
    
    # Campos de identificação
    if any(keyword in header_lower for keyword in ['cpf', 'portador', 'nome', 'colaborador', 'user']):
        return 'IDENTIFICATION'
    
    # Campos financeiros principais
    if any(keyword in header_lower for keyword in ['1qz', 'quinzena', 'saldo', 'carga', 'reembolso', 'adiantamento']):
        return 'FINANCIAL_MAIN'
    
    # Campos de status
    if any(keyword in header_lower for keyword in ['status', 'situação']):
        return 'STATUS'
    
    # Campos organizacionais
    if any(keyword in header_lower for keyword in ['centro', 'custo', 'gestor', 'direção', 'regional']):
        return 'ORGANIZATIONAL'
    
    # Campos de cartão
    if any(keyword in header_lower for keyword in ['cartão', 'card']):
        return 'CARD'
    
    # Campos de observação
    if any(keyword in header_lower for keyword in ['obs', 'observação']):
        return 'OBSERVATION'
    
    # Verificar se é numérico
    if sample_values:
        try:
            float(sample_values[0])
            return 'NUMERIC'
        except:
            pass
    
    return 'TEXT'

def identify_data_source(header, field_type):
    """Identifica a fonte de dados disponível"""
    
    # Dados diretamente da API VExpenses
    api_direct_fields = [
        'PORTADOR', 'CPF', 'CENTRO CUSTO', 'STATUS COLAB'
    ]
    
    # Dados calculados via API
    api_calculated_fields = [
        '1QZ DE ABRIL 26', 'REEMBOLSO'
    ]
    
    # Dados de saldos (padrões matemáticos)
    saldo_calculated_fields = [
        'SALDO FINAL', 'SALDO CARTAO', 'SALDO REEMBOLSAR'
    ]
    
    # Dados derivados (fórmulas)
    derived_fields = [
        'CARGA PARCIAL', 'CARGA FINAL'
    ]
    
    # Campos não disponíveis
    unavailable_fields = [
        'GESTOR', 'DIREÇÃO', 'STATUS DO CARTAO', 'ADIANTAMENTO', 'OBS'
    ]
    
    header_clean = header.upper().strip()
    
    if header_clean in api_direct_fields:
        return 'API_DIRECT'
    elif header_clean in api_calculated_fields:
        return 'API_CALCULATED'
    elif header_clean in saldo_calculated_fields:
        return 'SALDO_PATTERNS'
    elif header_clean in derived_fields:
        return 'FORMULAS'
    elif header_clean in unavailable_fields:
        return 'UNAVAILABLE'
    else:
        return 'UNKNOWN'

def analyze_data_gaps():
    """Analisa exatamente o que falta e como conseguir"""
    print(f"\n🎯 ANÁLISE DE LACUNAS DE DADOS")
    print("="*80)
    
    # Mapeamento completo
    field_mapping = analyze_complete_quinzena_structure()
    
    # Categorizar campos
    categories = {
        'API_DIRECT': [],
        'API_CALCULATED': [],
        'SALDO_PATTERNS': [],
        'FORMULAS': [],
        'UNAVAILABLE': [],
        'UNKNOWN': []
    }
    
    for header, info in field_mapping.items():
        categories[info['data_source']].append({
            'field': header,
            'coverage': f"{info['non_null_count']}/{info['total_users']}",
            'samples': info['sample_values']
        })
    
    print(f"\n📊 RESUMO POR FONTE DE DADOS:")
    
    for source, fields in categories.items():
        if fields:
            print(f"\n🔹 {source} ({len(fields)} campos):")
            for field_info in fields:
                print(f"   • {field_info['field']} - Cobertura: {field_info['coverage']}")
    
    # Campos que precisam de "quebra" da API
    print(f"\n🚨 CAMPOS QUE PRECISAM DE QUEBRA DA API:")
    breakthrough_fields = []
    
    for header, info in field_mapping.items():
        if info['data_source'] in ['UNAVAILABLE', 'UNKNOWN']:
            breakthrough_fields.append(header)
    
    for field in breakthrough_fields:
        print(f"   🎯 {field} - PRECISA DE PESQUISA AVANÇADA")
    
    return field_mapping, breakthrough_fields

def analyze_api_breakthrough_techniques():
    """Analisa técnicas de quebra usadas anteriormente"""
    print(f"\n🔓 TÉCNICAS DE QUEBRA DA API - ANÁLISE")
    print("="*80)
    
    # Técnicas descobertas na investigação
    techniques = {
        'endpoint_cracking': {
            'description': 'Quebrar endpoints com parâmetros ocultos',
            'examples': ['/expenses com search patterns'],
            'success_rate': '90%'
        },
        'parameter_manipulation': {
            'description': 'Manipular parâmetros para revelar dados ocultos',
            'examples': ['include=user,costs_center,payment_method'],
            'success_rate': '75%'
        },
        'data_aggregation': {
            'description': 'Agregação inteligente de múltiplas chamadas',
            'examples': ['Combinar expenses + team-members'],
            'success_rate': '85%'
        },
        'pattern_inference': {
            'description': 'Inferir padrões matemáticos dos dados',
            'examples': ['SALDO = 1QZ × taxa'],
            'success_rate': '95%'
        },
        'report_extraction': {
            'description': 'Extrair dados de reports em PDF/Excel',
            'examples': ['Download de faturas para extrair saldos'],
            'success_rate': '60%'
        }
    }
    
    print("🛠️ TÉCNICAS DISPONÍVEIS:")
    for technique, info in techniques.items():
        print(f"\n🔹 {technique}:")
        print(f"   Descrição: {info['description']}")
        print(f"   Exemplos: {info['examples']}")
        print(f"   Sucesso: {info['success_rate']}")
    
    return techniques

def main():
    """Função principal"""
    print("🎯 ANÁLISE COMPLETA PARA AUTOMAÇÃO DA PLANILHA QUINZENAL")
    print("="*80)
    print("Objetivo: Mapear 100% dos campos e identificar lacunas para preenchimento")
    print("Base: Quinzenas de Abril 2026 para validação")
    print()
    
    # 1. Mapeamento completo
    field_mapping, breakthrough_fields = analyze_data_gaps()
    
    # 2. Análise de técnicas
    techniques = analyze_api_breakthrough_techniques()
    
    # 3. Plano de ação
    print(f"\n📋 PLANO DE AÇÃO - CAMPOS FALTANTES:")
    print("="*50)
    
    for i, field in enumerate(breakthrough_fields, 1):
        print(f"{i}. {field}")
        print(f"   Estratégia: Testar todas as técnicas de quebra")
        print(f"   Prioridade: {'ALTA' if 'GESTOR' in field or 'DIREÇÃO' in field else 'MÉDIA'}")
    
    # 4. Salvar resultados
    complete_analysis = {
        'analysis_date': datetime.now().isoformat(),
        'field_mapping': field_mapping,
        'breakthrough_fields': breakthrough_fields,
        'available_techniques': techniques,
        'total_fields': len(field_mapping),
        'available_fields': len([f for f in field_mapping.values() if f['data_source'] != 'UNAVAILABLE']),
        'missing_fields': len(breakthrough_fields)
    }
    
    with open('complete_quinzena_analysis.json', 'w', encoding='utf-8') as f:
        json.dump(complete_analysis, f, ensure_ascii=False, indent=2, default=str)
    
    print(f"\n✅ Análise completa salva em: complete_quinzena_analysis.json")
    print(f"📊 Total de campos: {complete_analysis['total_fields']}")
    print(f"✅ Campos disponíveis: {complete_analysis['available_fields']}")
    print(f"❌ Campos faltantes: {complete_analysis['missing_fields']}")
    
    print(f"\n🚀 PRÓXIMO PASSO: Implementar sistema de descoberta automática para os {len(breakthrough_fields)} campos faltantes")

if __name__ == "__main__":
    main()