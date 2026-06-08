import pandas as pd
import openpyxl
import json
from datetime import datetime
import os

def analyze_saldo_columns_in_sheets():
    """Analisa colunas de SALDO em todas as planilhas para entender origem"""
    print("ANALISANDO ORIGEM DOS DADOS DE SALDO NAS PLANILHAS")
    print("="*60)
    
    # Diretório das planilhas
    sheets_dir = '/home/haumea/Projects/dashboard-aprovacao-leticia/data'
    
    # Planilhas disponíveis
    sheet_files = [
        '1QZ ABRIL 2026 - VEXPENSES (1).xlsx',
        'CONTROLE - VEXPENSES - ABRIL- 2026 (1).xlsb',
        'CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx'
    ]
    
    findings = {}
    
    for sheet_file in sheet_files:
        file_path = os.path.join(sheets_dir, sheet_file)
        
        if not os.path.exists(file_path):
            print(f"Arquivo não encontrado: {sheet_file}")
            continue
        
        print(f"\n📊 Analisando: {sheet_file}")
        print("-" * 40)
        
        try:
            if sheet_file.endswith('.xlsx'):
                findings[sheet_file] = analyze_xlsx_balance_file(file_path)
            elif sheet_file.endswith('.xlsb'):
                findings[sheet_file] = analyze_xlsb_file(file_path)
        except Exception as e:
            print(f"Erro ao analisar {sheet_file}: {e}")
            findings[sheet_file] = {'error': str(e)}
    
    return findings

def analyze_xlsx_file(file_path):
    """Analisa arquivo XLSB para dados de saldo"""
    print(f"Analisando XLSB: {os.path.basename(file_path)}")
    
    try:
        # Usar pandas para ler XLSB
        # Tentar ler todas as sheets
        xlsb_file = pd.ExcelFile(file_path)
        sheet_names = xlsb_file.sheet_names
        
        print(f"Sheets encontradas: {sheet_names}")
        
        findings = {'sheets': {}}
        
        for sheet_name in sheet_names:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                print(f"\nSheet: {sheet_name}")
                print(f"Dimensões: {df.shape}")
                print(f"Colunas: {list(df.columns)}")
                
                # Procurar colunas de saldo
                saldo_columns = []
                for col in df.columns:
                    if any(keyword in str(col).upper() for keyword in [
                        'SALDO', 'BALANCE', 'TOTAL', 'VALOR', 'CARTÃO', 'CARD'
                    ]):
                        saldo_columns.append(col)
                
                if saldo_columns:
                    print(f"Colunas de saldo encontradas: {saldo_columns}")
                    
                    # Analisar valores nessas colunas
                    sheet_findings = {
                        'shape': df.shape,
                        'saldo_columns': saldo_columns,
                        'sample_data': {}
                    }
                    
                    for col in saldo_columns:
                        if col in df.columns:
                            # Remover valores nulos e mostrar amostra
                            non_null_values = df[col].dropna()
                            
                            if len(non_null_values) > 0:
                                print(f"\n{col}:")
                                print(f"  Valores não nulos: {len(non_null_values)}")
                                
                                # Mostrar primeiros valores
                                for i, value in enumerate(non_null_values[:10]):
                                    if pd.notna(value):
                                        if isinstance(value, (int, float)):
                                            print(f"    {i+1}: R$ {value:.2f}")
                                        else:
                                            print(f"    {i+1}: {value}")
                                
                                # Verificar se são fórmulas ou valores fixos
                                unique_values = non_null_values.unique()
                                print(f"  Valores únicos: {len(unique_values)}")
                                
                                if len(unique_values) < 20:
                                    print(f"  Todos os valores: {list(unique_values)}")
                                
                                sheet_findings['sample_data'][col] = {
                                    'count': len(non_null_values),
                                    'unique_count': len(unique_values),
                                    'sample_values': list(non_null_values[:5])
                                }
                    
                    findings['sheets'][sheet_name] = sheet_findings
                else:
                    print("Nenhuma coluna de saldo encontrada")
                    findings['sheets'][sheet_name] = {'no_saldo_columns': True}
                
            except Exception as e:
                print(f"Erro ao ler sheet {sheet_name}: {e}")
                findings['sheets'][sheet_name] = {'error': str(e)}
        
        return findings
        
    except Exception as e:
        print(f"Erro geral ao analisar XLSB: {e}")
        return {'error': str(e)}

def analyze_xlsx_balance_file(file_path):
    """Analisa arquivo XLSX para dados de saldo"""
    print(f"Analisando XLSX: {os.path.basename(file_path)}")
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        findings = {'sheets': {}}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            print(f"\nSheet: {sheet_name}")
            print(f"Dimensões: {ws.max_row}x{ws.max_column}")
            
            # Procurar por células com "SALDO"
            saldo_cells = []
            
            for row in range(1, min(ws.max_row + 1, 100)):  # Primeiras 100 linhas
                for col in range(1, min(ws.max_column + 1, 20)):  # Primeiras 20 colunas
                    cell = ws.cell(row, col)
                    value = cell.value
                    
                    if value is not None:
                        # Procurar por células que contenham "SALDO"
                        if isinstance(value, str) and 'SALDO' in value.upper():
                            saldo_cells.append({
                                'cell': f"{row}{chr(64+col)}",
                                'content': value,
                                'row': row,
                                'col': col
                            })
                            
                            # Verificar células próximas (valores)
                            nearby_values = []
                            for dc in range(-2, 3):  # -2 to +2 columns
                                for dr in range(-2, 3):  # -2 to +2 rows
                                    if dc == 0 and dr == 0:
                                        continue
                                    
                                    nearby_row = row + dr
                                    nearby_col = col + dc
                                    
                                    if (1 <= nearby_row <= ws.max_row and 
                                        1 <= nearby_col <= ws.max_column):
                                        
                                        nearby_cell = ws.cell(nearby_row, nearby_col)
                                        nearby_value = nearby_cell.value
                                        
                                        if nearby_value is not None and isinstance(nearby_value, (int, float)):
                                            nearby_values.append({
                                                'cell': f"{nearby_row}{chr(64+nearby_col)}",
                                                'value': nearby_value
                                            })
                            
                            print(f"  {cell}: {value}")
                            if nearby_values:
                                print(f"    Valores próximos: {nearby_values[:3]}")
            
            # Procurar por valores numéricos altos (possíveis saldos)
            high_values = []
            
            for row in range(1, min(ws.max_row + 1, 200)):
                for col in range(1, min(ws.max_column + 1, 20)):
                    cell = ws.cell(row, col)
                    value = cell.value
                    
                    if isinstance(value, (int, float)) and abs(value) > 1000:  # Valores > 1000
                        # Verificar se é um saldo (positivo e razoável)
                        if 1000 <= abs(value) <= 50000:
                            high_values.append({
                                'cell': f"{row}{chr(64+col)}",
                                'value': value,
                                'row': row,
                                'col': col
                            })
            
            if high_values:
                print(f"Valores altos encontrados (possíveis saldos):")
                for hv in high_values[:10]:  # Primeiros 10
                    print(f"  {hv['cell']}: R$ {hv['value']:.2f}")
            
            findings['sheets'][sheet_name] = {
                'saldo_cells': saldo_cells,
                'high_values': high_values,
                'dimensions': f"{ws.max_row}x{ws.max_column}"
            }
        
        wb.close()
        return findings
        
    except Exception as e:
        print(f"Erro ao analisar XLSX: {e}")
        return {'error': str(e)}

def investigate_fatura_reports():
    """Investiga reports de FATURA para encontrar dados de saldo"""
    print("\nINVESTIGANDO REPORTS DE FATURA/CARTÃO")
    print("="*60)
    
    import requests
    
    API_KEY = "N2ntX8mUF7AvjVcyT6DZwnOQXRvYgel9pCirCGKKETO5R70HEZ8UdbQA0Nt8"
    BASE_URL = "https://api.vexpenses.com/v2"
    
    headers = {
        "Authorization": API_KEY,
        "Content-Type": "application/json",
        "Accept": "application/json"
    }
    
    try:
        url = f"{BASE_URL}/reports"
        params = {"paginate": "false", "per_page": 100}
        response = requests.get(url, headers=headers, params=params)
        
        if response.status_code == 200:
            data = response.json()
            if 'data' in data:
                reports = data['data']
                
                # Filtrar reports de FATURA
                fatura_reports = []
                for report in reports:
                    description = report.get('description', '').upper()
                    if 'FATURA' in description or 'CARTÃO' in description:
                        fatura_reports.append(report)
                
                print(f"Reports de FATURA/CARTÃO encontrados: {len(fatura_reports)}")
                
                # Analisar cada report
                for i, report in enumerate(fatura_reports[:5]):  # Primeiros 5
                    report_id = report.get('id')
                    description = report.get('description', '')
                    status = report.get('status', '')
                    
                    print(f"\nReport {i+1}:")
                    print(f"  ID: {report_id}")
                    print(f"  Description: {description}")
                    print(f"  Status: {status}")
                    
                    # Procurar campos financeiros
                    for key, value in report.items():
                        if key not in ['id', 'description', 'status', 'created_at', 'updated_at', 'user_id']:
                            if isinstance(value, (int, float)) and value > 0:
                                print(f"  {key}: R$ {value:.2f}")
                            elif isinstance(value, str) and len(value) < 100:
                                print(f"  {key}: {value}")
                
                return fatura_reports
                
    except Exception as e:
        print(f"Erro ao obter reports: {e}")
    
    return []

def main():
    """Função principal"""
    print("INVESTIGAÇÃO PROFUNDA - ORIGEM DOS DADOS DE SALDO")
    print("="*80)
    print("Analisando planilhas e reports de FATURA para encontrar fonte dos saldos")
    print("="*80)
    
    # 1. Analisar planilhas
    sheet_findings = analyze_saldo_columns_in_sheets()
    
    # 2. Investigar reports de FATURA
    fatura_reports = investigate_fatura_reports()
    
    # 3. Compilar resultados
    results = {
        'investigation_date': datetime.now().isoformat(),
        'sheet_analysis': sheet_findings,
        'fatura_reports': {
            'count': len(fatura_reports),
            'reports': fatura_reports[:5]  # Primeiros 5
        },
        'next_steps': [
            'Analisar arquivos Excel dos reports de FATURA',
            'Extrair dados de saldo dos relatórios',
            'Implementar cálculo preciso via API'
        ]
    }
    
    # Salvar resultados
    output_file = '/home/haumea/Projects/dashboard-aprovacao-leticia/investigation-docs/saldo_source_investigation.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    print(f"\n\nResultados salvos em: {output_file}")
    print("\n" + "="*80)
    print("INVESTIGAÇÃO CONCLUÍDA!")
    print("="*80)

if __name__ == "__main__":
    main()
