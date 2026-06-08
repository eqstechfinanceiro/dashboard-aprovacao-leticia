#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
🔍 ANÁLISE DE COLUNAS CALCULADAS NA BASE PREST
Identifica quais colunas são fórmulas vs dados da API
"""

import json
import os

def analisar_colunas_calculadas():
    """Analisa quais colunas da BASE PREST são calculadas"""
    
    print("🔍 ANALISANDO COLUNAS CALCULADAS NA BASE PREST")
    print("=" * 60)
    
    # Carregar arquivo de fórmulas
    caminho_formulas = '../formulas_controle.json' if os.path.exists('../formulas_controle.json') else 'formulas_controle.json'
    
    with open(caminho_formulas, 'r', encoding='utf-8') as f:
        formulas_data = json.load(f)
    
    # Encontrar BASE PREST
    base_prest_key = None
    for key in formulas_data.keys():
        if 'BASE PREST' in key.upper():
            base_prest_key = key
            break
    
    if not base_prest_key:
        print("❌ BASE PREST não encontrada no arquivo de fórmulas")
        return
    
    print(f"   ✅ Usando aba: {base_prest_key}")
    
    formulas = formulas_data[base_prest_key]
    
    # Mapear colunas com fórmulas
    colunas_com_formulas = set()
    
    for linha in formulas:
        if not linha:
            continue
        
        for celula in linha:
            if celula and isinstance(celula, dict) and 'formula' in celula:
                col = celula['col']
                colunas_com_formulas.add(col)
    
    print(f"\n📊 COLUNAS COM FÓRMULAS ({len(colunas_com_formulas)}):")
    for col in sorted(colunas_com_formulas):
        print(f"   Coluna {col}")
    
    # Carregar dados para obter nomes das colunas
    caminho_excel = '../sheets/CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsx'
    if not os.path.exists(caminho_excel):
        caminho_excel = 'sheets/CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsx'
    
    import openpyxl
    wb = openpyxl.load_workbook(caminho_excel, read_only=True, data_only=True)
    
    aba_nome = None
    for sheet_name in wb.sheetnames:
        if 'BASE PREST' in sheet_name.upper():
            aba_nome = sheet_name
            break
    
    ws = wb[aba_nome]
    
    # Encontrar cabeçalho
    cabecalho = None
    for row in ws.iter_rows(max_row=20):
        valores = [cell.value for cell in row]
        if any(v for v in valores if v and isinstance(v, str) and len(v) > 2):
            cabecalho = valores
            break
    
    print(f"\n📋 MAPEAMENTO COLUNA -> NOME:")
    colunas_calculadas_nomes = {}
    colunas_api_nomes = {}
    
    for i, col in enumerate(cabecalho):
        if col:
            if i in colunas_com_formulas:
                colunas_calculadas_nomes[i] = col
                print(f"   [CALCULADA] {i:2} - {col}")
            else:
                colunas_api_nomes[i] = col
                print(f"   [API]       {i:2} - {col}")
    
    print(f"\n📈 RESUMO:")
    print(f"   Total de colunas: {len([c for c in cabecalho if c])}")
    print(f"   Colunas calculadas: {len(colunas_calculadas_nomes)}")
    print(f"   Colunas da API: {len(colunas_api_nomes)}")
    
    wb.close()
    
    return colunas_calculadas_nomes, colunas_api_nomes

if __name__ == "__main__":
    analisar_colunas_calculadas()
