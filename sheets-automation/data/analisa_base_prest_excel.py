#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
🔍 ANÁLISE DA ABA BASE PREST (Excel direto)
Lê o arquivo Excel para obter os nomes corretos das colunas
"""

import openpyxl
import json
import random

def analisar_base_prest_excel():
    """Analisa a estrutura da aba BASE PREST do Excel"""
    
    print("🔍 ANALISANDO ABA BASE PREST (Excel)")
    print("=" * 60)
    
    # Carregar arquivo Excel
    caminho_excel = '../sheets/CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsx'
    
    try:
        wb = openpyxl.load_workbook(caminho_excel, read_only=True, data_only=True)
    except:
        # Tentar caminho alternativo
        caminho_excel = 'sheets/CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsx'
        wb = openpyxl.load_workbook(caminho_excel, read_only=True, data_only=True)
    
    # Encontrar aba BASE PREST
    aba_nome = None
    for sheet_name in wb.sheetnames:
        if 'BASE PREST' in sheet_name.upper():
            aba_nome = sheet_name
            break
    
    if not aba_nome:
        print("❌ Aba BASE PREST não encontrada")
        print(f"   Abas disponíveis: {wb.sheetnames}")
        return
    
    print(f"   ✅ Usando aba: {aba_nome}")
    
    ws = wb[aba_nome]
    
    # Ler cabeçalho (primeira linha não vazia)
    cabecalho = None
    row_idx = 0
    for row in ws.iter_rows(max_row=20):
        valores = [cell.value for cell in row]
        # Verificar se parece um cabeçalho (tem texto em várias colunas)
        if any(v for v in valores if v and isinstance(v, str) and len(v) > 2):
            cabecalho = valores
            row_idx = row[0].row
            print(f"   Cabeçalho encontrado na linha {row_idx}")
            break
    
    if not cabecalho:
        print("❌ Não foi possível encontrar cabeçalho")
        return
    
    print(f"\n📋 COLUNAS (linha {row_idx}):")
    for i, col in enumerate(cabecalho):
        if col:
            print(f"   {i:2}. {col}")
    
    # Contar total de linhas com dados
    total_linhas = 0
    for row in ws.iter_rows(min_row=row_idx+1):
        if any(cell.value for cell in row):
            total_linhas += 1
    
    print(f"\n📊 ESTRUTURA:")
    print(f"   Total de linhas com dados: {total_linhas}")
    print(f"   Total de colunas: {len(cabecalho)}")
    
    # Extrair amostra de 100 dados
    print(f"\n📦 EXTRAINDO AMOSTRA DE 100 DADOS...")
    
    linhas_dados = []
    for row in ws.iter_rows(min_row=row_idx+1, values_only=True):
        if any(v for v in row if v):
            linhas_dados.append(row)
    
    amostra = random.sample(linhas_dados, min(100, len(linhas_dados)))
    
    # Salvar amostra
    output_file = 'amostra_base_prest_100.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump({
            'aba': aba_nome,
            'cabecalho': cabecalho,
            'total_linhas': len(linhas_dados),
            'amostra_tamanho': len(amostra),
            'dados': amostra
        }, f, ensure_ascii=False, indent=2)
    
    print(f"   ✅ Amostra salva em: {output_file}")
    
    # Análise dos dados
    print(f"\n📈 ANÁLISE DOS DADOS:")
    
    for i, col in enumerate(cabecalho):
        if col:
            valores = [linha[i] if i < len(linha) else None for linha in amostra]
            valores_unicos = set(v for v in valores if v is not None and v != '')
            print(f"   {col:30} - {len(valores_unicos)} valores únicos")
    
    # Mostrar exemplos
    print(f"\n💡 EXEMPLOS DE DADOS (primeiros 3):")
    for i, linha in enumerate(amostra[:3], 1):
        print(f"\n   Exemplo {i}:")
        for j, (col, val) in enumerate(zip(cabecalho, linha)):
            if val and col:
                print(f"      {col:30}: {val}")
    
    wb.close()
    return cabecalho, amostra

if __name__ == "__main__":
    analisar_base_prest_excel()
