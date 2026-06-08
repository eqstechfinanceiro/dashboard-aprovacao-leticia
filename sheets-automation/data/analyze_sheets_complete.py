import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
import os
from datetime import datetime

def analyze_xlsx_sheet(sheet_name, file_path):
    """Analisa uma aba específica do arquivo XLSX incluindo fórmulas"""
    print(f"\n{'='*80}")
    print(f"ABA: {sheet_name}")
    print(f"{'='*80}")
    
    wb = load_workbook(file_path, data_only=False)
    ws = wb[sheet_name]
    
    analysis = {
        "sheet_name": sheet_name,
        "total_rows": ws.max_row,
        "total_cols": ws.max_column,
        "has_data": False,
        "headers": [],
        "sample_data": [],
        "formulas": [],
        "formula_cells": [],
        "value_cells": [],
        "column_analysis": {}
    }
    
    # Detectar linha de cabeçalho (primeira linha não vazia)
    header_row = None
    for row_idx in range(1, min(10, ws.max_row + 1)):
        row_values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        if any(v is not None for v in row_values):
            header_row = row_idx
            break
    
    if header_row:
        # Extrair cabeçalhos
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(header_row, col).value
            headers.append(str(cell_value) if cell_value else f"COL_{col}")
        analysis["headers"] = headers
        analysis["header_row"] = header_row
        
        # Analisar primeiras 10 linhas de dados
        data_start = header_row + 1
        for row_idx in range(data_start, min(data_start + 10, ws.max_row + 1)):
            row_data = {}
            has_formula = False
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row_idx, col_idx)
                if cell.value is not None:
                    row_data[header] = {
                        "value": cell.value,
                        "is_formula": cell.data_type == 'f',
                        "coordinate": cell.coordinate
                    }
                    if cell.data_type == 'f':
                        has_formula = True
                        analysis["formulas"].append({
                            "cell": cell.coordinate,
                            "formula": cell.value,
                            "header": header
                        })
                    else:
                        analysis["value_cells"].append({
                            "cell": cell.coordinate,
                            "value": cell.value,
                            "header": header
                        })
            if row_data:
                analysis["sample_data"].append(row_data)
                analysis["has_data"] = True
        
        # Análise por coluna
        for col_idx, header in enumerate(headers, 1):
            col_data = []
            col_formulas = []
            for row_idx in range(data_start, min(data_start + 50, ws.max_row + 1)):
                cell = ws.cell(row_idx, col_idx)
                if cell.value is not None:
                    if cell.data_type == 'f':
                        col_formulas.append(cell.value)
                    else:
                        col_data.append(cell.value)
            
            analysis["column_analysis"][header] = {
                "has_formulas": len(col_formulas) > 0,
                "formula_count": len(col_formulas),
                "sample_formulas": col_formulas[:5],
                "data_type": type(col_data[0]).__name__ if col_data else "unknown",
                "sample_values": col_data[:5]
            }
    
    wb.close()
    return analysis

def analyze_xlsb_sheet(sheet_name, file_path):
    """Analisa uma aba específica do arquivo XLSB (binário - sem fórmulas)"""
    print(f"\n{'='*80}")
    print(f"ABA: {sheet_name}")
    print(f"{'='*80}")
    
    # XLSB não suporta leitura de fórmulas com pandas/pyxlsb
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    
    analysis = {
        "sheet_name": sheet_name,
        "total_rows": len(df),
        "total_cols": len(df.columns),
        "has_data": False,
        "headers": [],
        "sample_data": [],
        "formulas": [],  # XLSB não expõe fórmulas facilmente
        "formula_cells": [],
        "column_analysis": {},
        "note": "XLSB format - formulas not accessible via standard libraries"
    }
    
    # Detectar linha de cabeçalho
    header_row = None
    for row_idx in range(min(10, len(df))):
        row_values = df.iloc[row_idx].dropna().tolist()
        if row_values:
            header_row = row_idx
            break
    
    if header_row is not None:
        headers = [str(v) if v is not None else f"COL_{i}" for i, v in enumerate(df.iloc[header_row])]
        analysis["headers"] = headers
        analysis["header_row"] = header_row
        
        # Primeiras 10 linhas de dados
        data_start = header_row + 1
        for row_idx in range(data_start, min(data_start + 10, len(df))):
            row_data = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(df.columns):
                    val = df.iloc[row_idx, col_idx]
                    if val is not None and pd.notna(val):
                        row_data[header] = {
                            "value": val,
                            "is_formula": False,
                            "coordinate": f"R{row_idx+1}C{col_idx+1}"
                        }
            if row_data:
                analysis["sample_data"].append(row_data)
                analysis["has_data"] = True
        
        # Análise por coluna
        for col_idx, header in enumerate(headers):
            if col_idx < len(df.columns):
                col_data = df.iloc[data_start:min(data_start+50, len(df)), col_idx].dropna().tolist()
                analysis["column_analysis"][header] = {
                    "has_formulas": False,
                    "formula_count": 0,
                    "sample_formulas": [],
                    "data_type": type(col_data[0]).__name__ if col_data else "unknown",
                    "sample_values": col_data[:5]
                }
    
    return analysis

def analyze_planilha1():
    """Analisa completamente a planilha 1QZ ABRIL 2026"""
    file_path = "/home/haumea/Projects/dashboard-aprovacao-leticia/data/1QZ ABRIL 2026 - VEXPENSES (1).xlsx"
    
    print("\n" + "="*80)
    print("ANÁLISE COMPLETA: PLANILHA 1 - 1QZ ABRIL 2026 - VEXPENSES (1).xlsx")
    print("="*80)
    
    wb = load_workbook(file_path, data_only=False)
    sheet_names = wb.sheetnames
    wb.close()
    
    print(f"\nTotal de abas: {len(sheet_names)}")
    print(f"Nomes das abas: {sheet_names}")
    
    full_analysis = {
        "file": "1QZ ABRIL 2026 - VEXPENSES (1).xlsx",
        "file_type": "XLSX",
        "total_sheets": len(sheet_names),
        "sheet_names": sheet_names,
        "sheets": {},
        "analysis_date": datetime.now().isoformat()
    }
    
    for sheet_name in sheet_names:
        try:
            sheet_analysis = analyze_xlsx_sheet(sheet_name, file_path)
            full_analysis["sheets"][sheet_name] = sheet_analysis
        except Exception as e:
            print(f"ERRO ao analisar aba {sheet_name}: {e}")
            full_analysis["sheets"][sheet_name] = {"error": str(e)}
    
    return full_analysis

def analyze_planilha2():
    """Analisa completamente a planilha CONTROLE"""
    file_path = "/home/haumea/Projects/dashboard-aprovacao-leticia/data/CONTROLE - VEXPENSES - ABRIL- 2026 (1).xlsb"
    
    print("\n" + "="*80)
    print("ANÁLISE COMPLETA: PLANILHA 2 - CONTROLE - VEXPENSES - ABRIL- 2026 (1).xlsb")
    print("="*80)
    
    # Ler com pandas para obter nomes das abas
    xls = pd.ExcelFile(file_path)
    sheet_names = xls.sheet_names
    xls.close()
    
    print(f"\nTotal de abas: {len(sheet_names)}")
    print(f"Nomes das abas: {sheet_names}")
    
    full_analysis = {
        "file": "CONTROLE - VEXPENSES - ABRIL- 2026 (1).xlsb",
        "file_type": "XLSB",
        "total_sheets": len(sheet_names),
        "sheet_names": sheet_names,
        "sheets": {},
        "analysis_date": datetime.now().isoformat()
    }
    
    for sheet_name in sheet_names:
        try:
            sheet_analysis = analyze_xlsb_sheet(sheet_name, file_path)
            full_analysis["sheets"][sheet_name] = sheet_analysis
        except Exception as e:
            print(f"ERRO ao analisar aba {sheet_name}: {e}")
            full_analysis["sheets"][sheet_name] = {"error": str(e)}
    
    return full_analysis

if __name__ == "__main__":
    # Analisar planilha 1
    analysis1 = analyze_planilha1()
    
    # Salvar resultado
    output1 = "/home/haumea/Projects/dashboard-aprovacao-leticia/investigation-docs/planilha1_complete_analysis.json"
    with open(output1, 'w', encoding='utf-8') as f:
        json.dump(analysis1, f, ensure_ascii=False, indent=2)
    print(f"\n\nAnálise da planilha 1 salva em: {output1}")
    
    # Analisar planilha 2
    analysis2 = analyze_planilha2()
    
    # Salvar resultado
    output2 = "/home/haumea/Projects/dashboard-aprovacao-leticia/investigation-docs/planilha2_complete_analysis.json"
    with open(output2, 'w', encoding='utf-8') as f:
        json.dump(analysis2, f, ensure_ascii=False, indent=2)
    print(f"\n\nAnálise da planilha 2 salva em: {output2}")
    
    print("\n" + "="*80)
    print("ANÁLISE CONCLUÍDA!")
    print("="*80)
