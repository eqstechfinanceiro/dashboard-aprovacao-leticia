import requests
import openpyxl
import json
from datetime import datetime
import os

API_KEY = "N2ntX8mUF7AvjVcyT6DZwnOQXRvYgel9pCirCGKKETO5R70HEZ8UdbQA0Nt8"
BASE_URL = "https://api.vexpenses.com/v2"

headers = {
    "Authorization": API_KEY,
    "Content-Type": "application/json",
    "Accept": "application/json"
}

def get_recent_reports():
    """Obtém reports mais recentes"""
    print("OBTENDO REPORTS MAIS RECENTES")
    print("="*60)
    
    params = {"paginate": "false", "per_page": 100}
    
    try:
        url = f"{BASE_URL}/reports"
        response = requests.get(url, headers=headers, params=params, timeout=30)
        
        if response.status_code == 200:
            data = response.json()
            if 'data' in data:
                reports = data['data']
                
                # Filtrar reports recentes (últimos 2 meses)
                current_date = datetime.now()
                recent_reports = []
                
                for report in reports:
                    created_at = report.get('created_at', '')
                    if created_at:
                        try:
                            report_date = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')
                            if (current_date - report_date).days <= 60:
                                recent_reports.append(report)
                        except:
                            continue
                
                print(f"Reports recentes: {len(recent_reports)}")
                return recent_reports
                
    except Exception as e:
        print(f"Erro: {e}")
    
    return []

def download_report_excel(report):
    """Baixa o arquivo Excel de um report"""
    excel_link = report.get('excel_link', '')
    
    if not excel_link:
        return None
    
    try:
        response = requests.get(excel_link, headers=headers, timeout=30)
        
        if response.status_code == 200:
            return response.content
    except Exception as e:
        print(f"Erro ao baixar: {e}")
    
    return None

def analyze_excel_for_saldo(excel_content, report_id):
    """Analisa arquivo Excel buscando dados de SALDO"""
    print(f"\nAnalisando Excel do report {report_id}")
    
    try:
        # Salvar arquivo temporariamente
        temp_file = f"/tmp/report_{report_id}.xlsx"
        with open(temp_file, 'wb') as f:
            f.write(excel_content)
        
        # Abrir com openpyxl
        wb = openpyxl.load_workbook(temp_file, data_only=True)
        
        print(f"Abas: {wb.sheetnames}")
        
        saldo_data = {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            print(f"\nAnalisando aba: {sheet_name}")
            print(f"Dimensões: {ws.max_row}x{ws.max_column}")
            
            # Procurar por cabeçalhos de SALDO
            saldo_headers = []
            
            for col in range(1, min(ws.max_column + 1, 30)):
                for row in range(1, min(ws.max_row + 1, 10)):
                    cell = ws.cell(row, col)
                    value = cell.value
                    
                    if value and isinstance(value, str):
                        value_upper = value.upper()
                        if 'SALDO' in value_upper:
                            saldo_headers.append({
                                'row': row,
                                'col': col,
                                'letter': chr(64 + col),
                                'header': value
                            })
            
            print(f"Cabeçalhos SALDO encontrados: {len(saldo_headers)}")
            
            # Extrair dados das colunas de SALDO
            for header_info in saldo_headers:
                col_letter = header_info['letter']
                header_row = header_info['row']
                header = header_info['header']
                
                print(f"\n{header} ({col_letter}):")
                
                values = []
                for row in range(header_row + 1, min(header_row + 20, ws.max_row + 1)):
                    cell = ws.cell(row, col_letter)
                    val = cell.value
                    
                    if val is not None and isinstance(val, (int, float)):
                        values.append(val)
                        print(f"  {col_letter}{row}: R$ {val:.2f}")
                
                if values:
                    saldo_data[f"{sheet_name}::{header}"] = values
        
        wb.close()
        
        # Remover arquivo temporário
        os.remove(temp_file)
        
        return saldo_data
        
    except Exception as e:
        print(f"Erro ao analisar Excel: {e}")
        import traceback
        traceback.print_exc()
        return {}

def analyze_all_numeric_values(excel_content, report_id):
    """Analisa TODOS os valores numéricos do Excel"""
    print(f"\nAnalisando TODOS os valores numéricos do report {report_id}")
    
    try:
        temp_file = f"/tmp/report_{report_id}.xlsx"
        with open(temp_file, 'wb') as f:
            f.write(excel_content)
        
        wb = openpyxl.load_workbook(temp_file, data_only=True)
        
        all_numeric_values = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            for row in range(1, min(ws.max_row + 1, 50)):
                for col in range(1, min(ws.max_column + 1, 30)):
                    cell = ws.cell(row, col)
                    val = cell.value
                    
                    if val is not None and isinstance(val, (int, float)) and abs(val) > 10:
                        all_numeric_values.append({
                            'sheet': sheet_name,
                            'cell': f"{chr(64 + col)}{row}",
                            'value': val
                        })
        
        wb.close()
        os.remove(temp_file)
        
        print(f"Total de valores numéricos: {len(all_numeric_values)}")
        
        # Mostrar os maiores valores
        all_numeric_values.sort(key=lambda x: abs(x['value']), reverse=True)
        
        print(f"\nTop 20 valores:")
        for i, v in enumerate(all_numeric_values[:20]):
            print(f"  {i+1}. {v['sheet']} - {v['cell']}: R$ {v['value']:.2f}")
        
        return all_numeric_values
        
    except Exception as e:
        print(f"Erro: {e}")
        return []

def search_target_values_in_excel(excel_content, report_id):
    """Procura pelos valores específicos da planilha"""
    print(f"\nProcurando valores alvo no report {report_id}")
    
    # Valores que procuramos
    target_values = [6945.16, 6626.04, 6504.20, -98.92, -428.82, 291.66, 18329.5, 20, 5, 1154.94]
    
    try:
        temp_file = f"/tmp/report_{report_id}.xlsx"
        with open(temp_file, 'wb') as f:
            f.write(excel_content)
        
        wb = openpyxl.load_workbook(temp_file, data_only=True)
        
        matches = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            for row in range(1, min(ws.max_row + 1, 100)):
                for col in range(1, min(ws.max_column + 1, 30)):
                    cell = ws.cell(row, col)
                    val = cell.value
                    
                    if val is not None and isinstance(val, (int, float)):
                        for target in target_values:
                            if abs(val - target) < 1:
                                matches.append({
                                    'sheet': sheet_name,
                                    'cell': f"{chr(64 + col)}{row}",
                                    'value': val,
                                    'target': target
                                })
                                
                                print(f"✅ ENCONTRADO: {val:.2f} em {sheet_name}!{chr(64 + col)}{row}")
        
        wb.close()
        os.remove(temp_file)
        
        print(f"\nTotal de correspondências: {len(matches)}")
        
        return matches
        
    except Exception as e:
        print(f"Erro: {e}")
        return []

def main():
    """Função principal"""
    print("BAIXANDO E ANALISANDO EXCELS DOS REPORTS")
    print("="*80)
    
    # 1. Obter reports recentes
    reports = get_recent_reports()
    
    if not reports:
        print("Nenhum report encontrado")
        return
    
    # 2. Filtrar reports com Excel
    excel_reports = [r for r in reports if r.get('excel_link')]
    
    print(f"Reports com Excel: {len(excel_reports)}")
    
    # 3. Analisar primeiros 5 reports
    results = {}
    
    for i, report in enumerate(excel_reports[:5]):
        report_id = report.get('id')
        description = report.get('description', '')
        
        print(f"\n{'='*60}")
        print(f"Report {i+1}/5: {description}")
        print(f"ID: {report_id}")
        print(f"{'='*60}")
        
        # Baixar Excel
        excel_content = download_report_excel(report)
        
        if excel_content:
            print(f"Excel baixado: {len(excel_content)} bytes")
            
            # Analisar por cabeçalhos SALDO
            saldo_data = analyze_excel_for_saldo(excel_content, report_id)
            
            # Analisar todos os valores numéricos
            numeric_values = analyze_all_numeric_values(excel_content, report_id)
            
            # Procurar valores alvo
            target_matches = search_target_values_in_excel(excel_content, report_id)
            
            results[f"report_{report_id}"] = {
                'description': description,
                'saldo_data': saldo_data,
                'numeric_values_count': len(numeric_values),
                'target_matches': target_matches
            }
        else:
            print("Erro ao baixar Excel")
    
    # 4. Salvar resultados
    output_file = '/home/haumea/Projects/dashboard-aprovacao-leticia/investigation-docs/report_excels_analysis.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    print(f"\n\nResultados salvos em: {output_file}")
    
    # 5. Conclusão
    total_matches = sum(len(r.get('target_matches', [])) for r in results.values())
    
    print("\n" + "="*80)
    print("CONCLUSÃO")
    print("="*80)
    
    if total_matches > 0:
        print(f"✅ ENCONTRADOS {total_matches} VALORES DE SALDO NOS EXCELS DOS REPORTS!")
        print("   Os dados de SALDO vieram dos arquivos Excel dos reports")
    else:
        print("⚠️  Valores de SALDO não encontrados nos Excels analisados")
        print("   Precisamos investigar mais reports ou outras fontes")

if __name__ == "__main__":
    main()
