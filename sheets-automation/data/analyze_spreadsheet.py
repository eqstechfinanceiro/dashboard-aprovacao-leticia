#!/usr/bin/env python3
"""
Script para analisar planilha Excel e identificar:
- O que é dado bruto vs calculado por fórmula
- Estrutura de cada aba
- Tipos de dados e dependências
"""

import json
import os
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


def analyze_workbook(filepath):
    """Analisa uma planilha Excel completa."""
    wb = openpyxl.load_workbook(filepath, data_only=False)
    
    analysis = {
        "arquivo": os.path.basename(filepath),
        "caminho_completo": str(filepath),
        "data_analise": datetime.now().isoformat(),
        "total_abas": len(wb.sheetnames),
        "abas": []
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_analysis = analyze_sheet(ws, sheet_name)
        analysis["abas"].append(sheet_analysis)
    
    return analysis


def analyze_sheet(ws, sheet_name):
    """Analisa uma aba específica da planilha."""
    print(f"Analisando aba: {sheet_name}...")
    
    # Limitar análise a um range razoável
    max_row = min(ws.max_row, 1000)  # Limitar a 1000 linhas
    max_col = min(ws.max_column, 100)   # Limitar a 100 colunas
    
    formulas = []
    raw_data = []
    headers = []
    
    # Identificar cabeçalhos (primeira linha com conteúdo)
    header_row = None
    for row in range(1, min(max_row + 1, 20)):  # Procurar nas primeiras 20 linhas
        has_content = False
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                has_content = True
                break
        if has_content:
            header_row = row
            break
    
    if header_row:
        for col in range(1, max_col + 1):
            cell = ws.cell(row=header_row, column=col)
            if cell.value is not None:
                headers.append({
                    "coluna": get_column_letter(col),
                    "valor": str(cell.value),
                    "tipo": type(cell.value).__name__,
                    "eh_formula": cell.value is not None and isinstance(cell.value, str) and cell.value.startswith('=')
                })
    
    # Analisar células
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue
            
            cell_ref = f"{get_column_letter(col)}{row}"
            cell_info = {
                "celula": cell_ref,
                "linha": row,
                "coluna": get_column_letter(col),
                "valor": str(cell.value)[:200] if cell.value is not None else None,  # Limitar tamanho
                "tipo_dado": type(cell.value).__name__,
            }
            
            # Verificar se é fórmula
            if isinstance(cell.value, str) and cell.value.startswith('='):
                cell_info["formula_original"] = cell.value
                cell_info["categoria"] = "formula"
                formulas.append(cell_info)
            else:
                cell_info["categoria"] = "dado_bruto"
                raw_data.append(cell_info)
    
    # Resumo estatístico
    summary = {
        "nome": sheet_name,
        "dimensao": f"{ws.max_row}x{ws.max_column}",
        "cabecalhos": headers,
        "total_celulas_com_dados": len(formulas) + len(raw_data),
        "total_formulas": len(formulas),
        "total_dados_brutos": len(raw_data),
        "formulas": formulas[:50],  # Limitar a 50 fórmulas para não gerar arquivo muito grande
        "dados_brutos_amostra": raw_data[:50],  # Amostra de dados brutos
    }
    
    return summary


def generate_markdown_report(analysis, output_path):
    """Gera relatório em formato Markdown."""
    lines = []
    lines.append(f"# Análise da Planilha: {analysis['arquivo']}")
    lines.append(f"**Data da análise:** {analysis['data_analise']}")
    lines.append(f"**Total de abas:** {analysis['total_abas']}")
    lines.append("")
    
    for sheet in analysis['abas']:
        lines.append(f"## Aba: {sheet['nome']}")
        lines.append(f"**Dimensão:** {sheet['dimensao']}")
        lines.append(f"**Total de células com dados:** {sheet['total_celulas_com_dados']}")
        lines.append(f"**Fórmulas:** {sheet['total_formulas']}")
        lines.append(f"**Dados brutos:** {sheet['total_dados_brutos']}")
        lines.append("")
        
        # Cabeçalhos
        if sheet['cabecalhos']:
            lines.append("### Cabeçalhos")
            for h in sheet['cabecalhos']:
                formula_marker = " (FÓRMULA)" if h['eh_formula'] else ""
                lines.append(f"- **{h['coluna']}**: {h['valor']}{formula_marker}")
            lines.append("")
        
        # Fórmulas
        if sheet['formulas']:
            lines.append("### Fórmulas Identificadas")
            for f in sheet['formulas']:
                lines.append(f"- **{f['celula']}**: `{f['formula_original']}`")
            lines.append("")
        
        # Dados brutos amostra
        if sheet['dados_brutos_amostra']:
            lines.append("### Amostra de Dados Brutos")
            for d in sheet['dados_brutos_amostra'][:20]:
                lines.append(f"- **{d['celula']}**: {d['valor']} (tipo: {d['tipo_dado']})")
            lines.append("")
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    
    print(f"Relatório Markdown gerado: {output_path}")


def main():
    # Caminho da planilha
    spreadsheet_path = Path(__file__).parent / "CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx"
    
    # Verificar se arquivo existe
    if not spreadsheet_path.exists():
        print(f"Erro: Arquivo não encontrado: {spreadsheet_path}")
        return
    
    print(f"Analisando: {spreadsheet_path}")
    
    # Analisar
    analysis = analyze_workbook(spreadsheet_path)
    
    # Salvar JSON completo
    json_path = spreadsheet_path.parent / "analise_planilha_completa.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(analysis, f, indent=2, ensure_ascii=False)
    print(f"JSON completo salvo: {json_path}")
    
    # Gerar relatório Markdown
    md_path = spreadsheet_path.parent / "analise_planilha_documentacao.md"
    generate_markdown_report(analysis, md_path)
    
    # Resumo
    print("\n" + "="*60)
    print("RESUMO DA ANÁLISE")
    print("="*60)
    print(f"Arquivo: {analysis['arquivo']}")
    print(f"Total de abas: {analysis['total_abas']}")
    for sheet in analysis['abas']:
        print(f"\nAba: {sheet['nome']}")
        print(f"  - Células com dados: {sheet['total_celulas_com_dados']}")
        print(f"  - Fórmulas: {sheet['total_formulas']}")
        print(f"  - Dados brutos: {sheet['total_dados_brutos']}")
    print("="*60)
    
    print("\nArquivos gerados com sucesso!")
    print(f"- JSON: {json_path}")
    print(f"- Markdown: {md_path}")


if __name__ == "__main__":
    main()
