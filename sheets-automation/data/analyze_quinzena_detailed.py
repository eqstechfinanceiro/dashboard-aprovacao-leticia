import pandas as pd
import openpyxl
import json
from datetime import datetime

print('ANALISANDO ESTRUTURA DETALHADA DA PLANILHA DE QUINZENA')
print('='*60)

# Analisar planilha CARGA 1 QZ MAIO 26
file_path = 'CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx'

try:
    # Primeiro tentar com openpyxl para ver a estrutura real
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f'Abas disponíveis: {wb.sheetnames}')
    
    # Analisar cada aba
    for sheet_name in wb.sheetnames:
        print(f'\n=== Analisando aba: {sheet_name} ===')
        ws = wb[sheet_name]
        print(f'Dimensões: {ws.max_row} linhas x {ws.max_column} colunas')
        
        # Procurar cabeçalhos nas primeiras 10 linhas
        headers_row = None
        for row_num in range(1, min(11, ws.max_row + 1)):
            row_data = []
            has_headers = False
            for col in range(1, min(20, ws.max_column + 1)):
                cell_value = ws.cell(row_num, col).value
                if cell_value and str(cell_value).strip():
                    row_data.append(str(cell_value).strip())
                    has_headers = True
                else:
                    row_data.append('')
            
            if has_headers and len([h for h in row_data if h]) > 5:  # Pelo menos 5 cabeçalhos
                headers_row = row_num
                print(f'Cabeçalhos encontrados na linha {row_num}:')
                for i, header in enumerate(row_data, 1):
                    if header:
                        print(f'  {i:2d}. {header}')
                break
        
        if headers_row:
            # Analisar dados abaixo dos cabeçalhos
            print(f'\nAnalisando dados a partir da linha {headers_row + 1}:')
            data_count = 0
            sample_rows = []
            
            for row_num in range(headers_row + 1, min(headers_row + 6, ws.max_row + 1)):
                row_data = {}
                has_data = False
                
                for col in range(1, len(row_data) + 1):
                    cell_value = ws.cell(row_num, col).value
                    if cell_value is not None:
                        header = ws.cell(headers_row, col).value
                        if header:
                            row_data[str(header)] = cell_value
                            has_data = True
                
                if has_data:
                    data_count += 1
                    if len(sample_rows) < 3:
                        sample_rows.append(row_data)
                    print(f'  Linha {row_num}: {json.dumps(row_data, indent=4, default=str)}')
            
            print(f'Total de linhas com dados analisadas: {data_count}')
        else:
            print('Nenhum cabeçalho encontrado nas primeiras 10 linhas')
            
            # Mostrar primeiras linhas como estão
            print('Primeiras 5 linhas do arquivo:')
            for row_num in range(1, min(6, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(10, ws.max_column + 1)):
                    cell_value = ws.cell(row_num, col).value
                    row_data.append(str(cell_value) if cell_value is not None else '')
                print(f'  Linha {row_num}: {row_data}')
    
    wb.close()
    
    # Tentar ler com pandas usando header=0 para comparar
    print(f'\n=== Tentando ler com pandas (header=0) ===')
    try:
        df = pd.read_excel(file_path, header=0)
        print(f'Colunas encontradas: {list(df.columns)}')
        print(f'Primeiras linhas:')
        print(df.head())
    except Exception as e:
        print(f'Erro com pandas: {e}')
    
    # Tentar ler sem header
    print(f'\n=== Tentando ler com pandas (sem header) ===')
    try:
        df = pd.read_excel(file_path, header=None)
        print(f'Dimensões: {df.shape}')
        print('Primeiras 5 linhas:')
        print(df.head())
    except Exception as e:
        print(f'Erro com pandas: {e}')

except Exception as e:
    print(f'Erro geral: {e}')
    import traceback
    traceback.print_exc()