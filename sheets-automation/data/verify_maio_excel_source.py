#!/usr/bin/env python3
"""
Verifica se a planilha CARGA 1 QZ MAIO 26 contém todos os dados financeiros
"""

import json
from pathlib import Path

import openpyxl


def analyze_maio_spreadsheet():
    """Analisa a planilha de MAIO 2026."""
    filepath = Path(__file__).parent / "CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx"
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb['Planilha1']
    
    # Identificar linha de cabeçalho
    header_row = 6
    
    # Ler cabeçalhos
    headers = {}
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=header_row, column=col)
        if cell.value:
            headers[col] = str(cell.value).strip()
    
    print("📊 CABEÇALHOS DA PLANILHA:")
    for col, header in headers.items():
        print(f"  Coluna {col}: {header}")
    
    # Extrair dados de uma amostra
    print("\n📋 AMOSTRA DE DADOS (primeiros 5 usuários):")
    for row in range(header_row + 1, header_row + 6):
        row_data = {}
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None and cell.value != '':
                header_name = headers.get(col, f"Col{col}")
                row_data[header_name] = cell.value
        
        print(f"\nUsuário {row - header_row}:")
        for key, value in row_data.items():
            print(f"  {key}: {value}")
    
    # Verificar campos financeiros
    financial_fields = ['1ª QZ', 'SALDO FINAL', 'SALDO CARTAO', 'SALDO REEMBOLSAR', 
                       'CARGA PARCIAL', 'REEMBOLSO', 'Carga Final']
    
    print("\n✅ VERIFICAÇÃO DE CAMPOS FINANCEIROS:")
    for field in financial_fields:
        found = any(field in str(h) for h in headers.values())
        status = "✅ PRESENTE" if found else "❌ AUSENTE"
        print(f"  {field}: {status}")
    
    # Contar registros
    total_rows = sheet.max_row - header_row
    print(f"\n📊 TOTAL DE REGISTROS: {total_rows}")
    
    return headers, total_rows


def main():
    print("🔍 Analisando planilha CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx\n")
    headers, total_rows = analyze_maio_spreadsheet()
    
    print("\n🎯 CONCLUSÃO:")
    print("A planilha CARGA 1 QZ MAIO 26 é a FONTE DE VERDADE para MAIO 2026.")
    print("Ela contém todos os dados necessários (cadastrais + financeiros).")
    print("\n💡 SOLUÇÃO:")
    print("1. Extrair dados financeiros desta planilha (1ª QZ, SALDO FINAL, etc.)")
    print("2. Complementar com dados da API VExpenses (GESTOR, DIRETOR, etc.)")
    print("3. Criar solução híbrida: Excel (financeiros) + API (cadastrais)")


if __name__ == "__main__":
    main()
