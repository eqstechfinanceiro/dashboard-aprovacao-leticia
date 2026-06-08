#!/usr/bin/env python3
"""
Investiga o padrão de geração de arquivos por quinzena
e tenta identificar a fonte original dos dados financeiros
"""

import json
from pathlib import Path

import openpyxl


def analyze_abril_files():
    """Analisa os arquivos de ABRIL 2026 para entender o padrão."""
    print("🔍 Analisando arquivos de ABRIL 2026\n")
    
    # Arquivo 1QZ ABRIL 2026
    filepath1 = Path(__file__).parent / "1QZ ABRIL 2026 - VEXPENSES.xlsx"
    if filepath1.exists():
        print("📁 Arquivo: 1QZ ABRIL 2026 - VEXPENSES.xlsx")
        wb1 = openpyxl.load_workbook(filepath1, data_only=True)
        print(f"  Abas: {list(wb1.sheetnames)}")
        
        # Verificar aba principal
        if '1 QZ VEXPENSES 04_2026' in wb1.sheetnames:
            sheet = wb1['1 QZ VEXPENSES 04_2026']
            print(f"  Linhas: {sheet.max_row}")
            print(f"  Colunas: {sheet.max_column}")
            
            # Ler cabeçalhos
            headers = []
            for col in range(1, min(20, sheet.max_column + 1)):
                cell = sheet.cell(row=1, column=col)
                if cell.value:
                    headers.append(str(cell.value))
            print(f"  Campos: {headers[:10]}...")
    
    # Arquivo CONTROLE
    filepath2 = Path(__file__).parent / "CONTROLE - VEXPENSES - ABRIL- 2026.xlsb"
    if filepath2.exists():
        print("\n📁 Arquivo: CONTROLE - VEXPENSES - ABRIL- 2026.xlsb")
        try:
            wb2 = openpyxl.load_workbook(filepath2, data_only=True)
            print(f"  Abas: {list(wb2.sheetnames)}")
            
            # Verificar aba SALDO CARTAO
            if 'SALDO CARTAO' in wb2.sheetnames:
                sheet = wb2['SALDO CARTAO']
                print(f"  Aba SALDO CARTAO: {sheet.max_row} linhas")
                
                # Amostra de dados
                print("  Amostra de dados:")
                for row in range(1, min(6, sheet.max_row + 1)):
                    row_data = []
                    for col in range(1, min(13, sheet.max_column + 1)):
                        cell = sheet.cell(row=row, column=col)
                        if cell.value is not None:
                            row_data.append(str(cell.value)[:20])
                    print(f"    Linha {row}: {row_data[:5]}")
        except Exception as e:
            print(f"  Erro ao ler: {e}")


def analyze_maio_files():
    """Analisa os arquivos de MAIO 2026."""
    print("\n🔍 Analisando arquivos de MAIO 2026\n")
    
    # Arquivo CARGA
    filepath = Path(__file__).parent / "CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx"
    if filepath.exists():
        print("📁 Arquivo: CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx")
        wb = openpyxl.load_workbook(filepath, data_only=True)
        print(f"  Abas: {list(wb.sheetnames)}")
        
        sheet = wb['Planilha1']
        print(f"  Linhas: {sheet.max_row}")
        print(f"  Colunas: {sheet.max_column}")


def check_for_maio_controle():
    """Verifica se existe arquivo CONTROLE para MAIO."""
    print("\n🔍 Verificando se existe CONTROLE para MAIO 2026\n")
    
    maio_files = list(Path(__file__).parent.glob("*MAIO*"))
    print(f"Arquivos com MAIO no nome: {len(maio_files)}")
    for f in maio_files:
        print(f"  - {f.name}")


def analyze_pattern():
    """Analisa o padrão de geração de arquivos."""
    print("\n🎯 ANÁLISE DO PADRÃO\n")
    
    print("Para ABRIL 2026:")
    print("  - Existe arquivo separado: 1QZ ABRIL 2026 - VEXPENSES.xlsx")
    print("  - Existe arquivo CONTROLE: CONTROLE - VEXPENSES - ABRIL- 2026.xlsb")
    print("  - Provavelmente o CONTROLE é a FONTE ORIGINAL dos dados de saldo")
    
    print("\nPara MAIO 2026:")
    print("  - NÃO existe arquivo separado 1QZ")
    print("  - NÃO existe arquivo CONTROLE")
    print("  - A planilha CARGA já contém todos os dados")
    
    print("\n🤔 HIPÓTESES:")
    print("1. O arquivo CONTROLE é gerado por um sistema externo (banco/financeiro)")
    print("2. A planilha CARGA é preenchida manualmente usando dados do CONTROLE")
    print("3. Para MAIO, alguém já consolidou tudo na planilha CARGA")
    print("4. Para quinzenas futuras, precisamos:")
    print("   a) Do arquivo CONTROLE do período, OU")
    print("   b) Integrar com o sistema que gera o CONTROLE, OU")
    print("   c) Descobrir se a API VExpenses tem endpoint não documentado")


def main():
    analyze_abril_files()
    analyze_maio_files()
    check_for_maio_controle()
    analyze_pattern()
    
    print("\n💡 PRÓXIMOS PASSOS SUGERIDOS:")
    print("1. Contatar quem gera o arquivo CONTROLE - VEXPENSES")
    print("2. Perguntar se existe API ou integração automatizada")
    print("3. Verificar se é possível receber dados via webhook/integração")
    print("4. Investigar se a VExpenses tem endpoint não documentado para saldos")
    print("5. Enquanto isso, usar abordagem híbrida: Excel + API")


if __name__ == "__main__":
    main()
