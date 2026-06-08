#!/usr/bin/env python3
"""
Replica o processo de geração do CONTROLE
Baixa reports da API VExpenses e extrai dados dos arquivos Excel
"""

import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl
import requests

API_KEY = "N2ntX8mUF7AvjVcyT6DZwnOQXRvYgel9pCirCGKKETO5R70HEZ8UdbQA0Nt8"
BASE_URL = "https://api.vexpenses.com/v2"

headers = {
    "Authorization": API_KEY,
    "Content-Type": "application/json",
    "Accept": "application/json"
}


def get_reports_by_period(year, month):
    """Obtém reports de um período específico"""
    print(f"📊 Obtendo reports de {month}/{year}...")
    
    try:
        url = f"{BASE_URL}/reports"
        params = {"paginate": "false", "per_page": 100}
        response = requests.get(url, headers=headers, params=params, timeout=30)
        
        if response.status_code == 200:
            data = response.json()
            if 'data' in data:
                reports = data['data']
                
                # Filtrar reports do período
                period_reports = []
                for report in reports:
                    created_at = report.get('created_at', '')
                    if created_at:
                        try:
                            report_date = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')
                            if report_date.year == year and report_date.month == month:
                                period_reports.append(report)
                        except:
                            continue
                
                print(f"✅ Reports encontrados: {len(period_reports)}")
                return period_reports
                
    except Exception as e:
        print(f"❌ Erro ao obter reports: {e}")
    
    return []


def download_report_excel(report):
    """Baixa o arquivo Excel de um report"""
    report_id = report.get('id')
    excel_link = report.get('excel_link', '')
    
    if not excel_link:
        return None
    
    try:
        response = requests.get(excel_link, headers=headers, timeout=30)
        
        if response.status_code == 200:
            return response.content
    except Exception as e:
        print(f"❌ Erro ao baixar Excel do report {report_id}: {e}")
    
    return None


def extract_1qz_from_excel(excel_content):
    """Extrai valor de 1QZ do arquivo Excel"""
    try:
        # Verificar se é um arquivo Excel válido
        if not excel_content or len(excel_content) < 100:
            return None
        
        # Verificar magic bytes do Excel
        if excel_content[:2] != b'PK':
            print(f"    ⚠️ Arquivo não é Excel válido")
            return None
        
        # Salvar arquivo temporariamente
        temp_file = Path(__file__).parent / f"temp_report.xlsx"
        with open(temp_file, 'wb') as f:
            f.write(excel_content)
        
        wb = openpyxl.load_workbook(temp_file, data_only=True)
        
        # Procurar por 1QZ em todas as células
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            for row in range(1, min(ws.max_row + 1, 50)):
                for col in range(1, min(ws.max_column + 1, 20)):
                    cell = ws.cell(row, col)
                    value = cell.value
                    
                    if value and isinstance(value, str):
                        # Procurar por padrões de 1QZ
                        if '1QZ' in value.upper() or '1ª QZ' in value.upper() or '1 QZ' in value.upper():
                            # Verificar células adjacentes para o valor
                            for c in range(col, min(col + 5, ws.max_column + 1)):
                                adj_cell = ws.cell(row, c)
                                if isinstance(adj_cell.value, (int, float)) and adj_cell.value > 0:
                                    wb.close()
                                    temp_file.unlink()
                                    return adj_cell.value
        
        wb.close()
        temp_file.unlink()
        
    except Exception as e:
        print(f"❌ Erro ao extrair 1QZ: {e}")
    
    return None


def extract_saldo_cartao_from_excel(excel_content):
    """Extrai saldo de cartão do arquivo Excel"""
    try:
        temp_file = Path(__file__).parent / f"temp_report.xlsx"
        with open(temp_file, 'wb') as f:
            f.write(excel_content)
        
        wb = openpyxl.load_workbook(temp_file, data_only=True)
        
        # Procurar por saldo de cartão
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            for row in range(1, min(ws.max_row + 1, 50)):
                for col in range(1, min(ws.max_column + 1, 20)):
                    cell = ws.cell(row, col)
                    value = cell.value
                    
                    if value and isinstance(value, str):
                        # Procurar por padrões de saldo cartão
                        if 'SALDO' in value.upper() and 'CARTÃO' in value.upper():
                            # Verificar células adjacentes para o valor
                            for c in range(col, min(col + 5, ws.max_column + 1)):
                                adj_cell = ws.cell(row, c)
                                if isinstance(adj_cell.value, (int, float)):
                                    wb.close()
                                    temp_file.unlink()
                                    return adj_cell.value
        
        wb.close()
        temp_file.unlink()
        
    except Exception as e:
        print(f"❌ Erro ao extrair saldo cartão: {e}")
    
    return None


def extract_financial_data_from_reports(reports):
    """Extrai dados financeiros de múltiplos reports"""
    print(f"\n🔍 Extraindo dados financeiros de {len(reports)} reports...")
    
    financial_data = {
        'quinzenas': {},
        'saldo_cartao': {},
        'adicionais': {}
    }
    
    for i, report in enumerate(reports):
        report_id = report.get('id')
        user_id = report.get('user_id')
        description = report.get('description', '')
        
        print(f"  [{i+1}/{len(reports)}] Report {report_id}: {description[:50]}")
        
        # Baixar Excel
        excel_content = download_report_excel(report)
        if not excel_content:
            continue
        
        # Extrair 1QZ
        qz_value = extract_1qz_from_excel(excel_content)
        if qz_value and user_id:
            if user_id not in financial_data['quinzenas']:
                financial_data['quinzenas'][user_id] = []
            financial_data['quinzenas'][user_id].append({
                'report_id': report_id,
                'value': qz_value,
                'description': description
            })
            print(f"    ✅ 1QZ: {qz_value}")
        
        # Extrair saldo cartão
        saldo_value = extract_saldo_cartao_from_excel(excel_content)
        if saldo_value and user_id:
            if user_id not in financial_data['saldo_cartao']:
                financial_data['saldo_cartao'][user_id] = []
            financial_data['saldo_cartao'][user_id].append({
                'report_id': report_id,
                'value': saldo_value,
                'description': description
            })
            print(f"    ✅ Saldo Cartão: {saldo_value}")
    
    return financial_data


def main():
    print("🎯 REPLICAÇÃO DO PROCESSO DE GERAÇÃO DO CONTROLE\n")
    print("=" * 60)
    
    # Definir período (ABRIL 2026)
    year = 2026
    month = 4
    
    # Obter reports do período
    reports = get_reports_by_period(year, month)
    
    if not reports:
        print("❌ Nenhum report encontrado para o período")
        return
    
    # Extrair dados financeiros
    financial_data = extract_financial_data_from_reports(reports)
    
    # Salvar resultados
    output_path = Path(__file__).parent / "financial_data_from_reports.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(financial_data, f, indent=2, ensure_ascii=False)
    
    print(f"\n📁 Dados salvos: {output_path}")
    
    print(f"\n📊 Resumo:")
    print(f"  Quinzenas extraídas: {len(financial_data['quinzenas'])} usuários")
    print(f"  Saldo cartão extraído: {len(financial_data['saldo_cartao'])} usuários")
    
    print("\n💡 PRÓXIMOS PASSOS:")
    print("1. Mapear user_id para CPF")
    print("2. Cruzar com dados do CONTROLE para validar")
    print("3. Refinar extração para capturar todos os campos")
    print("4. Automatizar processo para quinzenas futuras")


if __name__ == "__main__":
    main()
