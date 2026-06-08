import json
import openpyxl
from datetime import datetime

def extract_all_saldo_data():
    """Extrai todos os dados de SALDO da planilha"""
    print("EXTRAINDO TODOS OS DADOS DE SALDO")
    print("="*60)
    
    file_path = '1QZ ABRIL 2026 - VEXPENSES.xlsx'
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb["1 QZ VEXPENSES 04_2026"]
        
        all_data = []
        
        for row in range(6, min(100, ws.max_row + 1)):
            nome = ws.cell(row, 2).value
            saldo_reembolsar = ws.cell(row, 9).value
            saldo_final = ws.cell(row, 10).value
            quinzena_qz = ws.cell(row, 11).value
            saldo_cartao = ws.cell(row, 12).value
            
            if nome and isinstance(nome, str):
                # Converter valores
                def safe_float(val):
                    try:
                        if val is None:
                            return 0.0
                        if isinstance(val, (int, float)):
                            return float(val)
                        if isinstance(val, str):
                            clean = ''.join(c for c in val if c.isdigit() or c == '.' or c == '-')
                            if clean:
                                return float(clean)
                        return 0.0
                    except:
                        return 0.0
                
                data = {
                    'nome': nome,
                    'saldo_reembolsar': safe_float(saldo_reembolsar),
                    'saldo_final': safe_float(saldo_final),
                    'quinzena_qz': safe_float(quinzena_qz),
                    'saldo_cartao': safe_float(saldo_cartao),
                    'row': row
                }
                
                # Incluir apenas se tiver dados relevantes
                if data['quinzena_qz'] > 0:
                    all_data.append(data)
        
        wb.close()
        
        print(f"Total de usuários com dados: {len(all_data)}")
        
        return all_data
        
    except Exception as e:
        print(f"Erro: {e}")
        import traceback
        traceback.print_exc()
        return []

def analyze_saldo_patterns(all_data):
    """Analisa padrões matemáticos nos dados de SALDO"""
    print(f"\nANALISANDO PADRÕES MATEMÁTICOS")
    print("="*60)
    
    # Filtrar usuários com dados relevantes
    relevant_data = [d for d in all_data if d['quinzena_qz'] > 100]
    
    print(f"Usuários com 1QZ > 100: {len(relevant_data)}")
    
    if not relevant_data:
        print("Nenhum dado relevante encontrado")
        return None
    
    # Análise 1: SALDO FINAL vs 1QZ
    print(f"\nANÁLISE 1: SALDO FINAL vs 1QZ")
    saldo_final_ratios = []
    
    for data in relevant_data:
        if data['quinzena_qz'] > 0 and data['saldo_final'] != 0:
            ratio = data['saldo_final'] / data['quinzena_qz']
            saldo_final_ratios.append(ratio)
    
    if saldo_final_ratios:
        avg_ratio = sum(saldo_final_ratios) / len(saldo_final_ratios)
        print(f"  Razão média SALDO FINAL / 1QZ: {avg_ratio:.4f}")
        print(f"  Exemplos:")
        for i, data in enumerate(relevant_data[:3]):
            if data['quinzena_qz'] > 0:
                ratio = data['saldo_final'] / data['quinzena_qz'] if data['saldo_final'] != 0 else 0
                print(f"    {data['nome']}: {data['saldo_final']:.2f} / {data['quinzena_qz']:.2f} = {ratio:.4f}")
    
    # Análise 2: SALDO CARTAO vs 1QZ
    print(f"\nANÁLISE 2: SALDO CARTAO vs 1QZ")
    saldo_cartao_ratios = []
    
    for data in relevant_data:
        if data['quinzena_qz'] > 0 and data['saldo_cartao'] != 0:
            ratio = data['saldo_cartao'] / data['quinzena_qz']
            saldo_cartao_ratios.append(ratio)
    
    if saldo_cartao_ratios:
        avg_ratio = sum(saldo_cartao_ratios) / len(saldo_cartao_ratios)
        print(f"  Razão média SALDO CARTAO / 1QZ: {avg_ratio:.4f}")
        print(f"  Exemplos:")
        for i, data in enumerate(relevant_data[:3]):
            if data['quinzena_qz'] > 0 and data['saldo_cartao'] != 0:
                ratio = data['saldo_cartao'] / data['quinzena_qz']
                print(f"    {data['nome']}: {data['saldo_cartao']:.2f} / {data['quinzena_qz']:.2f} = {ratio:.4f}")
    
    # Análise 3: SALDO REEMBOLSAR vs 1QZ
    print(f"\nANÁLISE 3: SALDO REEMBOLSAR vs 1QZ")
    saldo_reembolsar_ratios = []
    
    for data in relevant_data:
        if data['quinzena_qz'] > 0 and data['saldo_reembolsar'] != 0:
            ratio = abs(data['saldo_reembolsar']) / data['quinzena_qz']
            saldo_reembolsar_ratios.append(ratio)
    
    if saldo_reembolsar_ratios:
        avg_ratio = sum(saldo_reembolsar_ratios) / len(saldo_reembolsar_ratios)
        print(f"  Razão média |SALDO REEMBOLSAR| / 1QZ: {avg_ratio:.4f}")
        print(f"  Exemplos:")
        for i, data in enumerate(relevant_data[:3]):
            if data['quinzena_qz'] > 0 and data['saldo_reembolsar'] != 0:
                ratio = abs(data['saldo_reembolsar']) / data['quinzena_qz']
                print(f"    {data['nome']}: {abs(data['saldo_reembolsar']):.2f} / {data['quinzena_qz']:.2f} = {ratio:.4f}")
    
    return {
        'saldo_final_avg_ratio': sum(saldo_final_ratios) / len(saldo_final_ratios) if saldo_final_ratios else 0,
        'saldo_cartao_avg_ratio': sum(saldo_cartao_ratios) / len(saldo_cartao_ratios) if saldo_cartao_ratios else 0,
        'saldo_reembolsar_avg_ratio': sum(saldo_reembolsar_ratios) / len(saldo_reembolsar_ratios) if saldo_reembolsar_ratios else 0,
        'relevant_users': len(relevant_data)
    }

def main():
    """Função principal"""
    print("DESCOBRINDO PADRÕES DE SALDO - VERSÃO WINDOWS")
    print("="*80)
    
    # 1. Extrair todos os dados
    all_data = extract_all_saldo_data()
    
    if not all_data:
        print("Nenhum dado extraído")
        return
    
    # 2. Analisar padrões
    patterns = analyze_saldo_patterns(all_data)
    
    if patterns:
        # 3. Salvar resultados
        results = {
            'investigation_date': datetime.now().isoformat(),
            'total_users': len(all_data),
            'patterns': patterns,
            'conclusion': 'PADRÕES MATEMÁTICOS ENCONTRADOS' if patterns['saldo_final_avg_ratio'] > 0 else 'NENHUM PADRÃO CLARO ENCONTRADO'
        }
        
        print(f"\n✅ RESUMO DOS PADRÕES ENCONTRADOS:")
        print(f"   SALDO FINAL = 1QZ * {patterns['saldo_final_avg_ratio']:.4f}")
        print(f"   SALDO CARTAO = 1QZ * {patterns['saldo_cartao_avg_ratio']:.4f}")
        print(f"   SALDO REEMBOLSAR = 1QZ * {patterns['saldo_reembolsar_avg_ratio']:.4f}")
        print(f"   USUÁRIOS RELEVANTES: {patterns['relevant_users']}")
        
        output_file = 'saldo_patterns_windows.json'
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        
        print(f"\nSalvo em: {output_file}")
    
    print("\n" + "="*80)
    print("INVESTIGAÇÃO CONCLUÍDA!")
    print("="*80)

if __name__ == "__main__":
    main()