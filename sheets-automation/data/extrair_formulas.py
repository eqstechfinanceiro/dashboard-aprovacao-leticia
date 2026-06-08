import json
from openpyxl import load_workbook
import os

def extrair_formulas_excel():
    """Extrai fórmulas do arquivo Excel de controle"""
    
    print("🔍 EXTRAINDO FÓRMULAS DO EXCEL")
    print("=" * 60)
    
    # Caminho do arquivo Excel (deve ser .xlsx para preservar fórmulas)
    xlsx_path = 'sheets/CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsx'
    
    if not os.path.exists(xlsx_path):
        print(f"❌ Arquivo não encontrado: {xlsx_path}")
        print(f"\n⚠️  O arquivo .xlsb não preserva fórmulas quando convertido via Python.")
        print(f"   Por favor, abra o arquivo 'CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsb' no Excel")
        print(f"   e salve como 'CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsx' na pasta sheets/")
        return None
    
    try:
        # Carregar workbook .xlsx com openpyxl em modo read_only (evita pivot tables)
        import warnings
        warnings.filterwarnings('ignore', category=UserWarning)
        wb = load_workbook(xlsx_path, data_only=False, read_only=True, keep_vba=False)
        
        formulas_data = {}
        
        for sheet_name in wb.sheetnames:
            print(f"\n📄 Processando aba: {sheet_name}")
            ws = wb[sheet_name]
            
            sheet_formulas = []
            
            # Percorrer todas as células
            for row in ws.iter_rows():
                row_formulas = []
                for cell in row:
                    if cell.data_type == 'f':  # Célula com fórmula
                        row_formulas.append({
                            'cell': cell.coordinate,
                            'formula': cell.value,
                            'row': cell.row,
                            'col': cell.column
                        })
                    else:
                        row_formulas.append(None)
                
                if any(f is not None for f in row_formulas):
                    sheet_formulas.append(row_formulas)
            
            if sheet_formulas:
                formulas_data[sheet_name] = sheet_formulas
                print(f"   ✅ {len(sheet_formulas)} linhas com fórmulas encontradas")
            else:
                print(f"   ℹ️  Nenhuma fórmula encontrada")
        
        # Salvar em JSON
        output_path = 'formulas_controle.json'
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(formulas_data, f, ensure_ascii=False, indent=2)
        
        print(f"\n💾 Fórmulas salvas em: {output_path}")
        
        # Estatísticas
        total_formulas = sum(len([f for row in sheet for f in row if f]) 
                            for sheet in formulas_data.values())
        print(f"\n📊 ESTATÍSTICAS:")
        print(f"   Total de abas com fórmulas: {len(formulas_data)}")
        print(f"   Total de células com fórmulas: {total_formulas}")
        
        return formulas_data
        
    except Exception as e:
        print(f"❌ Erro ao processar arquivo: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    extrair_formulas_excel()
