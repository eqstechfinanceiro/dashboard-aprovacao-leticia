import openpyxl
import json
from datetime import datetime

def analyze_saldo_by_exact_headers(file_path, sheet_name):
    """Analisa pelos cabeçalhos exatos que conhecemos"""
    print(f"Analisando: {file_path} - {sheet_name}")
    print("="*50)
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb[sheet_name]
        
        # Cabeçalhos que procuramos
        target_headers = ['SALDO REEMBOLSAR', 'SALDO FINAL', 'SALDO CARTAO', 'ADIANTAMENTO']
        
        # Encontrar colunas pelos cabeçalhos exatos
        header_map = {}
        
        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header and isinstance(header, str):
                header_upper = header.strip().upper()
                
                for target in target_headers:
                    if target in header_upper or header_upper in target:
                        header_map[target] = {
                            'col': col,
                            'letter': chr(64 + col),
                            'header': header
                        }
                        print(f"Encontrado: {header} na coluna {chr(64 + col)}")
        
        print(f"Cabeçalhos encontrados: {len(header_map)}")
        
        analysis = {}
        
        for target, col_info in header_map.items():
            letter = col_info['letter']
            header = col_info['header']
            
            formulas = 0
            values = 0
            
            print(f"\n{header} ({letter}):")
            
            for row in range(2, min(22, ws.max_row + 1)):  # Primeiras 20 linhas
                cell = ws.cell(row, letter)
                val = cell.value
                
                if val is not None:
                    if isinstance(val, str) and val.startswith('='):
                        formulas += 1
                        print(f"  {letter}{row}: FÓRMULA = {val[:60]}...")
                    else:
                        values += 1
                        print(f"  {letter}{row}: VALOR = {val}")
            
            analysis[target] = {
                'formulas': formulas,
                'values': values,
                'type': 'formula' if formulas > values else 'value' if values > formulas else 'mixed',
                'header': header
            }
        
        wb.close()
        return analysis
        
    except Exception as e:
        print(f"Erro: {e}")
        import traceback
        traceback.print_exc()
        return {}

def analyze_all_headers(file_path, sheet_name):
    """Analisa TODOS os cabeçalhos para entender a estrutura"""
    print(f"\nAnalisando TODOS os cabeçalhos: {sheet_name}")
    print("="*50)
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb[sheet_name]
        
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header:
                headers.append({
                    'col': col,
                    'letter': chr(64 + col),
                    'header': str(header)
                })
        
        print(f"Total de cabeçalhos: {len(headers)}")
        print("Cabeçalhos encontrados:")
        for h in headers:
            print(f"  {h['letter']}: {h['header']}")
        
        wb.close()
        return headers
        
    except Exception as e:
        print(f"Erro: {e}")
        return []

def main():
    print("INVESTIGAÇÃO PROFUNDA - ORIGEM DOS DADOS DE SALDO")
    print("="*80)
    
    # Arquivos
    files = [
        '/home/haumea/Projects/dashboard-aprovacao-leticia/data/CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx',
        '/home/haumea/Projects/dashboard-aprovacao-leticia/data/1QZ ABRIL 2026 - VEXPENSES.xlsx'
    ]
    
    all_results = {}
    
    for file_path in files:
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
            
            for sheet_name in wb.sheetnames:
                print(f"\n{'='*60}")
                print(f"ARQUIVO: {file_path.split('/')[-1]}")
                print(f"ABA: {sheet_name}")
                print(f"{'='*60}")
                
                # Primeiro, analisar todos os cabeçalhos
                headers = analyze_all_headers(file_path, sheet_name)
                
                # Depois, analisar os campos de SALDO específicos
                saldo_analysis = analyze_saldo_by_exact_headers(file_path, sheet_name)
                
                if saldo_analysis:
                    all_results[f"{file_path.split('/')[-1]}::{sheet_name}"] = {
                        'headers': [h['header'] for h in headers],
                        'saldo_analysis': saldo_analysis
                    }
            
            wb.close()
        except Exception as e:
            print(f"Erro em {file_path}: {e}")
            import traceback
            traceback.print_exc()
    
    # Conclusão
    print(f"\n{'='*80}")
    print("CONCLUSÃO FINAL")
    print(f"{'='*80}")
    
    total_formulas = 0
    total_values = 0
    
    for key, data in all_results.items():
        print(f"\n{key}:")
        print(f"  Cabeçalhos: {len(data['headers'])}")
        
        if 'saldo_analysis' in data:
            for field, stats in data['saldo_analysis'].items():
                total_formulas += stats['formulas']
                total_values += stats['values']
                
                print(f"  {field} ({stats['header']}):")
                print(f"    Fórmulas: {stats['formulas']}")
                print(f"    Valores: {stats['values']}")
                print(f"    Tipo: {stats['type']}")
    
    print(f"\nTOTAL GERAL: Fórmulas={total_formulas}, Valores={total_values}")
    
    if total_values > total_formulas:
        print("\n✅ CONCLUSÃO: Dados de SALDO são VALORES ESTÁTICOS")
        print("   VIERAM DE: Reports da API (dados importados)")
        print("   PRÓXIMO PASSO: Investigar reports como fonte")
    elif total_formulas > total_values:
        print("\n✅ CONCLUSÃO: Dados de SALDO são FÓRMULAS")
        print("   VIERAM DE: Cálculos")
        print("   PRÓXIMO PASSO: Analisar lógica das fórmulas")
    else:
        print("\n⚠️ CONCLUSÃO: Mistura ou nenhum dado encontrado")
    
    # Salvar
    output = '/home/haumea/Projects/dashboard-aprovacao-leticia/investigation-docs/saldo_origin_complete.json'
    with open(output, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)
    
    print(f"\nSalvo em: {output}")

if __name__ == "__main__":
    main()
