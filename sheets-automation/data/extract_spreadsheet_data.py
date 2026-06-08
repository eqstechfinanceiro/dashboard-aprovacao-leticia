#!/usr/bin/env python3
"""
Extrai dados da planilha CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx
Cria arquivo JSON com todos os dados brutos (sem fórmulas)
"""

import json
import os
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


def extract_spreadsheet_data(filepath):
    """Extrai dados brutos da planilha."""
    wb = openpyxl.load_workbook(filepath, data_only=True)  # data_only=True para pegar valores calculados
    
    sheet = wb['Planilha1']
    
    # Identificar linha de cabeçalho
    header_row = 6  # Baseado na análise anterior
    
    # Ler cabeçalhos
    headers = {}
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=header_row, column=col)
        if cell.value:
            headers[col] = str(cell.value).strip()
    
    # Extrair dados de cada linha
    data = []
    for row in range(header_row + 1, sheet.max_row + 1):
        row_data = {}
        has_data = False
        
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            value = cell.value
            
            if value is not None and value != '':
                has_data = True
                col_letter = get_column_letter(col)
                header_name = headers.get(col, col_letter)
                
                # Converter para tipos apropriados
                if isinstance(value, (int, float)):
                    row_data[header_name] = float(value)
                else:
                    row_data[header_name] = str(value).strip()
        
        if has_data:
            data.append(row_data)
    
    return data


def main():
    spreadsheet_path = Path(__file__).parent / "CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx"
    
    if not spreadsheet_path.exists():
        print(f"Erro: Arquivo não encontrado: {spreadsheet_path}")
        return
    
    print(f"Extraindo dados da planilha: {spreadsheet_path}")
    
    data = extract_spreadsheet_data(spreadsheet_path)
    
    # Salvar JSON
    output_path = spreadsheet_path.parent / "dados_planilha_brutos.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump({
            "arquivo": spreadsheet_path.name,
            "data_extracao": datetime.now().isoformat(),
            "total_registros": len(data),
            "dados": data
        }, f, indent=2, ensure_ascii=False)
    
    print(f"✅ Dados extraídos: {len(data)} registros")
    print(f"📁 Arquivo salvo: {output_path}")
    
    # Mostrar amostra
    if data:
        print("\n📊 Amostra dos dados:")
        print(json.dumps(data[0], indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
