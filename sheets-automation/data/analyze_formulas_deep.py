#!/usr/bin/env python3
"""
Analisa fórmulas do arquivo CONTROLE convertido para .xlsx
Identifica origem dos dados (API vs cálculo vs manual)
"""

import json
import re
from pathlib import Path

import openpyxl


def analyze_sheet_formulas(sheet_name, sheet):
    """Analisa fórmulas de uma aba específica."""
    formulas = []
    
    # Analisar primeiras 50 linhas e 20 colunas
    for row_idx in range(1, min(51, sheet.max_row + 1)):
        for col_idx in range(1, min(21, sheet.max_column + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value
                formulas.append({
                    'cell': f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}",
                    'formula': formula,
                    'row': row_idx,
                    'col': col_idx
                })
    
    return formulas


def identify_data_sources(formulas):
    """Identifica possíveis fontes de dados nas fórmulas."""
    sources = {
        'api_references': [],
        'sheet_references': [],
        'external_references': [],
        'calculations': []
    }
    
    for item in formulas:
        formula = item['formula']
        
        # Referências a outras abas (pode indicar dados da API)
        sheet_refs = re.findall(r"'([^']+)'!", formula)
        for ref in sheet_refs:
            if ref not in sources['sheet_references']:
                sources['sheet_references'].append(ref)
        
        # Padrões que indicam dados da API
        if any(keyword in formula.upper() for keyword in ['VEXPENSES', 'API', 'IMPORT', 'QUERY']):
            sources['api_references'].append(item)
        
        # Referências externas
        if '[' in formula or ']' in formula:
            sources['external_references'].append(item)
        
        # Cálculos matemáticos
        if re.search(r'[+\-*/^]', formula):
            sources['calculations'].append(item)
    
    return sources


def analyze_all_sheets():
    """Analisa todas as abas do arquivo CONTROLE."""
    filepath = Path(__file__).parent / "CONTROLE - VEXPENSES - ABRIL- 2026.xlsx"
    
    if not filepath.exists():
        print(f"❌ Arquivo não encontrado: {filepath}")
        return None
    
    print("🔍 Analisando fórmulas do CONTROLE - VEXPENSES - ABRIL- 2026.xlsx\n")
    
    wb = openpyxl.load_workbook(filepath, data_only=False)  # data_only=False para ler fórmulas
    
    all_formulas = {}
    all_sources = {}
    
    for sheet_name in wb.sheetnames:
        print(f"\n📊 Aba: {sheet_name}")
        sheet = wb[sheet_name]
        
        # Analisar fórmulas
        formulas = analyze_sheet_formulas(sheet_name, sheet)
        print(f"  Fórmulas encontradas: {len(formulas)}")
        
        if formulas:
            # Mostrar primeiras 5 fórmulas
            print(f"  Amostra de fórmulas:")
            for f in formulas[:5]:
                print(f"    {f['cell']}: {f['formula'][:80]}")
        
        # Identificar fontes
        sources = identify_data_sources(formulas)
        print(f"  Referências a outras abas: {sources['sheet_references']}")
        print(f"  Possíveis referências API: {len(sources['api_references'])}")
        print(f"  Referências externas: {len(sources['external_references'])}")
        print(f"  Cálculos matemáticos: {len(sources['calculations'])}")
        
        all_formulas[sheet_name] = formulas
        all_sources[sheet_name] = sources
    
    wb.close()
    
    return all_formulas, all_sources


def analyze_key_sheets():
    """Analisa em detalhe as abas principais."""
    filepath = Path(__file__).parent / "CONTROLE - VEXPENSES - ABRIL- 2026.xlsx"
    wb = openpyxl.load_workbook(filepath, data_only=False)
    
    key_sheets = ['PAINEL', 'QUINZENAS', 'SALDO CARTAO', 'ADICIONAIS', 'EXTRATO']
    
    print("\n" + "="*60)
    print("🎯 ANÁLISE DETALHADA DAS ABAS PRINCIPAIS")
    print("="*60)
    
    for sheet_name in key_sheets:
        if sheet_name in wb.sheetnames:
            print(f"\n📋 {sheet_name}:")
            sheet = wb[sheet_name]
            
            # Ler cabeçalhos
            headers = []
            for col in range(1, min(20, sheet.max_column + 1)):
                cell = sheet.cell(row=1, column=col)
                if cell.value:
                    headers.append(str(cell.value))
            
            print(f"  Cabeçalhos: {headers[:10]}")
            
            # Analisar primeiras 10 linhas
            print(f"  Amostra de dados (primeiras 5 linhas):")
            for row_idx in range(2, min(7, sheet.max_row + 1)):
                row_data = []
                for col in range(1, min(15, sheet.max_column + 1)):
                    cell = sheet.cell(row=row_idx, column=col)
                    if cell.value:
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            row_data.append(f"[FÓRMULA: {cell.value[:30]}]")
                        else:
                            row_data.append(str(cell.value)[:20])
                    else:
                        row_data.append("")
                print(f"    Linha {row_idx}: {row_data[:5]}")
    
    wb.close()


def main():
    # Análise geral
    all_formulas, all_sources = analyze_all_sheets()
    
    # Análise detalhada
    analyze_key_sheets()
    
    # Salvar resultados
    if all_formulas:
        output_path = Path(__file__).parent / "controle_formulas_analysis.json"
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump({
                'formulas': all_formulas,
                'sources': all_sources
            }, f, indent=2, ensure_ascii=False)
        print(f"\n📁 Análise salva: {output_path}")
    
    print("\n💡 PRÓXIMOS PASSOS:")
    print("1. Analisar dependências entre abas")
    print("2. Identificar quais abas contêm dados brutos da API")
    print("3. Mapear o fluxo: API → dados brutos → fórmulas → valores finais")
    print("4. Replicar o fluxo via código")


if __name__ == "__main__":
    main()
