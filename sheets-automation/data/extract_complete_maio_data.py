#!/usr/bin/env python3
"""
Extrai dados completos da planilha CARGA 1 QZ MAIO 26
Esta é a FONTE DE VERDADE para MAIO 2026
"""

import json
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


def extract_complete_data():
    """Extrai todos os dados da planilha."""
    filepath = Path(__file__).parent / "CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx"
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb['Planilha1']
    
    # Identificar linha de cabeçalho
    header_row = 6
    
    # Ler cabeçalhos
    headers = {}
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=header_row, column=col)
        if cell.value:
            headers[col] = str(cell.value).strip()
    
    # Extrair dados de cada linha
    data = []
    for row in range(header_row + 1, sheet.max_row + 1):
        row_data = {}
        has_data = False
        
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            value = cell.value
            
            if value is not None and value != '':
                has_data = True
                header_name = headers.get(col, f"Col{col}")
                
                # Converter para tipos apropriados
                if isinstance(value, (int, float)):
                    row_data[header_name] = float(value)
                else:
                    row_data[header_name] = str(value).strip()
        
        if has_data:
            data.append(row_data)
    
    return data, headers


def main():
    print("📊 Extraindo dados completos de CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx\n")
    
    data, headers = extract_complete_data()
    
    # Salvar JSON completo
    output_path = Path(__file__).parent / "dados_completos_maio_2026.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump({
            "arquivo": "CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx",
            "data_extracao": datetime.now().isoformat(),
            "periodo": "MAIO 2026 (1ª quinzena)",
            "total_campos": len(headers),
            "campos": list(headers.values()),
            "total_registros": len(data),
            "dados": data
        }, f, indent=2, ensure_ascii=False)
    
    print(f"✅ Dados extraídos: {len(data)} registros")
    print(f"📁 Arquivo salvo: {output_path}")
    print(f"\n📋 Campos disponíveis ({len(headers)}):")
    for i, header in enumerate(headers.values(), 1):
        print(f"  {i}. {header}")
    
    # Mostrar amostra
    if data:
        print(f"\n📊 Amostra do primeiro registro:")
        print(json.dumps(data[0], indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
