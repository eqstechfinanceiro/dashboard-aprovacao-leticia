import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from datetime import datetime

def extract_all_formulas_xlsx(file_path, sheet_name):
    """Extrai TODAS as fórmulas de uma aba XLSX"""
    print(f"\n{'='*80}")
    print(f"EXTRAINDO FÓRMULAS: {sheet_name}")
    print(f"{'='*80}")
    
    wb = load_workbook(file_path, data_only=False)
    ws = wb[sheet_name]
    
    formulas = {
        "sheet_name": sheet_name,
        "total_cells_with_formulas": 0,
        "formulas_by_column": {},
        "formulas_by_type": {},
        "all_formulas": []
    }
    
    # Iterar por todas as células
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # célula com fórmula
                col_letter = cell.column_letter
                formula_str = cell.value
                
                # Categorizar fórmula
                formula_type = "UNKNOWN"
                if "XLOOKUP" in formula_str or "VLOOKUP" in formula_str:
                    formula_type = "LOOKUP"
                elif "SUBTOTAL" in formula_str:
                    formula_type = "SUBTOTAL"
                elif "IF(" in formula_str or "IFS(" in formula_str:
                    formula_type = "CONDITIONAL"
                elif "SUM" in formula_str or "+" in formula_str:
                    formula_type = "CALCULATION"
                
                formula_info = {
                    "cell": cell.coordinate,
                    "column": col_letter,
                    "row": cell.row,
                    "formula": formula_str,
                    "type": formula_type
                }
                
                formulas["all_formulas"].append(formula_info)
                formulas["total_cells_with_formulas"] += 1
                
                # Agrupar por coluna
                if col_letter not in formulas["formulas_by_column"]:
                    formulas["formulas_by_column"][col_letter] = []
                formulas["formulas_by_column"][col_letter].append(formula_info)
                
                # Agrupar por tipo
                if formula_type not in formulas["formulas_by_type"]:
                    formulas["formulas_by_type"][formula_type] = []
                formulas["formulas_by_type"][formula_type].append(formula_info)
    
    wb.close()
    return formulas

def analyze_main_sheet_structure(file_path, sheet_name):
    """Analisa em detalhe a estrutura da aba principal"""
    print(f"\n{'='*80}")
    print(f"ANÁLISE ESTRUTURAL: {sheet_name}")
    print(f"{'='*80}")
    
    wb = load_workbook(file_path, data_only=False)
    ws = wb[sheet_name]
    
    # Encontrar linha de cabeçalho (primeira linha com texto)
    header_row = None
    for row_idx in range(1, min(20, ws.max_row + 1)):
        row_values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        if any(v and isinstance(v, str) and len(v) > 2 for v in row_values):
            header_row = row_idx
            break
    
    if not header_row:
        header_row = 6  # fallback baseado na análise anterior
    
    structure = {
        "sheet_name": sheet_name,
        "header_row": header_row,
        "data_start_row": header_row + 1,
        "total_rows": ws.max_row,
        "total_cols": ws.max_column,
        "columns": {}
    }
    
    # Analisar cada coluna
    for col_idx in range(1, ws.max_column + 1):
        col_letter = chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + (col_idx - 26))}"
        
        header_cell = ws.cell(header_row, col_idx)
        header_value = header_cell.value if header_cell.value else f"COL_{col_letter}"
        
        # Verificar se tem fórmula no cabeçalho
        header_has_formula = header_cell.data_type == 'f'
        
        # Amostrar 10 células de dados
        sample_values = []
        sample_formulas = []
        data_count = 0
        
        for row_idx in range(header_row + 1, min(header_row + 11, ws.max_row + 1)):
            cell = ws.cell(row_idx, col_idx)
            if cell.value is not None:
                data_count += 1
                if cell.data_type == 'f':
                    sample_formulas.append(cell.value)
                else:
                    sample_values.append(cell.value)
        
        structure["columns"][col_letter] = {
            "header": header_value,
            "header_has_formula": header_has_formula,
            "header_formula": header_cell.value if header_has_formula else None,
            "data_sample_count": data_count,
            "sample_values": sample_values[:5],
            "sample_formulas": sample_formulas[:3],
            "has_formulas": len(sample_formulas) > 0
        }
    
    wb.close()
    return structure

if __name__ == "__main__":
    # Planilha 1
    file1 = "/home/haumea/Projects/dashboard-aprovacao-leticia/data/1QZ ABRIL 2026 - VEXPENSES (1).xlsx"
    main_sheet = "1 QZ VEXPENSES 04_2026"
    
    # Extrair fórmulas da aba principal
    formulas = extract_all_formulas_xlsx(file1, main_sheet)
    
    output_formulas = "/home/haumea/Projects/dashboard-aprovacao-leticia/investigation-docs/planilha1_formulas.json"
    with open(output_formulas, 'w', encoding='utf-8') as f:
        json.dump(formulas, f, ensure_ascii=False, indent=2)
    print(f"\nFórmulas salvas em: {output_formulas}")
    
    # Analisar estrutura da aba principal
    structure = analyze_main_sheet_structure(file1, main_sheet)
    
    output_structure = "/home/haumea/Projects/dashboard-aprovacao-leticia/investigation-docs/planilha1_structure.json"
    with open(output_structure, 'w', encoding='utf-8') as f:
        json.dump(structure, f, ensure_ascii=False, indent=2)
    print(f"Estrutura salva em: {output_structure}")
    
    print("\n" + "="*80)
    print("ANÁLISE DETALHADA CONCLUÍDA!")
    print("="*80)
