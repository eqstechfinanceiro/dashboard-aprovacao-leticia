import openpyxl
import json
import subprocess
from collections import defaultdict

# Load BASE PREST per CPF
wb = openpyxl.load_workbook('CONTROLE - VEXPENSES - AGOSTO 2026.xlsx', read_only=True, data_only=True)
ws = wb['BASE PREST ']

bp_per_cpf = defaultdict(float)
bp_fatura_per_cpf = defaultdict(float)
bp_reports = defaultdict(lambda: defaultdict(lambda: {'total': 0, 'count': 0, 'status': '', 'name': ''}))
for row in ws.iter_rows(min_row=4, values_only=True):
    cpf = str(row[9]).strip() if row[9] else ''
    valor = float(row[26]) if row[26] is not None else 0
    report_id = str(row[1]).strip() if row[1] else ''
    report_name = str(row[2]).strip().upper() if row[2] else ''
    status = str(row[10]).strip() if row[10] else ''
    if not cpf or cpf == 'None':
        continue
    bp_per_cpf[cpf] += valor
    if 'FATURA' in report_name or 'CARTAO' in report_name or 'CARTAO' in report_name:
        bp_fatura_per_cpf[cpf] += valor
    if report_id and report_id != 'None':
        bp_reports[cpf][report_id]['total'] += valor
        bp_reports[cpf][report_id]['count'] += 1
        bp_reports[cpf][report_id]['status'] = status
        bp_reports[cpf][report_id]['name'] = report_name

# Load API prestacao debug (cpf_summary only - this has all CPFs)
with open('prestacao_debug.json', 'r', encoding='utf-8-sig') as f:
    debug = json.load(f)

api_per_cpf = {}
for item in debug.get('cpf_summary', []):
    cpf = item['user_cpf'].strip()
    api_per_cpf[cpf] = {
        'prestacao_total': float(item['prestacao_total']),
        'total_expenses': float(item['total_expenses']),
        'report_count': int(item['report_count']),
    }

# Load PAINEL names
ws_painel = wb['PAINEL']
painel_names = {}
for row in ws_painel.iter_rows(min_row=12, values_only=True):
    cpf = str(row[2]).strip() if row[2] else ''
    nome = str(row[1]).strip() if row[1] else ''
    if cpf and cpf != 'None':
        painel_names[cpf] = nome

# Find diffs
diffs = []
for cpf in set(bp_per_cpf.keys()) | set(api_per_cpf.keys()):
    bp_total = bp_per_cpf.get(cpf, 0) - bp_fatura_per_cpf.get(cpf, 0)
    api_total = api_per_cpf.get(cpf, {}).get('prestacao_total', 0)
    diff = bp_total - api_total
    if abs(diff) > 1:
        diffs.append((cpf, painel_names.get(cpf, ''), bp_total, api_total, diff))

diffs.sort(key=lambda x: abs(x[4]), reverse=True)

total_diff = sum(d[4] for d in diffs)
print(f"Total remaining diff: {total_diff:.2f}")
print(f"Users with diff: {len(diffs)}")
print()
print("Top 30 diffs (BASE_PREST_normal - API_prestacao):")
for cpf, nome, bp, api, diff in diffs[:30]:
    print(f"  {nome[:35]:35s} cpf={cpf} BP={bp:.2f} API={api:.2f} diff={diff:+.2f}")

# For top 10, fetch per-CPF debug data from the API
print("\n\n=== REPORT-LEVEL ANALYSIS FOR TOP 10 ===")
for cpf, nome, bp, api, diff in diffs[:10]:
    print(f"\n--- {nome} (cpf={cpf}) BP={bp:.2f} API={api:.2f} diff={diff:+.2f} ---")

    # Fetch per-CPF debug data
    result = subprocess.run(
        ['powershell', '-NoProfile', '-ExecutionPolicy', 'Bypass', '-File',
         'c:\\Users\\italo.medrado\\Desktop\\Projects\\Análise de dados\\Leticia\\dashboard-test\\fetch_cpf_debug.ps1',
         cpf],
        capture_output=True, text=True, timeout=30
    )
    # We need to modify the script to accept a CPF parameter
    # Instead, let's use inline PowerShell
    ps_script = f'''
$session = New-Object Microsoft.PowerShell.Commands.WebRequestSession
$body = @{{ email = "italo.medrado@eqsengenharia.com.br"; password = "EQSeng4292@" }} | ConvertTo-Json
Invoke-RestMethod -Uri "http://localhost:3000/api/auth/login" -Method POST -Body $body -ContentType "application/json" -WebSession $session | Out-Null
$r = Invoke-RestMethod -Uri "http://localhost:3000/api/debug-prestacao-cpf?cpf={cpf}" -Method GET -WebSession $session
$r | ConvertTo-Json -Depth 5
'''
    result = subprocess.run(
        ['powershell', '-NoProfile', '-Command', ps_script],
        capture_output=True, text=True, timeout=30
    )

    try:
        api_data = json.loads(result.stdout)
        api_reps = {r['report_id']: r for r in api_data.get('api_expenses', [])}
    except:
        print(f"  ERROR parsing API data: {result.stderr[:200]}")
        continue

    bp_reps = bp_reports.get(cpf, {})

    # Compare each report
    all_rids = set(bp_reps.keys()) | set(str(r) for r in api_reps.keys())

    def sort_key(x):
        bp_v = float(bp_reps.get(x, {}).get('total', 0))
        api_v = float(api_reps.get(int(x) if str(x).isdigit() else -1, {}).get('total_value', 0))
        return abs(bp_v - api_v)

    for rid in sorted(all_rids, key=sort_key, reverse=True):
        bp_r = bp_reps.get(rid, {'total': 0, 'count': 0, 'status': '', 'name': ''})
        api_r = api_reps.get(int(rid) if str(rid).isdigit() else -1, {'report_name': '', 'status': '', 'expense_count': 0, 'total_value': 0, 'excluded_value': 0})
        rdiff = float(bp_r['total']) - float(api_r['total_value'])
        if abs(rdiff) > 0.01:
            in_api = 'YES' if int(api_r.get('expense_count', 0)) > 0 else 'NO'
            print(f"  rid={rid} bp_name={bp_r['name'][:25]} bp_status={bp_r['status']} bp_total={float(bp_r['total']):.2f} api_name={str(api_r.get('report_name',''))[:25]} api_status={api_r.get('status','')} api_count={api_r.get('expense_count',0)} api_total={float(api_r.get('total_value',0)):.2f} api_excl={float(api_r.get('excluded_value',0)):.2f} diff={rdiff:+.2f} in_api={in_api}")
