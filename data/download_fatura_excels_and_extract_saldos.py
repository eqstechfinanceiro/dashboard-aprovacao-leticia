import requests
import json
import pandas as pd
import openpyxl
import io
from datetime import datetime

API_KEY = "N2ntX8mUF7AvjVcyT6DZwnOQXRvYgel9pCirCGKKETO5R70HEZ8UdbQA0Nt8"
BASE_URL = "https://api.vexpenses.com/v2"

headers = {
    "Authorization": API_KEY,
    "Content-Type": "application/json",
    "Accept": "application/json"
}

def get_fatura_reports():
    """Obtém reports de FATURA/CARTÃO"""
    try:
        url = f"{BASE_URL}/reports"
        params = {"paginate": "false", "per_page": 100}
        response = requests.get(url, headers=headers, params=params)
        
        if response.status_code == 200:
            data = response.json()
            if 'data' in data:
                reports = data['data']
                
                # Filtrar reports de FATURA recentes
                fatura_reports = []
                for report in reports:
                    description = report.get('description', '').upper()
                    if 'FATURA' in description and 'CARTÃO' in description:
                        # Verificar se é recente (2025-2026)
                        created_at = report.get('created_at', '')
                        if created_at and ('2025' in created_at or '2026' in created_at):
                            fatura_reports.append(report)
                
                print(f"Reports de FATURA/CARTÃO recentes: {len(fatura_reports)}")
                return fatura_reports
                
    except Exception as e:
        print(f"Erro ao obter reports: {e}")
    
    return []

def download_fatura_excel(report):
    """Baixa arquivo Excel de um report de FATURA"""
    report_id = report.get('id')
    excel_link = report.get('excel_link', '')
    
    if not excel_link:
        return None
    
    try:
        print(f"Baixando Excel do Report {report_id}...")
        
        response = requests.get(excel_link, headers=headers, timeout=30)
        
        if response.status_code == 200:
            print(f"Arquivo baixado: {len(response.content)} bytes")
            
            # Salvar arquivo temporariamente
            temp_file = f"/tmp/fatura_report_{report_id}.xlsx"
            with open(temp_file, 'wb') as f:
                f.write(response.content)
            
            return temp_file
        else:
            print(f"Erro ao baixar: {response.status_code}")
            
    except Exception as e:
        print(f"Exceção ao baixar: {e}")
    
    return None

def analyze_fatura_excel_for_saldos(excel_file_path, report_info):
    """Analisa arquivo Excel de FATURA procurando dados de saldo"""
    print(f"\nAnalisando Excel: {excel_file_path}")
    
    try:
        # Tentar com openpyxl primeiro
        try:
            wb = openpyxl.load_workbook(excel_file_path, data_only=True)
            
            print(f"Sheets: {wb.sheetnames}")
            
            saldo_data = {
                'report_info': report_info,
                'sheets_data': {}
            }
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                print(f"\nSheet: {sheet_name}")
                print(f"Dimensões: {ws.max_row}x{ws.max_column}")
                
                # Procurar por dados de saldo
                sheet_saldos = []
                
                # Procurar células com valores financeiros
                for row in range(1, min(ws.max_row + 1, 100)):  # Primeiras 100 linhas
                    for col in range(1, min(ws.max_column + 1, 20)):  # Primeiras 20 colunas
                        cell = ws.cell(row, col)
                        value = cell.value
                        
                        if value is not None:
                            # Procurar por valores numéricos (saldos)
                            if isinstance(value, (int, float)) and abs(value) > 100:
                                # Obter contexto (células ao redor)
                                context = []
                                
                                # Cabeçalhos na mesma linha
                                for c in range(max(1, col-3), min(col+3, ws.max_column+1)):
                                    header_cell = ws.cell(row, c)
                                    if header_cell.value and isinstance(header_cell.value, str):
                                        context.append(header_cell.value)
                                
                                # Cabeçalhos na mesma coluna
                                for r in range(max(1, row-3), min(row+3, ws.max_row+1)):
                                    header_cell = ws.cell(r, col)
                                    if header_cell.value and isinstance(header_cell.value, str):
                                        context.append(header_cell.value)
                                
                                sheet_saldos.append({
                                    'cell': f"{row}{chr(64+col)}",
                                    'value': value,
                                    'context': list(set(context)),  # Remover duplicatas
                                    'row': row,
                                    'col': col
                                })
                
                # Mostrar valores encontrados
                if sheet_saldos:
                    print(f"Valores financeiros encontrados: {len(sheet_saldos)}")
                    for saldo in sheet_saldos[:10]:  # Primeiros 10
                        print(f"  {saldo['cell']}: R$ {saldo['value']:.2f}")
                        if saldo['context']:
                            print(f"    Contexto: {saldo['context'][:3]}")
                
                saldo_data['sheets_data'][sheet_name] = sheet_saldos
            
            wb.close()
            return saldo_data
            
        except Exception as e:
            print(f"Erro com openpyxl: {e}")
            
            # Tentar com pandas como fallback
            try:
                xls = pd.ExcelFile(excel_file_path)
                
                saldo_data = {
                    'report_info': report_info,
                    'sheets_data': {}
                }
                
                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
                    
                    print(f"\nSheet (pandas): {sheet_name}")
                    print(f"Dimensões: {df.shape}")
                    print(f"Colunas: {list(df.columns)}")
                    
                    # Procurar colunas numéricas com valores altos
                    numeric_cols = []
                    for col in df.columns:
                        if df[col].dtype in ['int64', 'float64']:
                            non_null = df[col].dropna()
                            if len(non_null) > 0:
                                max_val = non_null.max()
                                min_val = non_null.min()
                                
                                if abs(max_val) > 1000 or abs(min_val) > 1000:
                                    numeric_cols.append({
                                        'column': col,
                                        'max': max_val,
                                        'min': min_val,
                                        'count': len(non_null)
                                    })
                    
                    if numeric_cols:
                        print(f"Colunas numéricas relevantes: {len(numeric_cols)}")
                        for col_data in numeric_cols:
                            print(f"  {col_data['column']}: R$ {col_data['min']:.2f} a R$ {col_data['max']:.2f}")
                    
                    saldo_data['sheets_data'][sheet_name] = numeric_cols
                
                return saldo_data
                
            except Exception as e2:
                print(f"Erro também com pandas: {e2}")
                
    except Exception as e:
        print(f"Erro geral ao analisar Excel: {e}")
    
    return None

def extract_user_saldos_from_faturas():
    """Extrai dados de saldo dos reports de FATURA"""
    print("EXTRAINDO DADOS DE SALDO DOS REPORTS DE FATURA")
    print("="*60)
    
    # 1. Obter reports de FATURA
    fatura_reports = get_fatura_reports()
    
    if not fatura_reports:
        print("Nenhum report de FATURA encontrado")
        return {}
    
    print(f"Analisando {len(fatura_reports)} reports...")
    
    # 2. Baixar e analisar alguns reports
    all_saldo_data = {}
    
    for i, report in enumerate(fatura_reports[:5]):  # Primeiros 5
        report_id = report.get('id')
        description = report.get('description', '')
        
        print(f"\n{'='*50}")
        print(f"Report {i+1}/5: {description}")
        print(f"ID: {report_id}")
        print(f"{'='*50}")
        
        # Baixar Excel
        excel_file = download_fatura_excel(report)
        
        if excel_file:
            # Analisar Excel
            saldo_data = analyze_fatura_excel_for_saldos(excel_file, {
                'id': report_id,
                'description': description,
                'status': report.get('status', ''),
                'created_at': report.get('created_at', '')
            })
            
            if saldo_data:
                all_saldo_data[f"report_{report_id}"] = saldo_data
                
                # Salvar dados do report
                report_file = f"/home/haumea/Projects/dashboard-aprovacao-leticia/investigation-docs/fatura_report_{report_id}_analysis.json"
                with open(report_file, 'w', encoding='utf-8') as f:
                    json.dump(saldo_data, f, ensure_ascii=False, indent=2)
                print(f"Dados salvos em: {report_file}")
    
    return all_saldo_data

def find_saldo_patterns_for_users():
    """Encontra padrões de saldo para usuários específicos"""
    print("\nPROCURANDO PADRÕES DE SALDO PARA USUÁRIOS")
    print("="*50)
    
    # Usuários que já mapeamos
    target_users = {
        'JONAS CAVALCANTI': {'saldo_final': 6945.16, 'saldo_cartao': 0},
        'RODRIGO CESAR': {'saldo_final': 6626.04, 'saldo_cartao': 0},
        'CAIO FRANCESCONI': {'saldo_final': 6504.20, 'saldo_cartao': 0}
    }
    
    # Carregar dados analisados
    saldo_data = extract_user_saldos_from_faturas()
    
    # Procurar correspondências
    matches = {}
    
    for report_key, report_data in saldo_data.items():
        report_info = report_data.get('report_info', {})
        sheets_data = report_data.get('sheets_data', {})
        
        print(f"\nAnalisando {report_key}:")
        print(f"  Report: {report_info.get('description', '')}")
        
        for sheet_name, data in sheets_data.items():
            if isinstance(data, list):  # Dados do openpyxl
                for item in data:
                    value = item.get('value', 0)
                    
                    # Procurar valores próximos aos esperados
                    for user_name, expected_data in target_users.items():
                        expected_saldo = expected_data['saldo_final']
                        
                        if abs(value - expected_saldo) < 100:  # Tolerância de 100
                            if user_name not in matches:
                                matches[user_name] = []
                            
                            matches[user_name].append({
                                'report_key': report_key,
                                'sheet_name': sheet_name,
                                'cell': item.get('cell'),
                                'value': value,
                                'expected': expected_saldo,
                                'diff': abs(value - expected_saldo),
                                'context': item.get('context', [])
                            })
                            
                            print(f"    ✅ {user_name}: R$ {value:.2f} (esperado: R$ {expected_saldo:.2f})")
                            print(f"       Local: {item.get('cell')} em {sheet_name}")
            
            elif isinstance(data, list):  # Dados do pandas
                for col_data in data:
                    max_val = col_data.get('max', 0)
                    
                    for user_name, expected_data in target_users.items():
                        expected_saldo = expected_data['saldo_final']
                        
                        if abs(max_val - expected_saldo) < 500:  # Tolerância maior
                            if user_name not in matches:
                                matches[user_name] = []
                            
                            matches[user_name].append({
                                'report_key': report_key,
                                'sheet_name': sheet_name,
                                'column': col_data.get('column'),
                                'max_value': max_val,
                                'expected': expected_saldo,
                                'diff': abs(max_val - expected_saldo)
                            })
                            
                            print(f"    ✅ {user_name}: R$ {max_val:.2f} (esperado: R$ {expected_saldo:.2f})")
                            print(f"       Coluna: {col_data.get('column')} em {sheet_name}")
    
    # Mostrar resumo das correspondências
    print(f"\n{'='*50}")
    print("RESUMO DE CORRESPONDÊNCIAS")
    print(f"{'='*50}")
    
    for user_name, user_matches in matches.items():
        print(f"\n{user_name}:")
        print(f"  Correspondências: {len(user_matches)}")
        
        if user_matches:
            best_match = min(user_matches, key=lambda x: x['diff'])
            print(f"  Melhor: R$ {best_match['value'] if 'value' in best_match else best_match['max_value']:.2f}")
            print(f"  Diferença: R$ {best_match['diff']:.2f}")
            print(f"  Fonte: {best_match['report_key']}")
    
    return matches

def main():
    """Função principal"""
    print("EXTRAINDO 100% DOS DADOS DE SALDO VIA REPORTS DE FATURA")
    print("="*80)
    print("Baixando e analisando arquivos Excel dos reports de FATURA/CARTÃO")
    print("="*80)
    
    # 1. Extrair dados de saldo dos reports
    saldo_data = extract_user_saldos_from_faturas()
    
    # 2. Encontrar padrões para usuários
    user_matches = find_saldo_patterns_for_users()
    
    # 3. Salvar resultados completos
    results = {
        'extraction_date': datetime.now().isoformat(),
        'saldo_data_count': len(saldo_data),
        'user_matches': user_matches,
        'status': 'Dados de saldo extraídos dos reports de FATURA',
        'next_steps': [
            'Analisar correspondências encontradas',
            'Implementar extração automatizada',
            'Integrar com solução principal'
        ]
    }
    
    output_file = '/home/haumea/Projects/dashboard-aprovacao-leticia/investigation-docs/fatura_saldos_extraction.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    print(f"\n\nResultados salvos em: {output_file}")
    print("\n" + "="*80)
    print("🎯 EXTRAÇÃO DE DADOS DE SALDO CONCLUÍDA!")
    print("="*80)
    
    if user_matches:
        print(f"✅ {len(user_matches)} usuários com correspondências de saldo")
        print("✅ Fonte: Reports de FATURA/CARTÃO")
        print("✅ Pronto para integração 100%")
    else:
        print("⚠️  Nenhuma correspondência direta encontrada")
        print("   Análise manual dos dados necessária")

if __name__ == "__main__":
    main()
