import openpyxl
import json
from datetime import datetime

def analyze_saldo_cells_in_sheet(file_path, sheet_name):
    """Analisa especificamente as células de SALDO para entender se são fórmulas ou valores"""
    print(f"Analisando dados de SALDO em: {file_path} - {sheet_name}")
    print("="*60)
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)  # data_only=False para ver fórmulas
        ws = wb[sheet_name]
        
        # Procurar colunas de SALDO
        saldo_columns = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header and isinstance(header, str):
                header_upper = header.upper()
                if 'SALDO' in header_upper or 'SALDO' in header:
                    saldo_columns.append({
                        'col': col,
                        'header': header,
                        'col_letter': chr(64 + col)
                    })
        
        print(f"Colunas de SALDO encontradas: {len(saldo_columns)}")
        for col_info in saldo_columns:
            print(f"  Coluna {col_info['col_letter']}: {col_info['header']}")
        
        # Analisar as células de cada coluna SALDO
        saldo_data = {}
        
        for col_info in saldo_columns:
            col_letter = col_info['col_letter']
            col = col_info['col']
            header = col_info['header']
            
            print(f"\nAnalisando coluna: {header} ({col_letter})")
            
            # Verificar primeiras 20 linhas de dados
            cells_analysis = []
            
            for row in range(2, min(22, ws.max_row + 1)):  # Linhas 2-21
                cell = ws.cell(row, col)
                value = cell.value
                
                # Verificar se é fórmula
                is_formula = False
                formula = None
                
                if value and isinstance(value, str):
                    if value.startswith('='):
                        is_formula = True
                        formula = value
                
                cells_analysis.append({
                    'row': row,
                    'cell_ref': f"{col_letter}{row}",
                    'value': value,
                    'is_formula': is_formula,
                    'formula': formula,
                    'data_type': type(value).__name__
                })
            
            # Contar fórmulas vs valores estáticos
            formula_count = sum(1 for c in cells_analysis if c['is_formula'])
            value_count = sum(1 for c in cells_analysis if not c['is_formula'] and c['value'] is not None)
            
            print(f"  Total células analisadas: {len(cells_analysis)}")
            print(f"  Fórmulas: {formula_count}")
            print(f"  Valores estáticos: {value_count}")
            
            # Mostrar exemplos
            if formula_count > 0:
                print(f"  Exemplos de fórmulas:")
                for c in cells_analysis[:3]:
                    if c['is_formula']:
                        print(f"    {c['cell_ref']}: {c['formula']}")
            
            if value_count > 0:
                print(f"  Exemplos de valores estáticos:")
                for c in cells_analysis[:3]:
                    if not c['is_formula'] and c['value'] is not None:
                        print(f"    {c['cell_ref']}: {c['value']} ({c['data_type']})")
            
            saldo_data[header] = {
                'col_letter': col_letter,
                'formula_count': formula_count,
                'value_count': value_count,
                'is_mostly_formulas': formula_count > value_count,
                'is_mostly_values': value_count > formula_count,
                'samples': cells_analysis[:5]
            }
        
        wb.close()
        return saldo_data
        
    except Exception as e:
        print(f"Erro ao analisar: {e}")
        return {}

def analyze_all_sheets_for_saldo():
    """Analisa todas as planilhas disponíveis para dados de SALDO"""
    print("ANALISANDO ORIGEM DOS DADOS DE SALDO")
    print("="*80)
    
    # Planilhas disponíveis
    sheet_files = [
        '/home/haumea/Projects/dashboard-aprovacao-leticia/data/CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx',
        '/home/haumea/Projects/dashboard-aprovacao-leticia/data/1QZ ABRIL 2026 - VEXPENSES (1).xlsx',
        '/home/haumea/Projects/dashboard-aprovacao-leticia/data/CONTROLE - VEXPENSES - ABRIL- 2026 (1).xlsb'
    ]
    
    all_saldo_data = {}
    
    for file_path in sheet_files:
        print(f"\n{'='*60}")
        print(f"Arquivo: {file_path}")
        print(f"{'='*60}")
        
        try:
            if file_path.endswith('.xlsb'):
                print("Arquivo XLSB - não suportado pelo openpyxl")
                continue
                
            wb = openpyxl.load_workbook(file_path, data_only=False)
            
            print(f"Abas disponíveis: {wb.sheetnames}")
            
            for sheet_name in wb.sheetnames:
                # Analisar apenas abas que parecem ter dados de SALDO
                sheet_name_upper = sheet_name.upper()
                if any(keyword in sheet_name_upper for keyword in ['VEXPENSES', 'CARGA', '1QZ', 'SALDO', 'CONTROLE']):
                    saldo_data = analyze_saldo_cells_in_sheet(file_path, sheet_name)
                    
                    if saldo_data:
                        all_saldo_data[f"{file_path}::{sheet_name}"] = saldo_data
            
            wb.close()
            
        except Exception as e:
            print(f"Erro ao processar arquivo: {e}")
    
    return all_saldo_data

def conclude_saldo_origin(all_saldo_data):
    """Conclui a origem dos dados de SALDO"""
    print(f"\nCONCLUSÃO SOBRE ORIGEM DOS DADOS DE SALDO")
    print("="*60)
    
    if not all_saldo_data:
        print("Nenhum dado de SALDO encontrado")
        return {}
    
    total_formulas = 0
    total_values = 0
    
    for file_sheet, data in all_saldo_data.items():
        print(f"\n{file_sheet}:")
        
        for col_header, col_data in data.items():
            print(f"  {col_header}:")
            print(f"    Fórmulas: {col_data['formula_count']}")
            print(f"    Valores: {col_data['value_count']}")
            print(f"    Tipo: {'FÓRMULAS' if col_data['is_mostly_formulas'] else 'VALORES ESTÁTICOS' if col_data['is_mostly_values'] else 'MISTO'}")
            
            total_formulas += col_data['formula_count']
            total_values += col_data['value_count']
    
    print(f"\nTOTAL GERAL:")
    print(f"  Fórmulas: {total_formulas}")
    print(f"  Valores estáticos: {total_values}")
    
    if total_values > total_formulas:
        print(f"  CONCLUSÃO: Os dados de SALDO são VALORES ESTÁTICOS (vieram de outra fonte)")
        print(f"  IMPLICAÇÃO: Devem vir dos REPORTS da API")
    elif total_formulas > total_values:
        print(f"  CONCLUSÃO: Os dados de SALDO são FÓRMULAS (calculados)")
        print(f"  IMPLICAÇÃO: Precisamos entender a lógica das fórmulas")
    else:
        print(f"  CONCLUSÃO: Mistura de fórmulas e valores")
    
    return {
        'total_formulas': total_formulas,
        'total_values': total_values,
        'conclusion': 'values' if total_values > total_formulas else 'formulas' if total_formulas > total_values else 'mixed'
    }

def main():
    """Função principal"""
    print("INVESTIGAÇÃO PROFUNDA - ORIGEM DOS DADOS DE SALDO")
    print("="*80)
    
    # 1. Analisar planilhas para origem dos dados de SALDO
    all_saldo_data = analyze_all_sheets_for_saldo()
    
    # 2. Concluir sobre a origem
    conclusion = conclude_saldo_origin(all_saldo_data)
    
    # 3. Salvar resultados
    results = {
        'investigation_date': datetime.now().isoformat(),
        'saldo_data_analysis': all_saldo_data,
        'conclusion': conclusion,
        'next_steps': []
    }
    
    if conclusion['conclusion'] == 'values':
        results['next_steps'] = [
            'Investigar reports da API como fonte dos valores',
            'Baixar arquivos Excel dos reports',
            'Extrair dados de SALDO dos reports',
            'Mapear campos de SALDO para API'
        ]
    elif conclusion['conclusion'] == 'formulas':
        results['next_steps'] = [
            'Analisar as fórmulas de SALDO',
            'Entender a lógica de cálculo',
            'Implementar as fórmulas',
            'Validar com dados reais'
        ]
    else:
        results['next_steps'] = [
            'Investigar ambas as fontes',
            'Analisar fórmulas e valores',
            'Determinar a origem principal'
        ]
    
    output_file = '/home/haumea/Projects/dashboard-aprovacao-leticia/investigation-docs/saldo_origin_investigation.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    print(f"\n\nResultados salvos em: {output_file}")
    print("\n" + "="*80)
    print("INVESTIGAÇÃO CONCLUÍDA!")
    print("="*80)

if __name__ == "__main__":
    main()
