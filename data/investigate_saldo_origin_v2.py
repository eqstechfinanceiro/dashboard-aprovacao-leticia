import openpyxl
import json
from datetime import datetime

def analyze_saldo_cells_detailed(file_path, sheet_name):
    """Analisa células de SALDO para determinar se são fórmulas ou valores"""
    print(f"Analisando: {file_path} - {sheet_name}")
    print("="*50)
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb[sheet_name]
        
        # Encontrar colunas de SALDO
        saldo_cols = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header and isinstance(header, str):
                if 'SALDO' in header.upper():
                    saldo_cols.append({
                        'col': col,
                        'letter': chr(64 + col),
                        'header': header
                    })
        
        print(f"Colunas SALDO: {len(saldo_cols)}")
        
        analysis = {}
        
        for col_info in saldo_cols:
            letter = col_info['letter']
            header = col_info['header']
            
            formulas = 0
            values = 0
            
            print(f"\n{header} ({letter}):")
            
            for row in range(2, min(12, ws.max_row + 1)):
                cell = ws.cell(row, letter)
                val = cell.value
                
                if val:
                    if isinstance(val, str) and val.startswith('='):
                        formulas += 1
                        print(f"  {letter}{row}: FÓRMULA = {val[:50]}...")
                    else:
                        values += 1
                        print(f"  {letter}{row}: VALOR = {val}")
            
            analysis[header] = {
                'formulas': formulas,
                'values': values,
                'type': 'formula' if formulas > values else 'value' if values > formulas else 'mixed'
            }
        
        wb.close()
        return analysis
        
    except Exception as e:
        print(f"Erro: {e}")
        return {}

def main():
    print("INVESTIGANDO ORIGEM DOS DADOS DE SALDO")
    print("="*60)
    
    # Arquivos corretos
    files = [
        '/home/haumea/Projects/dashboard-aprovacao-leticia/data/CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx',
        '/home/haumea/Projects/dashboard-aprovacao-leticia/data/1QZ ABRIL 2026 - VEXPENSES.xlsx'
    ]
    
    all_results = {}
    
    for file_path in files:
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
            
            for sheet_name in wb.sheetnames:
                result = analyze_saldo_cells_detailed(file_path, sheet_name)
                if result:
                    all_results[f"{file_path}::{sheet_name}"] = result
            
            wb.close()
        except Exception as e:
            print(f"Erro em {file_path}: {e}")
    
    # Conclusão
    print(f"\nCONCLUSÃO:")
    print("="*50)
    
    total_formulas = 0
    total_values = 0
    
    for key, data in all_results.items():
        for col, stats in data.items():
            total_formulas += stats['formulas']
            total_values += stats['values']
            
            print(f"{key} - {col}:")
            print(f"  Fórmulas: {stats['formulas']}")
            print(f"  Valores: {stats['values']}")
            print(f"  Tipo: {stats['type']}")
    
    print(f"\nTOTAL: Fórmulas={total_formulas}, Valores={total_values}")
    
    if total_values > total_formulas:
        print("CONCLUSÃO: Dados de SALDO são VALORES ESTÁTICOS")
        print("VIERAM DE: Reports da API")
        print("AÇÃO: Investigar reports como fonte")
    elif total_formulas > total_values:
        print("CONCLUSÃO: Dados de SALDO são FÓRMULAS")
        print("VIERAM DE: Cálculos")
        print("AÇÃO: Analisar lógica das fórmulas")
    
    # Salvar
    output = '/home/haumea/Projects/dashboard-aprovacao-leticia/investigation-docs/saldo_origin.json'
    with open(output, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)
    
    print(f"\nSalvo em: {output}")

if __name__ == "__main__":
    main()
