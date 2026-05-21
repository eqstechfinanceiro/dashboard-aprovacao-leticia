import pandas as pd
import json
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import xlrd

def analyze_sheet_structure(file_path, sheet_name=None):
    """Analisa estrutura completa de uma planilha"""
    print(f"\n{'='*80}")
    print(f"ANALISANDO: {file_path}")
    if sheet_name:
        print(f"Aba: {sheet_name}")
    print(f"{'='*80}")
    
    try:
        # Detectar formato do arquivo
        if file_path.endswith('.xlsx'):
            wb = load_workbook(file_path, data_only=False)
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                # Analisar primeira aba
                ws = wb.active
                sheet_name = ws.title
            
            print(f"Formato: XLSX")
            print(f"Aba analisada: {sheet_name}")
            print(f"Total linhas: {ws.max_row}")
            print(f"Total colunas: {ws.max_column}")
            
            # Encontrar linha de cabeçalho
            header_row = None
            for row_idx in range(1, min(20, ws.max_row + 1)):
                row_values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
                if any(v and isinstance(v, str) and len(v.strip()) > 2 for v in row_values):
                    header_row = row_idx
                    break
            
            if header_row:
                print(f"Linha de cabeçalho: {header_row}")
                
                # Extrair cabeçalhos
                headers = []
                for col_idx in range(1, ws.max_column + 1):
                    header_cell = ws.cell(header_row, col_idx)
                    header_value = header_cell.value or f"COL_{col_idx}"
                    headers.append(str(header_value).strip())
                
                print(f"Cabeçalhos encontrados: {len(headers)}")
                for i, header in enumerate(headers):
                    if header:  # Só mostrar cabeçalhos não vazios
                        print(f"  Coluna {i+1}: {header}")
                
                # Analisar fórmulas
                formulas = []
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':  # célula com fórmula
                            formulas.append({
                                'cell': cell.coordinate,
                                'formula': cell.value,
                                'column': cell.column_letter
                            })
                
                print(f"Fórmulas encontradas: {len(formulas)}")
                if formulas:
                    print("Primeiras fórmulas:")
                    for formula in formulas[:5]:
                        print(f"  {formula['cell']}: {formula['formula']}")
                
                # Amostra de dados
                print("\nAmostra de dados (primeiras 5 linhas):")
                data_start_row = header_row + 1
                for row_idx in range(data_start_row, min(data_start_row + 5, ws.max_row + 1)):
                    row_data = {}
                    has_data = False
                    for col_idx, header in enumerate(headers):
                        if col_idx < ws.max_column:
                            cell = ws.cell(row_idx, col_idx + 1)
                            value = cell.value
                            if value is not None:
                                row_data[header] = value
                                has_data = True
                    
                    if has_data:
                        print(f"  Linha {row_idx}: {row_data}")
                
                # Salvar análise
                analysis = {
                    'file_path': file_path,
                    'sheet_name': sheet_name,
                    'format': 'xlsx',
                    'total_rows': ws.max_row,
                    'total_columns': ws.max_column,
                    'header_row': header_row,
                    'headers': headers,
                    'formulas_count': len(formulas),
                    'formulas': formulas[:10],  # Primeiras 10 fórmulas
                    'data_sample': []  # Preenchido abaixo
                }
                
                # Adicionar amostra de dados
                for row_idx in range(data_start_row, min(data_start_row + 3, ws.max_row + 1)):
                    row_data = {}
                    for col_idx, header in enumerate(headers):
                        if col_idx < ws.max_column:
                            cell = ws.cell(row_idx, col_idx + 1)
                            value = cell.value
                            row_data[header] = value
                    
                    analysis['data_sample'].append(row_data)
                
                wb.close()
                return analysis
                
            else:
                print("Não foi possível encontrar linha de cabeçalho")
                return None
                
        elif file_path.endswith('.xlsb'):
            print(f"Formato: XLSB (não suportado para análise de fórmulas)")
            
            # Tentar ler com pandas para dados básicos
            try:
                if sheet_name:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                else:
                    df = pd.read_excel(file_path)
                
                print(f"Linhas: {len(df)}")
                print(f"Colunas: {len(df.columns)}")
                print(f"Colunas: {list(df.columns)}")
                
                # Amostra de dados
                print("\nAmostra de dados:")
                print(df.head().to_string())
                
                analysis = {
                    'file_path': file_path,
                    'sheet_name': sheet_name or 'default',
                    'format': 'xlsb',
                    'total_rows': len(df),
                    'total_columns': len(df.columns),
                    'headers': list(df.columns),
                    'formulas_count': 0,
                    'formulas': [],
                    'data_sample': df.head().to_dict('records')
                }
                
                return analysis
                
            except Exception as e:
                print(f"Erro ao ler XLSB: {e}")
                return None
                
    except Exception as e:
        print(f"Erro ao analisar planilha: {e}")
        return None

def analyze_all_new_sheets():
    """Analisa todas as novas planilhas"""
    print("ANÁLISE DAS NOVAS PLANILHAS")
    print("="*80)
    
    # Novas planilhas identificadas
    new_sheets = [
        '/home/haumea/Projects/dashboard-aprovacao-leticia/data/1QZ ABRIL 2026 - VEXPENSES.xlsx',
        '/home/haumea/Projects/dashboard-aprovacao-leticia/data/CARGA 1 QZ MAIO 26 VEXPENSES EQS.xlsx'
    ]
    
    analyses = []
    
    for file_path in new_sheets:
        print(f"\nAnalisando arquivo: {file_path}")
        
        try:
            if file_path.endswith('.xlsx'):
                wb = load_workbook(file_path, data_only=False)
                sheet_names = wb.sheetnames
                print(f"Abas encontradas: {sheet_names}")
                
                for sheet_name in sheet_names:
                    analysis = analyze_sheet_structure(file_path, sheet_name)
                    if analysis:
                        analyses.append(analysis)
                
                wb.close()
            else:
                analysis = analyze_sheet_structure(file_path)
                if analysis:
                    analyses.append(analysis)
                    
        except Exception as e:
            print(f"Erro ao processar {file_path}: {e}")
    
    return analyses

def compare_with_api_data(analyses):
    """Compara dados das planilhas com dados da API"""
    print("\n" + "="*80)
    print("COMPARANDO COM DADOS DA API")
    print("="*80)
    
    # Carregar dados da API que já temos
    try:
        with open('/home/haumea/Projects/dashboard-aprovacao-leticia/investigation-docs/expenses_breakthrough.json', 'r', encoding='utf-8') as f:
            api_data = json.load(f)
        
        print("Dados da API carregados")
        
        # Analisar períodos encontrados nas planilhas
        for analysis in analyses:
            print(f"\nComparando planilha: {analysis['sheet_name']}")
            
            # Procurar campos financeiros
            financial_headers = []
            for header in analysis['headers']:
                header_lower = header.lower()
                if any(keyword in header_lower for keyword in ['saldo', '1qz', 'carga', 'reembolso', 'adiant', 'valor']):
                    financial_headers.append(header)
            
            print(f"  Campos financeiros encontrados: {financial_headers}")
            
            # Verificar amostra de dados
            if analysis['data_sample']:
                print(f"  Amostra de dados:")
                for i, row in enumerate(analysis['data_sample'][:3]):
                    print(f"    Linha {i+1}:")
                    for header, value in row.items():
                        if header in financial_headers and value is not None:
                            print(f"      {header}: {value}")
            
    except Exception as e:
        print(f"Erro ao carregar dados da API: {e}")

def extract_financial_patterns(analyses):
    """Extrai padrões financeiros das novas planilhas"""
    print("\n" + "="*80)
    print("EXTRAINDO PADRÕES FINANCEIROS")
    print("="*80)
    
    patterns = {
        'financial_fields': set(),
        'calculation_formulas': [],
        'periods': set(),
        'users': set()
    }
    
    for analysis in analyses:
        print(f"\nAnalisando padrões em: {analysis['sheet_name']}")
        
        # Campos financeiros
        for header in analysis['headers']:
            header_lower = header.lower()
            if any(keyword in header_lower for keyword in ['saldo', '1qz', 'carga', 'reembolso', 'adiant', 'valor']):
                patterns['financial_fields'].add(header)
        
        # Fórmulas
        for formula in analysis.get('formulas', []):
            if any(keyword in formula['formula'].lower() for keyword in ['sum', 'subtotal', 'if', 'vlookup', 'xlookup']):
                patterns['calculation_formulas'].append(formula)
        
        # Períodos (nos dados)
        for row in analysis.get('data_sample', []):
            for header, value in row.items():
                if isinstance(value, str):
                    value_lower = value.lower()
                    if any(keyword in value_lower for keyword in ['abril', 'maio', '2026', 'quinzena', '1qz']):
                        patterns['periods'].add(value)
        
        # Usuários (nos dados)
        for row in analysis.get('data_sample', []):
            for header, value in row.items():
                if isinstance(value, str) and len(value) > 3:
                    # Provável nome de usuário
                    if any(keyword in header.lower() for keyword in ['nome', 'portador', 'usuário', 'colaborador']):
                        patterns['users'].add(value)
    
    # Converter sets para lists
    patterns['financial_fields'] = list(patterns['financial_fields'])
    patterns['periods'] = list(patterns['periods'])
    patterns['users'] = list(patterns['users'])
    
    print(f"\nResumo dos padrões encontrados:")
    print(f"  Campos financeiros: {patterns['financial_fields']}")
    print(f"  Períodos: {patterns['periods']}")
    print(f"  Usuários: {patterns['users'][:5]}...")  # Primeiros 5
    print(f"  Fórmulas de cálculo: {len(patterns['calculation_formulas'])}")
    
    return patterns

def main():
    """Função principal"""
    print("ANÁLISE DAS NOVAS PLANILHAS - ATUALIZADAS HOJE")
    print("="*80)
    print("Investigando estrutura e dados das planilhas recentes")
    print("="*80)
    
    # 1. Analisar estrutura das novas planilhas
    analyses = analyze_all_new_sheets()
    
    # 2. Comparar com dados da API
    compare_with_api_data(analyses)
    
    # 3. Extrair padrões financeiros
    patterns = extract_financial_patterns(analyses)
    
    # 4. Salvar resultados
    results = {
        'analysis_date': datetime.now().isoformat(),
        'sheets_analyzed': len(analyses),
        'analyses': analyses,
        'patterns': patterns,
        'summary': {
            'total_sheets': len(analyses),
            'financial_fields_found': len(patterns['financial_fields']),
            'formulas_found': len(patterns['calculation_formulas']),
            'periods_identified': len(patterns['periods'])
        }
    }
    
    output_file = '/home/haumea/Projects/dashboard-aprovacao-leticia/investigation-docs/new_sheets_analysis.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    print(f"\n\nAnálise salva em: {output_file}")
    print("\n" + "="*80)
    print("ANÁLISE DAS NOVAS PLANILHAS CONCLUÍDA!")
    print("="*80)

if __name__ == "__main__":
    main()
