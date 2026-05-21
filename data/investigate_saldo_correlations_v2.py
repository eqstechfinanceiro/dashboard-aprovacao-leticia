import requests
import json
import openpyxl
from datetime import datetime

API_KEY = "N2ntX8mUF7AvjVcyT6DZwnOQXRvYgel9pCirCGKKETO5R70HEZ8UdbQA0Nt8"
BASE_URL = "https://api.vexpenses.com/v2"

headers = {
    "Authorization": API_KEY,
    "Content-Type": "application/json",
    "Accept": "application/json"
}

def get_expenses_for_user(user_id, start_date, end_date):
    """Obtém expenses para um usuário específico"""
    params = {
        "search": f"date:{start_date},{end_date}",
        "searchFields": "date:between",
        "searchJoin": "and",
        "paginate": "true",
        "page": "1",
        "per_page": "200",
        "include": "expense_type,costs_center,payment_method,user"
    }
    
    try:
        url = f"{BASE_URL}/expenses"
        response = requests.get(url, headers=headers, params=params, timeout=30)
        
        if response.status_code == 200:
            data = response.json()
            if 'data' in data:
                all_expenses = data['data']
                user_expenses = [exp for exp in all_expenses if exp.get('user_id') == user_id]
                return user_expenses
    except Exception as e:
        print(f"Erro: {e}")
    
    return []

def safe_float(value):
    """Converte valor para float de forma segura"""
    try:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            # Remove caracteres não numéricos
            clean = ''.join(c for c in value if c.isdigit() or c == '.' or c == '-')
            if clean:
                return float(clean)
        return 0.0
    except:
        return 0.0

def extract_saldo_values_from_sheet():
    """Extrai valores de SALDO da planilha"""
    print("EXTRAINDO VALORES DE SALDO DA PLANILHA")
    print("="*60)
    
    file_path = '/home/haumea/Projects/dashboard-aprovacao-leticia/data/1QZ ABRIL 2026 - VEXPENSES.xlsx'
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb["1 QZ VEXPENSES 04_2026"]
        
        user_saldo_data = {}
        
        for row in range(6, min(56, ws.max_row + 1)):
            nome = ws.cell(row, 2).value
            saldo_reembolsar = ws.cell(row, 9).value
            saldo_final = ws.cell(row, 10).value
            saldo_cartao = ws.cell(row, 12).value
            quinzena_qz = ws.cell(row, 11).value
            
            if nome and isinstance(nome, str):
                user_saldo_data[nome] = {
                    'saldo_reembolsar': safe_float(saldo_reembolsar),
                    'saldo_final': safe_float(saldo_final),
                    'saldo_cartao': safe_float(saldo_cartao),
                    'quinzena_qz': safe_float(quinzena_qz),
                    'row': row
                }
        
        wb.close()
        
        print(f"Usuários extraídos: {len(user_saldo_data)}")
        
        # Mostrar exemplos
        for nome, data in list(user_saldo_data.items())[:5]:
            print(f"\n{nome}:")
            print(f"  SALDO REEMBOLSAR: R$ {data['saldo_reembolsar']:.2f}")
            print(f"  SALDO FINAL: R$ {data['saldo_final']:.2f}")
            print(f"  SALDO CARTAO: R$ {data['saldo_cartao']:.2f}")
            print(f"  1QZ: R$ {data['quinzena_qz']:.2f}")
        
        return user_saldo_data
        
    except Exception as e:
        print(f"Erro: {e}")
        import traceback
        traceback.print_exc()
        return {}

def analyze_correlations(user_saldo_data):
    """Analisa correlações entre SALDO e métricas da API"""
    print(f"\nANALISANDO CORRELAÇÕES")
    print("="*60)
    
    correlations = []
    
    # Usar usuários mapeados anteriormente
    user_id_mapping = {
        'JONAS CAVALCANTI': 895945,
        'RODRIGO CESAR': 895946,
        'CAIO FRANCESCONI': 895947
    }
    
    for planilha_name, user_id in user_id_mapping.items():
        # Procurar usuário na planilha
        matched_user = None
        for nome, saldo_data in user_saldo_data.items():
            if planilha_name in nome.upper() or nome.upper() in planilha_name:
                matched_user = (nome, saldo_data)
                break
        
        if not matched_user:
            continue
        
        nome, saldo_data = matched_user
        print(f"\nUsuário: {nome} -> ID {user_id}")
        
        # Obter expenses
        expenses = get_expenses_for_user(user_id, '2026-01-01', '2026-04-30')
        
        if not expenses:
            print(f"  Nenhuma expense encontrada")
            continue
        
        # Calcular métricas
        annual_total = sum(exp.get('value', 0) for exp in expenses if exp.get('value', 0) > 0)
        
        reimbursable_expenses = [exp for exp in expenses if exp.get('reimbursable', False)]
        reimbursable_total = sum(exp.get('value', 0) for exp in reimbursable_expenses if exp.get('value', 0) > 0)
        
        payment_methods = {}
        for exp in expenses:
            pm = exp.get('payment_method', {})
            pm_name = pm.get('name', 'Unknown')
            pm_value = exp.get('value', 0)
            
            if pm_name not in payment_methods:
                payment_methods[pm_name] = 0
            payment_methods[pm_name] += pm_value
        
        print(f"  Total anual: R$ {annual_total:.2f}")
        print(f"  Total reembolsável: R$ {reimbursable_total:.2f}")
        print(f"  Métodos: {payment_methods}")
        
        # Calcular correlações
        saldo_final = saldo_data['saldo_final']
        saldo_cartao = saldo_data['saldo_cartao']
        saldo_reembolsar = saldo_data['saldo_reembolsar']
        
        saldo_final_ratio = saldo_final / annual_total if annual_total > 0 else 0
        saldo_reembolsar_ratio = saldo_reembolsar / reimbursable_total if reimbursable_total > 0 else 0
        
        card_total = 0
        for pm_name, pm_value in payment_methods.items():
            if 'CARTÃO' in pm_name.upper() or 'CARD' in pm_name.upper():
                card_total += pm_value
        
        saldo_cartao_ratio = saldo_cartao / card_total if card_total > 0 else 0
        
        print(f"\n  CORRELAÇÕES:")
        print(f"    SALDO FINAL / Anual: {saldo_final_ratio:.4f}")
        print(f"    SALDO REEMBOLSAR / Reembolsável: {saldo_reembolsar_ratio:.4f}")
        print(f"    SALDO CARTÃO / Cartão: {saldo_cartao_ratio:.4f}")
        
        correlations.append({
            'user_id': user_id,
            'nome': nome,
            'saldo_final': saldo_final,
            'saldo_cartao': saldo_cartao,
            'saldo_reembolsar': saldo_reembolsar,
            'annual_total': annual_total,
            'reimbursable_total': reimbursable_total,
            'card_total': card_total,
            'saldo_final_ratio': saldo_final_ratio,
            'saldo_cartao_ratio': saldo_cartao_ratio,
            'saldo_reembolsar_ratio': saldo_reembolsar_ratio
        })
    
    return correlations

def main():
    """Função principal"""
    print("INVESTIGAÇÃO DE CORRELAÇÕES")
    print("="*80)
    
    user_saldo_data = extract_saldo_values_from_sheet()
    
    if not user_saldo_data:
        print("Nenhum dado extraído")
        return
    
    correlations = analyze_correlations(user_saldo_data)
    
    results = {
        'investigation_date': datetime.now().isoformat(),
        'user_saldo_data': user_saldo_data,
        'correlations': correlations
    }
    
    output_file = '/home/haumea/Projects/dashboard-aprovacao-leticia/investigation-docs/saldo_correlations.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    print(f"\nSalvo em: {output_file}")

if __name__ == "__main__":
    main()
