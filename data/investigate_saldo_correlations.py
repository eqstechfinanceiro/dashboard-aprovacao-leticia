import requests
import json
import openpyxl
from datetime import datetime

API_KEY = "N2ntX8mUF7AvjVcyT6DZwnOQXRvYgel9pCirCGKKETO5R70HEZ8UdbQA0Nt8"
BASE_URL = "https://api.vexpenses.com/v2"

headers = {
    "Authorization": API_KEY,
    "Content-Type": "application/json",
    "Accept": "application/json"
}

def get_expenses_for_user(user_id, start_date, end_date):
    """Obtém expenses para um usuário específico"""
    params = {
        "search": f"date:{start_date},{end_date}",
        "searchFields": "date:between",
        "searchJoin": "and",
        "paginate": "true",
        "page": "1",
        "per_page": "200",
        "include": "expense_type,costs_center,payment_method,user"
    }
    
    try:
        url = f"{BASE_URL}/expenses"
        response = requests.get(url, headers=headers, params=params, timeout=30)
        
        if response.status_code == 200:
            data = response.json()
            if 'data' in data:
                all_expenses = data['data']
                # Filtrar por usuário no cliente
                user_expenses = [exp for exp in all_expenses if exp.get('user_id') == user_id]
                return user_expenses
    except Exception as e:
        print(f"Erro: {e}")
    
    return []

def extract_saldo_values_from_sheet():
    """Extrai valores de SALDO da planilha para usuários específicos"""
    print("EXTRAINDO VALORES DE SALDO DA PLANILHA")
    print("="*60)
    
    file_path = '/home/haumea/Projects/dashboard-aprovacao-leticia/data/1QZ ABRIL 2026 - VEXPENSES.xlsx'
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb["1 QZ VEXPENSES 04_2026"]
        
        # Encontrar colunas de SALDO (baseado na estrutura conhecida)
        # Coluna I = SALDO REEMBOLSAR
        # Coluna J = SALDO FINAL
        # Coluna L = SALDO CARTAO
        
        user_saldo_data = {}
        
        # Ler primeiras 50 linhas de dados
        for row in range(6, min(56, ws.max_row + 1)):
            nome = ws.cell(row, 2).value  # Coluna B = PORTADOR
            saldo_reembolsar = ws.cell(row, 9).value  # Coluna I
            saldo_final = ws.cell(row, 10).value  # Coluna J
            saldo_cartao = ws.cell(row, 12).value  # Coluna L
            quinzena_qz = ws.cell(row, 11).value  # Coluna K = 1QZ
            
            if nome and isinstance(nome, str):
                # Extrair CPF se possível
                cpf = ws.cell(row, 3).value  # Coluna C = CPF
                
                user_saldo_data[nome] = {
                    'cpf': cpf,
                    'saldo_reembolsar': saldo_reembolsar if saldo_reembolsar is not None else 0,
                    'saldo_final': saldo_final if saldo_final is not None else 0,
                    'saldo_cartao': saldo_cartao if saldo_cartao is not None else 0,
                    'quinzena_qz': quinzena_qz if quinzena_qz is not None else 0,
                    'row': row
                }
        
        wb.close()
        
        print(f"Usuários extraídos: {len(user_saldo_data)}")
        
        # Mostrar exemplos
        for nome, data in list(user_saldo_data.items())[:5]:
            print(f"\n{nome}:")
            print(f"  SALDO REEMBOLSAR: R$ {data['saldo_reembolsar']:.2f}")
            print(f"  SALDO FINAL: R$ {data['saldo_final']:.2f}")
            print(f"  SALDO CARTAO: R$ {data['saldo_cartao']:.2f}")
            print(f"  1QZ: R$ {data['quinzena_qz']:.2f}")
        
        return user_saldo_data
        
    except Exception as e:
        print(f"Erro: {e}")
        import traceback
        traceback.print_exc()
        return {}

def analyze_correlations(user_saldo_data):
    """Analisa correlações entre SALDO e métricas da API"""
    print(f"\nANALISANDO CORRELAÇÕES")
    print("="*60)
    
    correlations = []
    
    for nome, saldo_data in user_saldo_data.items():
        # Usar os usuários que conseguimos mapear anteriormente
        # Jonas Cavalcanti (895945), Rodrigo Cesar (895946), Caio Francesconi (895947)
        
        user_id_mapping = {
            'JONAS CAVALCANTI': 895945,
            'RODRIGO CESAR': 895946,
            'CAIO FRANCESCONI': 895947
        }
        
        # Tentar encontrar correspondência por nome
        for planilha_name, user_id in user_id_mapping.items():
            if planilha_name in nome.upper() or nome.upper() in planilha_name:
                print(f"\nUsuário mapeado: {nome} -> ID {user_id}")
                
                # Obter expenses do usuário
                expenses = get_expenses_for_user(user_id, '2026-01-01', '2026-04-30')
                
                if not expenses:
                    print(f"  Nenhuma expense encontrada")
                    continue
                
                # Calcular métricas
                annual_total = sum(exp.get('value', 0) for exp in expenses if exp.get('value', 0) > 0)
                
                # Expenses reembolsáveis
                reimbursable_expenses = [exp for exp in expenses if exp.get('reimbursable', False)]
                reimbursable_total = sum(exp.get('value', 0) for exp in reimbursable_expenses if exp.get('value', 0) > 0)
                
                # Expenses por método de pagamento
                payment_methods = {}
                for exp in expenses:
                    pm = exp.get('payment_method', {})
                    pm_name = pm.get('name', 'Unknown')
                    pm_value = exp.get('value', 0)
                    
                    if pm_name not in payment_methods:
                        payment_methods[pm_name] = 0
                    payment_methods[pm_name] += pm_value
                
                print(f"  Total anual: R$ {annual_total:.2f}")
                print(f"  Total reembolsável: R$ {reimbursable_total:.2f}")
                print(f"  Métodos de pagamento: {payment_methods}")
                
                # Calcular correlações
                saldo_final = saldo_data['saldo_final']
                saldo_cartao = saldo_data['saldo_cartao']
                saldo_reembolsar = saldo_data['saldo_reembolsar']
                
                # Correlação SALDO FINAL vs anual
                if annual_total > 0:
                    saldo_final_ratio = saldo_final / annual_total
                    print(f"\n  CORRELAÇÃO SALDO FINAL:")
                    print(f"    Saldo Final: R$ {saldo_final:.2f}")
                    print(f"    Anual: R$ {annual_total:.2f}")
                    print(f"    Razão: {saldo_final_ratio:.4f}")
                
                # Correlação SALDO CARTAO vs métodos de pagamento
                if 'Cartão' in str(payment_methods):
                    card_total = payment_methods.get('Cartão', 0)
                    if card_total > 0:
                        saldo_cartao_ratio = saldo_cartao / card_total
                        print(f"\n  CORRELAÇÃO SALDO CARTÃO:")
                        print(f"    Saldo Cartão: R$ {saldo_cartao:.2f}")
                        print(f"    Total Cartão: R$ {card_total:.2f}")
                        print(f"    Razão: {saldo_cartao_ratio:.4f}")
                
                # Correlação SALDO REEMBOLSAR vs reembolsável
                if reimbursable_total > 0:
                    saldo_reembolsar_ratio = saldo_reembolsar / reimbursable_total
                    print(f"\n  CORRELAÇÃO SALDO REEMBOLSAR:")
                    print(f"    Saldo Reembolsar: R$ {saldo_reembolsar:.2f}")
                    print(f"    Reembolsável: R$ {reimbursable_total:.2f}")
                    print(f"    Razão: {saldo_reembolsar_ratio:.4f}")
                
                correlations.append({
                    'user_id': user_id,
                    'nome': nome,
                    'saldo_final': saldo_final,
                    'saldo_cartao': saldo_cartao,
                    'saldo_reembolsar': saldo_reembolsar,
                    'annual_total': annual_total,
                    'reimbursable_total': reimbursable_total,
                    'saldo_final_ratio': saldo_final_ratio if annual_total > 0 else 0,
                    'saldo_cartao_ratio': saldo_cartao_ratio if 'Cartão' in str(payment_methods) and card_total > 0 else 0,
                    'saldo_reembolsar_ratio': saldo_reembolsar_ratio if reimbursable_total > 0 else 0
                })
                
                break
    
    return correlations

def calculate_formulas_from_correlations(correlations):
    """Calcula fórmulas baseadas nas correlações encontradas"""
    print(f"\nCALCULANDO FÓRMULAS DAS CORRELAÇÕES")
    print("="*60)
    
    if not correlations:
        print("Nenhuma correlação encontrada")
        return {}
    
    # Calcular médias das razões
    avg_saldo_final_ratio = sum(c['saldo_final_ratio'] for c in correlations) / len(correlations)
    avg_saldo_cartao_ratio = sum(c['saldo_cartao_ratio'] for c in correlations if c['saldo_cartao_ratio'] > 0) / len([c for c in correlations if c['saldo_cartao_ratio'] > 0]) if any(c['saldo_cartao_ratio'] > 0 for c in correlations) else 0
    avg_saldo_reembolsar_ratio = sum(c['saldo_reembolsar_ratio'] for c in correlations if c['saldo_reembolsar_ratio'] > 0) / len([c for c in correlations if c['saldo_reembolsar_ratio'] > 0]) if any(c['saldo_reembolsar_ratio'] > 0 for c in correlations) else 0
    
    print(f"\nFÓRMULAS CALCULADAS:")
    print(f"  SALDO FINAL = annual_total * {avg_saldo_final_ratio:.4f}")
    print(f"  SALDO CARTAO = card_total * {avg_saldo_cartao_ratio:.4f}")
    print(f"  SALDO REEMBOLSAR = reimbursable_total * {avg_saldo_reembolsar_ratio:.4f}")
    
    return {
        'saldo_final_formula': f'annual_total * {avg_saldo_final_ratio:.4f}',
        'saldo_cartao_formula': f'card_total * {avg_saldo_cartao_ratio:.4f}',
        'saldo_reembolsar_formula': f'reimbursable_total * {avg_saldo_reembolsar_ratio:.4f}',
        'avg_saldo_final_ratio': avg_saldo_final_ratio,
        'avg_saldo_cartao_ratio': avg_saldo_cartao_ratio,
        'avg_saldo_reembolsar_ratio': avg_saldo_reembolsar_ratio
    }

def main():
    """Função principal"""
    print("INVESTIGAÇÃO DE CORRELAÇÕES - DADOS DE SALDO")
    print("="*80)
    
    # 1. Extrair valores de SALDO da planilha
    user_saldo_data = extract_saldo_values_from_sheet()
    
    if not user_saldo_data:
        print("Nenhum dado de SALDO extraído")
        return
    
    # 2. Analisar correlações
    correlations = analyze_correlations(user_saldo_data)
    
    # 3. Calcular fórmulas
    formulas = calculate_formulas_from_correlations(correlations)
    
    # 4. Salvar resultados
    results = {
        'investigation_date': datetime.now().isoformat(),
        'user_saldo_data': user_saldo_data,
        'correlations': correlations,
        'formulas': formulas,
        'conclusion': ''
    }
    
    if correlations:
        results['conclusion'] = 'CORRELAÇÕES ENCONTRADAS - FÓRMULAS CALCULADAS'
    else:
        results['conclusion'] = 'NENHUMA CORRELAÇÃO ENCONTRADA'
    
    output_file = '/home/haumea/Projects/dashboard-aprovacao-leticia/investigation-docs/saldo_correlations.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    print(f"\n\nResultados salvos em: {output_file}")
    
    print("\n" + "="*80)
    print("INVESTIGAÇÃO CONCLUÍDA!")
    print("="*80)

if __name__ == "__main__":
    main()
