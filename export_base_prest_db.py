import json
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Fetch data from API - save to file to avoid stdout truncation
json_file = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\base_prest_dump.json"
ps_script = f'''
$session = New-Object Microsoft.PowerShell.Commands.WebRequestSession
$body = @{{ email = "italo.medrado@eqsengenharia.com.br"; password = "EQSeng4292@" }} | ConvertTo-Json
Invoke-RestMethod -Uri "http://localhost:3000/api/auth/login" -Method POST -Body $body -ContentType "application/json" -WebSession $session | Out-Null
$r = Invoke-RestMethod -Uri "http://localhost:3000/api/debug-base-prest-dump" -Method GET -WebSession $session
$r | ConvertTo-Json -Depth 10 | Out-File -FilePath "{json_file}" -Encoding UTF8
'''
print("Fetching data from API...")
subprocess.run(['powershell.exe', '-NoProfile', '-Command', ps_script], timeout=300)
with open(json_file, 'r', encoding='utf-8-sig') as f:
    data = json.load(f)

expenses = data['expenses']
cpf_summary = data['cpf_summary']
report_summary = data['report_summary']

print(f"Got {len(expenses)} expenses, {len(cpf_summary)} CPFs, {len(report_summary)} reports")

# Create workbook
wb = Workbook()

# ===== Sheet 1: CPF Summary =====
ws1 = wb.active
ws1.title = "CPF Summary"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

headers1 = ["CPF", "Nome", "Qtd Relatórios", "Qtd Despesas", "Total Prestação"]
for col, h in enumerate(headers1, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

for row_idx, row in enumerate(cpf_summary, 2):
    ws1.cell(row=row_idx, column=1, value=row['user_cpf']).border = thin_border
    ws1.cell(row=row_idx, column=2, value=row['user_name']).border = thin_border
    ws1.cell(row=row_idx, column=3, value=int(row['report_count'])).border = thin_border
    ws1.cell(row=row_idx, column=4, value=int(row['expense_count'])).border = thin_border
    val_cell = ws1.cell(row=row_idx, column=5, value=float(row['total_prestacao']))
    val_cell.number_format = '#,##0.00'
    val_cell.border = thin_border

ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 35
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 15
ws1.column_dimensions['E'].width = 18
ws1.auto_filter.ref = f"A1:E{len(cpf_summary)+1}"

# ===== Sheet 2: Report Summary =====
ws2 = wb.create_sheet("Report Summary")
headers2 = ["Report ID", "Report Name", "Status", "CPF", "Nome", "Approval Date", "Qtd Despesas", "Total Valor"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

for row_idx, row in enumerate(report_summary, 2):
    ws2.cell(row=row_idx, column=1, value=int(row['report_id'])).border = thin_border
    ws2.cell(row=row_idx, column=2, value=row['report_name']).border = thin_border
    ws2.cell(row=row_idx, column=3, value=row['status']).border = thin_border
    ws2.cell(row=row_idx, column=4, value=row['user_cpf']).border = thin_border
    ws2.cell(row=row_idx, column=5, value=row['user_name']).border = thin_border
    ws2.cell(row=row_idx, column=6, value=row['approval_date']).border = thin_border
    ws2.cell(row=row_idx, column=7, value=int(row['expense_count'])).border = thin_border
    val_cell = ws2.cell(row=row_idx, column=8, value=float(row['total_value']))
    val_cell.number_format = '#,##0.00'
    val_cell.border = thin_border

for col in range(1, 9):
    ws2.column_dimensions[get_column_letter(col)].width = [12, 25, 12, 15, 30, 22, 12, 15][col-1]
ws2.auto_filter.ref = f"A1:H{len(report_summary)+1}"

# ===== Sheet 3: All Expenses (BASE PREST) =====
ws3 = wb.create_sheet("BASE PREST (DB)")
headers3 = [
    "Report ID", "Report Name", "Status", "CPF", "Nome",
    "Created At", "Updated At", "Approval Date", "Report Total",
    "Expense ID", "Expense Description", "Expense Value", "Expense Date", "Expense Status"
]
for col, h in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

for row_idx, row in enumerate(expenses, 2):
    ws3.cell(row=row_idx, column=1, value=int(row['report_id']) if row['report_id'] else None).border = thin_border
    ws3.cell(row=row_idx, column=2, value=row['report_name']).border = thin_border
    ws3.cell(row=row_idx, column=3, value=row['status']).border = thin_border
    ws3.cell(row=row_idx, column=4, value=row['user_cpf']).border = thin_border
    ws3.cell(row=row_idx, column=5, value=row['user_name']).border = thin_border
    ws3.cell(row=row_idx, column=6, value=str(row['created_at']) if row['created_at'] else None).border = thin_border
    ws3.cell(row=row_idx, column=7, value=str(row['updated_at']) if row['updated_at'] else None).border = thin_border
    ws3.cell(row=row_idx, column=8, value=row['approval_date']).border = thin_border
    if row['report_total'] is not None:
        c = ws3.cell(row=row_idx, column=9, value=float(row['report_total']))
        c.number_format = '#,##0.00'
        c.border = thin_border
    else:
        ws3.cell(row=row_idx, column=9).border = thin_border
    ws3.cell(row=row_idx, column=10, value=int(row['expense_id']) if row['expense_id'] else None).border = thin_border
    ws3.cell(row=row_idx, column=11, value=row['expense_description']).border = thin_border
    if row['expense_value'] is not None:
        c = ws3.cell(row=row_idx, column=12, value=float(row['expense_value']))
        c.number_format = '#,##0.00'
        c.border = thin_border
    else:
        ws3.cell(row=row_idx, column=12).border = thin_border
    ws3.cell(row=row_idx, column=13, value=str(row['expense_date']) if row['expense_date'] else None).border = thin_border
    ws3.cell(row=row_idx, column=14, value=row['expense_status']).border = thin_border

widths3 = [12, 25, 12, 15, 30, 22, 22, 22, 15, 12, 30, 15, 14, 15]
for col in range(1, len(headers3)+1):
    ws3.column_dimensions[get_column_letter(col)].width = widths3[col-1]
ws3.auto_filter.ref = f"A1:N{len(expenses)+1}"
ws3.freeze_panes = "A2"

# Save
output_path = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\BASE_PREST_DB_EXPORT.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
