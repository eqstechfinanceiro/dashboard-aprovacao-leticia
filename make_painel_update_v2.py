import psycopg2
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DB_URL = "postgresql://neondb_owner:npg_iItZN95svyEG@ep-restless-voice-amrrz188-pooler.c-5.us-east-1.aws.neon.tech/neondb?sslmode=require"
CONTROLE = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - AGOSTO 2026.xlsx"
OUTPUT = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\PRESTACAO_PAINEL_UPDATE_V2.xlsx"

# ITAU payment_method_id = 627401
ITAU_PMID = '627401'

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
money_fmt = 'R$ #,##0.00;[Red]-R$ #,##0.00'

print("Querying DB (excluding Itau payment_method_id=627401)...")
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()

cur.execute("""
    SELECT r.user_cpf,
           r.user_name,
           COALESCE(SUM(e.value), 0) as total_prestacao,
           COUNT(e.id) as expense_count,
           COUNT(DISTINCT r.id) as report_count
    FROM prestacao_reports r
    LEFT JOIN prestacao_expenses e ON e.report_id = r.id
    WHERE r.user_cpf IS NOT NULL
      AND (r.status ILIKE 'Aprovado' OR r.status ILIKE 'Enviado')
      AND NOT (r.name ILIKE '%FATURA%' OR r.name ILIKE '%CARTAO%' OR r.name ILIKE '%CARTAO%')
      AND (e.raw_data->>'payment_method_id' IS NULL OR e.raw_data->>'payment_method_id' != '627401')
    GROUP BY r.user_cpf, r.user_name
    ORDER BY r.user_name
""")

db_data = {}
for row in cur.fetchall():
    cpf = str(row[0]).strip()
    db_data[cpf] = {
        "nome": row[1],
        "total": float(row[2]),
        "expense_count": int(row[3]),
        "report_count": int(row[4]),
    }
cur.close()
conn.close()
print(f"  {len(db_data)} CPFs from DB")

# Load PAINEL
print("Loading PAINEL...")
wb_pn = openpyxl.load_workbook(CONTROLE, read_only=True, data_only=True)
ws_pn = wb_pn['PAINEL']
painel_rows = []
for row in ws_pn.iter_rows(min_row=12, values_only=True):
    cpf = str(row[2]).strip() if len(row) > 2 and row[2] else ''
    if not cpf or cpf == 'None':
        continue
    painel_rows.append({
        "cpf": cpf,
        "nome": str(row[1]).strip() if row[1] else '',
        "painel_prestacao": float(row[16]) if row[16] is not None else 0,
    })
wb_pn.close()
print(f"  {len(painel_rows)} CPFs in PAINEL")

# Build output
print("Building output...")
wb = Workbook()

# Sheet 1: Ready to paste (in PAINEL order)
ws1 = wb.active
ws1.title = "PASTE_COL"
headers = ["CPF", "NOME", "PRESTAÇÃO DB (NOVO)"]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center')
    c.border = thin_border

r = 2
for p in painel_rows:
    cpf = p["cpf"]
    db = db_data.get(cpf)
    nome = db["nome"] if db else p["nome"]
    total = db["total"] if db else 0
    ws1.cell(row=r, column=1, value=cpf).border = thin_border
    ws1.cell(row=r, column=2, value=nome).border = thin_border
    c = ws1.cell(row=r, column=3, value=total)
    c.number_format = money_fmt
    c.border = thin_border
    r += 1

# Add CPFs in DB but not in PAINEL
extra = set(db_data.keys()) - set(p["cpf"] for p in painel_rows)
if extra:
    r += 1
    ws1.cell(row=r, column=1, value="--- CPFs no DB mas não no PAINEL ---").font = Font(bold=True, italic=True)
    r += 1
    for cpf in sorted(extra, key=lambda c: db_data[c]["nome"]):
        d = db_data[cpf]
        ws1.cell(row=r, column=1, value=cpf).border = thin_border
        ws1.cell(row=r, column=2, value=d["nome"]).border = thin_border
        c = ws1.cell(row=r, column=3, value=d["total"])
        c.number_format = money_fmt
        c.border = thin_border
        r += 1

ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 35
ws1.column_dimensions['C'].width = 22
ws1.freeze_panes = "A2"

# Sheet 2: Comparison side-by-side
ws2 = wb.create_sheet("COMPARACAO")
headers2 = ["CPF", "NOME", "PAINEL (R$)", "DB NOVO (R$)", "DIFERENÇA (R$)"]
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center')
    c.border = thin_border

r = 2
total_painel = 0
total_db = 0
for p in painel_rows:
    cpf = p["cpf"]
    db = db_data.get(cpf)
    pn_val = p["painel_prestacao"]
    db_val = db["total"] if db else 0
    diff = db_val - pn_val
    nome = db["nome"] if db else p["nome"]
    total_painel += pn_val
    total_db += db_val

    ws2.cell(row=r, column=1, value=cpf).border = thin_border
    ws2.cell(row=r, column=2, value=nome).border = thin_border
    c = ws2.cell(row=r, column=3, value=pn_val); c.number_format = money_fmt; c.border = thin_border
    c = ws2.cell(row=r, column=4, value=db_val); c.number_format = money_fmt; c.border = thin_border
    c = ws2.cell(row=r, column=5, value=diff); c.number_format = money_fmt; c.border = thin_border
    if abs(diff) > 0.01:
        fill_color = "FFC7CE" if diff > 0 else "C6EFCE"
        c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    r += 1

# Totals row
c = ws2.cell(row=r, column=2, value="TOTAL"); c.font = Font(bold=True); c.border = thin_border
c = ws2.cell(row=r, column=3, value=total_painel); c.number_format = money_fmt; c.font = Font(bold=True); c.border = thin_border
c = ws2.cell(row=r, column=4, value=total_db); c.number_format = money_fmt; c.font = Font(bold=True); c.border = thin_border
c = ws2.cell(row=r, column=5, value=total_db - total_painel); c.number_format = money_fmt; c.font = Font(bold=True); c.border = thin_border

ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 18
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:E{r-1}"

print(f"Saving to {OUTPUT}...")
wb.save(OUTPUT)
print(f"Done! PAINEL: R$ {total_painel:,.2f} | DB: R$ {total_db:,.2f} | Diff: R$ {total_db - total_painel:+,.2f}")
