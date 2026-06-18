#!/usr/bin/env python3
"""
Script rápido para analisar a planilha Excel do VExpenses
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import os

# Caminho do arquivo
file_path = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\controle-api\data\CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsx"

print("=" * 80)
print("ANÁLISE DA PLANILHA: CONTROLE - VEXPENSES - MAIO - 2026")
print("=" * 80)
print(f"\nCaminho: {file_path}")
print(f"Existe: {os.path.exists(file_path)}")
print()

if not os.path.exists(file_path):
    print("ERRO: Arquivo não encontrado!")
    exit(1)

# ============================================================
# 1. LISTAR TODAS AS ABAS DISPONÍVEIS (usando pandas - rápido)
# ============================================================
print("=" * 80)
print("1. ABAS DISPONÍVEIS NA PLANILHA")
print("=" * 80)

xl_file = pd.ExcelFile(file_path)
sheet_names = xl_file.sheet_names

for i, name in enumerate(sheet_names, 1):
    print(f"  {i}. {name}")

print(f"\nTotal de abas: {len(sheet_names)}")
print()

# ============================================================
# 2. IDENTIFICAR LINHA DO CABEÇALHO EM CADA ABA
# ============================================================
print("=" * 80)
print("2. LINHA DO CABEÇALHO POR ABA")
print("=" * 80)

for sheet_name in sheet_names:
    try:
        # Ler as primeiras 10 linhas para detectar cabeçalho
        df_test = pd.read_excel(file_path, sheet_name=sheet_name, nrows=10, header=None)

        if df_test.empty:
            print(f"\n  📄 {sheet_name}: Aba vazia")
            continue

        # Verificar cada linha
        header_row = None
        for idx in range(min(10, len(df_test))):
            row = df_test.iloc[idx]
            non_empty = row.notna().sum()
            text_count = sum(1 for v in row if isinstance(v, str))

            # Se tiver 2+ valores não-vazios e algum texto, provavelmente é cabeçalho
            if non_empty >= 2 and text_count >= 1:
                header_row = idx + 1  # +1 porque índice começa em 0
                sample_values = [str(v) for v in row if pd.notna(v)][:8]
                break

        if header_row:
            print(f"\n  📄 {sheet_name}:")
            print(f"     Cabeçalho na linha: {header_row}")
            print(f"     Amostra de colunas: {sample_values}")
        else:
            print(f"\n  📄 {sheet_name}: Cabeçalho na linha 1 (padrão)")

    except Exception as e:
        print(f"\n  📄 {sheet_name}: Erro ao ler - {e}")

print()

# ============================================================
# 3. FOCAR NA ABA PAINEL - LISTAR COLUNAS
# ============================================================
print("=" * 80)
print("3. ANÁLISE DA ABA 'PAINEL' - COLUNAS")
print("=" * 80)

if 'PAINEL' in sheet_names:
    try:
        df_painel = pd.read_excel(file_path, sheet_name='PAINEL', nrows=20)

        if not df_painel.empty:
            print(f"\nTotal de colunas detectadas: {len(df_painel.columns)}")
            print(f"Cabeçalho detectado automaticamente: {df_painel.columns[0] if len(df_painel.columns) > 0 else 'N/A'}")
            print()
            print("LISTA DE COLUNAS (índice: nome):")
            print("-" * 50)

            for i, col in enumerate(df_painel.columns, 1):
                print(f"  {i:2d}. {col}")

            painel_headers = list(df_painel.columns)
        else:
            print("ABA PAINEL está vazia!")
            painel_headers = []
    except Exception as e:
        print(f"Erro ao ler aba PAINEL: {e}")
        painel_headers = []
else:
    print("ABA 'PAINEL' NÃO ENCONTRADA!")
    painel_headers = []

print()

# ============================================================
# 4. VERIFICAR FÓRMULAS VS VALORES (openpyxl - carregar rápido)
# ============================================================
print("=" * 80)
print("4. ANÁLISE DE FÓRMULAS VS VALORES")
print("=" * 80)

# Colunas a verificar - variações de nomes
search_terms = ['SALDO FINAL', 'SALDO CARTAO', 'SALDO CARTÃO', 'SALDO PRESTAÇÃO', 'SALDO PRESTACAO']

if 'PAINEL' in sheet_names and painel_headers:
    print("\nProcurando colunas de SALDO...")

    # Mapear colunas encontradas
    found_cols = {}
    for idx, col in enumerate(painel_headers, 1):
        col_upper = str(col).upper() if col else ""
        for term in search_terms:
            if term.upper() in col_upper:
                found_cols[col] = idx
                break

    if found_cols:
        print(f"Colunas encontradas: {list(found_cols.keys())}")
    else:
        print("Nenhuma coluna de SALDO encontrada pelo nome.")
        print("Buscando colunas que contenham 'SALDO'...")
        for idx, col in enumerate(painel_headers, 1):
            if col and 'SALDO' in str(col).upper():
                found_cols[col] = idx
        print(f"Encontradas: {list(found_cols.keys())}")

    # Carregar com openpyxl para ver fórmulas
    print("\nCarregando openpyxl para verificar fórmulas...")
    wb = load_workbook(file_path, data_only=False)
    ws = wb['PAINEL']

    print(f"Dimensões da aba PAINEL: {ws.max_row} linhas x {ws.max_column} colunas")

    # Encontrar linha do cabeçalho
    header_row = 1
    for row_idx in range(1, min(10, ws.max_row + 1)):
        first_cell = ws.cell(row=row_idx, column=1).value
        if first_cell and str(first_cell) in [str(h) for h in painel_headers[:3]]:
            header_row = row_idx
            break

    print(f"Cabeçalho na linha: {header_row}")
    print()

    data_start = header_row + 1

    for col_name, col_idx in found_cols.items():
        print(f"\n📊 Coluna: '{col_name}' (coluna {col_idx})")
        print("-" * 40)

        # Verificar primeiras 20 células de dados
        formulas_found = 0
        values_found = 0
        empty_found = 0
        samples = []

        for row_idx in range(data_start, min(data_start + 20, ws.max_row + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)

            # Verificar se é fórmula
            if cell.data_type == 'f':  # 'f' = formula
                formulas_found += 1
                if len(samples) < 3:
                    samples.append(f"FÓRMULA Linha {row_idx}: {cell.value}")
            elif cell.value is None:
                empty_found += 1
            else:
                values_found += 1
                if len(samples) < 3:
                    samples.append(f"VALOR   Linha {row_idx}: {cell.value} (tipo: {type(cell.value).__name__})")

        print(f"  Células vazias: {empty_found}")
        print(f"  Fórmulas detectadas: {formulas_found}")
        print(f"  Valores detectados: {values_found}")

        if samples:
            print("  Amostras:")
            for s in samples:
                print(f"    {s}")

        if formulas_found > 0:
            print(f"  ✅ RESULTADO: Contém FÓRMULAS (calculadas automaticamente)")
        elif values_found > 0:
            print(f"  ⚠️  RESULTADO: Contém VALORES ESTÁTICOS (inseridos manualmente)")
        else:
            print(f"  ❓ RESULTADO: Apenas células vazias encontradas")

    wb.close()

print()

# ============================================================
# 5. VERIFICAR SE DADOS PARECEM ATUALIZADOS
# ============================================================
print("=" * 80)
print("5. VERIFICAÇÃO DE ATUALIZAÇÃO DOS DADOS")
print("=" * 80)

if 'PAINEL' in sheet_names:
    print("\nAnalisando valores calculados (data_only=True)...")
    wb_data = load_workbook(file_path, data_only=True)
    ws_data = wb_data['PAINEL']

    # Estatísticas das colunas de SALDO
    for col_name, col_idx in found_cols.items():
        numeric_values = []
        dates_in_sheet = []

        for row_idx in range(data_start, min(data_start + 500, ws_data.max_row + 1)):
            # Valor da coluna de saldo
            cell_val = ws_data.cell(row=row_idx, column=col_idx).value
            if isinstance(cell_val, (int, float)):
                numeric_values.append(cell_val)

            # Tentar encontrar datas nas primeiras colunas
            for dc in range(1, min(6, ws_data.max_column + 1)):
                dcell = ws_data.cell(row=row_idx, column=dc).value
                if isinstance(dcell, datetime):
                    dates_in_sheet.append(dcell)

        if numeric_values:
            print(f"\n📊 '{col_name}':")
            print(f"  Total de valores numéricos: {len(numeric_values)}")
            print(f"  Valor mínimo: {min(numeric_values):,.2f}")
            print(f"  Valor máximo: {max(numeric_values):,.2f}")
            print(f"  Média: {sum(numeric_values)/len(numeric_values):,.2f}")

            # Últimos 5 valores (do final dos dados)
            last_values = numeric_values[-5:] if len(numeric_values) >= 5 else numeric_values
            print(f"  Últimos valores: {[f'{v:,.2f}' for v in last_values]}")

            # Verificar zeros
            zeros = sum(1 for v in numeric_values if v == 0)
            if zeros == len(numeric_values) and len(numeric_values) > 0:
                print(f"  🚨 ALERTA: TODOS os valores são ZERO - dados NÃO atualizados!")
            elif zeros > len(numeric_values) * 0.3:
                print(f"  ⚠️  ATENÇÃO: {zeros}/{len(numeric_values)} valores são zero ({zeros/len(numeric_values)*100:.1f}%)")
            else:
                print(f"  ✅ Apenas {zeros}/{len(numeric_values)} valores são zero")
        else:
            print(f"\n📊 '{col_name}': Nenhum valor numérico encontrado")

    # Análise de datas
    if dates_in_sheet:
        print(f"\n📅 DATAS ENCONTRADAS:")
        print(f"  Total de datas: {len(dates_in_sheet)}")
        print(f"  Primeira data: {min(dates_in_sheet).strftime('%d/%m/%Y')}")
        print(f"  Última data: {max(dates_in_sheet).strftime('%d/%m/%Y')}")

        # Verificar se há dados de maio/2026
        maio_2026 = [d for d in dates_in_sheet if d.year == 2026 and d.month == 5]
        if maio_2026:
            print(f"  ✅ Encontradas {len(maio_2026)} datas de Maio/2026")
            print(f"     Período: {min(maio_2026).strftime('%d/%m/%Y')} a {max(maio_2026).strftime('%d/%m/%Y')}")

    wb_data.close()

# Informações do arquivo
print("\n" + "=" * 80)
print("INFORMAÇÕES DO ARQUIVO")
print("=" * 80)
stat = os.stat(file_path)
print(f"  Nome: {os.path.basename(file_path)}")
print(f"  Tamanho: {stat.st_size:,} bytes ({stat.st_size/1024:.1f} KB)")
print(f"  Modificado em: {datetime.fromtimestamp(stat.st_mtime).strftime('%d/%m/%Y %H:%M:%S')}")
print(f"  Criado em: {datetime.fromtimestamp(stat.st_ctime).strftime('%d/%m/%Y %H:%M:%S')}")

print("\n" + "=" * 80)
print("ANÁLISE CONCLUÍDA")
print("=" * 80)
