import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from collections import defaultdict
import time

SOURCE = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\BASE_PREST_DB_EXPORT_V2.xlsx"
OUTPUT = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\PRESTACAO_CONTAS.xlsx"

# Styles
title_font = Font(bold=True, size=16, color="1F4E79")
subtitle_font = Font(bold=True, size=11, color="4472C4")
header_font = Font(bold=True, color="FFFFFF", size=10)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
cpf_header_font = Font(bold=True, size=11, color="1F4E79")
cpf_header_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
report_header_font = Font(bold=True, size=10, color="2E75B6")
report_header_fill = PatternFill(start_color="E7F0F9", end_color="E7F0F9", fill_type="solid")
subtotal_font = Font(bold=True, size=10)
subtotal_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
grandtotal_font = Font(bold=True, size=12, color="FFFFFF")
grandtotal_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
medium_border = Border(left=Side(style='medium'), right=Side(style='medium'),
                       top=Side(style='medium'), bottom=Side(style='medium'))
center = Alignment(horizontal='center', vertical='center')
left = Alignment(horizontal='left', vertical='center')
right = Alignment(horizontal='right', vertical='center')
money_fmt = 'R$ #,##0.00;[Red]-R$ #,##0.00'
date_fmt = 'DD/MM/YYYY'

print("Loading source data...")
t0 = time.time()
wb_src = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)

# Load report summary
ws_rep = wb_src['Report Summary']
reports = []
for i, row in enumerate(ws_rep.iter_rows(values_only=True)):
    if i == 0: continue
    reports.append({
        'id': row[0], 'name': row[1], 'status': row[2], 'cpf': str(row[3] or '').strip(),
        'nome': row[4], 'approval_date': row[5], 'expense_count': row[6], 'total_value': float(row[7] or 0)
    })
print(f"  {len(reports)} reports loaded")

# Load all expenses
ws_exp = wb_src['BASE PREST (DB)']
expenses = []
for i, row in enumerate(ws_exp.iter_rows(values_only=True)):
    if i == 0: continue
    expenses.append({
        'report_id': row[0], 'report_name': row[1], 'status': row[2],
        'cpf': str(row[3] or '').strip(), 'nome': row[4],
        'approval_date': row[7], 'report_total': row[8],
        'expense_id': row[9], 'description': row[10], 'value': float(row[11] or 0),
        'date': row[12], 'expense_status': row[13]
    })
print(f"  {len(expenses)} expenses loaded in {time.time()-t0:.1f}s")
wb_src.close()

# Group by CPF -> reports -> expenses
cpf_data = defaultdict(lambda: {'nome': '', 'reports': defaultdict(lambda: {'name': '', 'status': '', 'approval_date': '', 'expenses': []})})
for e in expenses:
    cpf = e['cpf']
    if not cpf: continue
    cpf_data[cpf]['nome'] = e['nome']
    rid = e['report_id']
    r = cpf_data[cpf]['reports'][rid]
    r['name'] = e['report_name']
    r['status'] = e['status']
    r['approval_date'] = e['approval_date']
    r['expenses'].append(e)

# Sort CPFs by nome
sorted_cpfs = sorted(cpf_data.keys(), key=lambda c: cpf_data[c]['nome'])

print(f"Building Prestação de Contas ({len(sorted_cpfs)} CPFs)...")

wb = Workbook()
ws = wb.active
ws.title = "Prestação de Contas"
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

# Title
r = 1
ws.merge_cells(f'A{r}:H{r}')
c = ws.cell(row=r, column=1, value="PRESTAÇÃO DE CONTAS - VEXPENSES")
c.font = title_font
c.alignment = center
ws.row_dimensions[r].height = 28
r += 2

# Column headers
headers = ["CPF", "Nome", "Relatório", "Data Aprovação", "Despesa", "Data", "Valor", "Status"]
col_widths = [14, 32, 22, 20, 38, 12, 15, 12]
for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    c = ws.cell(row=r, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border
    ws.column_dimensions[get_column_letter(col)].width = w
header_row = r
r += 1

grand_total = 0.0
grand_count = 0

for cpf in sorted_cpfs:
    data = cpf_data[cpf]
    nome = data['nome']
    cpf_total = 0.0
    cpf_count = 0
    cpf_start_row = r

    # CPF header row
    ws.merge_cells(f'A{r}:H{r}')
    c = ws.cell(row=r, column=1, value=f"CPF: {cpf}   |   {nome}")
    c.font = cpf_header_font
    c.fill = cpf_header_fill
    c.alignment = left
    c.border = thin_border
    ws.row_dimensions[r].height = 22
    r += 1

    # Reports sorted by id
    sorted_reports = sorted(data['reports'].keys())
    for rid in sorted_reports:
        rep = data['reports'][rid]
        rep_total = 0.0
        rep_start_row = r

        # Report header
        ws.merge_cells(f'C{r}:H{r}')
        c = ws.cell(row=r, column=3, value=f"Relatório {rid}: {rep['name']} ({rep['status']})")
        c.font = report_header_font
        c.fill = report_header_fill
        c.alignment = left
        c.border = thin_border
        r += 1

        # Expenses
        for e in sorted(rep['expenses'], key=lambda x: x['expense_id'] or 0):
            vals = [
                cpf, nome,
                rep['name'],
                rep['approval_date'],
                e['description'],
                e['date'],
                e['value'],
                e['expense_status'],
            ]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=v)
                c.border = thin_border
                c.font = Font(size=9)
                if col == 6 and v is not None:
                    c.number_format = date_fmt
                if col == 7:
                    c.number_format = money_fmt
                if col == 8:
                    c.alignment = center
            rep_total += e['value']
            cpf_total += e['value']
            cpf_count += 1
            grand_count += 1
            r += 1

        # Report subtotal
        ws.merge_cells(f'A{r}:F{r}')
        c = ws.cell(row=r, column=1, value=f"  Subtotal Relatório {rid}:")
        c.font = subtotal_font
        c.fill = subtotal_fill
        c.alignment = right
        c.border = thin_border
        c = ws.cell(row=r, column=7, value=rep_total)
        c.font = subtotal_font
        c.fill = subtotal_fill
        c.number_format = money_fmt
        c.alignment = right
        c.border = thin_border
        c = ws.cell(row=r, column=8, value=f"{len(rep['expenses'])} desp")
        c.font = Font(size=8, italic=True)
        c.fill = subtotal_fill
        c.alignment = center
        c.border = thin_border
        r += 1

    # CPF subtotal
    ws.merge_cells(f'A{r}:F{r}')
    c = ws.cell(row=r, column=1, value=f"  TOTAL {nome}:")
    c.font = Font(bold=True, size=11, color="1F4E79")
    c.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    c.alignment = right
    c.border = medium_border
    c = ws.cell(row=r, column=7, value=cpf_total)
    c.font = Font(bold=True, size=11, color="1F4E79")
    c.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    c.number_format = money_fmt
    c.alignment = right
    c.border = medium_border
    c = ws.cell(row=r, column=8, value=f"{cpf_count} desp")
    c.font = Font(bold=True, size=9, color="1F4E79")
    c.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    c.alignment = center
    c.border = medium_border
    ws.row_dimensions[r].height = 24
    r += 1

    # Empty separator row
    r += 1
    grand_total += cpf_total

# Grand total
r += 1
ws.merge_cells(f'A{r}:F{r}')
c = ws.cell(row=r, column=1, value="  TOTAL GERAL PRESTAÇÃO DE CONTAS:")
c.font = grandtotal_font
c.fill = grandtotal_fill
c.alignment = right
c.border = medium_border
c = ws.cell(row=r, column=7, value=grand_total)
c.font = grandtotal_font
c.fill = grandtotal_fill
c.number_format = money_fmt
c.alignment = right
c.border = medium_border
c = ws.cell(row=r, column=8, value=f"{grand_count} desp")
c.font = grandtotal_font
c.fill = grandtotal_fill
c.alignment = center
c.border = medium_border
ws.row_dimensions[r].height = 30
r += 2

# Summary box
ws.merge_cells(f'A{r}:H{r}')
c = ws.cell(row=r, column=1, value=f"Resumo: {len(sorted_cpfs)} colaboradores | {len(reports)} relatórios | {grand_count} despesas | Total: R$ {grand_total:,.2f}")
c.font = Font(italic=True, size=10, color="4472C4")
c.alignment = center

# Freeze panes
ws.freeze_panes = f"A{header_row + 1}"

# Auto filter
ws.auto_filter.ref = f"A{header_row}:H{header_row}"

print(f"Saving to {OUTPUT}...")
wb.save(OUTPUT)
print(f"Done! {len(sorted_cpfs)} CPFs, {grand_count} expenses, total R$ {grand_total:,.2f}")
