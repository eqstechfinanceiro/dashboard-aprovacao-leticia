#!/usr/bin/env python3
"""
Script com output em arquivo
"""

import pandas as pd
import openpyxl
import os
import sys
import traceback

output_file = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\output_analise.txt"
file_path = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\controle-api\data\CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsx"

def log(msg):
    print(msg)
    with open(output_file, 'a', encoding='utf-8') as f:
        f.write(msg + '\n')

# Limpar arquivo anterior
with open(output_file, 'w', encoding='utf-8') as f:
    f.write('')

try:
    log("=" * 80)
    log("ANÁLISE DA PLANILHA")
    log("=" * 80)

    log(f"\nArquivo: {file_path}")
    log(f"Existe: {os.path.exists(file_path)}")

    if not os.path.exists(file_path):
        log("ERRO: Arquivo não existe!")
        sys.exit(1)

    # 1. Listar abas
    log("\n" + "-" * 40)
    log("1. ABAS DISPONÍVEIS")
    log("-" * 40)

    xl = pd.ExcelFile(file_path)
    for i, name in enumerate(xl.sheet_names, 1):
        log(f"  {i}. {name}")

    log(f"\nTotal: {len(xl.sheet_names)} abas")

    # 2. Para cada aba, detectar cabeçalho
    log("\n" + "-" * 40)
    log("2. CABEÇALHO POR ABA")
    log("-" * 40)

    for sheet in xl.sheet_names:
        try:
            df = pd.read_excel(file_path, sheet_name=sheet, nrows=5, header=None)
            if df.empty:
                log(f"\n  📄 {sheet}: (vazia)")
                continue

            log(f"\n  📄 {sheet}:")
            for i in range(min(5, len(df))):
                vals = [str(v)[:25] for v in df.iloc[i] if pd.notna(v)][:6]
                marker = " <-- provável cabeçalho" if i < 3 and len(vals) >= 2 else ""
                log(f"     Linha {i+1}: {vals}{marker}")
        except Exception as e2:
            log(f"\n  📄 {sheet}: ERRO - {e2}")

    # 3. Análise da aba PAINEL
    if 'PAINEL' in xl.sheet_names:
        log("\n" + "=" * 40)
        log("3. ABA PAINEL - COLUNAS")
        log("=" * 40)

        df_painel = pd.read_excel(file_path, sheet_name='PAINEL', nrows=30)
        log(f"\nTotal de colunas: {len(df_painel.columns)}")
        log("\nColunas:")
        for i, col in enumerate(df_painel.columns, 1):
            log(f"  {i:2d}. {col}")

        # 4. Verificar colunas de SALDO
        log("\n" + "=" * 40)
        log("4. ANÁLISE DE COLUNAS DE SALDO")
        log("=" * 40)

        saldo_cols = [c for c in df_painel.columns if 'SALDO' in str(c).upper()]
        log(f"\nColunas com 'SALDO' no nome: {saldo_cols}")

        # Carregar com openpyxl para ver fórmulas
        log("\nCarregando com openpyxl...")
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb['PAINEL']

        log(f"Dimensões: {ws.max_row} linhas x {ws.max_column} colunas")

        # Encontrar cabeçalho
        header_row = 1
        for r in range(1, min(15, ws.max_row + 1)):
            cell_val = ws.cell(row=r, column=1).value
            if cell_val and str(cell_val).strip() == str(df_painel.columns[0]).strip():
                header_row = r
                break

        log(f"Cabeçalho detectado na linha: {header_row}")

        # Mapear índices das colunas de SALDO
        col_mapping = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=header_row, column=col_idx).value
            if cell_val:
                for saldo_col in saldo_cols:
                    if str(cell_val).strip() == str(saldo_col).strip():
                        col_mapping[saldo_col] = col_idx

        log(f"\nMapeamento: {col_mapping}")

        data_start = header_row + 1

        for col_name, col_idx in col_mapping.items():
            log(f"\n📊 Analisando: '{col_name}' (coluna {col_idx})")

            formulas = 0
            values = 0
            empty = 0
            sample_data = []

            for row_idx in range(data_start, min(data_start + 20, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)

                if cell.data_type == 'f':
                    formulas += 1
                    if len(sample_data) < 2:
                        sample_data.append(f"FÓRMULA L{row_idx}: {cell.value}")
                elif cell.value is None:
                    empty += 1
                else:
                    values += 1
                    if len(sample_data) < 2:
                        sample_data.append(f"VALOR   L{row_idx}: {cell.value}")

            log(f"  Fórmulas: {formulas}")
            log(f"  Valores: {values}")
            log(f"  Vazios: {empty}")
            for s in sample_data:
                log(f"    {s}")

            if formulas > 0:
                log(f"  ✅ RESULTADO: Contém FÓRMULAS")
            elif values > 0:
                log(f"  ⚠️  RESULTADO: VALORES ESTÁTICOS (sem fórmulas detectadas)")
            else:
                log(f"  ❓ RESULTADO: Sem dados nas primeiras linhas")

        wb.close()

    log("\n" + "=" * 80)
    log("ANÁLISE CONCLUÍDA COM SUCESSO")
    log("=" * 80)

except Exception as e:
    log(f"\nERRO: {e}")
    log(traceback.format_exc())
    log("\nANÁLISE FALHOU")

log(f"\nOutput salvo em: {output_file}")
