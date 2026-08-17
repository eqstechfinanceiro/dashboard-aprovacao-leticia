import openpyxl
import json
import requests
import sys

wb = openpyxl.load_workbook('FECHAMENTO - BEATRIZ SILVA DE SIQUEIRA 1.xlsx', data_only=True)
ws = wb['FECHAMENTO']

print("=" * 80)
print("FECHAMENTO SHEET vs API COMPARISON")
print("=" * 80)

# Sheet structure: Row 7=headers, 8-22=monthly, 23=TOTAL, 25-30=Resumo, K-L=Status
sheet_months = []
for r in range(8, 23):
    ano = ws.cell(r, 2).value
    mes = ws.cell(r, 3).value
    if ano is None:
        continue
    sheet_months.append({
        'ano': int(ano), 'mes': str(mes).upper(),
        'carga': float(ws.cell(r, 4).value or 0),
        'transferencia': float(ws.cell(r, 5).value or 0),
        'taxa': float(ws.cell(r, 6).value or 0),
        'prestacao_contas': float(ws.cell(r, 7).value or 0),
        'saldo': float(ws.cell(r, 8).value or 0),
        'acumulado': float(ws.cell(r, 9).value or 0),
    })

sheet_totals = {
    'carga': float(ws.cell(23, 4).value or 0),
    'transferencia': float(ws.cell(23, 5).value or 0),
    'taxa': float(ws.cell(23, 6).value or 0),
    'prestacao_contas': float(ws.cell(23, 7).value or 0),
    'saldo': float(ws.cell(23, 8).value or 0),
}

sheet_resumo = {
    'saldoFinal': float(ws.cell(25, 8).value or 0),
    'saldoDisponivel': float(ws.cell(26, 8).value or 0),
    'prestacaoContas': float(ws.cell(27, 8).value or 0),
    'fechamentoPrestacao': float(ws.cell(28, 8).value or 0),
    'saldoCartao': float(ws.cell(29, 8).value or 0),
    'fechamentoFinal': float(ws.cell(30, 8).value or 0),
}

sheet_status = {
    'aberto': float(ws.cell(8, 12).value or 0),
    'aprovado': float(ws.cell(9, 12).value or 0),
    'totalgeral': float(ws.cell(10, 12).value or 0),
}

print("\n--- SHEET months ---")
for m in sheet_months:
    print(f"  {m['ano']} {m['mes']:12s} | Carga: {m['carga']:>10.2f} | Transf: {m['transferencia']:>8.2f} | Taxa: {m['taxa']:>7.2f} | Prest: {m['prestacao_contas']:>10.2f} | Saldo: {m['saldo']:>10.2f} | Acum: {m['acumulado']:>10.2f}")

print(f"\nSheet TOTALS: {sheet_totals}")
print(f"Sheet RESUMO: {sheet_resumo}")
print(f"Sheet STATUS: {sheet_status}")

# --- Fetch API ---
print("\n" + "=" * 80)
print("API DATA")
print("=" * 80)

TOKEN = "eyJhbGciOiJIUzI1NiJ9.eyJpZCI6MSwiZW1haWwiOiJpdGFsby5tZWRyYWRvQGVxc2VuZ2VuaGFyaWEuY29tLmJyIiwibmFtZSI6Ikl0YWxvIE1lZHJhZG8iLCJqb2JfdGl0bGUiOiJBZG1pbmlzdHJhZG9yIiwicm9sZSI6ImFkbWluIiwibW9kdWxlcyI6W10sIm11c3RfY2hhbmdlX3Bhc3N3b3JkIjpmYWxzZSwiaWF0IjoxNzg2NzI5NzM3LCJleHAiOjE3ODczMzQ1Mzd9.-pseYfUui7R0AeBNj2rXiXja8kT9owM1CqfJj1FuWxI"
resp = requests.get("http://localhost:3000/api/fechamento?userId=895999", timeout=60, cookies={"vexp_auth_token": TOKEN})
print(f"API status: {resp.status_code}")
if resp.status_code != 200:
    print(f"Error: {resp.text[:500]}")
    sys.exit(1)

api = resp.json()
print(f"\nColaborador: {api['colaborador']}")
print(f"Resumo: {json.dumps(api['resumo'], indent=2)}")
print(f"Status: {json.dumps(api['statusPanel'], indent=2)}")

print(f"\nAPI months ({len(api['fechamento'])}):")
for m in api['fechamento']:
    print(f"  {m['ano']} {m['mes']:12s} | Carga: {m['carga']:>10.2f} | Transf: {m['transferencia']:>8.2f} | Taxa: {m['taxa']:>7.2f} | Prest: {m['prestacao_contas']:>10.2f} | Saldo: {m['saldo']:>10.2f} | Acum: {m['acumulado']:>10.2f}")

api_totals = {
    'carga': sum(m['carga'] for m in api['fechamento']),
    'transferencia': sum(m['transferencia'] for m in api['fechamento']),
    'taxa': sum(m['taxa'] for m in api['fechamento']),
    'prestacao_contas': sum(m['prestacao_contas'] for m in api['fechamento']),
    'saldo': sum(m['saldo'] for m in api['fechamento']),
}

# --- COMPARISON ---
print("\n" + "=" * 80)
print("COMPARISON")
print("=" * 80)

all_ok = True

print("\n--- Monthly ---")
sheet_by_key = {(m['ano'], m['mes']): m for m in sheet_months}
api_by_key = {(m['ano'], m['mes']): m for m in api['fechamento']}
all_keys = sorted(set(list(sheet_by_key.keys()) + list(api_by_key.keys())), key=lambda x: (x[0], x[1]))

for key in all_keys:
    s = sheet_by_key.get(key)
    a = api_by_key.get(key)
    if s and a:
        diffs = []
        for field in ['carga', 'transferencia', 'taxa', 'prestacao_contas', 'saldo', 'acumulado']:
            if abs(s[field] - a[field]) > 0.01:
                diffs.append(f"{field}: sheet={s[field]:.2f} api={a[field]:.2f} diff={a[field]-s[field]:.2f}")
                all_ok = False
        status = "❌" if diffs else "✅"
        detail = " | ".join(diffs) if diffs else "MATCH"
        print(f"  {status} {key[0]} {key[1]}: {detail}")
    elif s:
        print(f"  ⚠️  {key[0]} {key[1]}: only in SHEET")
        all_ok = False
    elif a:
        print(f"  ⚠️  {key[0]} {key[1]}: only in API")
        all_ok = False

print("\n--- Totals ---")
for field in ['carga', 'transferencia', 'taxa', 'prestacao_contas', 'saldo']:
    sv = sheet_totals.get(field, 0)
    av = api_totals.get(field, 0)
    if abs(sv - av) > 0.01:
        print(f"  ❌ {field}: sheet={sv:.2f} api={av:.2f} diff={av-sv:.2f}")
        all_ok = False
    else:
        print(f"  ✅ {field}: sheet={sv:.2f} = api={av:.2f}")

print("\n--- Resumo ---")
for key in ['saldoFinal', 'saldoDisponivel', 'prestacaoContas', 'fechamentoPrestacao', 'saldoCartao', 'fechamentoFinal']:
    sv = sheet_resumo.get(key, 0)
    av = api['resumo'].get(key, 0)
    if abs(sv - av) > 0.01:
        print(f"  ❌ {key}: sheet={sv:.2f} api={av:.2f} diff={av-sv:.2f}")
        all_ok = False
    else:
        print(f"  ✅ {key}: sheet={sv:.2f} = api={av:.2f}")

print("\n--- Status Panel ---")
for key, sheet_key in [('aberto', 'aberto'), ('aprovado', 'aprovado'), ('totalGeral', 'totalgeral')]:
    sv = sheet_status.get(sheet_key, 0)
    av = api['statusPanel'].get(key, 0)
    if abs(sv - av) > 0.01:
        print(f"  ❌ {key}: sheet={sv:.2f} api={av:.2f} diff={av-sv:.2f}")
        all_ok = False
    else:
        print(f"  ✅ {key}: sheet={sv:.2f} = api={av:.2f}")

print("\n" + "=" * 80)
if all_ok:
    print("✅ ALL DATA MATCHES!")
else:
    print("❌ DISCREPANCIES FOUND")
print("=" * 80)
