#!/usr/bin/env python3
"""
Script simples para analisar a planilha Excel
"""

import pandas as pd
import openpyxl
import os
import traceback

file_path = r"C:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\controle-api\data\CONTROLE - VEXPENSES - MAIO - 2026 (1).xlsx"

print("=" * 80)
print("ANÁLISE DA PLANILHA")
print("=" * 80)

try:
    print(f"\nArquivo: {file_path}")
    print(f"Existe: {os.path.exists(file_path)}")

    if not os.path.exists(file_path):
        print("ERRO: Arquivo não existe!")
        exit(1)

    # 1. Listar abas
    print("\n" + "-" * 40)
    print("1. ABAS DISPONÍVEIS")
    print("-" * 40)

    xl = pd.ExcelFile(file_path)
    for i, name in enumerate(xl.sheet_names, 1):
        print(f"  {i}. {name}")

    # 2. Para cada aba, detectar cabeçalho
    print("\n" + "-" * 40)
    print("2. CABEÇALHO POR ABA")
    print("-" * 40)

    for sheet in xl.sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet, nrows=5, header=None)
        if not df.empty:
            print(f"\n  📄 {sheet}:")
            # Mostrar primeiras linhas para identificar cabeçalho
            for i in range(min(5, len(df))):
                vals = [str(v) for v in df.iloc[i] if pd.notna(v)][:6]
                print(f"     Linha {i+1}: {vals}")

    # 3. Análise da aba PAINEL
    if 'PAINEL' in xl.sheet_names:
        print("\n" + "-" * 40)
        print("3. ABA PAINEL - COLUNAS")
        print("-" * 40)

        df_painel = pd.read_excel(file_path, sheet_name='PAINEL', nrows=20)
        print(f"\nTotal de colunas: {len(df_painel.columns)}")
        print("\nColunas:")
        for i, col in enumerate(df_painel.columns, 1):
            print(f"  {i:2d}. {col}")

        # 4. Verificar colunas de SALDO
        print("\n" + "-" * 40)
        print("4. ANÁLISE DE COLUNAS DE SALDO")
        print("-" * 40)

        saldo_cols = [c for c in df_painel.columns if 'SALDO' in str(c).upper()]
        print(f"\nColunas com 'SALDO' no nome: {saldo_cols}")

        # Carregar com openpyxl para ver fórmulas
        print("\nCarregando com openpyxl para verificar fórmulas...")
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb['PAINEL']

        print(f"Dimensões: {ws.max_row} linhas x {ws.max_column} colunas")

        # Encontrar cabeçalho
        header_row = 1
        for r in range(1, min(10, ws.max_row + 1)):
            cell_val = ws.cell(row=r, column=1).value
            if cell_val and str(cell_val) == str(df_painel.columns[0]):
                header_row = r
                break

        print(f"Cabeçalho na linha: {header_row}")

        # Mapear índices das colunas de SALDO
        col_mapping = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=header_row, column=col_idx).value
            if cell_val:
                for saldo_col in saldo_cols:
                    if str(cell_val) == str(saldo_col):
                        col_mapping[saldo_col] = col_idx

        print(f"\nMapeamento de colunas: {col_mapping}")

        data_start = header_row + 1

        for col_name, col_idx in col_mapping.items():
            print(f"\n📊 Analisando: '{col_name}' (coluna {col_idx})")

            formulas = 0
            values = 0
            empty = 0
            sample_data = []

            for row_idx in range(data_start, min(data_start + 15, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)

                if cell.data_type == 'f':
                    formulas += 1
                    if len(sample_data) < 2:
                        sample_data.append(f"FÓRMULA L{row_idx}: {cell.value}")
                elif cell.value is None:
                    empty += 1
                else:
                    values += 1
                    if len(sample_data) < 2:
                        sample_data.append(f"VALOR   L{row_idx}: {cell.value}")

            print(f"  Fórmulas: {formulas}")
            print(f"  Valores: {values}")
            print(f"  Vazios: {empty}")
            for s in sample_data:
                print(f"    {s}")

            if formulas > 0:
                print(f"  ✅ FÓRMULAS detectadas")
            else:
                print(f"  ⚠️  VALORES ESTÁTICOS (sem fórmulas)")

        wb.close()

    print("\n" + "=" * 80)
    print("ANÁLISE CONCLUÍDA")
    print("=" * 80)

except Exception as e:
    print(f"\nERRO: {e}")
    print(traceback.format_exc())
