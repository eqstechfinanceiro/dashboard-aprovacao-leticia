import openpyxl
import json
import requests
import sys
import os
import unicodedata

TOKEN = "eyJhbGciOiJIUzI1NiJ9.eyJpZCI6MSwiZW1haWwiOiJpdGFsby5tZWRyYWRvQGVxc2VuZ2VuaGFyaWEuY29tLmJyIiwibmFtZSI6Ikl0YWxvIE1lZHJhZG8iLCJqb2JfdGl0bGUiOiJBZG1pbmlzdHJhZG9yIiwicm9sZSI6ImFkbWluIiwibW9kdWxlcyI6W10sIm11c3RfY2hhbmdlX3Bhc3N3b3JkIjpmYWxzZSwiaWF0IjoxNzg2NzI5NzM3LCJleHAiOjE3ODczMzQ1Mzd9.-pseYfUui7R0AeBNj2rXiXja8kT9owM1CqfJj1FuWxI"
API = "http://localhost:3000"
FECHAMENTOS_DIR = "fechamentos"

USER_IDS = {
    'ADAN LEONARDO SOUZA BATISTA': 923558,
    'ANDRE VALERIO DE PAIVA': 895985,
    'CARLOS NASCIMENTO NONATO JUNIOR': 896018,
    'DHIEGO RIBEIRO DINIZ': 896053,
}

MESES_PT = [
    'JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO',
    'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO'
]

MESE_NUM = {
    '01': 'JANEIRO', '02': 'FEVEREIRO', '03': 'MARÇO', '04': 'ABRIL',
    '05': 'MAIO', '06': 'JUNHO', '07': 'JULHO', '08': 'AGOSTO',
    '09': 'SETEMBRO', '10': 'OUTUBRO', '11': 'NOVEMBRO', '12': 'DEZEMBRO',
}


def parse_sheet(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['FECHAMENTO']
    
    header_row = None
    for r in range(1, min(20, ws.max_row + 1)):
        vals = [str(ws.cell(r, c).value or '').upper() for c in range(1, min(13, ws.max_column + 1))]
        if 'ANO' in vals and 'MÊS' in vals:
            header_row = r
            break
        if 'MÊS' in vals and 'CARGAS' in vals:
            header_row = r
            break
    
    if header_row is None:
        return None
    
    headers = [str(ws.cell(header_row, c).value or '').upper().strip() for c in range(1, ws.max_column + 1)]
    is_format_b = 'CARGAS' in headers
    
    colaborador = str(ws.cell(6, 2).value or '').strip()
    months = []
    
    if is_format_b:
        col_map = {h: c + 1 for c, h in enumerate(headers)}
        col_mes = col_map.get('MÊS', 2)
        col_carga = col_map.get('CARGAS', 3)
        col_transf = col_map.get('TRANSF.', 4)
        col_taxa = col_map.get('TAXA SAQ.', 5)
        col_prest = col_map.get('PREST. CONTAS', 6)
        col_saldo = col_map.get('SALDO ', 7) or col_map.get('SALDO', 7)
        col_acum = col_map.get('ACUMULADO', 8)
        
        for r in range(header_row + 1, ws.max_row + 1):
            mes_val = ws.cell(r, col_mes).value
            if mes_val is None:
                continue
            mes_str = str(mes_val).strip()
            if '/' not in mes_str:
                first = str(ws.cell(r, 1).value or '').upper()
                if 'TOTAL' in first or mes_str == '':
                    break
                continue
            parts = mes_str.split('/')
            mes_num = parts[0].strip()
            ano = int(parts[1].strip())
            mes_nome = MESE_NUM.get(mes_num, mes_str)
            months.append({
                'ano': ano, 'mes': mes_nome,
                'carga': float(ws.cell(r, col_carga).value or 0),
                'transferencia': float(ws.cell(r, col_transf).value or 0),
                'taxa': float(ws.cell(r, col_taxa).value or 0),
                'prestacao_contas': float(ws.cell(r, col_prest).value or 0),
                'saldo': float(ws.cell(r, col_saldo).value or 0),
                'acumulado': float(ws.cell(r, col_acum).value or 0),
            })
    else:
        col_map = {h: c + 1 for c, h in enumerate(headers)}
        col_ano = col_map.get('ANO', 2)
        col_mes = col_map.get('MÊS', 3)
        col_carga = col_map.get('CARGA', 4)
        col_transf = col_map.get('TRANSFERÊNCIA', 5)
        col_taxa = col_map.get('TAXA', 6)
        col_prest = col_map.get('PREST. CONTAS', 7)
        col_saldo = col_map.get('SALDO', 8)
        col_acum = col_map.get('ACUMULADO', 9)
        
        for r in range(header_row + 1, ws.max_row + 1):
            ano = ws.cell(r, col_ano).value
            mes = ws.cell(r, col_mes).value
            if ano is None or not isinstance(ano, (int, float)):
                first = str(ws.cell(r, col_ano).value or '').upper()
                if 'TOTAL' in first:
                    break
                continue
            months.append({
                'ano': int(ano), 'mes': str(mes).upper(),
                'carga': float(ws.cell(r, col_carga).value or 0),
                'transferencia': float(ws.cell(r, col_transf).value or 0),
                'taxa': float(ws.cell(r, col_taxa).value or 0),
                'prestacao_contas': float(ws.cell(r, col_prest).value or 0),
                'saldo': float(ws.cell(r, col_saldo).value or 0),
                'acumulado': float(ws.cell(r, col_acum).value or 0),
            })
    
    resumo = {}
    status = {}
    val_col = 8 if not is_format_b else 7
    
    for r in range(1, ws.max_row + 1):
        label_b = str(ws.cell(r, 2).value or '').upper().strip()
        if 'SALDO FINAL' in label_b:
            resumo['saldoFinal'] = float(ws.cell(r, val_col).value or 0)
        elif 'SALDO DISPONIVEL' in label_b or 'SALDO DISPONÍVEL' in label_b:
            resumo['saldoDisponivel'] = float(ws.cell(r, val_col).value or 0)
        elif 'PRESTAÇÃO DE CONTAS' in label_b or 'PRESTACAO DE CONTAS' in label_b:
            resumo['prestacaoContas'] = float(ws.cell(r, val_col).value or 0)
        elif 'FECHAMENTO' in label_b and 'CART' not in label_b and 'FINAL' not in label_b:
            resumo['fechamentoPrestacao'] = float(ws.cell(r, val_col).value or 0)
        elif 'SALDO CART' in label_b:
            resumo['saldoCartao'] = float(ws.cell(r, val_col).value or 0)
        elif 'FECHAMENTO FINAL' in label_b:
            resumo['fechamentoFinal'] = float(ws.cell(r, val_col).value or 0)
        if label_b in ['ABERTO', 'APROVADO', 'TOTAL GERAL']:
            for c in [12, 11, 8]:
                v = ws.cell(r, c).value
                if v is not None and isinstance(v, (int, float)):
                    status[label_b.lower().replace(' ', '')] = float(v)
                    break
    
    return {'months': months, 'resumo': resumo, 'status': status, 'colaborador': colaborador, 'format': 'B' if is_format_b else 'A'}


def compare_sheet_vs_api(sheet_data, api_data, filename):
    print(f"\n{'=' * 80}")
    print(f"FILE: {filename}")
    print(f"Colaborador: {sheet_data['colaborador']} (format {sheet_data['format']})")
    print(f"{'=' * 80}")
    
    all_ok = True
    
    print("\n--- Monthly ---")
    sheet_by_key = {(m['ano'], m['mes']): m for m in sheet_data['months']}
    api_by_key = {(m['ano'], m['mes']): m for m in api_data['fechamento']}
    all_keys = sorted(set(list(sheet_by_key.keys()) + list(api_by_key.keys())), key=lambda x: (x[0], MESES_PT.index(x[1]) if x[1] in MESES_PT else 0))
    
    for key in all_keys:
        s = sheet_by_key.get(key)
        a = api_by_key.get(key)
        if s and a:
            diffs = []
            for field in ['carga', 'transferencia', 'taxa', 'prestacao_contas', 'saldo', 'acumulado']:
                if abs(s[field] - a[field]) > 0.01:
                    diffs.append(f"{field}: sheet={s[field]:.2f} api={a[field]:.2f} diff={a[field]-s[field]:.2f}")
                    all_ok = False
            status = "❌" if diffs else "✅"
            detail = " | ".join(diffs) if diffs else "MATCH"
            print(f"  {status} {key[0]} {key[1]}: {detail}")
        elif s:
            print(f"  ⚠️  {key[0]} {key[1]}: only in SHEET")
            all_ok = False
        elif a:
            print(f"  ⚠️  {key[0]} {key[1]}: only in API")
            all_ok = False
    
    print("\n--- Resumo ---")
    for key in ['saldoFinal', 'saldoDisponivel', 'prestacaoContas', 'fechamentoPrestacao', 'saldoCartao', 'fechamentoFinal']:
        sv = sheet_data['resumo'].get(key, None)
        av = api_data['resumo'].get(key, 0)
        if sv is None:
            print(f"  ⚠️  {key}: not found in sheet")
            continue
        if abs(sv - av) > 0.01:
            print(f"  ❌ {key}: sheet={sv:.2f} api={av:.2f} diff={av-sv:.2f}")
            all_ok = False
        else:
            print(f"  ✅ {key}: sheet={sv:.2f} = api={av:.2f}")
    
    print("\n--- Status Panel ---")
    for key, sheet_key in [('aberto', 'aberto'), ('aprovado', 'aprovado'), ('totalGeral', 'totalgeral')]:
        sv = sheet_data['status'].get(sheet_key, None)
        av = api_data['statusPanel'].get(key, 0)
        if sv is None:
            print(f"  ⚠️  {key}: not found in sheet")
            continue
        if abs(sv - av) > 0.01:
            print(f"  ❌ {key}: sheet={sv:.2f} api={av:.2f} diff={av-sv:.2f}")
            all_ok = False
        else:
            print(f"  ✅ {key}: sheet={sv:.2f} = api={av:.2f}")
    
    return all_ok


files = sorted(os.listdir(FECHAMENTOS_DIR))
overall_ok = True

for fname in files:
    if not fname.endswith('.xlsx'):
        continue
    filepath = os.path.join(FECHAMENTOS_DIR, fname)
    sheet_data = parse_sheet(filepath)
    if sheet_data is None:
        print(f"\n❌ Could not parse: {fname}")
        overall_ok = False
        continue
    
    colab_clean = unicodedata.normalize('NFKD', sheet_data['colaborador']).encode('ASCII', 'ignore').decode().upper().strip()
    user_id = None
    for name, uid in USER_IDS.items():
        name_clean = unicodedata.normalize('NFKD', name).encode('ASCII', 'ignore').decode().upper().strip()
        if name_clean == colab_clean or colab_clean in name_clean or name_clean in colab_clean:
            user_id = uid
            break
    
    if not user_id:
        print(f"\n❌ Could not find user ID for '{sheet_data['colaborador']}'")
        overall_ok = False
        continue
    
    resp = requests.get(f"{API}/api/fechamento?userId={user_id}", timeout=60, cookies={"vexp_auth_token": TOKEN})
    if resp.status_code != 200:
        print(f"\n❌ API error for {sheet_data['colaborador']}: {resp.status_code}")
        overall_ok = False
        continue
    
    api_data = resp.json()
    ok = compare_sheet_vs_api(sheet_data, api_data, fname)
    if not ok:
        overall_ok = False

print(f"\n{'=' * 80}")
if overall_ok:
    print("✅ ALL 4 REPORTS MATCH!")
else:
    print("❌ DISCREPANCIES FOUND")
print("=" * 80)
