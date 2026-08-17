import openpyxl
import json
import subprocess
from collections import defaultdict

# Load BASE PREST per CPF and per report
wb = openpyxl.load_workbook('CONTROLE - VEXPENSES - AGOSTO 2026.xlsx', read_only=True, data_only=True)
ws = wb['BASE PREST ']

bp_per_cpf = defaultdict(float)
bp_fatura_per_cpf = defaultdict(float)
bp_reports = defaultdict(lambda: defaultdict(lambda: {'total': 0, 'count': 0, 'status': '', 'name': ''}))
for row in ws.iter_rows(min_row=4, values_only=True):
    cpf = str(row[9]).strip() if row[9] else ''
    valor = float(row[26]) if row[26] is not None else 0
    report_id = str(row[1]).strip() if row[1] else ''
    report_name = str(row[2]).strip().upper() if row[2] else ''
    status = str(row[10]).strip() if row[10] else ''
    if not cpf or cpf == 'None':
        continue
    bp_per_cpf[cpf] += valor
    if 'FATURA' in report_name or 'CARTAO' in report_name or 'CARTAO' in report_name:
        bp_fatura_per_cpf[cpf] += valor
    if report_id and report_id != 'None':
        bp_reports[cpf][report_id]['total'] += valor
        bp_reports[cpf][report_id]['count'] += 1
        bp_reports[cpf][report_id]['status'] = status
        bp_reports[cpf][report_id]['name'] = report_name

# Load API debug
with open('prestacao_debug.json', 'r', encoding='utf-8-sig') as f:
    debug = json.load(f)

api_per_cpf = {}
for item in debug.get('cpf_summary', []):
    cpf = item['user_cpf'].strip()
    api_per_cpf[cpf] = float(item['prestacao_total'])

# Load PAINEL names
ws_painel = wb['PAINEL']
painel_names = {}
for row in ws_painel.iter_rows(min_row=12, values_only=True):
    cpf = str(row[2]).strip() if row[2] else ''
    nome = str(row[1]).strip() if row[1] else ''
    if cpf and cpf != 'None':
        painel_names[cpf] = nome

# Find all CPFs with diff
diffs = []
for cpf in set(bp_per_cpf.keys()) | set(api_per_cpf.keys()):
    bp_total = bp_per_cpf.get(cpf, 0) - bp_fatura_per_cpf.get(cpf, 0)
    api_total = api_per_cpf.get(cpf, 0)
    diff = bp_total - api_total
    if abs(diff) > 1:
        diffs.append((cpf, painel_names.get(cpf, ''), bp_total, api_total, diff))

# For top 20 diffs, fetch per-CPF data and categorize
cat1_fatura = 0.0  # FATURA reports in API but not in BASE PREST (negative diff)
cat2_reprovado = 0.0  # REPROVADO/ABERTO reports in API but not in BASE PREST (negative diff)
cat3_partial = 0.0  # ENVIADO/APROVADO reports with partial expenses (positive diff)
cat_other = 0.0

for cpf, nome, bp, api, diff in sorted(diffs, key=lambda x: abs(x[4]), reverse=True)[:30]:
    ps_script = f'''
$session = New-Object Microsoft.PowerShell.Commands.WebRequestSession
$body = @{{ email = "italo.medrado@eqsengenharia.com.br"; password = "EQSeng4292@" }} | ConvertTo-Json
Invoke-RestMethod -Uri "http://localhost:3000/api/auth/login" -Method POST -Body $body -ContentType "application/json" -WebSession $session | Out-Null
$r = Invoke-RestMethod -Uri "http://localhost:3000/api/debug-prestacao-cpf?cpf={cpf}" -Method GET -WebSession $session
$r | ConvertTo-Json -Depth 5
'''
    result = subprocess.run(['powershell.exe', '-NoProfile', '-Command', ps_script], capture_output=True, text=True, timeout=30)
    try:
        api_data = json.loads(result.stdout)
        api_reps = {r['report_id']: r for r in api_data.get('api_expenses', [])}
    except:
        print(f"  ERROR for {nome}: {result.stderr[:100]}")
        cat_other += diff
        continue

    bp_reps = bp_reports.get(cpf, {})
    cpf_fatura = 0.0
    cpf_reprovado = 0.0
    cpf_partial = 0.0
    cpf_other = 0.0

    all_rids = set(bp_reps.keys()) | set(str(r) for r in api_reps.keys())
    for rid in all_rids:
        bp_r = bp_reps.get(rid, {'total': 0, 'count': 0, 'status': '', 'name': ''})
        api_r = api_reps.get(int(rid) if str(rid).isdigit() else -1, {'report_name': '', 'status': '', 'expense_count': 0, 'total_value': 0, 'excluded_value': 0})
        rdiff = float(bp_r['total']) - float(api_r.get('total_value', 0))

        if abs(rdiff) < 0.01:
            continue

        api_name = str(api_r.get('report_name', '')).upper()
        api_status = str(api_r.get('status', '')).upper()
        bp_name = bp_r['name'].upper()

        # Category 1: FATURA/CARTAO in API but not in BASE PREST
        if ('FATURA' in api_name or 'CARTAO' in api_name or 'CARTAO' in api_name) and bp_r['total'] == 0:
            cpf_fatura += rdiff  # negative
        # Category 2: REPROVADO/ABERTO in API but not in BASE PREST
        elif api_status in ('REPROVADO', 'ABERTO') and bp_r['total'] == 0:
            cpf_reprovado += rdiff  # negative
        # Category 3: ENVIADO/APROVADO with partial expenses (BP has more than API)
        elif rdiff > 0 and ('CAIXA' in bp_name or 'CAIXA' in api_name):
            cpf_partial += rdiff  # positive
        else:
            cpf_other += rdiff

    cat1_fatura += cpf_fatura
    cat2_reprovado += cpf_reprovado
    cat3_partial += cpf_partial
    cat_other += cpf_other

    print(f"{nome[:35]:35s} diff={diff:+.2f} | fatura={cpf_fatura:+.2f} reprovado={cpf_reprovado:+.2f} partial={cpf_partial:+.2f} other={cpf_other:+.2f}")

print(f"\n{'='*80}")
print(f"Category 1 - FATURA/CARTAO (API excludes, BASE PREST doesn't have): {cat1_fatura:+.2f}")
print(f"Category 2 - REPROVADO/ABERTO (API excludes, BASE PREST doesn't have): {cat2_reprovado:+.2f}")
print(f"Category 3 - Partial ENVIADO/APROVADO expenses (API has less than BP): {cat3_partial:+.2f}")
print(f"Other/uncategorized: {cat_other:+.2f}")
print(f"Sum of categories: {cat1_fatura+cat2_reprovado+cat3_partial+cat_other:+.2f}")
print(f"Total diff (all 193 users): {sum(d[4] for d in diffs):+.2f}")
print(f"Analyzed (top 30): {sum(d[4] for d in sorted(diffs, key=lambda x: abs(x[4]), reverse=True)[:30]):+.2f}")
