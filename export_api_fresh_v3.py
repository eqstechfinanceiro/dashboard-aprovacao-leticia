import requests
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

API_URL = "https://api.vexpenses.com"
API_KEY = "N2ntX8mUF7AvjVcyT6DZwnOQXRvYgel9pCirCGKKETO5R70HEZ8UdbQA0Nt8"
OUTPUT = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\BASE_PREST_API_FRESH_V2.xlsx"
ITAU_PMID = "627401"
MAX_WORKERS = 4

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
money_fmt = 'R$ #,##0.00;[Red]-R$ #,##0.00'
date_fmt = 'DD/MM/YYYY'
api_headers = {"Authorization": API_KEY, "Accept": "application/json"}

# Step 1: Fetch all reports and filter
print("[1/3] Fetching reports...", flush=True)
t0 = time.time()
resp = requests.get(f"{API_URL}/v2/reports?include=user", headers=api_headers, timeout=120)
resp.raise_for_status()
all_reports = resp.json().get("data", [])
print(f"  Got {len(all_reports)} reports in {time.time()-t0:.1f}s", flush=True)

def get_user_info(report):
    user = report.get("user", {})
    user_data = user.get("data", {}) if isinstance(user, dict) else {}
    return user_data.get("cpf", ""), user_data.get("name", "")

def is_fatura(name):
    n = (name or "").upper()
    return "FATURA" in n or "CARTAO" in n or "CARTÃO" in n

def is_aprovado(status):
    s = (status or "").upper()
    return "APROVADO" in s or "ENVIADO" in s

filtered = []
for r in all_reports:
    cpf, nome = get_user_info(r)
    if not cpf:
        continue
    if not is_aprovado(r.get("status")):
        continue
    if is_fatura(r.get("description")):
        continue
    r["_cpf"] = str(cpf).strip()
    r["_nome"] = nome
    r["_name"] = r.get("description") or ""
    filtered.append(r)

print(f"  Filtered: {len(filtered)} reports", flush=True)

# Step 2: Fetch expenses per report with 8 workers
print(f"[2/3] Fetching expenses ({MAX_WORKERS} workers)...", flush=True)
expenses_all = []
errors = []
lock = threading.Lock()
done_count = 0
t1 = time.time()
last_log = time.time()

def fetch_expenses(report):
    rid = report["id"]
    for attempt in range(3):
        try:
            r = requests.get(
                f"{API_URL}/v2/reports/{rid}?include=expenses",
                headers=api_headers,
                timeout=30
            )
            if r.status_code == 429:
                with lock:
                    errors.append(f"Report {rid}: 429 rate limited")
                time.sleep(5 * (attempt + 1))
                continue
            if r.status_code == 404:
                return []
            r.raise_for_status()
            data = r.json()
            report_expenses = data.get("data", {}).get("expenses", {}).get("data", [])
            result = []
            for e in report_expenses:
                pmid = str(e.get("payment_method_id", ""))
                if pmid == ITAU_PMID:
                    continue
                result.append({
                    "report_id": rid,
                    "report_name": report.get("_name", ""),
                    "status": report.get("status", ""),
                    "cpf": report.get("_cpf", ""),
                    "nome": report.get("_nome", ""),
                    "approval_date": report.get("approval_date", ""),
                    "report_total": None,
                    "expense_id": e.get("id"),
                    "description": e.get("title") or "",
                    "value": float(e.get("value") or 0),
                    "date": e.get("date"),
                    "expense_status": e.get("validate", ""),
                    "payment_method_id": pmid,
                    "payment_method_name": "",
                })
            return result
        except Exception as e:
            if attempt < 2:
                time.sleep(2 * (attempt + 1))
                continue
            with lock:
                errors.append(f"Report {rid}: {str(e)[:60]}")
            return []
    return []

with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
    futures = {executor.submit(fetch_expenses, r): r for r in filtered}
    for future in as_completed(futures):
        result = future.result()
        with lock:
            expenses_all.extend(result)
            done_count += 1
            if done_count % 50 == 0 or done_count == len(filtered) or (time.time() - last_log) > 15:
                elapsed = time.time() - t1
                pct = done_count / len(filtered) * 100
                rate = done_count / elapsed * 60
                eta = elapsed / done_count * (len(filtered) - done_count)
                print(f"  [{done_count}/{len(filtered)}] {pct:.0f}% | {len(expenses_all)} expenses | {rate:.0f} rep/min | {elapsed:.0f}s, ETA {eta:.0f}s | err={len(errors)}", flush=True)
                last_log = time.time()

print(f"  Done: {len(expenses_all)} expenses in {time.time()-t1:.1f}s | errors={len(errors)}", flush=True)

# Step 3: Build Excel
print("[3/3] Building Excel...", flush=True)
expenses_all.sort(key=lambda e: (e["cpf"], e["report_id"], e["expense_id"] or 0))

cpf_summary = defaultdict(lambda: {"nome": "", "reports": set(), "count": 0, "total": 0.0})
report_summary = defaultdict(lambda: {"name": "", "status": "", "cpf": "", "nome": "", "approval_date": "", "count": 0, "total": 0.0})
for e in expenses_all:
    cpf_summary[e["cpf"]]["nome"] = e["nome"]
    cpf_summary[e["cpf"]]["reports"].add(e["report_id"])
    cpf_summary[e["cpf"]]["count"] += 1
    cpf_summary[e["cpf"]]["total"] += e["value"]
    rs = report_summary[e["report_id"]]
    rs["name"] = e["report_name"]; rs["status"] = e["status"]; rs["cpf"] = e["cpf"]
    rs["nome"] = e["nome"]; rs["approval_date"] = e["approval_date"]
    rs["count"] += 1; rs["total"] += e["value"]

wb = Workbook()
ws1 = wb.active; ws1.title = "CPF Summary"
h1 = ["CPF", "Nome", "Qtd Relatórios", "Qtd Despesas", "Total Prestação"]
for col, h in enumerate(h1, 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal='center'); c.border = thin_border
for i, cpf in enumerate(sorted(cpf_summary.keys(), key=lambda c: cpf_summary[c]["nome"]), 2):
    d = cpf_summary[cpf]
    for col, v in enumerate([cpf, d["nome"], len(d["reports"]), d["count"], d["total"]], 1):
        c = ws1.cell(row=i, column=col, value=v); c.border = thin_border
        if col == 5: c.number_format = money_fmt
for col, w in enumerate([15, 35, 15, 15, 18], 1):
    ws1.column_dimensions[get_column_letter(col)].width = w
ws1.auto_filter.ref = f"A1:E{len(cpf_summary)+1}"; ws1.freeze_panes = "A2"

ws2 = wb.create_sheet("Report Summary")
h2 = ["Report ID", "Report Name", "Status", "CPF", "Nome", "Approval Date", "Qtd Despesas", "Total Valor"]
for col, h in enumerate(h2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal='center'); c.border = thin_border
for i, rid in enumerate(sorted(report_summary.keys(), key=lambda r: (report_summary[r]["cpf"], r)), 2):
    d = report_summary[rid]
    for col, v in enumerate([rid, d["name"], d["status"], d["cpf"], d["nome"], d["approval_date"], d["count"], d["total"]], 1):
        c = ws2.cell(row=i, column=col, value=v); c.border = thin_border
        if col == 8: c.number_format = money_fmt
for col, w in enumerate([12, 25, 12, 15, 30, 22, 12, 15], 1):
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.auto_filter.ref = f"A1:H{len(report_summary)+1}"; ws2.freeze_panes = "A2"

ws3 = wb.create_sheet("BASE PREST (API)")
h3 = ["Report ID", "Report Name", "Status", "CPF", "Nome", "Approval Date", "Report Total",
      "Expense ID", "Expense Description", "Expense Value", "Expense Date", "Expense Status",
      "Payment Method ID", "Payment Method Name"]
for col, h in enumerate(h3, 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal='center'); c.border = thin_border
for i, e in enumerate(expenses_all, 2):
    vals = [e["report_id"], e["report_name"], e["status"], e["cpf"], e["nome"], e["approval_date"], e["report_total"],
            e["expense_id"], e["description"], e["value"], e["date"], e["expense_status"], e["payment_method_id"], e["payment_method_name"]]
    for col, v in enumerate(vals, 1):
        c = ws3.cell(row=i, column=col, value=v); c.border = thin_border
        if col == 10: c.number_format = money_fmt
        if col == 11 and v: c.number_format = date_fmt
for col, w in enumerate([12, 25, 12, 15, 30, 22, 15, 12, 30, 15, 14, 15, 15, 20], 1):
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.auto_filter.ref = f"A1:N{len(expenses_all)+1}"; ws3.freeze_panes = "A2"

wb.save(OUTPUT)
total_val = sum(e["value"] for e in expenses_all)
print(f"DONE! {len(expenses_all)} expenses, {len(report_summary)} reports, {len(cpf_summary)} CPFs, Total: R$ {total_val:,.2f}", flush=True)
