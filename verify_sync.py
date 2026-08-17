import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import time

OLD_FILE = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\BASE_PREST_DB_EXPORT.xlsx"
NEW_FILE = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\BASE_PREST_DB_EXPORT_V2.xlsx"
CONTROLE_FILE = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\CONTROLE - VEXPENSES - AGOSTO 2026.xlsx"
OUTPUT = r"c:\Users\italo.medrado\Desktop\Projects\Análise de dados\Leticia\dashboard-test\SYNC_VERIFICATION.xlsx"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

def load_expenses(filepath, sheet_name="BASE PREST (DB)"):
    print(f"Loading {filepath} [{sheet_name}]...")
    t0 = time.time()
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]
    expenses = {}  # expense_id -> row dict
    report_totals = defaultdict(lambda: {"count": 0, "value": 0.0, "name": "", "cpf": "", "nome": "", "status": ""})
    cpf_totals = defaultdict(lambda: {"count": 0, "value": 0.0, "reports": set()})
    
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
            continue
        row_dict = dict(zip(headers, row))
        eid = row_dict.get("Expense ID")
        rid = row_dict.get("Report ID")
        cpf = str(row_dict.get("CPF") or "").strip()
        val = float(row_dict.get("Expense Value") or 0)
        
        if eid is not None:
            expenses[eid] = row_dict
        
        if rid is not None:
            report_totals[rid]["count"] += 1
            report_totals[rid]["value"] += val
            report_totals[rid]["name"] = row_dict.get("Report Name") or ""
            report_totals[rid]["cpf"] = cpf
            report_totals[rid]["nome"] = row_dict.get("Nome") or ""
            report_totals[rid]["status"] = row_dict.get("Status") or ""
        
        if cpf:
            cpf_totals[cpf]["count"] += 1
            cpf_totals[cpf]["value"] += val
            if rid:
                cpf_totals[cpf]["reports"].add(rid)
    
    wb.close()
    print(f"  Loaded {len(expenses)} expenses, {len(report_totals)} reports, {len(cpf_totals)} CPFs in {time.time()-t0:.1f}s")
    return expenses, report_totals, cpf_totals

def load_cpf_summary(filepath, sheet_name="CPF Summary"):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]
    summary = {}
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
            continue
        d = dict(zip(headers, row))
        cpf = str(d.get("CPF") or "").strip()
        if cpf:
            summary[cpf] = {
                "nome": d.get("Nome") or "",
                "report_count": int(d.get("Qtd Relatórios") or 0),
                "expense_count": int(d.get("Qtd Despesas") or 0),
                "total_prestacao": float(d.get("Total Prestação") or 0),
            }
    wb.close()
    return summary

def load_controle_painel(filepath):
    print(f"Loading CONTROLE PAINEL...")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['PAINEL']
    painel = {}
    for row in ws.iter_rows(min_row=12, values_only=True):
        cpf = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        if not cpf or cpf == 'None':
            continue
        painel[cpf] = {
            'nome': str(row[1]).strip() if row[1] else '',
            'carga': float(row[13]) if row[13] is not None else 0,
            'transferencia': float(row[14]) if row[14] is not None else 0,
            'tarifa': float(row[15]) if row[15] is not None else 0,
            'prestacao': float(row[16]) if row[16] is not None else 0,
            'saldo_prestacao': float(row[17]) if row[17] is not None else 0,
            'saldo_cartao': float(row[18]) if row[18] is not None else 0,
            'saldo_final': float(row[19]) if row[19] is not None else 0,
        }
    wb.close()
    print(f"  PAINEL: {len(painel)} CPFs loaded")
    return painel

# ===== Load all data =====
old_exp, old_reports, old_cpf = load_expenses(OLD_FILE)
new_exp, new_reports, new_cpf = load_expenses(NEW_FILE)
old_cpf_summary = load_cpf_summary(OLD_FILE)
new_cpf_summary = load_cpf_summary(NEW_FILE)
painel = load_controle_painel(CONTROLE_FILE)

# ===== Compute diffs =====
old_ids = set(old_exp.keys())
new_ids = set(new_exp.keys())
deleted_ids = old_ids - new_ids
inserted_ids = new_ids - old_ids
common_ids = old_ids & new_ids

updated_ids = []
for eid in common_ids:
    old_val = float(old_exp[eid].get("Expense Value") or 0)
    new_val = float(new_exp[eid].get("Expense Value") or 0)
    if abs(old_val - new_val) > 0.01:
        updated_ids.append(eid)

print(f"\n=== SYNC DIFFS ===")
print(f"Deleted: {len(deleted_ids)}")
print(f"Inserted: {len(inserted_ids)}")
print(f"Updated (value changed): {len(updated_ids)}")
print(f"Unchanged: {len(common_ids) - len(updated_ids)}")

# Report-level diffs
report_diffs = []
all_report_ids = set(old_reports.keys()) | set(new_reports.keys())
for rid in sorted(all_report_ids):
    old_r = old_reports.get(rid, {"count": 0, "value": 0, "name": "", "cpf": "", "nome": "", "status": ""})
    new_r = new_reports.get(rid, {"count": 0, "value": 0, "name": "", "cpf": "", "nome": "", "status": ""})
    count_diff = new_r["count"] - old_r["count"]
    value_diff = new_r["value"] - old_r["value"]
    if abs(count_diff) > 0 or abs(value_diff) > 0.01:
        report_diffs.append({
            "rid": rid,
            "name": new_r["name"] or old_r["name"],
            "cpf": new_r["cpf"] or old_r["cpf"],
            "nome": new_r["nome"] or old_r["nome"],
            "old_count": old_r["count"],
            "new_count": new_r["count"],
            "count_diff": count_diff,
            "old_value": old_r["value"],
            "new_value": new_r["value"],
            "value_diff": value_diff,
        })

report_diffs.sort(key=lambda x: abs(x["value_diff"]), reverse=True)
print(f"Reports with changes: {len(report_diffs)}")

# CPF-level diffs (old vs new DB)
cpf_diffs_db = []
all_cpfs_db = set(old_cpf_summary.keys()) | set(new_cpf_summary.keys())
for cpf in sorted(all_cpfs_db):
    old_c = old_cpf_summary.get(cpf, {"nome": "", "report_count": 0, "expense_count": 0, "total_prestacao": 0})
    new_c = new_cpf_summary.get(cpf, {"nome": "", "report_count": 0, "expense_count": 0, "total_prestacao": 0})
    exp_diff = new_c["expense_count"] - old_c["expense_count"]
    val_diff = new_c["total_prestacao"] - old_c["total_prestacao"]
    if abs(exp_diff) > 0 or abs(val_diff) > 0.01:
        cpf_diffs_db.append({
            "cpf": cpf,
            "nome": new_c["nome"] or old_c["nome"],
            "old_expenses": old_c["expense_count"],
            "new_expenses": new_c["expense_count"],
            "expense_diff": exp_diff,
            "old_total": old_c["total_prestacao"],
            "new_total": new_c["total_prestacao"],
            "value_diff": val_diff,
        })

cpf_diffs_db.sort(key=lambda x: abs(x["value_diff"]), reverse=True)
print(f"CPFs with DB changes: {len(cpf_diffs_db)}")

# CPF-level diffs (DB vs PAINEL)
cpf_diffs_painel = []
all_cpfs_painel = set(new_cpf_summary.keys()) | set(painel.keys())
for cpf in sorted(all_cpfs_painel):
    db_c = new_cpf_summary.get(cpf, {"nome": "", "total_prestacao": 0})
    pn = painel.get(cpf, {"nome": "", "prestacao": 0})
    db_val = db_c["total_prestacao"]
    pn_val = pn["prestacao"]
    diff = db_val - pn_val
    if abs(diff) > 0.01:
        cpf_diffs_painel.append({
            "cpf": cpf,
            "nome": db_c["nome"] or pn["nome"],
            "db_total": db_val,
            "painel_total": pn_val,
            "diff": diff,
        })

cpf_diffs_painel.sort(key=lambda x: abs(x["diff"]), reverse=True)
print(f"CPFs with DB vs PAINEL diffs: {len(cpf_diffs_painel)}")

# ===== Build Excel =====
print(f"\nBuilding verification Excel...")
wb = Workbook()

# ---- Sheet 1: Investigation (detailed) ----
ws1 = wb.active
ws1.title = "Investigation"

# Section 1: Expense-level diffs
r = 1
c = ws1.cell(row=r, column=1, value="=== EXPENSE-LEVEL DIFFS (V1 vs V2) ===")
c.font = Font(bold=True, size=14)
r += 2

# Deleted expenses
c = ws1.cell(row=r, column=1, value=f"DELETED ({len(deleted_ids)} expenses removed from DB)")
c.font = Font(bold=True, color="FF0000")
r += 1
del_headers = ["Expense ID", "Report ID", "Report Name", "CPF", "Nome", "Description", "Old Value"]
for col, h in enumerate(del_headers, 1):
    cell = ws1.cell(row=r, column=col, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
r += 1
for eid in sorted(deleted_ids):
    e = old_exp[eid]
    vals = [eid, e.get("Report ID"), e.get("Report Name"), e.get("CPF"), e.get("Nome"),
            e.get("Expense Description"), float(e.get("Expense Value") or 0)]
    for col, v in enumerate(vals, 1):
        cell = ws1.cell(row=r, column=col, value=v)
        cell.border = thin_border
        if col == 7: cell.number_format = '#,##0.00'; cell.fill = red_fill
    r += 1

r += 2
# Inserted expenses
c = ws1.cell(row=r, column=1, value=f"INSERTED ({len(inserted_ids)} expenses added to DB)")
c.font = Font(bold=True, color="008000")
r += 1
ins_headers = ["Expense ID", "Report ID", "Report Name", "CPF", "Nome", "Description", "New Value"]
for col, h in enumerate(ins_headers, 1):
    cell = ws1.cell(row=r, column=col, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
r += 1
for eid in sorted(inserted_ids):
    e = new_exp[eid]
    vals = [eid, e.get("Report ID"), e.get("Report Name"), e.get("CPF"), e.get("Nome"),
            e.get("Expense Description"), float(e.get("Expense Value") or 0)]
    for col, v in enumerate(vals, 1):
        cell = ws1.cell(row=r, column=col, value=v)
        cell.border = thin_border
        if col == 7: cell.number_format = '#,##0.00'; cell.fill = green_fill
    r += 1

r += 2
# Updated expenses
c = ws1.cell(row=r, column=1, value=f"UPDATED ({len(updated_ids)} expenses with value changes)")
c.font = Font(bold=True, color="FFA500")
r += 1
upd_headers = ["Expense ID", "Report ID", "Report Name", "CPF", "Nome", "Description", "Old Value", "New Value", "Diff"]
for col, h in enumerate(upd_headers, 1):
    cell = ws1.cell(row=r, column=col, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
r += 1
for eid in sorted(updated_ids):
    oe = old_exp[eid]
    ne = new_exp[eid]
    old_v = float(oe.get("Expense Value") or 0)
    new_v = float(ne.get("Expense Value") or 0)
    vals = [eid, ne.get("Report ID"), ne.get("Report Name"), ne.get("CPF"), ne.get("Nome"),
            ne.get("Expense Description"), old_v, new_v, new_v - old_v]
    for col, v in enumerate(vals, 1):
        cell = ws1.cell(row=r, column=col, value=v)
        cell.border = thin_border
        if col in (7, 8, 9): cell.number_format = '#,##0.00'; cell.fill = yellow_fill
    r += 1

r += 2
# Report-level diffs
c = ws1.cell(row=r, column=1, value=f"=== REPORT-LEVEL DIFFS ({len(report_diffs)} reports changed) ===")
c.font = Font(bold=True, size=14)
r += 1
rep_headers = ["Report ID", "Report Name", "CPF", "Nome", "Old Count", "New Count", "Count Diff", "Old Value", "New Value", "Value Diff"]
for col, h in enumerate(rep_headers, 1):
    cell = ws1.cell(row=r, column=col, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
r += 1
for d in report_diffs:
    vals = [d["rid"], d["name"], d["cpf"], d["nome"], d["old_count"], d["new_count"],
            d["count_diff"], d["old_value"], d["new_value"], d["value_diff"]]
    for col, v in enumerate(vals, 1):
        cell = ws1.cell(row=r, column=col, value=v)
        cell.border = thin_border
        if col in (8, 9, 10): cell.number_format = '#,##0.00'
        if col == 10 and abs(v) > 0.01: cell.fill = yellow_fill
    r += 1

r += 2
# CPF-level DB diffs
c = ws1.cell(row=r, column=1, value=f"=== CPF-LEVEL DB DIFFS (V1 vs V2: {len(cpf_diffs_db)} CPFs changed) ===")
c.font = Font(bold=True, size=14)
r += 1
cpf_db_headers = ["CPF", "Nome", "Old Expenses", "New Expenses", "Expense Diff", "Old Total", "New Total", "Value Diff"]
for col, h in enumerate(cpf_db_headers, 1):
    cell = ws1.cell(row=r, column=col, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
r += 1
for d in cpf_diffs_db:
    vals = [d["cpf"], d["nome"], d["old_expenses"], d["new_expenses"], d["expense_diff"],
            d["old_total"], d["new_total"], d["value_diff"]]
    for col, v in enumerate(vals, 1):
        cell = ws1.cell(row=r, column=col, value=v)
        cell.border = thin_border
        if col in (6, 7, 8): cell.number_format = '#,##0.00'
        if col == 8 and abs(v) > 0.01: cell.fill = yellow_fill
    r += 1

r += 2
# CPF-level PAINEL diffs
c = ws1.cell(row=r, column=1, value=f"=== CPF-LEVEL DB vs PAINEL DIFFS ({len(cpf_diffs_painel)} CPFs differ) ===")
c.font = Font(bold=True, size=14)
r += 1
cpf_pn_headers = ["CPF", "Nome", "DB Total", "PAINEL Total", "Diff (DB-PAINEL)"]
for col, h in enumerate(cpf_pn_headers, 1):
    cell = ws1.cell(row=r, column=col, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
r += 1
for d in cpf_diffs_painel:
    vals = [d["cpf"], d["nome"], d["db_total"], d["painel_total"], d["diff"]]
    for col, v in enumerate(vals, 1):
        cell = ws1.cell(row=r, column=col, value=v)
        cell.border = thin_border
        if col in (3, 4, 5): cell.number_format = '#,##0.00'
        if col == 5 and abs(v) > 0.01: cell.fill = yellow_fill
    r += 1

# Column widths for investigation
for col, w in enumerate([14, 25, 12, 15, 30, 30, 15, 15, 15, 15], 1):
    ws1.column_dimensions[get_column_letter(col)].width = w

# ---- Sheet 2: User View (clean summary) ----
ws2 = wb.create_sheet("Resumo")

r = 1
c = ws2.cell(row=r, column=1, value="VERIFICAÇÃO PÓS-SYNC")
c.font = Font(bold=True, size=16)
r += 2

# Summary stats
c = ws2.cell(row=r, column=1, value="RESUMO DO SYNC")
c.font = Font(bold=True, size=13)
r += 1
summary = [
    ("Total de despesas (antes)", len(old_exp)),
    ("Total de despesas (depois)", len(new_exp)),
    ("Despesas removidas", len(deleted_ids)),
    ("Despesas adicionadas", len(inserted_ids)),
    ("Despesas com valor alterado", len(updated_ids)),
    ("Relatórios com mudança", len(report_diffs)),
    ("CPFs com mudança no DB", len(cpf_diffs_db)),
    ("CPFs com diff vs PAINEL", len(cpf_diffs_painel)),
]
for label, val in summary:
    ws2.cell(row=r, column=1, value=label).font = Font(bold=True)
    ws2.cell(row=r, column=2, value=val)
    r += 1

r += 2
# Top CPF diffs vs PAINEL
c = ws2.cell(row=r, column=1, value="TOP 30 DIFERENÇAS: DB vs PAINEL (CONTROLE)")
c.font = Font(bold=True, size=13)
r += 1
pn_headers = ["CPF", "Nome", "DB Total (R$)", "PAINEL Total (R$)", "Diferença (R$)"]
for col, h in enumerate(pn_headers, 1):
    cell = ws2.cell(row=r, column=col, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
r += 1
for d in cpf_diffs_painel[:30]:
    vals = [d["cpf"], d["nome"], d["db_total"], d["painel_total"], d["diff"]]
    for col, v in enumerate(vals, 1):
        cell = ws2.cell(row=r, column=col, value=v)
        cell.border = thin_border
        if col in (3, 4, 5): cell.number_format = '#,##0.00'
        if col == 5:
            if v > 0.01: cell.fill = red_fill
            elif v < -0.01: cell.fill = green_fill
    r += 1

r += 2
# Top report diffs
c = ws2.cell(row=r, column=1, value="TOP 30 RELATÓRIOS COM MAIOR MUDANÇA DE VALOR")
c.font = Font(bold=True, size=13)
r += 1
rep_headers2 = ["Report ID", "Report Name", "CPF", "Nome", "Qtd Antes", "Qtd Depois", "Valor Antes (R$)", "Valor Depois (R$)", "Diff (R$)"]
for col, h in enumerate(rep_headers2, 1):
    cell = ws2.cell(row=r, column=col, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
r += 1
for d in report_diffs[:30]:
    vals = [d["rid"], d["name"], d["cpf"], d["nome"], d["old_count"], d["new_count"],
            d["old_value"], d["new_value"], d["value_diff"]]
    for col, v in enumerate(vals, 1):
        cell = ws2.cell(row=r, column=col, value=v)
        cell.border = thin_border
        if col in (7, 8, 9): cell.number_format = '#,##0.00'
        if col == 9:
            if v > 0.01: cell.fill = green_fill
            elif v < -0.01: cell.fill = red_fill
    r += 1

r += 2
# Top CPF DB diffs
c = ws2.cell(row=r, column=1, value="TOP 30 CPFs COM MAIOR MUDANÇA NO DB (V1 vs V2)")
c.font = Font(bold=True, size=13)
r += 1
cpf_db_headers2 = ["CPF", "Nome", "Desp Antes", "Desp Depois", "Total Antes (R$)", "Total Depois (R$)", "Diff (R$)"]
for col, h in enumerate(cpf_db_headers2, 1):
    cell = ws2.cell(row=r, column=col, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
r += 1
for d in cpf_diffs_db[:30]:
    vals = [d["cpf"], d["nome"], d["old_expenses"], d["new_expenses"],
            d["old_total"], d["new_total"], d["value_diff"]]
    for col, v in enumerate(vals, 1):
        cell = ws2.cell(row=r, column=col, value=v)
        cell.border = thin_border
        if col in (5, 6, 7): cell.number_format = '#,##0.00'
        if col == 7:
            if v > 0.01: cell.fill = green_fill
            elif v < -0.01: cell.fill = red_fill
    r += 1

# Column widths for user view
for col, w in enumerate([15, 30, 15, 15, 15, 15, 18, 18, 15], 1):
    ws2.column_dimensions[get_column_letter(col)].width = w

# Save
print(f"Saving to {OUTPUT}...")
wb.save(OUTPUT)
print("Done!")
