import openpyxl

files = [
    'fechamentos/F Carlos Nascimento Nonato.xlsx',
    'fechamentos/F DHIEGO RIBEIRO.xlsx',
    'fechamentos/FECHAMENTO - ADAN LEONARDO SOUZA BATISTA.xlsx',
    'fechamentos/FECHAMENTO - ANDRE VALERIO DE PAIVA.xlsx',
]

for f in files:
    print(f"\n{'=' * 60}")
    print(f"FILE: {f}")
    print(f"{'=' * 60}")
    wb = openpyxl.load_workbook(f, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    # Check if there's a PREST. CONTAS sheet
    for sn in wb.sheetnames:
        if 'PREST' in sn.upper() or 'CONTAS' in sn.upper():
            ws = wb[sn]
            print(f"\n  Sheet '{sn}': {ws.max_row} rows x {ws.max_column} cols")
            # Print header row
            for r in range(1, min(5, ws.max_row + 1)):
                vals = [str(ws.cell(r, c).value or '') for c in range(1, min(15, ws.max_column + 1))]
                print(f"    R{r}: {vals}")
            
            # Check if there's a payment method column
            headers = [str(ws.cell(1, c).value or '').upper() for c in range(1, ws.max_column + 1)]
            # Also check row 2, 3
            for r in [2, 3]:
                headers += [str(ws.cell(r, c).value or '').upper() for c in range(1, ws.max_column + 1)]
            
            if any('PAGAMENTO' in h or 'PAYMENT' in h or 'ITAU' in h or 'CARTAO' in h or 'CARTÃO' in h for h in headers):
                print(f"    *** HAS PAYMENT METHOD COLUMN ***")
            else:
                print(f"    No payment method column found in headers")
            
            # Print first few data rows
            for r in range(2, min(8, ws.max_row + 1)):
                vals = [str(ws.cell(r, c).value or '') for c in range(1, min(15, ws.max_column + 1))]
                print(f"    R{r}: {vals}")
